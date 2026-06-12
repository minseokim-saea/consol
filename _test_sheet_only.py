"""
시트 보호 검증:
- 파일은 비밀번호 없이 열려야 함
- TB(m) 시트만 protection.sheet = False
- 나머지 모든 시트는 protection.sheet = True
"""
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import distribute_builder as db
from openpyxl import load_workbook

state = json.loads(Path("uploads/_state.json").read_text(encoding="utf-8"))

tpl = db.get_template_path("2026")
out_dir = Path("results/distribute/_test_sheet_only")
out_dir.mkdir(parents=True, exist_ok=True)

for p in out_dir.glob("*.xlsm"):
    p.unlink()

res = db.build_distribution_package(
    template_path=tpl,
    output_dir=out_dir,
    company="글로벌세아",
    target_year=2026,
    target_quarter=1,
    uploaded_files=state,
    file_password=None,
    sheet_protect_password=None,
)

print(f"ok={res['ok']}, output={res['output_path']}")
print(f"시트 보호 암호: {res['file_password']}")
print(f"wrote_bs={res['wrote_bs']}, wrote_pl={res['wrote_pl']}")

print("\n=== 비밀번호 없이 파일 열기 ===")
wb = load_workbook(res["output_path"], keep_vba=True, data_only=False)
print(f"[OK] 비밀번호 없이 열림. 시트 {len(wb.sheetnames)}개")

print("\n=== 시트별 보호 상태 ===")
protected, unprotected = [], []
for sn in wb.sheetnames:
    ws = wb[sn]
    locked = bool(ws.protection.sheet)
    (protected if locked else unprotected).append(sn)
    flag = "[LOCK]" if locked else "[OPEN]"
    print(f"  {flag} {sn}")

print(f"\n보호된 시트: {len(protected)}개")
print(f"보호 안된 시트: {unprotected}")

# 검증
expected_open = ["TB(m)"]
if unprotected == expected_open:
    print(f"\n[PASS] TB(m)만 열려있고 나머지({len(protected)}개)는 모두 보호됨")
else:
    print(f"\n[FAIL] 기대: {expected_open}, 실제: {unprotected}")
    sys.exit(1)

wb.close()
