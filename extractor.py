"""
25YE PKG 재무보고서 환산값(KRW) 추출.

- 회사명: Cover 시트 D11 (국문 회사명)
- 통화: Cover 시트 B28 (예: 'KRW', 'USD', 'VND')
- 환율: Master 시트 당기 Fx 표 L열(통화) / N열(평균환율)
- BS: 207행부터 데이터, G열 (KRW 환산값)
- PL_MF: 10행부터 데이터, N열 (In KRW 환산값)
- CF: 2행부터, B열(항목명) / C열(값) / F열(참조코드)
       · 환산값 없음 → 해당 통화의 Master 평균환율을 곱해 KRW 환산
"""

import json
import os
import re
from collections import OrderedDict
from openpyxl import load_workbook

SHEET_CONFIG = {
    'BS': {
        'code_col': 2, 'kor_col': 3, 'eng_col': 4,
        'local_col': 6,         # F열 (당기 현지통화 — 중앙 환율 적용용)
        'value_col': 7,         # G열 (당기 환산값 KRW)
        'reason_col': 12,       # L열 (당기 변동사유 — 회사별 비교 출력용)
        'rate_type': 'spot_current',  # BS는 기말 현물환율로 환산
        'data_start_row': 207,
    },
    'PL_MF': {
        'code_col': 2, 'kor_col': 3, 'eng_col': 4,
        'local_col': 7,         # G열 (Adjusted K-GAAP — 회계정책 조정 후, 중앙 환율 적용용)
        'value_col': 14,        # N열 (In KRW)
        'reason_col': 12,       # L열 (당기 변동사유)
        'rate_type': 'avg',     # PL은 평균환율로 환산
        'data_start_row': 10,
    },
}


def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _round_krw(x):
    """KRW 환산값을 원(₩) 단위로 반올림 — 소수점 이하 제거.
    KRW 는 소수 단위가 없으므로 현지통화×환율 결과는 항상 정수 원으로 맞춘다.
    0.5 는 Excel ROUND 과 동일하게 절댓값이 큰 쪽으로(half-away-from-zero).
    현지통화 raw 값에는 적용하지 않는다(원 단위 환산값에만)."""
    if not _is_num(x):
        try:
            x = float(x)
        except (TypeError, ValueError):
            return x
    return float(int(x + 0.5)) if x >= 0 else float(-int(-x + 0.5))


# ─────────────────────────────────────────────────────────────
# BS 합계/소계 계정: '하위계정 환산값의 합'으로 재계산
#   합계계정을 개별 환산(로컬합계×환율)하면, 하위계정을 각각 반올림한 합과
#   단수(1~2원)가 어긋난다. 합계는 반드시 하위합과 일치해야 하므로
#   consol_template.json 의 합계 구조를 이용해 하위합으로 덮어쓴다.
# ─────────────────────────────────────────────────────────────
_BS_SUBTOTAL_MAP = None


def _sum_rows_of(row):
    """합계 행이 참조하는 하위 행 번호 목록. sum_range(잎 소계) 또는
    formula '=SUM(R60:R78)' / '=SUM(R44,R59,...)'(중첩 소계) 둘 다 해석."""
    if row.get('sum_range'):
        return list(row['sum_range'])
    f = row.get('formula') or ''
    rows = []
    for m in re.finditer(r'R(\d+):R(\d+)', f):        # 범위 Ra:Rb
        rows.extend(range(int(m.group(1)), int(m.group(2)) + 1))
    for m in re.finditer(r'R(\d+)', re.sub(r'R\d+:R\d+', '', f)):  # 개별 Rc
        rows.append(int(m.group(1)))
    return rows


def _load_bs_subtotal_map():
    """consol_template.json BS 섹션에서 [(합계코드, [하위코드...]), ...] (템플릿 순서). 캐시."""
    global _BS_SUBTOTAL_MAP
    if _BS_SUBTOTAL_MAP is not None:
        return _BS_SUBTOTAL_MAP
    result = []
    try:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'consol_template.json')
        with open(path, encoding='utf-8') as fp:
            tpl = json.load(fp)
        rows = tpl.get('rows') or []
        by_row = {r['row']: r for r in rows if 'row' in r}
        for r in rows:
            if r.get('kind') == 'subtotal' and r.get('section') == 'BS' and r.get('code'):
                kids = [str(by_row[rw]['code']) for rw in _sum_rows_of(r)
                        if rw in by_row and by_row[rw].get('code')]
                if kids:
                    result.append((str(r['code']), kids))
    except Exception:
        result = []
    _BS_SUBTOTAL_MAP = result
    return result


def _recompute_bs_subtotals(bs):
    """BS 합계/소계 계정의 KRW 환산값(value)을 하위계정 환산값의 합으로 재계산.
    중첩 소계(예: 비유동자산 = 유형자산 + 무형자산 + ...)는 다회 패스로 하위→상위 순 해소."""
    submap = _load_bs_subtotal_map()
    if not submap or not bs:
        return
    for _ in range(8):
        changed = False
        for sub_code, kids in submap:
            if sub_code not in bs:
                continue
            s = _round_krw(sum((bs[c].get('value', 0) or 0) for c in kids if c in bs))
            if bs[sub_code].get('value') != s:
                bs[sub_code]['value'] = s
                changed = True
        if not changed:
            break


# 대차 반올림 잔단 흡수 대상 (BS 총계 코드 + 흡수 계정)
_BS_ASSET_TOTAL = '1000000'      # 자산총계
_BS_LIAB_TOTAL = '2000000'       # 부채총계
_BS_EQUITY_TOTAL = '3000000'     # 자본총계
_BS_FX_TRANSLATION = '3400104'   # 해외사업환산손익(누적환산조정) — 잔단 흡수
_BS_BALANCE_PLUG_MAX = 200       # 이 이하(원)만 흡수. 초과는 실제 불일치로 보고 건드리지 않음


def _balance_bs_rounding(bs):
    """소계를 하위합으로 맞춘 뒤 남는 대차(자산총계 − (부채총계+자본총계)) 반올림 잔단을
    해외사업환산손익(3400104)에 흡수해 대차를 0으로 만든다.
    잔단이 반올림 규모를 넘으면(실제 데이터 불일치 가능) 손대지 않는다."""
    a = bs.get(_BS_ASSET_TOTAL, {}).get('value')
    l = bs.get(_BS_LIAB_TOTAL, {}).get('value')
    e = bs.get(_BS_EQUITY_TOTAL, {}).get('value')
    if a is None or l is None or e is None:
        return
    r = a - (l + e)
    if abs(r) < 0.5 or abs(r) > _BS_BALANCE_PLUG_MAX:
        return
    if _BS_FX_TRANSLATION not in bs:
        return
    bs[_BS_FX_TRANSLATION]['value'] = (bs[_BS_FX_TRANSLATION].get('value', 0) or 0) + r
    _recompute_bs_subtotals(bs)   # 자본 소계·자본총계 재합산 → 대차 0


def _get_company_name(wb):
    if 'Cover' in wb.sheetnames:
        v = wb['Cover']['D11'].value
        if v is not None:
            # 개행/탭/연속 공백 정리 → URL 안전 + 일관된 회사명
            s = ' '.join(str(v).split())
            if s:
                return s
    return None


def _get_index_error_count(wb):
    """Index 시트 C12의 Error 개수 반환. 읽지 못하면 None. (시트명 대소문자 무관)"""
    # 대소문자 무관하게 'Index' 시트 찾기
    target = None
    for name in wb.sheetnames:
        if name.strip().lower() == 'index':
            target = name
            break
    if target is None:
        return None

    v = wb[target]['C12'].value
    if _is_num(v):
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        if s.isdigit():
            return int(s)
    return None


def _get_year_quarter(wb):
    """Cover 시트의 결산연도(C9)와 분기(F9)를 읽어 반환.
    반환: {'year': 'YYYY' or None, 'quarter': '1'~'4' or None}
    """
    out = {'year': None, 'quarter': None}
    if 'Cover' not in wb.sheetnames:
        return out
    ws = wb['Cover']

    y = ws['C9'].value
    if y is not None:
        ys = str(y).strip()
        import re as _re
        m = _re.search(r'(\d{4})', ys)
        if m:
            out['year'] = m.group(1)

    q = ws['F9'].value
    if q is not None:
        qs = str(q).strip()
        import re as _re
        # "1", "1Q", "1분기", "Q1" 등 다양한 표기 허용
        m = _re.search(r'[1-4]', qs)
        if m:
            out['quarter'] = m.group(0)

    return out


def _get_currency(wb):
    """Cover B28 — 해당 회사의 통화 코드."""
    if 'Cover' in wb.sheetnames:
        v = wb['Cover']['B28'].value
        if v is not None:
            return str(v).strip().upper()
    return None


def _read_fx_table(ws, start_row, max_scan=25):
    """
    한 개 Fx 블록 읽기. 열:
      J(10)=Nation, K(11)=Currency, L(12)=Spot, M(13)=Avg
    'Fx rate' 헤더가 다시 나오면 중단.
    반환: {currency: {'spot': float, 'avg': float}}

    주의: 같은 통화코드가 여러 행에 걸쳐 등장할 수 있음
    (예: USA/USD와 Cambodia/USD가 별도 행에 존재).
    이 경우 None이 아닌 유효한 값이 이미 저장된 항목을
    None 값으로 덮어쓰지 않는다 (첫 번째 유효값 우선).
    """
    table = {}
    for r in range(start_row, start_row + max_scan):
        nation = ws.cell(r, 10).value
        if isinstance(nation, str) and ('Fx rate' in nation or 'Prior' in nation):
            break
        currency = ws.cell(r, 11).value
        spot = ws.cell(r, 12).value
        avg = ws.cell(r, 13).value
        if currency:
            key = str(currency).strip().upper()
            spot_val = float(spot) if _is_num(spot) else None
            avg_val  = float(avg)  if _is_num(avg)  else None
            if key not in table:
                table[key] = {'spot': spot_val, 'avg': avg_val}
            else:
                # 중복 통화코드: None이 아닌 값만 보완 (유효한 기존값 보존)
                if spot_val is not None:
                    table[key]['spot'] = spot_val
                if avg_val is not None:
                    table[key]['avg'] = avg_val
    return table


def _get_fx_rate_tables(wb):
    """
    Master 시트에서 당기/전기 환율 두 블록을 추출.
    - 당기 블록: J열 'Fx rate' 첫 등장 이후 2행 뒤부터 (예: 행 6~21)
    - 전기 블록: J열 'Fx rate' 두 번째 등장 이후 2행 뒤부터 (예: 행 29~44)
    반환: {'current': {cur: {'spot','avg'}}, 'prior': {cur: {'spot','avg'}}}
    """
    result = {'current': {}, 'prior': {}}
    if 'Master' not in wb.sheetnames:
        return result
    ws = wb['Master']

    fx_header_rows = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 10).value  # J열
        if isinstance(v, str) and v.strip() == 'Fx rate':
            fx_header_rows.append(r)

    # 각 'Fx rate' 헤더 다음 2행(헤더 스킵)부터 데이터 시작
    if len(fx_header_rows) >= 1:
        result['current'] = _read_fx_table(ws, fx_header_rows[0] + 2)
    if len(fx_header_rows) >= 2:
        result['prior']   = _read_fx_table(ws, fx_header_rows[1] + 2)
    return result


def _get_fx_rates(wb):
    """(하위호환) 당기 평균환율 사전만 반환."""
    tables = _get_fx_rate_tables(wb)
    return {c: info['avg'] for c, info in tables['current'].items() if info['avg'] is not None}


def _load_coa_eng(wb):
    """COA 시트에서 {code: 영문명} 매핑 반환.

    COA 시트 구조: A=Code | B=한글명 | C=Account Name (영문) | D=Description | E=표시과목 | F=분류
    BS/PL/CF 모든 코드 (4500xxx, 5xxxxxx, CF1xxxxxx ~ CF6xxxxxx 등) 망라.
    """
    if 'COA' not in wb.sheetnames:
        return {}
    ws = wb['COA']
    mapping = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        eng  = ws.cell(r, 3).value
        if code is None or eng is None:
            continue
        code_str = str(code).strip()
        eng_str  = str(eng).strip()
        if code_str and eng_str:
            mapping[code_str] = eng_str
    return mapping


def _extract_coded_sheet(ws, cfg, rate=None):
    """BS / PL_MF: 계정코드 기반 환산값 추출.

    rate가 주어지면 local_col × rate를 사용 (중앙 환율 우선 모드).
    rate=None 또는 local 값이 없으면 value_col(패키지 환산값) 폴백.
    """
    result = OrderedDict()
    use_central = (rate is not None) and ('local_col' in cfg)
    for r in range(cfg['data_start_row'], ws.max_row + 1):
        code = ws.cell(r, cfg['code_col']).value
        if code is None:
            continue
        code_str = str(code).strip()
        if not code_str or not code_str[0].isdigit():
            continue
        # 계정코드는 표준 7자리 (최소 4자리). 발맥스기술 같은 패키지는 헤더영역(R147)에
        # B=0, C=회사명 형태가 데이터 시작행 이후에 있어서 잘못 잡히는 경우 방지.
        if len(code_str) < 4:
            continue
        if code_str in result:
            continue

        kor = ws.cell(r, cfg['kor_col']).value
        eng = ws.cell(r, cfg['eng_col']).value

        val = 0
        local_raw = None
        if use_central:
            local = ws.cell(r, cfg['local_col']).value
            if _is_num(local):
                local_raw = float(local)
                val = local_raw * rate
            else:
                # local 값 없거나 숫자 아니면 패키지 환산값 폴백
                pkg_v = ws.cell(r, cfg['value_col']).value
                val = float(pkg_v) if _is_num(pkg_v) else 0
        else:
            # 중앙환율 미적용 모드: local 셀이 있으면 raw로 보관, 환산값은 패키지값
            if 'local_col' in cfg:
                lv = ws.cell(r, cfg['local_col']).value
                if _is_num(lv):
                    local_raw = float(lv)
            v = ws.cell(r, cfg['value_col']).value
            val = float(v) if _is_num(v) else 0
            # KRW 회사 등으로 local이 비어 있으면 환산값=로컬값(동일)
            if local_raw is None:
                local_raw = val

        # 당기 변동사유 (L열) — 회사별 비교 출력에 사용
        reason = ''
        if 'reason_col' in cfg:
            rv = ws.cell(r, cfg['reason_col']).value
            if rv is not None:
                reason = str(rv).strip()

        result[code_str] = {
            'kor': str(kor).strip() if kor else '',
            'eng': str(eng).strip() if eng else '',
            'value': _round_krw(val),
            'local_value': local_raw if local_raw is not None else 0,
            'compare': 0,         # 본체에서 채움 (KRW)
            'compare_local': 0,   # 본체에서 채움 (현지통화 raw)
            'reason': reason,
        }
    return result


def _extract_bs_prior_local(ws, start_row=14, end_row=197,
                            code_col=2, prior_local_col=8):
    """BS 시트의 전기 영역(기본 14~197행)에서 전기 현지통화 raw 추출.

    양식: 전기 BS와 당기 BS가 같은 시트 안에 두 블록으로 분리.
      · 14~197행: 전기 BS  (B열=코드, H열=전기 로컬 raw, PY 시트 A~B열에서 참조)
      · 207행~ : 당기 BS  (_extract_coded_sheet이 처리)

    반환: {code_str: float}
    """
    result = {}
    max_r = min(end_row, ws.max_row)
    for r in range(start_row, max_r + 1):
        code = ws.cell(r, code_col).value
        if code is None:
            continue
        code_str = str(code).strip()
        if not code_str or not code_str[0].isdigit() or len(code_str) < 4:
            continue
        if code_str in result:
            continue   # 같은 코드 중복 등장 시 첫 행 우선
        v = ws.cell(r, prior_local_col).value
        if _is_num(v):
            result[code_str] = float(v)
    return result


def _code_str(v):
    """PY 시트 코드 셀 값을 문자열로 변환. 1000000.0 → '1000000'."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s if s else None


def _extract_py_compare(wb, avg_rate_prior, bs_rescale_ratio=1.0, spot_prior=None):
    """
    PY 시트에서 전기 비교값 추출.
      BS: D(4)=코드, E(5)=금액 (이미 KRW → bs_rescale_ratio 적용)
      PL: G(7)=코드, H(8)=금액 (로컬통화 × 전기 avg rate → KRW 환산)
    데이터 시작행: 5 (헤더 2행, 4행)

    bs_rescale_ratio:
      = (중앙 spot_prior / 패키지 spot_prior) — 중앙 환율로 재조정 시 사용
      = 1.0 (기본) — 그대로 사용

    spot_prior:
      BS 로컬 raw 추정용 환율. PY-BS는 패키지 양식상 KRW만 보관하므로
      로컬통화 비교가 필요한 출력물은 KRW / spot_prior로 역산해 사용.
      None/0이면 로컬 raw는 0.

    반환: {
      'bs': {code: krw_value},                # bs_rescale_ratio 적용된 KRW (합산용)
      'bs_pkg_raw': {code: krw_value},        # PY 시트 E열 KRW 원본 (rescale 안 함)
      'bs_local': {code: local_raw_value},    # spot_prior로 역산 (추정)
      'pl': {code: krw_value},
      'pl_local': {code: local_raw_value},    # H열 raw 그대로
    }
    """
    result = {'bs': {}, 'bs_pkg_raw': {}, 'bs_local': {}, 'pl': {}, 'pl_local': {}}
    if 'PY' not in wb.sheetnames:
        return result
    ws = wb['PY']
    sp = float(spot_prior) if (spot_prior and _is_num(spot_prior) and float(spot_prior) != 0) else None
    for r in range(5, ws.max_row + 1):
        # BS KRW (D=4 code, E=5 amount — already KRW; 중앙 환율 재조정 비율 적용)
        bs_c = _code_str(ws.cell(r, 4).value)
        if bs_c and bs_c[0].isdigit() and bs_c not in result['bs']:
            amt = ws.cell(r, 5).value
            base = float(amt) if _is_num(amt) else 0
            krw = _round_krw(base * bs_rescale_ratio)
            result['bs'][bs_c] = krw
            result['bs_pkg_raw'][bs_c] = base    # PY 시트 KRW 원본 (회사별 비교용)
            # 로컬 raw 역산 (KRW / 적용 spot_prior) — sp가 1.0이면 KRW와 동일(KRW 회사)
            result['bs_local'][bs_c] = (krw / sp) if sp else krw

        # PL local (G=7 code, H=8 amount — local currency × prior avg rate)
        pl_c = _code_str(ws.cell(r, 7).value)
        if pl_c and pl_c[0].isdigit() and pl_c not in result['pl']:
            amt = ws.cell(r, 8).value
            raw = float(amt) if _is_num(amt) else 0
            result['pl_local'][pl_c] = raw
            result['pl'][pl_c] = _round_krw(raw * avg_rate_prior)

    return result


def _extract_cf1_sheet(ws, avg_rate, spot_current, spot_prior, value_col=7, coa=None):
    """
    CF1 '1. 증감내역(Details of Changes)' 섹션에서 각 계정의 증감항목을 추출.

    구조 (헤더 R6):
      B: 계정코드 (블록 첫 행)
      C: 증감항목 한글 (기초금액 / 신규대여 / ... / 기말금액)
      D: 증감항목 영문
      G: Total 금액

    환산 규칙:
      · '기초금액'  → 전기 Spot rate
      · '기말금액'  → 당기 Spot rate
      · 나머지      → 당기 Avg rate
      · '해외사업환산손익' = 기말금액 − (기초 + 기초수정 + ... + 외화평가감소)  [자동 계산]

    반환: OrderedDict[ '{code}::{label}' → {'kor','eng','value'} ]
    """
    # 1) 섹션 경계 찾기
    section_start, section_end = None, ws.max_row + 1
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if not isinstance(b, str):
            continue
        bs = b.strip()
        if section_start is None and bs.startswith('1.') and '증감' in bs:
            section_start = r
        elif section_start and (bs.startswith('2.') or bs.startswith('3.')):
            section_end = r
            break
    if section_start is None:
        return OrderedDict()

    # 2) 계정 블록별로 데이터 수집 (원본 G열 값)
    blocks = OrderedDict()      # code → {'kor','eng', 'rows': OrderedDict[label → raw_value]}
    current_code = None
    for r in range(section_start + 1, section_end):
        b_val = ws.cell(r, 2).value
        c_val = ws.cell(r, 3).value

        if c_val is None:
            continue
        label = str(c_val).strip()
        if not label:
            continue

        # 블록 시작 감지: C열='기초금액' + B열이 숫자 계정코드
        if '기초금액' in label and _is_num(b_val):
            code_num = int(b_val) if isinstance(b_val, float) and b_val.is_integer() else b_val
            current_code = str(code_num).strip()
            # 계정명은 블록 첫 행 다음 두 행의 B열
            kor_name = ws.cell(r + 1, 2).value or ''
            eng_name = ws.cell(r + 2, 2).value or ''
            blocks[current_code] = {
                'kor': str(kor_name).strip(),
                'eng': str(eng_name).strip(),
                'rows': OrderedDict(),
            }

        if current_code is None:
            continue

        raw_cell = ws.cell(r, value_col).value
        v = float(raw_cell) if _is_num(raw_cell) else 0.0
        # 같은 라벨이 중복되면 첫 번째만
        if label not in blocks[current_code]['rows']:
            blocks[current_code]['rows'][label] = v

    # 3) 환산 + 해외사업환산손익 자동 계산 (raw는 별도 보존)
    for code, block in blocks.items():
        raw_rows = block['rows']     # 원래 raw (현지통화)
        converted = OrderedDict()
        for label, raw in raw_rows.items():
            if '기초금액' in label:
                rate = spot_prior
            elif '기말금액' in label:
                rate = spot_current
            else:
                rate = avg_rate
            converted[label] = _round_krw(raw * rate)

        # 해외사업환산손익 = 기말금액 − (기초부터 외화평가감소까지의 합)
        ending = 0
        sum_to_fx_dec = 0
        fx_gain_key = None
        for label, val in converted.items():
            if '기말금액' in label:
                ending = val
            elif '해외사업환산' in label:
                fx_gain_key = label  # 덮어쓸 키 기록
            else:
                sum_to_fx_dec += val

        if fx_gain_key is not None:
            converted[fx_gain_key] = ending - sum_to_fx_dec
            # 해외사업환산손익의 raw 등가 = 잔차 / 평균환율 (KRW 잔차를 평균환율로 역환산)
            raw_rows[fx_gain_key] = (converted[fx_gain_key] / avg_rate) if avg_rate else 0

        block['rows_converted'] = converted
        block['rows_local']     = raw_rows

    # 4) 평탄화(flat dict)로 변환 — eng는 시트값 우선, 없으면 COA 매핑
    flat = OrderedDict()
    for code, block in blocks.items():
        kor = block['kor']
        eng = block['eng'] or (coa.get(code) if coa else '') or ''
        for label, val in block['rows_converted'].items():
            key = f'{code}::{label}'
            flat[key] = {
                'kor': f'{kor} / {label}' if kor else label,
                'eng': eng,
                'value': val,
                'local_value': block['rows_local'].get(label, 0),
                'compare': 0,
                'compare_local': 0,
            }
    return flat


def _extract_cf4_section(ws, header_row, end_row, avg_rate, spot_current, spot_prior, coa=None):
    """
    CF4의 한 섹션(취득가액/감가상각/손상차손) 추출.

    header_row: 'Code'가 B열에 있는 헤더 행
    end_row: 섹션 종료 경계(exclusive)
    환산 규칙:
      · 기초잔액/기초수정금액 → 전기 spot
      · 기말잔액              → 당기 spot
      · 그 사이 모든 변동 컬럼 → 당기 avg
      · 해외사업환산손익 = 기말잔액 − (기초~기타변동 합)  [자동 삽입]
    """
    # 1) 헤더에서 각 컬럼 유형 분류
    beginning_cols = []   # [(col_idx, label)]
    middle_cols = []      # [(col_idx, label)]
    ending_col = None     # (col_idx, label)
    category_col = None   # 용도구분처럼 라벨성 컬럼 (환산 대상 아님)

    for c in range(4, 16):  # D ~ O
        v = ws.cell(header_row, c).value
        if not isinstance(v, str):
            continue
        text = v
        label = text.split('\n')[0].strip()
        if not label:
            continue
        if '기초잔액' in text or 'Beginning' in text:
            beginning_cols.append((c, label))
        elif '기초수정' in text:
            beginning_cols.append((c, label))
        elif '기말잔액' in text or 'Ending' in text:
            ending_col = (c, label)
        elif '용도구분' in text or ('Account name' in text and c == 4):
            category_col = c
        else:
            middle_cols.append((c, label))

    if ending_col is None or not beginning_cols:
        return OrderedDict()

    # 2) 데이터 행 처리 (B열에 숫자 계정코드가 있는 행만)
    result = OrderedDict()
    for r in range(header_row + 1, end_row):
        code = ws.cell(r, 2).value
        if not _is_num(code):
            continue
        code_str = str(int(code) if isinstance(code, float) and code.is_integer() else code).strip()
        kor = ws.cell(r, 3).value
        kor_str = str(kor).strip() if kor else ''

        # 용도구분(카테고리) 값 → 같은 계정코드 중복 구분용
        category = ''
        if category_col is not None:
            cv = ws.cell(r, category_col).value
            if cv is not None and not _is_num(cv):
                category = str(cv).strip()

        # 환산 및 합계 (raw 별도 보존)
        converted = OrderedDict()
        local_raws = OrderedDict()
        beginning_plus_middle = 0.0

        for c, label in beginning_cols:
            raw = ws.cell(r, c).value
            raw_f = float(raw) if _is_num(raw) else 0.0
            local_raws[label] = raw_f
            val = _round_krw(raw_f * spot_prior)
            converted[label] = val
            beginning_plus_middle += val

        for c, label in middle_cols:
            raw = ws.cell(r, c).value
            raw_f = float(raw) if _is_num(raw) else 0.0
            local_raws[label] = raw_f
            val = _round_krw(raw_f * avg_rate)
            converted[label] = val
            beginning_plus_middle += val

        ending_raw = ws.cell(r, ending_col[0]).value
        ending_raw_f = float(ending_raw) if _is_num(ending_raw) else 0.0
        ending_val = _round_krw(ending_raw_f * spot_current)

        # 해외사업환산손익 = 기말 − (기초~기타변동 합)  → 기말잔액 바로 앞에 삽입
        fx_diff = ending_val - beginning_plus_middle
        converted['해외사업환산손익'] = fx_diff
        local_raws['해외사업환산손익'] = (fx_diff / avg_rate) if avg_rate else 0
        converted[ending_col[1]] = ending_val
        local_raws[ending_col[1]] = ending_raw_f

        # 같은 계정코드 여러 행(예: 감가상각 SGA/MF)은 합산
        eng_name = (coa.get(code_str) if coa else '') or ''
        for label, val in converted.items():
            key = f'{code_str}::{label}'
            lv = local_raws.get(label, 0)
            if key in result:
                result[key]['value'] += val
                result[key]['local_value'] += lv
            else:
                result[key] = {
                    'kor': f'{kor_str} / {label}' if kor_str else label,
                    'eng': eng_name,
                    'value': val,
                    'local_value': lv,
                    'compare': 0,
                    'compare_local': 0,
                }

    return result


def _extract_cf4_sheet(ws, avg_rate, spot_current, spot_prior, coa=None):
    """
    CF4 시트에서 3개 섹션을 각각 추출.
    첫 3개의 'Code' 헤더 행이 각각 취득가액/감가상각/손상차손 섹션.
    반환: {'CF4_취득가액': OrderedDict, 'CF4_감가상각': OrderedDict, 'CF4_손상차손': OrderedDict}
    """
    section_keys = ('CF4_취득가액', 'CF4_감가상각', 'CF4_손상차손')

    code_rows = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip() == 'Code':
            code_rows.append(r)

    result = {k: OrderedDict() for k in section_keys}
    for i, key in enumerate(section_keys):
        if i >= len(code_rows):
            continue
        header_row = code_rows[i]
        end_row = code_rows[i + 1] if i + 1 < len(code_rows) else ws.max_row + 1
        result[key] = _extract_cf4_section(ws, header_row, end_row,
                                           avg_rate, spot_current, spot_prior, coa=coa)
    return result


def _adjust_cf_fx_translation(cf_dict):
    """비-KRW 회사의 CF 환율변동효과(Ⅴ)를 차액으로 자동 채움.

    공식: Ⅴ = Ⅶ.기말의현금 - Ⅵ.기초의현금 - Ⅳ.현금의증감

    각 회사의 환율(spot_prior, spot_current, avg)이 서로 다르기 때문에
    KRW 환산 후 Ⅳ + Ⅵ ≠ Ⅶ 가 발생 → 그 차액이 환율변동효과의 의미.

    매칭:
      Ⅳ : 로마자 'Ⅳ'(U+2163)로 시작
      Ⅴ : 로마자 'Ⅴ'(U+2164)로 시작 + '환율' 포함
      Ⅵ : 로마자 'Ⅵ'(U+2165)로 시작
      Ⅶ : 로마자 'Ⅶ'(U+2166)로 시작

    반환: True (조정됨) / False (필요 행을 못 찾음)
    """
    iv_key = v_key = vi_key = vii_key = None
    for k, info in cf_dict.items():
        kor = (info.get('kor') or '').strip()
        if not kor:
            continue
        if iv_key is None and kor.startswith('Ⅳ'):
            iv_key = k
        elif v_key is None and kor.startswith('Ⅴ') and '환율' in kor:
            v_key = k
        elif vi_key is None and kor.startswith('Ⅵ'):
            vi_key = k
        elif vii_key is None and kor.startswith('Ⅶ'):
            vii_key = k
        if iv_key and v_key and vi_key and vii_key:
            break

    if not (iv_key and vi_key and vii_key and v_key):
        return False  # 4개 행 모두 있어야 조정 가능

    iv_val  = cf_dict[iv_key].get('value', 0) or 0
    vi_val  = cf_dict[vi_key].get('value', 0) or 0
    vii_val = cf_dict[vii_key].get('value', 0) or 0
    fx_effect = vii_val - vi_val - iv_val

    cf_dict[v_key]['value'] = fx_effect
    cf_dict[v_key]['rate_kind'] = 'auto-fx-diff'
    return True


def _extract_cf_sheet(ws, avg_rate, spot_current, spot_prior, coa=None):
    """
    CF 환산:
      · Ⅵ. 기초의현금 → 전기 Spot rate
      · Ⅶ. 기말의현금 → 당기 Spot rate
      · 그 외 모든 행  → 당기 Avg rate
    열: B(항목명) / C(원본값) / F(참조코드).

    참조코드(F열) 패턴:
      · 숫자 시작 (4500xxx, 5xxxxxx 등): 손익계정 직접 참조 (가산/차감 섹션)
      · "CF1xxxxxx" ~ "CF6xxxxxx": 자산부채변동/투자/재무 섹션 코드
        - CF1/CF2: 영업활동 자산·부채의 변동
        - CF3/CF4: 투자활동 유입/유출액
        - CF5/CF6: 재무활동 유입/유출액
    """
    result = OrderedDict()
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 2).value
        value = ws.cell(r, 3).value
        ref_code = ws.cell(r, 6).value

        if label is None:
            continue
        label_str = str(label).strip()
        if not label_str:
            continue

        v = float(value) if _is_num(value) else 0

        # 적용 환율 결정
        if '기초의현금' in label_str:
            rate = spot_prior
            rate_kind = 'spot(전기)'
        elif '기말의현금' in label_str:
            rate = spot_current
            rate_kind = 'spot(당기)'
        else:
            rate = avg_rate
            rate_kind = 'avg'

        converted = _round_krw(v * rate)

        # 코드 인식: 숫자 시작 또는 "CF"로 시작 (CF1xxx~CF6xxx)
        code_str = str(ref_code).strip() if ref_code is not None else ''
        is_code = bool(code_str) and (
            code_str[0].isdigit() or code_str.upper().startswith('CF')
        )
        key = code_str if is_code else f'LBL::{label_str}'

        if key in result:
            continue

        # 영문명: COA 시트의 Account Name (코드별) 우선, 없으면 빈문자열
        eng_name = ''
        if is_code and coa:
            eng_name = coa.get(code_str, '')

        result[key] = {
            'kor': label_str,
            'eng': eng_name,
            'value': converted,
            'local_value': v,
            'compare': 0,
            'compare_local': 0,
            'rate_kind': rate_kind,
        }
    return result


def _extract_wce_local_full(wb):
    """
    WCE 시트 1~132행(현지통화 섹션)에서 6개 테이블 전체 값 추출.

    반환: {table_id: {code: {korean_label: raw_local_value}}}
      - korean_label은 컬럼 B의 텍스트에서 괄호(영문) 부분을 제거한 값
      - 환율 미적용 — 원본 현지통화 그대로
    """
    # wce_schema 모듈에서 레이아웃 가져오기 (런타임 import — 순환 의존 방지)
    from wce_schema import WCE_LOCAL_LAYOUT

    if 'WCE' not in wb.sheetnames:
        return {tid: {} for tid in WCE_LOCAL_LAYOUT}
    ws = wb['WCE']

    # 테이블 5(이익잉여금) 회계 귀속 규칙:
    # - col 7(3500105 Current Net Income): '당기순이익'만 귀속
    #   · '보험수리적손익'은 Unappropriated R/E(3500104)에 귀속 → col 7에선 0 (수기 입력)
    #   · '지분법이익잉여금(R/E조정)'은 Unappropriated R/E(3500104)에 귀속 → col 7에선 0
    #   · 그 외 행은 col 6과 동일한 합계 → col 7로 읽으면 이중계산 → 0
    # - col 6(3500104 Unappropriated R/E): '당기순이익'은 별도 계정(3500105)에 귀속 → col 6에선 0
    #   · '보험수리적손익', '지분법이익잉여금'은 col 6에 정상 표시 (Unappropriated R/E 변동의 일부)
    T5_COL7_VALID_LABELS = {'당기순이익'}
    T5_COL6_ZERO_LABELS  = {'당기순이익'}

    result = {}
    for tid, layout in WCE_LOCAL_LAYOUT.items():
        table_data = {}
        for col_num, code in layout['cols'].items():
            cell_data = {}
            for r in layout['data_rows']:
                label = ws.cell(r, 2).value
                if not label:
                    continue
                # 한글 라벨만 추출 (괄호 안 영문 제거)
                label_str = str(label).split('(')[0].strip()
                if not label_str:
                    continue
                v = ws.cell(r, col_num).value
                val = float(v) if _is_num(v) else 0.0
                # 테이블 5, col 7(3500105): 위 화이트리스트 외 모두 0
                if tid == '5' and code == '3500105' and label_str not in T5_COL7_VALID_LABELS:
                    val = 0.0
                # 테이블 5, col 6(3500104): 당기순이익은 0 (Current Net Income으로 귀속됨)
                if tid == '5' and code == '3500104' and label_str in T5_COL6_ZERO_LABELS:
                    val = 0.0
                # 같은 라벨이 여러 행에 등장할 수 있음 (Excel 템플릿 잔여 행)
                # → 의미있는(non-zero) 값을 우선; 0/빈 값으로 덮어쓰지 않음
                existing = cell_data.get(label_str)
                if existing is None:
                    cell_data[label_str] = val
                elif existing == 0 and val != 0:
                    cell_data[label_str] = val
                # else: 기존값 유지 (non-zero existing은 zero/empty로 덮어쓰지 않음)
            table_data[code] = cell_data
        result[tid] = table_data

    return result


def _extract_wce_local_re(wb, avg_rate):
    """
    WCE 시트 1~132행(현지통화 섹션) 중 5번 이익잉여금 테이블의
    '당기순이익', '보험수리적손익', 'R/E조정' 행을 코드별로 추출하여 KRW 환산.

    구조:
      R81 col 3~7 = 코드 (3500101 Legal / 3500102 Voluntary / 3500103 Actuarial /
                          3500104 Unappropriated / 3500105 Current Net Income/Total)
      R84 = 당기순이익(Profit for the Year)
      R85 = 보험수리적손익(Remeasurements)
      R86 = 지분법이익잉여금(R/E Adjustment Of Equity Method)

    반환: {row_key: {code: krw_value}}
    """
    # 자동 KRW 환산 대상: 당기순이익만
    # (보험수리적손익·지분법이익잉여금은 수기 입력)
    result = {'당기순이익': {}}
    if 'WCE' not in wb.sheetnames:
        return result
    ws = wb['WCE']

    codes = ['3500101', '3500102', '3500103', '3500104', '3500105']
    row_map = {'당기순이익': 84}
    row_target = {'당기순이익': '3500105'}

    for row_key, r in row_map.items():
        target = row_target[row_key]
        primary_col, fallback_col = 7, 6
        vp = ws.cell(r, primary_col).value
        total_local = float(vp) if _is_num(vp) else 0.0
        if not total_local:
            vf = ws.cell(r, fallback_col).value
            total_local = float(vf) if _is_num(vf) else 0.0
        for code in codes:
            result[row_key][code] = (total_local * (avg_rate or 0)) if code == target else 0.0

    return result


def extract(file_path, central_rates=None, central_rates_lookup=None):
    """
    패키지 파일에서 BS/PL/CF/WCE 등 모든 데이터를 추출.

    중앙 환율 우선순위:
      1) central_rates (dict) — 직접 지정
      2) central_rates_lookup(currency) — 통화 추출 후 호출되는 콜러블
      3) None → 패키지 환율만 사용 (기존 동작)

    central_rates 형태: {'avg', 'spot_current', 'spot_prior', 'avg_prior'}
      값이 None인 키는 패키지 환율로 폴백.

    환산 적용:
      - BS: 현지 F열 × spot_current
      - PL_MF: 현지 E열 × avg
      - CF/CF1-4: avg/spot_current/spot_prior 매개변수 그대로 사용
      - PY-BS: 패키지 KRW × (사용 spot_prior / 패키지 spot_prior) 비율 보정
      - PY-PL: 현지 × 사용 avg_prior
      - WCE local RE: avg
    """
    # xlsm 파일은 VBA 포함 가능 → keep_vba=True
    # 값만 필요하므로 data_only=True로 캐시된 계산 결과 사용
    try:
        wb = load_workbook(file_path, keep_vba=True, data_only=True)
    except Exception as e:
        raise RuntimeError(f'엑셀 파일을 열 수 없습니다 ({type(e).__name__}): {e}')

    try:
        company = _get_company_name(wb) or str(file_path).replace('\\', '/').split('/')[-1]
        currency = _get_currency(wb) or 'KRW'
        yq = _get_year_quarter(wb)
        index_error_count = _get_index_error_count(wb)
        tables = _get_fx_rate_tables(wb)

        # 1) 패키지 환율 (Master 시트에서 추출한 원본)
        if currency == 'KRW':
            pkg_avg = pkg_spot_current = pkg_spot_prior = pkg_avg_prior = 1.0
        else:
            cur = tables['current'].get(currency, {})
            pri = tables['prior'].get(currency, {})
            pkg_avg          = cur.get('avg')  or 1.0
            pkg_spot_current = cur.get('spot') or 1.0
            pkg_spot_prior   = pri.get('spot') or pkg_spot_current
            pkg_avg_prior    = pri.get('avg')  or pkg_avg

        # 2) 중앙 환율 결정 — dict 직접 지정 > lookup 호출 > 없음
        if central_rates is None and central_rates_lookup is not None and currency != 'KRW':
            try:
                central_rates = central_rates_lookup(currency)
            except Exception:
                central_rates = None

        # 3) 사용할 환율 결정 (중앙 관리 환율 우선, 없으면 패키지 환율)
        use_central = bool(central_rates) and currency != 'KRW'
        if use_central:
            avg_rate       = central_rates.get('avg')          or pkg_avg
            spot_current   = central_rates.get('spot_current') or pkg_spot_current
            spot_prior     = central_rates.get('spot_prior')   or pkg_spot_prior
            avg_rate_prior = central_rates.get('avg_prior')    or pkg_avg_prior
        else:
            avg_rate       = pkg_avg
            spot_current   = pkg_spot_current
            spot_prior     = pkg_spot_prior
            avg_rate_prior = pkg_avg_prior

        # 3) PY-BS 재조정 비율: 사용 spot_prior / 패키지 spot_prior
        bs_py_ratio = (spot_prior / pkg_spot_prior) if pkg_spot_prior else 1.0

        # PY 시트에서 전기 비교값 추출 (BS는 비율 보정, PL은 사용 prior_avg)
        # spot_prior 함께 전달 → BS 로컬 raw 역산용
        py_compare = _extract_py_compare(wb, avg_rate_prior,
                                         bs_rescale_ratio=bs_py_ratio,
                                         spot_prior=spot_prior)

        # COA 시트에서 영문명 매핑 로드 (CF 등에서 fallback으로 사용)
        coa = _load_coa_eng(wb)

        sheets = {}
        for sheet_name, cfg in SHEET_CONFIG.items():
            if sheet_name in wb.sheetnames:
                # 사용 환율 모드면 시트별 rate_type에 맞는 환율로 재환산
                rate = None
                if currency != 'KRW':
                    rt = cfg.get('rate_type')
                    if rt == 'spot_current':
                        rate = spot_current
                    elif rt == 'avg':
                        rate = avg_rate
                data = _extract_coded_sheet(wb[sheet_name], cfg, rate=rate)
                # 비교값 주입 (KRW + 로컬 raw)
                cmp_map = py_compare['bs'] if sheet_name == 'BS' else py_compare['pl']
                cmp_local_map = py_compare['bs_local'] if sheet_name == 'BS' else py_compare['pl_local']
                cmp_pkg_raw_map = py_compare.get('bs_pkg_raw', {}) if sheet_name == 'BS' else {}
                # BS 전기 영역(14~197행 H열) 별도 추출
                bs_prior_local_map = (
                    _extract_bs_prior_local(wb[sheet_name]) if sheet_name == 'BS' else {}
                )
                for code, entry in data.items():
                    entry['compare'] = cmp_map.get(code, 0)
                    if sheet_name == 'BS':
                        # BS 전기 로컬: 시트 14~197행 H열 우선, 없으면 PY 역산값으로 폴백
                        direct = bs_prior_local_map.get(code)
                        if direct is not None:
                            entry['compare_local'] = direct
                        else:
                            entry['compare_local'] = cmp_local_map.get(code, 0)
                        entry['compare_pkg_raw'] = cmp_pkg_raw_map.get(code, 0)
                    else:
                        # PL: PY 시트 H열 raw 그대로
                        entry['compare_local'] = cmp_local_map.get(code, 0)
                    # eng가 비어있으면 COA에서 보완
                    if not entry.get('eng') and coa:
                        entry['eng'] = coa.get(code, '') or entry.get('eng', '')
                sheets[sheet_name] = data
            else:
                sheets[sheet_name] = OrderedDict()

        # BS 합계/소계 계정은 개별 환산이 아니라 하위계정 환산값의 합으로 맞춘다
        # (외화 회사에서 개별 반올림으로 합계≠하위합 단수차이가 생기는 것 방지)
        # 이어서 남는 대차 반올림 잔단은 해외사업환산손익으로 흡수 → 대차 0
        if currency != 'KRW' and sheets.get('BS'):
            _recompute_bs_subtotals(sheets['BS'])
            _balance_bs_rounding(sheets['BS'])

        if 'CF' in wb.sheetnames:
            sheets['CF'] = _extract_cf_sheet(wb['CF'], avg_rate, spot_current, spot_prior, coa=coa)
            # 비-KRW 회사: 환율변동효과(Ⅴ) = Ⅶ.기말의현금 − Ⅵ.기초의현금 − Ⅳ.현금의증감
            # 으로 자동 보정 → CF가 정확히 균형 맞도록
            if currency != 'KRW':
                _adjust_cf_fx_translation(sheets['CF'])
        else:
            sheets['CF'] = OrderedDict()

        for cf_sheet in ('CF1', 'CF2', 'CF3'):
            if cf_sheet in wb.sheetnames:
                ws_cf = wb[cf_sheet]
                # E열(5)=연결범위회사, F열(6)=제3자
                sheets[f'{cf_sheet}_연결'] = _extract_cf1_sheet(ws_cf, avg_rate, spot_current, spot_prior, value_col=5, coa=coa)
                sheets[f'{cf_sheet}_제3자'] = _extract_cf1_sheet(ws_cf, avg_rate, spot_current, spot_prior, value_col=6, coa=coa)
            else:
                sheets[f'{cf_sheet}_연결'] = OrderedDict()
                sheets[f'{cf_sheet}_제3자'] = OrderedDict()

        if 'CF4' in wb.sheetnames:
            cf4_sections = _extract_cf4_sheet(wb['CF4'], avg_rate, spot_current, spot_prior, coa=coa)
            for key, data in cf4_sections.items():
                sheets[key] = data
        else:
            for key in ('CF4_취득가액', 'CF4_감가상각', 'CF4_손상차손'):
                sheets[key] = OrderedDict()

        # WCE 로컬 섹션 - 5번 이익잉여금 RE 자동값 (당기 avg rate 환산)
        wce_local_re = _extract_wce_local_re(wb, avg_rate)
        # WCE 로컬 섹션 전체 (1~132행, 환율 미적용 raw 값)
        wce_local_full = _extract_wce_local_full(wb)

        return {
            'company': company,
            'currency': currency,
            'cover_year': yq['year'],
            'cover_quarter': yq['quarter'],
            'index_error_count': index_error_count,
            'fx_rate': avg_rate,                 # 하위호환(주 평균환율)
            'fx_avg': avg_rate,
            'fx_spot_current': spot_current,
            'fx_spot_prior': spot_prior,
            'fx_avg_prior': avg_rate_prior,
            'fx_source': 'central' if use_central else 'package',
            'fx_package': {                      # 패키지 원본 환율 (참고용)
                'avg':          pkg_avg,
                'spot_current': pkg_spot_current,
                'spot_prior':   pkg_spot_prior,
                'avg_prior':    pkg_avg_prior,
            } if currency != 'KRW' else None,
            'file': str(file_path),
            'sheets': sheets,
            'wce_local_re': wce_local_re,
            'wce_local_full': wce_local_full,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


# ─── 부호·정합성 검증 (업로드 후 경고용) ─────────────────────────────────────

def validate_local_vs_value_signs(extracted: dict,
                                  sheets_to_check=('BS', 'PL_MF'),
                                  krw_amount_tolerance: float = 1.0) -> list[dict]:
    """추출 결과에서 local_value(현지통화 raw)와 value(KRW 환산값)의 정합성을 검사.

    검출 케이스(모두 사용자 입력·수식 오류 후보):
      1) 부호 반전: 둘 다 0이 아닌데 sign(local) != sign(value)
      2) 한쪽만 0: 한쪽만 non-zero (일반적으로 입력 누락 또는 수식 오류)
      3) KRW 회사 절댓값 불일치: 환산 안 하는데도 local ≠ value (tolerance 1원 초과)

    업로드 차단은 아니고 단순 경고용. KRW 회사 / 외화 회사 모두 잡는다.

    반환:
      [
        {
          'sheet':       'BS' or 'PL_MF',
          'code':        '2299901',
          'kor':         '기타비유동부채',
          'local_value': 0,
          'value':       4567098467,
          'kind':        'one_side_zero' | 'sign_flip' | 'krw_amount_mismatch',
          'detail':      사람이 읽을 수 있는 짧은 메시지,
        },
        ...
      ]
    """
    out = []
    currency = (extracted.get('currency') or 'KRW').upper()
    is_krw = (currency == 'KRW')
    sheets = extracted.get('sheets') or {}

    for sheet_name in sheets_to_check:
        rows = sheets.get(sheet_name) or {}
        for code, info in rows.items():
            if not isinstance(info, dict):
                continue
            # BS 자본항목(3xxxxxx)은 패키지 셀이 잠겨있고 WCE 시트 입력값으로
            # 덮어써지므로 자회사 측에서 수정 불가 → 경고 대상에서 제외
            if sheet_name == 'BS' and str(code).startswith('3'):
                continue
            # 소계/공식 코드는 패키지 BS/PL의 셀에 직접 입력되지 않고
            # 수식이거나 합산 결과이므로 부호 검증의 대상이 다름. 다만 어차피
            # 잘못된 수식이면 표시되므로 일단 모두 검사한다.
            lv = info.get('local_value', None)
            v  = info.get('value', None)
            try:
                lv_f = float(lv) if lv is not None else 0.0
                v_f  = float(v)  if v  is not None else 0.0
            except (TypeError, ValueError):
                continue

            # 둘 다 0이면 검사 대상 아님
            if lv_f == 0 and v_f == 0:
                continue

            kor = info.get('kor', '') or ''

            # 한쪽만 0인 경우
            if (lv_f == 0) != (v_f == 0):  # XOR
                out.append({
                    'sheet': sheet_name,
                    'code': str(code),
                    'kor': kor,
                    'local_value': lv_f,
                    'value': v_f,
                    'kind': 'one_side_zero',
                    'detail': (f'{"현지통화" if lv_f == 0 else "환산값(KRW)"}이 0인데 '
                               f'{"환산값(KRW)" if lv_f == 0 else "현지통화"}만 입력됨'),
                })
                continue

            # 둘 다 non-zero — 부호 비교
            if (lv_f > 0) != (v_f > 0):  # 부호 반전
                out.append({
                    'sheet': sheet_name,
                    'code': str(code),
                    'kor': kor,
                    'local_value': lv_f,
                    'value': v_f,
                    'kind': 'sign_flip',
                    'detail': f'현지통화({lv_f:+,.0f})와 환산값({v_f:+,.0f}) 부호 반대',
                })
                continue

            # KRW 회사면 절댓값까지 같아야 함 (환율 1.0이므로)
            if is_krw and abs(lv_f - v_f) > krw_amount_tolerance:
                out.append({
                    'sheet': sheet_name,
                    'code': str(code),
                    'kor': kor,
                    'local_value': lv_f,
                    'value': v_f,
                    'kind': 'krw_amount_mismatch',
                    'detail': (f'KRW 회사인데 현지통화({lv_f:,.0f})와 환산값({v_f:,.0f})이 '
                               f'다름 (차이 {v_f - lv_f:+,.0f})'),
                })

    return out
