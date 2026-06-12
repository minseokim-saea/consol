"""
PY 시트 J~T 영역과 PL_MF 145행~, BS F열, CF1~3 시트 구조 정확히 확인.
- 빈 템플릿: distribute_template/2026.xlsm
- 소스 패키지: 글로벌세아 2025-4Q 패키지
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ─── 1) 빈 템플릿 PY 시트 모든 컬럼 헤더 확인 ───────────────────────
print("=" * 70)
print("① 빈 템플릿 PY 시트 1~6행 모든 열")
print("=" * 70)
tpl = Path("distribute_template/2026.xlsm")
wb = load_workbook(str(tpl), keep_vba=True, data_only=False)
py = wb["PY"]
print(f"PY max_row={py.max_row}, max_col={py.max_column}")
for r in range(1, 7):
    cells = []
    for c in range(1, py.max_column + 1):
        v = py.cell(r, c).value
        if v is not None:
            cells.append(f"{get_column_letter(c)}({c})={v!r}")
    if cells:
        print(f"  row {r:2d}: " + " | ".join(cells))

# 빈 템플릿 PY 시트의 형식·수식이 있을 수 있으니 J열~T열을 4~24행 정도 확인
print("\n--- PY 시트 J~U열 4~24행 (수식 포함) ---")
for r in range(4, 25):
    cells = []
    for c in range(10, 22):   # J(10) ~ U(21)
        cell = py.cell(r, c)
        v = cell.value
        if v is not None:
            cells.append(f"{get_column_letter(c)}={v!r}")
    if cells:
        print(f"  row {r:2d}: " + " | ".join(cells))
wb.close()

# ─── 2) 소스 패키지: PL_MF 145행~, BS F열, CF1~3 ─────────────────
print("\n" + "=" * 70)
print("② 글로벌세아 2025-4Q 패키지 — PL_MF 145행~, BS F열, CF1~3")
print("=" * 70)
state = json.loads(Path("uploads/_state.json").read_text(encoding="utf-8"))
src = next((f for f in state if f.get("year") == "2025-4Q" and "글로벌세아" in (f.get("company") or "")), None)
if not src:
    print("[FAIL] 소스 패키지 없음")
    exit(1)

print(f"src: {src['path']}")
wb = load_workbook(src["path"], data_only=False)
print(f"Sheets: {wb.sheetnames[:8]}...")

# PL_MF 145~165행 (구조 파악)
print("\n--- PL_MF 시트 140~165행 (헤더/데이터) ---")
if "PL_MF" in wb.sheetnames:
    ws = wb["PL_MF"]
    print(f"max_row={ws.max_row}, max_col={ws.max_column}")
    for r in list(range(140, min(166, ws.max_row + 1))):
        cells = []
        for c in range(1, min(ws.max_column + 1, 16)):
            v = ws.cell(r, c).value
            if v is not None:
                cells.append(f"{get_column_letter(c)}={v!r}")
        if cells:
            print(f"  row {r:3d}: " + " | ".join(cells)[:200])

# BS F열 (GAAP 차이) — 헤더 확인 + 207행~ 샘플
print("\n--- BS 시트 1~10행, 207~210행 (모든 열 헤더/데이터) ---")
if "BS" in wb.sheetnames:
    ws = wb["BS"]
    print(f"max_row={ws.max_row}, max_col={ws.max_column}")
    for r in [1, 2, 3, 4, 5, 6, 7, 8, 207, 208, 209, 210]:
        if r > ws.max_row:
            continue
        cells = []
        for c in range(1, min(ws.max_column + 1, 12)):
            v = ws.cell(r, c).value
            if v is not None:
                cells.append(f"{get_column_letter(c)}={v!r}")
        if cells:
            print(f"  row {r:3d}: " + " | ".join(cells)[:200])

# CF1~3 시트 T열 수식 + 헤더
for sn in ["CF1", "CF2", "CF3"]:
    print(f"\n--- {sn} 시트 1~10행 + T열 수식 (15~30행) ---")
    if sn not in wb.sheetnames:
        print(f"  (시트 없음)")
        continue
    ws = wb[sn]
    print(f"max_row={ws.max_row}, max_col={ws.max_column}")
    for r in range(1, 11):
        cells = []
        for c in range(1, min(ws.max_column + 1, 22)):
            v = ws.cell(r, c).value
            if v is not None:
                cells.append(f"{get_column_letter(c)}={v!r}")
        if cells:
            print(f"  row {r:3d}: " + " | ".join(cells)[:200])
    # T열 수식
    print(f"  T열 수식 샘플 (20=T):")
    for r in range(2, min(40, ws.max_row + 1)):
        v = ws.cell(r, 20).value
        if v is not None and str(v).startswith("="):
            print(f"    T{r}: {v!r}")
            if r > 20:
                break

wb.close()
print("\n[OK] 점검 완료")
