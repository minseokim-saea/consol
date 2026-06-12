"""
신규 회사 시나리오 검증:
  1) 이전 dry-run의 _test_full/글로벌세아.xlsm 을 '신규회사_TEST'의 2026-1Q 패키지로 위장
  2) state에 그 항목만 추가하고 전년 4Q 항목은 없도록 함
  3) resolve_py_data(2026-2Q)가 신규 회사 경로로 진입하여 1Q PY × 2 적용 검증
"""
import sys, json, copy
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import distribute_builder as db
from openpyxl import load_workbook

# 위장 state — 신규회사_TEST의 2026-1Q만 존재
fake_1q_pkg = Path("results/distribute/_test_full/글로벌세아.xlsm")
if not fake_1q_pkg.exists():
    print("[FAIL] dry-run 결과 파일이 없음. _verify_full_py.py 먼저 실행")
    sys.exit(1)

NEW_CO = "신규회사_TEST"
fake_state = [
    {
        'id': 'fake_1q',
        'company': NEW_CO,
        'year': '2026-1Q',
        'path': str(fake_1q_pkg),
        'uploaded_at': '2026-04-01 09:00:00',
    }
]
print(f"가짜 state: 1개 항목 (신규회사_TEST 2026-1Q = {fake_1q_pkg.name})\n")

# Case 1: 2026 1Q 자체 (신규 회사) — 빈 PY로 진입
print("=" * 60)
print("Case 1: 2026 1Q (신규 회사 첫 분기 — 빈 PY)")
print("=" * 60)
res = db.resolve_py_data(fake_state, NEW_CO, 2026, 1)
print(f"ok={res['ok']}, is_new_company={res.get('is_new_company')}")
print(f"source_bs={res['source_bs']}")
print(f"source_pl={res['source_pl']}")
print(f"bs={len(res['bs'])}, pl={len(res['pl'])}")

# Case 2: 2026 2Q — 1Q PY × 2 적용
print("\n" + "=" * 60)
print("Case 2: 2026 2Q (신규 회사 — 1Q PY × 2)")
print("=" * 60)
res = db.resolve_py_data(fake_state, NEW_CO, 2026, 2)
print(f"ok={res['ok']}, is_new_company={res.get('is_new_company')}")
print(f"source_bs={res['source_bs']}")
print(f"source_pl={res['source_pl']}")
print(f"pl_scale={res['pl_scale']}")
print(f"bs={len(res['bs'])}, bs_local={len(res['bs_local'])}, "
      f"pl={len(res['pl'])}, pl_mf_secondary={len(res['pl_mf_secondary'])}, "
      f"n_by_row={len(res.get('py_n_by_row') or {})}, s_by_row={len(res.get('py_s_by_row') or {})}")

# 1Q 패키지에서 직접 추출한 값과 비교 (×2 확인)
py1 = db.extract_py_sheet(fake_1q_pkg)
print("\n--- PL 샘플 (1Q vs 2Q resolved, ratio) ---")
for code in list(py1['pl'].keys())[:5]:
    v1 = py1['pl'][code]
    v2 = res['pl'].get(code, 0)
    ratio = v2 / v1 if v1 else 0
    print(f"  {code}: 1Q={v1:>15,.2f}  →  2Q={v2:>15,.2f}  (×{ratio:.2f})")

print("\n--- BS 샘플 (1Q vs 2Q resolved, 그대로) ---")
for code in list(py1['bs'].keys())[:5]:
    v1 = py1['bs'][code]
    v2 = res['bs'].get(code, 0)
    same = "OK" if abs(v1 - v2) < 0.5 else "DIFF"
    print(f"  {code}: 1Q={v1:>18,.0f}  →  2Q={v2:>18,.0f}  [{same}]")

# Case 3: 2026 3Q — ×3 확인
print("\n" + "=" * 60)
print("Case 3: 2026 3Q (신규 회사 — ×3)")
print("=" * 60)
res = db.resolve_py_data(fake_state, NEW_CO, 2026, 3)
print(f"pl_scale={res['pl_scale']}")
sample = next(iter(res['pl'].items()))
print(f"PL 샘플: {sample[0]} = {sample[1]:,.0f} "
      f"(1Q {py1['pl'][sample[0]]:,.0f} × 3 = {py1['pl'][sample[0]]*3:,.0f})")

# Case 4: build_distribution_package로 실제 파일 생성 (2Q 신규)
print("\n" + "=" * 60)
print("Case 4: 실제 파일 생성 (신규회사_TEST 2026-2Q)")
print("=" * 60)
out_dir = Path("results/distribute/_test_newco")
out_dir.mkdir(parents=True, exist_ok=True)
for p in out_dir.glob("*.xlsm"):
    p.unlink()
tpl = db.get_template_path("2026")
r = db.build_distribution_package(
    template_path=tpl, output_dir=out_dir, company=NEW_CO,
    target_year=2026, target_quarter=2,
    uploaded_files=fake_state, file_password=None, sheet_protect_password=None,
)
print(f"ok={r['ok']}")
print(f"wrote_bs={r['wrote_bs']}, wrote_pl={r['wrote_pl']}, "
      f"wrote_pl_mf_secondary={r.get('wrote_pl_mf_secondary')}, "
      f"wrote_n_refs={r.get('wrote_n_refs')}, wrote_s_refs={r.get('wrote_s_refs')}")
print(f"output: {r['output_path']}")

# 검증: 생성 파일의 PY!E5 (BS) ↔ 1Q의 PY!E5
wb = load_workbook(r['output_path'], keep_vba=True, data_only=False)
py = wb['PY']
print("\n--- 생성 파일 PY 시트 5~10행 (D/E/G/H 기준, 1Q 값과 비교) ---")
wb1 = load_workbook(str(fake_1q_pkg), keep_vba=True, data_only=False)
py1ws = wb1['PY']
for row in range(5, 11):
    d2 = py.cell(row, 4).value; e2 = py.cell(row, 5).value
    g2 = py.cell(row, 7).value; h2 = py.cell(row, 8).value
    d1 = py1ws.cell(row, 4).value; e1 = py1ws.cell(row, 5).value
    g1 = py1ws.cell(row, 7).value; h1 = py1ws.cell(row, 8).value
    print(f"  row {row}: "
          f"D={d2} E1Q={e1} →E2Q={e2} ({'OK' if e1==e2 else 'DIFF'}) | "
          f"G={g2} H1Q={h1} →H2Q={h2} (ratio={h2/h1 if h1 else 'NA'})")
wb1.close(); wb.close()
print("\n[OK] 신규 회사 시나리오 검증 완료")
