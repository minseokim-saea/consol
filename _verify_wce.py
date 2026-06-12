"""WCE 자본 적용 검증 — 패키지 자본 vs WCE 자본 비교"""
import sys, json
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import distribute_builder as db

state = json.loads(Path("uploads/_state.json").read_text(encoding="utf-8"))

# 1) 글로벌세아 2025-4Q WCE 데이터에서 자본 기말잔액 계산
wce_all = db._load_wce_overrides()
wce_rec = wce_all.get("2025-4Q::글로벌세아")
print(f"WCE record exists: {wce_rec is not None}")

if wce_rec:
    wce_bs = db.compute_wce_equity_bs(wce_rec)
    print(f"\n=== WCE 기반 자본 BS 코드 (총 {len(wce_bs)}개) ===")
    for code, val in sorted(wce_bs.items()):
        print(f"  {code}: {val:>20,.0f}")

# 2) 패키지에서 추출한 BS 자본부와 비교
from pathlib import Path as P
src_q4 = next(f for f in state if f.get("year") == "2025-4Q" and "글로벌세아" in (f.get("company") or ""))
pkg = db.extract_bs_pl_from_package(P(src_q4["path"]))
print(f"\n=== 패키지 BS 자본 코드 (3으로 시작) ===")
pkg_equity = {c: v for c, v in pkg["bs"].items() if c.startswith("3")}
for code, val in sorted(pkg_equity.items()):
    print(f"  {code}: {val:>20,.0f}")

# 3) resolve_py_data 호출 후 비교
print("\n=== resolve_py_data (2026-1Q, 글로벌세아) ===")
res = db.resolve_py_data(state, "글로벌세아", 2026, 1)
print(f"wce_overrides_applied: {res.get('wce_overrides_applied')}")
print(f"source_bs: {res['source_bs']}")

print(f"\n=== resolve_py_data BS 자본 코드 (덮어써짐) ===")
out_equity = {c: v for c, v in res["bs"].items() if c.startswith("3")}
for code in sorted(out_equity):
    pkg_v = pkg_equity.get(code, 0)
    out_v = out_equity[code]
    diff = "(WCE 덮어씀)" if abs(out_v - pkg_v) > 0.5 else "(동일)"
    print(f"  {code}: pkg={pkg_v:>18,.0f}  →  out={out_v:>18,.0f}  {diff}")

# 4) 최종 파일 생성하여 PY!D5+ 자본 코드 확인
print("\n=== 최종 파일 생성 ===")
out_dir = Path("results/distribute/_test_wce")
out_dir.mkdir(parents=True, exist_ok=True)
for p in out_dir.glob("*.xlsm"):
    p.unlink()
tpl = db.get_template_path("2026")
r = db.build_distribution_package(
    template_path=tpl, output_dir=out_dir, company="글로벌세아",
    target_year=2026, target_quarter=1,
    uploaded_files=state, file_password=None, sheet_protect_password=None,
)
print(f"ok={r['ok']}, wrote_bs={r['wrote_bs']}, output={r['output_path']}")

from openpyxl import load_workbook
wb = load_workbook(r["output_path"], keep_vba=True, data_only=False)
py = wb["PY"]
print("\n=== PY!D/E 자본부 일부 (3xxxxxx) ===")
for row in range(5, py.max_row+1):
    code = py.cell(row, 4).value
    amt = py.cell(row, 5).value
    if isinstance(code, str) and code.startswith("3"):
        marker = "  *" if abs(amt - pkg_equity.get(code, 0)) > 0.5 else "   "
        print(f"{marker}row {row:3d}: D={code}, E={amt:>20,.0f}")
wb.close()
print("\n[* = WCE 값으로 덮어써진 코드]")
