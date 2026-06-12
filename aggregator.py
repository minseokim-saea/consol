"""
환산값(KRW) 기준 회사별 합산.
회사별 1열씩 + 마지막 합계 열.
"""

import re
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime

# CF 시트에서 "현금유출 없는 비용 등의 가산" 섹션으로 모아야 할 비용 코드 패턴
# - 4500xxx: 영업외비용류 (외화환산손실, 평가손실, 처분손실 등)
# - 5xxxxxx: 판매관리비/제조원가류 (소모품비, 감가상각비 등)
CF_EXPENSE_CODE_RE = re.compile(r'^(4500\d{3}|5\d{6})$')

SHEET_DISPLAY = OrderedDict([
    ('BS', '대차대조표 (환산 KRW)'),
    ('PL_MF', '손익계산서 (환산 KRW)'),
    ('CF', '현금흐름표 (환산 KRW)'),
    ('CF1_연결', 'CF1_연결범위회사 (환산 KRW)'),
    ('CF1_제3자', 'CF1_제3자 (환산 KRW)'),
    ('CF2_연결', 'CF2_연결범위회사 (환산 KRW)'),
    ('CF2_제3자', 'CF2_제3자 (환산 KRW)'),
    ('CF3_연결', 'CF3_연결범위회사 (환산 KRW)'),
    ('CF3_제3자', 'CF3_제3자 (환산 KRW)'),
    ('CF4_취득가액', 'CF4_취득가액변동 (환산 KRW)'),
    ('CF4_감가상각', 'CF4_감가상각누계액변동 (환산 KRW)'),
    ('CF4_손상차손', 'CF4_손상차손누계액변동 (환산 KRW)'),
])

# 출력 시 '계정코드' 열에 표시할 값 정리용
def _display_code(key):
    if isinstance(key, str) and key.startswith('LBL::'):
        return ''  # 라벨 기반 행은 코드 열 비움
    if isinstance(key, str) and '::' in key:
        return key.split('::', 1)[0]  # CF1: '{code}::{label}' → code
    return key

HDR_FILL = PatternFill('solid', start_color='1F3864')
HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
DATA_FONT = Font(name='Arial', size=10)
TOTAL_FILL = PatternFill('solid', start_color='FFE699')
TOTAL_FONT = Font(bold=True, name='Arial', size=10, color='9C5700')
NUM_FMT = '#,##0;(#,##0);"-"'


def _reorder_cf_expense_codes(code_order):
    """CF 시트에서 4500xxx/5xxxxx 비용 코드를 '2. 현금유출 없는 비용 등의 가산' 섹션 끝으로 재배치.

    원본 패키지 양식상 비용 코드는 이미 가산 섹션에 위치하지만, 일부 회사만 보유한 코드는
    합산 시 첫 등장 순서 정렬 때문에 다른 위치(보통 맨 뒤)에 떠버림.
    이 함수는 그런 흩어진 코드들을 정상 위치로 끌어옴.

    반환: 새 OrderedDict (가산 섹션 식별 실패 또는 이동 대상 없으면 입력 그대로).
    """
    keys = list(code_order.keys())

    # 섹션 경계 탐색
    # 실제 라벨 예: "2. 현금의 유출이 없는 비용등의 가산" / "3. 현금의 유입이 없는 수익등의 차감"
    # → '유출'+'가산' / '유입'+'차감' 조합으로 매칭 (띄어쓰기·표현 변형 흡수)
    sec2_idx = None       # 가산 섹션 헤더 위치
    sec_after_idx = None  # 그 다음 섹션 (차감 / 자산부채변동 / 투자활동 등) 위치
    for i, k in enumerate(keys):
        if not isinstance(k, str) or not k.startswith('LBL::'):
            continue
        label = k[5:].strip()
        if sec2_idx is None and ('유출' in label) and ('가산' in label):
            sec2_idx = i
            continue
        if sec2_idx is not None:
            if ((('유입' in label) and ('차감' in label))
                    or ('자산부채' in label)
                    or ('투자활동' in label) or ('재무활동' in label)
                    or label.startswith('Ⅱ') or label.startswith('Ⅲ')
                    or label.startswith('II.') or label.startswith('III.')):
                sec_after_idx = i
                break

    if sec2_idx is None:
        return code_order  # 가산 섹션 헤더가 없음 → 그대로

    if sec_after_idx is None:
        sec_after_idx = len(keys)

    # 이동 대상: 4500xxx/5xxxxx인데 가산 섹션 밖에 있는 코드
    in_section_2 = set(keys[sec2_idx + 1:sec_after_idx])
    to_move = sorted(
        k for k in keys
        if isinstance(k, str) and CF_EXPENSE_CODE_RE.match(k) and k not in in_section_2
    )

    if not to_move:
        return code_order

    # 새 순서 구축: 이동 대상 제거 후 '다음 섹션' 직전에 정렬된 순으로 삽입
    move_set = set(to_move)
    new_keys = [k for k in keys if k not in move_set]

    if sec_after_idx < len(keys):
        anchor = keys[sec_after_idx]
        try:
            insert_idx = new_keys.index(anchor)
        except ValueError:
            insert_idx = len(new_keys)
    else:
        insert_idx = len(new_keys)

    for c in reversed(to_move):
        new_keys.insert(insert_idx, c)

    return OrderedDict((k, True) for k in new_keys)


def aggregate(extracted_list):
    """
    반환:
    {
      'companies': [회사1, 회사2, ...],
      'sheets': {
        'BS': OrderedDict[code → {'kor', 'eng', 'by_company': {회사: num}, 'total': num}]
      }
    }
    """
    companies = [e['company'] for e in extracted_list]
    sheets_agg = {}

    for sheet_name in SHEET_DISPLAY:
        code_order = OrderedDict()
        meta = {}

        for e in extracted_list:
            for code, info in e['sheets'].get(sheet_name, {}).items():
                code_order[code] = True
                if code not in meta:
                    meta[code] = (info['kor'], info['eng'])

        # CF 시트만: 4500xxx/5xxxxx 비용 코드를 "현금유출 없는 비용 등의 가산" 섹션으로 재배치
        if sheet_name == 'CF':
            code_order = _reorder_cf_expense_codes(code_order)

        result = OrderedDict()
        for code in code_order:
            kor, eng = meta[code]
            by_co = {}
            cmp_by_co = {}
            total = 0.0
            compare_total = 0.0
            for e in extracted_list:
                info = e['sheets'].get(sheet_name, {}).get(code, {})
                v = info.get('value', 0) or 0
                by_co[e['company']] = v
                total += v
                cmp_v = info.get('compare', 0) or 0
                cmp_by_co[e['company']] = cmp_v
                compare_total += cmp_v
            result[code] = {
                'kor': kor, 'eng': eng,
                'by_company': by_co, 'total': total,
                'compare_by_company': cmp_by_co,
                'compare_total': compare_total,
            }
        sheets_agg[sheet_name] = result

    return {'companies': companies, 'sheets': sheets_agg}


def _write_sheet(wb, sheet_name, companies, sheet_data):
    if not sheet_data:
        return
    ws = wb.create_sheet(f"{sheet_name} ({SHEET_DISPLAY[sheet_name].split('(')[0].strip()})")

    # 헤더
    headers = ['계정코드', '한글명', '영문명'] + list(companies) + ['합   계']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    total_col = len(headers)

    # 데이터
    row_idx = 2
    for code, info in sheet_data.items():
        ws.cell(row_idx, 1, _display_code(code)).font = DATA_FONT
        ws.cell(row_idx, 2, info['kor']).font = DATA_FONT
        ws.cell(row_idx, 3, info['eng']).font = DATA_FONT

        for i, company in enumerate(companies, 4):
            v = info['by_company'].get(company, 0)
            cell = ws.cell(row_idx, i, v if v != 0 else None)
            cell.font = DATA_FONT
            cell.number_format = NUM_FMT

        t = info['total']
        cell = ws.cell(row_idx, total_col, t if t != 0 else None)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.number_format = NUM_FMT
        row_idx += 1

    # 서식
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 32
    for c in range(4, total_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'D2'


def write_excel(agg_result, extracted_list, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    companies = agg_result['companies']

    # 요약 시트
    ws = wb.create_sheet('요약', 0)
    ws['A1'] = '연결 재무보고 통합 결과 (환산 KRW 기준)'
    ws['A1'].font = Font(bold=True, size=16, name='Arial')
    ws['A2'] = f'생성일시: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(size=10, name='Arial', color='666666')
    ws['A4'] = f'참여 회사 수: {len(companies)} 개사'
    ws['A4'].font = Font(bold=True, size=12, name='Arial')

    headers = ['#', '회사명 (국문)', '통화', '평균환율 (vs KRW)', '파일명']
    for i, h in enumerate(headers, 1):
        c = ws.cell(6, i, h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
    for i, e in enumerate(extracted_list, 1):
        ws.cell(6 + i, 1, i)
        ws.cell(6 + i, 2, e['company'])
        ws.cell(6 + i, 3, e.get('currency', '-'))
        rate_cell = ws.cell(6 + i, 4, e.get('fx_rate', 1))
        rate_cell.number_format = '#,##0.0000'
        ws.cell(6 + i, 5, e['file'].split('\\')[-1].split('/')[-1])

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 60

    for sheet_name in SHEET_DISPLAY:
        _write_sheet(wb, sheet_name, companies, agg_result['sheets'].get(sheet_name, {}))

    wb.save(output_path)
    return output_path
