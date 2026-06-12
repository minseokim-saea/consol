"""모든 PY 영역(A/B, D/E, G/H, J/K, N, S) 채움 검증"""
import sys, json
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import distribute_builder as db
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

state = json.loads(Path("uploads/_state.json").read_text(encoding="utf-8"))
tpl = db.get_template_path("2026")

out_dir = Path("results/distribute/_test_full")
out_dir.mkdir(parents=True, exist_ok=True)
for p in out_dir.glob("*.xlsm"):
    p.unlink()

res = db.build_distribution_package(
    template_path=tpl,
    output_dir=out_dir,
    company="글로벌세아",
    target_year=2026,
    target_quarter=1,
    uploaded_files=state,
    file_password=None,
    sheet_protect_password=None,
)
print(f"ok={res['ok']}")
print(f"wrote_bs={res['wrote_bs']}, wrote_pl={res['wrote_pl']}")
print(f"wrote_pl_mf_secondary={res.get('wrote_pl_mf_secondary')}")
print(f"wrote_n_refs={res.get('wrote_n_refs')}, wrote_s_refs={res.get('wrote_s_refs')}")
print(f"output: {res['output_path']}")

# PY 시트 검증
wb = load_workbook(res["output_path"], keep_vba=True, data_only=False)
py = wb["PY"]
print(f"\nPY max_row={py.max_row}")

print("\n=== PY 4~14행 (모든 컬럼) ===")
for r in range(4, 15):
    cells = []
    for c in range(1, py.max_column + 1):
        v = py.cell(r, c).value
        if v is None: continue
        s = repr(v)
        if len(s) > 35: s = s[:35] + "..."
        cells.append(f"{get_column_letter(c)}={s}")
    if cells:
        print(f"  row {r:3d}: " + " | ".join(cells))

# N열·S열 모두
print("\n=== PY 4~30행 J/K/N/S 컬럼만 ===")
for r in range(4, 31):
    j = py.cell(r, 10).value
    k = py.cell(r, 11).value
    n = py.cell(r, 14).value
    s = py.cell(r, 19).value
    pieces = []
    if j is not None: pieces.append(f"J={j}")
    if k is not None: pieces.append(f"K={k}")
    if n is not None: pieces.append(f"N={n}")
    if s is not None: pieces.append(f"S={s}")
    if pieces:
        print(f"  row {r:3d}: " + " | ".join(pieces))

# J/K 마지막 데이터 행 확인
print("\n=== J/K열 마지막 데이터 행 (3개) ===")
last_jk = None
for r in range(5, py.max_row + 1):
    if py.cell(r, 10).value is not None or py.cell(r, 11).value is not None:
        last_jk = r
if last_jk:
    for r in range(max(5, last_jk-2), last_jk+1):
        j = py.cell(r, 10).value
        k = py.cell(r, 11).value
        print(f"  row {r:3d}: J={j}, K={k}")

wb.close()
print("\n[DONE]")
