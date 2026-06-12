"""
주석 합산 모듈.

L1(단기차입금) 등 패키지 주석 시트를 zipfile + xml stream 파싱으로 빠르게 추출.
extract_l1_borrowings(): 4종 단기차입금의 명세 행을 종류별로 반환.

섹션 식별 (행 추가에 안전한 라벨 기반):
  · A열에 "N." 패턴(예: '1.', '2.') + B열에 종류명 + D열에 계정코드 → 섹션 시작
  · 다음 섹션 시작 또는 B열 'Total' 라벨 → 명세 종료
  · 명세 영역: B(대주구분), C(대주명), D(이자율), E(금액) — E가 숫자인 행만 채택

자회사가 한 섹션 내에 명세 행을 추가해도 그대로 합산 (Total 라벨 전까지).
"""

import re
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_NS_MAIN = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_NS_R    = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_TAG_ROW = f'{{{_NS_MAIN}}}row'
_TAG_C   = f'{{{_NS_MAIN}}}c'
_TAG_V   = f'{{{_NS_MAIN}}}v'
_TAG_IS  = f'{{{_NS_MAIN}}}is'
_TAG_T   = f'{{{_NS_MAIN}}}t'
_TAG_SI  = f'{{{_NS_MAIN}}}si'
_TAG_SHEET = f'{{{_NS_MAIN}}}sheet'

_CELL_REF_RE = re.compile(r'^([A-Z]+)(\d+)$')
_SECTION_RE  = re.compile(r'^\s*(\d+)\.\s*$')
_L4_START_RE = re.compile(r'^\s*1\s*-\s*1\s*\.')
_L4_END_RE   = re.compile(r'^\s*2\s*\.')
_L4_Q2_RE    = re.compile(r'^\s*2\s*\.(?!\d)')   # "2." 단, "2-1" 은 제외
_L4_Q21_RE   = re.compile(r'^\s*2\s*-\s*1')
_L4_Q3_RE    = re.compile(r'^\s*3\s*\.(?!\d)')
_L4_Q31_RE   = re.compile(r'^\s*3\s*-\s*1')
_L4_Q4_RE    = re.compile(r'^\s*4\s*\.(?!\d)')
_L4_Q41_RE   = re.compile(r'^\s*4\s*-\s*1')
_L4_Q5_RE    = re.compile(r'^\s*5\s*\.(?!\d)')
_L4_Q52_RE   = re.compile(r'^\s*5\s*-\s*2')
_L4_Q6_RE    = re.compile(r'^\s*6\s*\.(?!\d)')
_L4_Q61_RE   = re.compile(r'^\s*6\s*-\s*1')
_L4_Q7_RE    = re.compile(r'^\s*7\s*\.(?!\d)')
_L4_Q71_RE   = re.compile(r'^\s*7\s*-\s*1')
_L4_Q8_RE    = re.compile(r'^\s*8\s*\.(?!\d)')
_L4_Q81_RE   = re.compile(r'^\s*8\s*-\s*1')
_L4_Q82_RE   = re.compile(r'^\s*8\s*-\s*2')
_L4_Q9_RE    = re.compile(r'^\s*9\s*\.(?!\d)')
_L4_Q91_RE   = re.compile(r'^\s*9\s*-\s*1')
_L4_Q10_RE   = re.compile(r'^\s*10\s*\.(?!\d)')


def _col_letters_to_index(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def _split_cell_ref(ref):
    m = _CELL_REF_RE.match(ref or '')
    if not m:
        return None, None
    return m.group(1), int(m.group(2))


def _load_shared_strings(zf):
    if 'xl/sharedStrings.xml' not in zf.namelist():
        return []
    shared = []
    with zf.open('xl/sharedStrings.xml') as f:
        for event, elem in ET.iterparse(f, events=('end',)):
            if elem.tag != _TAG_SI:
                continue
            texts = [t.text or '' for t in elem.iter(_TAG_T)]
            shared.append(''.join(texts))
            elem.clear()
    return shared


def _find_sheet_path(zf, sheet_name):
    if 'xl/workbook.xml' not in zf.namelist():
        return None
    with zf.open('xl/workbook.xml') as f:
        wb_tree = ET.parse(f)
    rid = None
    for sheet in wb_tree.getroot().iter(_TAG_SHEET):
        if sheet.get('name') == sheet_name:
            rid = sheet.get(f'{{{_NS_R}}}id')
            break
    if not rid:
        return None
    if 'xl/_rels/workbook.xml.rels' not in zf.namelist():
        return None
    with zf.open('xl/_rels/workbook.xml.rels') as f:
        rels_tree = ET.parse(f)
    target = None
    for rel in rels_tree.getroot():
        if rel.get('Id') == rid:
            target = rel.get('Target')
            break
    if not target:
        return None
    return target if target.startswith('xl/') else 'xl/' + target


def _cell_value(c_elem, shared):
    t = c_elem.get('t')
    if t == 's':
        v = c_elem.find(_TAG_V)
        if v is not None and v.text is not None:
            try:
                idx = int(v.text)
                if 0 <= idx < len(shared):
                    return shared[idx]
            except ValueError:
                pass
        return None
    if t == 'inlineStr':
        is_el = c_elem.find(_TAG_IS)
        if is_el is not None:
            texts = [te.text or '' for te in is_el.iter(_TAG_T)]
            return ''.join(texts) or None
        return None
    if t == 'str':
        # 수식 결과. 숫자로 변환 가능하면 숫자로 — A2 시트가 cost/book을 t='str'로 저장하는 케이스 대응
        v = c_elem.find(_TAG_V)
        if v is None or v.text is None:
            return None
        txt = v.text
        try:
            fv = float(txt)
            if fv == int(fv):
                return int(fv)
            return fv
        except ValueError:
            return txt
    v = c_elem.find(_TAG_V)
    if v is None or v.text is None:
        return None
    try:
        fv = float(v.text)
        if fv == int(fv):
            return int(fv)
        return fv
    except ValueError:
        return v.text


def extract_l1_borrowings(file_path, sheet_name='L1', max_row_scan=200):
    """L1 시트에서 차입금 종류별 명세 행 추출.

    반환: {
      'sheet_found': bool,
      'categories': [
        {'key':'1', 'name':'원재료차입금', 'code':'2100205',
         'rows':[{'creditor_type','creditor','rate','amount'}]},
        ...
      ],
      'error': str | None,
    }
    """
    try:
        zf = zipfile.ZipFile(file_path)
    except Exception as e:
        return {'sheet_found': False, 'categories': [], 'error': f'파일 열기 실패: {e}'}

    try:
        sheet_path = _find_sheet_path(zf, sheet_name)
        if not sheet_path or sheet_path not in zf.namelist():
            return {'sheet_found': False, 'categories': [], 'error': None}

        shared = _load_shared_strings(zf)

        # 1) row 단위로 A~E 컬럼 값 수집
        rows_data = {}   # r → {col_idx: value}
        with zf.open(sheet_path) as f:
            for event, elem in ET.iterparse(f, events=('end',)):
                if elem.tag != _TAG_ROW:
                    continue
                r_attr = elem.get('r')
                if not r_attr:
                    elem.clear()
                    continue
                r = int(r_attr)
                if r > max_row_scan:
                    elem.clear()
                    continue
                vals = {}
                for c in elem.findall(_TAG_C):
                    letters, _ = _split_cell_ref(c.get('r'))
                    if not letters:
                        continue
                    col_idx = _col_letters_to_index(letters)
                    if col_idx > 5:   # A~E만
                        continue
                    vals[col_idx] = _cell_value(c, shared)
                if vals:
                    rows_data[r] = vals
                elem.clear()
    finally:
        zf.close()

    if not rows_data:
        return {'sheet_found': True, 'categories': [], 'error': None}

    sorted_rows = sorted(rows_data.keys())

    # 2) 섹션 헤더 찾기: A열 "N." + B열 종류명 + D열 코드
    section_starts = []
    for r in sorted_rows:
        v = rows_data[r]
        a = v.get(1)
        if a is None:
            continue
        m = _SECTION_RE.match(str(a).strip())
        if not m:
            continue
        key = m.group(1)
        name = str(v.get(2) or '').strip() or f'종류 {key}'
        code = v.get(4)
        code_str = str(code).strip() if code is not None else ''
        section_starts.append((r, key, name, code_str))

    # 3) 각 섹션 명세 추출
    categories = []
    for i, (start_r, key, name, code) in enumerate(section_starts):
        end_r = section_starts[i + 1][0] if i + 1 < len(section_starts) else 10**9
        rows = []
        for r in sorted_rows:
            if r <= start_r or r >= end_r:
                continue
            v = rows_data[r]
            b = v.get(2)
            # B열 'Total' 라벨 → 명세 종료
            if isinstance(b, str) and 'total' in b.lower():
                break
            # E열이 숫자가 아니면(헤더/메타 행) 스킵
            e = v.get(5)
            if not isinstance(e, (int, float)) or isinstance(e, bool):
                continue
            rows.append({
                'creditor_type': str(v.get(2) or '').strip(),
                'creditor':      str(v.get(3) or '').strip(),
                'rate':          v.get(4),
                'amount':        float(e),
            })
        categories.append({
            'key': key, 'name': name, 'code': code,
            'rows': rows,
        })

    return {'sheet_found': True, 'categories': categories, 'error': None}


def _extract_l4_table_section(file_path, sheet_name, start_re, end_re,
                              col_map, max_col, not_found_msg,
                              amount_col=6, skip_total=False,
                              max_row_scan=200):
    """L4 시트의 '다행 명세 + 금액 컬럼' 패턴 섹션 공통 추출 헬퍼.

    구조 (1-1 대출한도, 4-1/5-2/8-1/9-1 보증·담보, 7-1/8-2 등):
      · B열 start_re 매칭 → 섹션 시작 (라벨 행)
      · B열 end_re   매칭 → 섹션 종료 (직전 행까지 채택)
      · 사이 영역에서 amount_col이 숫자인 행만 채택 (헤더/주석 행 자동 스킵)
      · skip_total=True 면 C열(col 3) 값이 'total'(대소문자 무시)인 행 제외

    col_map: {col_idx(1-base): key_name}. 'amount'/'currency' 키는 자동 처리.
             col_map의 'amount' 키 위치는 amount_col과 일치해야 함.
    max_col: 워크시트에서 읽을 최대 컬럼 인덱스
    amount_col: 금액 컬럼 인덱스 (기본 6=F열). 7-1·8-2는 5=E열.

    반환: {sheet_found, section_label, rows[dict], error}
    """
    try:
        zf = zipfile.ZipFile(file_path)
    except Exception as e:
        return {'sheet_found': False, 'section_label': None, 'rows': [],
                'error': f'파일 열기 실패: {e}'}

    try:
        sheet_path = _find_sheet_path(zf, sheet_name)
        if not sheet_path or sheet_path not in zf.namelist():
            return {'sheet_found': False, 'section_label': None, 'rows': [],
                    'error': None}

        shared = _load_shared_strings(zf)

        rows_data = {}
        with zf.open(sheet_path) as f:
            for event, elem in ET.iterparse(f, events=('end',)):
                if elem.tag != _TAG_ROW:
                    continue
                r_attr = elem.get('r')
                if not r_attr:
                    elem.clear(); continue
                r = int(r_attr)
                if r > max_row_scan:
                    elem.clear(); continue
                vals = {}
                for c in elem.findall(_TAG_C):
                    letters, _ = _split_cell_ref(c.get('r'))
                    if not letters:
                        continue
                    col_idx = _col_letters_to_index(letters)
                    if col_idx > max_col:
                        continue
                    vals[col_idx] = _cell_value(c, shared)
                if vals:
                    rows_data[r] = vals
                elem.clear()
    finally:
        zf.close()

    if not rows_data:
        return {'sheet_found': True, 'section_label': None, 'rows': [],
                'error': None}

    sorted_rows = sorted(rows_data.keys())

    start_r = end_r = None
    section_label = None
    for r in sorted_rows:
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        b_str = b.strip()
        if start_r is None and start_re.match(b_str):
            start_r = r
            section_label = b_str
            continue
        if start_r is not None and end_re.match(b_str):
            end_r = r
            break

    if start_r is None:
        return {'sheet_found': True, 'section_label': None, 'rows': [],
                'error': not_found_msg}
    if end_r is None:
        end_r = 10**9

    rows = []
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        detect = v.get(amount_col)
        if not isinstance(detect, (int, float)) or isinstance(detect, bool):
            continue
        # Total 행 제외 (6-1 소송 등)
        if skip_total:
            c_val = v.get(3)
            if isinstance(c_val, str) and c_val.strip().lower() == 'total':
                continue
        row_dict = {}
        for col_idx, key in col_map.items():
            cell_val = v.get(col_idx)
            if key == 'amount':
                # col_map의 'amount' 키가 가리키는 컬럼의 값을 채택
                # (amount_col과 다른 컬럼일 수 있음 — 6-1처럼)
                if isinstance(cell_val, (int, float)) and not isinstance(cell_val, bool):
                    row_dict['amount'] = float(cell_val)
                else:
                    row_dict['amount'] = None
            elif key == 'currency':
                row_dict['currency'] = (str(cell_val).strip().upper()
                                        if cell_val not in (None, '')
                                        else 'KRW') or 'KRW'
            else:
                row_dict[key] = (str(cell_val).strip()
                                 if cell_val not in (None, '') else '')
        rows.append(row_dict)

    return {'sheet_found': True, 'section_label': section_label,
            'rows': rows, 'error': None}


def extract_l4_loan_facility(file_path, sheet_name='L4', max_row_scan=200):
    """L4 시트의 1-1 (대출한도 약정 내용) 섹션 명세 행 추출.

    섹션 식별 (행 추가에 안전):
      · B열 '1-1.' prefix → 시작,  B열 '2.' prefix → 종료
      · 사이 영역 F열 숫자 행을 채택
      · 컬럼: C=종류(type), D=금융기관(institution), E=통화(currency), F=금액(amount)

    반환: {sheet_found, section_label, rows[{type,institution,currency,amount}], error}
    """
    return _extract_l4_table_section(
        file_path, sheet_name,
        start_re=_L4_START_RE, end_re=_L4_END_RE,
        col_map={3: 'type', 4: 'institution', 5: 'currency', 6: 'amount'},
        max_col=6,
        not_found_msg='1-1 섹션을 찾지 못함',
        max_row_scan=max_row_scan,
    )


def extract_l4_guarantees_received(file_path, sheet_name='L4', max_row_scan=200):
    """L4 시트의 4-1 (제공받은 보증 내용) 섹션 명세 행 추출.

    컬럼: C=제공자, D=보증종류, E=통화, F=금액, G=관련계정, H=설명
    """
    return _extract_l4_table_section(
        file_path, sheet_name,
        start_re=_L4_Q41_RE, end_re=_L4_Q5_RE,
        col_map={3: 'guarantor', 4: 'type', 5: 'currency',
                 6: 'amount',    7: 'account', 8: 'description'},
        max_col=8,
        not_found_msg='4-1 섹션을 찾지 못함',
        max_row_scan=max_row_scan,
    )


def extract_l4_guarantees_provided(file_path, sheet_name='L4', max_row_scan=200):
    """L4 시트의 5-2 (제공한 보증 내용) 섹션 명세.

    컬럼: C=제공받는자, D=보증종류, E=통화, F=금액, G=채권자, H=설명
    """
    return _extract_l4_table_section(
        file_path, sheet_name,
        start_re=_L4_Q52_RE, end_re=_L4_Q6_RE,
        col_map={3: 'beneficiary', 4: 'type', 5: 'currency',
                 6: 'amount', 7: 'guaranteed_creditor', 8: 'description'},
        max_col=8,
        not_found_msg='5-2 섹션을 찾지 못함',
        max_row_scan=max_row_scan,
    )


def extract_l4_lawsuits(file_path, sheet_name='L4', max_row_scan=200):
    """L4 시트의 6-1 (소송중인 사건 내용) 섹션 명세.

    특수형 — 통화 없음, Total 행은 합산에서 제외:
      C=Type(피고/원고),  D=소송건수(count),
      E=소송금액(claim_amount),  F=인식한 충당부채(provision_amount, =amount)

    D열(소송건수)이 숫자인 행을 채택 — 원고처럼 충당부채(F열)가 비어있어도 잡힘.
    C='Total' 행은 자동 제외.
    """
    return _extract_l4_table_section(
        file_path, sheet_name,
        start_re=_L4_Q61_RE, end_re=_L4_Q7_RE,
        col_map={3: 'type', 4: 'count',
                 5: 'claim_amount', 6: 'amount'},  # F열 = provision_amount
        max_col=6,
        not_found_msg='6-1 섹션을 찾지 못함',
        amount_col=4,   # ← D열(소송건수) 기준으로 검출 (F가 비어도 채택)
        skip_total=True,
        max_row_scan=max_row_scan,
    )


def extract_l4_restricted_financial(file_path, sheet_name='L4', max_row_scan=200):
    """L4 시트의 7-1 (사용제한 금융상품 내용) 섹션 명세.

    E열이 금액인 패턴:
      C=계정과목, D=통화, E=금액, F=제한내용(description)
    """
    return _extract_l4_table_section(
        file_path, sheet_name,
        start_re=_L4_Q71_RE, end_re=_L4_Q8_RE,
        col_map={3: 'account', 4: 'currency',
                 5: 'amount',  6: 'description'},
        max_col=6,
        not_found_msg='7-1 섹션을 찾지 못함',
        amount_col=5,
        max_row_scan=max_row_scan,
    )


def extract_l4_insured_ppe(file_path, sheet_name='L4', max_row_scan=200):
    """L4 시트의 8-1 (보험가입 유형자산 내용) 섹션 명세.

    컬럼: C=자산종류, D=보험사, E=통화, F=부보금액, G=설명
    """
    return _extract_l4_table_section(
        file_path, sheet_name,
        start_re=_L4_Q81_RE, end_re=_L4_Q82_RE,
        col_map={3: 'asset_type', 4: 'insurer', 5: 'currency',
                 6: 'amount',    7: 'description'},
        max_col=7,
        not_found_msg='8-1 섹션을 찾지 못함',
        max_row_scan=max_row_scan,
    )


def extract_l4_pledged_proceeds(file_path, sheet_name='L4', max_row_scan=200):
    """L4 시트의 8-2 (보험수익금 질권설정) 섹션 명세.

    E열이 금액인 패턴:
      C=질권자, D=통화, E=질권금액, F=설명
    """
    return _extract_l4_table_section(
        file_path, sheet_name,
        start_re=_L4_Q82_RE, end_re=_L4_Q9_RE,
        col_map={3: 'pledgee', 4: 'currency',
                 5: 'amount',  6: 'description'},
        max_col=6,
        not_found_msg='8-2 섹션을 찾지 못함',
        amount_col=5,
        max_row_scan=max_row_scan,
    )


_A2_START_RE = re.compile(r'^\s*1\s*\.\s*유가증권', re.IGNORECASE)
_A3_IP_START_RE = re.compile(r'^\s*1\s*\.\s*투자부동산')
_A3_IP_END_RE   = re.compile(r'^\s*2\s*\.(?!\d|\s*-)')   # "2." 단, "2-1" 제외
_A3_Q21_RE = re.compile(r'^\s*2\s*-\s*1')
_A3_Q31_RE = re.compile(r'^\s*3\s*-\s*1')
_A3_Q3_RE  = re.compile(r'^\s*3\s*\.(?!\d|\s*-)')
_A3_Q4_RE  = re.compile(r'^\s*4\s*\.(?!\d|\s*-)')

# 1번 투자부동산 관련 손익 — B열 한글 라벨 → 표준 키 매핑
_A3_IP_ITEM_KEYS = {
    '임대수익':              'rental_revenue',
    '직접 관련된 운영비용':  'operating_expenses',
    '감가상각비':            'depreciation',
    '공정가치 변동':         'fv_change',
    '기타':                  'others',
}

# A4 정규식
_A4_S1_RE = re.compile(r'^\s*1\s*\.\s*공사계약')
_A4_S2_RE = re.compile(r'^\s*2\s*\.\s*진행중')
_A4_S3_RE = re.compile(r'^\s*3\s*\.\s*계약자산')
_A4_END_RE = re.compile(r'^\s*\d+\s*\.\s*')   # 다음 섹션

# 공사 종류 라벨 (lowercase 부분 검색) → 표준 키
_A4_TYPE_KEYS = [
    ('건축',       'architecture'),
    ('토목',       'civil'),
    ('플랜트',     'plant'),
    ('수소충전소', 'hydrogen'),
    ('others',     'others'),    # B='Others' (영문)
]

def _a4_match_type(b_val):
    """B열 값을 공사 종류 표준 키로 매핑. 매칭 안 되면 None."""
    if not isinstance(b_val, str):
        return None
    s = b_val.strip().lower()
    for label, key in _A4_TYPE_KEYS:
        if label.lower() in s:
            return key
    return None


# 2번 (Pivot) — B열 한글 행 라벨 → 표준 키
_A4_PROFIT_ITEMS = {
    '누적공사수익': 'accumulated_revenue',
    '누적공사원가': 'accumulated_cost',
    '누적공사손익': 'accumulated_income',
}

# A5 정규식 + 라벨 매핑
_A5_S1_RE = re.compile(r'^\s*1\s*\.\s*사용권자산')
_A5_S2_RE = re.compile(r'^\s*2\s*\.\s*리스계약')

# A5 1번 자산종류 (B열 부분 검색 — 영문 키워드 우선)
_A5_ASSET_KEYS = [
    ('property',   ('properties', '부동산')),
    ('vehicle',    ('vehicles', '차량운반구')),
    ('equipment',  ('construction equipment', '건설장비')),
    ('others',     ('others', '기타')),
]

def _a5_match_asset(b_val):
    if not isinstance(b_val, str):
        return None
    s = b_val.strip().lower()
    if s == 'total':
        return None
    for key, keywords in _A5_ASSET_KEYS:
        for kw in keywords:
            if kw.lower() in s:
                return key
    return None

# A5 2번 손익 항목 (B열 한글 부분 검색)
# 주의: 라벨이 더 구체적인 것을 먼저 두어야 함.
# 예: '소액자산리스 관련비용\n(단기리스에 포함된 비용 제외)'에 '단기리스'가 substring으로
# 포함되므로 'low_value'를 'short_term'보다 먼저 검사해야 함.
_A5_PL_ITEMS = [
    ('depreciation', '사용권자산 감가상각비'),
    ('interest',     '리스부채 이자비용'),
    ('low_value',    '소액자산리스'),    # 'short_term'보다 먼저
    ('short_term',   '단기리스'),
    ('variable',     '변동리스료'),
    ('disposal_gain','리스처분이익'),
]

def _a5_match_pl_item(b_val):
    if not isinstance(b_val, str):
        return None
    s = b_val.strip()
    for key, label in _A5_PL_ITEMS:
        if label in s:
            return key
    return None

# A6 정규식
_A6_S1_RE = re.compile(r'^\s*1\s*\.\s*파생상품')

# L2 Verification 라벨
_L2_VERIFY_RE = re.compile(r'^\s*verification\s*$', re.IGNORECASE)

# L2 1번/2번 섹션 — A열에 '1.' / '2.' / '3.' 라벨
_L2_S1_RE = re.compile(r'^\s*1\s*\.\s*$')
_L2_S2_RE = re.compile(r'^\s*2\s*\.\s*$')
_L2_S3_RE = re.compile(r'^\s*3\s*\.\s*$')

# L3 시트 — 퇴직급여충당부채 (B열에 '1.', '2.', '3.', '4.' 시작 라벨)
_L3_S1_RE = re.compile(r'^\s*1\s*\.\s+퇴직급여')
_L3_S2_RE = re.compile(r'^\s*2\s*\.\s+퇴직연금운용자산의\s*변동')
_L3_S3_RE = re.compile(r'^\s*3\s*\.\s+퇴직연금운용자산의\s*구성')
_L3_S4_RE = re.compile(r'^\s*4\s*\.\s+퇴직연금운용자산의\s*운용사')

# L3 1번 항목 라벨 → 표준 키
_L3_S1_ITEMS = {
    '기초금액':               'beginning',
    '퇴직급여 설정액':        'provision',
    '급여지급액':             'payment',
    '관계사전출입액':         'transfer',
    '사업결합으로 인한 증가': 'business_combination',
    '기타증감':               'others',
    '기말금액':               'ending',
}

# L3 2번 항목 라벨 → 표준 키
_L3_S2_ITEMS = {
    '기초금액':               'beginning',
    '적립액':                 'contribution',
    '급여지급액':             'payment',
    '이자수익':               'interest_income',
    '관계사전출입액':         'transfer',
    '사업결합으로 인한 증가': 'business_combination',
    '기타증감':               'others',
    '기말금액':               'ending',
}

# L3 3번 자산종류 (B열 부분 검색, 영문 포함)
_L3_S3_ASSETS = [
    ('cash',       ('현금', 'cash')),
    ('deposit',    ('예금', 'deposit')),
    ('securities', ('주식', 'securities')),
    ('bond',       ('채권', 'bond')),
    ('others',     ('기타', 'others')),
]

# ── TX 시트 — 법인세 ──────────────────────────────────────────
# 섹션 식별: A열 번호 + B열 라벨 (B에 번호 prefix 없음).
# A열 값은 정수/문자열 모두 가능 — 매칭 시 str() 변환.
_TX_S1_NUM  = '1'         # B='이연법인세자산(부채) 증감내용 ...'
_TX_S2_NUM  = '2'         # 법정세율
_TX_S3_NUM  = '3'         # 법인세비용의 구성내역
_TX_S31_NUM = '3-1'       # 자본에 직접 부과되는 이연법인세
_TX_S4_NUM  = '4'         # Reconciliation
_TX_S5_NUM  = '5'         # 이연법인세로 인식되지 않은
_TX_S51_NUM = '5-1'       # 이월결손금 만기
_TX_S6_NUM  = '6'         # 보험수리적손익 검증

# TX 3번 — 법인세비용 구성내역 (4개 항목 + 합계). B열 한글 라벨 기준.
_TX_S3_ITEMS = {
    '당기법인세부담액':                      'current_tax',
    '일시적차이로 인한 이연법인세 변동액':    'deferred_temp_diff',
    '자본에 직접 부과되는 이연법인세 변동액': 'deferred_equity',
    '법인세추납액(환급액)':                  'additional_refund',
    '법인세비용':                            'total_expense',
}

# TX 3-1번 — 자본 직접 부과 이연법인세 변동액 명세 (8개 항목 + 합계).
_TX_S31_ITEMS = {
    '재평가이익':                       'revaluation',
    '보험수리적손익':                   'actuarial',
    '매도가능증권평가손익':             'afs_securities',
    '기타포괄손익-공정가치측정금융상품': 'fvoci',
    '지분법적용투자주식':               'equity_method',
    '해외사업환산손익':                 'fx_translation',
    '파생상품평가손익':                 'derivatives',
    '기타':                             'others',
}

# TX 4번 — Reconciliation (항목별 + 유효세율 별도).
_TX_S4_ITEMS = {
    '법인세비용차감전순이익':                 'pretax_income',
    '적용세율에 따른 세부담액':               'tax_at_statutory',
    '영구적차이':                            'permanent_diff',
    '세액공제':                              'tax_credit',
    '법인세추납액/환급액':                   'additional_refund',
    '인식하지 않은 일시적차이의 변동 등':    'unrecognized_change',
    '법인세비용':                            'total_expense',
}
_TX_S4_RATE_LABEL  = '유효세율'   # 별도 (회사별 표시)
_TX_S4_RATE_STAT_LABEL = '적용세율에 따른 세부담액'  # 별도 라벨 처리 X

# TX 5번 — 미인식 일시적차이 (2개 항목)
_TX_S5_ITEMS_KEYS = [
    ('loss_carryforward', '이월결손금'),
    ('others',            '기타'),
]

# L3-1 시트 — 확정급여부채
_L31_S1_RE = re.compile(r'^\s*1\s*\.\s+확정급여채무의')
_L31_S2_RE = re.compile(r'^\s*2\s*\.\s+사외적립자산의\s*공정가치')
_L31_S3_RE = re.compile(r'^\s*3\s*\.\s+보험수리적\s*평가')
_L31_S4_RE = re.compile(r'^\s*4\s*\.\s+보험수리적\s*가정의\s*변동')
_L31_S5_RE = re.compile(r'^\s*5\s*\.\s+사외적립자산의\s*구성')
_L31_S6_RE = re.compile(r'^\s*6\s*\.\s+사외적립자산의\s*운용사')

# L3-1 1번 (확정급여채무 변동) 11개 항목
_L31_S1_ITEMS = {
    '기초금액':                                  'beginning',
    '근무원가':                                  'current_service_cost',
    '이자비용':                                  'interest_cost',
    '확정급여제도의 재측정요소':                 'remeasurement',
    '인구통계가정 변동에 의한 보험수리적손익':   'demographic_gain_loss',
    '재무적가정 변동에 의한 보험수리적손익':     'financial_gain_loss',
    '확정급여채무에 대한 경험적 조정':           'experience_adjustment',
    '급여지급액':                                'payment',
    '사업결합으로 인한 증가':                    'business_combination',
    '기타증감':                                  'others',
    '기말금액':                                  'ending',
}

# L3-1 2번 (사외적립자산 공정가치 변동) 8개 항목
_L31_S2_ITEMS = {
    '기초금액':                                       'beginning',
    '사외적립자산의 이자수익':                        'interest_income',
    '순이자에 포함된 금액을 제외한 제도자산의 손익':  'return_excluding_interest',
    '고용인의 기여금':                                'employer_contribution',
    '급여지급액':                                     'payment',
    '사업결합으로 인한 증가':                         'business_combination',
    '기타증감':                                       'others',
    '기말금액':                                       'ending',
}

# L3-1 4번 가정 항목 (기대임금상승율, 할인율) × {1% 상승, 1% 하락}
_L31_S4_ASSUMPTIONS = {
    '기대임금상승율': 'wage_growth',
    '할인율':         'discount_rate',
}

# L3-1 3번 — 보험수리적 평가를 위한 주요 가정치 (기대임금상승률 / 할인율)
_L31_S3_ASSUMPTIONS = {
    '기대임금상승율': 'wage_growth',
    '기대임금상승률': 'wage_growth',
    '할인율':         'discount_rate',
}


def _parse_rate_value(raw):
    """L3-1 3번 가정치 값 파싱.

    반환: {'raw': 원본문자열, 'value': float|None,
           'value_min': float|None, 'value_max': float|None}
      - value: 비교용 단일 수치 (범위면 평균)
      - value_min/max: 범위인 경우 양 끝 값
      - 0 / 미입력 / 'NO' / 'Error!' → 모두 None 처리
    """
    out = {'raw': '', 'value': None, 'value_min': None, 'value_max': None}
    if raw is None:
        return out
    # 숫자(int/float) 직접 입력
    if isinstance(raw, (int, float)):
        v = float(raw)
        out['raw'] = f'{v*100:.2f}%' if 0 < v < 1 else str(raw)
        if v != 0:
            out['value'] = v
            out['value_min'] = v
            out['value_max'] = v
        return out
    s = str(raw).strip()
    out['raw'] = s
    if not s:
        return out
    upper = s.upper()
    if upper in ('NO', 'N/A', 'NA', '-', 'ERROR!') or upper.startswith('ERROR'):
        return out
    # 1. 순수 수치 (소수 또는 지수표기)
    try:
        v = float(s)
        if v == 0:
            return out
        out['value'] = v
        out['value_min'] = v
        out['value_max'] = v
        return out
    except (TypeError, ValueError):
        pass
    # 2. 숫자(%) 추출 — 범위 또는 단일/혼합 문자열 모두 처리
    nums = re.findall(r'(-?\d+(?:\.\d+)?)\s*%', s)
    pcts = [float(n) / 100.0 for n in nums]
    pcts = [p for p in pcts if p != 0]
    if not pcts:
        # %가 없는데 숫자만 있을 수 있음 (드물지만)
        nums2 = re.findall(r'-?\d+(?:\.\d+)?', s)
        try:
            vals = [float(n) for n in nums2 if float(n) != 0]
        except ValueError:
            vals = []
        if not vals:
            return out
        # 값이 1 미만이면 비율로, 1 이상이면 백분율 가정
        normed = [(v if 0 < v < 1 else v / 100.0) for v in vals]
        pcts = normed
    pcts.sort()
    out['value_min'] = pcts[0]
    out['value_max'] = pcts[-1]
    out['value'] = sum(pcts) / len(pcts)
    return out


def _l3_match_pl_item(b_val, item_dict):
    """L3 1/2번 B열 라벨 → 표준 키 (한글 라벨 정확 매칭)."""
    if not isinstance(b_val, str):
        return None
    s = b_val.strip()
    return item_dict.get(s)


def _l3_match_asset(b_val):
    if not isinstance(b_val, str):
        return None
    s = b_val.strip().lower()
    for key, keywords in _L3_S3_ASSETS:
        for kw in keywords:
            if kw.lower() in s:
                return key
    return None

# A7 정규식 + 종류 라벨 매핑
_A7_S1_RE = re.compile(r'^\s*1\s*\.\s*지분법')

def _a7_match_type(b_val):
    """B열 종류 라벨 → 표준 키.
    '종속회사' or 'Subsidiaries' → subsidiary
    '기타지분법' or 'Other Equity' → other
    """
    if not isinstance(b_val, str):
        return None
    s = b_val.strip().lower()
    if '종속회사' in b_val or 'subsidiaries' in s:
        return 'subsidiary'
    if '기타지분법' in b_val or 'other equity' in s:
        return 'other'
    return None


def extract_a2_securities(file_path, sheet_name='A2', max_row_scan=200):
    """A2 시트의 1. 유가증권 명세 추출.

    섹션 식별 (행 추가 안전):
      · A열에 '1. 유가증권' prefix 행 → 섹션 시작
      · D[r]='FS' 또는 E[r]='SUM' (대소문자 무시) → 검증 영역 시작 → 명세 종료
      · 사이 영역에서 F열(취득원가)이 숫자인 행만 채택

    컬럼:
      B=계정(account), C=피투자회사명(investee),
      D=주식수(shares), E=지분율(ownership_pct),
      F=취득원가(acquisition_cost), G=장부가액(book_amount)

    반환: {sheet_found, section_label, rows[dict], error}
    """
    try:
        zf = zipfile.ZipFile(file_path)
    except Exception as e:
        return {'sheet_found': False, 'section_label': None, 'rows': [],
                'error': f'파일 열기 실패: {e}'}

    try:
        sheet_path = _find_sheet_path(zf, sheet_name)
        if not sheet_path or sheet_path not in zf.namelist():
            return {'sheet_found': False, 'section_label': None, 'rows': [],
                    'error': None}

        shared = _load_shared_strings(zf)
        rows_data = {}
        with zf.open(sheet_path) as f:
            for event, elem in ET.iterparse(f, events=('end',)):
                if elem.tag != _TAG_ROW:
                    continue
                r_attr = elem.get('r')
                if not r_attr:
                    elem.clear(); continue
                r = int(r_attr)
                if r > max_row_scan:
                    elem.clear(); continue
                vals = {}
                for c in elem.findall(_TAG_C):
                    letters, _ = _split_cell_ref(c.get('r'))
                    if not letters:
                        continue
                    col_idx = _col_letters_to_index(letters)
                    if col_idx > 7:   # A~G
                        continue
                    vals[col_idx] = _cell_value(c, shared)
                if vals:
                    rows_data[r] = vals
                elem.clear()
    finally:
        zf.close()

    if not rows_data:
        return {'sheet_found': True, 'section_label': None, 'rows': [],
                'error': None}

    sorted_rows = sorted(rows_data.keys())

    # A열에 '1. 유가증권' 패턴 행 찾기
    start_r = None
    section_label = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if isinstance(a, str) and _A2_START_RE.match(a.strip()):
            start_r = r
            section_label = a.strip()
            break

    if start_r is None:
        return {'sheet_found': True, 'section_label': None, 'rows': [],
                'error': '유가증권 명세 섹션을 찾지 못함'}

    # 종료점: D='FS' or E='SUM' 행 (대소문자 무시)
    end_r = None
    for r in sorted_rows:
        if r <= start_r:
            continue
        v = rows_data[r]
        d = v.get(4); e = v.get(5)
        if isinstance(d, str) and d.strip().upper() == 'FS':
            end_r = r; break
        if isinstance(e, str) and e.strip().upper() == 'SUM':
            end_r = r; break
    if end_r is None:
        end_r = 10**9

    rows = []
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        cost = v.get(6)
        if not isinstance(cost, (int, float)) or isinstance(cost, bool):
            continue
        # 헤더 행 추가 안전: B열이 라벨 텍스트 같으면 (e.g. '계정\nAccount') 스킵
        b = v.get(2)
        if isinstance(b, str) and ('\n' in b or b.lower().startswith('account')):
            continue

        def _to_num(x):
            if isinstance(x, (int, float)) and not isinstance(x, bool):
                return float(x)
            return None

        rows.append({
            'account':          str(v.get(2) or '').strip(),
            'investee':         str(v.get(3) or '').strip(),
            'shares':           _to_num(v.get(4)),
            'ownership_pct':    _to_num(v.get(5)),
            'acquisition_cost': float(cost),
            'book_amount':      _to_num(v.get(7)),
        })

    return {'sheet_found': True, 'section_label': section_label,
            'rows': rows, 'error': None}


def _load_sheet_rows(file_path, sheet_name, max_row_scan=200, max_col=10):
    """임의 시트 행 데이터 로드 (zipfile/xml stream).

    반환: (rows_data | None, error_msg | None)
      rows_data: {row_num: {col_idx: value}}
      시트 없으면 rows_data=None
    """
    try:
        zf = zipfile.ZipFile(file_path)
    except Exception as e:
        return None, f'파일 열기 실패: {e}'
    try:
        sheet_path = _find_sheet_path(zf, sheet_name)
        if not sheet_path or sheet_path not in zf.namelist():
            return None, None
        shared = _load_shared_strings(zf)
        rows_data = {}
        with zf.open(sheet_path) as f:
            for event, elem in ET.iterparse(f, events=('end',)):
                if elem.tag != _TAG_ROW:
                    continue
                r_attr = elem.get('r')
                if not r_attr:
                    elem.clear(); continue
                r = int(r_attr)
                if r > max_row_scan:
                    elem.clear(); continue
                vals = {}
                for c in elem.findall(_TAG_C):
                    letters, _ = _split_cell_ref(c.get('r'))
                    if not letters:
                        continue
                    col_idx = _col_letters_to_index(letters)
                    if col_idx > max_col:
                        continue
                    vals[col_idx] = _cell_value(c, shared)
                if vals:
                    rows_data[r] = vals
                elem.clear()
    finally:
        zf.close()
    return rows_data, None


def _load_a3_rows(file_path, sheet_name='A3', max_row_scan=200, max_col=10):
    """A3 시트 행 데이터 로드. _load_sheet_rows의 A3 전용 래퍼."""
    return _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col)


def _to_float(v):
    """숫자/숫자 문자열을 float로 변환. 실패 시 None."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return None
    return None


def extract_a3_investment_property_pl(file_path, sheet_name='A3', max_row_scan=200):
    """A3 시트의 1. 투자부동산 관련 손익 추출.

    구조 (행 추가 없는 5개 고정 항목):
      A열에 '1. 투자부동산' prefix → 섹션 시작
      A열에 '2.' prefix → 종료
      사이 영역에서 B열 라벨 매칭 (임대수익/운영비용/감가상각비/공정가치 변동/기타)
      D열에 금액

    반환: {
      'sheet_found': bool,
      'items':   {key: float | None},   # 5개 표준 키
      'total':   float | None,           # 합계 (계산값)
      'error':   str | None,
    }
    """
    rows_data, err = _load_a3_rows(file_path, sheet_name, max_row_scan, max_col=6)
    if err:
        return {'sheet_found': False, 'items': {}, 'total': None, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'total': None, 'error': None}

    sorted_rows = sorted(rows_data.keys())

    start_r = end_r = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if not isinstance(a, str):
            continue
        s = a.strip()
        if start_r is None and _A3_IP_START_RE.match(s):
            start_r = r
            continue
        if start_r is not None and _A3_IP_END_RE.match(s):
            end_r = r
            break

    if start_r is None:
        return {'sheet_found': True, 'items': {}, 'total': None,
                'error': '1번 투자부동산 관련 손익 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 30

    items = {k: None for k in _A3_IP_ITEM_KEYS.values()}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        key = _A3_IP_ITEM_KEYS.get(b.strip())
        if not key:
            continue
        items[key] = _to_float(rows_data[r].get(4))

    total = sum(v for v in items.values() if v is not None)
    return {'sheet_found': True, 'items': items, 'total': total, 'error': None}


def extract_a3_land_value_investment(file_path, sheet_name='A3', max_row_scan=200):
    """A3 시트의 2-1. 투자부동산(토지) 공시지가 추출.

    구조: A열에 '2-1.' prefix 행 (r) → D[r] = 단일 금액 (선택적 YES/NO는 2번 행 E열)
          2번 행 E[r]에 응답이 있을 수 있어 yn으로 같이 반환.

    반환: {sheet_found, yn, amount, error}
    """
    rows_data, err = _load_a3_rows(file_path, sheet_name, max_row_scan, max_col=6)
    if err:
        return {'sheet_found': False, 'yn': None, 'amount': None, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'yn': None, 'amount': None, 'error': None}

    sorted_rows = sorted(rows_data.keys())

    # A열에서 "2." (2번 질문)과 "2-1" (명세) 찾기
    r2 = r21 = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if not isinstance(a, str):
            continue
        s = a.strip()
        if r2 is None and _A3_IP_END_RE.match(s):    # "2." (2-1 제외)
            r2 = r
        elif r21 is None and _A3_Q21_RE.match(s):
            r21 = r
            break

    yn = None
    if r2 is not None:
        e2 = rows_data[r2].get(5)
        if isinstance(e2, str) and e2.strip():
            yn = e2.strip().upper()

    amount = None
    if r21 is not None:
        amount = _to_float(rows_data[r21].get(4))

    if r21 is None and yn is None:
        return {'sheet_found': True, 'yn': None, 'amount': None,
                'error': '2-1 섹션을 찾지 못함'}

    return {'sheet_found': True, 'yn': yn, 'amount': amount, 'error': None}


def extract_a3_land_value_ppe(file_path, sheet_name='A3', max_row_scan=200):
    """A3 시트의 3-1. 유형자산(토지) 공시지가 추출.

    구조: A열에 '3-1.' prefix 행 (r) → D[r] = 단일 금액
          3번 행 E[r]에 YES/NO 응답.

    반환: {sheet_found, yn, amount, error}
    """
    rows_data, err = _load_a3_rows(file_path, sheet_name, max_row_scan, max_col=6)
    if err:
        return {'sheet_found': False, 'yn': None, 'amount': None, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'yn': None, 'amount': None, 'error': None}

    sorted_rows = sorted(rows_data.keys())

    r3 = r31 = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if not isinstance(a, str):
            continue
        s = a.strip()
        if r3 is None and _A3_Q3_RE.match(s):
            r3 = r
        elif r31 is None and _A3_Q31_RE.match(s):
            r31 = r
            break

    yn = None
    if r3 is not None:
        e3 = rows_data[r3].get(5)
        if isinstance(e3, str) and e3.strip():
            yn = e3.strip().upper()

    amount = None
    if r31 is not None:
        amount = _to_float(rows_data[r31].get(4))

    if r31 is None and yn is None:
        return {'sheet_found': True, 'yn': None, 'amount': None,
                'error': '3-1 섹션을 찾지 못함'}

    return {'sheet_found': True, 'yn': yn, 'amount': amount, 'error': None}


# ──────────────────────────────────────────────────────────────
# A4 시트 — 건설계약 주석 (1번 잔액 변동 / 2번 공사손익 / 3번 계약자산·부채)
# ──────────────────────────────────────────────────────────────

def _a4_find_section_bounds(rows_data, start_re, next_re):
    """A열에서 start_re 매칭 행 ~ next_re 매칭 행 사이 범위."""
    sorted_rows = sorted(rows_data.keys())
    start_r = end_r = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if not isinstance(a, str):
            continue
        s = a.strip()
        if start_r is None and start_re.match(s):
            start_r = r
            continue
        if start_r is not None and next_re.match(s) and not start_re.match(s):
            end_r = r
            break
    return start_r, end_r, sorted_rows


def extract_a4_construction_balance(file_path, sheet_name='A4', max_row_scan=80):
    """A4 시트의 1. 공사계약 잔액의 변동내역 추출.

    구조 (A5 시작, A19 종료):
      r7  헤더 — B=구분, C=기초잔액, D=증감액, E=공사수익, F=기타, G=기말잔액
      r8~r12 데이터 — 5개 공사 종류 (B열 한글 라벨로 식별)
      r13 Total (검증용 — 추출 결과에는 별도 표시하지 않고 계산값 사용)

    반환: {
      'sheet_found': bool,
      'items': {type_key: {beginning,variance,profit,others,ending}},
      'error': str | None,
    }
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _a4_find_section_bounds(
        rows_data, _A4_S1_RE, _A4_S2_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '1번 공사계약 잔액 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    items = {key: {'beginning': 0.0, 'variance': 0.0,
                   'profit': 0.0, 'others': 0.0, 'ending': 0.0}
             for _, key in _A4_TYPE_KEYS}

    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        type_key = _a4_match_type(v.get(2))
        if not type_key:
            continue
        # Total 라벨 행 제외 (B='Total')
        b = v.get(2)
        if isinstance(b, str) and b.strip().lower() == 'total':
            continue
        items[type_key] = {
            'beginning': _to_float(v.get(3)) or 0.0,
            'variance':  _to_float(v.get(4)) or 0.0,
            'profit':    _to_float(v.get(5)) or 0.0,
            'others':    _to_float(v.get(6)) or 0.0,
            'ending':    _to_float(v.get(7)) or 0.0,
        }

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_a4_construction_profit(file_path, sheet_name='A4', max_row_scan=80):
    """A4 시트의 2. 진행중인 건설계약 관련 공사손익 추출 (Pivot 형태).

    구조 (A19 시작, A27 종료):
      r21 헤더 — C=건축, D=토목, E=플랜트, F=수소충전소, G=Others, H=Total
      r22 누적공사수익, r23 누적공사원가, r24 누적공사손익
        (행 순서 고정, B열 라벨로 식별)

    반환: {
      'sheet_found': bool,
      'items': {item_key: {architecture,civil,plant,hydrogen,others,total}},
      'error': str | None,
    }
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=9)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _a4_find_section_bounds(
        rows_data, _A4_S2_RE, _A4_S3_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '2번 공사손익 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    items = {key: {'architecture': 0.0, 'civil': 0.0, 'plant': 0.0,
                   'hydrogen': 0.0, 'others': 0.0, 'total': 0.0}
             for key in _A4_PROFIT_ITEMS.values()}

    # B열 라벨로 행 식별
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        # 헤더 라벨의 한글 부분만 비교 ('누적공사수익\nAccumulated...')
        b_first_line = b.split('\n', 1)[0].strip()
        item_key = _A4_PROFIT_ITEMS.get(b_first_line)
        if not item_key:
            continue
        v = rows_data[r]
        items[item_key] = {
            'architecture': _to_float(v.get(3)) or 0.0,
            'civil':        _to_float(v.get(4)) or 0.0,
            'plant':        _to_float(v.get(5)) or 0.0,
            'hydrogen':     _to_float(v.get(6)) or 0.0,
            'others':       _to_float(v.get(7)) or 0.0,
            'total':        _to_float(v.get(8)) or 0.0,
        }

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_a4_contract_balance(file_path, sheet_name='A4', max_row_scan=80):
    """A4 시트의 3. 계약자산 및 계약부채 내역 추출.

    구조 (A27 시작, B열 'FS' 또는 'Total' 직후 종료):
      r30~r31 헤더 — C=미청구공사, D=초과청구공사, E=선수금
      r32~r36 데이터 — 5개 공사 종류
      r37 Total / r39 FS / r40 Diff (검증)

    반환: {
      'sheet_found': bool,
      'items': {type_key: {receivable,payable,advance}},
      'error': str | None,
    }
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=6)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    sorted_rows = sorted(rows_data.keys())
    start_r = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if isinstance(a, str) and _A4_S3_RE.match(a.strip()):
            start_r = r
            break
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '3번 계약자산·부채 섹션을 찾지 못함'}

    # 종료점: B='Total' 행 (검증 영역 시작)
    end_r = None
    for r in sorted_rows:
        if r <= start_r:
            continue
        b = rows_data[r].get(2)
        if isinstance(b, str) and b.strip().lower() == 'total':
            end_r = r
            break
    if end_r is None:
        end_r = start_r + 20

    items = {key: {'receivable': 0.0, 'payable': 0.0, 'advance': 0.0}
             for _, key in _A4_TYPE_KEYS}

    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        type_key = _a4_match_type(v.get(2))
        if not type_key:
            continue
        items[type_key] = {
            'receivable': _to_float(v.get(3)) or 0.0,
            'payable':    _to_float(v.get(4)) or 0.0,
            'advance':    _to_float(v.get(5)) or 0.0,
        }

    return {'sheet_found': True, 'items': items, 'error': None}


# ──────────────────────────────────────────────────────────────
# A5 시트 — 리스 (1번 사용권자산 변동 / 2번 리스 손익)
# ──────────────────────────────────────────────────────────────

def extract_a5_rou_changes(file_path, sheet_name='A5', max_row_scan=50):
    """A5 시트의 1. 사용권자산의 변동내역 추출.

    구조 (A5 시작, A17 종료):
      r7 헤더 — C=기초, D=취득, E=처분, F=상각, G=기타, H=기말
      r8~r11 데이터 — 부동산/차량운반구/건설장비/기타 (B열 라벨로 식별)

    반환: {sheet_found, items: {asset_key:{beginning,acquisition,disposal,
                                          depreciation,others,ending}}, error}
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _a4_find_section_bounds(
        rows_data, _A5_S1_RE, _A5_S2_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '1번 사용권자산 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    items = {key: {'beginning': 0.0, 'acquisition': 0.0, 'disposal': 0.0,
                   'depreciation': 0.0, 'others': 0.0, 'ending': 0.0}
             for key, _ in _A5_ASSET_KEYS}

    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        asset_key = _a5_match_asset(v.get(2))
        if not asset_key:
            continue
        items[asset_key] = {
            'beginning':    _to_float(v.get(3)) or 0.0,
            'acquisition':  _to_float(v.get(4)) or 0.0,
            'disposal':     _to_float(v.get(5)) or 0.0,
            'depreciation': _to_float(v.get(6)) or 0.0,
            'others':       _to_float(v.get(7)) or 0.0,
            'ending':       _to_float(v.get(8)) or 0.0,
        }

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_a5_lease_pl(file_path, sheet_name='A5', max_row_scan=60):
    """A5 시트의 2. 리스계약 관련 손익으로 인식된 금액 추출.

    구조 (A17 시작, 시트 끝까지):
      r20~r25 데이터 — 6개 항목 (B열 한글 라벨로 식별), D열에 금액

    반환: {sheet_found, items: {item_key: float}, total, error}
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=5)
    if err:
        return {'sheet_found': False, 'items': {}, 'total': None, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'total': None, 'error': None}

    sorted_rows = sorted(rows_data.keys())
    start_r = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if isinstance(a, str) and _A5_S2_RE.match(a.strip()):
            start_r = r
            break
    if start_r is None:
        return {'sheet_found': True, 'items': {}, 'total': None,
                'error': '2번 리스 손익 섹션을 찾지 못함'}

    items = {key: None for key, _ in _A5_PL_ITEMS}
    for r in sorted_rows:
        if r <= start_r:
            continue
        item_key = _a5_match_pl_item(rows_data[r].get(2))
        if not item_key:
            continue
        items[item_key] = _to_float(rows_data[r].get(4))

    total = sum(v for v in items.values() if v is not None)
    return {'sheet_found': True, 'items': items, 'total': total, 'error': None}


# ──────────────────────────────────────────────────────────────
# A6 시트 — 파생상품평가손익
# ──────────────────────────────────────────────────────────────

def extract_a7_equity_method(file_path, sheet_name='A7', max_row_scan=80):
    """A7 시트의 1. 지분법투자주식 명세 추출.

    구조 (행 추가에 안전):
      A열 '1. 지분법' prefix → 섹션 시작
      B열 'Total' → 섹션 종료
      사이 영역에서 B열에 '종속회사' 또는 '기타지분법' 라벨이 들어간 행 채택
      컬럼: B=종류(type), C=회사명(investee), D=지분율(ownership_pct),
            E=취득원가(acquisition_cost), F=순자산가액(net_asset_value),
            G=장부가액(book_value)

    반환: {
      'sheet_found': bool,
      'rows': [{'type','investee','ownership_pct','acquisition_cost',
                'net_asset_value','book_value'}],
      'error': str | None,
    }
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'rows': [], 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'rows': [], 'error': None}

    sorted_rows = sorted(rows_data.keys())
    start_r = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if isinstance(a, str) and _A7_S1_RE.match(a.strip()):
            start_r = r
            break
    if start_r is None:
        return {'sheet_found': True, 'rows': [],
                'error': '1번 지분법 섹션을 찾지 못함'}

    end_r = None
    for r in sorted_rows:
        if r <= start_r:
            continue
        b = rows_data[r].get(2)
        if isinstance(b, str) and b.strip().lower() == 'total':
            end_r = r
            break
    if end_r is None:
        end_r = start_r + 60

    rows_out = []
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        type_key = _a7_match_type(v.get(2))
        if not type_key:
            continue   # 헤더 라벨 / Total / 빈 행 자동 스킵
        investee = str(v.get(3) or '').strip()
        if not investee:
            continue
        rows_out.append({
            'type':             type_key,
            'investee':         investee,
            'ownership_pct':    _to_float(v.get(4)),
            'acquisition_cost': _to_float(v.get(5)) or 0.0,
            'net_asset_value':  _to_float(v.get(6)) or 0.0,
            'book_value':       _to_float(v.get(7)) or 0.0,
        })

    return {'sheet_found': True, 'rows': rows_out, 'error': None}


def _has_korean(s):
    """문자열에 한글 글자가 포함되어 있는지."""
    if not isinstance(s, str):
        return False
    return any('가' <= ch <= '힣' for ch in s)


def _extract_l2_balance_section(file_path, sheet_name, start_re, end_re,
                                not_found_msg, max_row_scan=80):
    """L2 시트의 'A=숫자.' 시작 → B='Total' 종료 다행 명세 추출.

    L2 1번(장기차입금) / 2번(사채) 공통 패턴:
      r_start: A열에 '1.' 또는 '2.' (B열 항목명)
      r_start+4 정도: 헤더 (B=대주구분/종류, C=대주명/주관사, D=이자율,
                              E=유동, F=비유동, G=Total)
      r_start+5~: 데이터 행 (G=Total이 0이 아닌 숫자인 행)
      B열 'Total' 행 → 종료

    반환: {
      'sheet_found': bool,
      'rows': [{'type1','type2','rate','current','non_current','total'}],
      'error': str | None,
    }
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'rows': [], 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'rows': [], 'error': None}

    sorted_rows = sorted(rows_data.keys())
    start_r = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if isinstance(a, str) and start_re.match(a.strip()):
            start_r = r
            break
    if start_r is None:
        return {'sheet_found': True, 'rows': [], 'error': not_found_msg}

    # 종료: 다음 섹션 (end_re) 또는 그 사이 첫 'Total' 행
    next_section_r = None
    for r in sorted_rows:
        if r <= start_r:
            continue
        a = rows_data[r].get(1)
        if isinstance(a, str) and end_re.match(a.strip()):
            next_section_r = r
            break

    # Total 행 (start_r ~ next_section_r 사이)
    upper = next_section_r if next_section_r is not None else (start_r + 30)
    end_r = None
    for r in sorted_rows:
        if r <= start_r or r >= upper:
            continue
        b = rows_data[r].get(2)
        if isinstance(b, str) and b.strip().lower() == 'total':
            end_r = r
            break
    if end_r is None:
        end_r = upper

    def _first_line(x):
        if not isinstance(x, str):
            return x
        return x.split('\n', 1)[0].strip()

    rows = []
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        # G(Total)을 숫자로 변환 시도 — 문자열 저장 케이스 대응
        # 헤더 행은 G='Total' 같은 문자열이라 None → 자동 스킵
        g_num = _to_float(v.get(7))
        if g_num is None or g_num == 0:
            continue
        rows.append({
            'type1':       _first_line(v.get(2)) or '',
            'type2':       _first_line(v.get(3)) or '',
            'rate':        _to_float(v.get(4)),
            'current':     _to_float(v.get(5)) or 0.0,
            'non_current': _to_float(v.get(6)) or 0.0,
            'total':       g_num,
        })

    return {'sheet_found': True, 'rows': rows, 'error': None}


def extract_l2_long_term_borrowings(file_path, sheet_name='L2', max_row_scan=80):
    """L2 시트의 1. 장기차입금 (유동성 포함) 명세 추출.

    컬럼: B=대주구분(type1), C=대주명(type2), D=이자율(rate),
          E=유동(current), F=비유동(non_current), G=Total
    """
    return _extract_l2_balance_section(
        file_path, sheet_name,
        start_re=_L2_S1_RE, end_re=_L2_S2_RE,
        not_found_msg='1번 장기차입금 섹션을 찾지 못함',
        max_row_scan=max_row_scan,
    )


def extract_l2_debentures(file_path, sheet_name='L2', max_row_scan=80):
    """L2 시트의 2. 사채 (유동성 포함) 명세 추출.

    컬럼: B=종류(type1), C=주관사(type2), D=이자율(rate),
          E=유동(current), F=비유동(non_current), G=Total
    """
    return _extract_l2_balance_section(
        file_path, sheet_name,
        start_re=_L2_S2_RE, end_re=_L2_S3_RE,
        not_found_msg='2번 사채 섹션을 찾지 못함',
        max_row_scan=max_row_scan,
    )


def _l3_find_section_bounds(rows_data, start_re, end_re):
    """B열에서 start_re ~ end_re 사이 범위 (L3 1~4번 공통)."""
    sorted_rows = sorted(rows_data.keys())
    start_r = end_r = None
    for r in sorted_rows:
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        s = b.strip()
        if start_r is None and start_re.match(s):
            start_r = r
            continue
        if start_r is not None and end_re and end_re.match(s):
            end_r = r
            break
    return start_r, end_r, sorted_rows


def extract_l3_severance_provision(file_path, sheet_name='L3', max_row_scan=60):
    """L3 시트의 1. 퇴직급여충당부채의 변동 추출.

    7개 항목 × 1금액 (B열 라벨로 매칭, C열 금액).
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=5)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _l3_find_section_bounds(
        rows_data, _L3_S1_RE, _L3_S2_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '1번 퇴직급여충당부채 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    items = {key: None for key in _L3_S1_ITEMS.values()}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        key = _l3_match_pl_item(rows_data[r].get(2), _L3_S1_ITEMS)
        if not key:
            continue
        items[key] = _to_float(rows_data[r].get(3))

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_l3_pension_funds_movement(file_path, sheet_name='L3', max_row_scan=60):
    """L3 시트의 2. 퇴직연금운용자산의 변동 추출.

    8개 항목 × 1금액 (B열 라벨, C열 금액).
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=5)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _l3_find_section_bounds(
        rows_data, _L3_S2_RE, _L3_S3_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '2번 퇴직연금운용자산 변동 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    items = {key: None for key in _L3_S2_ITEMS.values()}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        key = _l3_match_pl_item(rows_data[r].get(2), _L3_S2_ITEMS)
        if not key:
            continue
        items[key] = _to_float(rows_data[r].get(3))

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_l3_pension_breakdown(file_path, sheet_name='L3', max_row_scan=60):
    """L3 시트의 3. 퇴직연금운용자산의 구성내역 추출.

    5개 자산종류 × 1금액 (B열 라벨, C열 금액, D열 비고).
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=6)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _l3_find_section_bounds(
        rows_data, _L3_S3_RE, _L3_S4_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '3번 퇴직연금운용자산 구성내역 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    items = {key: {'amount': None, 'remarks': ''} for key, _ in _L3_S3_ASSETS}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        # 합계 행 스킵
        b = v.get(2)
        if isinstance(b, str) and '합계' in b.strip():
            continue
        asset_key = _l3_match_asset(b)
        if not asset_key:
            continue
        items[asset_key] = {
            'amount':  _to_float(v.get(3)),
            'remarks': str(v.get(4) or '').strip() if v.get(4) else '',
        }

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_l3_pension_managers(file_path, sheet_name='L3', max_row_scan=60):
    """L3 시트의 4. 퇴직연금운용자산의 운용사 추출 (다행 명세).

    B열 운용사명, C열 금액, D열 비고. 행 추가 안전.
    종료: B='합계' 행 또는 B에 '※' 시작.
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=6)
    if err:
        return {'sheet_found': False, 'rows': [], 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'rows': [], 'error': None}

    sorted_rows = sorted(rows_data.keys())
    start_r = None
    for r in sorted_rows:
        b = rows_data[r].get(2)
        if isinstance(b, str) and _L3_S4_RE.match(b.strip()):
            start_r = r
            break
    if start_r is None:
        return {'sheet_found': True, 'rows': [],
                'error': '4번 운용사 섹션을 찾지 못함'}

    # 종료: B='합계' 또는 '※' 시작
    end_r = None
    for r in sorted_rows:
        if r <= start_r:
            continue
        b = rows_data[r].get(2)
        if isinstance(b, str):
            s = b.strip()
            if '합계' in s or s.startswith('※'):
                end_r = r
                break
    if end_r is None:
        end_r = start_r + 20

    rows = []
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        name = v.get(2)
        amount = _to_float(v.get(3))
        remarks = v.get(4)
        # 헤더 라벨 행 자동 스킵 ('운용사명(Name)')
        if isinstance(name, str) and ('운용사명' in name or 'name' in name.lower()):
            continue
        # 금액 없으면 스킵 (단, 비고만 있는 경우는 유지하지 않음)
        if amount is None or amount == 0:
            continue
        rows.append({
            'name':    str(name or '').strip() if name else '',
            'amount':  amount,
            'remarks': str(remarks or '').strip() if remarks else '',
        })

    return {'sheet_found': True, 'rows': rows, 'error': None}


def extract_l31_dbo_changes(file_path, sheet_name='L3-1', max_row_scan=80):
    """L3-1 시트의 1. 확정급여채무의 변동 추출.

    11개 항목 × 1금액 (B열 라벨, C열 금액).
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=5)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _l3_find_section_bounds(
        rows_data, _L31_S1_RE, _L31_S2_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '1번 확정급여채무 변동 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 20

    items = {key: None for key in _L31_S1_ITEMS.values()}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        key = _l3_match_pl_item(rows_data[r].get(2), _L31_S1_ITEMS)
        if not key:
            continue
        items[key] = _to_float(rows_data[r].get(3))

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_l31_plan_asset_changes(file_path, sheet_name='L3-1', max_row_scan=80):
    """L3-1 시트의 2. 사외적립자산의 공정가치 변동 추출.

    8개 항목 × 1금액.
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=5)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _l3_find_section_bounds(
        rows_data, _L31_S2_RE, _L31_S3_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '2번 사외적립자산 공정가치 변동 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 20

    items = {key: None for key in _L31_S2_ITEMS.values()}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        key = _l3_match_pl_item(rows_data[r].get(2), _L31_S2_ITEMS)
        if not key:
            continue
        items[key] = _to_float(rows_data[r].get(3))

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_l31_assumptions(file_path, sheet_name='L3-1', max_row_scan=80):
    """L3-1 시트의 3. 보험수리적 평가를 위한 주요 가정치 추출.

    2개 가정(기대임금상승률/할인율). 값은 다양한 형식:
      - 0.0429 (소수 비율)
      - '4.29%' / '4.29% + 호봉률' (백분율 문자열, 혼합문)
      - '2.27%~2.97%' (범위)
      - 'NO', 'Error!', 0, 빈 값 (미입력)
    `_parse_rate_value()`로 정규화하여 {raw, value, value_min, value_max} 반환.

    반환: {
      'sheet_found': bool,
      'items': {
        'wage_growth':   {raw, value, value_min, value_max},
        'discount_rate': {raw, value, value_min, value_max},
      },
      'error': str | None,
    }
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=5)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _l3_find_section_bounds(
        rows_data, _L31_S3_RE, _L31_S4_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '3번 보험수리적 가정치 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 10

    items = {'wage_growth':   _parse_rate_value(None),
             'discount_rate': _parse_rate_value(None)}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        key = _L31_S3_ASSUMPTIONS.get(b.strip())
        if not key:
            continue
        items[key] = _parse_rate_value(rows_data[r].get(3))

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_l31_sensitivity(file_path, sheet_name='L3-1', max_row_scan=80):
    """L3-1 시트의 4. 보험수리적 가정의 변동에 의한 영향 추출.

    2개 가정(기대임금상승율/할인율) × {1% 상승, 1% 하락}.

    반환: {
      'sheet_found': bool,
      'items': {
        'wage_growth':   {'up': float|None, 'down': float|None},
        'discount_rate': {'up': float|None, 'down': float|None},
      },
      'error': str | None,
    }
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=5)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _l3_find_section_bounds(
        rows_data, _L31_S4_RE, _L31_S5_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '4번 민감도 분석 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 10

    items = {v: {'up': None, 'down': None} for v in _L31_S4_ASSUMPTIONS.values()}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        key = _L31_S4_ASSUMPTIONS.get(b.strip())
        if not key:
            continue
        items[key] = {
            'up':   _to_float(rows_data[r].get(3)),   # 1% 상승시
            'down': _to_float(rows_data[r].get(4)),   # 1% 하락시
        }

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_l31_plan_breakdown(file_path, sheet_name='L3-1', max_row_scan=80):
    """L3-1 시트의 5. 사외적립자산의 구성내역 추출.

    5개 자산종류 × 1금액 (B=구분, C=금액, D=비고).
    L3 3번과 동일 구조.
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=6)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _l3_find_section_bounds(
        rows_data, _L31_S5_RE, _L31_S6_RE)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '5번 사외적립자산 구성내역 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    items = {key: {'amount': None, 'remarks': ''} for key, _ in _L3_S3_ASSETS}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        b = v.get(2)
        if isinstance(b, str) and '합계' in b.strip():
            continue
        asset_key = _l3_match_asset(b)
        if not asset_key:
            continue
        items[asset_key] = {
            'amount':  _to_float(v.get(3)),
            'remarks': str(v.get(4) or '').strip() if v.get(4) else '',
        }

    return {'sheet_found': True, 'items': items, 'error': None}


def extract_l31_plan_managers(file_path, sheet_name='L3-1', max_row_scan=80):
    """L3-1 시트의 6. 사외적립자산의 운용사 추출 (다행 명세).

    B열 운용사명, C열 금액, D열 비고. 행 추가 안전.
    L3 4번과 동일 구조.
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=6)
    if err:
        return {'sheet_found': False, 'rows': [], 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'rows': [], 'error': None}

    sorted_rows = sorted(rows_data.keys())
    start_r = None
    for r in sorted_rows:
        b = rows_data[r].get(2)
        if isinstance(b, str) and _L31_S6_RE.match(b.strip()):
            start_r = r
            break
    if start_r is None:
        return {'sheet_found': True, 'rows': [],
                'error': '6번 운용사 섹션을 찾지 못함'}

    end_r = None
    for r in sorted_rows:
        if r <= start_r:
            continue
        b = rows_data[r].get(2)
        if isinstance(b, str):
            s = b.strip()
            if '합계' in s or s.startswith('※'):
                end_r = r
                break
    if end_r is None:
        end_r = start_r + 20

    rows = []
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        name = v.get(2)
        amount = _to_float(v.get(3))
        remarks = v.get(4)
        if isinstance(name, str) and ('운용사명' in name or 'name' in name.lower()):
            continue
        if amount is None or amount == 0:
            continue
        rows.append({
            'name':    str(name or '').strip() if name else '',
            'amount':  amount,
            'remarks': str(remarks or '').strip() if remarks else '',
        })

    return {'sheet_found': True, 'rows': rows, 'error': None}


def _tx_find_section_bounds(rows_data, start_num, end_num=None):
    """TX 시트 섹션 경계 찾기. A열의 섹션 번호로 식별.
    start_num/end_num: '1', '3', '3-1', '5-1' 등 문자열.
    A열 값이 정수/문자열 모두 가능 → str() 정규화.
    """
    sorted_rows = sorted(rows_data.keys())

    def _row_section_num(r):
        a = rows_data[r].get(1)
        if a is None:
            return None
        s = str(a).strip().rstrip('.').strip()  # '5.' → '5'
        return s if s else None

    start_r = None
    for r in sorted_rows:
        if _row_section_num(r) == start_num:
            start_r = r
            break
    if start_r is None:
        return None, None, sorted_rows
    if end_num is None:
        return start_r, None, sorted_rows
    end_r = None
    for r in sorted_rows:
        if r <= start_r:
            continue
        if _row_section_num(r) == end_num:
            end_r = r
            break
    return start_r, end_r, sorted_rows


def extract_tx_deferred_tax_changes(file_path, sheet_name='TX',
                                     max_row_scan=120):
    """TX 1. 이연법인세자산(부채) 증감내용.

    각 항목 행: C열=한글 계정명, D열=Current/Non-Current,
                E열=기초, F열=기말, G열=증감.
    'Total' 행에서 종료.

    법정세율(2번 항목): 섹션 2 행(보통 41행)의 D열. 분율(예 0.22) 저장.

    반환: {'sheet_found','rows':[{kor_label, current_flag,
                                  beginning, ending, change}],
            'statutory_rate': float|None, 'error'}
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name,
                                       max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'rows': [],
                'statutory_rate': None, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'rows': [],
                'statutory_rate': None, 'error': None}

    start_r, end_r, sorted_rows = _tx_find_section_bounds(
        rows_data, _TX_S1_NUM, _TX_S2_NUM)
    if start_r is None:
        return {'sheet_found': True, 'rows': [], 'statutory_rate': None,
                'error': '1번 이연법인세 증감 섹션을 찾지 못함'}
    sec2_row = end_r  # 섹션 2(법정세율) 행 — D열에 세율 (fallback 적용 전 보존)
    if end_r is None:
        end_r = start_r + 40

    rows = []
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        b = v.get(2)
        # Total → main table 끝 (이후 sub-table 등 skip)
        if isinstance(b, str) and b.strip() == 'Total':
            break
        # 헤더 행: B='Eng.' or C='Kor.'
        if isinstance(b, str) and b.strip() in ('Eng.', 'Account'):
            continue
        c = v.get(3)  # C열 한글
        if not isinstance(c, str) or not c.strip():
            continue
        kor = c.strip()
        d = v.get(4)
        cur_flag = (d.strip() if isinstance(d, str) else '') or ''
        beg = _to_float(v.get(5))
        end = _to_float(v.get(6))
        chg = _to_float(v.get(7))
        # 모두 None/0이면 skip
        if (beg in (None, 0)) and (end in (None, 0)) and (chg in (None, 0)):
            continue
        rows.append({
            'kor_label':    kor,
            'current_flag': cur_flag,
            'beginning':    beg or 0,
            'ending':       end or 0,
            'change':       chg or 0,
        })
    statutory_rate = (_to_float(rows_data.get(sec2_row, {}).get(4))
                      if sec2_row else None)
    return {'sheet_found': True, 'rows': rows,
            'statutory_rate': statutory_rate, 'error': None}


def extract_tx_income_tax_breakdown(file_path, sheet_name='TX',
                                     max_row_scan=120):
    """TX 3. 법인세비용의 구성내역. 4개 항목 + 합계 (법인세비용)."""
    rows_data, err = _load_sheet_rows(file_path, sheet_name,
                                       max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _tx_find_section_bounds(
        rows_data, _TX_S3_NUM, _TX_S31_NUM)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '3번 법인세비용 구성내역 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    items = {key: None for key in _TX_S3_ITEMS.values()}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        key = _TX_S3_ITEMS.get(b.strip())
        if not key:
            continue
        items[key] = _to_float(rows_data[r].get(4))  # D열 금액
    return {'sheet_found': True, 'items': items, 'error': None}


def extract_tx_equity_deferred_tax(file_path, sheet_name='TX',
                                    max_row_scan=120):
    """TX 3-1. 자본 직접 부과 이연법인세 변동액 (8개 항목 + 합계)."""
    rows_data, err = _load_sheet_rows(file_path, sheet_name,
                                       max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _tx_find_section_bounds(
        rows_data, _TX_S31_NUM, _TX_S4_NUM)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '3-1번 자본 직접 부과 이연법인세 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    items = {key: None for key in _TX_S31_ITEMS.values()}
    total = None
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        s = b.strip()
        if s == 'Total':
            total = _to_float(rows_data[r].get(4))
            continue
        key = _TX_S31_ITEMS.get(s)
        if not key:
            continue
        items[key] = _to_float(rows_data[r].get(4))  # D열 금액
    items['__total'] = total
    return {'sheet_found': True, 'items': items, 'error': None}


def extract_tx_reconciliation(file_path, sheet_name='TX',
                               max_row_scan=120):
    """TX 4. Reconciliation. 항목별 금액 + 유효세율 (비율) 별도."""
    rows_data, err = _load_sheet_rows(file_path, sheet_name,
                                       max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'items': {},
                'effective_rate': None, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {},
                'effective_rate': None, 'error': None}

    start_r, end_r, sorted_rows = _tx_find_section_bounds(
        rows_data, _TX_S4_NUM, _TX_S5_NUM)
    if start_r is None:
        return {'sheet_found': True, 'items': {}, 'effective_rate': None,
                'error': '4번 reconciliation 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    items = {key: None for key in _TX_S4_ITEMS.values()}
    effective_rate = None
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        s = b.strip()
        if s == _TX_S4_RATE_LABEL:
            effective_rate = _to_float(rows_data[r].get(4))
            continue
        key = _TX_S4_ITEMS.get(s)
        if not key:
            continue
        items[key] = _to_float(rows_data[r].get(4))
    return {'sheet_found': True, 'items': items,
            'effective_rate': effective_rate, 'error': None}


def extract_tx_unrecognized_temp_diff(file_path, sheet_name='TX',
                                       max_row_scan=120):
    """TX 5. 이연법인세로 인식되지 않은 일시적차이 (이월결손금/기타)."""
    rows_data, err = _load_sheet_rows(file_path, sheet_name,
                                       max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'items': {}, 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'items': {}, 'error': None}

    start_r, end_r, sorted_rows = _tx_find_section_bounds(
        rows_data, _TX_S5_NUM, _TX_S51_NUM)
    if start_r is None:
        return {'sheet_found': True, 'items': {},
                'error': '5번 미인식 일시적차이 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 10

    items = {k: None for k, _ in _TX_S5_ITEMS_KEYS}
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        s = b.strip()
        for key, kw in _TX_S5_ITEMS_KEYS:
            if kw in s:
                items[key] = _to_float(rows_data[r].get(3))  # C열 금액
                break
    return {'sheet_found': True, 'items': items, 'error': None}


def extract_tx_loss_carryforward_maturity(file_path, sheet_name='TX',
                                            max_row_scan=120):
    """TX 5-1. 이연법인세자산을 인식하지 아니한 이월결손금 만기.

    만기 라벨이 회계연도별로 다름 (2026/2027/.../After 2030).
    상대 위치(N+1, N+2, N+3, N+4, After N+5)로 정규화.
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name,
                                       max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'buckets': [], 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'buckets': [], 'error': None}

    start_r, end_r, sorted_rows = _tx_find_section_bounds(
        rows_data, _TX_S51_NUM, _TX_S6_NUM)
    if start_r is None:
        return {'sheet_found': True, 'buckets': [],
                'error': '5-1번 이월결손금 만기 섹션을 찾지 못함'}
    if end_r is None:
        end_r = start_r + 15

    buckets = []  # ordered [{label, amount}]
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        b = v.get(2)
        if b is None:
            continue
        # B는 연도(int) 또는 'After XXXX' 문자열일 수 있음
        if isinstance(b, (int, float)):
            s = str(int(b))
        else:
            s = str(b).strip()
        if not s or s in ('Total', 'Diff', '합계') or s.startswith('만기'):
            continue
        # 연도 형식 또는 'After...' 만 허용
        if not (re.fullmatch(r'\d{4}', s) or s.lower().startswith('after')):
            continue
        amount = _to_float(v.get(3)) or 0
        buckets.append({'label': s, 'amount': amount})

    return {'sheet_found': True, 'buckets': buckets, 'error': None}


def extract_l2_maturity_analysis(file_path, sheet_name='L2', max_row_scan=80):
    """L2 시트의 3. 부채성 금융상품의 만기 분석내역 추출.

    구조:
      A58 = '3.' (섹션 시작)
      r61 헤더: B=Account, C=대주구분, D=1년 미만, E=1년~2년,
                F=2년~5년, G=5년 초과, H=Total
      r62~ : 데이터 (한 항목당 한글/영문 2행씩 — 영문 행은 다른 대주구분)
      B='Total' 행 → 종료

    Account가 두 줄(한글/영문)로 되어있을 때 사용자 요청에 따라
    **위쪽 한글 계정명만 사용** (영문 행은 위쪽 한글 anchor 유지).

    반환: {
      'sheet_found': bool,
      'rows': [{'account','creditor_type','within_1y','within_2y',
                'within_5y','over_5y','total'}],
      'error': str | None,
    }
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'rows': [], 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'rows': [], 'error': None}

    sorted_rows = sorted(rows_data.keys())
    start_r = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if isinstance(a, str) and _L2_S3_RE.match(a.strip()):
            start_r = r
            break
    if start_r is None:
        return {'sheet_found': True, 'rows': [],
                'error': '3번 만기 분석 섹션을 찾지 못함'}

    # 종료: B열 'Total' 행
    end_r = None
    for r in sorted_rows:
        if r <= start_r:
            continue
        b = rows_data[r].get(2)
        if isinstance(b, str) and b.strip().lower() == 'total':
            end_r = r
            break
    if end_r is None:
        end_r = start_r + 25

    def _first_line(x):
        if not isinstance(x, str):
            return x
        return x.split('\n', 1)[0].strip()

    # 한글 계정명 anchor (영문 행에서도 위쪽 한글 라벨 유지)
    last_account_kor = ''
    rows = []
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        b = v.get(2)

        # B에 값이 있고 한글이면 anchor 갱신
        if isinstance(b, str) and b.strip():
            bs = b.strip()
            if _has_korean(bs) or not last_account_kor:
                last_account_kor = bs
            # 영문 라벨(Borrowings/Debentures 등)이면 last 유지

        # H(Total)이 숫자이고 0이 아닌 행만 채택
        h_num = _to_float(v.get(8))
        if h_num is None or h_num == 0:
            continue

        rows.append({
            'account':       last_account_kor,
            'creditor_type': _first_line(v.get(3)) or '',
            'within_1y':     _to_float(v.get(4)) or 0.0,
            'within_2y':     _to_float(v.get(5)) or 0.0,
            'within_5y':     _to_float(v.get(6)) or 0.0,
            'over_5y':       _to_float(v.get(7)) or 0.0,
            'total':         h_num,
        })

    return {'sheet_found': True, 'rows': rows, 'error': None}


def extract_l2_verification(file_path, sheet_name='L2', max_row_scan=200):
    """L2 시트의 verification 테이블에서 G열 'Y' 행 추출.

    구조:
      r? = B열 'Verification' 라벨 (보통 r81)
      r?+1 = 헤더 (B=계정명, C=종류, D=대주구분, E=금액, F=재무제표, G=오류, H=사유)
      r?+2 이후 = 데이터 행

    셀병합 처리:
      · B/C 열은 셀병합되어 일부 행이 비어 있을 수 있음 (fill-down)
      · B열은 같은 항목에 대해 한글/영문 두 anchor가 따로 있음
        (예: '차입금'과 'Borrowings'). 사용자 가독성을 위해 **한글 우선** —
        영문 라벨이 들어와도 한글 anchor 유지.

    반환: {sheet_found, verify_row, y_rows[{row,account,sub_type,creditor,
                                            amount,book,alert,reason}], error}
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=8)
    if err:
        return {'sheet_found': False, 'verify_row': None, 'y_rows': [],
                'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'verify_row': None, 'y_rows': [],
                'error': None}

    sorted_rows = sorted(rows_data.keys())
    verify_r = None
    for r in sorted_rows:
        b = rows_data[r].get(2)
        if isinstance(b, str) and _L2_VERIFY_RE.match(b.strip()):
            verify_r = r
            break

    if verify_r is None:
        return {'sheet_found': True, 'verify_row': None, 'y_rows': [],
                'error': 'Verification 라벨을 찾지 못함'}

    def _first_line(x):
        if not isinstance(x, str):
            return x
        return x.split('\n', 1)[0].strip()

    # 셀병합 fill-down용 anchor 추적
    last_account = ''
    last_sub_type = ''
    y_rows = []
    for r in sorted_rows:
        if r <= verify_r + 1:   # 라벨 행 + 헤더 행 스킵
            continue
        v = rows_data[r]
        b = v.get(2)
        c = v.get(3)

        # B(계정명) anchor 업데이트 — 한글 라벨 우선
        if isinstance(b, str) and b.strip():
            bs = b.strip()
            # 한글이 있거나, anchor가 아직 없으면 갱신
            if _has_korean(bs) or not last_account:
                last_account = bs
            # 영문 라벨(Borrowings 등)이고 한글 anchor 있으면 유지
        # C(종류) anchor 업데이트
        if isinstance(c, str) and c.strip():
            last_sub_type = _first_line(c) or ''

        g = v.get(7)
        if not isinstance(g, str) or g.strip().upper() != 'Y':
            continue

        y_rows.append({
            'row':      r,
            'account':  last_account,
            'sub_type': last_sub_type,
            'creditor': _first_line(v.get(4)) or '',
            'amount':   _to_float(v.get(5)),
            'book':     _to_float(v.get(6)),
            'alert':    g.strip().upper(),
            'reason':   str(v.get(8) or '').strip() if v.get(8) else '',
        })

    return {'sheet_found': True, 'verify_row': verify_r,
            'y_rows': y_rows, 'error': None}


def extract_a6_derivatives(file_path, sheet_name='A6', max_row_scan=80):
    """A6 시트의 1. 파생상품평가손익 내역 추출.

    구조 (행 추가에 안전):
      A열 '1. 파생상품' prefix → 섹션 시작
      B열 'Total' → 섹션 종료
      사이 영역: B=파생상품 종류, C=평가이익, D=평가손실
        C 또는 D 중 하나 이상이 숫자인 행 채택

    반환: {
      'sheet_found': bool,
      'rows': [{'type','gain','loss'}],
      'error': str | None,
    }
    """
    rows_data, err = _load_sheet_rows(file_path, sheet_name, max_row_scan, max_col=4)
    if err:
        return {'sheet_found': False, 'rows': [], 'error': err}
    if rows_data is None:
        return {'sheet_found': False, 'rows': [], 'error': None}

    sorted_rows = sorted(rows_data.keys())
    start_r = None
    for r in sorted_rows:
        a = rows_data[r].get(1)
        if isinstance(a, str) and _A6_S1_RE.match(a.strip()):
            start_r = r
            break
    if start_r is None:
        return {'sheet_found': True, 'rows': [],
                'error': '1번 파생상품 섹션을 찾지 못함'}

    # 종료: B열 'Total' 행
    end_r = None
    for r in sorted_rows:
        if r <= start_r:
            continue
        b = rows_data[r].get(2)
        if isinstance(b, str) and b.strip().lower() == 'total':
            end_r = r
            break
    if end_r is None:
        end_r = start_r + 20

    rows_out = []
    for r in sorted_rows:
        if r <= start_r or r >= end_r:
            continue
        v = rows_data[r]
        gain = _to_float(v.get(3))
        loss = _to_float(v.get(4))
        # 둘 다 None/0 이면 빈 행 → 스킵
        if (gain is None or gain == 0) and (loss is None or loss == 0):
            continue
        b = v.get(2)
        # 헤더 라벨 행 자동 스킵 (B에 '\n'이 포함된 다국어 헤더)
        if isinstance(b, str) and '\n' in b:
            continue
        rows_out.append({
            'type': str(b or '').strip(),
            'gain': gain or 0.0,
            'loss': loss or 0.0,
        })

    return {'sheet_found': True, 'rows': rows_out, 'error': None}


def extract_l4_other_commitments(file_path, sheet_name='L4', max_row_scan=500):
    """L4 시트의 11번 (그외 우발부채 및 약정사항) — 자유 텍스트 수집.

    구조:
      · B[r11] = "11. 그외 우발부채 및 약정사항을 아래에 자유롭게 기술해 주세요"
      · C[r11+1] = "Freely Describe any other contingencies and commitments below"
        (영문 안내 라벨 — 모든 회사 공통, 자동 제외)
      · r11+2 이후 행에 회사가 자유롭게 작성한 텍스트

    합산 아닌 단순 수집 — content가 있는 회사만 합산 결과에 포함.

    반환: {
      'sheet_found': bool,
      'content':     str | None,   # 자유 텍스트 (라벨 제외, 줄바꿈으로 join)
      'error':       str | None,
    }
    """
    LABEL_KEYWORDS = (
        'freely describe', 'describe any other',
        'have there been', 'if so',
        '있다면', '내용을', '자유롭게',
    )
    try:
        zf = zipfile.ZipFile(file_path)
    except Exception as e:
        return {'sheet_found': False, 'content': None,
                'error': f'파일 열기 실패: {e}'}

    try:
        sheet_path = _find_sheet_path(zf, sheet_name)
        if not sheet_path or sheet_path not in zf.namelist():
            return {'sheet_found': False, 'content': None, 'error': None}

        shared = _load_shared_strings(zf)
        rows_data = {}
        with zf.open(sheet_path) as f:
            for event, elem in ET.iterparse(f, events=('end',)):
                if elem.tag != _TAG_ROW:
                    continue
                r_attr = elem.get('r')
                if not r_attr:
                    elem.clear(); continue
                r = int(r_attr)
                if r > max_row_scan:
                    elem.clear(); continue
                vals = {}
                for c in elem.findall(_TAG_C):
                    letters, _ = _split_cell_ref(c.get('r'))
                    if not letters:
                        continue
                    col_idx = _col_letters_to_index(letters)
                    if col_idx > 8:   # A~H
                        continue
                    vals[col_idx] = _cell_value(c, shared)
                if vals:
                    rows_data[r] = vals
                elem.clear()
    finally:
        zf.close()

    if not rows_data:
        return {'sheet_found': True, 'content': None, 'error': None}

    sorted_rows = sorted(rows_data.keys())

    _Q11 = re.compile(r'^\s*11\s*\.(?!\d)')
    _Q12 = re.compile(r'^\s*12\s*\.(?!\d)')
    r11 = r12 = None
    for r in sorted_rows:
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        bs = b.strip()
        if r11 is None and _Q11.match(bs):
            r11 = r
        elif r11 is not None and r12 is None and _Q12.match(bs):
            r12 = r
            break

    if r11 is None:
        return {'sheet_found': True, 'content': None,
                'error': '11번 항목을 찾지 못함'}

    upper = r12 if r12 is not None else (sorted_rows[-1] + 1)

    content_lines = []
    for r in sorted_rows:
        if r <= r11 or r >= upper:
            continue
        c_val = rows_data[r].get(3)
        if not isinstance(c_val, str):
            continue
        s = c_val.strip()
        if not s:
            continue
        s_low = s.lower()
        if any(k in s_low for k in LABEL_KEYWORDS):
            continue
        content_lines.append(s)

    content = '\n'.join(content_lines) if content_lines else None
    return {'sheet_found': True, 'content': content, 'error': None}


def extract_l4_subsequent_events(file_path, sheet_name='L4', max_row_scan=400):
    """L4 시트의 10번 (보고기간일 이후 발생 중요 사건) 추출.

    합산 아닌 단순 응답 수집:
      · B[r10] = "10." 라벨 행 → G[r10] = YES/NO  (F가 아닌 G열에 응답)
      · YES인 경우: r10 ~ r11 사이 C열의 자유 텍스트들 (라벨 제외) 수집

    반환: {
      'sheet_found': bool,
      'yn':       'YES'/'NO'/None,
      'content':  str | None,   # YES일 때만 의미 있음 (라벨 제외 자유 텍스트)
      'error':    str | None,
    }
    """
    # 라벨로 인식해 제외할 패턴 (자유 텍스트가 아님)
    LABEL_KEYWORDS = (
        '있다면', 'if so', 'have there been',
        'describe any other contingencies',
    )
    try:
        zf = zipfile.ZipFile(file_path)
    except Exception as e:
        return {'sheet_found': False, 'yn': None, 'content': None,
                'error': f'파일 열기 실패: {e}'}

    try:
        sheet_path = _find_sheet_path(zf, sheet_name)
        if not sheet_path or sheet_path not in zf.namelist():
            return {'sheet_found': False, 'yn': None, 'content': None,
                    'error': None}

        shared = _load_shared_strings(zf)
        rows_data = {}
        with zf.open(sheet_path) as f:
            for event, elem in ET.iterparse(f, events=('end',)):
                if elem.tag != _TAG_ROW:
                    continue
                r_attr = elem.get('r')
                if not r_attr:
                    elem.clear(); continue
                r = int(r_attr)
                if r > max_row_scan:
                    elem.clear(); continue
                vals = {}
                for c in elem.findall(_TAG_C):
                    letters, _ = _split_cell_ref(c.get('r'))
                    if not letters:
                        continue
                    col_idx = _col_letters_to_index(letters)
                    if col_idx > 8:   # A~H만
                        continue
                    vals[col_idx] = _cell_value(c, shared)
                if vals:
                    rows_data[r] = vals
                elem.clear()
    finally:
        zf.close()

    if not rows_data:
        return {'sheet_found': True, 'yn': None, 'content': None, 'error': None}

    sorted_rows = sorted(rows_data.keys())

    # 10. ~ 11. 행 위치 탐색
    _Q10 = re.compile(r'^\s*10\s*\.(?!\d)')
    _Q11 = re.compile(r'^\s*11\s*\.(?!\d)')
    r10 = r11 = None
    for r in sorted_rows:
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        bs = b.strip()
        if r10 is None and _Q10.match(bs):
            r10 = r
        elif r10 is not None and r11 is None and _Q11.match(bs):
            r11 = r
            break

    if r10 is None:
        return {'sheet_found': True, 'yn': None, 'content': None,
                'error': '10번 항목을 찾지 못함'}

    # G[r10] = Yes/No 응답
    g10 = rows_data.get(r10, {}).get(7)
    yn = (str(g10).strip().upper()
          if isinstance(g10, str) and g10.strip() else None)

    # YES인 경우 r10~r11 사이 C열 자유 텍스트 수집 (라벨 제외)
    content_lines = []
    upper = r11 if r11 is not None else (r10 + 20)
    for r in sorted_rows:
        if r <= r10 or r >= upper:
            continue
        c_val = rows_data[r].get(3)
        if not isinstance(c_val, str):
            continue
        s = c_val.strip()
        if not s:
            continue
        # 라벨 키워드 포함 시 제외
        s_low = s.lower()
        if any(k in s_low for k in LABEL_KEYWORDS):
            continue
        content_lines.append(s)

    content = '\n'.join(content_lines) if content_lines else None

    return {'sheet_found': True, 'yn': yn, 'content': content, 'error': None}


def extract_l4_pledged_assets(file_path, sheet_name='L4', max_row_scan=200):
    """L4 시트의 9-1 (담보제공자산 내용) 섹션 명세.

    컬럼: C=채권자/저당권자, D=계정과목(담보자산), E=통화,
          F=금액, G=계정과목(관련부채), H=설명
    """
    return _extract_l4_table_section(
        file_path, sheet_name,
        start_re=_L4_Q91_RE, end_re=_L4_Q10_RE,
        col_map={3: 'creditor', 4: 'asset_account', 5: 'currency',
                 6: 'amount',  7: 'liability_account', 8: 'description'},
        max_col=8,
        not_found_msg='9-1 섹션을 찾지 못함',
        max_row_scan=max_row_scan,
    )


def _extract_l4_qna(file_path, sheet_name, q_re, qn_re, end_re,
                    not_found_msg, max_row_scan=200):
    """L4 시트의 'Yes/No + 금액(통화)' 패턴 섹션 공통 추출 헬퍼.

    구조 (수입신용장 2/2-1, 수출채권 3/3-1 등 동일 패턴):
      · q_re   매칭 행(r_q)        → F[r_q]   = Yes/No
      · qn_re  매칭 행(r_qn)       → r_qn+1 ~ end_r-1 구간에서 F열 첫 숫자 행:
                                       F = 금액, G = 통화
      · end_re 매칭 행(end_r)      → 종료점

    반환: {
      'sheet_found': bool,
      'yn':         'YES'/'NO'/원문 | None,   (대문자 정규화)
      'amount':     float | None,
      'currency':   str | None,
      'error':      str | None,
    }
    """
    try:
        zf = zipfile.ZipFile(file_path)
    except Exception as e:
        return {'sheet_found': False, 'yn': None, 'amount': None,
                'currency': None, 'error': f'파일 열기 실패: {e}'}

    try:
        sheet_path = _find_sheet_path(zf, sheet_name)
        if not sheet_path or sheet_path not in zf.namelist():
            return {'sheet_found': False, 'yn': None, 'amount': None,
                    'currency': None, 'error': None}

        shared = _load_shared_strings(zf)

        rows_data = {}
        with zf.open(sheet_path) as f:
            for event, elem in ET.iterparse(f, events=('end',)):
                if elem.tag != _TAG_ROW:
                    continue
                r_attr = elem.get('r')
                if not r_attr:
                    elem.clear(); continue
                r = int(r_attr)
                if r > max_row_scan:
                    elem.clear(); continue
                vals = {}
                for c in elem.findall(_TAG_C):
                    letters, _ = _split_cell_ref(c.get('r'))
                    if not letters:
                        continue
                    col_idx = _col_letters_to_index(letters)
                    if col_idx > 7:   # A~G만
                        continue
                    vals[col_idx] = _cell_value(c, shared)
                if vals:
                    rows_data[r] = vals
                elem.clear()
    finally:
        zf.close()

    if not rows_data:
        return {'sheet_found': True, 'yn': None, 'amount': None,
                'currency': None, 'error': None}

    sorted_rows = sorted(rows_data.keys())

    r_q = r_qn = end_r = None
    for r in sorted_rows:
        b = rows_data[r].get(2)
        if not isinstance(b, str):
            continue
        bs = b.strip()
        if r_q is None and q_re.match(bs) and not qn_re.match(bs):
            r_q = r
        elif r_qn is None and qn_re.match(bs):
            r_qn = r
        elif end_r is None and end_re.match(bs):
            end_r = r
            break

    if r_q is None:
        return {'sheet_found': True, 'yn': None, 'amount': None,
                'currency': None, 'error': not_found_msg}

    # F[r_q] = Yes/No
    fq = rows_data.get(r_q, {}).get(6)
    yn = str(fq).strip().upper() if isinstance(fq, str) and fq.strip() else None

    # 금액/통화: r_qn+1 ~ end_r-1 사이 F열 숫자 첫 행
    amount = None
    currency = None
    if r_qn is not None:
        upper = end_r if end_r is not None else (r_qn + 10)
        for r in sorted_rows:
            if r <= r_qn or r >= upper:
                continue
            v = rows_data[r]
            f = v.get(6)
            if isinstance(f, (int, float)) and not isinstance(f, bool):
                amount = float(f)
                g = v.get(7)
                currency = (str(g).strip().upper() if g not in (None, '') else None)
                break

    return {'sheet_found': True, 'yn': yn,
            'amount': amount, 'currency': currency, 'error': None}


def extract_l4_lc(file_path, sheet_name='L4', max_row_scan=200):
    """L4 시트의 2번(수입신용장 오픈 여부) + 2-1(미확정 지급보증 실행금액).

    1-1 테이블 행 증가로 r2/r21이 28/31에서 밀려나도 B열 라벨로 동적 탐색.

    반환: {sheet_found, lc_open, amount, currency, error}
    """
    r = _extract_l4_qna(file_path, sheet_name,
                        _L4_Q2_RE, _L4_Q21_RE, _L4_Q3_RE,
                        not_found_msg='2번 항목을 찾지 못함',
                        max_row_scan=max_row_scan)
    return {
        'sheet_found': r['sheet_found'],
        'lc_open':     r['yn'],
        'amount':      r['amount'],
        'currency':    r['currency'],
        'error':       r['error'],
    }


def extract_l4_export(file_path, sheet_name='L4', max_row_scan=200):
    """L4 시트의 3번(수출채권 할인 여부) + 3-1(만기 미도래 할인금액).

    1-1/2/2-1 영역 행 증가로 r3/r31이 34/37에서 밀려나도 B열 라벨로 동적 탐색.

    반환: {sheet_found, discount_done, amount, currency, error}
    """
    r = _extract_l4_qna(file_path, sheet_name,
                        _L4_Q3_RE, _L4_Q31_RE, _L4_Q4_RE,
                        not_found_msg='3번 항목을 찾지 못함',
                        max_row_scan=max_row_scan)
    return {
        'sheet_found':   r['sheet_found'],
        'discount_done': r['yn'],
        'amount':        r['amount'],
        'currency':      r['currency'],
        'error':         r['error'],
    }


# ──────────────────────────────────────────────────────────────
# 엑셀 빌더 (L1 단기차입금 합산 결과)
# ──────────────────────────────────────────────────────────────

_XL_HDR_FILL  = PatternFill('solid', start_color='1F3864')
_XL_HDR_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=11)
_XL_TITLE     = Font(bold=True, color='1F3864', name='Arial', size=14)
_XL_SEC_FILL  = PatternFill('solid', start_color='8FAADC')
_XL_SEC_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=11)
_XL_DATA      = Font(name='Arial', size=10)
_XL_TOTAL_FILL = PatternFill('solid', start_color='FFE699')
_XL_TOTAL_FONT = Font(bold=True, name='Arial', size=10, color='9C5700')
_XL_NUM_FMT   = '#,##0;(#,##0);"-"'
_XL_RATE_FMT  = '0.0000'

_XL_COLS = [
    ('회사',       28),
    ('대주구분',    18),
    ('대주명',      28),
    ('이자율',      10),
    ('통화',         8),
    ('로컬 금액',   16),
    ('Spot',        10),
    ('KRW 환산',   18),
]


def _write_l1_section(ws, row, category):
    """한 종류 블록을 ws에 그리고 다음 시작 row 반환."""
    # 섹션 헤더
    name = category.get('name') or ''
    key  = category.get('key') or ''
    code = category.get('code') or ''
    title = f"{key}. {name}" + (f"  [{code}]" if code else '')
    c = ws.cell(row, 1, title)
    c.font = _XL_SEC_FONT
    c.fill = _XL_SEC_FILL
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(_XL_COLS))
    ws.row_dimensions[row].height = 20
    row += 1

    # 컬럼 헤더
    for j, (label, _w) in enumerate(_XL_COLS, 1):
        cell = ws.cell(row, j, label)
        cell.font = _XL_HDR_FONT
        cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    # 데이터 행
    rows = category.get('rows') or []
    for r in rows:
        ws.cell(row, 1, r.get('company') or '').font = _XL_DATA
        ws.cell(row, 2, r.get('creditor_type') or '').font = _XL_DATA
        ws.cell(row, 3, r.get('creditor') or '').font = _XL_DATA
        rate = r.get('rate')
        rc = ws.cell(row, 4, rate if isinstance(rate, (int, float)) else (rate or ''))
        rc.font = _XL_DATA
        if isinstance(rate, (int, float)):
            rc.number_format = _XL_RATE_FMT
        ws.cell(row, 5, r.get('currency') or '').font = _XL_DATA
        lc = ws.cell(row, 6, r.get('local') or 0)
        lc.font = _XL_DATA; lc.number_format = _XL_NUM_FMT
        sc = ws.cell(row, 7, r.get('spot') or 0)
        sc.font = _XL_DATA; sc.number_format = _XL_RATE_FMT
        kc = ws.cell(row, 8, r.get('krw') or 0)
        kc.font = _XL_DATA; kc.number_format = _XL_NUM_FMT
        row += 1

    # 소계 행
    total_krw = category.get('total_krw') or 0
    by_cur = category.get('total_local_by_currency') or {}
    local_note = ' · '.join(f'{cur} {amt:,.0f}' for cur, amt in sorted(by_cur.items()))
    for j in range(1, len(_XL_COLS) + 1):
        ws.cell(row, j).fill = _XL_TOTAL_FILL
    ws.cell(row, 1, f'{key}. 소계').font = _XL_TOTAL_FONT
    ws.cell(row, 5, local_note).font = _XL_TOTAL_FONT
    tc = ws.cell(row, 8, total_krw)
    tc.font = _XL_TOTAL_FONT
    tc.number_format = _XL_NUM_FMT
    row += 2   # 섹션 사이 공백
    return row


def _write_l1_cover(wb, agg_data):
    """표지 시트."""
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 60

    ws.cell(1, 1, 'L1 단기차입금 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '종류 수').font = _XL_DATA
    ws.cell(5, 2, len(agg_data.get('categories') or [])).font = _XL_DATA
    ws.cell(6, 1, '전체 KRW 합계').font = _XL_DATA
    grand = ws.cell(6, 2, agg_data.get('grand_total_krw') or 0)
    grand.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    grand.number_format = _XL_NUM_FMT

    # 종류별 요약 표
    ws.cell(8, 1, '종류별 KRW 합계').font = Font(bold=True, color='1F3864', size=11)
    headers = ['번호', '종류', '계정코드', 'KRW 합계']
    for j, h in enumerate(headers, 1):
        c = ws.cell(9, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
        c.alignment = Alignment(horizontal='center')
    r = 10
    for cat in (agg_data.get('categories') or []):
        ws.cell(r, 1, cat.get('key') or '').font = _XL_DATA
        ws.cell(r, 2, cat.get('name') or '').font = _XL_DATA
        ws.cell(r, 3, cat.get('code') or '').font = _XL_DATA
        cc = ws.cell(r, 4, cat.get('total_krw') or 0)
        cc.font = _XL_DATA; cc.number_format = _XL_NUM_FMT
        r += 1

    # 경고: 누락된 회사
    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L1 시트 누락/오류 회사').font = Font(bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1


def _write_l1_combined_sheet(wb, agg_data):
    """모든 종류를 단일 시트에 적층."""
    ws = wb.create_sheet('상세')
    for j, (_lbl, w) in enumerate(_XL_COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A1'

    row = 1
    for cat in (agg_data.get('categories') or []):
        row = _write_l1_section(ws, row, cat)

    # 그랜드 토탈
    if agg_data.get('categories'):
        for j in range(1, len(_XL_COLS) + 1):
            ws.cell(row, j).fill = _XL_TOTAL_FILL
        ws.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        gc = ws.cell(row, 8, agg_data.get('grand_total_krw') or 0)
        gc.font = _XL_TOTAL_FONT
        gc.number_format = _XL_NUM_FMT


def build_l1_excel(agg_data, output_path):
    """L1 단기차입금 합산 결과(JSON 구조)를 엑셀로 저장.

    agg_data: admin_note_aggregate_l1 라우트가 만드는 dict
      {
        'year', 'scanned', 'grand_total_krw',
        'categories': [{'key','name','code','rows':[...],
                        'total_krw', 'total_local_by_currency'}],
        'errors': [{'company','reason'}],
      }
    """
    wb = Workbook()
    wb.remove(wb.active)
    _write_l1_cover(wb, agg_data)
    _write_l1_combined_sheet(wb, agg_data)
    wb.save(str(output_path))
    return str(output_path)


# ──────────────────────────────────────────────────────────────
# 엑셀 빌더 (L4 대출한도 약정 합산 결과)
# ──────────────────────────────────────────────────────────────

_XL_L4_COLS = [
    ('회사',         28),
    ('종류',         28),
    ('금융기관',     32),
    ('통화',          8),
    ('로컬 금액',    18),
    ('Spot',         10),
    ('KRW 환산',    20),
]


def _write_l4_cover(wb, agg_data):
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 60

    ws.cell(1, 1, 'L4 대출한도 약정 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '합산 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_rows') or 0).font = _XL_DATA
    ws.cell(6, 1, '전체 KRW 합계').font = _XL_DATA
    grand = ws.cell(6, 2, agg_data.get('grand_total_krw') or 0)
    grand.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    grand.number_format = _XL_NUM_FMT

    # 통화별 로컬 합계
    ws.cell(8, 1, '통화별 로컬 합계').font = Font(bold=True, color='1F3864', size=11)
    headers = ['통화', '로컬 합계', 'KRW 합계']
    for j, h in enumerate(headers, 1):
        c = ws.cell(9, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
        c.alignment = Alignment(horizontal='center')
    by_cur = agg_data.get('total_by_currency') or {}
    r = 10
    for cur in sorted(by_cur.keys()):
        ws.cell(r, 1, cur).font = _XL_DATA
        lc = ws.cell(r, 2, by_cur[cur].get('local') or 0)
        lc.font = _XL_DATA; lc.number_format = _XL_NUM_FMT
        kc = ws.cell(r, 3, by_cur[cur].get('krw') or 0)
        kc.font = _XL_DATA; kc.number_format = _XL_NUM_FMT
        r += 1

    # 경고
    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L4 시트 누락/오류 회사').font = Font(bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1


def _write_l4_combined_sheet(wb, agg_data):
    ws = wb.create_sheet('상세')
    for j, (_lbl, w) in enumerate(_XL_L4_COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'

    # 헤더
    for j, (label, _w) in enumerate(_XL_L4_COLS, 1):
        cell = ws.cell(1, j, label)
        cell.font = _XL_HDR_FONT
        cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        ws.cell(row, 1, r.get('company') or '').font = _XL_DATA
        ws.cell(row, 2, r.get('type') or '').font = _XL_DATA
        ws.cell(row, 3, r.get('institution') or '').font = _XL_DATA
        ws.cell(row, 4, r.get('currency') or '').font = _XL_DATA
        lc = ws.cell(row, 5, r.get('local') or 0)
        lc.font = _XL_DATA; lc.number_format = _XL_NUM_FMT
        sc = ws.cell(row, 6, r.get('spot') or 0)
        sc.font = _XL_DATA; sc.number_format = _XL_RATE_FMT
        kc = ws.cell(row, 7, r.get('krw') or 0)
        kc.font = _XL_DATA; kc.number_format = _XL_NUM_FMT
        row += 1

    # 그랜드 토탈
    if agg_data.get('rows'):
        for j in range(1, len(_XL_L4_COLS) + 1):
            ws.cell(row, j).fill = _XL_TOTAL_FILL
        ws.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        gc = ws.cell(row, 7, agg_data.get('grand_total_krw') or 0)
        gc.font = _XL_TOTAL_FONT
        gc.number_format = _XL_NUM_FMT


def build_l4_excel(agg_data, output_path):
    """L4 대출한도 약정 합산 결과(JSON 구조)를 엑셀로 저장.

    agg_data: admin_note_aggregate_l4 라우트가 만드는 dict
      {
        'year', 'scanned', 'with_rows',
        'rows': [{'company','type','institution','currency','local','spot','krw'}],
        'grand_total_krw', 'total_by_currency': {cur: {'local','krw'}},
        'errors': [{'company','reason'}],
      }
    """
    wb = Workbook()
    wb.remove(wb.active)
    _write_l4_cover(wb, agg_data)
    _write_l4_combined_sheet(wb, agg_data)
    wb.save(str(output_path))
    return str(output_path)


# ──────────────────────────────────────────────────────────────
# 엑셀 빌더 (L4 다행 명세 generic)
# 5종 (4-1·5-2·7-1·8-1·8-2·9-1)이 동일한 표지 + 컬럼 정의만 다른 상세 시트.
# ──────────────────────────────────────────────────────────────


def _write_l4_table_cover(wb, agg_data, title):
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 60

    ws.cell(1, 1, title).font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '합산 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_rows') or 0).font = _XL_DATA
    ws.cell(6, 1, '전체 KRW 합계').font = _XL_DATA
    grand = ws.cell(6, 2, agg_data.get('grand_total_krw') or 0)
    grand.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    grand.number_format = _XL_NUM_FMT

    ws.cell(8, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=11)
    headers = ['통화', '로컬 합계', 'KRW 합계']
    for j, h in enumerate(headers, 1):
        c = ws.cell(9, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
        c.alignment = Alignment(horizontal='center')
    by_cur = agg_data.get('total_by_currency') or {}
    r = 10
    for cur in sorted(by_cur.keys()):
        ws.cell(r, 1, cur).font = _XL_DATA
        lc = ws.cell(r, 2, by_cur[cur].get('local') or 0)
        lc.font = _XL_DATA; lc.number_format = _XL_NUM_FMT
        kc = ws.cell(r, 3, by_cur[cur].get('krw') or 0)
        kc.font = _XL_DATA; kc.number_format = _XL_NUM_FMT
        r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L4 시트 누락/오류 회사').font = Font(bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1


def _write_l4_table_detail(wb, agg_data, col_defs):
    """col_defs: [(label, width, key, fmt)] — fmt 'num'|'rate'|'text'."""
    ws = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'

    for j, (label, _w, _k, _f) in enumerate(col_defs, 1):
        cell = ws.cell(1, j, label)
        cell.font = _XL_HDR_FONT
        cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    krw_col_idx = None
    for j, (_lbl, _w, key, _f) in enumerate(col_defs, 1):
        if key == 'krw':
            krw_col_idx = j; break

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(col_defs, 1):
            val = r.get(key)
            if fmt == 'num':
                cell = ws.cell(row, j, val if isinstance(val, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ws.cell(row, j, val if isinstance(val, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ws.cell(row, j, val if val is not None else '')
            cell.font = _XL_DATA
        row += 1

    if agg_data.get('rows'):
        for j in range(1, len(col_defs) + 1):
            ws.cell(row, j).fill = _XL_TOTAL_FILL
        ws.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        if krw_col_idx:
            gc = ws.cell(row, krw_col_idx, agg_data.get('grand_total_krw') or 0)
            gc.font = _XL_TOTAL_FONT
            gc.number_format = _XL_NUM_FMT


def _build_l4_table_excel(agg_data, output_path, title, col_defs):
    wb = Workbook()
    wb.remove(wb.active)
    _write_l4_table_cover(wb, agg_data, title)
    _write_l4_table_detail(wb, agg_data, col_defs)
    wb.save(str(output_path))
    return str(output_path)


# 다행 명세 컬럼 정의 (label, width, key, fmt)
_L4_GR_COLS = [   # 4-1 받은 보증
    ('회사', 28, 'company', 'text'),
    ('제공자', 28, 'guarantor', 'text'),
    ('보증종류', 22, 'type', 'text'),
    ('통화', 8, 'currency', 'text'),
    ('로컬 금액', 18, 'local', 'num'),
    ('Spot', 10, 'spot', 'rate'),
    ('KRW 환산', 20, 'krw', 'num'),
    ('관련계정', 16, 'account', 'text'),
    ('Description', 40, 'description', 'text'),
]
_L4_GP_COLS = [   # 5-2 제공한 보증
    ('회사', 28, 'company', 'text'),
    ('제공받는자', 32, 'beneficiary', 'text'),
    ('보증종류', 22, 'type', 'text'),
    ('통화', 8, 'currency', 'text'),
    ('로컬 금액', 18, 'local', 'num'),
    ('Spot', 10, 'spot', 'rate'),
    ('KRW 환산', 20, 'krw', 'num'),
    ('채권자', 22, 'guaranteed_creditor', 'text'),
    ('Description', 40, 'description', 'text'),
]
_L4_RF_COLS = [   # 7-1 사용제한 금융상품
    ('회사', 28, 'company', 'text'),
    ('계정과목', 22, 'account', 'text'),
    ('통화', 8, 'currency', 'text'),
    ('로컬 금액', 18, 'local', 'num'),
    ('Spot', 10, 'spot', 'rate'),
    ('KRW 환산', 20, 'krw', 'num'),
    ('제한내용', 50, 'description', 'text'),
]
_L4_IP_COLS = [   # 8-1 보험가입 유형자산
    ('회사', 28, 'company', 'text'),
    ('자산종류', 22, 'asset_type', 'text'),
    ('보험사', 22, 'insurer', 'text'),
    ('통화', 8, 'currency', 'text'),
    ('부보금액(로컬)', 18, 'local', 'num'),
    ('Spot', 10, 'spot', 'rate'),
    ('KRW 환산', 20, 'krw', 'num'),
    ('Description', 40, 'description', 'text'),
]
_L4_PP_COLS = [   # 8-2 보험수익금 질권설정
    ('회사', 28, 'company', 'text'),
    ('질권자', 22, 'pledgee', 'text'),
    ('통화', 8, 'currency', 'text'),
    ('질권금액(로컬)', 18, 'local', 'num'),
    ('Spot', 10, 'spot', 'rate'),
    ('KRW 환산', 20, 'krw', 'num'),
    ('Description', 50, 'description', 'text'),
]
_L4_PA_COLS = [   # 9-1 담보제공자산
    ('회사', 28, 'company', 'text'),
    ('채권자/저당권자', 22, 'creditor', 'text'),
    ('담보자산', 22, 'asset_account', 'text'),
    ('통화', 8, 'currency', 'text'),
    ('담보금액(로컬)', 18, 'local', 'num'),
    ('Spot', 10, 'spot', 'rate'),
    ('KRW 환산', 20, 'krw', 'num'),
    ('관련부채', 18, 'liability_account', 'text'),
    ('Description', 40, 'description', 'text'),
]


def build_l4_guarantees_excel(agg_data, output_path):
    """4-1 받은 보증 합산 엑셀."""
    return _build_l4_table_excel(agg_data, output_path,
                                 'L4 4-1 받은 보증 내용 합산', _L4_GR_COLS)


def build_l4_guarantees_provided_excel(agg_data, output_path):
    """5-2 제공한 보증 합산 엑셀."""
    return _build_l4_table_excel(agg_data, output_path,
                                 'L4 5-2 제공한 보증 내용 합산', _L4_GP_COLS)


def build_l4_restricted_excel(agg_data, output_path):
    """7-1 사용제한 금융상품 합산 엑셀."""
    return _build_l4_table_excel(agg_data, output_path,
                                 'L4 7-1 사용제한 금융상품 합산', _L4_RF_COLS)


def build_l4_insured_ppe_excel(agg_data, output_path):
    """8-1 보험가입 유형자산 합산 엑셀."""
    return _build_l4_table_excel(agg_data, output_path,
                                 'L4 8-1 보험가입 유형자산 합산', _L4_IP_COLS)


def build_l4_pledged_proceeds_excel(agg_data, output_path):
    """8-2 보험수익금 질권설정 합산 엑셀."""
    return _build_l4_table_excel(agg_data, output_path,
                                 'L4 8-2 보험수익금 질권설정 합산', _L4_PP_COLS)


def build_l4_pledged_assets_excel(agg_data, output_path):
    """9-1 담보제공자산 합산 엑셀."""
    return _build_l4_table_excel(agg_data, output_path,
                                 'L4 9-1 담보제공자산 합산', _L4_PA_COLS)


# ──────────────────────────────────────────────────────────────
# 엑셀 빌더 (L4 6-1 소송 — 통화 없는 특수 패턴)
# ──────────────────────────────────────────────────────────────

_L4_LAW_COLS = [
    ('회사',                       30),
    ('구분',                       16),  # 피고/원고
    ('소송건수',                   12),
    ('통화',                        8),
    ('소송금액(로컬)',             22),
    ('충당부채(로컬)',             22),
    ('Spot',                       10),
    ('소송금액(KRW)',              22),
    ('충당부채(KRW)',              22),
]


def build_a2_securities_excel(agg_data, output_path):
    """A2 유가증권 명세 합산 결과(JSON)를 엑셀로 저장.

    agg_data: {
      'year', 'scanned', 'with_rows',
      'rows': [{'company','account','investee','shares','ownership_pct',
                'currency','local_cost','local_book','spot',
                'krw_cost','krw_book'}],
      'grand_cost_krw', 'grand_book_krw',
      'total_by_currency': {cur: {'cost':, 'book':, 'krw_cost':, 'krw_book':}},
      'errors': [{'company','reason'}],
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 60
    ws.cell(1, 1, 'A2 유가증권 명세 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '합산 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_rows') or 0).font = _XL_DATA
    ws.cell(6, 1, '취득원가 KRW 합계').font = _XL_DATA
    gc = ws.cell(6, 2, agg_data.get('grand_cost_krw') or 0)
    gc.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    gc.number_format = _XL_NUM_FMT
    ws.cell(7, 1, '장부가액 KRW 합계').font = _XL_DATA
    gb = ws.cell(7, 2, agg_data.get('grand_book_krw') or 0)
    gb.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    gb.number_format = _XL_NUM_FMT

    # 통화별 합계
    by_cur = agg_data.get('total_by_currency') or {}
    if by_cur:
        ws.cell(9, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=11)
        headers = ['통화', '취득원가 로컬', '장부가액 로컬',
                   '취득원가 KRW', '장부가액 KRW']
        for j, h in enumerate(headers, 1):
            c = ws.cell(10, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
            c.alignment = Alignment(horizontal='center')
        r = 11
        for cur in sorted(by_cur.keys()):
            ws.cell(r, 1, cur).font = _XL_DATA
            for j, key in enumerate(['cost', 'book', 'krw_cost', 'krw_book'], 2):
                cell = ws.cell(r, j, by_cur[cur].get(key) or 0)
                cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
            r += 1
    else:
        r = 11

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'A2 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세 시트
    cols = [
        ('회사',           30),
        ('계정',           18),
        ('피투자회사명',   34),
        ('주식수',         14),
        ('지분율(%)',      12),
        ('통화',           8),
        ('취득원가(로컬)', 18),
        ('장부가액(로컬)', 18),
        ('Spot',           10),
        ('취득원가(KRW)',  18),
        ('장부가액(KRW)',  18),
    ]
    ds = wb.create_sheet('상세')
    for j, (_lbl, w) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        ds.cell(row, 1, r.get('company') or '').font = _XL_DATA
        ds.cell(row, 2, r.get('account') or '').font = _XL_DATA
        ds.cell(row, 3, r.get('investee') or '').font = _XL_DATA
        sh = ds.cell(row, 4, r.get('shares') if r.get('shares') is not None else '')
        sh.font = _XL_DATA
        if isinstance(r.get('shares'), (int, float)):
            sh.number_format = '#,##0'
        op = ds.cell(row, 5, r.get('ownership_pct') if r.get('ownership_pct') is not None else '')
        op.font = _XL_DATA
        if isinstance(r.get('ownership_pct'), (int, float)):
            op.number_format = '0.0000%'
        ds.cell(row, 6, r.get('currency') or '').font = _XL_DATA
        for j, key in [(7, 'local_cost'), (8, 'local_book')]:
            cell = ds.cell(row, j, r.get(key) or 0)
            cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
        sc = ds.cell(row, 9, r.get('spot') or 0)
        sc.font = _XL_DATA; sc.number_format = _XL_RATE_FMT
        for j, key in [(10, 'krw_cost'), (11, 'krw_book')]:
            cell = ds.cell(row, j, r.get(key) or 0)
            cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
        row += 1

    if agg_data.get('rows'):
        for j in range(1, len(cols) + 1):
            ds.cell(row, j).fill = _XL_TOTAL_FILL
        ds.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        kc_total = ds.cell(row, 10, agg_data.get('grand_cost_krw') or 0)
        kc_total.font = _XL_TOTAL_FONT; kc_total.number_format = _XL_NUM_FMT
        kb_total = ds.cell(row, 11, agg_data.get('grand_book_krw') or 0)
        kb_total.font = _XL_TOTAL_FONT; kb_total.number_format = _XL_NUM_FMT

    wb.save(str(output_path))
    return str(output_path)


def build_a3_investment_pl_excel(agg_data, output_path):
    """A3 1번 투자부동산 관련 손익 합산 엑셀.

    agg_data: {
      'year', 'scanned', 'with_data_count',
      'rows': [{'company','currency','spot',
                'local_rental_revenue','local_operating_expenses',
                'local_depreciation','local_fv_change','local_others','local_total',
                'krw_rental_revenue','krw_operating_expenses',
                'krw_depreciation','krw_fv_change','krw_others','krw_total'}],
      'grand_total_krw',
      'totals_by_item_krw': {item_key: sum},
      'errors': [{'company','reason'}],
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    ws.cell(1, 1, 'A3 1번 투자부동산 관련 손익 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA
    ws.cell(6, 1, '합계 KRW').font = _XL_DATA
    gc = ws.cell(6, 2, agg_data.get('grand_total_krw') or 0)
    gc.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    gc.number_format = _XL_NUM_FMT

    # 항목별 KRW 합계
    item_labels = [
        ('rental_revenue',     '임대수익'),
        ('operating_expenses', '직접 관련된 운영비용'),
        ('depreciation',       '감가상각비'),
        ('fv_change',          '공정가치 변동'),
        ('others',             '기타'),
    ]
    ws.cell(8, 1, '항목별 KRW 합계').font = Font(bold=True, color='1F3864', size=11)
    totals = agg_data.get('totals_by_item_krw') or {}
    r = 9
    for key, lbl in item_labels:
        ws.cell(r, 1, lbl).font = _XL_DATA
        cell = ws.cell(r, 2, totals.get(key) or 0)
        cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
        r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'A3 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세 — 회사 / 통화 / 항목별 로컬 / Spot / 항목별 KRW
    item_keys = [k for k, _ in item_labels] + ['total']
    item_label_dict = dict(item_labels) | {'total': '합계'}
    cols = [('회사', 30, 'company', 'text'),
            ('통화', 8, 'currency', 'text')]
    for k in item_keys:
        cols.append((f'{item_label_dict[k]}(로컬)', 18, f'local_{k}', 'num'))
    cols.append(('Spot', 10, 'spot', 'rate'))
    for k in item_keys:
        cols.append((f'{item_label_dict[k]}(KRW)', 18, f'krw_{k}', 'num'))

    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ds.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1

    wb.save(str(output_path))
    return str(output_path)


def build_a3_land_value_excel(agg_data, output_path, title, amount_label):
    """A3 2-1 / 3-1 공시지가 단일 금액 합산 엑셀."""
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 40
    ws.cell(1, 1, title).font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, 'YES 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('yes_count') or 0).font = _XL_DATA
    ws.cell(6, 1, '입력 금액 회사 수').font = _XL_DATA
    ws.cell(6, 2, agg_data.get('with_amount_count') or 0).font = _XL_DATA
    ws.cell(7, 1, '전체 KRW 합계').font = _XL_DATA
    gc = ws.cell(7, 2, agg_data.get('grand_total_krw') or 0)
    gc.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    gc.number_format = _XL_NUM_FMT

    by_cur = agg_data.get('total_by_currency') or {}
    if by_cur:
        ws.cell(9, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=11)
        headers = ['통화', '로컬 합계', 'KRW 합계']
        for j, h in enumerate(headers, 1):
            c = ws.cell(10, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
        r = 11
        for cur in sorted(by_cur.keys()):
            ws.cell(r, 1, cur).font = _XL_DATA
            for j, key in enumerate(['local', 'krw'], 2):
                cell = ws.cell(r, j, by_cur[cur].get(key) or 0)
                cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
            r += 1
    else:
        r = 11

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'A3 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    cols = [
        ('회사',         30),
        ('해당여부',     12),
        ('통화',          8),
        (f'{amount_label}(로컬)', 22),
        ('Spot',         10),
        (f'{amount_label}(KRW)',  22),
    ]
    ds = wb.create_sheet('상세')
    for j, (_lbl, w) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        ds.cell(row, 1, r.get('company') or '').font = _XL_DATA
        ds.cell(row, 2, r.get('yn') or '').font = _XL_DATA
        ds.cell(row, 3, r.get('currency') or '').font = _XL_DATA
        lc = ds.cell(row, 4, r.get('local') if r.get('local') is not None else '')
        lc.font = _XL_DATA
        if isinstance(r.get('local'), (int, float)):
            lc.number_format = _XL_NUM_FMT
        sc = ds.cell(row, 5, r.get('spot') if r.get('spot') is not None else '')
        sc.font = _XL_DATA
        if isinstance(r.get('spot'), (int, float)):
            sc.number_format = _XL_RATE_FMT
        kc = ds.cell(row, 6, r.get('krw') if r.get('krw') is not None else '')
        kc.font = _XL_DATA
        if isinstance(r.get('krw'), (int, float)):
            kc.number_format = _XL_NUM_FMT
        row += 1

    if agg_data.get('rows'):
        for j in range(1, len(cols) + 1):
            ds.cell(row, j).fill = _XL_TOTAL_FILL
        ds.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        gc2 = ds.cell(row, 6, agg_data.get('grand_total_krw') or 0)
        gc2.font = _XL_TOTAL_FONT; gc2.number_format = _XL_NUM_FMT

    wb.save(str(output_path))
    return str(output_path)


def build_a3_land_investment_excel(agg_data, output_path):
    return build_a3_land_value_excel(
        agg_data, output_path,
        title='A3 2-1 투자부동산(토지) 공시지가 합산',
        amount_label='공시지가')


def build_a3_land_ppe_excel(agg_data, output_path):
    return build_a3_land_value_excel(
        agg_data, output_path,
        title='A3 3-1 유형자산(토지) 공시지가 합산',
        amount_label='공시지가')


# ──────────────────────────────────────────────────────────────
# A4 시트 빌더 — 건설계약 1/2/3
# ──────────────────────────────────────────────────────────────

_A4_TYPE_LABELS_KOR = {
    'architecture': '건축',
    'civil':        '토목',
    'plant':        '플랜트',
    'hydrogen':     '수소충전소',
    'others':       'Others',
}

_A4_S1_COLS = [   # 1번 잔액 변동 (avg 환율 사용)
    ('회사',         30, 'company',   'text'),
    ('공사 종류',    18, 'type_label','text'),
    ('통화',          8, 'currency',  'text'),
    ('기초(로컬)',   18, 'local_beginning', 'num'),
    ('증감(로컬)',   18, 'local_variance',  'num'),
    ('공사수익(로컬)', 18, 'local_profit', 'num'),
    ('기타(로컬)',   18, 'local_others', 'num'),
    ('기말(로컬)',   18, 'local_ending', 'num'),
    ('환율(Avg)',    11, 'spot',      'rate'),
    ('기초(KRW)',    18, 'krw_beginning', 'num'),
    ('증감(KRW)',    18, 'krw_variance',  'num'),
    ('공사수익(KRW)', 18, 'krw_profit', 'num'),
    ('기타(KRW)',    18, 'krw_others', 'num'),
    ('기말(KRW)',    18, 'krw_ending', 'num'),
]

_A4_S3_COLS = [   # 3번 계약자산·부채 (spot 환율 사용 — B/S 성격)
    ('회사',           30, 'company', 'text'),
    ('공사 종류',      18, 'type_label', 'text'),
    ('통화',            8, 'currency', 'text'),
    ('미청구공사(로컬)', 22, 'local_receivable', 'num'),
    ('초과청구공사(로컬)', 22, 'local_payable', 'num'),
    ('선수금(로컬)',   18, 'local_advance', 'num'),
    ('Spot',           10, 'spot', 'rate'),
    ('미청구공사(KRW)', 22, 'krw_receivable', 'num'),
    ('초과청구공사(KRW)', 22, 'krw_payable', 'num'),
    ('선수금(KRW)',    18, 'krw_advance', 'num'),
]


def _a4_write_cover(wb, agg_data, title, total_labels):
    """A4 1/3 공통 표지. total_labels: [(label, agg_key), ...] KRW 합계 항목들."""
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    ws.cell(1, 1, title).font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA

    r = 7
    ws.cell(r, 1, 'KRW 환산 합계').font = Font(bold=True, color='1F3864', size=12)
    r += 1
    for label, key in total_labels:
        ws.cell(r, 1, label).font = _XL_DATA
        cell = ws.cell(r, 2, agg_data.get(key) or 0)
        cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
        r += 1

    # 통화별 합계
    by_cur = agg_data.get('total_by_currency') or {}
    if by_cur:
        r += 1
        ws.cell(r, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=11)
        r += 1
        # 헤더: 통화 + 각 항목 (로컬 / KRW 모두)
        local_keys = [k.replace('krw_', 'local_') for _, k in total_labels]
        all_keys = []
        headers = ['통화']
        for label, k in total_labels:
            headers.append(label.replace('KRW', '로컬'))
            all_keys.append(k.replace('krw_', 'local_'))
        for label, k in total_labels:
            headers.append(label)
            all_keys.append(k)
        for j, h in enumerate(headers, 1):
            c = ws.cell(r, j, h)
            c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
            c.alignment = Alignment(horizontal='center')
        r += 1
        for cur in sorted(by_cur.keys()):
            ws.cell(r, 1, cur).font = _XL_DATA
            for j, k in enumerate(all_keys, 2):
                cell = ws.cell(r, j, by_cur[cur].get(k) or 0)
                cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
            r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'A4 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1


def _a4_write_detail(wb, agg_data, col_defs):
    """A4 1/3 공통 상세 시트 작성기."""
    ws = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(col_defs, 1):
        cell = ws.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(col_defs, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ws.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ws.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ws.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1


def build_a4_construction_balance_excel(agg_data, output_path):
    """A4 1. 공사계약 잔액 변동 합산 엑셀."""
    wb = Workbook()
    wb.remove(wb.active)
    _a4_write_cover(wb, agg_data,
                    title='A4 1. 공사계약 잔액의 변동내역 합산',
                    total_labels=[
                        ('기초 KRW',     'krw_beginning'),
                        ('증감 KRW',     'krw_variance'),
                        ('공사수익 KRW', 'krw_profit'),
                        ('기타 KRW',     'krw_others'),
                        ('기말 KRW',     'krw_ending'),
                    ])
    _a4_write_detail(wb, agg_data, _A4_S1_COLS)
    wb.save(str(output_path))
    return str(output_path)


def build_a4_construction_profit_excel(agg_data, output_path):
    """A4 2. 공사손익 합산 엑셀 (Pivot 형태)."""
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    ws.cell(1, 1, 'A4 2. 진행중인 건설계약 공사손익 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA

    # 항목별 × 공사종류별 KRW 합계 (피벗)
    item_labels = [
        ('accumulated_revenue', '누적공사수익 (Revenue)'),
        ('accumulated_cost',    '누적공사원가 (Cost)'),
        ('accumulated_income',  '누적공사손익 (Income)'),
    ]
    type_labels = [('architecture', '건축'), ('civil', '토목'),
                   ('plant', '플랜트'), ('hydrogen', '수소충전소'),
                   ('others', 'Others'), ('total', 'Total')]

    ws.cell(7, 1, '항목 × 공사종류별 KRW 합계').font = Font(
        bold=True, color='1F3864', size=12)
    # 헤더
    ws.cell(8, 1, '항목').font = _XL_HDR_FONT
    ws.cell(8, 1).fill = _XL_HDR_FILL
    for j, (_k, lbl) in enumerate(type_labels, 2):
        c = ws.cell(8, j, lbl); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(j)].width = 20

    totals_krw = agg_data.get('totals_krw') or {}
    r = 9
    for item_key, item_lbl in item_labels:
        ws.cell(r, 1, item_lbl).font = _XL_DATA
        item_totals = totals_krw.get(item_key) or {}
        for j, (type_key, _lbl) in enumerate(type_labels, 2):
            cell = ws.cell(r, j, item_totals.get(type_key) or 0)
            cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
        r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'A4 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세 — 회사별 × 항목 × 공사종류 (행: company × item, 열: 공사종류)
    # A4는 avg 환율 사용 (P&L 성격)
    ds = wb.create_sheet('상세')
    # 컬럼 정의: 회사 / 항목 / 통화 / [공사종류 6개 로컬] / Avg환율 / [공사종류 6개 KRW]
    cols = [('회사', 30), ('항목', 24), ('통화', 8)]
    for _k, lbl in type_labels:
        cols.append((f'{lbl}(로컬)', 18))
    cols.append(('환율(Avg)', 11))
    for _k, lbl in type_labels:
        cols.append((f'{lbl}(KRW)', 18))
    for j, (_lbl, w) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        ds.cell(row, 1, r.get('company') or '').font = _XL_DATA
        ds.cell(row, 2, r.get('item_label') or '').font = _XL_DATA
        ds.cell(row, 3, r.get('currency') or '').font = _XL_DATA
        # 로컬 6개
        for idx, (type_key, _lbl) in enumerate(type_labels):
            cell = ds.cell(row, 4 + idx, r.get(f'local_{type_key}') or 0)
            cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
        # Spot
        sc = ds.cell(row, 4 + len(type_labels), r.get('spot') or 0)
        sc.font = _XL_DATA; sc.number_format = _XL_RATE_FMT
        # KRW 6개
        for idx, (type_key, _lbl) in enumerate(type_labels):
            cell = ds.cell(row, 5 + len(type_labels) + idx,
                           r.get(f'krw_{type_key}') or 0)
            cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
        row += 1

    wb.save(str(output_path))
    return str(output_path)


def build_a4_contract_balance_excel(agg_data, output_path):
    """A4 3. 계약자산 및 계약부채 합산 엑셀."""
    wb = Workbook()
    wb.remove(wb.active)
    _a4_write_cover(wb, agg_data,
                    title='A4 3. 계약자산 및 계약부채 합산',
                    total_labels=[
                        ('미청구공사 KRW',   'krw_receivable'),
                        ('초과청구공사 KRW', 'krw_payable'),
                        ('선수금 KRW',       'krw_advance'),
                    ])
    _a4_write_detail(wb, agg_data, _A4_S3_COLS)
    wb.save(str(output_path))
    return str(output_path)


# ──────────────────────────────────────────────────────────────
# A5 시트 빌더 — 리스
# ──────────────────────────────────────────────────────────────

_A5_ASSET_LABELS = {
    'property':  '부동산',
    'vehicle':   '차량운반구',
    'equipment': '건설장비',
    'others':    '기타',
}

_A5_S1_COLS = [   # 1번 사용권자산 변동
    ('회사',         30, 'company',          'text'),
    ('자산 종류',    18, 'asset_label',      'text'),
    ('통화',          8, 'currency',         'text'),
    ('기초(로컬)',   18, 'local_beginning',    'num'),
    ('취득(로컬)',   18, 'local_acquisition',  'num'),
    ('처분(로컬)',   18, 'local_disposal',     'num'),
    ('상각(로컬)',   18, 'local_depreciation', 'num'),
    ('기타(로컬)',   18, 'local_others',       'num'),
    ('기말(로컬)',   18, 'local_ending',       'num'),
    ('Spot',         10, 'spot',             'rate'),
    ('기초(KRW)',    18, 'krw_beginning',    'num'),
    ('취득(KRW)',    18, 'krw_acquisition',  'num'),
    ('처분(KRW)',    18, 'krw_disposal',     'num'),
    ('상각(KRW)',    18, 'krw_depreciation', 'num'),
    ('기타(KRW)',    18, 'krw_others',       'num'),
    ('기말(KRW)',    18, 'krw_ending',       'num'),
]


def build_a5_rou_changes_excel(agg_data, output_path):
    """A5 1. 사용권자산 변동 합산 엑셀."""
    wb = Workbook()
    wb.remove(wb.active)
    _a4_write_cover(wb, agg_data,
                    title='A5 1. 사용권자산의 변동내역 합산',
                    total_labels=[
                        ('기초 KRW',     'krw_beginning'),
                        ('취득 KRW',     'krw_acquisition'),
                        ('처분 KRW',     'krw_disposal'),
                        ('상각 KRW',     'krw_depreciation'),
                        ('기타 KRW',     'krw_others'),
                        ('기말 KRW',     'krw_ending'),
                    ])
    _a4_write_detail(wb, agg_data, _A5_S1_COLS)
    wb.save(str(output_path))
    return str(output_path)


def build_l2_balance_excel(agg_data, output_path, title, col_labels):
    """L2 1번/2번 (장기차입금/사채) 다행 명세 합산 엑셀.

    col_labels: {'type1','type2'} — 컬럼 헤더 (예: 대주구분/대주명 또는 종류/주관사)
    agg_data: {
      'year', 'scanned', 'with_data_count',
      'rows': [{'company','type1','type2','rate','currency','spot',
                'local_current','local_non_current','local_total',
                'krw_current','krw_non_current','krw_total'}],
      'grand_total_krw', 'grand_current_krw', 'grand_non_current_krw',
      'total_by_currency': {cur: {...}},
      'errors': [{'company','reason'}],
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    ws.cell(1, 1, title).font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA
    ws.cell(6, 1, '유동 KRW 합계').font = _XL_DATA
    cc = ws.cell(6, 2, agg_data.get('grand_current_krw') or 0)
    cc.font = _XL_DATA; cc.number_format = _XL_NUM_FMT
    ws.cell(7, 1, '비유동 KRW 합계').font = _XL_DATA
    nc = ws.cell(7, 2, agg_data.get('grand_non_current_krw') or 0)
    nc.font = _XL_DATA; nc.number_format = _XL_NUM_FMT
    ws.cell(8, 1, '전체 KRW 합계').font = Font(bold=True, color='9C5700', name='Arial', size=11)
    tc = ws.cell(8, 2, agg_data.get('grand_total_krw') or 0)
    tc.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    tc.number_format = _XL_NUM_FMT

    # 통화별 합계
    by_cur = agg_data.get('total_by_currency') or {}
    r = 10
    if by_cur:
        ws.cell(r, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=11)
        r += 1
        headers = ['통화', '유동(로컬)', '비유동(로컬)', '합계(로컬)',
                   '유동(KRW)', '비유동(KRW)', '합계(KRW)']
        for j, h in enumerate(headers, 1):
            c = ws.cell(r, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
            c.alignment = Alignment(horizontal='center')
        r += 1
        for cur in sorted(by_cur.keys()):
            ws.cell(r, 1, cur).font = _XL_DATA
            for j, key in enumerate(['current', 'non_current', 'total',
                                     'krw_current', 'krw_non_current', 'krw_total'], 2):
                cell = ws.cell(r, j, by_cur[cur].get(key) or 0)
                cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
            r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L2 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세
    cols = [
        ('회사',            30, 'company',           'text'),
        (col_labels['type1'], 22, 'type1',          'text'),
        (col_labels['type2'], 28, 'type2',          'text'),
        ('이자율',          10, 'rate',              'rate'),
        ('통화',             8, 'currency',          'text'),
        ('유동(로컬)',      18, 'local_current',     'num'),
        ('비유동(로컬)',    18, 'local_non_current', 'num'),
        ('합계(로컬)',      18, 'local_total',       'num'),
        ('Spot',            10, 'spot',              'rate'),
        ('유동(KRW)',       18, 'krw_current',       'num'),
        ('비유동(KRW)',     18, 'krw_non_current',   'num'),
        ('합계(KRW)',       18, 'krw_total',         'num'),
    ]
    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ds.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1

    if agg_data.get('rows'):
        for j in range(1, len(cols) + 1):
            ds.cell(row, j).fill = _XL_TOTAL_FILL
        ds.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        tc2 = ds.cell(row, 12, agg_data.get('grand_total_krw') or 0)
        tc2.font = _XL_TOTAL_FONT; tc2.number_format = _XL_NUM_FMT

    wb.save(str(output_path))
    return str(output_path)


def build_l2_long_term_borrowings_excel(agg_data, output_path):
    return build_l2_balance_excel(
        agg_data, output_path,
        title='L2 1. 장기차입금 (유동성 포함) 합산',
        col_labels={'type1': '대주구분', 'type2': '대주명'})


def build_l2_debentures_excel(agg_data, output_path):
    return build_l2_balance_excel(
        agg_data, output_path,
        title='L2 2. 사채 (유동성 사채 포함) 합산',
        col_labels={'type1': '종류', 'type2': '주관사'})


def _build_l3_pivot_excel(agg_data, output_path, title, item_keys, item_labels,
                          extra_summary=None):
    """L3 1/2번 공통 빌더 — 회사별 항목 펼친 형태.

    agg_data: {
      'year','scanned','with_data_count',
      'rows': [{'company','currency','spot',
                'local_<key>','krw_<key>'} ...],
      'totals_by_item_krw': {key: sum},
      'grand_total_krw' (옵션, 기말금액 KRW 등 강조용),
      'total_by_currency': {cur: {...}},
      'errors': [...],
    }
    item_keys/item_labels: 항목 키와 라벨
    extra_summary: 추가 표지 요약 (key, label) 튜플 리스트
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    ws.cell(1, 1, title).font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA

    r = 7
    ws.cell(r, 1, '항목별 KRW 합계').font = Font(bold=True, color='1F3864', size=11)
    r += 1
    totals = agg_data.get('totals_by_item_krw') or {}
    for key, lbl in zip(item_keys, item_labels):
        ws.cell(r, 1, lbl).font = _XL_DATA
        cell = ws.cell(r, 2, totals.get(key) or 0)
        cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
        r += 1

    if extra_summary:
        r += 1
        for label, value in extra_summary:
            ws.cell(r, 1, label).font = Font(bold=True, color='9C5700',
                                            name='Arial', size=11)
            cell = ws.cell(r, 2, value or 0)
            cell.font = Font(bold=True, color='9C5700', name='Arial', size=11)
            cell.number_format = _XL_NUM_FMT
            r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L3 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 연결그룹별 KRW 합계 (집계에 포함된 경우만)
    group_subs = agg_data.get('group_subtotals') or []
    if group_subs:
        gs = wb.create_sheet('연결그룹별')
        gs.cell(1, 1, title).font = _XL_TITLE
        gs.cell(2, 1, '연결그룹별 KRW 합계').font = Font(
            bold=True, color='1F3864', size=11)
        gcols = [('연결그룹', 24), ('회사수', 10)] + [(l, 18) for l in item_labels]
        for j, (lbl, w) in enumerate(gcols, 1):
            gs.column_dimensions[get_column_letter(j)].width = w
            cell = gs.cell(4, j, lbl)
            cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
        gr = 5
        for sub in group_subs:
            gs.cell(gr, 1, sub.get('group') or '').font = _XL_DATA
            c = gs.cell(gr, 2, sub.get('company_count') or 0)
            c.font = _XL_DATA; c.number_format = '#,##0'
            krw = sub.get('krw') or {}
            for j, key in enumerate(item_keys, 3):
                cc = gs.cell(gr, j, krw.get(key) or 0)
                cc.font = _XL_DATA; cc.number_format = _XL_NUM_FMT
            gr += 1
        gs.freeze_panes = 'A5'

    # 상세
    cols = [
        ('회사', 30, 'company', 'text'),
        ('통화', 8, 'currency', 'text'),
    ]
    for key, lbl in zip(item_keys, item_labels):
        cols.append((f'{lbl}(로컬)', 18, f'local_{key}', 'num'))
    cols.append(('Spot', 10, 'spot', 'rate'))
    for key, lbl in zip(item_keys, item_labels):
        cols.append((f'{lbl}(KRW)', 18, f'krw_{key}', 'num'))

    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ds.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1

    wb.save(str(output_path))
    return str(output_path)


# L3 1번/2번 항목 정의 (표시 순서)
_L3_S1_KEYS = ['beginning', 'provision', 'payment', 'transfer',
               'business_combination', 'others', 'ending']
_L3_S1_LBLS = ['기초금액', '퇴직급여 설정액', '급여지급액', '관계사전출입액',
               '사업결합으로 인한 증가', '기타증감', '기말금액']

_L3_S2_KEYS = ['beginning', 'contribution', 'payment', 'interest_income',
               'transfer', 'business_combination', 'others', 'ending']
_L3_S2_LBLS = ['기초금액', '적립액', '급여지급액', '이자수익', '관계사전출입액',
               '사업결합으로 인한 증가', '기타증감', '기말금액']

# L3 3번 자산종류
_L3_S3_KEYS = ['cash', 'deposit', 'securities', 'bond', 'others']
_L3_S3_LBLS = ['현금', '예금', '주식', '채권', '기타']

# L3-1 1번 (확정급여채무 변동) 항목 표시 순서
_L31_S1_KEYS = ['beginning', 'current_service_cost', 'interest_cost',
                'remeasurement', 'demographic_gain_loss', 'financial_gain_loss',
                'experience_adjustment', 'payment',
                'business_combination', 'others', 'ending']
_L31_S1_LBLS = ['기초금액', '근무원가', '이자비용', '재측정요소',
                '인구통계가정 손익', '재무적가정 손익', '경험적 조정',
                '급여지급액', '사업결합 증가', '기타증감', '기말금액']

# L3-1 2번 (사외적립자산 공정가치 변동) 항목 표시 순서
_L31_S2_KEYS = ['beginning', 'interest_income', 'return_excluding_interest',
                'employer_contribution', 'payment',
                'business_combination', 'others', 'ending']
_L31_S2_LBLS = ['기초금액', '이자수익', '제도자산 손익(이자제외)',
                '고용인 기여금', '급여지급액',
                '사업결합 증가', '기타증감', '기말금액']


def _build_l3_fx_effect_excel(agg_data, output_path, title, var_keys, var_labels):
    """L3 1번/2번 공통 빌더 (항목별 환율 + 환율변동효과 컬럼).

    var_keys: 변동 항목 키 리스트 (avg 환율 적용 대상)
    var_labels: 표시용 라벨 (동일 길이)
    공통 항목: beginning(전기말 spot), ending(당기말 spot), fx_effect(계산값)
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ─ 표지 ─
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 32
    ws.cell(1, 1, title).font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA

    # 환율 정책 안내
    ws.cell(7, 1, '환율 정책').font = Font(bold=True, color='1F3864', size=12)
    policy = agg_data.get('rate_policy') or {}
    info_rows = [
        ('기초금액 환산', policy.get('beginning', '전기말 spot')),
        ('변동 항목 환산', policy.get('variance', '당기 avg')),
        ('기말금액 환산', policy.get('ending', '당기말 spot')),
        ('환율변동효과', policy.get('fx_effect', '기말 - (기초 + 변동분 합)')),
    ]
    r = 8
    for lbl, val in info_rows:
        ws.cell(r, 1, lbl).font = _XL_DATA
        ws.cell(r, 2, val).font = _XL_DATA
        r += 1

    # 항목별 KRW 합계
    r += 1
    ws.cell(r, 1, '항목별 KRW 합계').font = Font(bold=True, color='1F3864', size=12)
    r += 1
    totals = agg_data.get('totals_by_item_krw') or {}
    items_summary = [('beginning', '기초금액')]
    for k, lbl in zip(var_keys, var_labels):
        items_summary.append((k, lbl))
    items_summary.append(('fx_effect', '환율변동효과'))
    items_summary.append(('ending', '기말금액'))
    for key, lbl in items_summary:
        ws.cell(r, 1, lbl).font = _XL_DATA
        cell = ws.cell(r, 2, totals.get(key) or 0)
        if key == 'fx_effect':
            cell.font = Font(bold=True, color='1F3864', name='Arial', size=11)
        elif key == 'ending':
            cell.font = Font(bold=True, color='9C5700', name='Arial', size=11)
        else:
            cell.font = _XL_DATA
        cell.number_format = _XL_NUM_FMT
        r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L3 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # ─ 상세 ─
    # 회사 / 통화 / [기초, 변동..., 기말 로컬] / 3종 환율 /
    # [기초 KRW, 변동... KRW, 환율변동효과 KRW, 기말 KRW]
    item_keys = ['beginning'] + list(var_keys) + ['ending']
    item_lbls = ['기초금액'] + list(var_labels) + ['기말금액']

    cols = [
        ('회사', 30, 'company', 'text'),
        ('통화', 8, 'currency', 'text'),
    ]
    for k, lbl in zip(item_keys, item_lbls):
        cols.append((f'{lbl}(로컬)', 18, f'local_{k}', 'num'))
    cols.append(('전기말 Spot', 12, 'prior_spot', 'rate'))
    cols.append(('당기 Avg', 11, 'avg_rate', 'rate'))
    cols.append(('당기말 Spot', 12, 'spot', 'rate'))
    cols.append(('기초(KRW)\n[전기말 Spot]', 20, 'krw_beginning', 'num'))
    for k, lbl in zip(var_keys, var_labels):
        cols.append((f'{lbl}(KRW)\n[Avg]', 18, f'krw_{k}', 'num'))
    cols.append(('환율변동효과(KRW)', 18, 'krw_fx_effect', 'num'))
    cols.append(('기말(KRW)\n[당기말 Spot]', 20, 'krw_ending', 'num'))

    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    ds.row_dimensions[1].height = 32
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ds.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1

    wb.save(str(output_path))
    return str(output_path)


def build_l3_severance_excel(agg_data, output_path):
    """L3 1. 퇴직급여충당부채의 변동 (기초=전기말 spot / 변동=avg / 기말=당기말 spot)."""
    return _build_l3_fx_effect_excel(
        agg_data, output_path,
        title='L3 1. 퇴직급여충당부채의 변동 합산',
        var_keys=['provision', 'payment', 'transfer',
                  'business_combination', 'others'],
        var_labels=['퇴직급여 설정액', '급여지급액', '관계사전출입액',
                    '사업결합 증가', '기타증감'],
    )


def build_l3_pension_movement_excel(agg_data, output_path):
    """L3 2. 퇴직연금운용자산의 변동 (기초=전기말 spot / 변동=avg / 기말=당기말 spot)."""
    return _build_l3_fx_effect_excel(
        agg_data, output_path,
        title='L3 2. 퇴직연금운용자산의 변동 합산',
        var_keys=['contribution', 'payment', 'interest_income',
                  'transfer', 'business_combination', 'others'],
        var_labels=['적립액', '급여지급액', '이자수익', '관계사전출입액',
                    '사업결합 증가', '기타증감'],
    )


def build_l3_pension_breakdown_excel(agg_data, output_path):
    """L3 3. 퇴직연금운용자산의 구성내역 합산 엑셀."""
    return _build_l3_pivot_excel(
        agg_data, output_path,
        title='L3 3. 퇴직연금운용자산의 구성내역 합산',
        item_keys=_L3_S3_KEYS,
        item_labels=_L3_S3_LBLS,
        extra_summary=[('합계 KRW', agg_data.get('grand_total_krw') or 0)],
    )


def build_l31_dbo_excel(agg_data, output_path):
    """L3-1 1. 확정급여채무 변동 합산 엑셀."""
    return _build_l3_pivot_excel(
        agg_data, output_path,
        title='L3-1 1. 확정급여채무의 변동 합산',
        item_keys=_L31_S1_KEYS,
        item_labels=_L31_S1_LBLS,
        extra_summary=[('전체 기말금액 KRW 합계',
                        (agg_data.get('totals_by_item_krw') or {}).get('ending') or 0)],
    )


def build_l31_plan_asset_excel(agg_data, output_path):
    """L3-1 2. 사외적립자산 공정가치 변동 합산 엑셀."""
    return _build_l3_pivot_excel(
        agg_data, output_path,
        title='L3-1 2. 사외적립자산의 공정가치 변동 합산',
        item_keys=_L31_S2_KEYS,
        item_labels=_L31_S2_LBLS,
        extra_summary=[('전체 기말금액 KRW 합계',
                        (agg_data.get('totals_by_item_krw') or {}).get('ending') or 0)],
    )


def build_l31_assumptions_excel(agg_data, output_path):
    """L3-1 3. 보험수리적 평가를 위한 주요 가정치 — 최저/최고 입력 회사 엑셀.

    agg_data: {
      'year','scanned','with_data_count',
      'rows': [{'company','wage_growth_raw','wage_growth_value',
                'wage_growth_value_min','wage_growth_value_max',
                'discount_rate_raw',...}],
      'extremes': {
        'wage_growth':   {'min': {...row...}, 'max': {...row...}},
        'discount_rate': {'min': {...row...}, 'max': {...row...}},
      },
      'errors': [...]
    }
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 24
    ws.cell(1, 1, 'L3-1 3. 보험수리적 평가 가정치 — 최저/최고 입력 회사'
            ).font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA

    pairs = [('wage_growth', '기대임금상승률'),
             ('discount_rate', '할인율')]
    extremes = agg_data.get('extremes') or {}

    r = 7
    ws.cell(r, 1, '최저 / 최고 입력 회사').font = Font(
        bold=True, color='1F3864', size=11)
    r += 1
    for j, h in enumerate(['가정', '구분', '회사', '입력값'], 1):
        c = ws.cell(r, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
        c.alignment = Alignment(horizontal='center')
    r += 1
    for key, lbl in pairs:
        ex = extremes.get(key) or {}
        for kind, kind_lbl in [('min', '최저'), ('max', '최고')]:
            row_data = ex.get(kind) or {}
            ws.cell(r, 1, lbl).font = _XL_DATA
            ws.cell(r, 2, kind_lbl).font = _XL_DATA
            ws.cell(r, 3, row_data.get('company') or '-').font = _XL_DATA
            ws.cell(r, 4, row_data.get('raw') or '-').font = _XL_DATA
            r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L3-1 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세 — 회사별 raw + 정규화 값
    cols = [
        ('회사', 30, 'company', 'text'),
        ('기대임금상승률 (원문)', 22, 'wage_growth_raw',   'text'),
        ('기대임금상승률 (값)',    14, 'wage_growth_value', 'pct'),
        ('할인율 (원문)',         22, 'discount_rate_raw',   'text'),
        ('할인율 (값)',            14, 'discount_rate_value', 'pct'),
    ]
    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    rr = 2
    for r_row in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r_row.get(key)
            if fmt == 'pct':
                cell = ds.cell(rr, j, v if isinstance(v, (int, float)) else None)
                cell.number_format = '0.00%'
            else:
                cell = ds.cell(rr, j, v if v is not None else '')
            cell.font = _XL_DATA
        rr += 1

    wb.save(str(output_path))
    return str(output_path)


def build_l31_sensitivity_excel(agg_data, output_path):
    """L3-1 4. 보험수리적 가정의 변동에 의한 영향 합산 엑셀.

    agg_data: {
      'year','scanned','with_data_count',
      'rows': [{'company','currency','spot',
                'local_wage_growth_up', 'local_wage_growth_down',
                'local_discount_rate_up', 'local_discount_rate_down',
                'krw_*'}],
      'totals_by_item_krw': {key: sum},  # 4 keys
      'errors': [...]
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 28
    ws.cell(1, 1, 'L3-1 4. 보험수리적 가정의 변동 영향 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA

    # KRW 합계 — 가정 × 방향
    pairs = [
        ('wage_growth',   '기대임금상승율'),
        ('discount_rate', '할인율'),
    ]
    ws.cell(7, 1, 'KRW 환산 합계').font = Font(bold=True, color='1F3864', size=11)
    hdr = ['가정', '1% 상승 (KRW)', '1% 하락 (KRW)']
    for j, h in enumerate(hdr, 1):
        c = ws.cell(8, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
        c.alignment = Alignment(horizontal='center')
    totals = agg_data.get('totals_by_item_krw') or {}
    r = 9
    for key, lbl in pairs:
        ws.cell(r, 1, lbl).font = _XL_DATA
        u = ws.cell(r, 2, totals.get(f'{key}_up') or 0)
        u.font = _XL_DATA; u.number_format = _XL_NUM_FMT
        d = ws.cell(r, 3, totals.get(f'{key}_down') or 0)
        d.font = _XL_DATA; d.number_format = _XL_NUM_FMT
        r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L3-1 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세 — 회사 × (가정 × 방향)
    cols = [
        ('회사', 30, 'company', 'text'),
        ('통화', 8, 'currency', 'text'),
        ('기대임금상승율 1%↑(로컬)', 22, 'local_wage_growth_up',   'num'),
        ('기대임금상승율 1%↓(로컬)', 22, 'local_wage_growth_down', 'num'),
        ('할인율 1%↑(로컬)', 18, 'local_discount_rate_up',   'num'),
        ('할인율 1%↓(로컬)', 18, 'local_discount_rate_down', 'num'),
        ('Spot', 10, 'spot', 'rate'),
        ('기대임금상승율 1%↑(KRW)', 22, 'krw_wage_growth_up',   'num'),
        ('기대임금상승율 1%↓(KRW)', 22, 'krw_wage_growth_down', 'num'),
        ('할인율 1%↑(KRW)', 18, 'krw_discount_rate_up',   'num'),
        ('할인율 1%↓(KRW)', 18, 'krw_discount_rate_down', 'num'),
    ]
    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ds.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1

    wb.save(str(output_path))
    return str(output_path)


def build_l31_plan_breakdown_excel(agg_data, output_path):
    """L3-1 5. 사외적립자산 구성내역 합산 엑셀."""
    return _build_l3_pivot_excel(
        agg_data, output_path,
        title='L3-1 5. 사외적립자산의 구성내역 합산',
        item_keys=_L3_S3_KEYS,
        item_labels=_L3_S3_LBLS,
        extra_summary=[('합계 KRW', agg_data.get('grand_total_krw') or 0)],
    )


def build_l31_plan_managers_excel(agg_data, output_path):
    """L3-1 6. 사외적립자산 운용사 합산 엑셀 (L3 4번과 동일 빌더 재사용 가능)."""
    return build_l3_pension_managers_excel(agg_data, output_path)


def build_l3_pension_managers_excel(agg_data, output_path):
    """L3 4. 퇴직연금운용자산의 운용사 합산 엑셀 (다행 명세)."""
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 40
    ws.cell(1, 1, 'L3 4. 퇴직연금운용자산의 운용사 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA
    ws.cell(6, 1, '전체 KRW 합계').font = Font(bold=True, color='9C5700',
                                              name='Arial', size=11)
    gc = ws.cell(6, 2, agg_data.get('grand_total_krw') or 0)
    gc.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    gc.number_format = _XL_NUM_FMT

    by_cur = agg_data.get('total_by_currency') or {}
    r = 8
    if by_cur:
        ws.cell(r, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=11)
        r += 1
        headers = ['통화', '로컬 합계', 'KRW 합계']
        for j, h in enumerate(headers, 1):
            c = ws.cell(r, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
            c.alignment = Alignment(horizontal='center')
        r += 1
        for cur in sorted(by_cur.keys()):
            ws.cell(r, 1, cur).font = _XL_DATA
            lc = ws.cell(r, 2, by_cur[cur].get('local') or 0)
            lc.font = _XL_DATA; lc.number_format = _XL_NUM_FMT
            kc = ws.cell(r, 3, by_cur[cur].get('krw') or 0)
            kc.font = _XL_DATA; kc.number_format = _XL_NUM_FMT
            r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L3 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세
    cols = [
        ('회사', 30, 'company', 'text'),
        ('운용사', 22, 'name', 'text'),
        ('통화', 8, 'currency', 'text'),
        ('로컬 금액', 20, 'local', 'num'),
        ('Spot', 10, 'spot', 'rate'),
        ('KRW 환산', 20, 'krw', 'num'),
        ('비고', 40, 'remarks', 'text'),
    ]
    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ds.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1

    if agg_data.get('rows'):
        for j in range(1, len(cols) + 1):
            ds.cell(row, j).fill = _XL_TOTAL_FILL
        ds.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        tc = ds.cell(row, 6, agg_data.get('grand_total_krw') or 0)
        tc.font = _XL_TOTAL_FONT; tc.number_format = _XL_NUM_FMT

    wb.save(str(output_path))
    return str(output_path)


def build_l2_maturity_excel(agg_data, output_path):
    """L2 3. 부채성 금융상품 만기 분석 합산 엑셀."""
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    ws.cell(1, 1, 'L2 3. 부채성 금융상품 만기 분석 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA

    # 만기 구간별 KRW 합계
    intervals = [
        ('within_1y', '1년 미만'),
        ('within_2y', '1~2년'),
        ('within_5y', '2~5년'),
        ('over_5y',   '5년 초과'),
        ('total',     '합계'),
    ]
    r = 7
    ws.cell(r, 1, '만기 구간별 KRW 합계').font = Font(bold=True, color='1F3864', size=11)
    r += 1
    for key, lbl in intervals:
        ws.cell(r, 1, lbl).font = _XL_DATA
        cell = ws.cell(r, 2, agg_data.get(f'krw_{key}') or 0)
        cell.font = _XL_DATA
        if key == 'total':
            cell.font = Font(bold=True, color='9C5700', name='Arial', size=11)
        cell.number_format = _XL_NUM_FMT
        r += 1

    # 통화별 합계
    by_cur = agg_data.get('total_by_currency') or {}
    if by_cur:
        r += 1
        ws.cell(r, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=11)
        r += 1
        headers = ['통화'] + [f'{lbl}(로컬)' for _, lbl in intervals] \
                + [f'{lbl}(KRW)' for _, lbl in intervals]
        for j, h in enumerate(headers, 1):
            c = ws.cell(r, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
            c.alignment = Alignment(horizontal='center')
        r += 1
        for cur in sorted(by_cur.keys()):
            ws.cell(r, 1, cur).font = _XL_DATA
            j = 2
            for key, _l in intervals:
                cell = ws.cell(r, j, by_cur[cur].get(key) or 0)
                cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
                j += 1
            for key, _l in intervals:
                cell = ws.cell(r, j, by_cur[cur].get(f'krw_{key}') or 0)
                cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
                j += 1
            r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L2 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세
    cols = [
        ('회사',          30, 'company',        'text'),
        ('계정명',        18, 'account',        'text'),
        ('대주구분',      18, 'creditor_type',  'text'),
        ('통화',           8, 'currency',       'text'),
    ]
    for key, lbl in intervals:
        cols.append((f'{lbl}(로컬)', 18, f'local_{key}', 'num'))
    cols.append(('Spot', 10, 'spot', 'rate'))
    for key, lbl in intervals:
        cols.append((f'{lbl}(KRW)', 18, f'krw_{key}', 'num'))

    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ds.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1

    if agg_data.get('rows'):
        for j in range(1, len(cols) + 1):
            ds.cell(row, j).fill = _XL_TOTAL_FILL
        ds.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        # 마지막 컬럼이 합계(KRW)
        tc = ds.cell(row, len(cols), agg_data.get('krw_total') or 0)
        tc.font = _XL_TOTAL_FONT; tc.number_format = _XL_NUM_FMT

    wb.save(str(output_path))
    return str(output_path)


def build_a7_equity_method_excel(agg_data, output_path):
    """A7 1. 지분법투자주식 명세 합산 엑셀.

    agg_data: {
      'year', 'scanned', 'with_data_count',
      'rows': [{'company','type','type_label','investee','ownership_pct',
                'currency','spot',
                'local_cost','local_net_asset','local_book',
                'krw_cost','krw_net_asset','krw_book'}],
      'grand_cost_krw', 'grand_net_asset_krw', 'grand_book_krw',
      'subsidiary_count', 'other_count',
      'total_by_currency': {cur: {cost,net_asset,book,krw_cost,krw_net_asset,krw_book}},
      'errors': [{'company','reason'}],
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    ws.cell(1, 1, 'A7 1. 지분법투자주식 명세 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA
    ws.cell(6, 1, '종속회사 라인 수').font = _XL_DATA
    ws.cell(6, 2, agg_data.get('subsidiary_count') or 0).font = _XL_DATA
    ws.cell(7, 1, '기타지분법 라인 수').font = _XL_DATA
    ws.cell(7, 2, agg_data.get('other_count') or 0).font = _XL_DATA

    ws.cell(9, 1, '취득원가 KRW').font = _XL_DATA
    cc = ws.cell(9, 2, agg_data.get('grand_cost_krw') or 0)
    cc.font = _XL_DATA; cc.number_format = _XL_NUM_FMT
    ws.cell(10, 1, '순자산가액 KRW').font = _XL_DATA
    nc = ws.cell(10, 2, agg_data.get('grand_net_asset_krw') or 0)
    nc.font = _XL_DATA; nc.number_format = _XL_NUM_FMT
    ws.cell(11, 1, '장부가액 KRW').font = Font(bold=True, color='9C5700', name='Arial', size=11)
    bc = ws.cell(11, 2, agg_data.get('grand_book_krw') or 0)
    bc.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    bc.number_format = _XL_NUM_FMT

    # 통화별 합계
    by_cur = agg_data.get('total_by_currency') or {}
    r = 13
    if by_cur:
        ws.cell(r, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=11)
        r += 1
        headers = ['통화', '취득원가(로컬)', '순자산(로컬)', '장부가(로컬)',
                   '취득원가(KRW)', '순자산(KRW)', '장부가(KRW)']
        for j, h in enumerate(headers, 1):
            c = ws.cell(r, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
            c.alignment = Alignment(horizontal='center')
        r += 1
        for cur in sorted(by_cur.keys()):
            ws.cell(r, 1, cur).font = _XL_DATA
            for j, key in enumerate(['cost', 'net_asset', 'book',
                                     'krw_cost', 'krw_net_asset', 'krw_book'], 2):
                cell = ws.cell(r, j, by_cur[cur].get(key) or 0)
                cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
            r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'A7 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세 — 회사 / 종류 / 회사명 / 지분율 / 통화 / 로컬 3 / Spot / KRW 3
    cols = [
        ('회사',         30, 'company',         'text'),
        ('종류',         24, 'type_label',      'text'),
        ('회사명',       36, 'investee',        'text'),
        ('지분율(%)',    12, 'ownership_pct',   'pct'),
        ('통화',          8, 'currency',        'text'),
        ('취득원가(로컬)', 20, 'local_cost',      'num'),
        ('순자산(로컬)',  20, 'local_net_asset', 'num'),
        ('장부가(로컬)',  20, 'local_book',      'num'),
        ('Spot',         10, 'spot',            'rate'),
        ('취득원가(KRW)', 20, 'krw_cost',        'num'),
        ('순자산(KRW)',   20, 'krw_net_asset',   'num'),
        ('장부가(KRW)',   20, 'krw_book',        'num'),
    ]
    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            elif fmt == 'pct':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else '')
                if isinstance(v, (int, float)):
                    cell.number_format = '0.0000%'
            else:
                cell = ds.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1

    if agg_data.get('rows'):
        for j in range(1, len(cols) + 1):
            ds.cell(row, j).fill = _XL_TOTAL_FILL
        ds.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        for j, key in [(10, 'grand_cost_krw'), (11, 'grand_net_asset_krw'),
                       (12, 'grand_book_krw')]:
            tc = ds.cell(row, j, agg_data.get(key) or 0)
            tc.font = _XL_TOTAL_FONT; tc.number_format = _XL_NUM_FMT

    wb.save(str(output_path))
    return str(output_path)


def build_a6_derivatives_excel(agg_data, output_path):
    """A6 1. 파생상품평가손익 합산 엑셀.

    agg_data: {
      'year', 'scanned', 'with_data_count',
      'rows': [{'company','type','currency','spot',
                'local_gain','local_loss','krw_gain','krw_loss'}],
      'grand_gain_krw', 'grand_loss_krw', 'grand_net_krw',
      'total_by_currency': {cur: {gain,loss,gain_krw,loss_krw}},
      'errors': [{'company','reason'}],
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    ws.cell(1, 1, 'A6 1. 파생상품평가손익 내역 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA
    ws.cell(6, 1, '평가이익 KRW').font = _XL_DATA
    g = ws.cell(6, 2, agg_data.get('grand_gain_krw') or 0)
    g.font = _XL_DATA; g.number_format = _XL_NUM_FMT
    ws.cell(7, 1, '평가손실 KRW').font = _XL_DATA
    l = ws.cell(7, 2, agg_data.get('grand_loss_krw') or 0)
    l.font = _XL_DATA; l.number_format = _XL_NUM_FMT
    ws.cell(8, 1, '순손익 KRW').font = Font(bold=True, color='9C5700', name='Arial', size=11)
    n = ws.cell(8, 2, agg_data.get('grand_net_krw') or 0)
    n.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    n.number_format = _XL_NUM_FMT

    # 통화별 합계
    by_cur = agg_data.get('total_by_currency') or {}
    r = 10
    if by_cur:
        ws.cell(r, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=11)
        r += 1
        headers = ['통화', '평가이익(로컬)', '평가손실(로컬)',
                   '평가이익(KRW)', '평가손실(KRW)']
        for j, h in enumerate(headers, 1):
            c = ws.cell(r, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
            c.alignment = Alignment(horizontal='center')
        r += 1
        for cur in sorted(by_cur.keys()):
            ws.cell(r, 1, cur).font = _XL_DATA
            for j, key in enumerate(['gain', 'loss', 'gain_krw', 'loss_krw'], 2):
                cell = ws.cell(r, j, by_cur[cur].get(key) or 0)
                cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
            r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'A6 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세 — 회사 / 파생상품 종류 / 통화 / 로컬 / Spot / KRW
    cols = [
        ('회사',           30, 'company',    'text'),
        ('파생상품 종류',  36, 'type',       'text'),
        ('통화',            8, 'currency',   'text'),
        ('평가이익(로컬)', 20, 'local_gain', 'num'),
        ('평가손실(로컬)', 20, 'local_loss', 'num'),
        ('Spot',           10, 'spot',       'rate'),
        ('평가이익(KRW)',  20, 'krw_gain',   'num'),
        ('평가손실(KRW)',  20, 'krw_loss',   'num'),
    ]
    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ds.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1

    # 합계 행
    if agg_data.get('rows'):
        for j in range(1, len(cols) + 1):
            ds.cell(row, j).fill = _XL_TOTAL_FILL
        ds.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        gtot = ds.cell(row, 7, agg_data.get('grand_gain_krw') or 0)
        gtot.font = _XL_TOTAL_FONT; gtot.number_format = _XL_NUM_FMT
        ltot = ds.cell(row, 8, agg_data.get('grand_loss_krw') or 0)
        ltot.font = _XL_TOTAL_FONT; ltot.number_format = _XL_NUM_FMT

    wb.save(str(output_path))
    return str(output_path)


def build_a5_lease_pl_excel(agg_data, output_path):
    """A5 2. 리스계약 관련 손익 합산 엑셀.

    agg_data: {
      'year', 'scanned', 'with_data_count',
      'rows': [{'company','currency','spot',
                'local_<item>','krw_<item>' (6개 항목),
                'local_total','krw_total'}],
      'totals_by_item_krw': {item_key: sum},
      'grand_total_krw',
      'total_by_currency': {cur: {...}},
      'errors': [{'company','reason'}],
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    item_labels = [
        ('depreciation',  '사용권자산 감가상각비'),
        ('interest',      '리스부채 이자비용'),
        ('short_term',    '단기리스 관련비용'),
        ('low_value',     '소액자산리스 관련비용'),
        ('variable',      '변동리스료 관련비용'),
        ('disposal_gain', '리스처분이익'),
    ]

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    ws.cell(1, 1, 'A5 2. 리스계약 관련 손익 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA
    ws.cell(6, 1, '합계 KRW').font = _XL_DATA
    gc = ws.cell(6, 2, agg_data.get('grand_total_krw') or 0)
    gc.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    gc.number_format = _XL_NUM_FMT

    # 항목별 KRW 합계
    ws.cell(8, 1, '항목별 KRW 합계').font = Font(bold=True, color='1F3864', size=11)
    totals = agg_data.get('totals_by_item_krw') or {}
    r = 9
    for key, lbl in item_labels:
        ws.cell(r, 1, lbl).font = _XL_DATA
        cell = ws.cell(r, 2, totals.get(key) or 0)
        cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
        r += 1
    # 전체 합계 행
    ws.cell(r, 1, '합계').font = _XL_TOTAL_FONT
    tc = ws.cell(r, 2, agg_data.get('grand_total_krw') or 0)
    tc.font = _XL_TOTAL_FONT; tc.number_format = _XL_NUM_FMT
    r += 2

    errs = agg_data.get('errors') or []
    if errs:
        ws.cell(r, 1, 'A5 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세
    item_keys = [k for k, _ in item_labels]
    cols = [('회사', 30, 'company', 'text'),
            ('통화', 8, 'currency', 'text')]
    for k, lbl in item_labels:
        cols.append((f'{lbl}(로컬)', 22, f'local_{k}', 'num'))
    cols.append(('합계(로컬)', 22, 'local_total', 'num'))
    cols.append(('Spot', 10, 'spot', 'rate'))
    for k, lbl in item_labels:
        cols.append((f'{lbl}(KRW)', 22, f'krw_{k}', 'num'))
    cols.append(('합계(KRW)', 22, 'krw_total', 'num'))

    ds = wb.create_sheet('상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r.get(key)
            if fmt == 'num':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(row, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            else:
                cell = ds.cell(row, j, v if v is not None else '')
            cell.font = _XL_DATA
        row += 1

    wb.save(str(output_path))
    return str(output_path)


def build_l4_other_commitments_excel(agg_data, output_path):
    """11번 그외 우발부채 및 약정사항 — 내용 있는 회사 목록 엑셀.

    agg_data: {
      'year', 'scanned',
      'with_content_count', 'no_content_count',
      'rows': [{'company','content'}],   # 내용 있는 회사만
      'errors': [{'company','reason'}],
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 60
    ws.cell(1, 1, 'L4 11번 그외 우발부채 및 약정사항 (내용 있는 회사)').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '내용 있는 회사 수').font = Font(bold=True, color='9C5700',
                                              name='Arial', size=11)
    ws.cell(5, 2, agg_data.get('with_content_count') or 0).font = Font(
        bold=True, color='9C5700', name='Arial', size=11)
    ws.cell(6, 1, '내용 없는 회사 수').font = _XL_DATA
    ws.cell(6, 2, agg_data.get('no_content_count') or 0).font = _XL_DATA

    errs = agg_data.get('errors') or []
    if errs:
        r = 8
        ws.cell(r, 1, 'L4 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세 (내용 있는 회사만)
    ds = wb.create_sheet('내용')
    ds.column_dimensions['A'].width = 32
    ds.column_dimensions['B'].width = 110
    ds.freeze_panes = 'A2'
    headers = ['회사', '약정·우발부채 내용']
    for j, h in enumerate(headers, 1):
        cell = ds.cell(1, j, h)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        ds.cell(row, 1, r.get('company') or '').font = _XL_DATA
        cc = ds.cell(row, 2, r.get('content') or '')
        cc.font = _XL_DATA
        cc.alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    wb.save(str(output_path))
    return str(output_path)


def build_l4_subsequent_events_excel(agg_data, output_path):
    """10번 보고기간일 이후 사건 — YES 회사 목록 엑셀.

    agg_data: {
      'year', 'scanned', 'yes_count', 'no_count', 'empty_count',
      'rows': [{'company','yn','content'}],   # YES 회사만
      'errors': [{'company','reason'}],
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 60
    ws.cell(1, 1, 'L4 10번 보고기간일 이후 사건 (YES 회사)').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, 'YES 회사 수').font = Font(bold=True, color='9C5700',
                                           name='Arial', size=11)
    ws.cell(5, 2, agg_data.get('yes_count') or 0).font = Font(
        bold=True, color='9C5700', name='Arial', size=11)
    ws.cell(6, 1, 'NO 회사 수').font = _XL_DATA
    ws.cell(6, 2, agg_data.get('no_count') or 0).font = _XL_DATA
    ws.cell(7, 1, '미입력 회사 수').font = _XL_DATA
    ws.cell(7, 2, agg_data.get('empty_count') or 0).font = _XL_DATA

    errs = agg_data.get('errors') or []
    if errs:
        r = 9
        ws.cell(r, 1, 'L4 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 상세 (YES 회사만)
    ds = wb.create_sheet('YES 회사')
    ds.column_dimensions['A'].width = 32
    ds.column_dimensions['B'].width = 10
    ds.column_dimensions['C'].width = 100
    ds.freeze_panes = 'A2'
    headers = ['회사', '응답', '사건 내용']
    for j, h in enumerate(headers, 1):
        cell = ds.cell(1, j, h)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        ds.cell(row, 1, r.get('company') or '').font = _XL_DATA
        yc = ds.cell(row, 2, r.get('yn') or '')
        yc.font = Font(bold=True, color='9C5700', name='Arial', size=10)
        yc.alignment = Alignment(horizontal='center', vertical='top')
        cc = ds.cell(row, 3, r.get('content') or '')
        cc.font = _XL_DATA
        cc.alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    wb.save(str(output_path))
    return str(output_path)


def build_l4_lawsuits_excel(agg_data, output_path):
    """6-1 소송 합산 엑셀. 회사 패키지 통화/spot으로 KRW 환산.

    agg_data: {
      'year', 'scanned', 'with_rows',
      'rows': [{'company','type','count','currency','spot',
                'claim_amount','provision_amount',
                'claim_amount_krw','provision_amount_krw'}],
      'total_count', 'total_claim_krw', 'total_provision_krw',
      'plaintiff_count', 'plaintiff_claim_krw', 'plaintiff_provision_krw',
      'defendant_count', 'defendant_claim_krw',
      'total_by_currency': {cur: {count,claim,claim_krw,provision,provision_krw}},
      'errors': [{'company','reason'}],
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 표지
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    ws.cell(1, 1, 'L4 6-1 소송중인 사건 합산').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '소송 입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_rows') or 0).font = _XL_DATA

    # 구분별 합계 표 — KRW 환산 기준
    ws.cell(7, 1, '구분별 합계 (KRW 환산)').font = Font(bold=True, color='1F3864', size=12)

    headers = ['구분', '소송건수', '소송금액(KRW)', '인식한 충당부채(KRW)']
    for j, h in enumerate(headers, 1):
        c = ws.cell(8, j, h)
        c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')

    # 원고 행 (r9) — 충당부채 N/A
    ws.cell(9, 1, '원고 (Defendant)').font = Font(
        bold=True, color='1F3864', name='Arial', size=11)
    dc = ws.cell(9, 2, agg_data.get('defendant_count') or 0)
    dc.font = _XL_DATA; dc.number_format = '#,##0'
    dl = ws.cell(9, 3, agg_data.get('defendant_claim_krw') or 0)
    dl.font = _XL_DATA; dl.number_format = _XL_NUM_FMT
    nd = ws.cell(9, 4, '-')
    nd.font = Font(color='9aa9bd', name='Arial', size=10)
    nd.alignment = Alignment(horizontal='right')

    # 피고 행 (r10) — 충당부채 인식
    ws.cell(10, 1, '피고 (Plaintiff)').font = Font(
        bold=True, color='9C5700', name='Arial', size=11)
    pc = ws.cell(10, 2, agg_data.get('plaintiff_count') or 0)
    pc.font = _XL_DATA; pc.number_format = '#,##0'
    pl = ws.cell(10, 3, agg_data.get('plaintiff_claim_krw') or 0)
    pl.font = _XL_DATA; pl.number_format = _XL_NUM_FMT
    pp = ws.cell(10, 4, agg_data.get('plaintiff_provision_krw') or 0)
    pp.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    pp.number_format = _XL_NUM_FMT

    # 합계 행 (r11)
    for j in range(1, 5):
        ws.cell(11, j).fill = _XL_TOTAL_FILL
    ws.cell(11, 1, '전체 합계').font = _XL_TOTAL_FONT
    tc = ws.cell(11, 2, agg_data.get('total_count') or 0)
    tc.font = _XL_TOTAL_FONT; tc.number_format = '#,##0'
    tl = ws.cell(11, 3, agg_data.get('total_claim_krw') or 0)
    tl.font = _XL_TOTAL_FONT; tl.number_format = _XL_NUM_FMT
    tp = ws.cell(11, 4, agg_data.get('total_provision_krw') or 0)
    tp.font = _XL_TOTAL_FONT; tp.number_format = _XL_NUM_FMT

    # 통화별 합계 (로컬 / KRW)
    next_r = 13
    by_cur = agg_data.get('total_by_currency') or {}
    if by_cur:
        ws.cell(next_r, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=12)
        next_r += 1
        cur_headers = ['통화', '소송건수', '소송금액(로컬)', '충당부채(로컬)',
                       '소송금액(KRW)', '충당부채(KRW)']
        for j, h in enumerate(cur_headers, 1):
            c = ws.cell(next_r, j, h)
            c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
            c.alignment = Alignment(horizontal='center')
        next_r += 1
        for cur in sorted(by_cur.keys()):
            row_data = by_cur[cur]
            ws.cell(next_r, 1, cur).font = _XL_DATA
            cnt_cell = ws.cell(next_r, 2, row_data.get('count') or 0)
            cnt_cell.font = _XL_DATA; cnt_cell.number_format = '#,##0'
            for j, key in [(3, 'claim'), (4, 'provision'),
                           (5, 'claim_krw'), (6, 'provision_krw')]:
                cell = ws.cell(next_r, j, row_data.get(key) or 0)
                cell.font = _XL_DATA; cell.number_format = _XL_NUM_FMT
            next_r += 1
        next_r += 1

    errs = agg_data.get('errors') or []
    if errs:
        ws.cell(next_r, 1, 'L4 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        next_r += 1
        for e in errs:
            ws.cell(next_r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(next_r, 2, e.get('reason') or '').font = _XL_DATA
            next_r += 1

    # 상세
    ds = wb.create_sheet('상세')
    for j, (_lbl, w) in enumerate(_L4_LAW_COLS, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w) in enumerate(_L4_LAW_COLS, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        ds.cell(row, 1, r.get('company') or '').font = _XL_DATA
        ds.cell(row, 2, r.get('type') or '').font = _XL_DATA
        cc = ds.cell(row, 3, r.get('count') or 0)
        cc.font = _XL_DATA; cc.number_format = '#,##0'
        ds.cell(row, 4, r.get('currency') or '').font = _XL_DATA
        cl = ds.cell(row, 5, r.get('claim_amount') or 0)
        cl.font = _XL_DATA; cl.number_format = _XL_NUM_FMT
        pv = ds.cell(row, 6, r.get('provision_amount') or 0)
        pv.font = _XL_DATA; pv.number_format = _XL_NUM_FMT
        sc = ds.cell(row, 7, r.get('spot') or 0)
        sc.font = _XL_DATA; sc.number_format = _XL_RATE_FMT
        ckrw = ds.cell(row, 8, r.get('claim_amount_krw') or 0)
        ckrw.font = _XL_DATA; ckrw.number_format = _XL_NUM_FMT
        pkrw = ds.cell(row, 9, r.get('provision_amount_krw') or 0)
        pkrw.font = _XL_DATA; pkrw.number_format = _XL_NUM_FMT
        row += 1

    if agg_data.get('rows'):
        for j in range(1, len(_L4_LAW_COLS) + 1):
            ds.cell(row, j).fill = _XL_TOTAL_FILL
        ds.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        tc = ds.cell(row, 3, agg_data.get('total_count') or 0)
        tc.font = _XL_TOTAL_FONT; tc.number_format = '#,##0'
        tcl = ds.cell(row, 8, agg_data.get('total_claim_krw') or 0)
        tcl.font = _XL_TOTAL_FONT; tcl.number_format = _XL_NUM_FMT
        tpv = ds.cell(row, 9, agg_data.get('total_provision_krw') or 0)
        tpv.font = _XL_TOTAL_FONT; tpv.number_format = _XL_NUM_FMT

    wb.save(str(output_path))
    return str(output_path)


# ──────────────────────────────────────────────────────────────
# 엑셀 빌더 (L4 Yes/No + 금액 패턴 — 2/2-1, 3/3-1 공통)
# ──────────────────────────────────────────────────────────────

def _write_l4_qna_cover(wb, agg_data, title, yn_label):
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 50

    ws.cell(1, 1, title).font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, f'{yn_label} YES 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('yes_count') or 0).font = _XL_DATA
    ws.cell(6, 1, '전체 KRW 합계').font = _XL_DATA
    grand = ws.cell(6, 2, agg_data.get('grand_total_krw') or 0)
    grand.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    grand.number_format = _XL_NUM_FMT

    by_cur = agg_data.get('total_by_currency') or {}
    if by_cur:
        ws.cell(8, 1, '통화별 합계').font = Font(bold=True, color='1F3864', size=11)
        headers = ['통화', '로컬 합계', 'KRW 합계']
        for j, h in enumerate(headers, 1):
            c = ws.cell(9, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
            c.alignment = Alignment(horizontal='center')
        r = 10
        for cur in sorted(by_cur.keys()):
            ws.cell(r, 1, cur).font = _XL_DATA
            lc = ws.cell(r, 2, by_cur[cur].get('local') or 0)
            lc.font = _XL_DATA; lc.number_format = _XL_NUM_FMT
            kc = ws.cell(r, 3, by_cur[cur].get('krw') or 0)
            kc.font = _XL_DATA; kc.number_format = _XL_NUM_FMT
            r += 1
    else:
        r = 10

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'L4 시트 누락/오류 회사').font = Font(bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1


def _write_l4_qna_combined_sheet(wb, agg_data, yn_label, amount_label, yn_key):
    cols = [
        ('회사',         30),
        (yn_label,       18),
        (amount_label,   22),
        ('통화',          8),
        ('Spot',         10),
        ('KRW 환산',    20),
    ]
    ws = wb.create_sheet('상세')
    for j, (_lbl, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'

    for j, (label, _w) in enumerate(cols, 1):
        cell = ws.cell(1, j, label)
        cell.font = _XL_HDR_FONT
        cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for r in (agg_data.get('rows') or []):
        ws.cell(row, 1, r.get('company') or '').font = _XL_DATA
        ws.cell(row, 2, r.get(yn_key) or '').font = _XL_DATA
        lc = ws.cell(row, 3, r.get('amount') if r.get('amount') is not None else '')
        lc.font = _XL_DATA
        if isinstance(r.get('amount'), (int, float)):
            lc.number_format = _XL_NUM_FMT
        ws.cell(row, 4, r.get('currency') or '').font = _XL_DATA
        sc = ws.cell(row, 5, r.get('spot') if r.get('spot') is not None else '')
        sc.font = _XL_DATA
        if isinstance(r.get('spot'), (int, float)):
            sc.number_format = _XL_RATE_FMT
        kc = ws.cell(row, 6, r.get('krw') if r.get('krw') is not None else '')
        kc.font = _XL_DATA
        if isinstance(r.get('krw'), (int, float)):
            kc.number_format = _XL_NUM_FMT
        row += 1

    if agg_data.get('rows'):
        for j in range(1, len(cols) + 1):
            ws.cell(row, j).fill = _XL_TOTAL_FILL
        ws.cell(row, 1, '전체 합계').font = _XL_TOTAL_FONT
        gc = ws.cell(row, 6, agg_data.get('grand_total_krw') or 0)
        gc.font = _XL_TOTAL_FONT
        gc.number_format = _XL_NUM_FMT


def build_l4lc_excel(agg_data, output_path):
    """L4 수입신용장(2/2-1) 합산 결과(JSON)를 엑셀로 저장."""
    wb = Workbook()
    wb.remove(wb.active)
    _write_l4_qna_cover(wb, agg_data,
                        title='L4 수입신용장 / 미확정 지급보증 합산',
                        yn_label='수입신용장 오픈')
    _write_l4_qna_combined_sheet(wb, agg_data,
                                 yn_label='수입신용장 오픈 여부',
                                 amount_label='미확정 지급보증 실행금액',
                                 yn_key='lc_open')
    wb.save(str(output_path))
    return str(output_path)


def build_l4_export_excel(agg_data, output_path):
    """L4 수출채권 할인(3/3-1) 합산 결과(JSON)를 엑셀로 저장."""
    wb = Workbook()
    wb.remove(wb.active)
    _write_l4_qna_cover(wb, agg_data,
                        title='L4 수출채권 할인 합산',
                        yn_label='수출채권 할인')
    _write_l4_qna_combined_sheet(wb, agg_data,
                                 yn_label='수출채권 할인 여부',
                                 amount_label='만기 미도래 할인금액',
                                 yn_key='discount_done')
    wb.save(str(output_path))
    return str(output_path)


# ──────────────────────────────────────────────────────────────────────
# TX 시트 빌더 — 법인세 1/3/3-1/4/5/5-1
# ──────────────────────────────────────────────────────────────────────

def build_tx_deferred_tax_changes_excel(agg_data, output_path):
    """TX 1. 이연법인세자산(부채) 증감내용 합산 엑셀.

    카테고리 = (kor_label, current_flag).
    agg_data: {
      'year','scanned','with_data_count',
      'categories': [{'kor_label','current_flag',
                      'company_count',
                      'beginning_krw','ending_krw','change_krw'}],
      'totals_krw': {'beginning','ending','change'},
      'company_rows': [{'company','currency','spot','prior_spot',
                        'statutory_rate','kor_label','current_flag',
                        'local_beginning','local_ending','local_change',
                        'krw_beginning','krw_ending','krw_change',
                        'beginning_temp_diff','ending_temp_diff'}],
      'errors':[...]
    }
    기말 일시적차이 = 기말(KRW) ÷ 법정세율,
    기초 일시적차이 = 전년 4분기 주석의 기말 일시적차이(없으면 None).
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.cell(1, 1, 'TX 1. 이연법인세자산(부채) 증감내용 합산 '
                  '(기초=전기말 Spot / 기말=당기말 Spot)').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA

    totals = agg_data.get('totals_krw') or {}
    ws.cell(7, 1, '전체 KRW 합계').font = Font(bold=True, color='1F3864', size=11)
    for i, (lbl, key) in enumerate([('기초 KRW', 'beginning'),
                                     ('기말 KRW', 'ending'),
                                     ('증감 KRW', 'change')]):
        ws.cell(8 + i, 1, lbl).font = _XL_DATA
        c = ws.cell(8 + i, 2, totals.get(key) or 0)
        c.font = _XL_DATA; c.number_format = _XL_NUM_FMT

    # 카테고리별 합계 표
    r = 12
    ws.cell(r, 1, '카테고리별 합계 (계정명 × Current/Non-Current)').font = Font(
        bold=True, color='1F3864', size=11)
    r += 1
    headers = ['계정명', 'Current/Non-Current', '회사수',
               '기초(KRW)', '기말(KRW)', '증감(KRW)']
    for j, h in enumerate(headers, 1):
        cell = ws.cell(r, j, h); cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center')
    r += 1
    for cat in (agg_data.get('categories') or []):
        ws.cell(r, 1, cat.get('kor_label') or '').font = _XL_DATA
        ws.cell(r, 2, cat.get('current_flag') or '').font = _XL_DATA
        cc = ws.cell(r, 3, cat.get('company_count') or 0)
        cc.font = _XL_DATA; cc.number_format = '#,##0'
        for j, k in enumerate(['beginning_krw', 'ending_krw', 'change_krw'], 4):
            c = ws.cell(r, j, cat.get(k) or 0)
            c.font = _XL_DATA; c.number_format = _XL_NUM_FMT
        r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'TX 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 회사별 상세
    cols = [
        ('회사', 28, 'company', 'text'),
        ('통화', 8, 'currency', 'text'),
        ('계정명', 28, 'kor_label', 'text'),
        ('유동/비유동', 14, 'current_flag', 'text'),
        ('기초(로컬)', 18, 'local_beginning', 'num'),
        ('기말(로컬)', 18, 'local_ending', 'num'),
        ('증감(로컬)', 18, 'local_change', 'num'),
        ('전기말 Spot', 12, 'prior_spot', 'rate'),
        ('당기말 Spot', 12, 'spot', 'rate'),
        ('기초(KRW)', 18, 'krw_beginning', 'num'),
        ('기말(KRW)', 18, 'krw_ending', 'num'),
        ('증감(KRW)', 18, 'krw_change', 'num'),
        ('법정세율', 12, 'statutory_rate', 'pct'),
        ('기초 일시적차이', 20, 'beginning_temp_diff', 'numn'),
        ('기말 일시적차이', 20, 'ending_temp_diff', 'numn'),
    ]
    ds = wb.create_sheet('회사별 상세')
    for j, (_lbl, w, _k, _f) in enumerate(cols, 1):
        ds.column_dimensions[get_column_letter(j)].width = w
    ds.freeze_panes = 'A2'
    for j, (label, _w, _k, _f) in enumerate(cols, 1):
        cell = ds.cell(1, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    rr = 2
    for r_row in (agg_data.get('company_rows') or []):
        for j, (_lbl, _w, key, fmt) in enumerate(cols, 1):
            v = r_row.get(key)
            if fmt == 'num':
                cell = ds.cell(rr, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'numn':  # 값 없으면 빈 칸 (0과 구분)
                cell = ds.cell(rr, j, v if isinstance(v, (int, float)) else '')
                if isinstance(v, (int, float)):
                    cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ds.cell(rr, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            elif fmt == 'pct':
                cell = ds.cell(rr, j, v if isinstance(v, (int, float)) else '')
                if isinstance(v, (int, float)):
                    cell.number_format = '0.0000%'
            else:
                cell = ds.cell(rr, j, v if v is not None else '')
            cell.font = _XL_DATA
        rr += 1

    wb.save(str(output_path))
    return str(output_path)


def build_tx_income_tax_breakdown_excel(agg_data, output_path):
    """TX 3. 법인세비용의 구성내역 합산 (avg 환율)."""
    return _build_l3_pivot_excel(
        agg_data, output_path,
        title='TX 3. 법인세비용의 구성내역 합산 (Avg 환율)',
        item_keys=['current_tax', 'deferred_temp_diff', 'deferred_equity',
                   'additional_refund', 'total_expense'],
        item_labels=['당기법인세부담액', '일시적차이 이연법인세 변동액',
                     '자본 직접부과 이연법인세 변동액',
                     '법인세 추납액(환급액)', '법인세비용 총계'],
    )


def build_tx_equity_deferred_tax_excel(agg_data, output_path):
    """TX 3-1. 자본 직접 부과 이연법인세 변동액 명세 합산 (avg 환율)."""
    return _build_l3_pivot_excel(
        agg_data, output_path,
        title='TX 3-1. 자본 직접 부과 이연법인세 변동액 명세 합산 (Avg 환율)',
        item_keys=['revaluation', 'actuarial', 'afs_securities', 'fvoci',
                   'equity_method', 'fx_translation', 'derivatives',
                   'others', 'total'],
        item_labels=['재평가이익', '보험수리적손익', '매도가능증권평가손익',
                     'FVOCI 금융상품', '지분법적용투자주식',
                     '해외사업환산손익', '파생상품평가손익', '기타', '합계'],
    )


def build_tx_reconciliation_excel(agg_data, output_path):
    """TX 4. Reconciliation 합산 (avg 환율) + 회사별 유효세율 표시."""
    return _build_l3_pivot_excel(
        agg_data, output_path,
        title='TX 4. 법인세 reconciliation 합산 (Avg 환율)',
        item_keys=['pretax_income', 'tax_at_statutory', 'permanent_diff',
                   'tax_credit', 'additional_refund', 'unrecognized_change',
                   'total_expense'],
        item_labels=['법인세차감전 순이익', '적용세율 산출세액',
                     '영구적차이', '세액공제', '법인세 추납액/환급액',
                     '미인식 일시적차이 변동', '법인세비용'],
        extra_summary=[('* 유효세율은 회사별 상세 시트 참조', 0)],
    )


def build_tx_unrecognized_excel(agg_data, output_path):
    """TX 5. 이연법인세로 인식되지 않은 일시적차이 합산 (spot 환율)."""
    return _build_l3_pivot_excel(
        agg_data, output_path,
        title='TX 5. 미인식 일시적차이 합산 (Spot 환율)',
        item_keys=['loss_carryforward', 'others'],
        item_labels=['이월결손금', '기타'],
    )


def build_tx_loss_maturity_excel(agg_data, output_path):
    """TX 5-1. 이월결손금 만기 합산 (spot 환율).

    agg_data: {
      'year','scanned','with_data_count',
      'buckets': [{'label','position','total_krw','company_count'}],
      'grand_total_krw',
      'company_rows': [{'company','currency','spot',
                        'buckets':[{label,local,krw}]}],
      'errors': [...]
    }
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 14
    ws.cell(1, 1, 'TX 5-1. 이월결손금 만기 합산 (Spot 환율)').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, agg_data.get('year') or '').font = _XL_DATA
    ws.cell(4, 1, '스캔 회사 수').font = _XL_DATA
    ws.cell(4, 2, agg_data.get('scanned') or 0).font = _XL_DATA
    ws.cell(5, 1, '입력 회사 수').font = _XL_DATA
    ws.cell(5, 2, agg_data.get('with_data_count') or 0).font = _XL_DATA
    ws.cell(6, 1, '전체 KRW 합계').font = Font(bold=True, color='9C5700',
                                                name='Arial', size=11)
    gc = ws.cell(6, 2, agg_data.get('grand_total_krw') or 0)
    gc.font = Font(bold=True, color='9C5700', name='Arial', size=11)
    gc.number_format = _XL_NUM_FMT

    r = 8
    ws.cell(r, 1, '만기 buckets').font = Font(bold=True, color='1F3864', size=11)
    r += 1
    for j, h in enumerate(['만기', '회사수', 'KRW 합계'], 1):
        cell = ws.cell(r, j, h); cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center')
    r += 1
    for b in (agg_data.get('buckets') or []):
        ws.cell(r, 1, b.get('label') or '').font = _XL_DATA
        cc = ws.cell(r, 2, b.get('company_count') or 0)
        cc.font = _XL_DATA; cc.number_format = '#,##0'
        kc = ws.cell(r, 3, b.get('total_krw') or 0)
        kc.font = _XL_DATA; kc.number_format = _XL_NUM_FMT
        r += 1

    errs = agg_data.get('errors') or []
    if errs:
        r += 1
        ws.cell(r, 1, 'TX 시트 누락/오류 회사').font = Font(
            bold=True, color='b54200', size=11)
        r += 1
        for e in errs:
            ws.cell(r, 1, e.get('company') or '').font = _XL_DATA
            ws.cell(r, 2, e.get('reason') or '').font = _XL_DATA
            r += 1

    # 회사별 상세 — 만기 컬럼 동적
    bucket_labels = [b.get('label') for b in (agg_data.get('buckets') or [])]
    ds = wb.create_sheet('회사별 상세')
    headers = ['회사', '통화', 'Spot'] \
              + [f'{lbl}(로컬)' for lbl in bucket_labels] \
              + [f'{lbl}(KRW)' for lbl in bucket_labels]
    for j, h in enumerate(headers, 1):
        ds.column_dimensions[get_column_letter(j)].width = 16
        cell = ds.cell(1, j, h)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ds.freeze_panes = 'A2'
    rr = 2
    for row in (agg_data.get('company_rows') or []):
        ds.cell(rr, 1, row.get('company') or '').font = _XL_DATA
        ds.cell(rr, 2, row.get('currency') or '').font = _XL_DATA
        sc = ds.cell(rr, 3, row.get('spot') or 0)
        sc.font = _XL_DATA; sc.number_format = _XL_RATE_FMT
        bd = {b.get('label'): b for b in (row.get('buckets') or [])}
        for j, lbl in enumerate(bucket_labels):
            v = (bd.get(lbl) or {}).get('local') or 0
            c = ds.cell(rr, 4 + j, v); c.font = _XL_DATA
            c.number_format = _XL_NUM_FMT
        for j, lbl in enumerate(bucket_labels):
            v = (bd.get(lbl) or {}).get('krw') or 0
            c = ds.cell(rr, 4 + len(bucket_labels) + j, v); c.font = _XL_DATA
            c.number_format = _XL_NUM_FMT
        rr += 1

    wb.save(str(output_path))
    return str(output_path)


# ──────────────────────────────────────────────────────────────────────
# 통합 다운로드 빌더 — 모든 합산 결과를 단일 워크북에 시트별로 작성
# ──────────────────────────────────────────────────────────────────────

# Excel 시트명 31자 제한
def _safe_sheet_name(name):
    s = str(name)[:31]
    # Excel에서 금지되는 문자 치환
    for ch in '/\\?*[]:':
        s = s.replace(ch, '_')
    return s.strip()


def _aio_write_section(wb, sheet_name, title, info_pairs, col_defs, rows,
                       totals_row=None):
    """범용 시트 작성 헬퍼.
      title: 시트 1행 제목
      info_pairs: [(label, value, fmt)] — 상단 요약 (fmt='text'|'num'|'rate'|'int')
      col_defs: [(label, width, key, fmt)] — 표 컬럼 정의
      rows: [dict]
      totals_row: dict | None — 마지막 합계 행 (col key → value). None이면 생략.
    반환: ws (성공) | None.
    """
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))
    ws.cell(1, 1, title).font = _XL_TITLE

    # 상단 요약 (라벨/값 한 줄씩)
    r = 3
    for item in info_pairs:
        if not item:
            r += 1
            continue
        label, value, fmt = item
        ws.cell(r, 1, label).font = _XL_DATA
        cell = ws.cell(r, 2, value if value is not None else '')
        cell.font = _XL_DATA
        if fmt == 'num':
            cell.number_format = _XL_NUM_FMT
        elif fmt == 'rate':
            cell.number_format = _XL_RATE_FMT
        elif fmt == 'int':
            cell.number_format = '#,##0'
        r += 1

    # 헤더 행
    r += 1
    header_r = r
    for j, (label, w, _k, _f) in enumerate(col_defs, 1):
        cell = ws.cell(header_r, j, label)
        cell.font = _XL_HDR_FONT; cell.fill = _XL_HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(header_r + 1, 1).coordinate
    r += 1

    # 데이터
    for row in (rows or []):
        for j, (_lbl, _w, key, fmt) in enumerate(col_defs, 1):
            v = row.get(key)
            if fmt == 'num':
                cell = ws.cell(r, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'numn':  # 값 없으면 빈 칸 (0과 구분)
                cell = ws.cell(r, j, v if isinstance(v, (int, float)) else '')
                if isinstance(v, (int, float)):
                    cell.number_format = _XL_NUM_FMT
            elif fmt == 'rate':
                cell = ws.cell(r, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = _XL_RATE_FMT
            elif fmt == 'int':
                cell = ws.cell(r, j, v if isinstance(v, (int, float)) else 0)
                cell.number_format = '#,##0'
            elif fmt == 'pct':
                cell = ws.cell(r, j, v if isinstance(v, (int, float)) else '')
                if isinstance(v, (int, float)):
                    cell.number_format = '0.0000%'
            else:
                cell = ws.cell(r, j, v if v is not None else '')
            cell.font = _XL_DATA
        r += 1

    # 합계 행
    if totals_row and rows:
        for j in range(1, len(col_defs) + 1):
            ws.cell(r, j).fill = _XL_TOTAL_FILL
        for j, (_lbl, _w, key, fmt) in enumerate(col_defs, 1):
            v = totals_row.get(key)
            if v is None and j == 1:
                v = '전체 합계'
            if v is None:
                continue
            cell = ws.cell(r, j, v)
            cell.font = _XL_TOTAL_FONT
            if fmt == 'num':
                cell.number_format = _XL_NUM_FMT
            elif fmt == 'int':
                cell.number_format = '#,##0'

    return ws


def _aio_cover_sheet(wb, year, sections_meta):
    """전체 표지 시트 — 합산 항목 인덱스 + KRW 합계.
    sections_meta: [(sheet_name, title, krw_label, krw_value)] 또는
                   [(sheet_name, title, summary_text)]
    """
    ws = wb.create_sheet(_safe_sheet_name('전체 표지'), 0)
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 24

    ws.cell(1, 1, '주석합산 전체 다운로드').font = _XL_TITLE
    ws.cell(3, 1, '결산기간').font = _XL_DATA
    ws.cell(3, 2, year).font = _XL_DATA
    ws.cell(4, 1, '합산 항목 수').font = _XL_DATA
    ws.cell(4, 2, len(sections_meta)).font = _XL_DATA

    headers = ['시트', '항목', '주요 KRW 합계 / 요약']
    for j, h in enumerate(headers, 1):
        c = ws.cell(6, j, h); c.font = _XL_HDR_FONT; c.fill = _XL_HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')

    r = 7
    for meta in sections_meta:
        sheet_name, title = meta[0], meta[1]
        ws.cell(r, 1, sheet_name).font = _XL_DATA
        ws.cell(r, 2, title).font = _XL_DATA
        if len(meta) >= 4:
            krw_label, krw_value = meta[2], meta[3]
            text_cell = ws.cell(r, 3, krw_value if krw_value is not None else '')
            text_cell.font = _XL_DATA
            if isinstance(krw_value, (int, float)):
                text_cell.number_format = _XL_NUM_FMT
            if krw_label:
                ws.cell(r, 3).comment = None  # placeholder
        elif len(meta) == 3:
            ws.cell(r, 3, meta[2]).font = _XL_DATA
        r += 1


def _aio_write_table_multi(wb, sheet_name, title, agg, col_defs, total_keys=None):
    """다행 명세 패턴 (회사+컬럼들+KRW환산) generic.
      agg: {'rows':[..], 'scanned','with_data_count' or 'with_rows', ...}
      total_keys: [(key, label, fmt)] — 합계 행에 표시할 키들 (모두 num)
    """
    rows = agg.get('rows') or []
    info_pairs = [
        ('결산기간', agg.get('year'), 'text'),
        ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
        ('입력 회사 수', agg.get('with_data_count') or agg.get('with_rows') or 0, 'int'),
    ]
    # 추가 grand totals
    if total_keys:
        for key, label, _f in total_keys:
            info_pairs.append((label, agg.get(key) or 0, 'num'))

    totals_row = None
    if total_keys and rows:
        totals_row = {col_defs[0][2]: '전체 합계'}
        for key, _l, _f in total_keys:
            # col_defs에서 해당 합계 키와 매칭되는 컬럼 찾기 (key 끝부분)
            # 예: 'grand_total_krw' → col_defs key='krw'
            pass
        # 단순화: total_keys의 값을 마지막 컬럼에 표시. 컬럼 매칭은 호출자 책임.

    _aio_write_section(wb, sheet_name, title, info_pairs, col_defs, rows,
                       totals_row=None)


def build_all_in_one_excel(year, all_aggs, output_path):
    """모든 주석합산 결과를 단일 워크북에 시트별로 작성.

    all_aggs: dict — key별 agg_data
      keys: 'l1', 'l4_loan_facility', 'l4_lc', 'l4_export',
            'l4_guarantees_received', 'l4_guarantees_provided',
            'l4_restricted_financial', 'l4_insured_ppe',
            'l4_pledged_proceeds', 'l4_pledged_assets',
            'l4_lawsuits', 'l4_subsequent_events', 'l4_other_commitments',
            'a2_securities',
            'a3_investment_pl', 'a3_land_investment', 'a3_land_ppe',
            'a4_construction_balance', 'a4_construction_profit', 'a4_contract_balance',
            'a5_rou_changes', 'a5_lease_pl',
            'a6_derivatives',
            'a7_equity_method'
    """
    wb = Workbook()
    wb.remove(wb.active)

    sections_meta = []   # 표지용

    # ─── L1 단기차입금 ────────────────────────────────────────
    agg = all_aggs.get('l1') or {}
    if agg:
        sheet = 'L1_단기차입금'
        rows = []
        for cat in agg.get('categories') or []:
            for r in cat.get('rows') or []:
                rows.append({
                    'category':      f"{cat.get('key')}. {cat.get('name')}",
                    'company':       r.get('company') or '',
                    'creditor_type': r.get('creditor_type') or '',
                    'creditor':      r.get('creditor') or '',
                    'rate':          r.get('rate'),
                    'currency':      r.get('currency') or '',
                    'local':         r.get('local') or 0,
                    'spot':          r.get('spot') or 0,
                    'krw':           r.get('krw') or 0,
                })
        col_defs = [
            ('종류 (계정)', 30, 'category', 'text'),
            ('회사', 26, 'company', 'text'),
            ('대주구분', 18, 'creditor_type', 'text'),
            ('대주명', 26, 'creditor', 'text'),
            ('이자율', 10, 'rate', 'rate'),
            ('통화', 8, 'currency', 'text'),
            ('로컬 금액', 18, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
        ]
        _aio_write_section(wb, sheet, 'L1 단기차입금 명세 합산',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('종류 수', len(agg.get('categories') or []), 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, rows)
        sections_meta.append((sheet, '단기차입금', 'KRW 합계', agg.get('grand_total_krw') or 0))

    # ─── L2 1. 장기차입금 ─────────────────────────────────────
    def _add_l2_balance_sheet(key, sheet_name, title, lbl1, lbl2, short_name):
        agg = all_aggs.get(key) or {}
        if not agg:
            return
        col_defs = [
            ('회사', 28, 'company', 'text'),
            (lbl1, 22, 'type1', 'text'),
            (lbl2, 28, 'type2', 'text'),
            ('이자율', 10, 'rate', 'rate'),
            ('통화', 8, 'currency', 'text'),
            ('유동(로컬)', 18, 'local_current', 'num'),
            ('비유동(로컬)', 18, 'local_non_current', 'num'),
            ('합계(로컬)', 18, 'local_total', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('유동(KRW)', 18, 'krw_current', 'num'),
            ('비유동(KRW)', 18, 'krw_non_current', 'num'),
            ('합계(KRW)', 18, 'krw_total', 'num'),
        ]
        _aio_write_section(wb, sheet_name, title,
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('유동 KRW 합계', agg.get('grand_current_krw') or 0, 'num'),
             ('비유동 KRW 합계', agg.get('grand_non_current_krw') or 0, 'num'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet_name, short_name, 'KRW 합계',
                              agg.get('grand_total_krw') or 0))

    _add_l2_balance_sheet('l2_long_term_borrowings', 'L2_1_장기차입금',
                          'L2 1. 장기차입금 (유동성 포함)',
                          '대주구분', '대주명', '장기차입금')
    _add_l2_balance_sheet('l2_debentures', 'L2_2_사채',
                          'L2 2. 사채 (유동성 사채 포함)',
                          '종류', '주관사', '사채')

    # ─── L3 1. 퇴직급여충당부채 변동 ──────────────────────────
    def _add_l3_pivot_sheet(key, sheet_name, title, item_keys, item_lbls):
        agg2 = all_aggs.get(key) or {}
        if not agg2:
            return
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for ik, lbl in zip(item_keys, item_lbls):
            col_defs.append((f'{lbl}(로컬)', 18, f'local_{ik}', 'num'))
        col_defs.append(('Spot', 10, 'spot', 'rate'))
        for ik, lbl in zip(item_keys, item_lbls):
            col_defs.append((f'{lbl}(KRW)', 18, f'krw_{ik}', 'num'))
        totals = agg2.get('totals_by_item_krw') or {}
        _aio_write_section(wb, sheet_name, title,
            [('결산기간', agg2.get('year'), 'text'),
             ('스캔 회사 수', agg2.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg2.get('with_data_count') or 0, 'int'),
             ('기말금액 KRW 합계', totals.get('ending') or 0, 'num')],
            col_defs, agg2.get('rows') or [])
        short = title.split('. ', 1)[-1] if '. ' in title else title
        sections_meta.append((sheet_name, short, '기말 KRW',
                              totals.get('ending') or 0))

    # L3 1번/2번 — 항목별 환율(전기말 spot / avg / 당기말 spot) + 환율변동효과
    def _add_l3_fx_sheet(key, sheet_name, title, short_name,
                         var_keys, var_labels):
        agg2 = all_aggs.get(key) or {}
        if not agg2:
            return
        item_keys = ['beginning'] + list(var_keys) + ['ending']
        item_lbls = ['기초금액'] + list(var_labels) + ['기말금액']
        cols = [
            ('회사', 28, 'company', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for k, lbl in zip(item_keys, item_lbls):
            cols.append((f'{lbl}(로컬)', 16, f'local_{k}', 'num'))
        cols.append(('전기말 Spot', 12, 'prior_spot', 'rate'))
        cols.append(('당기 Avg', 11, 'avg_rate', 'rate'))
        cols.append(('당기말 Spot', 12, 'spot', 'rate'))
        cols.append(('기초(KRW)', 18, 'krw_beginning', 'num'))
        for k, lbl in zip(var_keys, var_labels):
            cols.append((f'{lbl}(KRW)', 18, f'krw_{k}', 'num'))
        cols.append(('환율변동효과(KRW)', 18, 'krw_fx_effect', 'num'))
        cols.append(('기말(KRW)', 18, 'krw_ending', 'num'))
        totals = agg2.get('totals_by_item_krw') or {}
        var_sum = sum((totals.get(k) or 0) for k in var_keys)
        _aio_write_section(wb, sheet_name,
            f'{title} (기초=전기말Spot / 변동=Avg / 기말=당기말Spot)',
            [('결산기간', agg2.get('year'), 'text'),
             ('스캔 회사 수', agg2.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg2.get('with_data_count') or 0, 'int'),
             ('기초 KRW 합계', totals.get('beginning') or 0, 'num'),
             ('변동 KRW 합계', var_sum, 'num'),
             ('환율변동효과 KRW', totals.get('fx_effect') or 0, 'num'),
             ('기말 KRW 합계', totals.get('ending') or 0, 'num')],
            cols, agg2.get('rows') or [])
        sections_meta.append((sheet_name, short_name, '기말 KRW',
                              totals.get('ending') or 0))

    _add_l3_fx_sheet('l3_severance', 'L3_1_퇴직급여충당부채',
        'L3 1. 퇴직급여충당부채의 변동', '퇴직급여충당부채 변동',
        ['provision', 'payment', 'transfer', 'business_combination', 'others'],
        ['설정액', '지급액', '전출입', '사업결합', '기타증감'])
    _add_l3_fx_sheet('l3_pension_movement', 'L3_2_퇴직연금자산변동',
        'L3 2. 퇴직연금운용자산의 변동', '퇴직연금자산 변동',
        ['contribution', 'payment', 'interest_income',
         'transfer', 'business_combination', 'others'],
        ['적립액', '지급액', '이자수익', '전출입', '사업결합', '기타증감'])

    # ─── L3 3. 퇴직연금운용자산 구성 ──────────────────────────
    agg = all_aggs.get('l3_pension_breakdown') or {}
    if agg:
        sheet = 'L3_3_퇴직연금구성'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for k, lbl in zip(_L3_S3_KEYS, _L3_S3_LBLS):
            col_defs.append((f'{lbl}(로컬)', 18, f'local_{k}', 'num'))
        col_defs.append(('Spot', 10, 'spot', 'rate'))
        for k, lbl in zip(_L3_S3_KEYS, _L3_S3_LBLS):
            col_defs.append((f'{lbl}(KRW)', 18, f'krw_{k}', 'num'))
        _aio_write_section(wb, sheet, 'L3 3. 퇴직연금운용자산의 구성내역',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '퇴직연금자산 구성', 'KRW 합계',
                              agg.get('grand_total_krw') or 0))

    # ─── L3 4. 퇴직연금운용자산 운용사 ────────────────────────
    agg = all_aggs.get('l3_pension_managers') or {}
    if agg:
        sheet = 'L3_4_퇴직연금운용사'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('운용사', 22, 'name', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('로컬 금액', 20, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
            ('비고', 36, 'remarks', 'text'),
        ]
        _aio_write_section(wb, sheet, 'L3 4. 퇴직연금운용자산의 운용사',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '퇴직연금 운용사', 'KRW 합계',
                              agg.get('grand_total_krw') or 0))

    # ─── L3-1 1번 확정급여채무 변동, 2번 사외적립자산 변동 ────
    _add_l3_pivot_sheet('l31_dbo', 'L3-1_1_확정급여채무',
        'L3-1 1. 확정급여채무의 변동',
        _L31_S1_KEYS, _L31_S1_LBLS)
    _add_l3_pivot_sheet('l31_plan_asset', 'L3-1_2_사외적립자산변동',
        'L3-1 2. 사외적립자산의 공정가치 변동',
        _L31_S2_KEYS, _L31_S2_LBLS)

    # ─── L3-1 3. 보험수리 가정치 (min/max) ────────────────────
    agg = all_aggs.get('l31_assumptions') or {}
    if agg:
        sheet = 'L3-1_3_보험수리가정치'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('기대임금상승률 (원문)', 22, 'wage_growth_raw', 'text'),
            ('기대임금상승률 (값)', 14, 'wage_growth_value', 'pct'),
            ('할인율 (원문)', 22, 'discount_rate_raw', 'text'),
            ('할인율 (값)', 14, 'discount_rate_value', 'pct'),
        ]
        ex = agg.get('extremes') or {}
        wg_min = (ex.get('wage_growth') or {}).get('min') or {}
        wg_max = (ex.get('wage_growth') or {}).get('max') or {}
        dr_min = (ex.get('discount_rate') or {}).get('min') or {}
        dr_max = (ex.get('discount_rate') or {}).get('max') or {}
        _aio_write_section(wb, sheet,
            'L3-1 3. 보험수리적 평가 가정치 (최저/최고 회사)',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('기대임금상승률 최저',
              f"{wg_min.get('company') or '-'} ({wg_min.get('raw') or '-'})", 'text'),
             ('기대임금상승률 최고',
              f"{wg_max.get('company') or '-'} ({wg_max.get('raw') or '-'})", 'text'),
             ('할인율 최저',
              f"{dr_min.get('company') or '-'} ({dr_min.get('raw') or '-'})", 'text'),
             ('할인율 최고',
              f"{dr_max.get('company') or '-'} ({dr_max.get('raw') or '-'})", 'text')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '보험수리 가정치', '입력 회사',
                              agg.get('with_data_count') or 0))

    # ─── L3-1 4. 민감도 분석 ──────────────────────────────────
    agg = all_aggs.get('l31_sensitivity') or {}
    if agg:
        sheet = 'L3-1_4_민감도'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('기대임금상승율 1%↑(로컬)', 22, 'local_wage_growth_up', 'num'),
            ('기대임금상승율 1%↓(로컬)', 22, 'local_wage_growth_down', 'num'),
            ('할인율 1%↑(로컬)', 18, 'local_discount_rate_up', 'num'),
            ('할인율 1%↓(로컬)', 18, 'local_discount_rate_down', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('기대임금상승율 1%↑(KRW)', 22, 'krw_wage_growth_up', 'num'),
            ('기대임금상승율 1%↓(KRW)', 22, 'krw_wage_growth_down', 'num'),
            ('할인율 1%↑(KRW)', 18, 'krw_discount_rate_up', 'num'),
            ('할인율 1%↓(KRW)', 18, 'krw_discount_rate_down', 'num'),
        ]
        totals = agg.get('totals_by_item_krw') or {}
        _aio_write_section(wb, sheet, 'L3-1 4. 보험수리적 가정의 변동 영향',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('기대임금상승율 1%↑ KRW', totals.get('wage_growth_up') or 0, 'num'),
             ('기대임금상승율 1%↓ KRW', totals.get('wage_growth_down') or 0, 'num'),
             ('할인율 1%↑ KRW', totals.get('discount_rate_up') or 0, 'num'),
             ('할인율 1%↓ KRW', totals.get('discount_rate_down') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '민감도 분석', '입력 회사',
                              agg.get('with_data_count') or 0))

    # ─── L3-1 5. 사외적립자산 구성내역 ────────────────────────
    agg = all_aggs.get('l31_plan_breakdown') or {}
    if agg:
        sheet = 'L3-1_5_사외적립구성'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for k, lbl in zip(_L3_S3_KEYS, _L3_S3_LBLS):
            col_defs.append((f'{lbl}(로컬)', 18, f'local_{k}', 'num'))
        col_defs.append(('Spot', 10, 'spot', 'rate'))
        for k, lbl in zip(_L3_S3_KEYS, _L3_S3_LBLS):
            col_defs.append((f'{lbl}(KRW)', 18, f'krw_{k}', 'num'))
        _aio_write_section(wb, sheet, 'L3-1 5. 사외적립자산의 구성내역',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '사외적립자산 구성', 'KRW 합계',
                              agg.get('grand_total_krw') or 0))

    # ─── L3-1 6. 사외적립자산 운용사 ──────────────────────────
    agg = all_aggs.get('l31_plan_managers') or {}
    if agg:
        sheet = 'L3-1_6_사외적립운용사'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('운용사', 22, 'name', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('로컬 금액', 20, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
            ('비고', 36, 'remarks', 'text'),
        ]
        _aio_write_section(wb, sheet, 'L3-1 6. 사외적립자산의 운용사',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '사외적립자산 운용사', 'KRW 합계',
                              agg.get('grand_total_krw') or 0))

    # ─── TX 시트 — 법인세 합산 (1 / 3 / 3-1 / 4 / 5 / 5-1) ─────
    def _add_tx_pivot_sheet(key, sheet_name, title, item_keys, item_lbls,
                            rate_label='Avg'):
        """TX 3 / 3-1 / 4 / 5 공통 pivot 시트 작성. ws 반환."""
        agg_tx = all_aggs.get(key) or {}
        if not agg_tx:
            return None
        cols = [
            ('회사', 26, 'company', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for ik, lbl in zip(item_keys, item_lbls):
            cols.append((f'{lbl}(로컬)', 18, f'local_{ik}', 'num'))
        cols.append((rate_label, 10, 'spot', 'rate'))
        for ik, lbl in zip(item_keys, item_lbls):
            cols.append((f'{lbl}(KRW)', 18, f'krw_{ik}', 'num'))
        totals = agg_tx.get('totals_by_item_krw') or {}
        info = [('결산기간', agg_tx.get('year'), 'text'),
                ('스캔 회사 수', agg_tx.get('scanned') or 0, 'int'),
                ('입력 회사 수', agg_tx.get('with_data_count') or 0, 'int')]
        for ik, lbl in zip(item_keys, item_lbls):
            info.append((f'{lbl} KRW 합계', totals.get(ik) or 0, 'num'))
        return _aio_write_section(wb, sheet_name, title, info,
                                  cols, agg_tx.get('rows') or []), cols, info

    # TX 1. 이연법인세자산(부채) 증감내용 — categories + company_rows
    agg = all_aggs.get('tx_deferred_tax_changes') or {}
    if agg:
        sheet = 'TX_1_이연법인세_증감'
        col_defs = [
            ('회사', 26, 'company', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('계정명', 30, 'kor_label', 'text'),
            ('유동/비유동', 14, 'current_flag', 'text'),
            ('기초(로컬)', 16, 'local_beginning', 'num'),
            ('기말(로컬)', 16, 'local_ending', 'num'),
            ('증감(로컬)', 16, 'local_change', 'num'),
            ('전기말 Spot', 12, 'prior_spot', 'rate'),
            ('당기말 Spot', 12, 'spot', 'rate'),
            ('기초(KRW)', 18, 'krw_beginning', 'num'),
            ('기말(KRW)', 18, 'krw_ending', 'num'),
            ('증감(KRW)', 18, 'krw_change', 'num'),
            ('법정세율', 12, 'statutory_rate', 'pct'),
            ('기초 일시적차이', 18, 'beginning_temp_diff', 'numn'),
            ('기말 일시적차이', 18, 'ending_temp_diff', 'numn'),
        ]
        totals = agg.get('totals_krw') or {}
        _aio_write_section(wb, sheet,
            'TX 1. 이연법인세자산(부채) 증감내용 합산 '
            '(기초=전기말 Spot / 기말=당기말 Spot)',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('기초 KRW 합계', totals.get('beginning') or 0, 'num'),
             ('기말 KRW 합계', totals.get('ending') or 0, 'num'),
             ('증감 KRW 합계', totals.get('change') or 0, 'num')],
            col_defs, agg.get('company_rows') or [])
        sections_meta.append((sheet, '이연법인세 증감', '기말 KRW',
                              totals.get('ending') or 0))

    # TX 3. 법인세비용의 구성내역
    res = _add_tx_pivot_sheet('tx_income_tax_breakdown', 'TX_3_법인세비용_구성',
        'TX 3. 법인세비용의 구성내역 합산 (Avg 환율)',
        ['current_tax', 'deferred_temp_diff', 'deferred_equity',
         'additional_refund', 'total_expense'],
        ['당기법인세부담액', '일시적차이 이연법인세 변동액',
         '자본 직접부과 이연법인세 변동액',
         '법인세 추납액(환급액)', '법인세비용 총계'])
    if res is not None:
        agg = all_aggs.get('tx_income_tax_breakdown') or {}
        totals = agg.get('totals_by_item_krw') or {}
        sections_meta.append(('TX_3_법인세비용_구성', '법인세비용 구성',
                              '법인세비용 KRW', totals.get('total_expense') or 0))
        # TX 3 — 연결그룹별 KRW 합계 (별도 시트)
        gsubs = agg.get('group_subtotals') or []
        if gsubs:
            _tx3_keys = ['current_tax', 'deferred_temp_diff', 'deferred_equity',
                         'additional_refund', 'total_expense']
            _tx3_lbls = ['당기법인세부담액', '일시적차이 이연법인세 변동액',
                         '자본 직접부과 이연법인세 변동액',
                         '법인세 추납액(환급액)', '법인세비용 총계']
            gcols = [('연결그룹', 24, 'group', 'text'),
                     ('회사수', 10, 'company_count', 'int')]
            for ik, lbl in zip(_tx3_keys, _tx3_lbls):
                gcols.append((f'{lbl}(KRW)', 18, f'krw_{ik}', 'num'))
            grows = []
            for s in gsubs:
                row = {'group': s.get('group') or '',
                       'company_count': s.get('company_count') or 0}
                krw = s.get('krw') or {}
                for ik in _tx3_keys:
                    row[f'krw_{ik}'] = krw.get(ik) or 0
                grows.append(row)
            _aio_write_section(wb, 'TX_3_연결그룹별',
                'TX 3. 법인세비용의 구성내역 — 연결그룹별 KRW 합계',
                [('결산기간', agg.get('year'), 'text'),
                 ('연결그룹 수', len(gsubs), 'int')],
                gcols, grows)
            sections_meta.append(('TX_3_연결그룹별',
                                  '법인세비용 구성 — 연결그룹별',
                                  '법인세비용 KRW', totals.get('total_expense') or 0))

    # TX 3-1. 자본 직접 부과 이연법인세 변동액
    res = _add_tx_pivot_sheet('tx_equity_deferred_tax', 'TX_3_1_자본_직접부과_이연',
        'TX 3-1. 자본 직접 부과 이연법인세 변동액 명세 합산 (Avg 환율)',
        ['revaluation', 'actuarial', 'afs_securities', 'fvoci',
         'equity_method', 'fx_translation', 'derivatives',
         'others', 'total'],
        ['재평가이익', '보험수리적손익', '매도가능증권평가손익',
         'FVOCI 금융상품', '지분법적용투자주식',
         '해외사업환산손익', '파생상품평가손익', '기타', '합계'])
    if res is not None:
        agg = all_aggs.get('tx_equity_deferred_tax') or {}
        totals = agg.get('totals_by_item_krw') or {}
        sections_meta.append(('TX_3_1_자본_직접부과_이연',
                              '자본직접부과 이연법인세',
                              '합계 KRW', totals.get('total') or 0))

    # TX 4. Reconciliation — pivot + 유효세율(수식: 법인세비용 ÷ 법인세차감전순이익)
    agg = all_aggs.get('tx_reconciliation') or {}
    if agg:
        sheet = 'TX_4_Reconciliation'
        tx4_item_keys = ['pretax_income', 'tax_at_statutory', 'permanent_diff',
                         'tax_credit', 'additional_refund',
                         'unrecognized_change', 'total_expense']
        tx4_item_lbls = ['법인세차감전 순이익', '적용세율 산출세액',
                         '영구적차이', '세액공제', '법인세 추납액/환급액',
                         '미인식 일시적차이 변동', '법인세비용']
        col_defs = [
            ('회사', 26, 'company', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for ik, lbl in zip(tx4_item_keys, tx4_item_lbls):
            col_defs.append((f'{lbl}(로컬)', 18, f'local_{ik}', 'num'))
        col_defs.append(('Avg', 10, 'spot', 'rate'))
        for ik, lbl in zip(tx4_item_keys, tx4_item_lbls):
            col_defs.append((f'{lbl}(KRW)', 18, f'krw_{ik}', 'num'))
        col_defs.append(('유효세율', 14, '_eff_rate_formula', 'text'))
        totals = agg.get('totals_by_item_krw') or {}
        info = [('결산기간', agg.get('year'), 'text'),
                ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
                ('입력 회사 수', agg.get('with_data_count') or 0, 'int')]
        for ik, lbl in zip(tx4_item_keys, tx4_item_lbls):
            info.append((f'{lbl} KRW 합계', totals.get(ik) or 0, 'num'))
        info.append(('* 유효세율 = 법인세비용 ÷ 법인세차감전순이익 (셀 수식)',
                     '', 'text'))
        rows_tx4 = agg.get('rows') or []
        ws_tx4 = _aio_write_section(wb, sheet,
            'TX 4. 법인세 reconciliation 합산 (Avg 환율)',
            info, col_defs, rows_tx4)
        # 유효세율 셀에 수식 후처리 — 데이터 시작 행 = 헤더행 + 1
        if ws_tx4 is not None and rows_tx4:
            header_r = 3 + len(info) + 1   # info 끝 다음 r+=1, header_r
            data_start_r = header_r + 1
            keys = [k for (_lbl, _w, k, _f) in col_defs]
            col_pretax = keys.index('krw_pretax_income') + 1
            col_total  = keys.index('krw_total_expense') + 1
            col_eff    = len(col_defs)
            for i in range(len(rows_tx4)):
                rr = data_start_r + i
                pretax_addr = f'{get_column_letter(col_pretax)}{rr}'
                total_addr  = f'{get_column_letter(col_total)}{rr}'
                cell = ws_tx4.cell(rr, col_eff,
                    f'=IFERROR({total_addr}/{pretax_addr},"")')
                cell.number_format = '0.0000%'
                cell.font = _XL_DATA
        sections_meta.append((sheet, '법인세 Reconciliation',
                              '법인세비용 KRW',
                              totals.get('total_expense') or 0))
        # TX 4 — 연결그룹별 KRW 합계 (별도 시트, 그룹 유효세율 포함)
        gsubs = agg.get('group_subtotals') or []
        if gsubs:
            gcols = [('연결그룹', 24, 'group', 'text'),
                     ('회사수', 10, 'company_count', 'int')]
            for ik, lbl in zip(tx4_item_keys, tx4_item_lbls):
                gcols.append((f'{lbl}(KRW)', 18, f'krw_{ik}', 'num'))
            gcols.append(('유효세율', 12, 'eff_rate', 'pct'))
            grows = []
            for s in gsubs:
                krw = s.get('krw') or {}
                pre = krw.get('pretax_income') or 0
                tot = krw.get('total_expense') or 0
                row = {'group': s.get('group') or '',
                       'company_count': s.get('company_count') or 0,
                       'eff_rate': (tot / pre) if pre else None}
                for ik in tx4_item_keys:
                    row[f'krw_{ik}'] = krw.get(ik) or 0
                grows.append(row)
            _aio_write_section(wb, 'TX_4_연결그룹별',
                'TX 4. 법인세 reconciliation — 연결그룹별 KRW 합계',
                [('결산기간', agg.get('year'), 'text'),
                 ('연결그룹 수', len(gsubs), 'int'),
                 ('* 유효세율 = 법인세비용 ÷ 법인세차감전순이익', '', 'text')],
                gcols, grows)
            sections_meta.append(('TX_4_연결그룹별',
                                  '법인세 Reconciliation — 연결그룹별',
                                  '법인세비용 KRW', totals.get('total_expense') or 0))

    # TX 5. 미인식 일시적차이
    res = _add_tx_pivot_sheet('tx_unrecognized', 'TX_5_미인식_일시적차이',
        'TX 5. 미인식 일시적차이 합산 (Spot 환율)',
        ['loss_carryforward', 'others'],
        ['이월결손금', '기타'],
        rate_label='Spot')
    if res is not None:
        agg = all_aggs.get('tx_unrecognized') or {}
        totals = agg.get('totals_by_item_krw') or {}
        sections_meta.append(('TX_5_미인식_일시적차이', '미인식 일시적차이',
                              'KRW 합계',
                              (totals.get('loss_carryforward') or 0)
                              + (totals.get('others') or 0)))

    # TX 5-1. 이월결손금 만기 — buckets 동적 컬럼
    agg = all_aggs.get('tx_loss_maturity') or {}
    if agg:
        sheet = 'TX_5_1_이월결손금_만기'
        buckets = agg.get('buckets') or []
        bucket_labels = [b.get('label') or '' for b in buckets]
        flat_rows = []
        for co in (agg.get('company_rows') or []):
            row = {'company': co.get('company') or '',
                   'currency': co.get('currency') or '',
                   'spot': co.get('spot') or 0}
            label_to_b = {b.get('label'): b for b in (co.get('buckets') or [])}
            for lbl in bucket_labels:
                b = label_to_b.get(lbl) or {}
                row[f'local__{lbl}'] = b.get('local') or 0
                row[f'krw__{lbl}']   = b.get('krw') or 0
            flat_rows.append(row)
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for lbl in bucket_labels:
            col_defs.append((f'{lbl}(로컬)', 16, f'local__{lbl}', 'num'))
        col_defs.append(('Spot', 10, 'spot', 'rate'))
        for lbl in bucket_labels:
            col_defs.append((f'{lbl}(KRW)', 16, f'krw__{lbl}', 'num'))
        info = [('결산기간', agg.get('year'), 'text'),
                ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
                ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
                ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')]
        for b in buckets:
            info.append((f"{b.get('label') or ''} KRW",
                         b.get('total_krw') or 0, 'num'))
        _aio_write_section(wb, sheet,
            'TX 5-1. 이월결손금 만기 합산 (Spot 환율)',
            info, col_defs, flat_rows)
        sections_meta.append((sheet, '이월결손금 만기',
                              'KRW 합계', agg.get('grand_total_krw') or 0))

    # ─── L2 3. 부채성 금융상품 만기 분석 ──────────────────────
    agg = all_aggs.get('l2_maturity') or {}
    if agg:
        sheet = 'L2_3_만기분석'
        interval_keys = ['within_1y', 'within_2y', 'within_5y', 'over_5y', 'total']
        interval_lbls = ['1년 미만', '1~2년', '2~5년', '5년 초과', '합계']
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('계정명', 18, 'account', 'text'),
            ('대주구분', 18, 'creditor_type', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for k, lbl in zip(interval_keys, interval_lbls):
            col_defs.append((f'{lbl}(로컬)', 18, f'local_{k}', 'num'))
        col_defs.append(('Spot', 10, 'spot', 'rate'))
        for k, lbl in zip(interval_keys, interval_lbls):
            col_defs.append((f'{lbl}(KRW)', 18, f'krw_{k}', 'num'))
        _aio_write_section(wb, sheet, 'L2 3. 부채성 금융상품 만기 분석',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('1년 미만 KRW', agg.get('krw_within_1y') or 0, 'num'),
             ('1~2년 KRW', agg.get('krw_within_2y') or 0, 'num'),
             ('2~5년 KRW', agg.get('krw_within_5y') or 0, 'num'),
             ('5년 초과 KRW', agg.get('krw_over_5y') or 0, 'num'),
             ('전체 KRW 합계', agg.get('krw_total') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '만기 분석', 'KRW 합계',
                              agg.get('krw_total') or 0))

    # ─── L4 1-1 대출한도 약정 ─────────────────────────────────
    agg = all_aggs.get('l4_loan_facility') or {}
    if agg:
        sheet = 'L4_1-1_대출한도'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('종류', 24, 'type', 'text'),
            ('금융기관', 28, 'institution', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('로컬 금액', 18, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
        ]
        _aio_write_section(wb, sheet, 'L4 1-1. 대출한도 약정',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_rows') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '대출한도 약정', 'KRW', agg.get('grand_total_krw') or 0))

    # ─── L4 2 수입신용장 / 미확정 지급보증 ────────────────────
    agg = all_aggs.get('l4_lc') or {}
    if agg:
        sheet = 'L4_2_수입신용장'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('L/C 오픈', 12, 'lc_open', 'text'),
            ('미확정 지급보증', 22, 'amount', 'num'),
            ('통화', 8, 'currency', 'text'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
        ]
        _aio_write_section(wb, sheet, 'L4 2. 수입신용장 / 미확정 지급보증',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('YES 회사 수', agg.get('yes_count') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '수입신용장 / 미확정 지급보증',
                              'KRW', agg.get('grand_total_krw') or 0))

    # ─── L4 3 수출채권 할인 ───────────────────────────────────
    agg = all_aggs.get('l4_export') or {}
    if agg:
        sheet = 'L4_3_수출채권할인'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('할인 여부', 12, 'discount_done', 'text'),
            ('만기 미도래', 22, 'amount', 'num'),
            ('통화', 8, 'currency', 'text'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
        ]
        _aio_write_section(wb, sheet, 'L4 3. 수출채권 할인',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('YES 회사 수', agg.get('yes_count') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '수출채권 할인', 'KRW', agg.get('grand_total_krw') or 0))

    # ─── L4 4-1 받은 보증 ────────────────────────────────────
    agg = all_aggs.get('l4_guarantees_received') or {}
    if agg:
        sheet = 'L4_4-1_받은보증'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('제공자', 26, 'guarantor', 'text'),
            ('보증종류', 22, 'type', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('로컬 금액', 18, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
            ('관련계정', 16, 'account', 'text'),
            ('Description', 36, 'description', 'text'),
        ]
        _aio_write_section(wb, sheet, 'L4 4-1. 제공받은 보증 내용',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_rows') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '제공받은 보증', 'KRW', agg.get('grand_total_krw') or 0))

    # ─── L4 5-2 제공한 보증 ──────────────────────────────────
    agg = all_aggs.get('l4_guarantees_provided') or {}
    if agg:
        sheet = 'L4_5-2_제공한보증'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('제공받는자', 30, 'beneficiary', 'text'),
            ('보증종류', 22, 'type', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('로컬 금액', 18, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
            ('채권자', 22, 'guaranteed_creditor', 'text'),
            ('Description', 36, 'description', 'text'),
        ]
        _aio_write_section(wb, sheet, 'L4 5-2. 제공한 보증 내용',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_rows') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '제공한 보증', 'KRW', agg.get('grand_total_krw') or 0))

    # ─── L4 6-1 소송 ────────────────────────────────────────
    agg = all_aggs.get('l4_lawsuits') or {}
    if agg:
        sheet = 'L4_6-1_소송'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('구분', 16, 'type', 'text'),
            ('소송건수', 12, 'count', 'int'),
            ('통화', 8, 'currency', 'text'),
            ('소송금액(로컬)', 22, 'claim_amount', 'num'),
            ('충당부채(로컬)', 22, 'provision_amount', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('소송금액(KRW)', 22, 'claim_amount_krw', 'num'),
            ('충당부채(KRW)', 22, 'provision_amount_krw', 'num'),
        ]
        _aio_write_section(wb, sheet, 'L4 6-1. 소송중인 사건',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('소송 입력 회사', agg.get('with_rows') or 0, 'int'),
             ('원고 건수', agg.get('defendant_count') or 0, 'int'),
             ('피고 건수', agg.get('plaintiff_count') or 0, 'int'),
             ('총 소송금액 KRW', agg.get('total_claim_krw') or 0, 'num'),
             ('총 충당부채 KRW', agg.get('total_provision_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '소송중인 사건', '충당부채 KRW',
                              agg.get('total_provision_krw') or 0))

    # ─── L4 7-1 사용제한 금융상품 ────────────────────────────
    agg = all_aggs.get('l4_restricted_financial') or {}
    if agg:
        sheet = 'L4_7-1_사용제한금융'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('계정과목', 22, 'account', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('로컬 금액', 18, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
            ('제한내용', 40, 'description', 'text'),
        ]
        _aio_write_section(wb, sheet, 'L4 7-1. 사용제한 금융상품',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_rows') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '사용제한 금융상품', 'KRW',
                              agg.get('grand_total_krw') or 0))

    # ─── L4 8-1 보험가입 유형자산 ────────────────────────────
    agg = all_aggs.get('l4_insured_ppe') or {}
    if agg:
        sheet = 'L4_8-1_보험가입자산'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('자산종류', 22, 'asset_type', 'text'),
            ('보험사', 22, 'insurer', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('부보금액(로컬)', 20, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
            ('Description', 36, 'description', 'text'),
        ]
        _aio_write_section(wb, sheet, 'L4 8-1. 보험가입 유형자산',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_rows') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '보험가입 유형자산', 'KRW',
                              agg.get('grand_total_krw') or 0))

    # ─── L4 8-2 보험수익금 질권설정 ─────────────────────────
    agg = all_aggs.get('l4_pledged_proceeds') or {}
    if agg:
        sheet = 'L4_8-2_보험질권'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('질권자', 22, 'pledgee', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('질권금액(로컬)', 20, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
            ('Description', 40, 'description', 'text'),
        ]
        _aio_write_section(wb, sheet, 'L4 8-2. 보험수익금 질권설정',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_rows') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '보험수익금 질권설정', 'KRW',
                              agg.get('grand_total_krw') or 0))

    # ─── L4 9-1 담보제공자산 ─────────────────────────────────
    agg = all_aggs.get('l4_pledged_assets') or {}
    if agg:
        sheet = 'L4_9-1_담보자산'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('채권자/저당권자', 22, 'creditor', 'text'),
            ('담보자산', 22, 'asset_account', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('담보금액(로컬)', 20, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('KRW 환산', 20, 'krw', 'num'),
            ('관련부채', 18, 'liability_account', 'text'),
            ('Description', 36, 'description', 'text'),
        ]
        _aio_write_section(wb, sheet, 'L4 9-1. 담보제공자산',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_rows') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '담보제공자산', 'KRW',
                              agg.get('grand_total_krw') or 0))

    # ─── L4 10 보고기간일 이후 사건 (YES 회사만) ──────────────
    agg = all_aggs.get('l4_subsequent_events') or {}
    if agg:
        sheet = 'L4_10_보고후사건'
        col_defs = [
            ('회사', 32, 'company', 'text'),
            ('응답', 10, 'yn', 'text'),
            ('사건 내용', 100, 'content', 'text'),
        ]
        _aio_write_section(wb, sheet, 'L4 10. 보고기간일 이후 사건 (YES 회사)',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('YES 회사 수', agg.get('yes_count') or 0, 'int'),
             ('NO 회사 수', agg.get('no_count') or 0, 'int'),
             ('미입력 회사 수', agg.get('empty_count') or 0, 'int')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '보고기간일 이후 사건',
                              'YES 회사', agg.get('yes_count') or 0))

    # ─── L4 11 그외 우발부채·약정 (내용 있는 회사만) ─────────
    agg = all_aggs.get('l4_other_commitments') or {}
    if agg:
        sheet = 'L4_11_약정사항'
        col_defs = [
            ('회사', 32, 'company', 'text'),
            ('약정·우발부채 내용', 110, 'content', 'text'),
        ]
        _aio_write_section(wb, sheet, 'L4 11. 그외 우발부채 및 약정사항',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('내용 있는 회사', agg.get('with_content_count') or 0, 'int'),
             ('내용 없는 회사', agg.get('no_content_count') or 0, 'int')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '그외 우발부채/약정',
                              '내용 입력', agg.get('with_content_count') or 0))

    # ─── A2 유가증권 명세 ────────────────────────────────────
    agg = all_aggs.get('a2_securities') or {}
    if agg:
        sheet = 'A2_유가증권'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('계정', 18, 'account', 'text'),
            ('피투자회사명', 30, 'investee', 'text'),
            ('주식수', 14, 'shares', 'int'),
            ('지분율', 12, 'ownership_pct', 'pct'),
            ('통화', 8, 'currency', 'text'),
            ('취득원가(로컬)', 18, 'local_cost', 'num'),
            ('장부가액(로컬)', 18, 'local_book', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('취득원가(KRW)', 18, 'krw_cost', 'num'),
            ('장부가액(KRW)', 18, 'krw_book', 'num'),
        ]
        _aio_write_section(wb, sheet, 'A2 1. 유가증권 명세',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_rows') or 0, 'int'),
             ('취득원가 KRW', agg.get('grand_cost_krw') or 0, 'num'),
             ('장부가액 KRW', agg.get('grand_book_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '유가증권 명세', '장부가액 KRW',
                              agg.get('grand_book_krw') or 0))

    # ─── A3 1. 투자부동산 관련 손익 ───────────────────────────
    agg = all_aggs.get('a3_investment_pl') or {}
    if agg:
        sheet = 'A3_1_부동산손익'
        item_keys = ['rental_revenue', 'operating_expenses', 'depreciation',
                     'fv_change', 'others', 'total']
        item_labels = ['임대수익','운영비용','감가상각비','공정가치 변동','기타','합계']
        col_defs = [('회사', 28, 'company', 'text'), ('통화', 8, 'currency', 'text')]
        for k, lbl in zip(item_keys, item_labels):
            col_defs.append((f'{lbl}(로컬)', 18, f'local_{k}', 'num'))
        col_defs.append(('Spot', 10, 'spot', 'rate'))
        for k, lbl in zip(item_keys, item_labels):
            col_defs.append((f'{lbl}(KRW)', 18, f'krw_{k}', 'num'))
        _aio_write_section(wb, sheet, 'A3 1. 투자부동산 관련 손익',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('합계 KRW', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '투자부동산 관련 손익', '합계 KRW',
                              agg.get('grand_total_krw') or 0))

    # ─── A3 2-1 투자부동산(토지) 공시지가 ────────────────────
    agg = all_aggs.get('a3_land_investment') or {}
    if agg:
        sheet = 'A3_2-1_투자토지'
        col_defs = [
            ('회사', 30, 'company', 'text'),
            ('해당여부', 12, 'yn', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('공시지가(로컬)', 22, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('공시지가(KRW)', 22, 'krw', 'num'),
        ]
        _aio_write_section(wb, sheet, 'A3 2-1. 투자부동산(토지) 공시지가',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('YES 회사', agg.get('yes_count') or 0, 'int'),
             ('금액 입력 회사', agg.get('with_amount_count') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '투자부동산 토지 공시지가', 'KRW',
                              agg.get('grand_total_krw') or 0))

    # ─── A3 3-1 유형자산(토지) 공시지가 ──────────────────────
    agg = all_aggs.get('a3_land_ppe') or {}
    if agg:
        sheet = 'A3_3-1_유형토지'
        col_defs = [
            ('회사', 30, 'company', 'text'),
            ('해당여부', 12, 'yn', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('공시지가(로컬)', 22, 'local', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('공시지가(KRW)', 22, 'krw', 'num'),
        ]
        _aio_write_section(wb, sheet, 'A3 3-1. 유형자산(토지) 공시지가',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('YES 회사', agg.get('yes_count') or 0, 'int'),
             ('금액 입력 회사', agg.get('with_amount_count') or 0, 'int'),
             ('전체 KRW 합계', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '유형자산 토지 공시지가', 'KRW',
                              agg.get('grand_total_krw') or 0))

    # ─── A4 1. 공사계약 잔액 변동 ─────────────────────────────
    agg = all_aggs.get('a4_construction_balance') or {}
    if agg:
        sheet = 'A4_1_공사잔액'
        item_keys = ['beginning', 'variance', 'profit', 'others', 'ending']
        item_lbls = ['기초','증감','공사수익','기타','기말']
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('공사 종류', 16, 'type_label', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for k, lbl in zip(item_keys, item_lbls):
            col_defs.append((f'{lbl}(로컬)', 18, f'local_{k}', 'num'))
        col_defs.append(('환율(Avg)', 11, 'spot', 'rate'))
        for k, lbl in zip(item_keys, item_lbls):
            col_defs.append((f'{lbl}(KRW)', 18, f'krw_{k}', 'num'))
        _aio_write_section(wb, sheet, 'A4 1. 공사계약 잔액의 변동내역 (Avg 환율)',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('기말잔액 KRW', agg.get('krw_ending') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '공사계약 잔액 변동', '기말 KRW',
                              agg.get('krw_ending') or 0))

    # ─── A4 2. 공사손익 (Pivot) ──────────────────────────────
    agg = all_aggs.get('a4_construction_profit') or {}
    if agg:
        sheet = 'A4_2_공사손익'
        type_keys = ['architecture','civil','plant','hydrogen','others','total']
        type_lbls = ['건축','토목','플랜트','수소충전소','Others','Total']
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('항목', 22, 'item_label', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for k, lbl in zip(type_keys, type_lbls):
            col_defs.append((f'{lbl}(로컬)', 18, f'local_{k}', 'num'))
        col_defs.append(('환율(Avg)', 11, 'spot', 'rate'))
        for k, lbl in zip(type_keys, type_lbls):
            col_defs.append((f'{lbl}(KRW)', 18, f'krw_{k}', 'num'))
        income_total = ((agg.get('totals_krw') or {}).get('accumulated_income') or {}).get('total') or 0
        _aio_write_section(wb, sheet, 'A4 2. 진행중인 건설계약 공사손익 (Avg 환율)',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('누적공사손익 Total KRW', income_total, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '건설계약 공사손익', '손익 Total KRW', income_total))

    # ─── A4 3. 계약자산·부채 ─────────────────────────────────
    agg = all_aggs.get('a4_contract_balance') or {}
    if agg:
        sheet = 'A4_3_계약자산부채'
        item_keys = ['receivable', 'payable', 'advance']
        item_lbls = ['미청구공사','초과청구공사','선수금']
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('공사 종류', 16, 'type_label', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for k, lbl in zip(item_keys, item_lbls):
            col_defs.append((f'{lbl}(로컬)', 22, f'local_{k}', 'num'))
        col_defs.append(('Spot', 10, 'spot', 'rate'))
        for k, lbl in zip(item_keys, item_lbls):
            col_defs.append((f'{lbl}(KRW)', 22, f'krw_{k}', 'num'))
        _aio_write_section(wb, sheet, 'A4 3. 계약자산 및 계약부채 (Spot 환율)',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('미청구공사 KRW', agg.get('krw_receivable') or 0, 'num'),
             ('초과청구공사 KRW', agg.get('krw_payable') or 0, 'num'),
             ('선수금 KRW', agg.get('krw_advance') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '계약자산 및 계약부채', '미청구 KRW',
                              agg.get('krw_receivable') or 0))

    # ─── A5 1. 사용권자산 변동 ───────────────────────────────
    agg = all_aggs.get('a5_rou_changes') or {}
    if agg:
        sheet = 'A5_1_사용권자산'
        item_keys = ['beginning','acquisition','disposal','depreciation','others','ending']
        item_lbls = ['기초','취득','처분','상각','기타','기말']
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('자산 종류', 16, 'asset_label', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for k, lbl in zip(item_keys, item_lbls):
            col_defs.append((f'{lbl}(로컬)', 18, f'local_{k}', 'num'))
        col_defs.append(('Spot', 10, 'spot', 'rate'))
        for k, lbl in zip(item_keys, item_lbls):
            col_defs.append((f'{lbl}(KRW)', 18, f'krw_{k}', 'num'))
        _aio_write_section(wb, sheet, 'A5 1. 사용권자산의 변동내역',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('기말잔액 KRW', agg.get('krw_ending') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '사용권자산 변동', '기말 KRW',
                              agg.get('krw_ending') or 0))

    # ─── A5 2. 리스 손익 ─────────────────────────────────────
    agg = all_aggs.get('a5_lease_pl') or {}
    if agg:
        sheet = 'A5_2_리스손익'
        item_keys = ['depreciation','interest','short_term','low_value',
                     'variable','disposal_gain','total']
        item_lbls = ['감가상각비','이자비용','단기리스','소액자산리스',
                     '변동리스료','처분이익','합계']
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('통화', 8, 'currency', 'text'),
        ]
        for k, lbl in zip(item_keys, item_lbls):
            col_defs.append((f'{lbl}(로컬)', 20, f'local_{k}', 'num'))
        col_defs.append(('Spot', 10, 'spot', 'rate'))
        for k, lbl in zip(item_keys, item_lbls):
            col_defs.append((f'{lbl}(KRW)', 20, f'krw_{k}', 'num'))
        _aio_write_section(wb, sheet, 'A5 2. 리스계약 관련 손익',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('합계 KRW', agg.get('grand_total_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '리스 손익', 'KRW',
                              agg.get('grand_total_krw') or 0))

    # ─── A6 파생상품평가손익 ─────────────────────────────────
    agg = all_aggs.get('a6_derivatives') or {}
    if agg:
        sheet = 'A6_파생상품'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('파생상품 종류', 34, 'type', 'text'),
            ('통화', 8, 'currency', 'text'),
            ('평가이익(로컬)', 20, 'local_gain', 'num'),
            ('평가손실(로컬)', 20, 'local_loss', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('평가이익(KRW)', 20, 'krw_gain', 'num'),
            ('평가손실(KRW)', 20, 'krw_loss', 'num'),
        ]
        _aio_write_section(wb, sheet, 'A6 1. 파생상품평가손익 내역',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('평가이익 KRW', agg.get('grand_gain_krw') or 0, 'num'),
             ('평가손실 KRW', agg.get('grand_loss_krw') or 0, 'num'),
             ('순손익 KRW', agg.get('grand_net_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '파생상품평가손익', '순손익 KRW',
                              agg.get('grand_net_krw') or 0))

    # ─── A7 지분법투자주식 ───────────────────────────────────
    agg = all_aggs.get('a7_equity_method') or {}
    if agg:
        sheet = 'A7_지분법투자'
        col_defs = [
            ('회사', 28, 'company', 'text'),
            ('종류', 22, 'type_label', 'text'),
            ('피투자회사명', 34, 'investee', 'text'),
            ('지분율', 12, 'ownership_pct', 'pct'),
            ('통화', 8, 'currency', 'text'),
            ('취득원가(로컬)', 20, 'local_cost', 'num'),
            ('순자산(로컬)', 20, 'local_net_asset', 'num'),
            ('장부가(로컬)', 20, 'local_book', 'num'),
            ('Spot', 10, 'spot', 'rate'),
            ('취득원가(KRW)', 20, 'krw_cost', 'num'),
            ('순자산(KRW)', 20, 'krw_net_asset', 'num'),
            ('장부가(KRW)', 20, 'krw_book', 'num'),
        ]
        _aio_write_section(wb, sheet, 'A7 1. 지분법투자주식 명세',
            [('결산기간', agg.get('year'), 'text'),
             ('스캔 회사 수', agg.get('scanned') or 0, 'int'),
             ('입력 회사 수', agg.get('with_data_count') or 0, 'int'),
             ('종속회사 라인', agg.get('subsidiary_count') or 0, 'int'),
             ('기타지분법 라인', agg.get('other_count') or 0, 'int'),
             ('취득원가 KRW', agg.get('grand_cost_krw') or 0, 'num'),
             ('장부가액 KRW', agg.get('grand_book_krw') or 0, 'num')],
            col_defs, agg.get('rows') or [])
        sections_meta.append((sheet, '지분법투자주식', '장부가 KRW',
                              agg.get('grand_book_krw') or 0))

    # ─── 표지 시트 (마지막에 작성하여 sections_meta 활용) ──────
    _aio_cover_sheet(wb, year, sections_meta)

    wb.save(str(output_path))
    return str(output_path)
