"""
연결정산 엔진
- 합산(각사 BS/PL_MF) + 분개(차/대) → 최종 산출
- 분개 엑셀 템플릿 생성/파싱
- 최종 연결재무제표 엑셀 생성
"""
from __future__ import annotations
import io
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from consol_schema import load_template

# 사용자 요청 — 분개 부호 강제: 차변(+) / 대변(-)
# 4900002 비지배지분순이익 (detail 행만 적용)
# 4900001 지배지분순이익은 공식(당기순이익-비지배지분순이익)로만 계산 (분개 직접 반영 X)
FORCE_DR_SIGN_CODES = {'4900002'}


# ─── 계산 엔진 ──────────────────────────────────────────────────────────────

def _empty_company_row(companies):
    return {c: 0.0 for c in companies}


def compute_with_rollup(agg, adj_entries, inter_entries, companies,
                        rolled_up_groups: list[dict] | None = None):
    """rollup 컬럼을 추가해 compute() 호출."""
    extended_companies = list(companies) + [r['name'] for r in (rolled_up_groups or [])]

    bs = dict((agg.get('sheets') or {}).get('BS') or {})
    pl = dict((agg.get('sheets') or {}).get('PL_MF') or {})

    new_bs = {}
    for code, info in bs.items():
        new_info = dict(info or {})
        new_info['by_company'] = dict(info.get('by_company', {}) or {})
        new_bs[code] = new_info
    new_pl = {}
    for code, info in pl.items():
        new_info = dict(info or {})
        new_info['by_company'] = dict(info.get('by_company', {}) or {})
        new_pl[code] = new_info

    for r in (rolled_up_groups or []):
        name = r['name']
        for code, val in (r.get('rows_by_code') or {}).items():
            target = new_bs if str(code).startswith(('1', '2', '3')) else new_pl
            if code not in target:
                target[code] = {'kor': '', 'eng': '', 'by_company': {}, 'total': 0}
            target[code]['by_company'][name] = val

    new_agg = {'sheets': {'BS': new_bs, 'PL_MF': new_pl}}
    return compute(new_agg, adj_entries, inter_entries, extended_companies)


def compute(agg, adj_entries, inter_entries, companies):
    """
    agg          : aggregator.aggregate() 결과
    adj_entries  : 연결조정 분개 리스트 (자본/투자 상계, 영업권, 지분법 등)
    inter_entries: 내부거래 분개 리스트 (매출-매출원가 상계, 채권채무, 미실현 등)
    companies    : 합산 컬럼 회사 순서

    반환: 'rows' = [{'row','kind','code','name','companies':{...},'sum',
                    'dr_adj','cr_adj','dr_int','cr_int','final'}]

    신 양식 (2026.1Q): 4개 분개 컬럼 N(연결조정차) O(연결조정대) P(내부거래차) Q(내부거래대)
      detail D-sign: final = M + N - O + P - Q
      detail C-sign: final = M - N + O - P + Q
      subtotal: SUM(children's final)
    """
    template = load_template()
    bs = (agg.get('sheets') or {}).get('BS') or {}
    pl = (agg.get('sheets') or {}).get('PL_MF') or {}

    # 1) 두 종류 분개를 각각 code별로 집계
    def _aggregate(entries):
        out = {}
        for e in entries or []:
            dc = str(e.get('debit_code') or '').strip()
            cc = str(e.get('credit_code') or '').strip()
            da = _to_num(e.get('debit_amt'))
            ca = _to_num(e.get('credit_amt'))
            if dc:
                out.setdefault(dc, {'dr': 0.0, 'cr': 0.0})['dr'] += da
            if cc:
                out.setdefault(cc, {'dr': 0.0, 'cr': 0.0})['cr'] += ca
        return out

    adj_by_code   = _aggregate(adj_entries)
    inter_by_code = _aggregate(inter_entries)

    # 2) 1차 패스: 각 행 초기화 + detail 행 계산
    rows_out = []
    row_to_final = {}    # row index → final value (P-ref 계산용)

    for row in template['rows']:
        kind = row.get('kind')
        out = {
            'row': row['row'],
            'kind': kind,
            'code': row.get('code'),
            'name': row.get('name'),
            'section': row.get('section'),
            'companies': _empty_company_row(companies),
            'sum': 0.0,
            'dr_adj': 0.0, 'cr_adj': 0.0,    # 연결조정 차/대
            'dr_int': 0.0, 'cr_int': 0.0,    # 내부거래 차/대
            'final': 0.0,
        }

        # 분개 4컬럼 — 자기 코드로 lookup
        code = row.get('code')
        if code:
            if row.get('n_adj_mode') == 'lookup':
                out['dr_adj'] = adj_by_code.get(str(code), {'dr': 0.0}).get('dr', 0.0)
            if row.get('o_adj_mode') == 'lookup':
                out['cr_adj'] = adj_by_code.get(str(code), {'cr': 0.0}).get('cr', 0.0)
            if row.get('p_int_mode') == 'lookup':
                out['dr_int'] = inter_by_code.get(str(code), {'dr': 0.0}).get('dr', 0.0)
            if row.get('q_int_mode') == 'lookup':
                out['cr_int'] = inter_by_code.get(str(code), {'cr': 0.0}).get('cr', 0.0)

        if kind == 'detail':
            code_s = str(row['code'])
            sign = row.get('sign', 'D')

            # 사용자 요청: 4900001 지배지분순이익, 4900002 비지배지분순이익은
            #              차변+/대변- (sign='D')로 강제
            if code_s in FORCE_DR_SIGN_CODES:
                sign = 'D'

            # 각사 수치 — BS / PL_MF에서 lookup
            info = bs.get(code_s) or pl.get(code_s) or {}
            by_co = info.get('by_company', {}) or {}
            for c in companies:
                v = by_co.get(c, 0) or 0
                out['companies'][c] = v
            out['sum'] = sum(out['companies'].values())

            # 최종 (4컬럼 양식)
            #   자산/비용: M + N - O + P - Q  (차변=증가)
            #   부채/자본/수익: M - N + O - P + Q  (대변=증가)
            if sign == 'D':
                out['final'] = (out['sum'] + out['dr_adj'] - out['cr_adj']
                                + out['dr_int'] - out['cr_int'])
            else:  # 'C'
                out['final'] = (out['sum'] - out['dr_adj'] + out['cr_adj']
                                - out['dr_int'] + out['cr_int'])

            row_to_final[row['row']] = out['final']

        rows_out.append(out)

    # 3) 2차 패스: subtotal / formula / check 행 — 의존성 해소를 위해 다회 반복
    tpl_by_row = {r['row']: r for r in template['rows']}
    out_by_row = {r['row']: r for r in rows_out}

    # check 행의 add/sub refs 파싱 (1회만) — 신 양식에서 R열 참조 사용
    import re as _re
    check_refs = {}
    for trow in template['rows']:
        if trow.get('kind') == 'check':
            f = (trow.get('p_formula') or '')[1:]
            add_r, sub_r = [], []
            # R열 참조 우선, 없으면 P열 (구버전 호환)
            for m in _re.finditer(r'([+\-]?)[RP](\d+)', f):
                ref = int(m.group(2))
                (sub_r if m.group(1) == '-' else add_r).append(ref)
            check_refs[trow['row']] = (add_r, sub_r)

    JOURNAL_COLS = [
        ('dr_adj', 'n_adj_mode', 'n_adj_refs'),
        ('cr_adj', 'o_adj_mode', 'o_adj_refs'),
        ('dr_int', 'p_int_mode', 'p_int_refs'),
        ('cr_int', 'q_int_mode', 'q_int_refs'),
    ]

    MAX_ITER = 8
    for _ in range(MAX_ITER):
        changed = False
        for out in rows_out:
            kind = out['kind']
            if kind not in ('subtotal', 'formula', 'check'):
                continue
            trow = tpl_by_row.get(out['row'])
            if not trow:
                continue

            if kind == 'subtotal':
                refs = trow.get('sum_range') or trow.get('sum_refs') or []
                add_refs, sub_refs = refs, []
            elif kind == 'formula':
                add_refs = trow.get('add_refs') or []
                sub_refs = trow.get('sub_refs') or []
            else:  # check
                add_refs, sub_refs = check_refs.get(out['row'], ([], []))

            # sum / companies / final 은 자식들로부터
            new_final = (sum(row_to_final.get(r, 0.0) for r in add_refs)
                         - sum(row_to_final.get(r, 0.0) for r in sub_refs))
            new_sum   = (sum((out_by_row.get(r) or {}).get('sum', 0.0) for r in add_refs)
                         - sum((out_by_row.get(r) or {}).get('sum', 0.0) for r in sub_refs))
            new_co = {}
            for c in companies:
                new_co[c] = (sum(((out_by_row.get(r) or {}).get('companies') or {}).get(c, 0.0) for r in add_refs)
                             - sum(((out_by_row.get(r) or {}).get('companies') or {}).get(c, 0.0) for r in sub_refs))

            # 4900001 지배지분순이익은 공식(=당기순이익-비지배지분순이익) 그대로 사용,
            # 분개 직접 반영하지 않음 (FORCE_DR_SIGN_CODES는 detail 행에만 적용)

            # 4개 분개 컬럼 결정: lookup → 이미 1차 패스에서 설정됨,
            #                     row_sum → 지정된 행들의 같은 컬럼 합,
            #                     none    → 자식들 합산
            new_journal = {}
            for col, mode_key, refs_key in JOURNAL_COLS:
                mode = trow.get(mode_key, 'none')
                if mode == 'lookup':
                    new_journal[col] = out[col]
                elif mode == 'row_sum':
                    refs_r = trow.get(refs_key) or []
                    new_journal[col] = sum((out_by_row.get(r) or {}).get(col, 0.0) for r in refs_r)
                else:  # none
                    new_journal[col] = (
                        sum((out_by_row.get(r) or {}).get(col, 0.0) for r in add_refs)
                        - sum((out_by_row.get(r) or {}).get(col, 0.0) for r in sub_refs))

            if (new_final != out['final'] or new_sum != out['sum']
                    or any(new_journal[c] != out[c] for c in new_journal)):
                changed = True
            out['final'] = new_final
            out['sum']   = new_sum
            for c, v in new_journal.items():
                out[c] = v
            out['companies'] = new_co
            row_to_final[out['row']] = new_final

        if not changed:
            break

    return {
        'companies': companies,
        'rows': rows_out,
    }


def _row_sum(rows_out, row_idx):
    for r in rows_out:
        if r['row'] == row_idx:
            return r.get('sum', 0.0)
    return 0.0


def _row_attr(rows_out, row_idx, attr):
    for r in rows_out:
        if r['row'] == row_idx:
            return r.get(attr, 0.0)
    return 0.0


def _row_co(rows_out, row_idx, company):
    for r in rows_out:
        if r['row'] == row_idx:
            return (r.get('companies') or {}).get(company, 0.0)
    return 0.0


def _to_num(v):
    if v in (None, ''):
        return 0.0
    try:
        return float(str(v).replace(',', '').strip())
    except Exception:
        return 0.0


# ─── 분개 엑셀 템플릿 생성 ──────────────────────────────────────────────────

JOURNAL_HEADERS = ['분개번호', '차변코드', '차변계정명', '차변금액', '대변코드', '대변계정명', '대변금액', '적요']
# 내부거래 시트 전용 — 차변/대변 코드 왼쪽에 '회사' 컬럼 추가
JOURNAL_HEADERS_WITH_CO = ['분개번호',
                           '차변회사', '차변코드', '차변계정명', '차변금액',
                           '대변회사', '대변코드', '대변계정명', '대변금액',
                           '적요']

HDR_FILL = PatternFill('solid', start_color='1F3864')
HDR_FONT = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=11)
DATA_FONT = Font(name='맑은 고딕', size=10)
TOTAL_FILL = PatternFill('solid', start_color='FFE699')
TOTAL_FONT = Font(bold=True, name='맑은 고딕', size=10, color='9C5700')
SUBHDR_FILL = PatternFill('solid', start_color='D9E1F2')
NUM_FMT = '#,##0;(#,##0);"-"'
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_journal_sheet(ws, title: str, subtitle: str, entries: list | None,
                         include_company: bool = False):
    """분개 시트 1개 작성.

    include_company=True (내부거래 시트): 차/대변 코드 왼쪽에 '차변회사'/'대변회사' 컬럼 추가.
      컬럼: 분개번호 | 차변회사 | 차변코드 | 차변계정명 | 차변금액 |
                       대변회사 | 대변코드 | 대변계정명 | 대변금액 | 적요
    """
    headers = JOURNAL_HEADERS_WITH_CO if include_company else JOURNAL_HEADERS
    n_cols = len(headers)
    last_col_letter = get_column_letter(n_cols)
    amt_cols = (5, 9) if include_company else (4, 7)   # 차변/대변금액 컬럼 위치

    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, name='맑은 고딕', color='1F3864')
    ws.merge_cells(f'A1:{last_col_letter}1')

    ws['A2'] = subtitle
    ws['A2'].font = Font(size=9, italic=True, name='맑은 고딕', color='666666')
    ws.merge_cells(f'A2:{last_col_letter}2')

    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

    r = 5
    if entries:
        for e in entries:
            col = 1
            ws.cell(r, col, e.get('no') or ''); col += 1
            if include_company:
                ws.cell(r, col, e.get('debit_company') or ''); col += 1
            ws.cell(r, col, e.get('debit_code') or ''); col += 1
            ws.cell(r, col, e.get('debit_name') or ''); col += 1
            cd = ws.cell(r, col, e.get('debit_amt') or None); cd.number_format = NUM_FMT; col += 1
            if include_company:
                ws.cell(r, col, e.get('credit_company') or ''); col += 1
            ws.cell(r, col, e.get('credit_code') or ''); col += 1
            ws.cell(r, col, e.get('credit_name') or ''); col += 1
            cc = ws.cell(r, col, e.get('credit_amt') or None); cc.number_format = NUM_FMT; col += 1
            ws.cell(r, col, e.get('memo') or '')
            for col in range(1, n_cols + 1):
                ws.cell(r, col).font = DATA_FONT
                ws.cell(r, col).border = BORDER
            r += 1

    for _ in range(500):
        for col in range(1, n_cols + 1):
            cell = ws.cell(r, col)
            cell.font = DATA_FONT
            cell.border = BORDER
            if col in amt_cols:
                cell.number_format = NUM_FMT
        r += 1

    if include_company:
        widths = [10, 14, 12, 24, 16, 14, 12, 24, 16, 30]
    else:
        widths = [10, 12, 24, 16, 12, 24, 16, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[4].height = 26
    ws.freeze_panes = 'A5'


def make_journal_template(group_name: str, period: str,
                          adjustment_entries: list | None = None,
                          intercompany_entries: list | None = None) -> bytes:
    """분개 입력용 엑셀 (2개 시트). 기존 entries 있으면 채워서 반환.

    시트1 '연결조정' : 자본/투자 상계, 영업권, 지분법 등
    시트2 '내부거래' : 매출-매출원가, 채권채무, 미실현이익 등
    """
    wb = Workbook()
    # 첫 시트는 '연결조정'으로 사용
    ws_adj = wb.active
    ws_adj.title = '연결조정'
    _write_journal_sheet(
        ws_adj,
        f'연결조정 분개 — {group_name} / {period}',
        '※ 자본/투자 상계, 영업권, 지분법, 비지배지분 등 표준 연결조정 분개. '
        '차변회사/대변회사는 분개 귀속 회사명 (선택 입력). 코드는 7자리 패키지 COA.',
        adjustment_entries,
        include_company=True,
    )

    ws_int = wb.create_sheet('내부거래')
    _write_journal_sheet(
        ws_int,
        f'내부거래 분개 — {group_name} / {period}',
        '※ 그룹 내부거래 상계 (매출-매출원가, 채권채무, 재고미실현 등). '
        '차변회사/대변회사는 거래 양 당사자 회사명 (선택 입력). 코드는 7자리 패키지 COA.',
        intercompany_entries,
        include_company=True,
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─── 분개 엑셀 파싱 ─────────────────────────────────────────────────────────

def _parse_journal_sheet(ws) -> list[dict]:
    """엑셀 시트 1개에서 분개 entries 파싱."""
    header_row = None
    for r in range(1, min(16, ws.max_row + 1)):
        vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        short_vals = [v for v in vals if 0 < len(v) <= 12]
        has_debit_code = any(('차변' in v and '코드' in v) for v in short_vals)
        has_credit_code = any(('대변' in v and '코드' in v) for v in short_vals)
        if has_debit_code and has_credit_code:
            header_row = r
            break
    if header_row is None:
        raise ValueError('헤더(차변/대변)를 찾을 수 없습니다. 양식을 확인하세요.')

    # 컬럼 매핑
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(header_row, c).value or '').strip()
        if not v:
            continue
        # 매칭 우선순위 — '회사'가 '계정명'보다 먼저 와야 한다 ('차변회사'가 'name' 패턴에 잡히지 않도록)
        if '분개' in v and '번호' in v:
            col_map['no'] = c
        elif '차변' in v and '회사' in v:
            col_map['debit_company'] = c
        elif '대변' in v and '회사' in v:
            col_map['credit_company'] = c
        elif '차변' in v and '코드' in v:
            col_map['debit_code'] = c
        elif '차변' in v and ('계정' in v or '명' in v) and '금액' not in v:
            col_map['debit_name'] = c
        elif '차변' in v and '금액' in v:
            col_map['debit_amt'] = c
        elif '대변' in v and '코드' in v:
            col_map['credit_code'] = c
        elif '대변' in v and ('계정' in v or '명' in v) and '금액' not in v:
            col_map['credit_name'] = c
        elif '대변' in v and '금액' in v:
            col_map['credit_amt'] = c
        elif '적요' in v or '비고' in v:
            col_map['memo'] = c

    if 'debit_code' not in col_map or 'credit_code' not in col_map:
        raise ValueError('차변코드/대변코드 컬럼을 찾을 수 없습니다.')

    entries = []
    for r in range(header_row + 1, ws.max_row + 1):
        def _get(key):
            c = col_map.get(key)
            return ws.cell(r, c).value if c else None

        no = _get('no')
        dc = _get('debit_code')
        dn = _get('debit_name')
        da = _get('debit_amt')
        cc = _get('credit_code')
        cn = _get('credit_name')
        ca = _get('credit_amt')
        memo = _get('memo')
        dco = _get('debit_company')
        cco = _get('credit_company')

        if all(v in (None, '') for v in (dc, da, cc, ca)):
            continue

        entries.append({
            'no': str(no) if no not in (None, '') else '',
            'debit_company': str(dco).strip() if dco else '',
            'debit_code': _code_str(dc),
            'debit_name': str(dn).strip() if dn else '',
            'debit_amt': _to_num(da),
            'credit_company': str(cco).strip() if cco else '',
            'credit_code': _code_str(cc),
            'credit_name': str(cn).strip() if cn else '',
            'credit_amt': _to_num(ca),
            'memo': str(memo).strip() if memo else '',
            '_row': r,                       # 엑셀 행 번호 (검증 오류 메시지용)
            '_sheet': ws.title or '',
        })
    return entries


def parse_journal_excel(file_path_or_bytes) -> dict:
    """업로드된 분개 엑셀 → {'adjustment_entries': [...], 'intercompany_entries': [...]}.

    시트명으로 분류:
      '연결조정' / 'adjustment' / 'adj' → adjustment_entries
      '내부거래' / 'intercompany' / 'inter' → intercompany_entries
    인식 안되는 시트는 첫번째 발견된 시트를 adjustment로 처리 (구버전 호환).
    """
    if isinstance(file_path_or_bytes, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(file_path_or_bytes), data_only=True)
    else:
        wb = load_workbook(file_path_or_bytes, data_only=True)

    adj_entries = []
    inter_entries = []
    matched = False

    for sname in wb.sheetnames:
        low = sname.lower().strip()
        ws = wb[sname]
        if '연결조정' in sname or 'adjustment' in low or low in ('adj',):
            try:
                adj_entries.extend(_parse_journal_sheet(ws))
                matched = True
            except ValueError:
                pass
        elif '내부거래' in sname or 'intercompany' in low or low in ('inter', '내부'):
            try:
                inter_entries.extend(_parse_journal_sheet(ws))
                matched = True
            except ValueError:
                pass

    # 매칭되는 시트 없으면 활성 시트를 adjustment로 (구 양식 호환)
    if not matched:
        try:
            adj_entries = _parse_journal_sheet(wb.active)
        except ValueError as e:
            raise ValueError(f'분개 시트를 찾을 수 없습니다 (시트명: "연결조정"/"내부거래" 필요). {e}')

    return {
        'adjustment_entries': adj_entries,
        'intercompany_entries': inter_entries,
    }


def auto_bridge_pl_bs(adj_entries: list[dict], inter_entries: list[dict],
                      tolerance: float = 0.5):
    """두 종류 분개 모두 통합한 BS/PL 차이를 자동 매듭 분개로 흡수.

    BS 측 계정: 3600101 비지배지분 (Non-controlling interest)
    PL 측 계정: 4700004 당기순이익
    매듭 분개는 '연결조정' 묶음에 추가.

    반환: (adj_bridge_entries, info)
    """
    def _sec(code):
        c = str(code or '')
        if c.startswith(('1', '2', '3')): return 'BS'
        if c.startswith(('4', '5')): return 'PL'
        return 'OTHER'

    bs_dr = bs_cr = pl_dr = pl_cr = 0.0
    for e in (list(adj_entries or []) + list(inter_entries or [])):
        da = float(e.get('debit_amt') or 0)
        ca = float(e.get('credit_amt') or 0)
        sec_d = _sec(e.get('debit_code'))
        sec_c = _sec(e.get('credit_code'))
        if sec_d == 'BS': bs_dr += da
        elif sec_d == 'PL': pl_dr += da
        if sec_c == 'BS': bs_cr += ca
        elif sec_c == 'PL': pl_cr += ca

    imbalance = bs_dr - bs_cr
    info = {
        'bs_dr': bs_dr, 'bs_cr': bs_cr,
        'pl_dr': pl_dr, 'pl_cr': pl_cr,
        'imbalance': imbalance,
        'bridge_amount': 0.0,
        'applied': False,
    }
    if abs(imbalance) <= tolerance:
        return [], info

    amount = abs(imbalance)
    if imbalance < 0:
        # BS 대변 우세 → 자본(당기순이익) 감소 필요: 차변 3500105 / 대변 4700004
        bridge = {
            'no': 'AUTO-BRIDGE',
            'debit_code': '3500105', 'debit_name': '당기순이익',
            'debit_amt': amount,
            'credit_code': '4700004', 'credit_name': '당기순이익',
            'credit_amt': amount,
            'memo': f'[자동 매듭] PL→BS 흐름 보정 (BS 차/대 차이 {imbalance:,.0f} 흡수)',
        }
    else:
        # BS 차변 우세 → 자본(당기순이익) 증가 필요: 차변 4700004 / 대변 3500105
        bridge = {
            'no': 'AUTO-BRIDGE',
            'debit_code': '4700004', 'debit_name': '당기순이익',
            'debit_amt': amount,
            'credit_code': '3500105', 'credit_name': '당기순이익',
            'credit_amt': amount,
            'memo': f'[자동 매듭] PL→BS 흐름 보정 (BS 차/대 차이 +{imbalance:,.0f} 흡수)',
        }
    info['bridge_amount'] = imbalance
    info['applied'] = True
    return [bridge], info


def validate_codes_present(entries: list[dict], tolerance: float = 0.5):
    """분개 검증: 금액이 있는데 코드가 비어 있는 셀이 있는지 확인.

    반환: {'ok': bool, 'missing': [{'sheet','row','no','side','amount'}], ...}
      - side: 'debit' or 'credit'
      - row: 엑셀의 1-base 행 번호 (없으면 0)
    tolerance 이상의 금액이 입력됐는데 같은 쪽 코드가 비어 있으면 missing에 추가.
    """
    missing = []
    for idx, e in enumerate(entries):
        da = float(e.get('debit_amt') or 0)
        ca = float(e.get('credit_amt') or 0)
        dc = (e.get('debit_code') or '').strip()
        cc = (e.get('credit_code') or '').strip()
        row = e.get('_row') or 0
        sheet = e.get('_sheet') or ''
        no = (e.get('no') or '').strip()
        if abs(da) > tolerance and not dc:
            missing.append({'sheet': sheet, 'row': row, 'index': idx + 1,
                            'no': no, 'side': 'debit', 'amount': da})
        if abs(ca) > tolerance and not cc:
            missing.append({'sheet': sheet, 'row': row, 'index': idx + 1,
                            'no': no, 'side': 'credit', 'amount': ca})
    return {'ok': len(missing) == 0, 'missing': missing}


def validate_balance(entries: list[dict], tolerance: float = 0.5):
    """분개 차/대 균형 검증.

    반환: {'ok': bool, 'total_debit': float, 'total_credit': float, 'diff': float,
           'unbalanced_journals': [{'no', 'debit', 'credit', 'diff'}], ...}
    - 전체 차/대 합이 tolerance 초과로 불일치하면 ok=False
    - 분개번호(no)가 부여된 경우 번호별로도 균형 검증해 불일치 목록 반환
    """
    total_d = 0.0
    total_c = 0.0
    by_no: dict[str, dict] = {}

    for e in entries:
        da = float(e.get('debit_amt') or 0)
        ca = float(e.get('credit_amt') or 0)
        total_d += da
        total_c += ca
        no = (e.get('no') or '').strip()
        if no:
            slot = by_no.setdefault(no, {'no': no, 'debit': 0.0, 'credit': 0.0})
            slot['debit'] += da
            slot['credit'] += ca

    unbalanced = []
    for no, slot in by_no.items():
        d = slot['debit'] - slot['credit']
        if abs(d) > tolerance:
            unbalanced.append({**slot, 'diff': d})
    unbalanced.sort(key=lambda x: -abs(x['diff']))

    diff = total_d - total_c
    return {
        'ok': abs(diff) <= tolerance,
        'total_debit': total_d,
        'total_credit': total_c,
        'diff': diff,
        'unbalanced_journals': unbalanced,
        'entry_count': len(entries),
    }


def _code_str(v):
    if v in (None, ''):
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ─── 최종 연결재무제표 엑셀 출력 ───────────────────────────────────────────

def write_consolidation_excel(result, group_name: str, period: str, output_path: str):
    """compute() 결과를 태림연결 양식 엑셀로 저장 (4개 분개 컬럼 + 선택적 전년 4Q 비교 컬럼)."""
    companies = result['companies']
    rows = result['rows']
    prior_period = result.get('prior_period')

    wb = Workbook()
    ws = wb.active
    ws.title = '연결정산'

    EXTRA_COLS = ['합산', '연결조정(차)', '연결조정(대)', '내부거래(차)', '내부거래(대)', '최종']
    if prior_period:
        EXTRA_COLS = EXTRA_COLS + ['', f'{prior_period} 최종']
    headers = ['계정코드', '계정명'] + list(companies) + EXTRA_COLS

    ws['A1'] = f'연결재무제표 — {group_name} / {period}'
    ws['A1'].font = Font(bold=True, size=14, name='맑은 고딕', color='1F3864')
    last_col_letter = get_column_letter(len(headers))
    ws.merge_cells(f'A1:{last_col_letter}1')

    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER

    for row_idx, r in enumerate(rows, start=4):
        kind = r['kind']
        if kind in ('blank',):
            continue
        if kind == 'header' or kind == 'subheader':
            if kind == 'subheader':
                ws.cell(row_idx, 1, 'Code').font = HDR_FONT
                ws.cell(row_idx, 2, '계정명').font = HDR_FONT
                for col in range(1, len(headers) + 1):
                    ws.cell(row_idx, col).fill = SUBHDR_FILL
                    ws.cell(row_idx, col).border = BORDER
            continue

        ws.cell(row_idx, 1, r.get('code') or '')
        ws.cell(row_idx, 2, r.get('name') or '')
        col = 3
        for c in companies:
            v = (r.get('companies') or {}).get(c, 0) or 0
            cell = ws.cell(row_idx, col, v if v else None)
            cell.number_format = NUM_FMT
            col += 1
        for key in ('sum', 'dr_adj', 'cr_adj', 'dr_int', 'cr_int', 'final'):
            v = r.get(key) or 0
            cell = ws.cell(row_idx, col, v if v else None)
            cell.number_format = NUM_FMT
            col += 1
        if prior_period:
            col += 1  # 구분용 빈 컬럼
            pv = r.get('prior_final') or 0
            cell = ws.cell(row_idx, col, pv if pv else None)
            cell.number_format = NUM_FMT
            col += 1

        is_total = kind in ('subtotal', 'formula', 'check')
        for cc in range(1, len(headers) + 1):
            cell = ws.cell(row_idx, cc)
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
            else:
                cell.font = DATA_FONT
            cell.border = BORDER

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 26
    for i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[3].height = 36
    ws.freeze_panes = 'C4'

    wb.save(output_path)
    return output_path
