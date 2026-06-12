"""distribute 기능 검증 스크립트"""
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import distribute_builder as db

# 1) 업로드된 패키지 목록 확인
state = Path("uploads/_state.json")
data = json.loads(state.read_text(encoding="utf-8"))
print(f"총 업로드 파일: {len(data)}개")
y25q4 = [f for f in data if f.get("year") == "2025-4Q"]
print(f"2025-4Q: {len(y25q4)}개")

# 글로벌세아 찾기
hits = [f for f in data if "글로벌세아" in (f.get("company") or "")]
print(f"\n글로벌세아 업로드: {len(hits)}개")
for f in hits:
    print(f"  year={f.get('year')}, path={f.get('path')}, exists={Path(f.get('path','')).exists()}")

# 2) 글로벌세아 2025-4Q 패키지에서 BS/PL 데이터 추출 테스트
g_q4 = next((f for f in data if f.get("year") == "2025-4Q" and "글로벌세아" in (f.get("company") or "")), None)
if not g_q4:
    print("\n[FAIL] 글로벌세아 2025-4Q 패키지 없음")
    sys.exit(1)

print(f"\n=== {g_q4['company']} / {g_q4['year']} 패키지에서 BS/PL 추출 ===")
ext = db.extract_bs_pl_from_package(Path(g_q4["path"]))
print(f"company: {ext['company']}")
print(f"currency: {ext['currency']}")
print(f"year/quarter: {ext['year']}/{ext['quarter']}")
print(f"bs(KRW) 코드 수: {len(ext['bs'])}")
print(f"bs_local 코드 수: {len(ext['bs_local'])}")
print(f"pl(local) 코드 수: {len(ext['pl'])}")

# 샘플
print("\nbs(KRW) 샘플 5개:")
for k, v in list(ext["bs"].items())[:5]:
    print(f"  {k}: {v:,.0f}")
print("\nbs_local 샘플 5개:")
for k, v in list(ext["bs_local"].items())[:5]:
    print(f"  {k}: {v:,.0f}")
print("\npl(local) 샘플 5개:")
for k, v in list(ext["pl"].items())[:5]:
    print(f"  {k}: {v:,.0f}")

# 3) resolve_py_data 테스트 (2026 1Q 대상)
print("\n=== resolve_py_data(target=2026 1Q, 글로벌세아) ===")
target_year = 2026
target_q = 1
resolved = db.resolve_py_data(data, "글로벌세아", target_year, target_q)
print(f"ok: {resolved['ok']}")
print(f"reason: {resolved['reason']}")
print(f"source_bs: {resolved['source_bs']}")
print(f"source_pl: {resolved['source_pl']}")
print(f"pl_scale: {resolved['pl_scale']}")
print(f"bs 코드 수: {len(resolved['bs'])}")
print(f"bs_local 코드 수: {len(resolved['bs_local'])}")
print(f"pl 코드 수: {len(resolved['pl'])}")
if resolved["bs"]:
    print(f"bs 첫번째: {next(iter(resolved['bs'].items()))}")
if resolved["pl"]:
    sample = next(iter(resolved['pl'].items()))
    print(f"pl 첫번째 (scale 적용 후): {sample}")

# 4) fill_template으로 실제 파일 생성 (암호화 없이)
print("\n=== fill_template 테스트 (암호화 없이) ===")
tpl = db.get_template_path("2026")
if not tpl:
    print("[FAIL] 2026 템플릿 없음")
    sys.exit(1)

out_dir = Path("results/distribute/_test_dryrun")
out_dir.mkdir(parents=True, exist_ok=True)
out_path = out_dir / "글로벌세아_TEST.xlsm"

fill_res = db.fill_template(
    template_path=tpl,
    output_path=out_path,
    company="글로벌세아",
    target_year="2026",
    target_quarter="1",
    bs_data=resolved["bs"],
    bs_local_data=resolved["bs_local"],
    pl_data=resolved["pl"],
    sheet_protect_password=None,  # 검증을 위해 시트 보호 없이
)
print(f"wrote_bs: {fill_res['wrote_bs']}")
print(f"wrote_bs_local: {fill_res['wrote_bs_local']}")
print(f"wrote_pl: {fill_res['wrote_pl']}")
print(f"파일: {out_path}, size={out_path.stat().st_size:,} bytes")

# 5) 생성된 파일의 PY 시트 확인
print("\n=== 생성된 파일의 PY 시트 검증 ===")
from openpyxl import load_workbook
wb = load_workbook(str(out_path), keep_vba=True, data_only=False)
print(f"Sheets: {wb.sheetnames[:8]}...")
py = wb["PY"]
print(f"PY max_row={py.max_row}")
print("\n5행~10행 데이터:")
for r in range(5, 11):
    a = py.cell(r, 1).value
    b = py.cell(r, 2).value
    d = py.cell(r, 4).value
    e = py.cell(r, 5).value
    g = py.cell(r, 7).value
    h = py.cell(r, 8).value
    print(f"  row {r}: A={a!r}, B={b!r:>20s if isinstance(b,(int,float)) else 20} | D={d!r}, E={e!r} | G={g!r}, H={h!r}")
wb.close()

# 6) Cover 시트 확인
print("\n=== Cover 시트 확인 ===")
wb = load_workbook(str(out_path), keep_vba=True, data_only=False)
cv = wb["Cover"]
print(f"D11 (회사명): {cv['D11'].value!r}")
print(f"C9 (연도):    {cv['C9'].value!r}")
print(f"F9 (분기):    {cv['F9'].value!r}")
wb.close()

print("\n[OK] 검증 완료")
