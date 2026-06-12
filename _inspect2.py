"""PL_MF 145+, BS F열 의미, CF1~3 시트 헤더 - UTF-8 출력"""
import sys, json
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

state = json.loads(Path("uploads/_state.json").read_text(encoding="utf-8"))
src = next(f for f in state if f.get("year") == "2025-4Q" and "글로벌세아" in (f.get("company") or ""))
print(f"src: {src['path']}\n")

wb = load_workbook(src["path"], data_only=False)

# PL_MF 140~170행
print("[A] PL_MF 140~170행 ─ 145+ 데이터 영역")
ws = wb["PL_MF"]
print(f"PL_MF max_row={ws.max_row}, max_col={ws.max_column}")
for r in range(140, min(171, ws.max_row+1)):
    cells = []
    for c in range(1, min(ws.max_column+1, 16)):
        v = ws.cell(r, c).value
        if v is not None:
            s = repr(v)
            if len(s) > 60: s = s[:60] + "..."
            cells.append(f"{get_column_letter(c)}={s}")
    if cells:
        print(f"  {r:3d}: " + " | ".join(cells))

# BS 시트 헤더 + F191 주변
print("\n[B] BS 시트 1~10행, 188~200행")
ws = wb["BS"]
print(f"BS max_row={ws.max_row}, max_col={ws.max_column}")
for r in list(range(1, 11)) + list(range(188, 201)):
    if r > ws.max_row: continue
    cells = []
    for c in range(1, min(ws.max_column+1, 12)):
        v = ws.cell(r, c).value
        if v is not None:
            s = repr(v)
            if len(s) > 50: s = s[:50] + "..."
            cells.append(f"{get_column_letter(c)}={s}")
    if cells:
        print(f"  {r:3d}: " + " | ".join(cells))

# CF1, CF2, CF3 - 1~10행 + E/F열 샘플
for sn in ["CF1", "CF2", "CF3"]:
    if sn not in wb.sheetnames:
        continue
    print(f"\n[C] {sn} 시트 1~10행 + E열·F열 17, 28, 39, 41, 50, 54행")
    ws = wb[sn]
    print(f"{sn} max_row={ws.max_row}")
    for r in range(1, 11):
        cells = []
        for c in range(1, min(ws.max_column+1, 10)):
            v = ws.cell(r, c).value
            if v is not None:
                s = repr(v)
                if len(s) > 50: s = s[:50] + "..."
                cells.append(f"{get_column_letter(c)}={s}")
        if cells:
            print(f"  {r:3d}: " + " | ".join(cells))
    for rr in [7, 17, 28, 39, 41, 50, 54]:
        if rr > ws.max_row: continue
        b = ws.cell(rr, 2).value
        e = ws.cell(rr, 5).value
        f = ws.cell(rr, 6).value
        print(f"  ROW {rr}: B(label)={b!r}, E={e!r}, F={f!r}")

wb.close()
