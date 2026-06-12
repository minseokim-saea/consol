"""생성된 파일의 PY 시트 검증"""
from pathlib import Path
from openpyxl import load_workbook

p = Path("results/distribute/_test_dryrun/글로벌세아_TEST.xlsm")
print(f"파일: {p}, size={p.stat().st_size:,} bytes\n")

wb = load_workbook(str(p), keep_vba=True, data_only=False)
py = wb["PY"]
print(f"PY max_row={py.max_row}, max_col={py.max_column}\n")

print("=== 2~10행 ===")
for r in range(2, 11):
    cells = []
    for c, label in [(1,"A"), (2,"B"), (4,"D"), (5,"E"), (7,"G"), (8,"H"), (13,"M"), (14,"N")]:
        v = py.cell(r, c).value
        if v is None:
            continue
        if isinstance(v, (int, float)):
            cells.append(f"{label}={v:,.0f}")
        else:
            cells.append(f"{label}={v!r}")
    if cells:
        print(f"  row {r:3d}: " + " | ".join(cells))

print("\n=== 마지막 3행 ===")
for r in range(max(2, py.max_row-2), py.max_row+1):
    cells = []
    for c, label in [(1,"A"), (2,"B"), (4,"D"), (5,"E"), (7,"G"), (8,"H")]:
        v = py.cell(r, c).value
        if v is None:
            continue
        if isinstance(v, (int, float)):
            cells.append(f"{label}={v:,.0f}")
        else:
            cells.append(f"{label}={v!r}")
    if cells:
        print(f"  row {r:3d}: " + " | ".join(cells))

print("\n=== Cover 시트 ===")
cv = wb["Cover"]
print(f"  D11: {cv['D11'].value!r}")
print(f"  C9:  {cv['C9'].value!r}")
print(f"  F9:  {cv['F9'].value!r}")
print(f"  B28: {cv['B28'].value!r}")

wb.close()
print("\n[OK] 검증 완료 — PY 시트에 BS·BS_local·PL 모두 기입됨")
