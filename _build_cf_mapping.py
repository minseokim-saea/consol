"""
cf_mapping.json 초안 생성기.
- 태림합산CF 엑셀의 '합산' 시트에서 CF 라인 구조 추출
- 연결조정CF매칭.xlsx에서 PL→CF 매핑/부호 추출
- BS 운전자본변동 / 투자 / 재무 섹션은 합산 시트의 A열 코드를 그대로 사용
"""
import json
import openpyxl
from collections import OrderedDict

SUM_PATH = r'C:/연결시스템(26.1분기)/25.4분기/태림합산CF_25YE_시스템용.xlsx'
MAP_PATH = r'C:/패키지프로그램/연결조정CF매칭.xlsx'
OUT_PATH = r'C:/패키지프로그램/cf_mapping_draft.json'


def is_section_header(b):
    if not isinstance(b, str):
        return False
    s = b.strip()
    return (s.startswith(('Ⅰ', 'Ⅱ', 'Ⅲ', 'Ⅳ', 'Ⅴ', 'Ⅵ', 'Ⅶ'))
            or s[:2] in ('1.', '2.', '3.', '4.', '5.'))


def parse_sum_sheet():
    """합산 시트에서 (섹션, CF라인) 트리 추출."""
    wb = openpyxl.load_workbook(SUM_PATH, data_only=False)
    ws = wb['합산']

    sections = []           # [{section1, section2, rows:[...]}]
    cur1 = cur2 = None
    rows_buffer = []

    def flush():
        if cur1 and rows_buffer:
            sections.append({
                'roman': cur1,
                'sub': cur2,
                'rows': list(rows_buffer),
            })

    for r in range(2, ws.max_row + 1):
        A = ws[f'A{r}'].value
        B = ws[f'B{r}'].value
        if isinstance(B, str) and B.strip().startswith(('Ⅰ', 'Ⅱ', 'Ⅲ', 'Ⅳ', 'Ⅴ', 'Ⅵ', 'Ⅶ')):
            flush()
            rows_buffer = []
            cur1 = B.strip()
            cur2 = None
            continue
        if isinstance(B, str) and B.strip()[:2] in ('1.', '2.', '3.', '4.', '5.'):
            flush()
            rows_buffer = []
            cur2 = B.strip()
            continue
        if (A is not None) and isinstance(B, str) and B.strip() and B.strip() != '검증':
            rows_buffer.append({'code': str(A).strip(), 'name': B.strip(), 'row': r})
    flush()
    return sections


def parse_pl_mapping():
    """연결조정CF매칭.xlsx → {pl_code: {cf_code, sign}}."""
    wb = openpyxl.load_workbook(MAP_PATH, data_only=False)
    ws = wb['Sheet2']
    out = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        pl_code, name, cf_code, sign = row[0], row[1], row[2], row[3]
        if pl_code is None:
            continue
        # cf_code가 '=A{n}' 인 경우는 자기 자신을 가리킴 → pl_code 그대로
        if isinstance(cf_code, str) and cf_code.startswith('='):
            cf_code = str(pl_code)
        out[str(pl_code)] = {
            'name': name,
            'cf_code': str(cf_code) if cf_code else str(pl_code),
            'sign': '+' if sign == '양' else ('-' if sign == '음' else '+'),
        }
    return out


def section_sign(roman, sub):
    """섹션 기본부호 — 합산 시트 수식과 일치."""
    if roman.startswith('Ⅰ') and sub and '가산' in sub:
        return '+'  # 2. 가산
    if roman.startswith('Ⅰ') and sub and '차감' in sub:
        return '-'  # 3. 차감
    if roman.startswith('Ⅰ') and sub and '자산부채' in sub:
        return 'WC'  # 4. 운전자본 (자산/부채 자동판정)
    if roman.startswith('Ⅱ') and sub and '유입' in sub:
        return '+'
    if roman.startswith('Ⅱ') and sub and '유출' in sub:
        return '-'
    if roman.startswith('Ⅲ') and sub and '유입' in sub:
        return '+'
    if roman.startswith('Ⅲ') and sub and '유출' in sub:
        return '-'
    return '+'


def bs_type_from_cf_code(cf_code):
    """CF1xxx / CF2xxx → asset / liability."""
    if cf_code.startswith('CF1') or cf_code.startswith('CF3') or cf_code.startswith('CF4'):
        return 'asset'
    if cf_code.startswith('CF2') or cf_code.startswith('CF5') or cf_code.startswith('CF6'):
        return 'liability'
    return None


def build():
    sections = parse_sum_sheet()
    pl_map = parse_pl_mapping()

    result = {
        'version': '2026.1Q-draft',
        'description': '현금정산표 매핑 초안. PL=연결조정CF매칭.xlsx, BS/CF구조=태림합산CF 합산시트',
        'sections': [],
    }

    for sec in sections:
        roman = sec['roman']
        sub = sec['sub']
        base_sign = section_sign(roman, sub)

        section_out = {
            'roman': roman,
            'sub': sub,
            'sign_rule': base_sign,  # '+'=가산, '-'=차감, 'WC'=운전자본(자산음/부채양)
            'rows': [],
        }

        for row in sec['rows']:
            code = row['code']
            name = row['name']

            # 가산/차감 섹션은 PL 코드 (4xxx/5xxx)
            if base_sign in ('+', '-') and code.isdigit():
                pl = pl_map.get(code, {})
                section_out['rows'].append({
                    'cf_code': code,
                    'name': name,
                    'lookup_codes': [code],
                    'sign': base_sign,  # 섹션 부호 그대로 적용
                    'source': 'PL',
                })
            # 운전자본/투자/재무 섹션은 CF 코드
            elif code.startswith('CF'):
                bs_type = bs_type_from_cf_code(code)
                # 운전자본은 자산음/부채양, 투자/재무는 섹션 부호 그대로
                if base_sign == 'WC':
                    sign = '-' if bs_type == 'asset' else '+'
                else:
                    sign = base_sign
                # CF 코드의 숫자부 → BS 코드 후보 (예: CF1110301 → 1110301)
                bs_code_guess = code[2:] if code[2:].isdigit() else None
                section_out['rows'].append({
                    'cf_code': code,
                    'name': name,
                    'lookup_codes': [bs_code_guess] if bs_code_guess else [],
                    'sign': sign,
                    'type': bs_type,
                    'source': 'BS',
                })
            else:
                section_out['rows'].append({
                    'cf_code': code,
                    'name': name,
                    'lookup_codes': [],
                    'sign': base_sign if base_sign != 'WC' else '+',
                    'source': 'UNKNOWN',
                })

        result['sections'].append(section_out)

    # PL 매핑 중 합산 시트 가산/차감에 안 들어간 코드는 별도 섹션으로 노출
    used_pl = set()
    for sec in result['sections']:
        for row in sec['rows']:
            if row.get('source') == 'PL':
                used_pl.update(row['lookup_codes'])
    unused = []
    for code, info in pl_map.items():
        if code not in used_pl:
            unused.append({'pl_code': code, 'cf_code': info['cf_code'],
                           'name': info['name'], 'sign': info['sign']})
    result['extra_pl_mappings'] = unused

    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f'생성: {OUT_PATH}')
    print(f'섹션 수: {len(result["sections"])}')
    total_rows = sum(len(s['rows']) for s in result['sections'])
    print(f'전체 매핑 행 수: {total_rows}')
    print(f'합산시트 외 추가 PL 매핑(extra): {len(unused)} 건')


if __name__ == '__main__':
    build()
