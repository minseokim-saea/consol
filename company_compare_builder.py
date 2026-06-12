"""
회사별 전기 비교 엑셀 빌더.

각 회사 1시트에 BS / PL_MF / CF / CF1~CF4 섹션을 적층.
컬럼: 코드 | 한글 | 영문 | 로컬 당기 | 로컬 전기 | 로컬 증감액 | 로컬 증감률(%)
                                  | KRW 당기  | KRW 전기  | KRW 증감액  | KRW 증감률(%)

전기 KRW 환율 적용 규칙:
  · BS    → 전년 동기 spot rate (same_q_fx[CUR].spot)
  · PL_MF → 전년 동기 avg  rate (same_q_fx[CUR].avg)
  · CF    → 라인별 rate_kind ('spot(전기)'/'spot(당기)'/'avg') 에 따라 동기 환율 적용
  · CF1~CF4 → 패키지 prior 환산값 그대로 표시 (PY 시트에 전기 raw가 없어 비교 비활성)

동기 환율이 등록되지 않은 통화는 패키지의 prior 환산값(compare)을 그대로 사용.
KRW 회사는 환율 적용 없이 로컬=KRW 동일값으로 표시.
"""

import re
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HDR_FILL_DARK  = PatternFill('solid', start_color='1F3864')
HDR_FILL_LOCAL = PatternFill('solid', start_color='4472C4')
HDR_FILL_KRW   = PatternFill('solid', start_color='2E7D32')
HDR_FONT_W     = Font(bold=True, color='FFFFFF', name='Arial', size=10)
TITLE_FONT     = Font(bold=True, color='1F3864', name='Arial', size=14)
SECTION_FONT   = Font(bold=True, color='FFFFFF', name='Arial', size=11)
SECTION_FILL   = PatternFill('solid', start_color='8FAADC')
DATA_FONT      = Font(name='Arial', size=10)
LABEL_FONT     = Font(italic=True, name='Arial', size=10, color='666666')
NUM_FMT        = '#,##0;(#,##0);"-"'
PCT_FMT        = '+0.0%;-0.0%;"-"'
THIN           = Side(style='thin', color='D0D0D0')
CELL_BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


SECTION_DISPLAY = OrderedDict([
    ('BS',    ('대차대조표 (BS)', True)),
    ('PL_MF', ('손익계산서 (PL)', True)),
])


_BAD_SHEET_CHARS = re.compile(r'[\[\]\\/?:*]')


def _safe_sheet_name(name, used):
    """엑셀 시트명 규칙(31자, 금지문자 없음, 중복불가) 보정."""
    base = _BAD_SHEET_CHARS.sub(' ', str(name or '회사')).strip() or '회사'
    base = base[:31]
    cand = base
    i = 2
    while cand in used:
        suffix = f' ({i})'
        cand = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(cand)
    return cand


def _display_code(key):
    """`{code}::{label}` → code, `LBL::...` → 빈문자, 그 외 → 그대로."""
    if isinstance(key, str):
        if key.startswith('LBL::'):
            return ''
        if '::' in key:
            return key.split('::', 1)[0]
    return key


def _is_label_only(key):
    return isinstance(key, str) and key.startswith('LBL::')


def _cf_rate_for(rate_kind, spot_prior, spot_current, avg):
    """CF 라인의 rate_kind → 적용 환율 반환."""
    if rate_kind == 'spot(전기)':
        return spot_prior
    if rate_kind == 'spot(당기)':
        return spot_current
    return avg


def _compute_krw_prior(sheet_name, entry, currency, same_q_rates):
    """전기 KRW 산출.
      · BS    → PY 시트 E열 KRW 원본 그대로 (회사가 전년말 결산 시 환산한 KRW)
      · PL_MF → 전년 동기(same_q) avg × 로컬 raw
      · CF    → 라인 rate_kind별 동기 환율 × 로컬 raw
    해당 환율 미등록 시 entry['compare'](패키지 prior 환산값) 폴백.
    KRW 회사는 compare 그대로.
    반환: (krw_prior_value, applied_kind: 'py_raw'|'same_q'|'pkg'|'krw')
    """
    compare_pkg = entry.get('compare', 0) or 0
    cur = (currency or '').upper()
    if cur == 'KRW':
        return compare_pkg, 'krw'

    if sheet_name == 'BS':
        # PY 시트의 KRW 원본을 그대로 사용 (재환산 없음)
        pkg_raw = entry.get('compare_pkg_raw')
        if pkg_raw is not None:
            return pkg_raw, 'py_raw'
        return compare_pkg, 'pkg'

    compare_local = entry.get('compare_local', 0) or 0

    if sheet_name == 'PL_MF':
        rate = (same_q_rates or {}).get(cur, {}).get('avg')
        if rate and compare_local:
            return compare_local * rate, 'same_q'
        return compare_pkg, 'pkg'

    if sheet_name == 'CF':
        rk = entry.get('rate_kind', 'avg')
        sq = (same_q_rates or {}).get(cur, {}) or {}
        rate = _cf_rate_for(rk, sq.get('spot'), None, sq.get('avg'))
        if rate and compare_local:
            return compare_local * rate, 'same_q'
        return compare_pkg, 'pkg'

    return compare_pkg, 'pkg'


def _compute_local_prior(entry, currency, same_q_rates, sheet_name):
    """전기 로컬 raw. 우선순위: extractor가 저장한 compare_local → KRW 회사면 compare 그대로 → 0.
    KRW 회사는 로컬=KRW.
    """
    if (currency or '').upper() == 'KRW':
        return entry.get('compare', 0) or 0
    return entry.get('compare_local', 0) or 0


def _put_kv(ws, row, k, v, *, font_k=None, font_v=None):
    a = ws.cell(row, 1, k)
    b = ws.cell(row, 2, v)
    a.font = font_k or DATA_FONT
    b.font = font_v or DATA_FONT


def _build_cover(wb, year, prior_label, files_for_year, same_q_rates, currencies_in_use):
    ws = wb.create_sheet('표지', 0)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 70

    ws.cell(1, 1, '회사별 전기 비교').font = TITLE_FONT
    _put_kv(ws, 3, '당기',          year)
    _put_kv(ws, 4, '전기 (동기)',   prior_label or '-')
    _put_kv(ws, 5, '대상 회사 수',  len(files_for_year))
    _put_kv(ws, 6, '비교 컬럼',     '로컬통화 / KRW 환산 — 당기·전기·증감액·증감률')

    # 적용 규칙 요약
    ws.cell(8, 1, '적용 규칙').font = Font(bold=True, color='1F3864', size=11)
    ws.cell(9, 1,  '· 당기 BS 로컬     : 패키지 BS F열 raw (자본항목 KRW는 WCE 자본입력값 적용)').font = DATA_FONT
    ws.cell(10, 1, '· 전기 BS 로컬     : 패키지 BS 14~197행 H열 raw (전년말 결산값)').font = DATA_FONT
    ws.cell(11, 1, '· 전기 BS KRW      : PY 시트 KRW 원본 그대로 (재환산 없음)').font = DATA_FONT
    ws.cell(12, 1, '· 전기 PL/MF       : 패키지 PY 시트 로컬 raw × 전년 동기 등록 avg').font = DATA_FONT
    ws.cell(13, 1, '· 변동사유         : 당기 패키지 BS/PL L열 raw').font = DATA_FONT

    # 동기 환율 표 (PL에 적용되는 값)
    ws.cell(15, 1, '동기 환율 (PL 적용)').font = Font(bold=True, color='1F3864', size=11)
    headers = ['통화', 'Spot(동기)', 'Avg(동기)', '비고']
    for j, h in enumerate(headers, 1):
        c = ws.cell(16, j, h); c.font = HDR_FONT_W; c.fill = HDR_FILL_DARK
        c.alignment = Alignment(horizontal='center')
    row = 17
    for cur in sorted(currencies_in_use):
        rates = (same_q_rates or {}).get(cur) or {}
        sp = rates.get('spot'); av = rates.get('avg')
        ws.cell(row, 1, cur).font = DATA_FONT
        ws.cell(row, 2, sp if sp else None).font = DATA_FONT
        ws.cell(row, 3, av if av else None).font = DATA_FONT
        note = '동기 환율 미등록 → 패키지 prior 환산값 사용' if (not sp and not av) else ''
        if cur == 'KRW':
            note = 'KRW 회사 — 환율 적용 없음'
        ws.cell(row, 4, note).font = LABEL_FONT
        row += 1

    # 회사 목록
    row += 2
    ws.cell(row, 1, '대상 회사').font = Font(bold=True, color='1F3864', size=11)
    row += 1
    headers = ['#', '회사명', '통화', '시트']
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h); c.font = HDR_FONT_W; c.fill = HDR_FILL_DARK
        c.alignment = Alignment(horizontal='center')


def _write_section(ws, row, sheet_name, display_name, compare_active,
                   data, currency, same_q_rates, override_current_krw=None):
    """한 회계 섹션(BS/PL)을 ws에 그리고, 다음 시작 row 반환."""
    # 섹션 헤더 (12컬럼)
    c = ws.cell(row, 1, display_name)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.row_dimensions[row].height = 20
    row += 1

    # 컬럼 헤더 (2단) — 마지막 12번째 컬럼은 '변동사유'
    hdr1 = ['', '', '', '로컬통화', '', '', '', 'KRW (환산)', '', '', '', '']
    hdr2 = ['코드', '한글', '영문',
            '당기', '전기', '증감액', '증감률',
            '당기', '전기', '증감액', '증감률',
            '변동사유']
    for j, h in enumerate(hdr1, 1):
        c = ws.cell(row, j, h); c.font = HDR_FONT_W
        c.alignment = Alignment(horizontal='center', vertical='center')
        if 4 <= j <= 7:
            c.fill = HDR_FILL_LOCAL
        elif 8 <= j <= 11:
            c.fill = HDR_FILL_KRW
        else:
            c.fill = HDR_FILL_DARK
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=11)
    row += 1
    for j, h in enumerate(hdr2, 1):
        c = ws.cell(row, j, h); c.font = HDR_FONT_W
        c.alignment = Alignment(horizontal='center', vertical='center')
        if 4 <= j <= 7:
            c.fill = HDR_FILL_LOCAL
        elif 8 <= j <= 11:
            c.fill = HDR_FILL_KRW
        else:
            c.fill = HDR_FILL_DARK
    row += 1

    # 데이터 행
    for code_key, entry in (data or {}).items():
        kor = entry.get('kor', '')
        eng = entry.get('eng', '')

        ws.cell(row, 1, _display_code(code_key)).font = DATA_FONT
        # 라벨 행은 한글 컬럼에만 라벨 표시, 숫자 컬럼은 빈칸
        if _is_label_only(code_key):
            c = ws.cell(row, 2, kor); c.font = LABEL_FONT
            ws.cell(row, 3, eng).font = LABEL_FONT
            row += 1
            continue

        ws.cell(row, 2, kor).font = DATA_FONT
        ws.cell(row, 3, eng).font = DATA_FONT

        local_cur = entry.get('local_value', 0) or 0
        # KRW 당기: 외부에서 override가 있으면(BS 자본=WCE 입력) 그 값 우선
        if override_current_krw is not None and code_key in override_current_krw:
            krw_cur = override_current_krw[code_key] or 0
        else:
            krw_cur = entry.get('value', 0) or 0
        is_krw = (currency or '').upper() == 'KRW'

        # KRW 회사는 패키지에 local 셀이 비어있어도 KRW=KRW이므로 안전한 폴백
        if is_krw and not local_cur and krw_cur:
            local_cur = krw_cur

        if compare_active:
            local_pri = _compute_local_prior(entry, currency, same_q_rates, sheet_name)
            krw_pri, _ = _compute_krw_prior(sheet_name, entry, currency, same_q_rates)
            if is_krw and not local_pri and krw_pri:
                local_pri = krw_pri
        else:
            local_pri = 0
            krw_pri = 0

        # 로컬통화 4컬럼
        for j, v in enumerate([local_cur, local_pri], start=4):
            c = ws.cell(row, j, v if v else None); c.font = DATA_FONT; c.number_format = NUM_FMT
        delta_local = local_cur - local_pri
        c = ws.cell(row, 6, delta_local if delta_local else None); c.font = DATA_FONT; c.number_format = NUM_FMT
        if local_pri:
            c = ws.cell(row, 7, delta_local / abs(local_pri))
            c.font = DATA_FONT; c.number_format = PCT_FMT

        # KRW 4컬럼
        for j, v in enumerate([krw_cur, krw_pri], start=8):
            c = ws.cell(row, j, v if v else None); c.font = DATA_FONT; c.number_format = NUM_FMT
        delta_krw = krw_cur - krw_pri
        c = ws.cell(row, 10, delta_krw if delta_krw else None); c.font = DATA_FONT; c.number_format = NUM_FMT
        if krw_pri:
            c = ws.cell(row, 11, delta_krw / abs(krw_pri))
            c.font = DATA_FONT; c.number_format = PCT_FMT

        # 변동사유 (당기 패키지의 L열 raw)
        reason = entry.get('reason') or ''
        if reason:
            c = ws.cell(row, 12, reason)
            c.font = DATA_FONT
            c.alignment = Alignment(wrap_text=True, vertical='top')

        row += 1

    # 섹션 사이 공백 1행
    row += 1
    return row


def _setup_company_sheet(wb, sheet_name):
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 34
    for col_idx in range(4, 12):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.column_dimensions['L'].width = 45   # 변동사유 (긴 텍스트 가능)
    ws.freeze_panes = 'D4'
    return ws


def build_company_compare(year, files_for_year, same_q_rates, prior_period_label, output_path,
                          aggregated_bs=None):
    """회사별 전기 비교 엑셀 생성.

    files_for_year: list of dicts; 각 dict는 'company', 'extracted'를 포함해야 한다.
                    extracted['sheets'] = {sheet_name → OrderedDict[code → entry]}
                    extracted['currency'] = 통화코드 (예: 'KRW', 'USD')
    same_q_rates:   {currency: {'spot', 'avg'}}
    prior_period_label: '2025-1Q' 등 (표지 표시용)
    output_path: 저장 경로(str/Path)
    aggregated_bs:  agg['sheets']['BS'] — 합산 후 WCE 적용된 BS 데이터.
                    각 코드의 by_company 값을 회사별 BS KRW 당기값에 우선 적용.
                    None이면 패키지 BS 값 그대로 사용.
    """
    wb = Workbook()
    # 기본 시트 제거 (표지를 곧 추가)
    wb.remove(wb.active)

    # 환율 표지/메타용 통화 수집
    currencies_in_use = set()
    for f in files_for_year:
        cur = ((f.get('extracted') or {}).get('currency') or '').upper()
        if cur:
            currencies_in_use.add(cur)

    _build_cover(wb, year, prior_period_label, files_for_year,
                 same_q_rates, currencies_in_use)
    cover = wb['표지']

    used_sheet_names = {'표지'}

    # 회사 시트
    list_row_start = None
    # 표지의 '대상 회사' 헤더 다음 행을 찾아서 회사명 리스트 채우기
    # (위에서 헤더 row를 미리 만든 위치를 다시 알아내기 위해 단순 스캔)
    for r in range(1, cover.max_row + 1):
        if cover.cell(r, 1).value == '대상 회사':
            list_row_start = r + 2   # 헤더 다음
            break

    list_row = list_row_start or (cover.max_row + 2)

    for idx, f in enumerate(files_for_year, 1):
        company = f.get('company') or f.get('extracted', {}).get('company') or f'회사_{idx}'
        extracted = f.get('extracted') or {}
        currency  = (extracted.get('currency') or '').strip().upper() or 'KRW'
        sheets    = extracted.get('sheets') or {}

        sheet_name = _safe_sheet_name(company, used_sheet_names)
        ws = _setup_company_sheet(wb, sheet_name)

        # 시트 상단 메타
        ws.cell(1, 1, company).font = TITLE_FONT
        ws.cell(2, 1, f'당기 {year}  |  전기(동기) {prior_period_label or "-"}  |  통화 {currency}').font = LABEL_FONT
        row = 4

        # BS만: 합산 결과의 by_company 값을 KRW 당기값 override로 사용 (WCE 자본 반영)
        bs_override = None
        if aggregated_bs:
            bs_override = {
                code: (info.get('by_company') or {}).get(company)
                for code, info in aggregated_bs.items()
            }

        for sect_key, (display_name, compare_active) in SECTION_DISPLAY.items():
            data = sheets.get(sect_key)
            if not data:
                continue
            override = bs_override if sect_key == 'BS' else None
            row = _write_section(ws, row, sect_key, display_name, compare_active,
                                 data, currency, same_q_rates,
                                 override_current_krw=override)

        # 표지의 회사 목록에 등록
        cover.cell(list_row, 1, idx).font = DATA_FONT
        cover.cell(list_row, 2, company).font = DATA_FONT
        cover.cell(list_row, 3, currency).font = DATA_FONT
        link_cell = cover.cell(list_row, 4, sheet_name)
        link_cell.font = Font(name='Arial', size=10, color='0563C1', underline='single')
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        list_row += 1

    wb.save(str(output_path))
    return str(output_path)
