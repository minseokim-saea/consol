"""
연결 재무보고 통합 시스템 - Flask 웹 서버
"""

import os
import io
import re
import sys
import time
import uuid
import json
import secrets
import threading
import functools
import traceback
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file, url_for, session, redirect
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.routing import BaseConverter
from filelock import FileLock

from extractor import extract, validate_local_vs_value_signs
from aggregator import aggregate, write_excel
from company_compare_builder import build_company_compare
from package_verify import (verify_wcf_diff, verify_wcf_accounts,
                            verify_wcf_signs, verify_wcf_code_positive,
                            verify_retirement_benefit,
                            verify_cf3_current_portion_new_borrowing,
                            verify_cf_other_transfer,
                            verify_cf4_other_transfer,
                            verify_cf41_other_transfer)
from note_aggregate import (
    extract_l1_borrowings, build_l1_excel,
    extract_l4_loan_facility, build_l4_excel,
    extract_l4_lc, build_l4lc_excel,
    extract_l4_export, build_l4_export_excel,
    extract_l4_guarantees_received, build_l4_guarantees_excel,
    extract_l4_guarantees_provided, build_l4_guarantees_provided_excel,
    extract_l4_lawsuits, build_l4_lawsuits_excel,
    extract_l4_restricted_financial, build_l4_restricted_excel,
    extract_l4_insured_ppe, build_l4_insured_ppe_excel,
    extract_l4_pledged_proceeds, build_l4_pledged_proceeds_excel,
    extract_l4_pledged_assets, build_l4_pledged_assets_excel,
    extract_l4_subsequent_events, build_l4_subsequent_events_excel,
    extract_l4_other_commitments, build_l4_other_commitments_excel,
    extract_a2_securities, build_a2_securities_excel,
    extract_a3_investment_property_pl, build_a3_investment_pl_excel,
    extract_a3_land_value_investment, build_a3_land_investment_excel,
    extract_a3_land_value_ppe, build_a3_land_ppe_excel,
    extract_a4_construction_balance, build_a4_construction_balance_excel,
    extract_a4_construction_profit, build_a4_construction_profit_excel,
    extract_a4_contract_balance, build_a4_contract_balance_excel,
    extract_a5_rou_changes, build_a5_rou_changes_excel,
    extract_a5_lease_pl, build_a5_lease_pl_excel,
    extract_a6_derivatives, build_a6_derivatives_excel,
    extract_a7_equity_method, build_a7_equity_method_excel,
    build_all_in_one_excel,
    extract_l2_verification,
    extract_l2_long_term_borrowings, build_l2_long_term_borrowings_excel,
    extract_l2_debentures, build_l2_debentures_excel,
    extract_l2_maturity_analysis, build_l2_maturity_excel,
    extract_l3_severance_provision, build_l3_severance_excel,
    extract_l3_pension_funds_movement, build_l3_pension_movement_excel,
    extract_l3_pension_breakdown, build_l3_pension_breakdown_excel,
    extract_l3_pension_managers, build_l3_pension_managers_excel,
    extract_l31_dbo_changes, build_l31_dbo_excel,
    extract_l31_plan_asset_changes, build_l31_plan_asset_excel,
    extract_l31_assumptions, build_l31_assumptions_excel,
    extract_l31_sensitivity, build_l31_sensitivity_excel,
    extract_l31_plan_breakdown, build_l31_plan_breakdown_excel,
    extract_l31_plan_managers, build_l31_plan_managers_excel,
    extract_tx_deferred_tax_changes, build_tx_deferred_tax_changes_excel,
    extract_tx_income_tax_breakdown, build_tx_income_tax_breakdown_excel,
    extract_tx_equity_deferred_tax, build_tx_equity_deferred_tax_excel,
    extract_tx_reconciliation, build_tx_reconciliation_excel,
    extract_tx_unrecognized_temp_diff, build_tx_unrecognized_excel,
    extract_tx_loss_carryforward_maturity, build_tx_loss_maturity_excel,
)
from distribute_token import verify_token as verify_dist_token
import totp_util
from concurrent.futures import ThreadPoolExecutor
from wce_schema import (WCE_TABLES, WCE_EQUITY_GROUPS, WCE_ALL_EQUITY_CODES,
                        get_table as wce_get_table,
                        empty_overrides as wce_empty_overrides,
                        to_local_code as wce_to_local_code,
                        to_local_label as wce_to_local_label)
from consol_schema import (load_template as consol_load_template,
                           list_groups as consol_list_groups,
                           get_group as consol_get_group,
                           upsert_group as consol_upsert_group,
                           delete_group as consol_delete_group,
                           effective_companies as consol_effective_companies,
                           get_journal as consol_get_journal,
                           set_journal as consol_set_journal,
                           set_journal_partial as consol_set_journal_partial,
                           delete_journal as consol_delete_journal,
                           get_prior as consol_get_prior,
                           set_prior as consol_set_prior,
                           prior_year_of, prior_period_of,
                           BS_FIELDS as PRIOR_BS_FIELDS,
                           PL_FIELDS as PRIOR_PL_FIELDS)
from consol_engine import (compute as consol_compute,
                           compute_with_rollup as consol_compute_with_rollup,
                           make_journal_template as consol_make_journal_template,
                           parse_journal_excel as consol_parse_journal_excel,
                           validate_balance as consol_validate_balance,
                           validate_codes_present as consol_validate_codes_present,
                           auto_bridge_pl_bs as consol_auto_bridge,
                           write_consolidation_excel as consol_write_excel)
from cf_engine import (compute_cf as cf_compute,
                       load_mapping as cf_load_mapping,
                       save_mapping_v2 as cf_save_mapping_v2,
                       write_cash_worksheet_excel as cf_write_excel,
                       prior_year_4q as cf_prior_year_4q,
                       compute_fund_adjustments_global_sae as cf_fund_adj_global_sae)
import distribute_builder as dbuilder

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB per file
app.config['TEMPLATES_AUTO_RELOAD'] = True  # 템플릿 파일 수정 시 재시작 없이 반영


class _UidConverter(BaseConverter):
    """업로드 파일 uid 전용 컨버터 — uuid4()[:8] 8자리 hex 만 매칭.

    이렇게 좁혀두면 향후 /files/<신규경로> 류 라우트가 변수 매칭에 빨려들어가
    405(Method Not Allowed)로 떨어지는 사고를 막을 수 있다.
    """
    regex = r'[0-9a-fA-F]{8}'


app.url_map.converters['uid'] = _UidConverter

# 서버 재시작 후에도 세션 유지를 위해 SECRET_KEY를 파일에 저장
_KEY_FILE = Path('secret_key.bin')
if _KEY_FILE.exists():
    app.config['SECRET_KEY'] = _KEY_FILE.read_bytes()
else:
    _key = os.urandom(24)
    _KEY_FILE.write_bytes(_key)
    app.config['SECRET_KEY'] = _key

UPLOAD_DIR = Path('uploads')
RESULTS_DIR = Path('results')
JOURNAL_DIR = UPLOAD_DIR / 'journals'           # 업로드된 분개 원본 파일 보관소
STATE_FILE = UPLOAD_DIR / '_state.json'
AUTH_FILE = Path('auth_config.json')
YEARS_FILE = Path('years_config.json')
WCE_FILE = Path('wce_overrides.json')
FX_RATES_FILE = Path('fx_rates.json')
SMTP_CONFIG_FILE = Path('smtp_config.json')     # 메일 발송 SMTP 설정 (관리자가 채움)
OTP_MANUAL_DIR = Path('otp_manual')             # 계정 안내 메일에 첨부할 OTP 등록 매뉴얼(고정 1개)
ALLOWED_EXT = {'.xlsm', '.xlsx', '.xls'}

UPLOAD_DIR.mkdir(exist_ok=True)
RESULTS_DIR.mkdir(exist_ok=True)
JOURNAL_DIR.mkdir(exist_ok=True)

uploaded_files: list[dict] = []

# ─── 파일 잠금 (멀티프로세스·멀티스레드 동시 쓰기 방지) ────────────────────────
# FileLock  : 프로세스 간 배타 잠금 (.lock 파일 기반, Windows/Linux 공통)
# threading.Lock: 동일 프로세스 내 스레드 간 추가 보호
_state_filelock = FileLock(str(STATE_FILE) + '.lock', timeout=10)
_state_threadlock = threading.Lock()
_auth_filelock = FileLock(str(AUTH_FILE) + '.lock', timeout=10)
_years_filelock = FileLock(str(YEARS_FILE) + '.lock', timeout=10)
_wce_filelock = FileLock(str(WCE_FILE) + '.lock', timeout=10)
_fx_rates_filelock = FileLock(str(FX_RATES_FILE) + '.lock', timeout=10)


# ─── 오류 응답 ────────────────────────────────────────────────────────────────

def _json_error(e, status=500):
    """예외 전체(traceback)는 서버 콘솔에만 기록하고, 클라이언트에는 짧은 메시지만 반환.
    내부 경로·파일명·스택 구조가 응답으로 새어나가 정찰 단서가 되는 것을 막는다."""
    print(f'[ERROR] {traceback.format_exc()}', file=sys.stderr)
    return jsonify({'error': str(e)}), status


# ─── 인증 ─────────────────────────────────────────────────────────────────────

def _load_credentials():
    """auth_config.json에서 사용자 정보 로드. 없으면 기본 계정 생성.

    사용자 레코드 형식: {'password': <hash>, 'is_admin': bool}
    이전 형식(문자열 해시)은 자동 마이그레이션.
    """
    if not AUTH_FILE.exists():
        # 널리 알려진 고정 비번(admin123) 대신 매번 무작위 초기 비번을 생성.
        # 비번은 이 콘솔에 단 한 번만 출력되며 파일에는 해시만 저장됨.
        init_pw = secrets.token_urlsafe(12)
        data = {'users': {'admin': {
            'password': generate_password_hash(init_pw),
            'is_admin': True,
        }}}
        with open(AUTH_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
        print('=' * 60)
        print('[인증] 기본 관리자 계정이 생성되었습니다.')
        print(f'[인증]   아이디  : admin')
        print(f'[인증]   초기비번: {init_pw}')
        print('[인증] 이 비밀번호는 지금 한 번만 표시됩니다. 즉시 기록 후')
        print('[인증] 로그인하여 /change-password 에서 변경하세요.')
        print('=' * 60)
        return data['users']

    with open(AUTH_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    users = data.get('users', {})

    # 이전 형식(문자열 해시) → 딕셔너리 마이그레이션
    migrated = False
    for u, v in list(users.items()):
        if isinstance(v, str):
            users[u] = {'password': v, 'is_admin': (u == 'admin'), 'assigned_companies': []}
            migrated = True
        elif isinstance(v, dict) and 'assigned_companies' not in v:
            v['assigned_companies'] = []
            migrated = True
    # 권한그룹(permission_group) 마이그레이션:
    # is_admin=True → 'system_admin', is_admin=False → 'finance_member'
    for u, v in users.items():
        if isinstance(v, dict) and not v.get('permission_group'):
            v['permission_group'] = 'system_admin' if v.get('is_admin') else 'finance_member'
            migrated = True
    if migrated:
        with open(AUTH_FILE, 'w', encoding='utf-8') as f:
            json.dump({'users': users}, f, indent=2)
        print('[인증] 사용자 데이터 형식 업그레이드 완료')
    return users


CREDENTIALS = _load_credentials()


def _save_credentials():
    with _auth_filelock:
        _atomic_write_json(AUTH_FILE, {'users': CREDENTIALS})


def _get_hash(username):
    rec = CREDENTIALS.get(username)
    if isinstance(rec, dict):
        return rec.get('password')
    return rec


def _is_admin(username):
    rec = CREDENTIALS.get(username)
    if isinstance(rec, dict):
        return bool(rec.get('is_admin'))
    return username == 'admin'


# ─── 2단계 인증(TOTP / Google OTP) ──────────────────────────────────────────
# 정책: 전체 강제('all') — 모든 사용자가 최초 로그인 시 OTP 등록 후 이용.
#       'off' 로 바꾸면 자율(설정한 사용자만 2FA 적용).
TWOFA_MODE = 'all'


def _user_rec(username):
    rec = CREDENTIALS.get(username)
    return rec if isinstance(rec, dict) else None


def _twofa_enabled(username):
    """해당 사용자가 2FA를 활성화(시크릿 등록 완료)했는지."""
    rec = _user_rec(username)
    return bool(rec and rec.get('totp_enabled') and rec.get('totp_secret'))


def _twofa_secret(username):
    rec = _user_rec(username)
    return (rec or {}).get('totp_secret')


def _twofa_required(username):
    """이 사용자가 2FA를 반드시 설정해야 하는지(미설정 시 강제 등록 유도)."""
    return TWOFA_MODE == 'all'


@app.before_request
def _enforce_2fa_setup():
    """2FA 강제 등록 대상이 아직 미설정이면 모든 페이지를 설정 화면으로 유도."""
    if not session.get('force_2fa_setup'):
        return
    ep = request.endpoint or ''
    if ep in ('twofa_setup', 'logout', 'static'):
        return
    if request.method in ('GET', 'HEAD'):
        return redirect(url_for('twofa_setup'))
    return jsonify({'error': '2단계 인증(OTP) 설정을 먼저 완료해주세요.'}), 403


@app.context_processor
def _inject_sidebar_perms():
    """모든 템플릿에서 perms/is_admin 자동 사용 가능하게 주입.
    개별 render_template 호출에서 같은 키를 명시 전달하면 그것이 우선됨.
    로그인 안 된 상태(로그인 페이지 등)는 빈 dict 반환.
    """
    uname = session.get('username')
    if not uname:
        return {}
    return {
        'username': uname,
        'is_admin': _is_admin(uname),
        'perms': {
            'users_manage':    _has_permission(uname, 'users.manage'),
            'years_manage':    _has_permission(uname, 'years.manage'),
            'wce_manage':      _has_permission(uname, 'wce.manage'),
            'fx_manage':       _has_permission(uname, 'fx.manage'),
            'consol_compute':  _has_permission(uname, 'consol.compute'),
            'consol_journal':  _has_permission(uname, 'consol.journal'),
            'files_upload':    _has_permission(uname, 'files.upload'),
            'files_delete':    _has_permission(uname, 'files.delete'),
            'files_reanalyze': _has_permission(uname, 'files.reanalyze'),
            'aggregate_run':   _has_permission(uname, 'aggregate.run'),
            'package_verify':  _has_permission(uname, 'package.verify'),
            'note_aggregate':  _has_permission(uname, 'note.aggregate'),
            'cash_compute':    _has_permission(uname, 'cash.compute'),
            'cash_mapping':    _has_permission(uname, 'cash.mapping'),
            'distribute_run':  _has_permission(uname, 'distribute.run'),
            'distribute_admin':_has_permission(uname, 'distribute.admin'),
            'coa_audit':       _has_permission(uname, 'coa.audit'),
        },
    }


def _assigned_companies(username):
    """담당 회사 목록 반환. 관리자이거나 미지정이면 None (무제한)."""
    if _is_admin(username):
        return None
    rec = CREDENTIALS.get(username)
    if not isinstance(rec, dict):
        return None
    lst = rec.get('assigned_companies')
    if not lst:          # 빈 리스트 또는 None → 미지정 = 무제한
        return None
    return list(lst)


def _norm_co(s):
    """회사명 정규화 — 비알파벳·언더스코어 제거 + casefold."""
    import re as _re
    return _re.sub(r'[\W_]+', '', str(s or '').casefold(), flags=_re.UNICODE)


def _peer_companies_in_groups(allowed_companies):
    """사용자의 담당회사들이 속한 모든 연결그룹의 직속 companies를 합쳐서 반환.
    동일 그룹 내 다른 회사에 접근하려고 할 때 사용.
    """
    if not allowed_companies:
        return set()
    allowed_set = {_norm_co(c) for c in allowed_companies}
    from consol_schema import list_groups as _list_groups
    peers = set()
    for g in _list_groups():
        group_companies = g.get('companies') or []
        group_set = {_norm_co(c) for c in group_companies}
        if allowed_set & group_set:    # 본인 담당 회사가 이 그룹에 있으면
            peers |= group_set         # 같은 그룹의 모든 직속 회사를 동료로 추가
    return peers


def _can_access_company(username, company_name):
    """해당 사용자가 company_name 에 접근 가능한지 확인.
    직접 담당이거나, 본인 담당회사가 속한 연결그룹의 동료 회사이면 허용.
    """
    allowed = _assigned_companies(username)
    if allowed is None:
        return True
    target = _norm_co(company_name)
    if any(_norm_co(c) == target for c in allowed):
        return True
    # 같은 연결그룹의 다른 회사도 허용 (연결담당자가 그룹 전체 패키지를 관리할 수 있도록)
    return target in _peer_companies_in_groups(allowed)


PERMISSION_GROUPS_FILE = Path('permission_groups.json')
_PG_CACHE = None

def _load_permission_groups(force=False):
    """권한그룹 정의를 디스크에서 로드. 단순 캐시."""
    global _PG_CACHE
    if _PG_CACHE is not None and not force:
        return _PG_CACHE
    if not PERMISSION_GROUPS_FILE.exists():
        _PG_CACHE = {'definitions': [], 'groups': {}}
        return _PG_CACHE
    try:
        with open(PERMISSION_GROUPS_FILE, 'r', encoding='utf-8') as f:
            _PG_CACHE = json.load(f) or {'definitions': [], 'groups': {}}
    except (json.JSONDecodeError, OSError):
        _PG_CACHE = {'definitions': [], 'groups': {}}
    return _PG_CACHE


def _save_permission_groups(data):
    """권한그룹 정의를 디스크에 저장 + 캐시 갱신."""
    global _PG_CACHE
    # _atomic_write_json 이 이미 ensure_ascii=False, indent=2 를 적용함 — 중복 전달 금지
    _atomic_write_json(PERMISSION_GROUPS_FILE, data)
    _PG_CACHE = data


def _user_permission_group_id(username):
    """사용자의 권한그룹 ID 반환.
    명시되어 있지 않으면 legacy `is_admin` 플래그를 보고 마이그레이션 기본값을 돌려줌."""
    rec = CREDENTIALS.get(username)
    if not isinstance(rec, dict):
        return 'finance_member'
    gid = rec.get('permission_group')
    if gid:
        return gid
    return 'system_admin' if rec.get('is_admin') else 'finance_member'


def _has_permission(username, key):
    """사용자가 주어진 권한 키를 보유했는지 확인."""
    if not username:
        return False
    rec = CREDENTIALS.get(username)
    if not isinstance(rec, dict):
        return False
    # legacy: is_admin=True 는 전체 권한 (시스템관리자 그룹과 동등)
    if rec.get('is_admin'):
        return True
    gid = _user_permission_group_id(username)
    pg = (_load_permission_groups().get('groups') or {}).get(gid) or {}
    return bool((pg.get('perms') or {}).get(key))


def require_permission(key):
    """특정 권한이 필요한 라우트용 데코레이터."""
    def deco(f):
        @functools.wraps(f)
        def decorated(*args, **kwargs):
            if not session.get('logged_in'):
                if request.method in ('GET', 'HEAD'):
                    return redirect(url_for('login'))
                return jsonify({'error': '로그인이 필요합니다.'}), 401
            if not _has_permission(session.get('username'), key):
                msg = f'권한이 부족합니다. (필요 권한: {key})'
                if request.method in ('GET', 'HEAD'):
                    return msg, 403
                return jsonify({'error': msg}), 403
            return f(*args, **kwargs)
        return decorated
    return deco


def _assigned_groups(username):
    """담당 연결그룹 ID 목록. 관리자/미설정이면 빈 리스트(그룹기준 제한 없음)."""
    if _is_admin(username):
        return []
    rec = CREDENTIALS.get(username)
    if not isinstance(rec, dict):
        return []
    return list(rec.get('assigned_groups') or [])


def _can_access_group(username, group_id):
    """연결정산표 조회·실행(및 대시보드·현금정산표 조회·시계열 등) 접근.
    실행/조회는 기능 권한(라우트의 @require_permission)으로만 통제하고, 그룹 단위로는
    제한하지 않는다 — 어느 연결그룹이든 조회·실행 가능.
    쓰기(분개 업로드/삭제)는 _can_manage_group_journal 로 담당 그룹 기준 별도 제한.
    """
    return True


def _can_manage_group_journal(username, group_id):
    """분개 업로드/삭제 권한 — 담당 연결그룹 또는 담당 회사 기준 제한.
    - 시스템관리자: 항상 허용
    - 담당회사·담당그룹 모두 미지정: 허용(무제한)
    - 담당 연결그룹에 지정된 그룹: 허용
    - 담당회사가 그룹 직속 companies 중 하나라도 포함: 허용
    """
    if _is_admin(username):
        return True
    rec = CREDENTIALS.get(username)
    companies = (rec.get('assigned_companies') if isinstance(rec, dict) else None) or []
    groups    = (rec.get('assigned_groups')    if isinstance(rec, dict) else None) or []
    # 둘 다 미지정 = 무제한
    if not companies and not groups:
        return True
    # 담당 그룹 직접 지정
    if group_id in groups:
        return True
    # 회사 기반 (그룹 직속 companies 교집합)
    if companies:
        from consol_schema import get_group as _get_group
        group_companies = (_get_group(group_id) or {}).get('companies') or []
        allowed_set = {_norm_co(c) for c in companies}
        if any(_norm_co(c) in allowed_set for c in group_companies):
            return True
    return False


def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.method in ('GET', 'HEAD'):
                return redirect(url_for('login'))
            return jsonify({'error': '로그인이 필요합니다.'}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        if not _is_admin(session.get('username')):
            if request.method in ('GET', 'HEAD'):
                return '관리자 권한이 필요합니다.', 403
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        return f(*args, **kwargs)
    return decorated


# ─── 로그인 무차별 대입 방어 (간이 rate limit) ─────────────────────────────────
# 동일 계정에 대한 비밀번호 추측을 제한. 내부망·단일 프로세스 기준 in-memory.
_LOGIN_MAX_FAILS = 5      # 집계 창 안에서 허용하는 실패 횟수
_LOGIN_WINDOW = 300       # 실패 횟수 집계 창 (초)
_LOGIN_LOCKOUT = 300      # 초과 시 잠금 시간 (초)
_login_fails: dict[str, list] = {}     # username -> [실패 timestamp, ...]
_login_fails_lock = threading.Lock()


def _login_locked_remaining(username):
    """현재 잠금 상태면 해제까지 남은 초(>=1), 아니면 0."""
    now = time.time()
    with _login_fails_lock:
        fails = [t for t in (_login_fails.get(username) or []) if now - t < _LOGIN_WINDOW]
        _login_fails[username] = fails
        if len(fails) >= _LOGIN_MAX_FAILS:
            return max(int(_LOGIN_LOCKOUT - (now - fails[-1])), 1)
        return 0


def _login_record_fail(username):
    now = time.time()
    with _login_fails_lock:
        fails = [t for t in (_login_fails.get(username) or []) if now - t < _LOGIN_WINDOW]
        fails.append(now)
        _login_fails[username] = fails


def _login_clear_fails(username):
    with _login_fails_lock:
        _login_fails.pop(username, None)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('logged_in'):
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        locked = _login_locked_remaining(username)
        if locked:
            error = f'로그인 시도가 너무 많습니다. 약 {locked}초 후 다시 시도하세요.'
            return render_template('login.html', error=error)
        hashed = _get_hash(username)
        if hashed and check_password_hash(hashed, password):
            _login_clear_fails(username)
            if _twofa_enabled(username):
                # 비밀번호 통과 → 2단계 인증 대기(로그인 미완료)
                session.clear()
                session['pre_2fa_user'] = username
                session['pre_2fa_at'] = time.time()
                return redirect(url_for('login_2fa'))
            # 2FA 미설정 사용자
            session['logged_in'] = True
            session['username'] = username
            if _twofa_required(username):
                # 전체 강제 모드 → 등록 화면으로 유도(등록 완료 전까지 다른 페이지 차단)
                session['force_2fa_setup'] = True
                return redirect(url_for('twofa_setup'))
            return redirect(url_for('index'))
        _login_record_fail(username)
        error = '아이디 또는 비밀번호가 올바르지 않습니다.'
    return render_template('login.html', error=error)


@app.route('/login/2fa', methods=['GET', 'POST'])
def login_2fa():
    """로그인 2단계: OTP 6자리 검증. pre_2fa_user 세션이 있어야 접근 가능."""
    username = session.get('pre_2fa_user')
    if not username:
        return redirect(url_for('login'))
    # 5분 경과 시 만료
    if time.time() - float(session.get('pre_2fa_at') or 0) > 300:
        session.clear()
        return redirect(url_for('login'))
    error = None
    if request.method == 'POST':
        code = request.form.get('otp', '')
        if totp_util.verify(_twofa_secret(username), code):
            session.clear()
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('index'))
        attempts = int(session.get('2fa_attempts') or 0) + 1
        session['2fa_attempts'] = attempts
        if attempts >= 5:
            session.clear()
            return redirect(url_for('login'))
        error = f'인증 코드가 올바르지 않습니다. (남은 시도 {5 - attempts}회)'
    return render_template('login_2fa.html', error=error, username=username)


@app.route('/2fa/setup', methods=['GET', 'POST'])
@login_required
def twofa_setup():
    """OTP 등록(인증 앱에 QR 스캔 후 코드 1회 확인). 강제 등록 대상도 여기로 유도됨."""
    username = session.get('username')
    error = None
    if _twofa_enabled(username):
        # 이미 등록됨 — 재설정은 관리자 초기화 후 가능
        return render_template('twofa_setup.html', already=True, force=False)

    if request.method == 'POST':
        secret = session.get('pending_2fa_secret')
        code = request.form.get('otp', '')
        if not secret:
            error = '세션이 만료되었습니다. 페이지를 새로고침해 다시 시도해주세요.'
        elif totp_util.verify(secret, code):
            rec = _user_rec(username)
            if rec is not None:
                rec['totp_secret'] = secret
                rec['totp_enabled'] = True
                _save_credentials()
            session.pop('pending_2fa_secret', None)
            session.pop('force_2fa_setup', None)
            return redirect(url_for('index'))
        else:
            error = '인증 코드가 올바르지 않습니다. 앱에 표시된 최신 6자리로 다시 입력하세요.'

    # GET 또는 검증 실패 후 재표시 — 동일 시크릿 유지
    secret = session.get('pending_2fa_secret')
    if not secret:
        secret = totp_util.generate_secret()
        session['pending_2fa_secret'] = secret
    uri = totp_util.provisioning_uri(secret, username)
    return render_template('twofa_setup.html', already=False,
                           force=bool(session.get('force_2fa_setup')),
                           secret=secret, otpauth_uri=uri, error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    msg = None
    error = None
    if request.method == 'POST':
        current = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        username = session.get('username')
        hashed = _get_hash(username)
        if not hashed or not check_password_hash(hashed, current):
            error = '현재 비밀번호가 올바르지 않습니다.'
        elif len(new_pw) < 6:
            error = '새 비밀번호는 6자 이상이어야 합니다.'
        elif new_pw != confirm:
            error = '새 비밀번호가 일치하지 않습니다.'
        else:
            CREDENTIALS[username]['password'] = generate_password_hash(new_pw)
            _save_credentials()
            msg = '비밀번호가 변경되었습니다.'
    return render_template('change_password.html', msg=msg, error=error)


# ─── 계정 안내 메일 (SMTP) ──────────────────────────────────────────────────
def _load_smtp_config():
    """smtp_config.json + 환경변수에서 SMTP 설정 로드. host 비어있으면 미설정."""
    cfg = {'host': '', 'port': 587, 'use_tls': True, 'use_ssl': False,
           'username': '', 'password': '', 'from_addr': '',
           'from_name': '연결 재무보고 시스템', 'login_url': ''}
    if SMTP_CONFIG_FILE.exists():
        try:
            with open(SMTP_CONFIG_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f) or {}
            for k in cfg:
                if k in data and data[k] not in (None, ''):
                    cfg[k] = data[k]
        except Exception:
            pass
    import os as _os
    for k, ev in [('host', 'SMTP_HOST'), ('port', 'SMTP_PORT'), ('username', 'SMTP_USER'),
                  ('password', 'SMTP_PASSWORD'), ('from_addr', 'SMTP_FROM'), ('login_url', 'APP_LOGIN_URL')]:
        v = _os.environ.get(ev)
        if v:
            cfg[k] = int(v) if k == 'port' else v
    return cfg


def _smtp_ready():
    return bool(_load_smtp_config().get('host'))


def _otp_manual_file():
    """등록된 OTP 매뉴얼 파일(첫 파일). 없으면 None."""
    if not OTP_MANUAL_DIR.exists():
        return None
    files = [p for p in sorted(OTP_MANUAL_DIR.iterdir()) if p.is_file()]
    return files[0] if files else None


def _smtp_send(cfg, msg):
    """cfg 설정으로 SMTP 연결해 msg 발송. 예외는 호출측에서 처리."""
    import smtplib, ssl
    port = int(cfg.get('port') or (465 if cfg.get('use_ssl') else 587))
    if cfg.get('use_ssl'):
        with smtplib.SMTP_SSL(cfg['host'], port, context=ssl.create_default_context(), timeout=20) as s:
            if cfg.get('username'):
                s.login(cfg['username'], cfg.get('password', ''))
            s.send_message(msg)
    else:
        with smtplib.SMTP(cfg['host'], port, timeout=20) as s:
            if cfg.get('use_tls'):
                s.starttls(context=ssl.create_default_context())
            if cfg.get('username'):
                s.login(cfg['username'], cfg.get('password', ''))
            s.send_message(msg)


def _send_credentials_email(to_addr, username, password):
    """계정 정보(아이디·비번) + OTP 매뉴얼 첨부 메일 발송. (ok, 메시지) 반환."""
    cfg = _load_smtp_config()
    if not cfg.get('host'):
        return False, 'SMTP 미설정 (smtp_config.json 을 채워주세요)'
    if not to_addr:
        return False, '이메일 주소 없음'
    import mimetypes
    from email.message import EmailMessage
    try:
        msg = EmailMessage()
        from_addr = cfg.get('from_addr') or cfg.get('username') or ''
        msg['From'] = (f"{cfg['from_name']} <{from_addr}>" if cfg.get('from_name') else from_addr)
        msg['To'] = to_addr
        msg['Subject'] = '[연결 재무보고 통합 시스템] 계정 정보 안내'
        login_url = cfg.get('login_url') or ''
        body = (
            "안녕하세요.\n\n"
            "연결 재무보고 통합 시스템 계정이 생성되었습니다.\n\n"
            "[ 계정 정보 ]\n"
            f"· 아이디: {username}\n"
            f"· 임시 비밀번호: {password}\n"
            + (f"· 접속 주소: {login_url}\n" if login_url else "")
            + "\n"
            "──────────────────────────────────────\n"
            "⚠ 보안 안내 — 최초 로그인 후 반드시 비밀번호를 변경해 주세요.\n"
            "   · 로그인 → 상단 메뉴 [비밀번호 변경] 에서 변경\n"
            "   · 위 임시 비밀번호는 최초 접속용이므로, 본인만 아는 새 비밀번호로 즉시 바꿔 주세요.\n"
            "──────────────────────────────────────\n\n"
            "2단계 인증(OTP) 등록 방법은 첨부된 매뉴얼을 참고해 주세요.\n"
        )
        msg.set_content(body)
        manual = _otp_manual_file()
        if manual:
            ctype, _ = mimetypes.guess_type(manual.name)
            maintype, subtype = (ctype.split('/', 1) if ctype else ('application', 'octet-stream'))
            msg.add_attachment(manual.read_bytes(), maintype=maintype, subtype=subtype, filename=manual.name)

        _smtp_send(cfg, msg)
        return True, f'{to_addr} 로 발송 완료'
    except Exception as e:
        return False, f'발송 실패: {type(e).__name__}: {e}'


@app.route('/admin/smtp-test', methods=['POST'])
@require_permission('users.manage')
def admin_smtp_test():
    """현재 SMTP 설정으로 지정 주소에 테스트 메일 1통 발송 (설정 확인용)."""
    data = request.get_json(silent=True) or {}
    to_addr = (data.get('email') or '').strip()
    if not to_addr:
        return jsonify({'ok': False, 'error': '받는 이메일 주소를 입력하세요.'}), 400
    cfg = _load_smtp_config()
    if not cfg.get('host'):
        return jsonify({'ok': False, 'error': 'SMTP 미설정 — smtp_config.json 을 먼저 채워주세요.'}), 400
    from email.message import EmailMessage
    try:
        msg = EmailMessage()
        from_addr = cfg.get('from_addr') or cfg.get('username') or ''
        msg['From'] = (f"{cfg['from_name']} <{from_addr}>" if cfg.get('from_name') else from_addr)
        msg['To'] = to_addr
        msg['Subject'] = '[연결 재무보고 통합 시스템] SMTP 테스트 메일'
        msg.set_content('이 메일이 보이면 SMTP 설정이 정상입니다.\n'
                        '계정 안내 메일도 이 설정으로 발송됩니다.')
        _smtp_send(cfg, msg)
        return jsonify({'ok': True, 'msg': f'{to_addr} 로 테스트 메일을 보냈습니다. 받은편지함을 확인하세요.'})
    except Exception as e:
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}'}), 500


@app.route('/admin/otp-manual', methods=['POST'])
@require_permission('users.manage')
def admin_upload_otp_manual():
    """계정 안내 메일에 첨부할 OTP 등록 매뉴얼(고정 1개) 업로드/교체."""
    f = request.files.get('file')
    if not f or not f.filename:
        return redirect(url_for('admin_users', error='매뉴얼 파일이 없습니다.'))
    OTP_MANUAL_DIR.mkdir(exist_ok=True)
    # 기존 매뉴얼 1개만 유지 → 모두 제거 후 새로 저장
    for p in OTP_MANUAL_DIR.iterdir():
        if p.is_file():
            try:
                p.unlink()
            except Exception:
                pass
    safe = os.path.basename(f.filename.replace('\\', '/'))
    f.save(str(OTP_MANUAL_DIR / safe))
    return redirect(url_for('admin_users', msg=f'OTP 매뉴얼 등록 완료: {safe}'))


@app.route('/admin/smtp-config', methods=['POST'])
@require_permission('users.manage')
def admin_save_smtp_config():
    """웹 화면에서 입력한 SMTP 설정을 서버의 smtp_config.json 에 저장.
    비밀번호 칸이 비어 있으면 기존 비밀번호를 유지한다.
    """
    cur = _load_smtp_config()

    def _b(v):
        return str(v).lower() in ('1', 'true', 'on', 'yes')

    try:
        port = int((request.form.get('port') or '587').strip())
    except ValueError:
        port = 587

    pw = request.form.get('password') or ''
    new = {
        'host':      (request.form.get('host') or '').strip(),
        'port':      port,
        'use_tls':   _b(request.form.get('use_tls')),
        'use_ssl':   _b(request.form.get('use_ssl')),
        'username':  (request.form.get('username') or '').strip(),
        'password':  pw if pw else cur.get('password', ''),   # 빈칸이면 기존 유지
        'from_addr': (request.form.get('from_addr') or '').strip(),
        'from_name': (request.form.get('from_name') or '연결 재무보고 통합 시스템').strip(),
        'login_url': (request.form.get('login_url') or '').strip(),
    }
    try:
        with open(SMTP_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(new, f, ensure_ascii=False, indent=2)
        return redirect(url_for('admin_users', msg='메일(SMTP) 설정을 저장했습니다. 테스트 메일로 확인해 보세요.'))
    except Exception as e:
        return redirect(url_for('admin_users', error=f'SMTP 설정 저장 실패: {e}'))


# ─── 관리자 전용 사용자 관리 ────────────────────────────────────────────────

@app.route('/admin/users', methods=['GET'])
@require_permission('users.manage')
def admin_users():
    pg_data = _load_permission_groups()
    pg_groups = pg_data.get('groups') or {}
    # 연결그룹(담당 그룹 지정용) id→name 맵
    consol_groups = [{'id': g.get('id'), 'name': g.get('name', g.get('id'))}
                     for g in consol_list_groups()]
    cg_name = {g['id']: g['name'] for g in consol_groups}
    users_view = []
    for u in sorted(CREDENTIALS.keys()):
        gid = _user_permission_group_id(u)
        rec = CREDENTIALS[u] if isinstance(CREDENTIALS[u], dict) else {}
        agroups = rec.get('assigned_groups', []) or []
        users_view.append({
            'username': u,
            'name': rec.get('name', ''),
            'email': rec.get('email', ''),
            'is_admin': _is_admin(u),
            'is_self': u == session.get('username'),
            'permission_group': gid,
            'permission_group_name': (pg_groups.get(gid) or {}).get('name', gid),
            'assigned_companies': rec.get('assigned_companies', []),
            'assigned_groups': agroups,
            'assigned_group_names': [cg_name.get(x, x) for x in agroups],
            'twofa_enabled': _twofa_enabled(u),
        })
    # 권한그룹 드롭다운 옵션
    group_options = [
        {'id': gid, 'name': g.get('name', gid)}
        for gid, g in pg_groups.items()
    ]
    _otp_m = _otp_manual_file()
    _scfg = _load_smtp_config()
    smtp_cfg_view = {k: _scfg.get(k) for k in
                     ('host', 'port', 'use_tls', 'use_ssl', 'username', 'from_addr', 'from_name', 'login_url')}
    return render_template('admin_users.html', users=users_view,
                           permission_groups=group_options,
                           consol_groups=consol_groups,
                           smtp_ready=_smtp_ready(),
                           smtp_cfg=smtp_cfg_view,
                           smtp_has_password=bool(_scfg.get('password')),
                           otp_manual_name=(_otp_m.name if _otp_m else None),
                           msg=request.args.get('msg'),
                           error=request.args.get('error'))


@app.route('/admin/users/create', methods=['POST'])
@require_permission('users.manage')
def admin_create_user():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    email = (request.form.get('email') or '').strip()
    name = (request.form.get('name') or '').strip()[:50]
    # 신규: 권한그룹 선택 (없으면 finance_member 기본)
    pg_id = (request.form.get('permission_group_create') or '').strip()

    if not re.fullmatch(r'[A-Za-z0-9_\-\.]{3,30}', username):
        return redirect(url_for('admin_users', error='아이디는 3~30자의 영문/숫자/_-. 만 사용 가능합니다.'))
    if len(password) < 6:
        return redirect(url_for('admin_users', error='비밀번호는 6자 이상이어야 합니다.'))
    if username in CREDENTIALS:
        return redirect(url_for('admin_users', error=f'이미 존재하는 아이디입니다: {username}'))

    pg_groups = (_load_permission_groups().get('groups') or {})
    if pg_id not in pg_groups:
        pg_id = 'finance_member'

    # 권한 상승 차단: 시스템관리자 그룹으로의 계정 생성은 관리자(is_admin)만 가능.
    if pg_id == 'system_admin' and not _is_admin(session.get('username')):
        return redirect(url_for('admin_users', error='시스템관리자 계정 생성은 시스템관리자만 할 수 있습니다.'))

    CREDENTIALS[username] = {
        'password': generate_password_hash(password),
        'is_admin': (pg_id == 'system_admin'),
        'permission_group': pg_id,
        'assigned_companies': [],
        'email': email,
        'name': name,
    }
    _save_credentials()
    pg_name = (pg_groups.get(pg_id) or {}).get('name', pg_id)

    # 이메일 입력 시 계정정보(아이디·비번) + OTP 매뉴얼 자동 발송
    mail_note = ''
    if email:
        ok, m = _send_credentials_email(email, username, password)
        mail_note = f' · 메일 {"✅" if ok else "⚠"} {m}'
    return redirect(url_for('admin_users', msg=f'계정 생성 완료: {username} (권한그룹: {pg_name}){mail_note}'))


@app.route('/admin/users/<username>/name', methods=['POST'])
@require_permission('users.manage')
def admin_set_user_name(username):
    """사용자 표시 이름/메모 저장 (누가 쓰는 아이디인지 식별용)."""
    if username not in CREDENTIALS:
        return redirect(url_for('admin_users', error='존재하지 않는 계정입니다.'))
    rec = CREDENTIALS[username]
    if not isinstance(rec, dict):
        return redirect(url_for('admin_users', error='잘못된 계정 형식입니다.'))
    name = (request.form.get('name') or '').strip()[:50]
    rec['name'] = name
    _save_credentials()
    return redirect(url_for('admin_users', msg=f'{username} 이름 저장: {name or "(비움)"}'))


@app.route('/admin/users/<username>/delete', methods=['POST'])
@require_permission('users.manage')
def admin_delete_user(username):
    if username == session.get('username'):
        return redirect(url_for('admin_users', error='자기 자신은 삭제할 수 없습니다.'))
    if username not in CREDENTIALS:
        return redirect(url_for('admin_users', error='존재하지 않는 계정입니다.'))
    # 최소 1명의 관리자는 유지
    if _is_admin(username):
        admin_count = sum(1 for u in CREDENTIALS if _is_admin(u))
        if admin_count <= 1:
            return redirect(url_for('admin_users', error='마지막 관리자는 삭제할 수 없습니다.'))
    del CREDENTIALS[username]
    _save_credentials()
    return redirect(url_for('admin_users', msg=f'계정 삭제됨: {username}'))


@app.route('/admin/users/<username>/reset', methods=['POST'])
@require_permission('users.manage')
def admin_reset_password(username):
    new_pw = request.form.get('new_password', '')
    if username not in CREDENTIALS:
        return redirect(url_for('admin_users', error='존재하지 않는 계정입니다.'))
    if len(new_pw) < 6:
        return redirect(url_for('admin_users', error='비밀번호는 6자 이상이어야 합니다.'))
    CREDENTIALS[username]['password'] = generate_password_hash(new_pw)
    _save_credentials()
    return redirect(url_for('admin_users', msg=f'{username} 비밀번호가 재설정되었습니다.'))


@app.route('/admin/users/<username>/reset-2fa', methods=['POST'])
@require_permission('users.manage')
def admin_reset_2fa(username):
    """사용자의 2FA(OTP) 초기화 — 기기 분실 등. 다음 로그인 시 재등록하게 됨."""
    rec = _user_rec(username)
    if rec is None:
        return redirect(url_for('admin_users', error='존재하지 않는 계정입니다.'))
    rec.pop('totp_secret', None)
    rec['totp_enabled'] = False
    _save_credentials()
    return redirect(url_for('admin_users',
                            msg=f'{username}의 2FA(OTP)가 초기화되었습니다. 다음 로그인 시 재등록합니다.'))


@app.route('/admin/users/<username>/toggle-admin', methods=['POST'])
@require_permission('users.manage')
def admin_toggle_admin(username):
    # 권한 상승 차단: 관리자 권한 부여/해제는 시스템관리자(is_admin)만 가능.
    # users.manage 만 가진 계정이 다른 계정을 관리자로 승격시키는 우회로를 막는다.
    if not _is_admin(session.get('username')):
        return redirect(url_for('admin_users', error='관리자 권한 변경은 시스템관리자만 할 수 있습니다.'))
    if username == session.get('username'):
        return redirect(url_for('admin_users', error='자기 자신의 권한은 변경할 수 없습니다.'))
    if username not in CREDENTIALS:
        return redirect(url_for('admin_users', error='존재하지 않는 계정입니다.'))
    rec = CREDENTIALS[username]
    if isinstance(rec, str):
        rec = {'password': rec, 'is_admin': False}
    new_val = not bool(rec.get('is_admin'))
    # 마지막 관리자 보호
    if not new_val:
        admin_count = sum(1 for u in CREDENTIALS if _is_admin(u))
        if admin_count <= 1:
            return redirect(url_for('admin_users', error='마지막 관리자의 권한은 해제할 수 없습니다.'))
    rec['is_admin'] = new_val
    CREDENTIALS[username] = rec
    _save_credentials()
    label = '부여' if new_val else '해제'
    return redirect(url_for('admin_users', msg=f'{username} 관리자 권한 {label}'))


@app.route('/admin/users/<username>/companies', methods=['GET'])
@require_permission('users.manage')
def admin_get_user_companies(username):
    if username not in CREDENTIALS:
        return jsonify({'error': '존재하지 않는 계정입니다.'}), 404
    rec = CREDENTIALS[username]
    return jsonify({'companies': rec.get('assigned_companies', []) if isinstance(rec, dict) else []})


@app.route('/admin/users/<username>/companies', methods=['POST'])
@require_permission('users.manage')
def admin_set_user_companies(username):
    if username not in CREDENTIALS:
        return jsonify({'error': '존재하지 않는 계정입니다.'}), 404
    if _is_admin(username):
        return jsonify({'error': '관리자 계정에는 담당 회사를 지정할 수 없습니다.'}), 400
    data = request.get_json(silent=True) or {}
    companies = [str(c).strip() for c in data.get('companies', []) if str(c).strip()]
    rec = CREDENTIALS[username]
    if not isinstance(rec, dict):
        rec = {'password': rec, 'is_admin': False, 'assigned_companies': []}
    rec['assigned_companies'] = companies
    CREDENTIALS[username] = rec
    _save_credentials()
    return jsonify({'ok': True, 'companies': companies})


@app.route('/admin/users/<username>/groups', methods=['GET'])
@require_permission('users.manage')
def admin_get_user_groups(username):
    if username not in CREDENTIALS:
        return jsonify({'error': '존재하지 않는 계정입니다.'}), 404
    rec = CREDENTIALS[username]
    return jsonify({'groups': rec.get('assigned_groups', []) if isinstance(rec, dict) else []})


@app.route('/admin/users/<username>/groups', methods=['POST'])
@require_permission('users.manage')
def admin_set_user_groups(username):
    """사용자의 담당 연결그룹 지정 — 해당 그룹의 연결실행·조회·분개업로드 권한 부여."""
    if username not in CREDENTIALS:
        return jsonify({'error': '존재하지 않는 계정입니다.'}), 404
    if _is_admin(username):
        return jsonify({'error': '관리자 계정은 모든 그룹에 접근 가능하여 담당 그룹 지정이 불필요합니다.'}), 400
    data = request.get_json(silent=True) or {}
    valid_ids = {g.get('id') for g in consol_list_groups()}
    group_ids = [str(g).strip() for g in data.get('groups', []) if str(g).strip() in valid_ids]
    rec = CREDENTIALS[username]
    if not isinstance(rec, dict):
        rec = {'password': rec, 'is_admin': False, 'assigned_companies': []}
    rec['assigned_groups'] = group_ids
    CREDENTIALS[username] = rec
    _save_credentials()
    return jsonify({'ok': True, 'groups': group_ids})


@app.route('/admin/all-companies', methods=['GET'])
@require_permission('users.manage')
def admin_all_companies():
    """시스템에 등록된 모든 회사명 목록 (업로드 파일 기준)."""
    seen = {}
    for f in uploaded_files:
        co = (f.get('company') or '').strip()
        if co:
            seen[_norm_company_name(co)] = co
    # 회사 마스터(업로드 대상 회사목록)도 포함 — 비활성 회사까지 전부
    for n in _company_required_names(active_only=False):
        seen.setdefault(_norm_company_name(n), n)
    return jsonify({'companies': sorted(seen.values())})


# ─── 권한그룹 관리 ───────────────────────────────────────────────────────────

@app.route('/admin/permission-groups')
@require_permission('users.manage')
def admin_permission_groups_page():
    """권한그룹 관리 페이지."""
    return render_template('admin_permission_groups.html',
                           username=session.get('username'),
                           is_admin=_is_admin(session.get('username')))


@app.route('/admin/permission-groups/data', methods=['GET'])
@require_permission('users.manage')
def admin_permission_groups_list():
    """권한그룹 정의 + 사용자 매핑 조회."""
    data = _load_permission_groups()
    # 그룹별 소속 사용자 수
    user_count_by_group = {}
    for u, rec in CREDENTIALS.items():
        if not isinstance(rec, dict):
            continue
        gid = _user_permission_group_id(u)
        user_count_by_group[gid] = user_count_by_group.get(gid, 0) + 1
    return jsonify({
        'ok': True,
        'definitions': data.get('definitions', []),
        'groups': data.get('groups', {}),
        'user_count_by_group': user_count_by_group,
    })


@app.route('/admin/permission-groups', methods=['POST'])
@require_permission('users.manage')
def admin_permission_groups_create():
    """신규 권한그룹 생성. body: {id, name, perms: {key: bool}}."""
    # 권한 체계 자체(권한그룹 정의)의 변경은 시스템관리자만 가능 — users.manage 계정이
    # 자기 그룹 권한을 임의로 켜서 전체 기능을 얻는 권한 상승을 차단.
    if not _is_admin(session.get('username')):
        return jsonify({'error': '권한그룹 관리는 시스템관리자만 할 수 있습니다.'}), 403
    body = request.get_json(silent=True) or {}
    gid = (body.get('id') or '').strip()
    name = (body.get('name') or '').strip()
    if not gid or not name:
        return jsonify({'error': 'id, name 은 필수입니다.'}), 400
    if not gid.replace('_', '').isalnum():
        return jsonify({'error': 'id 는 영문/숫자/언더스코어만 허용됩니다.'}), 400
    data = _load_permission_groups(force=True)
    groups = data.get('groups') or {}
    if gid in groups:
        return jsonify({'error': '이미 존재하는 ID 입니다.'}), 400
    # 정의된 권한 키만 받음 — 그 외는 무시
    valid_keys = {d['key'] for d in (data.get('definitions') or [])}
    perms_in = body.get('perms') or {}
    perms = {k: bool(perms_in.get(k)) for k in valid_keys}
    groups[gid] = {'name': name, 'system': False, 'perms': perms}
    data['groups'] = groups
    _save_permission_groups(data)
    return jsonify({'ok': True, 'id': gid})


@app.route('/admin/permission-groups/<gid>', methods=['PUT', 'POST'])
@require_permission('users.manage')
def admin_permission_groups_update(gid):
    """권한그룹 수정. body: {name?, perms?}. system 그룹은 perms 변경 가능, name 변경 가능."""
    if not _is_admin(session.get('username')):
        return jsonify({'error': '권한그룹 관리는 시스템관리자만 할 수 있습니다.'}), 403
    body = request.get_json(silent=True) or {}
    data = _load_permission_groups(force=True)
    groups = data.get('groups') or {}
    if gid not in groups:
        return jsonify({'error': '존재하지 않는 권한그룹'}), 404
    g = groups[gid]
    if 'name' in body:
        new_name = (body.get('name') or '').strip()
        if not new_name:
            return jsonify({'error': '이름은 비울 수 없습니다.'}), 400
        g['name'] = new_name
    if 'perms' in body:
        valid_keys = {d['key'] for d in (data.get('definitions') or [])}
        perms_in = body.get('perms') or {}
        g['perms'] = {k: bool(perms_in.get(k)) for k in valid_keys}
    groups[gid] = g
    data['groups'] = groups
    _save_permission_groups(data)
    return jsonify({'ok': True})


@app.route('/admin/permission-groups/<gid>', methods=['DELETE'])
@require_permission('users.manage')
def admin_permission_groups_delete(gid):
    """권한그룹 삭제. system 그룹은 삭제 불가, 사용자가 매핑되어 있어도 불가."""
    if not _is_admin(session.get('username')):
        return jsonify({'error': '권한그룹 관리는 시스템관리자만 할 수 있습니다.'}), 403
    data = _load_permission_groups(force=True)
    groups = data.get('groups') or {}
    if gid not in groups:
        return jsonify({'error': '존재하지 않는 권한그룹'}), 404
    if groups[gid].get('system'):
        return jsonify({'error': '시스템 기본 권한그룹은 삭제할 수 없습니다.'}), 400
    in_use = [u for u, rec in CREDENTIALS.items()
              if isinstance(rec, dict) and _user_permission_group_id(u) == gid]
    if in_use:
        return jsonify({
            'error': f'이 권한그룹을 사용 중인 사용자가 있어 삭제할 수 없습니다 ({len(in_use)}명). '
                     '먼저 사용자들의 권한그룹을 변경하세요.',
            'in_use_users': in_use,
        }), 400
    del groups[gid]
    data['groups'] = groups
    _save_permission_groups(data)
    return jsonify({'ok': True})


@app.route('/admin/users/<username>/permission-group', methods=['POST'])
@require_permission('users.manage')
def admin_set_user_permission_group(username):
    """사용자의 권한그룹 변경."""
    if username not in CREDENTIALS:
        return jsonify({'error': '존재하지 않는 계정입니다.'}), 404
    body = request.get_json(silent=True) or {}
    new_gid = (body.get('permission_group') or '').strip()
    pg_data = _load_permission_groups()
    if new_gid not in (pg_data.get('groups') or {}):
        return jsonify({'error': '존재하지 않는 권한그룹'}), 400
    # 권한 상승 차단:
    #  - 비관리자는 자기 자신의 권한그룹을 바꿀 수 없음(자가 승격 방지).
    #  - 시스템관리자 그룹으로의 변경은 관리자(is_admin)만 가능.
    caller_is_admin = _is_admin(session.get('username'))
    if not caller_is_admin and username == session.get('username'):
        return jsonify({'error': '자기 자신의 권한그룹은 변경할 수 없습니다.'}), 403
    if new_gid == 'system_admin' and not caller_is_admin:
        return jsonify({'error': '시스템관리자 권한 부여는 시스템관리자만 할 수 있습니다.'}), 403
    rec = CREDENTIALS[username]
    if not isinstance(rec, dict):
        rec = {'password': rec, 'is_admin': False, 'assigned_companies': []}
    # 시스템관리자 그룹을 떠나는 경우, 최소 1명 보장
    cur_gid = _user_permission_group_id(username)
    if cur_gid == 'system_admin' and new_gid != 'system_admin':
        admin_count = sum(
            1 for u, r in CREDENTIALS.items()
            if isinstance(r, dict) and _user_permission_group_id(u) == 'system_admin'
        )
        if admin_count <= 1:
            return jsonify({'error': '시스템관리자가 최소 1명은 유지되어야 합니다.'}), 400
    rec['permission_group'] = new_gid
    # legacy is_admin 동기화: system_admin이면 True, 그 외 False
    rec['is_admin'] = (new_gid == 'system_admin')
    CREDENTIALS[username] = rec
    _save_credentials()
    return jsonify({'ok': True, 'permission_group': new_gid})


# ─── WCE 본사 입력 (132행 이후 6개 테이블) ───────────────────────────────────

@app.route('/admin/wce')
@require_permission('wce.manage')
def admin_wce_index():
    """연도/회사 선택 화면. 입력 상태(입력완료/미입력)를 함께 표시.
    담당회사 제한이 있는 사용자에게는 본인 담당 회사 + 동일 연결그룹 동료 회사만 노출."""
    year = request.args.get('year') or YEARS_DATA.get('default')
    if not _valid_year(year):
        year = YEARS_DATA.get('default')

    uname = session.get('username')

    # 해당 연도 업로드된 회사 목록 (담당회사 제한 적용)
    companies = sorted({
        (f.get('company') or '').strip()
        for f in uploaded_files
        if f.get('year') == year and (f.get('company') or '').strip()
        and _can_access_company(uname, f.get('company') or '')
    })

    # 입력 상태 + 미입력 셀 카운트
    wce_data = _load_wce()
    rows = []
    for co in companies:
        rec = wce_data.get(_wce_key(year, co))
        # 미입력 검출용: 현재 저장된 환산값 + 로컬값
        local_info = _wce_local_full_for(year, co)
        current_tables = (rec or {}).get('tables') or wce_empty_overrides()
        missing = _compute_wce_missing(local_info['local'], local_info['fx_avg'], current_tables)
        rows.append({
            'company': co,
            'has_data': bool(rec),
            'has_local': bool(local_info['local']),
            'currency': local_info['currency'],
            'updated_at': (rec or {}).get('updated_at', ''),
            'updated_by': (rec or {}).get('updated_by', ''),
            'missing_count': missing['count'],
        })

    return render_template('wce_list.html',
                           year=year,
                           years=YEARS_DATA['years'],
                           rows=rows,
                           username=session.get('username'))


@app.route('/admin/wce/<year>/<path:company>', methods=['GET'])
@require_permission('wce.manage')
def admin_wce_edit(year, company):
    if not _valid_year(year):
        return f'유효하지 않은 결산기간: {year}', 400
    if not _can_access_company(session.get('username'), company):
        return f'{company} 회사 WCE에 접근 권한이 없습니다.', 403
    overrides = _get_wce_for(year, company)
    auto_cells_set = {f'{tid}:{code}:{rk}' for (tid, code, rk) in overrides.get('auto_cells', [])}
    missing_set = {f'{c["table_id"]}:{c["code"]}:{c["row_key"]}' for c in overrides['missing']['cells']}
    return render_template('wce_input.html',
                           year=year, company=company,
                           tables=WCE_TABLES,
                           values=overrides['tables'],
                           summary=overrides['summary'],
                           meta=overrides['meta'],
                           is_first_year=overrides['is_first_year'],
                           is_new_company=overrides.get('is_new_company', False),
                           prior_meta=overrides['prior_meta'],
                           auto_cells_set=auto_cells_set,
                           auto_re_meta=overrides['auto_re_meta'],
                           local_lookup=overrides['local_lookup'],
                           currency=overrides['currency'],
                           fx_avg=overrides['fx_avg'],
                           missing=overrides['missing'],
                           missing_set=missing_set,
                           has_local=overrides['has_local'],
                           username=session.get('username'))


@app.route('/admin/wce/<year>/<path:company>', methods=['POST'])
@require_permission('wce.manage')
def admin_wce_save(year, company):
    if not _valid_year(year):
        return jsonify({'error': '유효하지 않은 결산기간'}), 400
    if not _can_access_company(session.get('username'), company):
        return jsonify({'error': f'{company} 회사 WCE에 접근 권한이 없습니다.'}), 403
    payload = request.get_json(silent=True) or {}

    def _to_num(v):
        if v in (None, ''):
            return 0
        try:
            return float(str(v).replace(',', '').strip())
        except Exception:
            return 0

    # 테이블 값 정제 (스키마에 정의된 코드/행만 저장)
    clean_tables = {}
    raw_tables = payload.get('tables') or {}
    is_first_year = (year == WCE_FIRST_YEAR)
    prior_year = _prior_q4_period(year)

    # 신규 편입 회사 감지 (전년 패키지 파일 유무)
    target_norm = _norm_company_name(company)
    is_new_company = False
    if not is_first_year and prior_year:
        is_new_company = not any(
            f.get('year') == prior_year
            and _norm_company_name(f.get('company', '')) == target_norm
            for f in uploaded_files
        )

    prior_endings_by_table = {}
    if not is_first_year and prior_year and not is_new_company:
        prior_rec = _load_wce().get(_wce_key(prior_year, company)) or {}
        prior_tables = prior_rec.get('tables') or {}
        for t in WCE_TABLES:
            prior_endings_by_table[str(t['id'])] = _compute_table_ending(
                t, prior_tables.get(str(t['id'])) or {}
            )

    # 5번 이익잉여금의 자동 RE 셀 (당기순이익/보험수리적손익/R/E조정) — 항상 추출값으로 덮어쓰기
    auto_re = _wce_auto_re_cells_for(year, company)

    for t in WCE_TABLES:
        tid = str(t['id'])
        col_codes = {c['code'] for c in t['columns']}
        row_keys = {r['key'] for r in t['rows']}
        clean = {}
        for code, row_dict in (raw_tables.get(tid) or {}).items():
            if code not in col_codes:
                continue
            clean[code] = {
                k: _to_num(v) for k, v in (row_dict or {}).items()
                if k in row_keys
            }
            # 비-첫해 기초금액: 자동값(prior Q4 기말)이 의미있을 때만 덮어쓰기
            # → 자동값 0이면 사용자 입력값 보존 (전년 데이터 비어있는 셀)
            if not is_first_year:
                auto_beg = (prior_endings_by_table.get(tid) or {}).get(code, 0)
                if auto_beg:
                    clean[code]['기초금액'] = auto_beg

            # 5번 테이블 자동 RE 셀은 항상 패키지 추출값으로 덮어쓰기
            if tid == '5' and code in auto_re:
                for auto_row_key, auto_val in auto_re[code].items():
                    if auto_row_key in row_keys:
                        clean[code][auto_row_key] = auto_val
        # 5번 이익잉여금 특별 처리:
        #   3500104 기초 = 전년 3500104 기말 + 전년 3500105 기말 (합산값이 있을 때만)
        #   3500105 기초 = 0 (Current Net Income은 매년 리셋)
        if not is_first_year and tid == '5':
            t5_endings = prior_endings_by_table.get('5') or {}
            merged_104 = (t5_endings.get('3500104', 0) or 0) + (t5_endings.get('3500105', 0) or 0)
            if merged_104:
                clean.setdefault('3500104', {row['key']: 0 for row in t['rows']})['기초금액'] = merged_104
            clean.setdefault('3500105', {row['key']: 0 for row in t['rows']})['기초금액'] = 0
        clean_tables[tid] = clean

    # 저장
    data = _load_wce()
    data[_wce_key(year, company)] = {
        'year': year,
        'company': company,
        'tables': clean_tables,
        'summary': {},
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'updated_by': session.get('username') or '',
    }
    _save_wce(data)
    return jsonify({'ok': True, 'updated_at': data[_wce_key(year, company)]['updated_at']})


@app.route('/admin/wce/<year>/<path:company>', methods=['DELETE'])
@require_permission('wce.manage')
def admin_wce_delete(year, company):
    if not _can_access_company(session.get('username'), company):
        return jsonify({'error': f'{company} 회사 WCE에 접근 권한이 없습니다.'}), 403
    data = _load_wce()
    key = _wce_key(year, company)
    if key in data:
        del data[key]
        _save_wce(data)
    return jsonify({'ok': True})


# ─── WCE 회사별 합산 ──────────────────────────────────────────────────────────

@app.route('/admin/wce/aggregate')
@require_permission('wce.manage')
def admin_wce_aggregate_page():
    """WCE 회사별 합산 페이지 (HTML)."""
    years_list = YEARS_DATA.get('years', [])
    year = (request.args.get('year') or '').strip()
    if not year or not _valid_year(year):
        year = YEARS_DATA.get('default') or (years_list[0] if years_list else '')
    return render_template(
        'wce_aggregate.html',
        year=year,
        years=years_list,
        username=session.get('username'),
    )


@app.route('/admin/wce/aggregate/data')
@require_permission('wce.manage')
def admin_wce_aggregate_data():
    """WCE 합산 데이터 (JSON).
    Query params:
      year=YYYY-NQ   (필수)
      ids=<JSON list of file ids>  (선택; 미지정 시 해당 연도 전체)

    반환:
      {
        'year': str,
        'companies': [{'id', 'company', 'currency', 'fx_avg', 'has_wce'}, ...],
        'tables': [
          {
            'id': 1, 'title_ko': '자본금', 'title_en': '...',
            'columns': [{'code','name'}],
            'rows': [{'key','name_en'}],
            'cells': {
              '<code>::<row_key>': {
                'by_company': {'<company>': {'local': float, 'krw': float}},
                'total_krw': float,
              }
            },
            'ending': {  # 자동 계산된 기말금액 (환산효과 제외)
              '<code>': {
                'by_company': {'<company>': {'local': float, 'krw': float}},
                'total_krw': float,
              }
            },
          },
          ...
        ]
      }
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효하지 않은 기간'}), 400

    ids_arg = request.args.get('ids')
    selected_ids = None
    if ids_arg:
        try:
            selected_ids = set(json.loads(ids_arg))
        except Exception:
            selected_ids = None

    # 대상 회사 결정 (중복 제거 — 같은 회사 다중 업로드 시 최신만)
    uname = session.get('username')
    seen = set()
    targets = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        if selected_ids is not None and f.get('id') not in selected_ids:
            continue
        company = f.get('company')
        if not company:
            continue
        # 담당회사 제한 적용 (연결담당자는 본인 그룹 회사만)
        if not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        ex = f.get('extracted') or {}
        targets.append({
            'id': f.get('id'),
            'company': company,
            'currency': ex.get('currency') or 'KRW',
            'fx_avg': ex.get('fx_avg') or 1.0,
        })
    targets.sort(key=lambda x: x['company'])

    # 회사별 WCE 데이터 및 로컬값을 한 번씩만 미리 캐싱
    wce_all = _load_wce()
    company_data = {}
    for co in targets:
        cname = co['company']
        rec = _get_wce_for(year, cname)  # 자동 채움 + 기초 자동 등 모두 반영된 값
        local_info = _wce_local_full_for(year, cname)
        company_data[cname] = {
            'tables': rec.get('tables') or {},
            'local': local_info.get('local') or {},
            'has_wce': bool(wce_all.get(_wce_key(year, cname))),
        }

    # 회사별 has_wce 표시
    for co in targets:
        co['has_wce'] = company_data[co['company']]['has_wce']

    # 각 테이블 셀별 집계
    out_tables = []
    for t in WCE_TABLES:
        tid = str(t['id'])
        cells = {}
        for col in t['columns']:
            code = col['code']
            for row in t['rows']:
                rk = row['key']
                cell_key = f'{code}::{rk}'
                by_company = {}
                total_krw = 0.0
                for co in targets:
                    cname = co['company']
                    cd = company_data[cname]
                    krw = ((cd['tables'].get(tid) or {}).get(code) or {}).get(rk, 0) or 0
                    local_v = _lookup_local(cd['local'], int(tid), code, rk)
                    by_company[cname] = {
                        'local': float(local_v or 0),
                        'krw':   float(krw or 0),
                    }
                    total_krw += float(krw or 0)
                cells[cell_key] = {
                    'by_company': by_company,
                    'total_krw': total_krw,
                }

        # 기말금액 자동 계산 (환산효과 제외)
        ending = {}
        for col in t['columns']:
            code = col['code']
            ending_co = {}
            ending_total = 0.0
            for co in targets:
                cname = co['company']
                local_sum = 0.0
                krw_sum = 0.0
                for row in t['rows']:
                    if row['key'] == '환산효과':
                        continue
                    cell = cells.get(f'{code}::{row["key"]}', {})
                    bc = (cell.get('by_company') or {}).get(cname, {})
                    local_sum += bc.get('local', 0) or 0
                    krw_sum   += bc.get('krw',   0) or 0
                ending_co[cname] = {'local': local_sum, 'krw': krw_sum}
                ending_total += krw_sum
            ending[code] = {
                'by_company': ending_co,
                'total_krw':  ending_total,
            }

        out_tables.append({
            'id':      t['id'],
            'title_ko': t['title_ko'],
            'title_en': t['title_en'],
            'columns':  t['columns'],
            'rows':     t['rows'],
            'cells':    cells,
            'ending':   ending,
        })

    return jsonify({
        'year': year,
        'companies': targets,
        'tables': out_tables,
    })


# ─── 패키지 검증 ────────────────────────────────────────────────────────────

@app.route('/admin/package-verify')
@require_permission('package.verify')
def admin_package_verify_page():
    """패키지 검증 페이지 (HTML). 향후 다양한 검증 항목 카드 형태로 확장."""
    years_list = YEARS_DATA.get('years', [])
    year = (request.args.get('year') or '').strip()
    if not year or not _valid_year(year):
        year = YEARS_DATA.get('default') or (years_list[0] if years_list else '')
    return render_template(
        'package_verify.html',
        year=year,
        years=years_list,
        username=session.get('username'),
    )


@app.route('/admin/package-verify/wcf-diff')
@require_permission('package.verify')
def admin_package_verify_wcf_diff():
    """WCF 시트 Diff(N열) 검증 결과 JSON.

    Diff != 0 인 행이 있는 회사 + WCF 시트 누락/오류 회사만 반환.
    Diff가 모두 0인 회사는 결과에서 제외 (정상).
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    # 같은 회사 다중 업로드는 최신 1개만
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    # 병렬 검증 — I/O 바운드(openpyxl zip 압축해제)라 ThreadPool로 효과적
    def _verify_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        currency = (f.get('extracted') or {}).get('currency')
        ver = verify_wcf_diff(path)
        dt = _time.time() - t0
        print(f'[패키지 검증] {company} - {dt:.1f}초'
              + (f' (ERROR: {ver.get("error")})' if ver.get('error') else ''),
              flush=True)

        if ver.get('error'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'error', 'message': ver['error'], 'rows': [],
            }
        if not ver.get('sheet_found'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'no_sheet', 'message': 'WCF 시트 없음', 'rows': [],
            }
        rows = ver.get('rows') or []
        if not rows:
            return None   # 정상: 결과에서 제외
        rows_missing_reason = sum(1 for r in rows if not (r.get('o') or '').strip())
        return {
            'company': company, 'currency': currency, 'file_id': f.get('id'),
            'status': 'diff' if rows_missing_reason == 0 else 'diff_no_reason',
            'rows_count': len(rows),
            'rows_missing_reason': rows_missing_reason,
            'rows': rows,
        }

    valid_files = [f for f in files if f.get('path') and f.get('company')]
    scanned = len(valid_files)

    # max_workers: 파일 수와 8 중 작은 값. 너무 크면 디스크 I/O 경합
    workers = min(8, max(1, scanned)) if scanned else 1
    t_total = _time.time()
    print(f'[패키지 검증] WCF Diff 시작: {scanned}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        all_results = list(ex.map(_verify_one, valid_files))
    results = [r for r in all_results if r is not None]
    print(f'[패키지 검증] 완료: {_time.time()-t_total:.1f}초, 특이사항 {len(results)}개사', flush=True)

    # 정렬: error → no_sheet → diff_no_reason → diff → 회사명순
    status_rank = {'error': 0, 'no_sheet': 1, 'diff_no_reason': 2, 'diff': 3}
    results.sort(key=lambda r: (status_rank.get(r.get('status'), 9), r.get('company') or ''))

    return jsonify({
        'year': year,
        'scanned': scanned,
        'with_issues': len(results),
        'companies': results,
    })


@app.route('/admin/package-verify/wcf-accounts')
@require_permission('package.verify')
def admin_package_verify_wcf_accounts():
    """WCF 시트 J/M 계정 입력 오류 검증 결과 JSON.

    J(=Adjustment expense)에 수익코드(41/44/46) 또는
    M(=Adjustment revenue)에 비용코드(42/43/45/48, 5xxx) 가
    잘못 입력된 행을 회사별로 집계.
    오류 행이 없는 회사는 결과에서 제외 (WCF 시트 누락/오류 회사는 포함).
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    def _verify_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        currency = (f.get('extracted') or {}).get('currency')
        ver = verify_wcf_accounts(path)
        dt = _time.time() - t0
        print(f'[패키지 검증/계정] {company} - {dt:.1f}초'
              + (f' (ERROR: {ver.get("error")})' if ver.get('error') else ''),
              flush=True)

        if ver.get('error'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'error', 'message': ver['error'], 'rows': [],
            }
        if not ver.get('sheet_found'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'no_sheet', 'message': 'WCF 시트 없음', 'rows': [],
            }
        rows = ver.get('rows') or []
        if not rows:
            return None
        j_wrong = sum(1 for r in rows if r.get('side') == 'J')
        m_wrong = sum(1 for r in rows if r.get('side') == 'M')
        return {
            'company': company, 'currency': currency, 'file_id': f.get('id'),
            'status': 'account_misplaced',
            'rows_count': len(rows),
            'j_wrong': j_wrong,    # J(비용)에 수익코드
            'm_wrong': m_wrong,    # M(수익)에 비용코드
            'rows': rows,
        }

    valid_files = [f for f in files if f.get('path') and f.get('company')]
    scanned = len(valid_files)
    workers = min(8, max(1, scanned)) if scanned else 1
    t_total = _time.time()
    print(f'[패키지 검증/계정] 시작: {scanned}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        all_results = list(ex.map(_verify_one, valid_files))
    results = [r for r in all_results if r is not None]
    print(f'[패키지 검증/계정] 완료: {_time.time()-t_total:.1f}초, 특이사항 {len(results)}개사', flush=True)

    status_rank = {'error': 0, 'no_sheet': 1, 'account_misplaced': 2}
    results.sort(key=lambda r: (status_rank.get(r.get('status'), 9), r.get('company') or ''))

    return jsonify({
        'year': year,
        'scanned': scanned,
        'with_issues': len(results),
        'companies': results,
    })


@app.route('/admin/package-verify/wcf-signs')
@require_permission('package.verify')
def admin_package_verify_wcf_signs():
    """WCF U/X/AB/AE 컬럼의 CF 코드별 합계 음수 검증 결과 JSON.

    각 컬럼은 양수만 들어가야 하는 Cash-in/out 금액 컬럼이므로,
    동일 CF 코드 합계가 음수이면 부호 오입력으로 판정.
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    def _verify_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        currency = (f.get('extracted') or {}).get('currency')
        ver = verify_wcf_signs(path)
        dt = _time.time() - t0
        print(f'[패키지 검증/부호] {company} - {dt:.1f}초'
              + (f' (ERROR: {ver.get("error")})' if ver.get('error') else ''),
              flush=True)

        if ver.get('error'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'error', 'message': ver['error'], 'items': [],
            }
        if not ver.get('sheet_found'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'no_sheet', 'message': 'WCF 시트 없음', 'items': [],
            }
        items = ver.get('items') or []
        if not items:
            return None
        # 컬럼별 카운트
        by_col = {}
        for it in items:
            by_col[it['column']] = by_col.get(it['column'], 0) + 1
        return {
            'company': company, 'currency': currency, 'file_id': f.get('id'),
            'status': 'sign_negative',
            'items_count': len(items),
            'by_column': by_col,
            'items': items,
        }

    valid_files = [f for f in files if f.get('path') and f.get('company')]
    scanned = len(valid_files)
    workers = min(8, max(1, scanned)) if scanned else 1
    t_total = _time.time()
    print(f'[패키지 검증/부호] 시작: {scanned}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        all_results = list(ex.map(_verify_one, valid_files))
    results = [r for r in all_results if r is not None]
    print(f'[패키지 검증/부호] 완료: {_time.time()-t_total:.1f}초, 특이사항 {len(results)}개사', flush=True)

    status_rank = {'error': 0, 'no_sheet': 1, 'sign_negative': 2}
    results.sort(key=lambda r: (status_rank.get(r.get('status'), 9), r.get('company') or ''))

    return jsonify({
        'year': year,
        'scanned': scanned,
        'with_issues': len(results),
        'companies': results,
    })


@app.route('/admin/package-verify/wcf-severance')
@require_permission('package.verify')
def admin_package_verify_wcf_severance():
    """WCF P열 CF2200401(퇴직금의 지급)의 R열 합계가 양수인 회사 검증.

    해당 항목은 Cash-out 성격이라 음수가 정상.
    합계가 0 보다 크면 부호 오입력으로 판정.
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    def _verify_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        currency = (f.get('extracted') or {}).get('currency')
        ver = verify_wcf_code_positive(path)
        dt = _time.time() - t0
        print(f'[패키지 검증/퇴직금] {company} - {dt:.1f}초'
              + (f' (ERROR: {ver.get("error")})' if ver.get('error') else ''),
              flush=True)

        if ver.get('error'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'error', 'message': ver['error'], 'rows': [],
            }
        if not ver.get('sheet_found'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'no_sheet', 'message': 'WCF 시트 없음', 'rows': [],
            }
        if not ver.get('is_positive'):
            return None   # 정상(음수 or 0 or 코드 없음)
        return {
            'company': company, 'currency': currency, 'file_id': f.get('id'),
            'status': 'sign_positive',
            'target_code': ver.get('target_code'),
            'name': ver.get('name') or '',
            'sum': ver.get('sum'),
            'rows': ver.get('rows') or [],
        }

    valid_files = [f for f in files if f.get('path') and f.get('company')]
    scanned = len(valid_files)
    workers = min(8, max(1, scanned)) if scanned else 1
    t_total = _time.time()
    print(f'[패키지 검증/퇴직금] 시작: {scanned}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        all_results = list(ex.map(_verify_one, valid_files))
    results = [r for r in all_results if r is not None]
    print(f'[패키지 검증/퇴직금] 완료: {_time.time()-t_total:.1f}초, 특이사항 {len(results)}개사', flush=True)

    status_rank = {'error': 0, 'no_sheet': 1, 'sign_positive': 2}
    results.sort(key=lambda r: (status_rank.get(r.get('status'), 9), r.get('company') or ''))

    return jsonify({
        'year': year,
        'scanned': scanned,
        'with_issues': len(results),
        'companies': results,
        'target_code': 'CF2200401',
        'target_name': '퇴직금의 지급',
    })


@app.route('/admin/package-verify/cf3-liquidity')
@require_permission('package.verify')
def admin_package_verify_cf3_liquidity():
    """CF3 유동성장기차입금(2100201) 블록의 '신규차입' Total 금액이
    0이 아닌 회사 검증.

    유동성장기차입금은 장기차입금의 유동성대체분이라 신규차입이 잡히면 이상치.
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    def _verify_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        currency = (f.get('extracted') or {}).get('currency')
        ver = verify_cf3_current_portion_new_borrowing(path)
        dt = _time.time() - t0
        print(f'[패키지 검증/유동성차입] {company} - {dt:.1f}초'
              + (f' (ERROR: {ver.get("error")})' if ver.get('error') else ''),
              flush=True)

        if ver.get('error'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'error', 'message': ver['error'],
            }
        if not ver.get('sheet_found'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'no_sheet', 'message': 'CF3 시트 없음',
            }
        if not ver.get('is_flagged'):
            return None   # 정상(신규차입 0 or 블록/행 없음)
        return {
            'company': company, 'currency': currency, 'file_id': f.get('id'),
            'status': 'new_borrow_positive',
            'row': ver.get('row'),
            'new_borrow': ver.get('new_borrow'),
            'affiliates': ver.get('affiliates'),
            'third_party': ver.get('third_party'),
        }

    valid_files = [f for f in files if f.get('path') and f.get('company')]
    scanned = len(valid_files)
    workers = min(8, max(1, scanned)) if scanned else 1
    t_total = _time.time()
    print(f'[패키지 검증/유동성차입] 시작: {scanned}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        all_results = list(ex.map(_verify_one, valid_files))
    results = [r for r in all_results if r is not None]
    print(f'[패키지 검증/유동성차입] 완료: {_time.time()-t_total:.1f}초, 특이사항 {len(results)}개사', flush=True)

    status_rank = {'error': 0, 'no_sheet': 1, 'new_borrow_positive': 2}
    results.sort(key=lambda r: (status_rank.get(r.get('status'), 9), r.get('company') or ''))

    return jsonify({
        'year': year,
        'scanned': scanned,
        'with_issues': len(results),
        'companies': results,
        'target_code': '2100201',
        'target_name': '유동성장기차입금 / 신규차입',
    })


@app.route('/admin/package-verify/cf-other-transfer')
@require_permission('package.verify')
def admin_package_verify_cf_other_transfer():
    """CF1/CF2/CF3 "2. 기타증감 내용" 섹션에 기재된 내용이 있는 회사 검증.

    세 시트 중 하나라도 기타증감 입력행이 있으면 표시.
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    def _verify_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        currency = (f.get('extracted') or {}).get('currency')
        ver = verify_cf_other_transfer(path)
        dt = _time.time() - t0
        print(f'[패키지 검증/기타증감] {company} - {dt:.1f}초'
              + (f' (ERROR: {ver.get("error")})' if ver.get('error') else ''),
              flush=True)

        if ver.get('error'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'error', 'message': ver['error'],
            }
        if not ver.get('sheet_found'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'no_sheet', 'message': 'CF1/CF2/CF3 시트 없음',
            }
        if not ver.get('is_flagged'):
            return None   # 정상(기타증감 내용 없음)
        return {
            'company': company, 'currency': currency, 'file_id': f.get('id'),
            'status': 'has_other_transfer',
            'by_sheet': ver.get('by_sheet') or {},
            'sheets_with_content': ver.get('sheets_with_content') or [],
            'entries_count': ver.get('entries_count') or 0,
        }

    valid_files = [f for f in files if f.get('path') and f.get('company')]
    scanned = len(valid_files)
    workers = min(8, max(1, scanned)) if scanned else 1
    t_total = _time.time()
    print(f'[패키지 검증/기타증감] 시작: {scanned}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        all_results = list(ex.map(_verify_one, valid_files))
    results = [r for r in all_results if r is not None]
    print(f'[패키지 검증/기타증감] 완료: {_time.time()-t_total:.1f}초, 특이사항 {len(results)}개사', flush=True)

    status_rank = {'error': 0, 'no_sheet': 1, 'has_other_transfer': 2}
    results.sort(key=lambda r: (status_rank.get(r.get('status'), 9), r.get('company') or ''))

    return jsonify({
        'year': year,
        'scanned': scanned,
        'with_issues': len(results),
        'companies': results,
        'target_name': 'CF1/CF2/CF3 기타증감 내용',
    })


@app.route('/admin/package-verify/cf4-other-transfer')
@require_permission('package.verify')
def admin_package_verify_cf4_other_transfer():
    """CF4 "6. 기타변동 내용" 섹션에 기재된 내용이 있는 회사 검증 (유형자산)."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    def _verify_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        currency = (f.get('extracted') or {}).get('currency')
        ver = verify_cf4_other_transfer(path)
        dt = _time.time() - t0
        print(f'[패키지 검증/CF4 기타변동] {company} - {dt:.1f}초'
              + (f' (ERROR: {ver.get("error")})' if ver.get('error') else ''),
              flush=True)

        if ver.get('error'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'error', 'message': ver['error'],
            }
        if not ver.get('sheet_found'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'no_sheet', 'message': 'CF4 시트 없음',
            }
        if not ver.get('is_flagged'):
            return None   # 정상(기타변동 내용 없음)
        return {
            'company': company, 'currency': currency, 'file_id': f.get('id'),
            'status': 'has_cf4_transfer',
            'entries': ver.get('entries') or [],
            'entries_count': ver.get('entries_count') or 0,
        }

    valid_files = [f for f in files if f.get('path') and f.get('company')]
    scanned = len(valid_files)
    workers = min(8, max(1, scanned)) if scanned else 1
    t_total = _time.time()
    print(f'[패키지 검증/CF4 기타변동] 시작: {scanned}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        all_results = list(ex.map(_verify_one, valid_files))
    results = [r for r in all_results if r is not None]
    print(f'[패키지 검증/CF4 기타변동] 완료: {_time.time()-t_total:.1f}초, 특이사항 {len(results)}개사', flush=True)

    status_rank = {'error': 0, 'no_sheet': 1, 'has_cf4_transfer': 2}
    results.sort(key=lambda r: (status_rank.get(r.get('status'), 9), r.get('company') or ''))

    return jsonify({
        'year': year,
        'scanned': scanned,
        'with_issues': len(results),
        'companies': results,
        'target_name': 'CF4 기타변동 내용 (유형자산)',
    })


@app.route('/admin/package-verify/cf41-other-transfer')
@require_permission('package.verify')
def admin_package_verify_cf41_other_transfer():
    """CF4-1 "3. 기타변동 내용" 섹션에 기재된 내용이 있는 회사 검증 (무형자산)."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    def _verify_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        currency = (f.get('extracted') or {}).get('currency')
        ver = verify_cf41_other_transfer(path)
        dt = _time.time() - t0
        print(f'[패키지 검증/CF4-1 기타변동] {company} - {dt:.1f}초'
              + (f' (ERROR: {ver.get("error")})' if ver.get('error') else ''),
              flush=True)

        if ver.get('error'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'error', 'message': ver['error'],
            }
        if not ver.get('sheet_found'):
            return {
                'company': company, 'currency': currency, 'file_id': f.get('id'),
                'status': 'no_sheet', 'message': 'CF4-1 시트 없음',
            }
        if not ver.get('is_flagged'):
            return None   # 정상(기타변동 내용 없음)
        return {
            'company': company, 'currency': currency, 'file_id': f.get('id'),
            'status': 'has_cf4_transfer',
            'entries': ver.get('entries') or [],
            'entries_count': ver.get('entries_count') or 0,
        }

    valid_files = [f for f in files if f.get('path') and f.get('company')]
    scanned = len(valid_files)
    workers = min(8, max(1, scanned)) if scanned else 1
    t_total = _time.time()
    print(f'[패키지 검증/CF4-1 기타변동] 시작: {scanned}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        all_results = list(ex.map(_verify_one, valid_files))
    results = [r for r in all_results if r is not None]
    print(f'[패키지 검증/CF4-1 기타변동] 완료: {_time.time()-t_total:.1f}초, 특이사항 {len(results)}개사', flush=True)

    status_rank = {'error': 0, 'no_sheet': 1, 'has_cf4_transfer': 2}
    results.sort(key=lambda r: (status_rank.get(r.get('status'), 9), r.get('company') or ''))

    return jsonify({
        'year': year,
        'scanned': scanned,
        'with_issues': len(results),
        'companies': results,
        'target_name': 'CF4-1 기타변동 내용 (무형자산)',
    })


@app.route('/admin/package-verify/l2-verification')
@require_permission('package.verify')
def admin_package_verify_l2_verification():
    """L2 시트 verification 테이블 — G='Y' 회사 + H 사유 표시.

    각 회사의 L2 시트에 있는 'Verification' 라벨 아래 표에서
    G열(오류 Alert)이 'Y'인 행을 찾는다. Y가 있으면 H열(사유 Reason)을 표시.
    H가 비어있는 경우 'reason_missing' 플래그로 강조.

    주석검증이므로 4분기(4Q)에만 실행 가능.
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    if not year.endswith('-4Q'):
        return jsonify({
            'error': '주석 검증은 4분기(4Q)에만 가능합니다. '
                     '결산기간을 4분기로 선택해주세요.',
            'reason': 'quarter_restricted',
        }), 400

    files = _collect_files_for_year(year)
    import time as _time

    def _verify_one(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        currency = (f.get('extracted') or {}).get('currency')
        t0 = _time.time()
        ver = extract_l2_verification(path)
        dt = _time.time() - t0
        print(f'[패키지 검증/L2] {company} - {dt:.1f}초'
              + (f' (ERROR: {ver.get("error")})' if ver.get('error') else ''),
              flush=True)
        if ver.get('error') and not ver.get('sheet_found'):
            return {'company': company, 'currency': currency, 'file_id': f.get('id'),
                    'status': 'error', 'message': ver['error'], 'y_rows': []}
        if not ver.get('sheet_found'):
            return {'company': company, 'currency': currency, 'file_id': f.get('id'),
                    'status': 'no_sheet', 'message': 'L2 시트 없음', 'y_rows': []}
        y_rows = ver.get('y_rows') or []
        if not y_rows:
            return None   # G='Y' 없으면 정상 → 결과 제외
        # 사유 미입력 분리
        has_missing = any(not r.get('reason') for r in y_rows)
        return {
            'company': company, 'currency': currency, 'file_id': f.get('id'),
            'status': 'reason_missing' if has_missing else 'has_alert',
            'verify_row': ver.get('verify_row'),
            'y_rows': y_rows,
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[패키지 검증/L2] 시작: {len(valid)}개 패키지, 워커 {workers}개',
          flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        all_results = list(exr.map(_verify_one, valid))
    results = [r for r in all_results if r is not None]
    print(f'[패키지 검증/L2] 완료: {_time.time()-t_total:.1f}초,'
          f' 특이사항 {len(results)}개사', flush=True)

    # 정렬: error / no_sheet / reason_missing / has_alert / 회사명
    rank = {'error': 0, 'no_sheet': 1, 'reason_missing': 2, 'has_alert': 3}
    results.sort(key=lambda r: (rank.get(r.get('status'), 9),
                                r.get('company') or ''))

    return jsonify({
        'year': year,
        'scanned': len(valid),
        'with_issues': len(results),
        'companies': results,
    })


@app.route('/admin/package-verify/retirement-benefit')
@require_permission('package.verify')
def admin_package_verify_retirement_benefit():
    """L3/L3-1 시트의 입력값이 회계기준(Cover!B27)과 맞지 않는 회사 표시.

    IFRS(K-IFRS) 회사는 L3-1을 사용해야 하므로 L3에 입력값이 있으면 오입력.
    K-GAAP 회사는 L3을 사용해야 하므로 L3-1에 입력값이 있으면 오입력.

    주석검증이므로 4분기(4Q)에만 실행 가능.
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    if not year.endswith('-4Q'):
        return jsonify({
            'error': '주석 검증은 4분기(4Q)에만 가능합니다. '
                     '결산기간을 4분기로 선택해주세요.',
            'reason': 'quarter_restricted',
        }), 400

    files = _collect_files_for_year(year)
    import time as _time

    def _verify_one(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        currency = (f.get('extracted') or {}).get('currency')
        t0 = _time.time()
        ver = verify_retirement_benefit(path)
        dt = _time.time() - t0
        print(f'[패키지 검증/퇴직급여] {company} - {dt:.1f}초'
              + (f' (ERROR: {ver.get("error")})' if ver.get('error') else ''),
              flush=True)

        if ver.get('error'):
            return {'company': company, 'currency': currency,
                    'file_id': f.get('id'),
                    'status': 'error', 'message': ver['error'],
                    'rows': []}
        if not ver.get('cover_found'):
            return {'company': company, 'currency': currency,
                    'file_id': f.get('id'),
                    'status': 'no_sheet', 'message': 'Cover 시트 없음',
                    'rows': []}
        if ver.get('standard') is None:
            return {'company': company, 'currency': currency,
                    'file_id': f.get('id'),
                    'status': 'unknown_standard',
                    'message': f'회계기준 판별 불가 (Cover!B27='
                               f'{ver.get("standard_raw") or "(빈 값)"})',
                    'standard_raw': ver.get('standard_raw'),
                    'rows': []}
        if not ver.get('is_misplaced'):
            return None   # 정상
        return {
            'company': company, 'currency': currency, 'file_id': f.get('id'),
            'status': 'misplaced',
            'standard': ver.get('standard'),
            'standard_raw': ver.get('standard_raw'),
            'target_sheet': ver.get('target_sheet'),
            'rows': ver.get('rows') or [],
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[패키지 검증/퇴직급여] 시작: {len(valid)}개 패키지, '
          f'워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        all_results = list(exr.map(_verify_one, valid))
    results = [r for r in all_results if r is not None]
    print(f'[패키지 검증/퇴직급여] 완료: {_time.time()-t_total:.1f}초, '
          f'특이사항 {len(results)}개사', flush=True)

    rank = {'error': 0, 'no_sheet': 1, 'unknown_standard': 2, 'misplaced': 3}
    results.sort(key=lambda r: (rank.get(r.get('status'), 9),
                                r.get('company') or ''))

    return jsonify({
        'year': year,
        'scanned': len(valid),
        'with_issues': len(results),
        'companies': results,
    })


# ─── 주석 합산 ───────────────────────────────────────────────────────────────

@app.route('/admin/note-aggregate')
@require_permission('note.aggregate')
def admin_note_aggregate_page():
    """주석 합산 페이지 (HTML). 주석은 4분기에만 해당."""
    years_list = YEARS_DATA.get('years', [])
    year = (request.args.get('year') or '').strip()
    if not year or not _valid_year(year):
        # 4Q 우선 선택 (주석은 4분기에만)
        q4 = [y for y in years_list if y.endswith('-4Q')]
        year = q4[0] if q4 else (YEARS_DATA.get('default') or
                                 (years_list[0] if years_list else ''))
    # 4Q만 노출 (주석 비대상 분기는 드롭다운에서 제외)
    q4_years = [y for y in years_list if y.endswith('-4Q')]
    return render_template(
        'note_aggregate.html',
        year=year,
        years=q4_years or years_list,
        username=session.get('username'),
    )


@app.route('/admin/note-aggregate/l1-borrowings')
@require_permission('note.aggregate')
def admin_note_aggregate_l1():
    """L1 단기차입금 종류별 회사 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    # 중앙 환율(2025-4Q current) 우선
    central_current = _get_current_fx_for_period(year)

    def _process_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l1_borrowings(path)
        dt = _time.time() - t0
        print(f'[주석합산 L1] {company} - {dt:.1f}초'
              + (f' (ERROR: {result.get("error")})' if result.get('error') else ''),
              flush=True)
        return {
            'company': company, 'currency': currency, 'spot': spot,
            'file_id': f.get('id'),
            'sheet_found': result.get('sheet_found'),
            'error': result.get('error'),
            'categories': result.get('categories') or [],
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[주석합산 L1] 시작: {len(valid)}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process_one, valid) if r is not None]
    print(f'[주석합산 L1] 완료: {_time.time()-t_total:.1f}초', flush=True)

    # 종류(key)별로 재배열
    cats = {}
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        spot = co.get('spot') or 0
        for cat in co['categories']:
            k = cat['key']
            if k not in cats:
                cats[k] = {'key': k, 'name': cat['name'], 'code': cat['code'], 'rows': []}
            for row in cat['rows']:
                local = row['amount'] or 0
                krw = (local * spot) if spot else 0
                cats[k]['rows'].append({
                    'company': co['company'],
                    'currency': co['currency'],
                    'spot': spot,
                    'creditor_type': row['creditor_type'],
                    'creditor': row['creditor'],
                    'rate': row['rate'],
                    'local': local,
                    'krw': krw,
                })

    # key 숫자순 정렬
    categories_out = []
    for k in sorted(cats.keys(), key=lambda x: int(x) if str(x).isdigit() else 999):
        c = cats[k]
        # 카테고리 내 회사명순 정렬
        c['rows'].sort(key=lambda r: (r['company'] or '', r['creditor'] or ''))
        total_krw = sum(r['krw'] for r in c['rows'])
        by_cur = {}
        for r in c['rows']:
            by_cur[r['currency']] = by_cur.get(r['currency'], 0) + r['local']
        categories_out.append({
            'key': c['key'], 'name': c['name'], 'code': c['code'],
            'rows': c['rows'],
            'total_krw': total_krw,
            'total_local_by_currency': by_cur,
        })

    grand_total_krw = sum(c['total_krw'] for c in categories_out)
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]

    agg_data = {
        'year': year,
        'scanned': len(per_company),
        'categories': categories_out,
        'grand_total_krw': grand_total_krw,
        'errors': [{
            'company': r['company'],
            'reason': r.get('error') or 'L1 시트 없음',
        } for r in errors],
    }

    # 엑셀 산출물 생성 (results/note_aggregate/ 하위)
    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'L1_단기차입금_{year}_{len(per_company)}개사.xlsx'
        out_path = out_dir / excel_filename
        build_l1_excel(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 L1] 엑셀 생성 실패: {e}', flush=True)

    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/l4-loan-facility')
@require_permission('note.aggregate')
def admin_note_aggregate_l4():
    """L4 대출한도 약정(1-1) 회사 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l4_loan_facility(path)
        dt = _time.time() - t0
        print(f'[주석합산 L4] {company} - {dt:.1f}초'
              + (f' (ERROR: {result.get("error")})' if result.get('error') else ''),
              flush=True)
        return {
            'company': company, 'currency': currency, 'spot': spot,
            'file_id': f.get('id'),
            'sheet_found': result.get('sheet_found'),
            'error': result.get('error'),
            'section_label': result.get('section_label'),
            'rows': result.get('rows') or [],
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[주석합산 L4] 시작: {len(valid)}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process_one, valid) if r is not None]
    print(f'[주석합산 L4] 완료: {_time.time()-t_total:.1f}초', flush=True)

    # 행 단위로 평탄화 (회사명 + KRW 환산)
    rows_out = []
    by_cur = {}
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        spot = co.get('spot') or 0
        # 회사의 패키지 통화: 명세 라인에 통화가 따로 적혀 있으면 그 통화로
        # 환산하기 위해 라인별 currency 사용 (KRW 라인은 spot=1.0 처리)
        for row in co['rows']:
            line_cur = (row.get('currency') or 'KRW').strip().upper() or 'KRW'
            if line_cur == 'KRW':
                line_spot = 1.0
            else:
                # 라인 통화가 패키지 통화와 같으면 패키지 spot 사용
                if line_cur == co.get('currency'):
                    line_spot = spot
                else:
                    # 다른 통화면 중앙환율에서 조회 (없으면 0)
                    line_spot = (central_current.get(line_cur) or {}).get('spot') or 0
            local = row.get('amount') or 0
            krw = (local * line_spot) if line_spot else 0
            rows_out.append({
                'company':     co['company'],
                'type':        row.get('type') or '',
                'institution': row.get('institution') or '',
                'currency':    line_cur,
                'local':       local,
                'spot':        line_spot,
                'krw':         krw,
            })
            t = by_cur.setdefault(line_cur, {'local': 0.0, 'krw': 0.0})
            t['local'] += local
            t['krw']   += krw

    # 회사명 → 통화 → 금융기관 순 정렬
    rows_out.sort(key=lambda r: (r['company'] or '', r['currency'] or '',
                                 r['institution'] or '', r['type'] or ''))
    grand_total_krw = sum(r['krw'] for r in rows_out)
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    with_rows = sum(1 for co in per_company if (co.get('rows') or []))

    agg_data = {
        'year': year,
        'scanned': len(per_company),
        'with_rows': with_rows,
        'rows': rows_out,
        'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{
            'company': r['company'],
            'reason': r.get('error') or 'L4 시트 없음',
        } for r in errors],
    }

    # 엑셀 산출물
    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'L4_대출한도약정_{year}_{len(per_company)}개사.xlsx'
        out_path = out_dir / excel_filename
        build_l4_excel(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 L4] 엑셀 생성 실패: {e}', flush=True)

    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/l4-lc')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_lc():
    """L4 2/2-1 (수입신용장 + 미확정 지급보증) 회사별 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        pkg_currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        pkg_spot = (central_current.get(pkg_currency) or {}).get('spot') \
                   or ex.get('fx_spot_current') or (1.0 if pkg_currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l4_lc(path)
        dt = _time.time() - t0
        print(f'[주석합산 L4-LC] {company} - {dt:.1f}초'
              + (f' (ERROR: {result.get("error")})' if result.get('error') else ''),
              flush=True)
        return {
            'company': company,
            'pkg_currency': pkg_currency, 'pkg_spot': pkg_spot,
            'file_id': f.get('id'),
            'sheet_found': result.get('sheet_found'),
            'error': result.get('error'),
            'lc_open': result.get('lc_open'),
            'amount': result.get('amount'),
            'currency': result.get('currency'),
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[주석합산 L4-LC] 시작: {len(valid)}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process_one, valid) if r is not None]
    print(f'[주석합산 L4-LC] 완료: {_time.time()-t_total:.1f}초', flush=True)

    # 회사별 1행 생성
    rows_out = []
    by_cur = {}
    yes_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        if (co.get('lc_open') or '').upper() == 'YES':
            yes_count += 1
        amount = co.get('amount')
        # 통화 결정: 셀에 통화가 있으면 그것, 없으면 패키지 통화
        cur = (co.get('currency') or co.get('pkg_currency') or 'KRW').strip().upper() or 'KRW'
        # spot 결정
        if amount is None:
            spot = None
            krw = None
        else:
            if cur == 'KRW':
                spot = 1.0
            elif cur == co.get('pkg_currency'):
                spot = co.get('pkg_spot')
            else:
                spot = (central_current.get(cur) or {}).get('spot')
            krw = (amount * spot) if (spot is not None) else None

        rows_out.append({
            'company':  co['company'],
            'lc_open':  co.get('lc_open') or '',
            'amount':   amount,
            'currency': cur if amount is not None else '',
            'spot':     spot,
            'krw':      krw,
        })
        if amount is not None and krw is not None:
            t = by_cur.setdefault(cur, {'local': 0.0, 'krw': 0.0})
            t['local'] += amount
            t['krw']   += krw

    rows_out.sort(key=lambda r: (r['company'] or ''))
    grand_total_krw = sum(r['krw'] for r in rows_out if r.get('krw') is not None)
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]

    agg_data = {
        'year': year,
        'scanned': len(per_company),
        'yes_count': yes_count,
        'rows': rows_out,
        'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{
            'company': r['company'],
            'reason': r.get('error') or 'L4 시트 없음',
        } for r in errors],
    }

    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'L4_수입신용장_{year}_{len(per_company)}개사.xlsx'
        out_path = out_dir / excel_filename
        build_l4lc_excel(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 L4-LC] 엑셀 생성 실패: {e}', flush=True)

    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/l4-export')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_export():
    """L4 3/3-1 (수출채권 할인 여부 + 만기 미도래 할인금액) 회사별 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        pkg_currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        pkg_spot = (central_current.get(pkg_currency) or {}).get('spot') \
                   or ex.get('fx_spot_current') or (1.0 if pkg_currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l4_export(path)
        dt = _time.time() - t0
        print(f'[주석합산 L4-EXPORT] {company} - {dt:.1f}초'
              + (f' (ERROR: {result.get("error")})' if result.get('error') else ''),
              flush=True)
        return {
            'company': company,
            'pkg_currency': pkg_currency, 'pkg_spot': pkg_spot,
            'file_id': f.get('id'),
            'sheet_found': result.get('sheet_found'),
            'error': result.get('error'),
            'discount_done': result.get('discount_done'),
            'amount': result.get('amount'),
            'currency': result.get('currency'),
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[주석합산 L4-EXPORT] 시작: {len(valid)}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process_one, valid) if r is not None]
    print(f'[주석합산 L4-EXPORT] 완료: {_time.time()-t_total:.1f}초', flush=True)

    rows_out = []
    by_cur = {}
    yes_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        if (co.get('discount_done') or '').upper() == 'YES':
            yes_count += 1
        amount = co.get('amount')
        cur = (co.get('currency') or co.get('pkg_currency') or 'KRW').strip().upper() or 'KRW'
        if amount is None:
            spot = None
            krw = None
        else:
            if cur == 'KRW':
                spot = 1.0
            elif cur == co.get('pkg_currency'):
                spot = co.get('pkg_spot')
            else:
                spot = (central_current.get(cur) or {}).get('spot')
            krw = (amount * spot) if (spot is not None) else None

        rows_out.append({
            'company':       co['company'],
            'discount_done': co.get('discount_done') or '',
            'amount':        amount,
            'currency':      cur if amount is not None else '',
            'spot':          spot,
            'krw':           krw,
        })
        if amount is not None and krw is not None:
            t = by_cur.setdefault(cur, {'local': 0.0, 'krw': 0.0})
            t['local'] += amount
            t['krw']   += krw

    rows_out.sort(key=lambda r: (r['company'] or ''))
    grand_total_krw = sum(r['krw'] for r in rows_out if r.get('krw') is not None)
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]

    agg_data = {
        'year': year,
        'scanned': len(per_company),
        'yes_count': yes_count,
        'rows': rows_out,
        'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{
            'company': r['company'],
            'reason': r.get('error') or 'L4 시트 없음',
        } for r in errors],
    }

    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'L4_수출채권할인_{year}_{len(per_company)}개사.xlsx'
        out_path = out_dir / excel_filename
        build_l4_export_excel(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 L4-EXPORT] 엑셀 생성 실패: {e}', flush=True)

    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/l4-guarantees-received')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_guarantees():
    """L4 4-1 (제공받은 보증 내용) 회사 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l4_guarantees_received(path)
        dt = _time.time() - t0
        print(f'[주석합산 L4-4-1] {company} - {dt:.1f}초'
              + (f' (ERROR: {result.get("error")})' if result.get('error') else ''),
              flush=True)
        return {
            'company': company, 'currency': currency, 'spot': spot,
            'file_id': f.get('id'),
            'sheet_found': result.get('sheet_found'),
            'error': result.get('error'),
            'section_label': result.get('section_label'),
            'rows': result.get('rows') or [],
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[주석합산 L4-4-1] 시작: {len(valid)}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process_one, valid) if r is not None]
    print(f'[주석합산 L4-4-1] 완료: {_time.time()-t_total:.1f}초', flush=True)

    # 평탄화 + KRW 환산 (라인 통화 기반)
    rows_out = []
    by_cur = {}
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        for row in co['rows']:
            line_cur = (row.get('currency') or 'KRW').strip().upper() or 'KRW'
            if line_cur == 'KRW':
                line_spot = 1.0
            elif line_cur == co.get('currency'):
                line_spot = co.get('spot') or 0
            else:
                line_spot = (central_current.get(line_cur) or {}).get('spot') or 0
            local = row.get('amount') or 0
            krw = (local * line_spot) if line_spot else 0
            rows_out.append({
                'company':     co['company'],
                'guarantor':   row.get('guarantor') or '',
                'type':        row.get('type') or '',
                'currency':    line_cur,
                'local':       local,
                'spot':        line_spot,
                'krw':         krw,
                'account':     row.get('account') or '',
                'description': row.get('description') or '',
            })
            t = by_cur.setdefault(line_cur, {'local': 0.0, 'krw': 0.0})
            t['local'] += local
            t['krw']   += krw

    rows_out.sort(key=lambda r: (r['company'] or '', r['currency'] or '',
                                 r['guarantor'] or '', r['type'] or ''))
    grand_total_krw = sum(r['krw'] for r in rows_out)
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    with_rows = sum(1 for co in per_company if (co.get('rows') or []))

    agg_data = {
        'year': year,
        'scanned': len(per_company),
        'with_rows': with_rows,
        'rows': rows_out,
        'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{
            'company': r['company'],
            'reason': r.get('error') or 'L4 시트 없음',
        } for r in errors],
    }

    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'L4_받은보증_{year}_{len(per_company)}개사.xlsx'
        out_path = out_dir / excel_filename
        build_l4_guarantees_excel(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 L4-4-1] 엑셀 생성 실패: {e}', flush=True)

    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename
    return jsonify(agg_data)


def _aggregate_l4_table_section(year, extractor, log_tag, row_keys):
    """L4 다행 명세 섹션 공통 집계기.

    년도 파일을 수집하여 병렬로 extractor를 호출하고, 회사명·KRW 환산을
    포함한 평탄화된 행 리스트를 만들어 agg_data dict로 반환한다.

    row_keys: 데이터 행에서 보존할 키 목록 (각 추출 함수가 만드는 키 이름)
    """
    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '',
                    reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extractor(path)
        dt = _time.time() - t0
        print(f'[주석합산 {log_tag}] {company} - {dt:.1f}초'
              + (f' (ERROR: {result.get("error")})' if result.get('error') else ''),
              flush=True)
        return {
            'company': company, 'currency': currency, 'spot': spot,
            'sheet_found': result.get('sheet_found'),
            'error': result.get('error'),
            'rows': result.get('rows') or [],
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[주석합산 {log_tag}] 시작: {len(valid)}개 패키지, 워커 {workers}개',
          flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process_one, valid) if r is not None]
    print(f'[주석합산 {log_tag}] 완료: {_time.time()-t_total:.1f}초', flush=True)

    rows_out = []
    by_cur = {}
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        for row in co['rows']:
            line_cur = (row.get('currency') or 'KRW').strip().upper() or 'KRW'
            if line_cur == 'KRW':
                line_spot = 1.0
            elif line_cur == co.get('currency'):
                line_spot = co.get('spot') or 0
            else:
                line_spot = (central_current.get(line_cur) or {}).get('spot') or 0
            local = row.get('amount') or 0
            krw = (local * line_spot) if line_spot else 0
            out = {
                'company': co['company'],
                'currency': line_cur,
                'local': local,
                'spot': line_spot,
                'krw': krw,
            }
            for k in row_keys:
                out[k] = row.get(k, '')
            rows_out.append(out)
            t = by_cur.setdefault(line_cur, {'local': 0.0, 'krw': 0.0})
            t['local'] += local
            t['krw']   += krw

    rows_out.sort(key=lambda r: (r['company'] or '', r['currency'] or ''))
    grand_total_krw = sum(r['krw'] for r in rows_out)
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    with_rows = sum(1 for co in per_company if (co.get('rows') or []))

    return {
        'year': year, 'scanned': len(per_company), 'with_rows': with_rows,
        'rows': rows_out, 'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{
            'company': r['company'],
            'reason': r.get('error') or 'L4 시트 없음',
        } for r in errors],
    }


def _attach_l4_excel(agg_data, year, builder, filename_prefix, log_tag):
    """build_*_excel 호출 + download_url 반환."""
    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'{filename_prefix}_{year}_{agg_data["scanned"]}개사.xlsx'
        out_path = out_dir / excel_filename
        builder(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 {log_tag}] 엑셀 생성 실패: {e}', flush=True)
    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename


@app.route('/admin/note-aggregate/l4-guarantees-provided')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_guarantees_provided():
    """L4 5-2 (제공한 보증 내용) 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    agg = _aggregate_l4_table_section(
        year, extract_l4_guarantees_provided, '5-2',
        ['beneficiary', 'type', 'guaranteed_creditor', 'description'])
    _attach_l4_excel(agg, year, build_l4_guarantees_provided_excel,
                     'L4_제공한보증', '5-2')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l4-restricted-financial')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_restricted():
    """L4 7-1 (사용제한 금융상품) 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    agg = _aggregate_l4_table_section(
        year, extract_l4_restricted_financial, '7-1',
        ['account', 'description'])
    _attach_l4_excel(agg, year, build_l4_restricted_excel,
                     'L4_사용제한금융상품', '7-1')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l4-insured-ppe')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_insured_ppe():
    """L4 8-1 (보험가입 유형자산) 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    agg = _aggregate_l4_table_section(
        year, extract_l4_insured_ppe, '8-1',
        ['asset_type', 'insurer', 'description'])
    _attach_l4_excel(agg, year, build_l4_insured_ppe_excel,
                     'L4_보험가입유형자산', '8-1')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l4-pledged-proceeds')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_pledged_proceeds():
    """L4 8-2 (보험수익금 질권설정) 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    agg = _aggregate_l4_table_section(
        year, extract_l4_pledged_proceeds, '8-2',
        ['pledgee', 'description'])
    _attach_l4_excel(agg, year, build_l4_pledged_proceeds_excel,
                     'L4_보험수익금질권', '8-2')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l4-pledged-assets')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_pledged_assets():
    """L4 9-1 (담보제공자산) 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    agg = _aggregate_l4_table_section(
        year, extract_l4_pledged_assets, '9-1',
        ['creditor', 'asset_account', 'liability_account', 'description'])
    _attach_l4_excel(agg, year, build_l4_pledged_assets_excel,
                     'L4_담보제공자산', '9-1')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l4-lawsuits')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_lawsuits():
    """L4 6-1 (소송중인 사건) 합산 JSON. 회사 패키지 통화/spot으로 KRW 환산."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '',
                    reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l4_lawsuits(path)
        dt = _time.time() - t0
        print(f'[주석합산 6-1] {company} - {dt:.1f}초'
              + (f' (ERROR: {result.get("error")})' if result.get('error') else ''),
              flush=True)
        return {
            'company': company,
            'currency': currency,
            'spot': spot,
            'sheet_found': result.get('sheet_found'),
            'error': result.get('error'),
            'rows': result.get('rows') or [],
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[주석합산 6-1] 시작: {len(valid)}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process_one, valid) if r is not None]
    print(f'[주석합산 6-1] 완료: {_time.time()-t_total:.1f}초', flush=True)

    def _as_int(v):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return 0

    def _as_float(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    rows_out = []
    total_count = 0
    total_claim_krw = 0.0
    total_prov_krw = 0.0
    # 원고/피고 구분 별 집계 (KRW 환산 기준)
    # (패키지 라벨: '원고(Defendant)', '피고(Plaintiff)' — 한국어 기준으로 분류)
    plaintiff_count = 0
    plaintiff_claim_krw = 0.0
    plaintiff_provision_krw = 0.0
    defendant_count = 0
    defendant_claim_krw = 0.0
    # 통화별 합계 (로컬·KRW)
    by_cur = {}

    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        cur = (co.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for row in co['rows']:
            cnt = _as_int(row.get('count'))
            clm = _as_float(row.get('claim_amount'))
            prov = _as_float(row.get('amount'))   # F열 = provision_amount
            clm_krw = clm * spot if spot else 0
            prov_krw = prov * spot if spot else 0
            type_str = row.get('type') or ''
            rows_out.append({
                'company':              co['company'],
                'type':                 type_str,
                'count':                cnt,
                'currency':             cur,
                'spot':                 spot,
                'claim_amount':         clm,
                'provision_amount':     prov,
                'claim_amount_krw':     clm_krw,
                'provision_amount_krw': prov_krw,
            })
            total_count     += cnt
            total_claim_krw += clm_krw
            total_prov_krw  += prov_krw
            # 한국어 라벨 기준 구분
            if '피고' in type_str:
                plaintiff_count         += cnt
                plaintiff_claim_krw     += clm_krw
                plaintiff_provision_krw += prov_krw
            elif '원고' in type_str:
                defendant_count     += cnt
                defendant_claim_krw += clm_krw
            # 통화별 누적
            t = by_cur.setdefault(cur, {
                'count': 0, 'claim': 0.0, 'provision': 0.0,
                'claim_krw': 0.0, 'provision_krw': 0.0,
            })
            t['count']         += cnt
            t['claim']         += clm
            t['provision']     += prov
            t['claim_krw']     += clm_krw
            t['provision_krw'] += prov_krw

    rows_out.sort(key=lambda r: (r['company'] or '', r['type'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    with_rows = sum(1 for co in per_company if (co.get('rows') or []))

    agg_data = {
        'year': year,
        'scanned': len(per_company),
        'with_rows': with_rows,
        'rows': rows_out,
        'total_count':         total_count,
        'total_claim_krw':     total_claim_krw,
        'total_provision_krw': total_prov_krw,
        # 원고/피고 별 집계 (KRW)
        'plaintiff_count':         plaintiff_count,
        'plaintiff_claim_krw':     plaintiff_claim_krw,
        'plaintiff_provision_krw': plaintiff_provision_krw,
        'defendant_count':         defendant_count,
        'defendant_claim_krw':     defendant_claim_krw,
        # 통화별 합계
        'total_by_currency': by_cur,
        'errors': [{
            'company': r['company'],
            'reason': r.get('error') or 'L4 시트 없음',
        } for r in errors],
    }

    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'L4_소송_{year}_{len(per_company)}개사.xlsx'
        out_path = out_dir / excel_filename
        build_l4_lawsuits_excel(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 6-1] 엑셀 생성 실패: {e}', flush=True)

    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/l4-subsequent-events')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_subsequent_events():
    """L4 10번 (보고기간일 이후 사건) — 합산 아닌 YES 회사 목록."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '',
                    reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    def _process_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        result = extract_l4_subsequent_events(path)
        dt = _time.time() - t0
        print(f'[주석합산 10] {company} - {dt:.1f}초'
              + (f' (ERROR: {result.get("error")})' if result.get('error') else ''),
              flush=True)
        return {
            'company': company,
            'sheet_found': result.get('sheet_found'),
            'error': result.get('error'),
            'yn': result.get('yn'),
            'content': result.get('content'),
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[주석합산 10] 시작: {len(valid)}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process_one, valid) if r is not None]
    print(f'[주석합산 10] 완료: {_time.time()-t_total:.1f}초', flush=True)

    yes_rows = []
    yes_count = no_count = empty_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        yn = (co.get('yn') or '').upper()
        if yn == 'YES':
            yes_count += 1
            yes_rows.append({
                'company': co['company'],
                'yn':      'YES',
                'content': co.get('content') or '',
            })
        elif yn == 'NO':
            no_count += 1
        else:
            empty_count += 1

    yes_rows.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]

    agg_data = {
        'year': year,
        'scanned': len(per_company),
        'yes_count': yes_count,
        'no_count': no_count,
        'empty_count': empty_count,
        'rows': yes_rows,
        'errors': [{
            'company': r['company'],
            'reason': r.get('error') or 'L4 시트 없음',
        } for r in errors],
    }

    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'L4_10_보고후사건_{year}_{len(per_company)}개사.xlsx'
        out_path = out_dir / excel_filename
        build_l4_subsequent_events_excel(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 10] 엑셀 생성 실패: {e}', flush=True)

    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/l4-other-commitments')
@require_permission('note.aggregate')
def admin_note_aggregate_l4_other_commitments():
    """L4 11번 (그외 우발부채 및 약정사항) — 내용 있는 회사만 표시."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '',
                    reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time

    def _process_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        result = extract_l4_other_commitments(path)
        dt = _time.time() - t0
        print(f'[주석합산 11] {company} - {dt:.1f}초'
              + (f' (ERROR: {result.get("error")})' if result.get('error') else ''),
              flush=True)
        return {
            'company': company,
            'sheet_found': result.get('sheet_found'),
            'error': result.get('error'),
            'content': result.get('content'),
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[주석합산 11] 시작: {len(valid)}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process_one, valid) if r is not None]
    print(f'[주석합산 11] 완료: {_time.time()-t_total:.1f}초', flush=True)

    rows_out = []
    no_content_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        content = co.get('content')
        if content:
            rows_out.append({
                'company': co['company'],
                'content': content,
            })
        else:
            no_content_count += 1

    rows_out.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]

    agg_data = {
        'year': year,
        'scanned': len(per_company),
        'with_content_count': len(rows_out),
        'no_content_count': no_content_count,
        'rows': rows_out,
        'errors': [{
            'company': r['company'],
            'reason': r.get('error') or 'L4 시트 없음',
        } for r in errors],
    }

    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'L4_11_약정사항_{year}_{len(per_company)}개사.xlsx'
        out_path = out_dir / excel_filename
        build_l4_other_commitments_excel(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 11] 엑셀 생성 실패: {e}', flush=True)

    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/a2-securities')
@require_permission('note.aggregate')
def admin_note_aggregate_a2_securities():
    """A2 유가증권 명세 회사 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '',
                    reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)

    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process_one(f):
        path = f.get('path')
        company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_a2_securities(path)
        dt = _time.time() - t0
        print(f'[주석합산 A2] {company} - {dt:.1f}초'
              + (f' (ERROR: {result.get("error")})' if result.get('error') else ''),
              flush=True)
        return {
            'company': company, 'currency': currency, 'spot': spot,
            'sheet_found': result.get('sheet_found'),
            'error': result.get('error'),
            'rows': result.get('rows') or [],
        }

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    t_total = _time.time()
    print(f'[주석합산 A2] 시작: {len(valid)}개 패키지, 워커 {workers}개', flush=True)
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process_one, valid) if r is not None]
    print(f'[주석합산 A2] 완료: {_time.time()-t_total:.1f}초', flush=True)

    # 행 평탄화 + KRW 환산 (회사 통화 기준 — 유가증권 명세에 통화 컬럼 없음)
    rows_out = []
    by_cur = {}
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for row in co['rows']:
            local_cost = row.get('acquisition_cost') or 0
            local_book = row.get('book_amount') or 0
            krw_cost = local_cost * spot if spot else 0
            krw_book = local_book * spot if spot else 0
            rows_out.append({
                'company':       co['company'],
                'account':       row.get('account') or '',
                'investee':      row.get('investee') or '',
                'shares':        row.get('shares'),
                'ownership_pct': row.get('ownership_pct'),
                'currency':      cur,
                'local_cost':    local_cost,
                'local_book':    local_book,
                'spot':          spot,
                'krw_cost':      krw_cost,
                'krw_book':      krw_book,
            })
            t = by_cur.setdefault(cur, {'cost': 0.0, 'book': 0.0,
                                        'krw_cost': 0.0, 'krw_book': 0.0})
            t['cost']     += local_cost
            t['book']     += local_book
            t['krw_cost'] += krw_cost
            t['krw_book'] += krw_book

    rows_out.sort(key=lambda r: (r['company'] or '', r['account'] or '',
                                 r['investee'] or ''))
    grand_cost_krw = sum(r['krw_cost'] for r in rows_out)
    grand_book_krw = sum(r['krw_book'] for r in rows_out)
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    with_rows = sum(1 for co in per_company if (co.get('rows') or []))

    agg_data = {
        'year': year,
        'scanned': len(per_company),
        'with_rows': with_rows,
        'rows': rows_out,
        'grand_cost_krw': grand_cost_krw,
        'grand_book_krw': grand_book_krw,
        'total_by_currency': by_cur,
        'errors': [{
            'company': r['company'],
            'reason': r.get('error') or 'A2 시트 없음',
        } for r in errors],
    }

    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'A2_유가증권_{year}_{len(per_company)}개사.xlsx'
        out_path = out_dir / excel_filename
        build_a2_securities_excel(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 A2] 엑셀 생성 실패: {e}', flush=True)

    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename
    return jsonify(agg_data)


def _collect_files_for_year(year):
    """년도별 패키지 파일 수집 (중복 제거 + 권한 필터)."""
    uname = session.get('username')
    seen = set()
    files = []
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '',
                    reverse=True):
        if f.get('year') != year:
            continue
        company = f.get('company')
        if not company or not _can_access_company(uname, company):
            continue
        norm = _norm_company_name(company)
        if norm in seen:
            continue
        seen.add(norm)
        files.append(f)
    return files


@app.route('/admin/note-aggregate/a3-investment-pl')
@require_permission('note.aggregate')
def admin_note_aggregate_a3_investment_pl():
    """A3 1. 투자부동산 관련 손익 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_a3_investment_property_pl(path)
        dt = _time.time() - t0
        print(f'[주석합산 A3-1] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'items': result.get('items') or {},
                'total': result.get('total')}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    item_keys = ['rental_revenue', 'operating_expenses', 'depreciation',
                 'fv_change', 'others']
    rows_out = []
    totals_krw = {k: 0.0 for k in item_keys}
    totals_krw['total'] = 0.0
    grand_total_krw = 0.0
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        if not any(v not in (None, 0) for v in items.values()):
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        row = {'company': co['company'], 'currency': cur, 'spot': spot}
        local_total = 0.0
        krw_total = 0.0
        for k in item_keys:
            local = items.get(k) or 0
            krw = local * spot if spot else 0
            row[f'local_{k}'] = local
            row[f'krw_{k}']   = krw
            totals_krw[k] += krw
            local_total += local
            krw_total += krw
        row['local_total'] = local_total
        row['krw_total']   = krw_total
        totals_krw['total'] += krw_total
        grand_total_krw += krw_total
        rows_out.append(row)

    rows_out.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company
              if r.get('error') or not r.get('sheet_found')]

    agg_data = {
        'year': year,
        'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'grand_total_krw': grand_total_krw,
        'totals_by_item_krw': totals_krw,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'A3 시트 없음'} for r in errors],
    }

    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'A3_투자부동산손익_{year}_{len(per_company)}개사.xlsx'
        out_path = out_dir / excel_filename
        build_a3_investment_pl_excel(agg_data, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 A3-1] 엑셀 생성 실패: {e}', flush=True)

    agg_data['download_url'] = download_url
    agg_data['excel_filename'] = excel_filename
    return jsonify(agg_data)


def _aggregate_a3_land_value(year, extractor, log_tag):
    """A3 2-1 / 3-1 공통: 단일 금액 + YN 합산기."""
    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extractor(path)
        dt = _time.time() - t0
        print(f'[주석합산 {log_tag}] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'yn': result.get('yn'),
                'amount': result.get('amount')}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    rows_out = []
    by_cur = {}
    yes_count = with_amount_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        yn = (co.get('yn') or '').upper()
        amount = co.get('amount')
        if yn == 'YES':
            yes_count += 1
        # 금액이 있고 0이 아닌 회사만 합산 결과에 포함
        if amount is None or amount == 0:
            continue
        with_amount_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        krw = amount * spot if spot else 0
        rows_out.append({
            'company': co['company'], 'yn': co.get('yn') or '',
            'currency': cur, 'local': amount, 'spot': spot, 'krw': krw,
        })
        t = by_cur.setdefault(cur, {'local': 0.0, 'krw': 0.0})
        t['local'] += amount; t['krw'] += krw

    rows_out.sort(key=lambda r: r['company'] or '')
    grand_total_krw = sum(r['krw'] for r in rows_out)
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]

    return {
        'year': year, 'scanned': len(per_company),
        'yes_count': yes_count,
        'with_amount_count': with_amount_count,
        'rows': rows_out, 'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'A3 시트 없음'} for r in errors],
    }


@app.route('/admin/note-aggregate/a3-land-investment')
@require_permission('note.aggregate')
def admin_note_aggregate_a3_land_investment():
    """A3 2-1. 투자부동산(토지) 공시지가 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    agg = _aggregate_a3_land_value(year, extract_a3_land_value_investment, '2-1')

    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'A3_투자부동산토지공시지가_{year}_{agg["scanned"]}개사.xlsx'
        out_path = out_dir / excel_filename
        build_a3_land_investment_excel(agg, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 A3-2-1] 엑셀 생성 실패: {e}', flush=True)
    agg['download_url'] = download_url
    agg['excel_filename'] = excel_filename
    return jsonify(agg)


@app.route('/admin/note-aggregate/a3-land-ppe')
@require_permission('note.aggregate')
def admin_note_aggregate_a3_land_ppe():
    """A3 3-1. 유형자산(토지) 공시지가 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    agg = _aggregate_a3_land_value(year, extract_a3_land_value_ppe, '3-1')

    download_url = None
    excel_filename = None
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_filename = f'A3_유형자산토지공시지가_{year}_{agg["scanned"]}개사.xlsx'
        out_path = out_dir / excel_filename
        build_a3_land_ppe_excel(agg, out_path)
        download_url = url_for('download_note_aggregate', filename=excel_filename)
    except Exception as e:
        print(f'[주석합산 A3-3-1] 엑셀 생성 실패: {e}', flush=True)
    agg['download_url'] = download_url
    agg['excel_filename'] = excel_filename
    return jsonify(agg)


_A4_TYPE_KEYS_LIST = ['architecture', 'civil', 'plant', 'hydrogen', 'others']
_A4_TYPE_LABELS = {'architecture': '건축', 'civil': '토목', 'plant': '플랜트',
                   'hydrogen': '수소충전소', 'others': 'Others'}


def _a4_process_packages(year, extractor, log_tag, rate_type='avg'):
    """A4 회사별 추출 공통 헬퍼.

    rate_type:
      'avg'  — P&L 성격 (1번 잔액 변동, 2번 공사손익)
                avg 우선순위: 중앙 fx[avg] → 패키지 fx_avg → KRW면 1.0
      'spot' — 재무상태표 성격 (3번 계약자산·부채)
                spot 우선순위: 중앙 fx[spot] → 패키지 fx_spot_current → KRW면 1.0

    반환 키는 호환성을 위해 'spot'으로 유지, 'rate_type'으로 실제 환율 종류 표시.
    """
    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        if rate_type == 'spot':
            rate = (central_current.get(currency) or {}).get('spot') \
                   or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        else:
            rate = (central_current.get(currency) or {}).get('avg') \
                   or ex.get('fx_avg') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extractor(path)
        dt = _time.time() - t0
        print(f'[주석합산 {log_tag}] {company} - {dt:.1f}초 ({rate_type}={rate})',
              flush=True)
        return {'company': company, 'currency': currency,
                'spot': rate, 'rate_type': rate_type,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'items': result.get('items') or {}}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]
    return per_company


@app.route('/admin/note-aggregate/a4-construction-balance')
@require_permission('note.aggregate')
def admin_note_aggregate_a4_construction_balance():
    """A4 1. 공사계약 잔액 변동 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    per_company = _a4_process_packages(
        year, extract_a4_construction_balance, 'A4-1')

    rows_out = []
    by_cur = {}
    item_keys = ['beginning', 'variance', 'profit', 'others', 'ending']
    totals_krw = {k: 0.0 for k in item_keys}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        # 데이터가 모두 0이면 입력 안 한 것으로 간주
        has_any = any(any(v != 0 for v in d.values()) for d in items.values())
        if not has_any:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for type_key in _A4_TYPE_KEYS_LIST:
            d = items.get(type_key) or {}
            # 해당 공사종류가 모두 0이면 행 생략 (가독성)
            if not any(d.get(k) for k in item_keys):
                continue
            row = {'company': co['company'],
                   'type_key':  type_key,
                   'type_label': _A4_TYPE_LABELS.get(type_key, type_key),
                   'currency':  cur, 'spot': spot}
            for k in item_keys:
                local = d.get(k) or 0
                krw = local * spot if spot else 0
                row[f'local_{k}'] = local
                row[f'krw_{k}']   = krw
                totals_krw[k] += krw
                t = by_cur.setdefault(cur, {f'local_{x}': 0.0 for x in item_keys})
                t.setdefault(f'krw_{k}', 0.0)
                t.setdefault(f'local_{k}', 0.0)
                t[f'local_{k}'] += local
                t[f'krw_{k}']   += krw
            rows_out.append(row)

    rows_out.sort(key=lambda r: (r['company'] or '', r['type_key'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg_data = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'krw_beginning': totals_krw['beginning'],
        'krw_variance':  totals_krw['variance'],
        'krw_profit':    totals_krw['profit'],
        'krw_others':    totals_krw['others'],
        'krw_ending':    totals_krw['ending'],
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'A4 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg_data, year, build_a4_construction_balance_excel,
                     'A4_공사잔액변동', 'A4-1')
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/a4-construction-profit')
@require_permission('note.aggregate')
def admin_note_aggregate_a4_construction_profit():
    """A4 2. 공사손익 합산 JSON (Pivot)."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    per_company = _a4_process_packages(
        year, extract_a4_construction_profit, 'A4-2')

    type_keys = ['architecture', 'civil', 'plant', 'hydrogen', 'others', 'total']
    item_keys = ['accumulated_revenue', 'accumulated_cost', 'accumulated_income']
    item_labels = {
        'accumulated_revenue': '누적공사수익',
        'accumulated_cost':    '누적공사원가',
        'accumulated_income':  '누적공사손익',
    }

    rows_out = []
    totals_krw = {item: {tk: 0.0 for tk in type_keys} for item in item_keys}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        has_any = any(any(v != 0 for v in (items.get(ik) or {}).values())
                      for ik in item_keys)
        if not has_any:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for ik in item_keys:
            d = items.get(ik) or {}
            if not any(d.get(tk) for tk in type_keys):
                continue
            row = {'company': co['company'],
                   'item_key': ik, 'item_label': item_labels[ik],
                   'currency': cur, 'spot': spot}
            for tk in type_keys:
                local = d.get(tk) or 0
                krw = local * spot if spot else 0
                row[f'local_{tk}'] = local
                row[f'krw_{tk}']   = krw
                totals_krw[ik][tk] += krw
            rows_out.append(row)

    rows_out.sort(key=lambda r: (r['company'] or '', r['item_key'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg_data = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'totals_krw': totals_krw,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'A4 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg_data, year, build_a4_construction_profit_excel,
                     'A4_공사손익', 'A4-2')
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/a4-contract-balance')
@require_permission('note.aggregate')
def admin_note_aggregate_a4_contract_balance():
    """A4 3. 계약자산 및 계약부채 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    # A4-3 계약자산·부채는 재무상태표 성격이므로 spot 환율 사용
    per_company = _a4_process_packages(
        year, extract_a4_contract_balance, 'A4-3', rate_type='spot')

    item_keys = ['receivable', 'payable', 'advance']
    rows_out = []
    by_cur = {}
    totals_krw = {k: 0.0 for k in item_keys}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        has_any = any(any(v != 0 for v in d.values()) for d in items.values())
        if not has_any:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for type_key in _A4_TYPE_KEYS_LIST:
            d = items.get(type_key) or {}
            if not any(d.get(k) for k in item_keys):
                continue
            row = {'company': co['company'],
                   'type_key': type_key,
                   'type_label': _A4_TYPE_LABELS.get(type_key, type_key),
                   'currency': cur, 'spot': spot}
            for k in item_keys:
                local = d.get(k) or 0
                krw = local * spot if spot else 0
                row[f'local_{k}'] = local
                row[f'krw_{k}']   = krw
                totals_krw[k] += krw
                t = by_cur.setdefault(cur, {})
                t.setdefault(f'local_{k}', 0.0)
                t.setdefault(f'krw_{k}', 0.0)
                t[f'local_{k}'] += local
                t[f'krw_{k}']   += krw
            rows_out.append(row)

    rows_out.sort(key=lambda r: (r['company'] or '', r['type_key'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg_data = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'krw_receivable': totals_krw['receivable'],
        'krw_payable':    totals_krw['payable'],
        'krw_advance':    totals_krw['advance'],
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'A4 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg_data, year, build_a4_contract_balance_excel,
                     'A4_계약자산부채', 'A4-3')
    return jsonify(agg_data)


_A5_ASSET_KEYS_LIST = ['property', 'vehicle', 'equipment', 'others']
_A5_ASSET_LABELS = {'property': '부동산', 'vehicle': '차량운반구',
                    'equipment': '건설장비', 'others': '기타'}


@app.route('/admin/note-aggregate/a5-rou-changes')
@require_permission('note.aggregate')
def admin_note_aggregate_a5_rou_changes():
    """A5 1. 사용권자산의 변동내역 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    per_company = _a4_process_packages(  # A4용 헬퍼 재사용 (구조 동일)
        year, extract_a5_rou_changes, 'A5-1')

    rows_out = []
    by_cur = {}
    item_keys = ['beginning', 'acquisition', 'disposal',
                 'depreciation', 'others', 'ending']
    totals_krw = {k: 0.0 for k in item_keys}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        has_any = any(any(v != 0 for v in d.values()) for d in items.values())
        if not has_any:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for asset_key in _A5_ASSET_KEYS_LIST:
            d = items.get(asset_key) or {}
            if not any(d.get(k) for k in item_keys):
                continue
            row = {'company': co['company'],
                   'asset_key': asset_key,
                   'asset_label': _A5_ASSET_LABELS.get(asset_key, asset_key),
                   'currency': cur, 'spot': spot}
            for k in item_keys:
                local = d.get(k) or 0
                krw = local * spot if spot else 0
                row[f'local_{k}'] = local
                row[f'krw_{k}']   = krw
                totals_krw[k] += krw
                t = by_cur.setdefault(cur, {})
                t.setdefault(f'local_{k}', 0.0)
                t.setdefault(f'krw_{k}', 0.0)
                t[f'local_{k}'] += local
                t[f'krw_{k}']   += krw
            rows_out.append(row)

    rows_out.sort(key=lambda r: (r['company'] or '', r['asset_key'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg_data = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'krw_beginning':    totals_krw['beginning'],
        'krw_acquisition':  totals_krw['acquisition'],
        'krw_disposal':     totals_krw['disposal'],
        'krw_depreciation': totals_krw['depreciation'],
        'krw_others':       totals_krw['others'],
        'krw_ending':       totals_krw['ending'],
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'A5 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg_data, year, build_a5_rou_changes_excel,
                     'A5_사용권자산변동', 'A5-1')
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/a5-lease-pl')
@require_permission('note.aggregate')
def admin_note_aggregate_a5_lease_pl():
    """A5 2. 리스계약 관련 손익 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_a5_lease_pl(path)
        dt = _time.time() - t0
        print(f'[주석합산 A5-2] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'items': result.get('items') or {},
                'total': result.get('total')}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    item_keys = ['depreciation', 'interest', 'short_term', 'low_value',
                 'variable', 'disposal_gain']
    rows_out = []
    totals_krw = {k: 0.0 for k in item_keys}
    grand_total_krw = 0.0
    by_cur = {}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        if not any(v not in (None, 0) for v in items.values()):
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        row = {'company': co['company'], 'currency': cur, 'spot': spot}
        local_total = 0.0
        krw_total = 0.0
        for k in item_keys:
            local = items.get(k) or 0
            krw = local * spot if spot else 0
            row[f'local_{k}'] = local
            row[f'krw_{k}']   = krw
            totals_krw[k] += krw
            local_total += local
            krw_total   += krw
            t = by_cur.setdefault(cur, {})
            t.setdefault(f'local_{k}', 0.0)
            t.setdefault(f'krw_{k}', 0.0)
            t[f'local_{k}'] += local
            t[f'krw_{k}']   += krw
        row['local_total'] = local_total
        row['krw_total']   = krw_total
        grand_total_krw += krw_total
        rows_out.append(row)

    rows_out.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg_data = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'totals_by_item_krw': totals_krw,
        'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'A5 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg_data, year, build_a5_lease_pl_excel,
                     'A5_리스손익', 'A5-2')
    return jsonify(agg_data)


@app.route('/admin/note-aggregate/a6-derivatives')
@require_permission('note.aggregate')
def admin_note_aggregate_a6_derivatives():
    """A6 1. 파생상품평가손익 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_a6_derivatives(path)
        dt = _time.time() - t0
        print(f'[주석합산 A6] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'rows': result.get('rows') or []}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    rows_out = []
    by_cur = {}
    grand_gain_krw = 0.0
    grand_loss_krw = 0.0
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        co_rows = co.get('rows') or []
        if not co_rows:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for row in co_rows:
            gain = row.get('gain') or 0
            loss = row.get('loss') or 0
            gain_krw = gain * spot if spot else 0
            loss_krw = loss * spot if spot else 0
            rows_out.append({
                'company':    co['company'],
                'type':       row.get('type') or '',
                'currency':   cur,
                'spot':       spot,
                'local_gain': gain,
                'local_loss': loss,
                'krw_gain':   gain_krw,
                'krw_loss':   loss_krw,
            })
            grand_gain_krw += gain_krw
            grand_loss_krw += loss_krw
            t = by_cur.setdefault(cur, {'gain': 0.0, 'loss': 0.0,
                                        'gain_krw': 0.0, 'loss_krw': 0.0})
            t['gain']     += gain
            t['loss']     += loss
            t['gain_krw'] += gain_krw
            t['loss_krw'] += loss_krw

    rows_out.sort(key=lambda r: (r['company'] or '', r['type'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg_data = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'grand_gain_krw': grand_gain_krw,
        'grand_loss_krw': grand_loss_krw,
        'grand_net_krw':  grand_gain_krw - grand_loss_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'A6 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg_data, year, build_a6_derivatives_excel,
                     'A6_파생상품평가', 'A6')
    return jsonify(agg_data)


_A7_TYPE_LABELS = {'subsidiary': '종속회사 (Subsidiaries)',
                   'other':      '기타지분법 (Other Equity Method)'}


@app.route('/admin/note-aggregate/a7-equity-method')
@require_permission('note.aggregate')
def admin_note_aggregate_a7_equity_method():
    """A7 1. 지분법투자주식 명세 합산 JSON."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_a7_equity_method(path)
        dt = _time.time() - t0
        print(f'[주석합산 A7] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'rows': result.get('rows') or []}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    rows_out = []
    by_cur = {}
    grand_cost_krw = 0.0
    grand_net_krw = 0.0
    grand_book_krw = 0.0
    with_data_count = 0
    subsidiary_count = 0
    other_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        co_rows = co.get('rows') or []
        if not co_rows:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for row in co_rows:
            cost = row.get('acquisition_cost') or 0
            net = row.get('net_asset_value') or 0
            book = row.get('book_value') or 0
            cost_krw = cost * spot if spot else 0
            net_krw  = net * spot if spot else 0
            book_krw = book * spot if spot else 0
            type_key = row.get('type') or ''
            if type_key == 'subsidiary':
                subsidiary_count += 1
            elif type_key == 'other':
                other_count += 1
            rows_out.append({
                'company':         co['company'],
                'type':            type_key,
                'type_label':      _A7_TYPE_LABELS.get(type_key, type_key),
                'investee':        row.get('investee') or '',
                'ownership_pct':   row.get('ownership_pct'),
                'currency':        cur,
                'spot':            spot,
                'local_cost':      cost,
                'local_net_asset': net,
                'local_book':      book,
                'krw_cost':        cost_krw,
                'krw_net_asset':   net_krw,
                'krw_book':        book_krw,
            })
            grand_cost_krw += cost_krw
            grand_net_krw  += net_krw
            grand_book_krw += book_krw
            t = by_cur.setdefault(cur, {
                'cost': 0.0, 'net_asset': 0.0, 'book': 0.0,
                'krw_cost': 0.0, 'krw_net_asset': 0.0, 'krw_book': 0.0,
            })
            t['cost']          += cost
            t['net_asset']     += net
            t['book']          += book
            t['krw_cost']      += cost_krw
            t['krw_net_asset'] += net_krw
            t['krw_book']      += book_krw

    # 회사 → 종류(종속 우선) → 회사명 정렬
    type_order = {'subsidiary': 0, 'other': 1}
    rows_out.sort(key=lambda r: (r['company'] or '',
                                 type_order.get(r['type'], 9),
                                 r['investee'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg_data = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'subsidiary_count': subsidiary_count,
        'other_count': other_count,
        'rows': rows_out,
        'grand_cost_krw':       grand_cost_krw,
        'grand_net_asset_krw':  grand_net_krw,
        'grand_book_krw':       grand_book_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'A7 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg_data, year, build_a7_equity_method_excel,
                     'A7_지분법투자주식', 'A7')
    return jsonify(agg_data)


def _aggregate_l2_balance(year, extractor, log_tag, rate_type='spot'):
    """L2 1번/2번 (장기차입금/사채) 회사별 추출 + KRW 환산.

    rate_type='spot' (재무상태표 성격) 기본값.
    """
    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        if rate_type == 'avg':
            rate = (central_current.get(currency) or {}).get('avg') \
                   or ex.get('fx_avg') or (1.0 if currency == 'KRW' else None)
        else:
            rate = (central_current.get(currency) or {}).get('spot') \
                   or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extractor(path)
        dt = _time.time() - t0
        print(f'[주석합산 {log_tag}] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': rate,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'rows': result.get('rows') or []}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    rows_out = []
    by_cur = {}
    grand_current_krw = 0.0
    grand_non_current_krw = 0.0
    grand_total_krw = 0.0
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        co_rows = co.get('rows') or []
        if not co_rows:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for row in co_rows:
            cur_local = row.get('current') or 0
            ncur_local = row.get('non_current') or 0
            tot_local = row.get('total') or 0
            cur_krw = cur_local * spot if spot else 0
            ncur_krw = ncur_local * spot if spot else 0
            tot_krw = tot_local * spot if spot else 0
            rows_out.append({
                'company':           co['company'],
                'type1':             row.get('type1') or '',
                'type2':             row.get('type2') or '',
                'rate':              row.get('rate'),
                'currency':          cur,
                'spot':              spot,
                'local_current':     cur_local,
                'local_non_current': ncur_local,
                'local_total':       tot_local,
                'krw_current':       cur_krw,
                'krw_non_current':   ncur_krw,
                'krw_total':         tot_krw,
            })
            grand_current_krw += cur_krw
            grand_non_current_krw += ncur_krw
            grand_total_krw += tot_krw
            t = by_cur.setdefault(cur, {
                'current': 0.0, 'non_current': 0.0, 'total': 0.0,
                'krw_current': 0.0, 'krw_non_current': 0.0, 'krw_total': 0.0,
            })
            t['current']         += cur_local
            t['non_current']     += ncur_local
            t['total']           += tot_local
            t['krw_current']     += cur_krw
            t['krw_non_current'] += ncur_krw
            t['krw_total']       += tot_krw

    rows_out.sort(key=lambda r: (r['company'] or '', r['type1'] or '',
                                 r['type2'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    return {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'grand_current_krw':     grand_current_krw,
        'grand_non_current_krw': grand_non_current_krw,
        'grand_total_krw':       grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'L2 시트 없음'} for r in errors],
    }


@app.route('/admin/note-aggregate/l2-long-term-borrowings')
@require_permission('note.aggregate')
def admin_note_aggregate_l2_long_term_borrowings():
    """L2 1. 장기차입금 (유동성 포함) 합산."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    agg = _aggregate_l2_balance(year, extract_l2_long_term_borrowings, 'L2-1')
    _attach_l4_excel(agg, year, build_l2_long_term_borrowings_excel,
                     'L2_장기차입금', 'L2-1')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l2-debentures')
@require_permission('note.aggregate')
def admin_note_aggregate_l2_debentures():
    """L2 2. 사채 (유동성 사채 포함) 합산."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    agg = _aggregate_l2_balance(year, extract_l2_debentures, 'L2-2')
    _attach_l4_excel(agg, year, build_l2_debentures_excel,
                     'L2_사채', 'L2-2')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l2-maturity-analysis')
@require_permission('note.aggregate')
def admin_note_aggregate_l2_maturity():
    """L2 3. 부채성 금융상품의 만기 분석 합산.

    재무상태표 성격(B/S) — spot 환율 사용.
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l2_maturity_analysis(path)
        dt = _time.time() - t0
        print(f'[주석합산 L2-3] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'rows': result.get('rows') or []}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    interval_keys = ['within_1y', 'within_2y', 'within_5y', 'over_5y', 'total']
    rows_out = []
    by_cur = {}
    totals_krw = {k: 0.0 for k in interval_keys}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        co_rows = co.get('rows') or []
        if not co_rows:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for row in co_rows:
            out = {
                'company':        co['company'],
                'account':        row.get('account') or '',
                'creditor_type':  row.get('creditor_type') or '',
                'currency':       cur,
                'spot':           spot,
            }
            for k in interval_keys:
                local = row.get(k) or 0
                krw = local * spot if spot else 0
                out[f'local_{k}'] = local
                out[f'krw_{k}']   = krw
                totals_krw[k] += krw
                t = by_cur.setdefault(cur, {})
                t.setdefault(k, 0.0); t.setdefault(f'krw_{k}', 0.0)
                t[k] += local
                t[f'krw_{k}'] += krw
            rows_out.append(out)

    rows_out.sort(key=lambda r: (r['company'] or '', r['account'] or '',
                                 r['creditor_type'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'krw_within_1y': totals_krw['within_1y'],
        'krw_within_2y': totals_krw['within_2y'],
        'krw_within_5y': totals_krw['within_5y'],
        'krw_over_5y':   totals_krw['over_5y'],
        'krw_total':     totals_krw['total'],
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'L2 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg, year, build_l2_maturity_excel,
                     'L2_만기분석', 'L2-3')
    return jsonify(agg)


def _company_group_map(period):
    """{norm_company: 연결그룹명} 매핑 + 그룹 표시순서 리스트 반환.

    consol_schema의 연결그룹 정의에서 해당 period에 멤버인 직속 회사를 매핑.
    """
    from consol_schema import (list_groups as _list_groups,
                               effective_companies as _eff)
    gmap = {}
    order = []
    for g in _list_groups():
        name = (g.get('name') or '').strip() or '(이름없음)'
        if name not in order:
            order.append(name)
        try:
            members = _eff(g, period)
        except Exception:
            members = g.get('companies') or []
        for co in members:
            gmap[_norm_co(co)] = name
    return gmap, order


def _build_group_subtotals(period, rows, item_keys):
    """회사별 rows(krw_<key> 포함)를 연결그룹별 KRW 소계로 집계.

    반환: [{'group','company_count','krw':{item:val}}] — 그룹 정의 순서,
          그룹 미지정 회사는 '(그룹 미지정)' 버킷으로 맨 뒤.
    """
    gmap, order = _company_group_map(period)
    UNASSIGNED = '(그룹 미지정)'
    acc = {}
    for r in rows:
        gname = gmap.get(_norm_co(r.get('company'))) or UNASSIGNED
        e = acc.setdefault(gname, {'companies': set(),
                                   'krw': {k: 0.0 for k in item_keys}})
        e['companies'].add(r.get('company'))
        for k in item_keys:
            e['krw'][k] += r.get(f'krw_{k}') or 0
    ordered = [g for g in order if g in acc]
    if UNASSIGNED in acc:
        ordered.append(UNASSIGNED)
    return [{'group': g, 'company_count': len(acc[g]['companies']),
             'krw': acc[g]['krw']} for g in ordered]


def _l3_process_pivot(year, extractor, item_keys, log_tag, rate_type='spot',
                      with_groups=False):
    """L3 1/2번 (항목별 단일 금액 피벗) 회사별 추출 + 합산.

    with_groups=True 이면 결과에 'group_subtotals'(연결그룹별 KRW 소계) 포함.
    """
    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        if rate_type == 'avg':
            rate = (central_current.get(currency) or {}).get('avg') \
                   or ex.get('fx_avg') or (1.0 if currency == 'KRW' else None)
        else:
            rate = (central_current.get(currency) or {}).get('spot') \
                   or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extractor(path)
        dt = _time.time() - t0
        print(f'[주석합산 {log_tag}] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': rate,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'items': result.get('items') or {}}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    rows_out = []
    by_cur = {}
    totals_krw = {k: 0.0 for k in item_keys}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        if not any(v not in (None, 0) for v in items.values()):
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        out = {'company': co['company'], 'currency': cur, 'spot': spot}
        for k in item_keys:
            local = items.get(k) or 0
            krw = local * spot if spot else 0
            out[f'local_{k}'] = local
            out[f'krw_{k}']   = krw
            totals_krw[k] += krw
            t = by_cur.setdefault(cur, {})
            t.setdefault(f'local_{k}', 0.0); t.setdefault(f'krw_{k}', 0.0)
            t[f'local_{k}'] += local
            t[f'krw_{k}']   += krw
        rows_out.append(out)

    rows_out.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    result = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'totals_by_item_krw': totals_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'L3 시트 없음'} for r in errors],
    }
    if with_groups:
        result['group_subtotals'] = _build_group_subtotals(
            year, rows_out, item_keys)
    return result


def _l3_process_with_fx_effect(year, extractor, var_keys, log_tag):
    """L3 1번/2번 공통: 항목별 환율(기초=전기말 spot / 변동=avg / 기말=당기말 spot)
    + 환율변동효과 계산.

    var_keys: 변동 항목 키 리스트 (기초/기말 제외, avg 환율 적용 대상)
    """
    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)
    central_prior   = _get_prior_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        if currency == 'KRW':
            prior_spot = 1.0; avg = 1.0; spot = 1.0
        else:
            cc = central_current.get(currency) or {}
            cp = central_prior.get(currency) or {}
            spot       = cc.get('spot') or ex.get('fx_spot_current') or 0
            avg        = cc.get('avg')  or ex.get('fx_avg') or 0
            prior_spot = cp.get('spot') or ex.get('fx_spot_prior') or 0
        t0 = _time.time()
        result = extractor(path)
        dt = _time.time() - t0
        print(f'[주석합산 {log_tag}] {company} - {dt:.1f}초 '
              f'(prior={prior_spot}, avg={avg}, spot={spot})', flush=True)
        return {'company': company, 'currency': currency,
                'prior_spot': prior_spot, 'avg': avg, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'items': result.get('items') or {}}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    rows_out = []
    totals_krw = {k: 0.0 for k in (['beginning', 'ending'] + var_keys
                                   + ['fx_effect'])}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        if not any(v not in (None, 0) for v in items.values()):
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        prior_spot = co.get('prior_spot') or 0
        avg        = co.get('avg') or 0
        spot       = co.get('spot') or 0

        out = {'company': co['company'], 'currency': cur,
               'prior_spot': prior_spot, 'avg_rate': avg, 'spot': spot}

        beg_local = items.get('beginning') or 0
        end_local = items.get('ending') or 0
        beg_krw = beg_local * prior_spot
        end_krw = end_local * spot
        out['local_beginning'] = beg_local
        out['krw_beginning']   = beg_krw
        out['local_ending']    = end_local
        out['krw_ending']      = end_krw
        totals_krw['beginning'] += beg_krw
        totals_krw['ending']    += end_krw

        var_krw_sum = 0.0
        for k in var_keys:
            local = items.get(k) or 0
            krw = local * avg
            out[f'local_{k}'] = local
            out[f'krw_{k}']   = krw
            totals_krw[k] += krw
            var_krw_sum += krw

        fx_effect = end_krw - (beg_krw + var_krw_sum)
        out['krw_fx_effect'] = fx_effect
        totals_krw['fx_effect'] += fx_effect

        rows_out.append(out)

    rows_out.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    return {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'totals_by_item_krw': totals_krw,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'L3 시트 없음'} for r in errors],
        'rate_policy': {
            'beginning': '전기말 spot',
            'variance':  '당기 avg',
            'ending':    '당기말 spot',
            'fx_effect': '기말 - (기초 + 변동분)',
        },
    }


@app.route('/admin/note-aggregate/l3-severance')
@require_permission('note.aggregate')
def admin_note_aggregate_l3_severance():
    """L3 1. 퇴직급여충당부채의 변동.

    환율: 기초=전기말 spot / 변동(설정~기타증감)=avg / 기말=당기말 spot.
    환율변동효과 = 기말 - (기초 + Σ 변동분).
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    var_keys = ['provision', 'payment', 'transfer',
                'business_combination', 'others']
    agg = _l3_process_with_fx_effect(
        year, extract_l3_severance_provision, var_keys, 'L3-1')
    _attach_l4_excel(agg, year, build_l3_severance_excel,
                     'L3_퇴직급여충당부채', 'L3-1')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l3-pension-movement')
@require_permission('note.aggregate')
def admin_note_aggregate_l3_pension_movement():
    """L3 2. 퇴직연금운용자산의 변동.

    환율: 기초=전기말 spot / 변동(적립~기타증감)=avg / 기말=당기말 spot.
    환율변동효과 = 기말 - (기초 + Σ 변동분).
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    var_keys = ['contribution', 'payment', 'interest_income',
                'transfer', 'business_combination', 'others']
    agg = _l3_process_with_fx_effect(
        year, extract_l3_pension_funds_movement, var_keys, 'L3-2')
    _attach_l4_excel(agg, year, build_l3_pension_movement_excel,
                     'L3_퇴직연금자산변동', 'L3-2')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l3-pension-breakdown')
@require_permission('note.aggregate')
def admin_note_aggregate_l3_pension_breakdown():
    """L3 3. 퇴직연금운용자산의 구성내역 합산. 재무상태표(B/S) 성격이라 spot."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l3_pension_breakdown(path)
        dt = _time.time() - t0
        print(f'[주석합산 L3-3] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'items': result.get('items') or {}}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    asset_keys = ['cash', 'deposit', 'securities', 'bond', 'others']
    rows_out = []
    by_cur = {}
    totals_krw = {k: 0.0 for k in asset_keys}
    grand_total_krw = 0.0
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        if not any(d.get('amount') not in (None, 0) for d in items.values()):
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        out = {'company': co['company'], 'currency': cur, 'spot': spot}
        co_total_krw = 0.0
        for k in asset_keys:
            local = (items.get(k) or {}).get('amount') or 0
            remarks = (items.get(k) or {}).get('remarks') or ''
            krw = local * spot if spot else 0
            out[f'local_{k}'] = local
            out[f'krw_{k}']   = krw
            out[f'remarks_{k}'] = remarks
            totals_krw[k] += krw
            co_total_krw += krw
            t = by_cur.setdefault(cur, {})
            t.setdefault(f'local_{k}', 0.0); t.setdefault(f'krw_{k}', 0.0)
            t[f'local_{k}'] += local
            t[f'krw_{k}']   += krw
        grand_total_krw += co_total_krw
        rows_out.append(out)

    rows_out.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'totals_by_item_krw': totals_krw,
        'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'L3 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg, year, build_l3_pension_breakdown_excel,
                     'L3_퇴직연금구성', 'L3-3')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l3-pension-managers')
@require_permission('note.aggregate')
def admin_note_aggregate_l3_pension_managers():
    """L3 4. 퇴직연금운용자산의 운용사 합산. 재무상태표(B/S) 성격이라 spot."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l3_pension_managers(path)
        dt = _time.time() - t0
        print(f'[주석합산 L3-4] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'rows': result.get('rows') or []}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    rows_out = []
    by_cur = {}
    grand_total_krw = 0.0
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        co_rows = co.get('rows') or []
        if not co_rows:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for row in co_rows:
            local = row.get('amount') or 0
            krw = local * spot if spot else 0
            rows_out.append({
                'company':  co['company'],
                'name':     row.get('name') or '',
                'currency': cur,
                'spot':     spot,
                'local':    local,
                'krw':      krw,
                'remarks':  row.get('remarks') or '',
            })
            grand_total_krw += krw
            t = by_cur.setdefault(cur, {'local': 0.0, 'krw': 0.0})
            t['local'] += local
            t['krw']   += krw

    rows_out.sort(key=lambda r: (r['company'] or '', r['name'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'L3 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg, year, build_l3_pension_managers_excel,
                     'L3_퇴직연금운용사', 'L3-4')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l31-dbo')
@require_permission('note.aggregate')
def admin_note_aggregate_l31_dbo():
    """L3-1 1. 확정급여채무 변동 합산. P&L 성격 → avg."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    item_keys = ['beginning', 'current_service_cost', 'interest_cost',
                 'remeasurement', 'demographic_gain_loss', 'financial_gain_loss',
                 'experience_adjustment', 'payment',
                 'business_combination', 'others', 'ending']
    agg = _l3_process_pivot(year, extract_l31_dbo_changes,
                            item_keys, 'L3-1-1', rate_type='avg')
    _attach_l4_excel(agg, year, build_l31_dbo_excel,
                     'L3-1_확정급여채무변동', 'L3-1-1')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l31-plan-asset')
@require_permission('note.aggregate')
def admin_note_aggregate_l31_plan_asset():
    """L3-1 2. 사외적립자산 공정가치 변동 합산. P&L 성격 → avg."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    item_keys = ['beginning', 'interest_income', 'return_excluding_interest',
                 'employer_contribution', 'payment',
                 'business_combination', 'others', 'ending']
    agg = _l3_process_pivot(year, extract_l31_plan_asset_changes,
                            item_keys, 'L3-1-2', rate_type='avg')
    _attach_l4_excel(agg, year, build_l31_plan_asset_excel,
                     'L3-1_사외적립자산변동', 'L3-1-2')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l31-assumptions')
@require_permission('note.aggregate')
def admin_note_aggregate_l31_assumptions():
    """L3-1 3. 보험수리적 평가 가정치 — 합산 X, 최저/최고 입력 회사 표시.

    환율 환산 없음 (비율 자체이므로). 값은 raw 텍스트와 정규화된 비율을
    모두 반환. 다양한 형식 (0.0429, '4.29%', '2.27%~2.97%', '5.16% + 호봉률')은
    `_parse_rate_value()`로 정규화됨.
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        t0 = _time.time()
        result = extract_l31_assumptions(path)
        dt = _time.time() - t0
        print(f'[주석합산 L3-1-3] {company} - {dt:.1f}초', flush=True)
        return {'company': company,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'items': result.get('items') or {}}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    rows_out = []
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        wg = items.get('wage_growth') or {}
        dr = items.get('discount_rate') or {}
        has_wg = wg.get('value') is not None
        has_dr = dr.get('value') is not None
        if not (has_wg or has_dr):
            continue
        with_data_count += 1
        rows_out.append({
            'company': co['company'],
            'wage_growth_raw':       wg.get('raw') or '',
            'wage_growth_value':     wg.get('value'),
            'wage_growth_value_min': wg.get('value_min'),
            'wage_growth_value_max': wg.get('value_max'),
            'discount_rate_raw':       dr.get('raw') or '',
            'discount_rate_value':     dr.get('value'),
            'discount_rate_value_min': dr.get('value_min'),
            'discount_rate_value_max': dr.get('value_max'),
        })

    rows_out.sort(key=lambda r: r['company'] or '')

    # 최저/최고 — 비교는 value_min(최저용) / value_max(최고용) 사용 (범위 입력 보정)
    extremes = {}
    for key in ('wage_growth', 'discount_rate'):
        min_key = f'{key}_value_min'
        max_key = f'{key}_value_max'
        min_candidates = [r for r in rows_out if r.get(min_key) is not None]
        max_candidates = [r for r in rows_out if r.get(max_key) is not None]
        min_row = min(min_candidates, key=lambda r: r[min_key]) \
                  if min_candidates else None
        max_row = max(max_candidates, key=lambda r: r[max_key]) \
                  if max_candidates else None
        extremes[key] = {
            'min': {
                'company': (min_row or {}).get('company'),
                'raw':     (min_row or {}).get(f'{key}_raw'),
                'value':   (min_row or {}).get(min_key),
            } if min_row else None,
            'max': {
                'company': (max_row or {}).get('company'),
                'raw':     (max_row or {}).get(f'{key}_raw'),
                'value':   (max_row or {}).get(max_key),
            } if max_row else None,
        }

    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'extremes': extremes,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'L3-1 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg, year, build_l31_assumptions_excel,
                     'L3-1_보험수리가정치', 'L3-1-3')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l31-sensitivity')
@require_permission('note.aggregate')
def admin_note_aggregate_l31_sensitivity():
    """L3-1 4. 보험수리적 가정의 변동 영향 합산. B/S 시점 영향 → spot."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l31_sensitivity(path)
        dt = _time.time() - t0
        print(f'[주석합산 L3-1-4] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'items': result.get('items') or {}}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    # 4개 키: wage_growth_up/down, discount_rate_up/down
    sub_keys = ['wage_growth_up', 'wage_growth_down',
                'discount_rate_up', 'discount_rate_down']
    rows_out = []
    by_cur = {}
    totals_krw = {k: 0.0 for k in sub_keys}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        # 데이터 있는지: 어느 하나라도 not None and not 0
        has_any = any(
            ((items.get(g) or {}).get(d) not in (None, 0))
            for g in ('wage_growth', 'discount_rate')
            for d in ('up', 'down')
        )
        if not has_any:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        out = {'company': co['company'], 'currency': cur, 'spot': spot}
        for g in ('wage_growth', 'discount_rate'):
            for d in ('up', 'down'):
                local = (items.get(g) or {}).get(d) or 0
                krw = local * spot if spot else 0
                out[f'local_{g}_{d}'] = local
                out[f'krw_{g}_{d}']   = krw
                totals_krw[f'{g}_{d}'] += krw
                t = by_cur.setdefault(cur, {})
                t.setdefault(f'local_{g}_{d}', 0.0)
                t.setdefault(f'krw_{g}_{d}', 0.0)
                t[f'local_{g}_{d}'] += local
                t[f'krw_{g}_{d}']   += krw
        rows_out.append(out)

    rows_out.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'totals_by_item_krw': totals_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'L3-1 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg, year, build_l31_sensitivity_excel,
                     'L3-1_민감도', 'L3-1-4')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l31-plan-breakdown')
@require_permission('note.aggregate')
def admin_note_aggregate_l31_plan_breakdown():
    """L3-1 5. 사외적립자산 구성내역 합산. B/S → spot."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l31_plan_breakdown(path)
        dt = _time.time() - t0
        print(f'[주석합산 L3-1-5] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'items': result.get('items') or {}}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    asset_keys = ['cash', 'deposit', 'securities', 'bond', 'others']
    rows_out = []
    by_cur = {}
    totals_krw = {k: 0.0 for k in asset_keys}
    grand_total_krw = 0.0
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        if not any(d.get('amount') not in (None, 0) for d in items.values()):
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        out = {'company': co['company'], 'currency': cur, 'spot': spot}
        co_total = 0.0
        for k in asset_keys:
            local = (items.get(k) or {}).get('amount') or 0
            remarks = (items.get(k) or {}).get('remarks') or ''
            krw = local * spot if spot else 0
            out[f'local_{k}'] = local
            out[f'krw_{k}']   = krw
            out[f'remarks_{k}'] = remarks
            totals_krw[k] += krw
            co_total += krw
            t = by_cur.setdefault(cur, {})
            t.setdefault(f'local_{k}', 0.0); t.setdefault(f'krw_{k}', 0.0)
            t[f'local_{k}'] += local
            t[f'krw_{k}']   += krw
        grand_total_krw += co_total
        rows_out.append(out)

    rows_out.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'totals_by_item_krw': totals_krw,
        'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'L3-1 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg, year, build_l31_plan_breakdown_excel,
                     'L3-1_사외적립구성', 'L3-1-5')
    return jsonify(agg)


@app.route('/admin/note-aggregate/l31-plan-managers')
@require_permission('note.aggregate')
def admin_note_aggregate_l31_plan_managers():
    """L3-1 6. 사외적립자산 운용사 합산 (L3 4번과 동일 패턴). B/S → spot."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_l31_plan_managers(path)
        dt = _time.time() - t0
        print(f'[주석합산 L3-1-6] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'rows': result.get('rows') or []}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    rows_out = []
    by_cur = {}
    grand_total_krw = 0.0
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        co_rows = co.get('rows') or []
        if not co_rows:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        for row in co_rows:
            local = row.get('amount') or 0
            krw = local * spot if spot else 0
            rows_out.append({
                'company':  co['company'],
                'name':     row.get('name') or '',
                'currency': cur,
                'spot':     spot,
                'local':    local,
                'krw':      krw,
                'remarks':  row.get('remarks') or '',
            })
            grand_total_krw += krw
            t = by_cur.setdefault(cur, {'local': 0.0, 'krw': 0.0})
            t['local'] += local; t['krw'] += krw

    rows_out.sort(key=lambda r: (r['company'] or '', r['name'] or ''))
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'grand_total_krw': grand_total_krw,
        'total_by_currency': by_cur,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'L3-1 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg, year, build_l31_plan_managers_excel,
                     'L3-1_사외적립운용사', 'L3-1-6')
    return jsonify(agg)


# ──────────────────────────────────────────────────────────────────────
# TX 시트 — 법인세 합산 라우트 (1 / 3 / 3-1 / 4 / 5 / 5-1)
# ──────────────────────────────────────────────────────────────────────

@app.route('/admin/note-aggregate/tx-deferred-tax-changes')
@require_permission('note.aggregate')
def admin_note_aggregate_tx_deferred_tax_changes():
    """TX 1. 이연법인세자산(부채) 증감내용 합산.

    각 행 = (kor_label, current_flag).
    환율: 기초 = 전기말(전년 4분기) spot / 기말 = 당기말 spot,
          증감(KRW) = 기말 KRW − 기초 KRW (환율변동효과 포함).
    카테고리 = (kor_label, current_flag) 키로 그룹화하여
    회사 수 + 기초/기말/증감 KRW 합계 표시.
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)
    central_prior   = _get_prior_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        if currency == 'KRW':
            spot = 1.0; prior_spot = 1.0
        else:
            cc = central_current.get(currency) or {}
            cp = central_prior.get(currency) or {}
            spot       = cc.get('spot') or ex.get('fx_spot_current') or None
            prior_spot = cp.get('spot') or ex.get('fx_spot_prior') or 0
        t0 = _time.time()
        result = extract_tx_deferred_tax_changes(path)
        dt = _time.time() - t0
        print(f'[주석합산 TX-1] {company} - {dt:.1f}초 '
              f'(prior={prior_spot}, spot={spot})', flush=True)
        return {'company': company, 'currency': currency,
                'spot': spot, 'prior_spot': prior_spot,
                'statutory_rate': result.get('statutory_rate'),
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'rows': result.get('rows') or []}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    # ─── 기초 일시적차이 = 전년도 4분기 주석의 '기말 일시적차이' ───────────
    #   전년 Q4 패키지 파일이 있으면: (그 파일 기말 잔액 × 전기말 spot) ÷ 그 파일 법정세율
    #   해당 분기 파일이 없으면(첫해 등) 비워둠 — 세팅만 되어 있고 값은 None.
    prior_year = _prior_q4_period(year)
    prior_temp_diff = {}  # (norm_company, kor_label, current_flag) → 기말 일시적차이(KRW)
    if prior_year:
        cur_norms = {_norm_company_name(f.get('company')) for f in valid}

        def _process_prior(f):
            path = f.get('path'); company = f.get('company')
            if not path or not company:
                return None
            norm = _norm_company_name(company)
            if norm not in cur_norms:
                return None
            pex = f.get('extracted') or {}
            pcur = (pex.get('currency') or 'KRW').strip().upper() or 'KRW'
            if pcur == 'KRW':
                pspot = 1.0
            else:
                pspot = (central_prior.get(pcur) or {}).get('spot') \
                        or pex.get('fx_spot_current') or 0
            pres = extract_tx_deferred_tax_changes(path)
            return {'norm': norm, 'spot': pspot,
                    'rate': pres.get('statutory_rate'),
                    'rows': pres.get('rows') or [],
                    'sheet_found': pres.get('sheet_found'),
                    'error': pres.get('error')}

        prior_files = _collect_files_for_year(prior_year)
        pvalid = [f for f in prior_files if f.get('path') and f.get('company')]
        if pvalid:
            pworkers = min(8, max(1, len(pvalid)))
            with ThreadPoolExecutor(max_workers=pworkers) as exr:
                prior_results = [r for r in exr.map(_process_prior, pvalid) if r]
            for pr in prior_results:
                if pr.get('error') or not pr.get('sheet_found'):
                    continue
                prate = pr.get('rate')
                pspot = pr.get('spot') or 0
                if not prate:
                    continue
                for it in pr['rows']:
                    end_krw = (it.get('ending') or 0) * pspot
                    key = (pr['norm'], it.get('kor_label'),
                           it.get('current_flag') or '')
                    prior_temp_diff[key] = end_krw / prate

    company_rows = []
    cat_map = {}  # (kor, cur_flag) → {company_count, beg, end, chg}
    totals_krw = {'beginning': 0.0, 'ending': 0.0, 'change': 0.0}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('rows') or []
        if not items:
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        prior_spot = co.get('prior_spot') or (1.0 if cur == 'KRW' else 0)
        rate = co.get('statutory_rate') or None  # 분율(0.22) / 0·None은 미적용
        co_norm = _norm_company_name(co['company'])
        for it in items:
            beg = it.get('beginning') or 0
            end = it.get('ending') or 0
            chg = it.get('change') or 0
            beg_krw = beg * prior_spot if prior_spot else 0
            end_krw = end * spot if spot else 0
            chg_krw = end_krw - beg_krw
            # 기말 일시적차이 = 기말 환산액 ÷ 법정세율
            end_temp = (end_krw / rate) if rate else None
            # 기초 일시적차이 = 전년도 4분기 주석의 기말 일시적차이 (없으면 None)
            beg_temp = prior_temp_diff.get(
                (co_norm, it['kor_label'], it.get('current_flag') or ''))
            company_rows.append({
                'company': co['company'], 'currency': cur,
                'spot': spot, 'prior_spot': prior_spot,
                'statutory_rate': rate,
                'kor_label': it['kor_label'],
                'current_flag': it.get('current_flag') or '',
                'local_beginning': beg, 'local_ending': end, 'local_change': chg,
                'krw_beginning':   beg_krw,
                'krw_ending':      end_krw,
                'krw_change':      chg_krw,
                'beginning_temp_diff': beg_temp,
                'ending_temp_diff':    end_temp,
            })
            totals_krw['beginning'] += beg_krw
            totals_krw['ending']    += end_krw
            totals_krw['change']    += chg_krw
            ck = (it['kor_label'], it.get('current_flag') or '')
            entry = cat_map.setdefault(ck, {
                'kor_label': ck[0], 'current_flag': ck[1],
                'company_count': 0,
                'beginning_krw': 0.0, 'ending_krw': 0.0, 'change_krw': 0.0,
                '_seen_companies': set(),
            })
            if co['company'] not in entry['_seen_companies']:
                entry['_seen_companies'].add(co['company'])
                entry['company_count'] += 1
            entry['beginning_krw'] += beg_krw
            entry['ending_krw']    += end_krw
            entry['change_krw']    += chg_krw

    categories = []
    for cat in cat_map.values():
        cat.pop('_seen_companies', None)
        categories.append(cat)
    categories.sort(key=lambda c: (c['kor_label'], c['current_flag']))
    company_rows.sort(key=lambda r: (r['company'] or '', r['kor_label'] or ''))

    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'categories': categories,
        'totals_krw': totals_krw,
        'company_rows': company_rows,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'TX 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg, year, build_tx_deferred_tax_changes_excel,
                     'TX_이연법인세_변동', 'TX-1')
    return jsonify(agg)


@app.route('/admin/note-aggregate/tx-income-tax-breakdown')
@require_permission('note.aggregate')
def admin_note_aggregate_tx_income_tax_breakdown():
    """TX 3. 법인세비용의 구성내역 합산. P&L → avg."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    item_keys = ['current_tax', 'deferred_temp_diff', 'deferred_equity',
                 'additional_refund', 'total_expense']
    agg = _l3_process_pivot(year, extract_tx_income_tax_breakdown,
                            item_keys, 'TX-3', rate_type='avg',
                            with_groups=True)
    _attach_l4_excel(agg, year, build_tx_income_tax_breakdown_excel,
                     'TX_법인세구성', 'TX-3')
    return jsonify(agg)


@app.route('/admin/note-aggregate/tx-equity-deferred-tax')
@require_permission('note.aggregate')
def admin_note_aggregate_tx_equity_deferred_tax():
    """TX 3-1. 자본 직접 부과 이연법인세 변동액 명세 합산. avg 환율."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    item_keys = ['revaluation', 'actuarial', 'afs_securities', 'fvoci',
                 'equity_method', 'fx_translation', 'derivatives',
                 'others', 'total']

    # extractor가 items dict에 __total을 'total'로 매핑하도록 어댑터
    def _adapter(path):
        r = extract_tx_equity_deferred_tax(path)
        if r.get('items'):
            r['items']['total'] = r['items'].pop('__total', None)
        return r

    agg = _l3_process_pivot(year, _adapter, item_keys, 'TX-3-1', rate_type='avg')
    _attach_l4_excel(agg, year, build_tx_equity_deferred_tax_excel,
                     'TX_자본직접부과', 'TX-3-1')
    return jsonify(agg)


@app.route('/admin/note-aggregate/tx-reconciliation')
@require_permission('note.aggregate')
def admin_note_aggregate_tx_reconciliation():
    """TX 4. Reconciliation 합산. P&L → avg. 유효세율은 회사별 표시."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    item_keys = ['pretax_income', 'tax_at_statutory', 'permanent_diff',
                 'tax_credit', 'additional_refund', 'unrecognized_change',
                 'total_expense']

    # 유효세율을 별도로 보존
    rate_by_company = {}

    def _adapter(path):
        r = extract_tx_reconciliation(path)
        # rate_by_company는 후처리에서 채움
        r['_effective_rate'] = r.get('effective_rate')
        return r

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        rate = (central_current.get(currency) or {}).get('avg') \
               or ex.get('fx_avg') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_tx_reconciliation(path)
        dt = _time.time() - t0
        print(f'[주석합산 TX-4] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': rate,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'items': result.get('items') or {},
                'effective_rate': result.get('effective_rate')}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    rows_out = []
    totals_krw = {k: 0.0 for k in item_keys}
    with_data_count = 0
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        items = co.get('items') or {}
        # 유효세율이라도 있으면 표시
        has_any = any(v not in (None, 0) for v in items.values())
        if not has_any and co.get('effective_rate') in (None, 0):
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        rate = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        out = {'company': co['company'], 'currency': cur, 'spot': rate,
               'effective_rate': co.get('effective_rate')}
        for k in item_keys:
            local = items.get(k) or 0
            krw = local * rate if rate else 0
            out[f'local_{k}'] = local
            out[f'krw_{k}']   = krw
            totals_krw[k] += krw
        rows_out.append(out)

    rows_out.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'rows': rows_out,
        'totals_by_item_krw': totals_krw,
        'group_subtotals': _build_group_subtotals(year, rows_out, item_keys),
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'TX 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg, year, build_tx_reconciliation_excel,
                     'TX_Reconciliation', 'TX-4')
    return jsonify(agg)


@app.route('/admin/note-aggregate/tx-unrecognized')
@require_permission('note.aggregate')
def admin_note_aggregate_tx_unrecognized():
    """TX 5. 미인식 일시적차이 합산. 잔액 → spot."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400
    item_keys = ['loss_carryforward', 'others']
    agg = _l3_process_pivot(year, extract_tx_unrecognized_temp_diff,
                            item_keys, 'TX-5', rate_type='spot')
    _attach_l4_excel(agg, year, build_tx_unrecognized_excel,
                     'TX_미인식일시적차이', 'TX-5')
    return jsonify(agg)


@app.route('/admin/note-aggregate/tx-loss-maturity')
@require_permission('note.aggregate')
def admin_note_aggregate_tx_loss_maturity():
    """TX 5-1. 이월결손금 만기 합산. 잔액 → spot.

    만기 라벨이 회계연도별로 다름 → 라벨 그대로 사용해서 동일 라벨 회사들만
    같은 bucket으로 합산.
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    files = _collect_files_for_year(year)
    import time as _time
    central_current = _get_current_fx_for_period(year)

    def _process(f):
        path = f.get('path'); company = f.get('company')
        if not path or not company:
            return None
        ex = f.get('extracted') or {}
        currency = (ex.get('currency') or 'KRW').strip().upper() or 'KRW'
        spot = (central_current.get(currency) or {}).get('spot') \
               or ex.get('fx_spot_current') or (1.0 if currency == 'KRW' else None)
        t0 = _time.time()
        result = extract_tx_loss_carryforward_maturity(path)
        dt = _time.time() - t0
        print(f'[주석합산 TX-5-1] {company} - {dt:.1f}초', flush=True)
        return {'company': company, 'currency': currency, 'spot': spot,
                'sheet_found': result.get('sheet_found'),
                'error': result.get('error'),
                'buckets': result.get('buckets') or []}

    valid = [f for f in files if f.get('path') and f.get('company')]
    workers = min(8, max(1, len(valid))) if valid else 1
    with ThreadPoolExecutor(max_workers=workers) as exr:
        per_company = [r for r in exr.map(_process, valid) if r is not None]

    bucket_map = {}  # label → {'total_krw','company_count','_companies'}
    company_rows = []
    grand_total_krw = 0.0
    with_data_count = 0
    label_order = []  # 첫 등장 순서 유지
    for co in per_company:
        if co.get('error') or not co.get('sheet_found'):
            continue
        buckets = co.get('buckets') or []
        if not any((b.get('amount') or 0) != 0 for b in buckets):
            continue
        with_data_count += 1
        cur = co.get('currency') or 'KRW'
        spot = co.get('spot') or (1.0 if cur == 'KRW' else 0)
        co_buckets = []
        for b in buckets:
            lbl = b.get('label')
            amount = b.get('amount') or 0
            krw = amount * spot if spot else 0
            co_buckets.append({'label': lbl, 'local': amount, 'krw': krw})
            entry = bucket_map.get(lbl)
            if entry is None:
                entry = bucket_map[lbl] = {
                    'label': lbl, 'total_krw': 0.0,
                    'company_count': 0, '_companies': set(),
                }
                label_order.append(lbl)
            if amount != 0:
                if co['company'] not in entry['_companies']:
                    entry['_companies'].add(co['company'])
                    entry['company_count'] += 1
                entry['total_krw'] += krw
                grand_total_krw += krw
        company_rows.append({
            'company': co['company'], 'currency': cur, 'spot': spot,
            'buckets': co_buckets,
        })

    # label_order 순서 유지 (After XXXX 는 항상 마지막으로 정렬)
    def _ord_key(lbl):
        s = str(lbl)
        if s.lower().startswith('after'):
            return (1, s)
        return (0, s)
    label_order_sorted = sorted(label_order, key=_ord_key)
    buckets_list = []
    for lbl in label_order_sorted:
        e = bucket_map[lbl]
        e.pop('_companies', None)
        buckets_list.append(e)

    company_rows.sort(key=lambda r: r['company'] or '')
    errors = [r for r in per_company if r.get('error') or not r.get('sheet_found')]
    agg = {
        'year': year, 'scanned': len(per_company),
        'with_data_count': with_data_count,
        'buckets': buckets_list,
        'grand_total_krw': grand_total_krw,
        'company_rows': company_rows,
        'errors': [{'company': r['company'],
                    'reason': r.get('error') or 'TX 시트 없음'} for r in errors],
    }
    _attach_l4_excel(agg, year, build_tx_loss_maturity_excel,
                     'TX_이월결손금만기', 'TX-5-1')
    return jsonify(agg)


# 통합 다운로드 — 시트별 sub-route 목록
_NOTE_AGGREGATE_SECTIONS = [
    ('l1',                       'l1-borrowings'),
    ('l2_long_term_borrowings',  'l2-long-term-borrowings'),
    ('l2_debentures',            'l2-debentures'),
    ('l2_maturity',              'l2-maturity-analysis'),
    ('l3_severance',             'l3-severance'),
    ('l3_pension_movement',      'l3-pension-movement'),
    ('l3_pension_breakdown',     'l3-pension-breakdown'),
    ('l3_pension_managers',      'l3-pension-managers'),
    ('l31_dbo',                  'l31-dbo'),
    ('l31_plan_asset',           'l31-plan-asset'),
    ('l31_assumptions',          'l31-assumptions'),
    ('l31_sensitivity',          'l31-sensitivity'),
    ('l31_plan_breakdown',       'l31-plan-breakdown'),
    ('l31_plan_managers',        'l31-plan-managers'),
    ('tx_deferred_tax_changes',  'tx-deferred-tax-changes'),
    ('tx_income_tax_breakdown',  'tx-income-tax-breakdown'),
    ('tx_equity_deferred_tax',   'tx-equity-deferred-tax'),
    ('tx_reconciliation',        'tx-reconciliation'),
    ('tx_unrecognized',          'tx-unrecognized'),
    ('tx_loss_maturity',         'tx-loss-maturity'),
    ('l4_loan_facility',         'l4-loan-facility'),
    ('l4_lc',                    'l4-lc'),
    ('l4_export',                'l4-export'),
    ('l4_guarantees_received',   'l4-guarantees-received'),
    ('l4_guarantees_provided',   'l4-guarantees-provided'),
    ('l4_lawsuits',              'l4-lawsuits'),
    ('l4_restricted_financial',  'l4-restricted-financial'),
    ('l4_insured_ppe',           'l4-insured-ppe'),
    ('l4_pledged_proceeds',      'l4-pledged-proceeds'),
    ('l4_pledged_assets',        'l4-pledged-assets'),
    ('l4_subsequent_events',     'l4-subsequent-events'),
    ('l4_other_commitments',     'l4-other-commitments'),
    ('a2_securities',            'a2-securities'),
    ('a3_investment_pl',         'a3-investment-pl'),
    ('a3_land_investment',       'a3-land-investment'),
    ('a3_land_ppe',              'a3-land-ppe'),
    ('a4_construction_balance',  'a4-construction-balance'),
    ('a4_construction_profit',   'a4-construction-profit'),
    ('a4_contract_balance',      'a4-contract-balance'),
    ('a5_rou_changes',           'a5-rou-changes'),
    ('a5_lease_pl',              'a5-lease-pl'),
    ('a6_derivatives',           'a6-derivatives'),
    ('a7_equity_method',         'a7-equity-method'),
]


@app.route('/admin/note-aggregate/all-in-one')
@require_permission('note.aggregate')
def admin_note_aggregate_all_in_one():
    """모든 주석합산 결과를 단일 워크북에 시트별로 작성한 엑셀 다운로드 URL 반환."""
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    import time as _time
    t0 = _time.time()
    print(f'[주석합산 ALL] 시작: {year}, 섹션 {len(_NOTE_AGGREGATE_SECTIONS)}개,'
          f' 현재 사용자={session.get("username")!r}', flush=True)

    # test_client로 자기 자신 호출 — 현재 세션을 복사해 인증 유지
    client = app.test_client()
    # 현재 세션 데이터를 sub-route에 그대로 전달
    current_session = dict(session)
    with client.session_transaction() as sub_sess:
        for k, v in current_session.items():
            sub_sess[k] = v

    all_aggs = {}
    for key, route in _NOTE_AGGREGATE_SECTIONS:
        url = f'/admin/note-aggregate/{route}?year={year}'
        ts = _time.time()
        resp = client.get(url)
        dt = _time.time() - ts
        if resp.status_code == 200:
            try:
                agg = resp.get_json() or {}
                all_aggs[key] = agg
                # 진단 로그: rows / categories 수
                if 'categories' in agg:
                    n = sum(len(c.get('rows') or []) for c in agg.get('categories') or [])
                else:
                    n = len(agg.get('rows') or [])
                print(f'[주석합산 ALL]   {key:30s} status=200 rows={n:4d} '
                      f'scanned={agg.get("scanned",0)} {dt:5.1f}초', flush=True)
            except Exception as e:
                print(f'[주석합산 ALL] {key} JSON 파싱 실패: {e}', flush=True)
                all_aggs[key] = {}
        else:
            body = resp.get_data(as_text=True)[:200]
            print(f'[주석합산 ALL] {key} 실패 status={resp.status_code} '
                  f'body={body!r}', flush=True)
            all_aggs[key] = {}

    print(f'[주석합산 ALL] 모든 sub-route 완료: {_time.time()-t0:.1f}초', flush=True)
    # 어떤 섹션이 데이터 있고 어떤 게 비어있는지 진단
    nonempty_keys = sorted([k for k, v in all_aggs.items()
                            if isinstance(v, dict) and (
                                v.get('rows') or v.get('categories'))])
    print(f'[주석합산 ALL] 데이터 있는 섹션 ({len(nonempty_keys)}/{len(all_aggs)}):'
          f' {nonempty_keys}', flush=True)

    # 통합 엑셀 생성 — 파일명에 timestamp 포함하여 브라우저 캐시 회피
    download_url = None
    excel_filename = None
    sheet_count = 0
    try:
        out_dir = RESULTS_DIR / 'note_aggregate'
        out_dir.mkdir(parents=True, exist_ok=True)
        timestamp = _time.strftime('%Y%m%d_%H%M%S')
        excel_filename = f'주석합산_전체_{year}_{timestamp}.xlsx'
        out_path = out_dir / excel_filename
        build_all_in_one_excel(year, all_aggs, out_path)
        # 실제 생성된 시트 개수 확인
        try:
            from openpyxl import load_workbook
            _wb = load_workbook(str(out_path), read_only=True)
            sheet_count = len(_wb.sheetnames)
            print(f'[주석합산 ALL] 생성된 시트 ({sheet_count}개): {_wb.sheetnames}',
                  flush=True)
            _wb.close()
        except Exception:
            pass
        download_url = url_for('download_note_aggregate', filename=excel_filename)
        print(f'[주석합산 ALL] 엑셀 생성 완료: {_time.time()-t0:.1f}초', flush=True)
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f'[주석합산 ALL] 엑셀 생성 실패: {e}', flush=True)
        return jsonify({'error': f'엑셀 생성 실패: {e}'}), 500

    return jsonify({
        'year': year,
        'sections_count': len(all_aggs),
        'sheets_count': sheet_count,
        'download_url': download_url,
        'excel_filename': excel_filename,
        'elapsed_seconds': round(_time.time() - t0, 1),
    })


@app.route('/admin/note-aggregate/download/<path:filename>')
@require_permission('note.aggregate')
def download_note_aggregate(filename):
    """주석 합산 결과 엑셀 다운로드."""
    # 경로 탈출 방지: filename은 단순 파일명만 허용
    if '/' in filename or '\\' in filename or '..' in filename:
        return '잘못된 파일명입니다.', 400
    path = RESULTS_DIR / 'note_aggregate' / filename
    if not path.exists():
        return '파일을 찾을 수 없습니다.', 404
    return send_file(str(path.resolve()), as_attachment=True, download_name=filename)


# ─── 환율 관리 라우트 ─────────────────────────────────────────────────────────

@app.route('/admin/fx-rates/periods', methods=['GET'])
@login_required
def admin_fx_rates_periods():
    """환율 모달 분기 드롭다운용 목록.
    결산기간(years)과 비교용 보조분기(fx_only_years)를 구분해 반환.
    보조분기는 결산 흐름에 노출되지 않고 환율 입력만 가능.
    """
    return jsonify({
        'years':         list(YEARS_DATA.get('years', [])),
        'fx_only_years': list(YEARS_DATA.get('fx_only_years', [])),
        'default':       YEARS_DATA.get('default'),
    })


@app.route('/admin/fx-rates/<year>/data', methods=['GET'])
@login_required
def admin_fx_rates_data(year):
    """환율 데이터 조회 API (메인 페이지 모달에서 사용)."""
    if not _valid_fx_year(year):
        return jsonify({'error': '유효하지 않은 기간'}), 400

    current_rates = _get_current_fx_for_period(year)
    is_first_year = (year == WCE_FIRST_YEAR)
    prior_year_label = _prior_q4_period(year)
    prior_rates = _get_prior_fx_for_period(year)
    pkg_currencies = _get_currencies_from_packages(year)
    all_currencies = sorted(
        (set(current_rates.keys()) | set(pkg_currencies) | set(prior_rates.keys())) - {'KRW'}
    )

    return jsonify({
        'year': year,
        'current_rates': current_rates,
        'prior_rates': prior_rates,
        'prior_auto': not is_first_year,
        'prior_year_label': prior_year_label,
        'is_first_year': is_first_year,
        'all_currencies': all_currencies,
    })


@app.route('/admin/fx-rates/<year>', methods=['POST'])
@require_permission('fx.manage')
def admin_fx_rates_save(year):
    """환율 저장 API."""
    if not _valid_fx_year(year):
        return jsonify({'error': '유효하지 않은 기간'}), 400

    payload = request.get_json(silent=True) or {}

    def _to_float(v):
        if v in (None, ''):
            return None
        try:
            return float(str(v).replace(',', '').strip())
        except Exception:
            return None

    fx_all = _load_fx_rates()
    if year not in fx_all:
        fx_all[year] = {}

    # 당기 환율 저장
    new_current = {}
    for cur_key, rates in (payload.get('current') or {}).items():
        key = str(cur_key).strip().upper()
        if not key or key == 'KRW':
            continue
        spot = _to_float(rates.get('spot'))
        avg  = _to_float(rates.get('avg'))
        if spot is not None or avg is not None:
            new_current[key] = {'spot': spot, 'avg': avg}
    fx_all[year]['current'] = new_current

    # 전기 환율 (WCE_FIRST_YEAR만 수동 저장 허용)
    if year == WCE_FIRST_YEAR:
        new_prior = {}
        for cur_key, rates in (payload.get('prior') or {}).items():
            key = str(cur_key).strip().upper()
            if not key or key == 'KRW':
                continue
            spot = _to_float(rates.get('spot'))
            avg  = _to_float(rates.get('avg'))
            if spot is not None or avg is not None:
                new_prior[key] = {'spot': spot, 'avg': avg}
        fx_all[year]['prior'] = new_prior

    _save_fx_rates(fx_all)
    return jsonify({'ok': True, 'year': year, 'currencies_saved': len(new_current)})


@app.route('/admin/fx-rates/<year>/reapply', methods=['POST'])
@require_permission('fx.manage')
def admin_fx_rates_reapply(year):
    """이 기간 업로드된 패키지를 중앙 환율로 재환산.

    payload (선택): {'ids': ['<file_id>', ...]}
      - 지정 시: 해당 ID의 파일만 처리
      - 미지정/빈 배열: 이 기간 전체 파일 처리

    - currency==KRW 파일은 스킵
    - 중앙 환율이 전혀 없는 통화도 스킵 (변경할 게 없음)
    """
    if not _valid_fx_year(year):
        return jsonify({'error': '유효하지 않은 기간'}), 400

    payload = request.get_json(silent=True) or {}
    selected_ids = payload.get('ids') or []
    id_set = set(selected_ids) if selected_ids else None

    targets = [
        e for e in uploaded_files
        if e.get('year') == year
        and (id_set is None or e.get('id') in id_set)
    ]
    ok_list, skipped_list, failed_list = [], [], []

    for entry in targets:
        cur = (entry.get('extracted') or {}).get('currency')
        if not cur or cur.upper() == 'KRW':
            skipped_list.append({'name': entry.get('original_name'), 'reason': 'KRW'})
            continue
        rates = _get_central_rates_for(year, cur)
        if not rates:
            skipped_list.append({'name': entry.get('original_name'), 'reason': '중앙 환율 미입력'})
            continue
        try:
            data = extract(entry['path'], central_rates=rates)
            entry['extracted'] = data
            entry['company'] = data['company']
            entry['sheet_summary'] = {
                s: len(rows) for s, rows in data['sheets'].items() if rows
            }
            ok_list.append({
                'name': entry.get('original_name'),
                'currency': cur,
                'avg': data.get('fx_avg'),
                'spot_current': data.get('fx_spot_current'),
            })
        except Exception as e:
            print(f'[재환산 실패] {entry.get("original_name")}: {e}',
                  file=sys.stderr, flush=True)
            failed_list.append({'name': entry.get('original_name'), 'error': str(e)})

    _save_state()
    return jsonify({
        'ok': True,
        'updated': len(ok_list),
        'skipped': len(skipped_list),
        'failed': len(failed_list),
        'details': {'ok': ok_list, 'skipped': skipped_list, 'failed': failed_list},
    })


@app.route('/admin/fx-rates/template', methods=['GET'])
@require_permission('fx.manage')
def admin_fx_rates_template():
    """환율 일괄 업로드용 엑셀 양식(.xlsx) 다운로드.

    쿼리: ?year=YYYY-NQ  (선택)
      · 지정 시: 그 분기에 저장된 환율을 prefill. 데이터 없으면 헤더만.
      · 미지정: 헤더만 (빈 양식).

    시트:
      · 당기환율  — A=통화코드, B=Spot Rate, C=Avg Rate (2행부터 데이터)
      · 전기환율  — 동일 양식. 최초연도(WCE_FIRST_YEAR)에서만 업로드 시 반영.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    year = (request.args.get('year') or '').strip()
    current_rows, prior_rows = [], []
    suffix = ''
    if year and _valid_fx_year(year):
        cur = _get_current_fx_for_period(year) or {}
        pri = _get_prior_fx_for_period(year) or {}
        for k in sorted(cur.keys()):
            current_rows.append((k, cur[k].get('spot'), cur[k].get('avg')))
        for k in sorted(pri.keys()):
            prior_rows.append((k, pri[k].get('spot'), pri[k].get('avg')))
        suffix = f'_{year}'

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F3864')
    center = Alignment(horizontal='center', vertical='center')

    def _build_sheet(ws, title, rows):
        ws.title = title
        ws['A1'] = '통화코드'
        ws['B1'] = 'Spot Rate'
        ws['C1'] = 'Avg Rate'
        for col in ('A1', 'B1', 'C1'):
            ws[col].font = header_font
            ws[col].fill = header_fill
            ws[col].alignment = center
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        for i, (cur, spot, avg) in enumerate(rows, start=2):
            ws.cell(i, 1).value = cur
            ws.cell(i, 2).value = spot
            ws.cell(i, 3).value = avg
        ws.freeze_panes = 'A2'

    _build_sheet(wb.active, '당기환율', current_rows)
    _build_sheet(wb.create_sheet(), '전기환율', prior_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'환율_업로드양식{suffix}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/admin/fx-rates/<year>/upload', methods=['POST'])
@require_permission('fx.manage')
def admin_fx_rates_upload(year):
    """엑셀 양식으로 환율 일괄 업로드.

    - 시트명에 '당기' 또는 'current' 포함 → 당기 환율로 저장
    - 시트명에 '전기' 또는 'prior'  포함 → 전기 환율로 저장 (WCE_FIRST_YEAR만 반영)
    - 칼럼: A=통화코드, B=Spot, C=Avg (1행 헤더, 2행부터 데이터)
    - 빈 값은 무시하고, 둘 다 비어 있으면 해당 통화는 스킵
    - 기존 저장값은 업로드 분으로 **대체**됨
    """
    if not _valid_fx_year(year):
        return jsonify({'error': '유효하지 않은 기간'}), 400

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': '파일이 첨부되지 않았습니다.'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': '엑셀(.xlsx/.xlsm) 파일만 업로드 가능합니다.'}), 400

    def _to_float(v):
        if v in (None, ''):
            return None
        try:
            return float(str(v).replace(',', '').strip())
        except Exception:
            return None

    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(file, data_only=True, read_only=True)
    except Exception as e:
        return jsonify({'error': f'엑셀 파일을 열 수 없습니다: {e}'}), 400

    def _parse_sheet(ws):
        out = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cur = row[0]
            if cur is None:
                continue
            key = str(cur).strip().upper()
            if not key or key == 'KRW' or key == '통화코드':
                continue
            spot = _to_float(row[1] if len(row) > 1 else None)
            avg  = _to_float(row[2] if len(row) > 2 else None)
            if spot is None and avg is None:
                continue
            out[key] = {'spot': spot, 'avg': avg}
        return out

    current_in, prior_in = {}, {}
    found_current = False
    found_prior = False
    for sn in wb.sheetnames:
        low = sn.lower()
        if ('당기' in sn) or ('current' in low):
            current_in = _parse_sheet(wb[sn])
            found_current = True
        elif ('전기' in sn) or ('prior' in low):
            prior_in = _parse_sheet(wb[sn])
            found_prior = True

    # 시트명이 규칙에 안 맞으면 첫 시트를 당기로 처리
    if not found_current and not found_prior and wb.sheetnames:
        current_in = _parse_sheet(wb[wb.sheetnames[0]])

    fx_all = _load_fx_rates()
    if year not in fx_all:
        fx_all[year] = {}
    fx_all[year]['current'] = current_in

    saved_prior = 0
    if year == WCE_FIRST_YEAR and prior_in:
        fx_all[year]['prior'] = prior_in
        saved_prior = len(prior_in)

    _save_fx_rates(fx_all)
    return jsonify({
        'ok': True,
        'year': year,
        'currencies_saved': len(current_in),
        'prior_saved': saved_prior,
        'is_first_year': (year == WCE_FIRST_YEAR),
    })


# ─── 회사 마스터 관리 라우트 (관리자 전용) ────────────────────────────────────
# 업로드 대상 회사목록 + 회사별 통화 + 통화별 적용 환율을 한 화면에서 관리.

@app.route('/admin/company-master')
@admin_required
def admin_company_master_page():
    """회사 마스터 관리 화면."""
    uname = session.get('username')
    return render_template('admin_company_master.html',
                           username=uname,
                           is_admin=_is_admin(uname))


@app.route('/admin/company-master/data', methods=['GET'])
@admin_required
def admin_company_master_data():
    """회사 마스터 + (분기 지정 시) 그 분기의 통화별 당기 환율 번들 반환."""
    year = (request.args.get('year') or '').strip()
    companies = list(_load_company_master().get('companies') or [])

    resp = {'companies': companies, 'year': year, 'valid': False}
    if year and _valid_fx_year(year):
        current_rates = _get_current_fx_for_period(year)
        co_curs = {(c.get('currency') or '').strip().upper()
                   for c in companies if (c.get('currency') or '').strip()}
        pkg_curs = set(_get_currencies_from_packages(year))
        all_currencies = sorted(
            (co_curs | pkg_curs | set(current_rates.keys())) - {'KRW', ''}
        )
        resp.update({
            'valid': True,
            'current_rates': current_rates,
            'all_currencies': all_currencies,
        })
    return jsonify(resp)


@app.route('/admin/company-master', methods=['POST'])
@admin_required
def admin_company_master_save():
    """회사 목록(이름·통화·활성) 저장."""
    payload = request.get_json(silent=True) or {}
    companies = payload.get('companies')
    if not isinstance(companies, list):
        return jsonify({'error': '회사 목록 형식이 올바르지 않습니다.'}), 400
    saved = _save_company_master(companies)
    return jsonify({'ok': True, 'count': len(saved)})


@app.route('/admin/company-master/rates', methods=['POST'])
@admin_required
def admin_company_master_save_rates():
    """선택 분기의 '당기(current)' 통화별 환율만 저장 — prior 는 건드리지 않음.
    (전기/최초연도 환율은 기존 환율관리 모달에서 별도 처리)
    """
    year = (request.args.get('year') or '').strip()
    if not _valid_fx_year(year):
        return jsonify({'error': '유효하지 않은 기간'}), 400
    payload = request.get_json(silent=True) or {}

    def _to_float(v):
        if v in (None, ''):
            return None
        try:
            return float(str(v).replace(',', '').strip())
        except Exception:
            return None

    fx_all = _load_fx_rates()
    fx_all.setdefault(year, {})
    new_current = {}
    for cur_key, rates in (payload.get('current') or {}).items():
        key = str(cur_key).strip().upper()
        if not key or key == 'KRW':
            continue
        spot = _to_float((rates or {}).get('spot'))
        avg = _to_float((rates or {}).get('avg'))
        if spot is not None or avg is not None:
            new_current[key] = {'spot': spot, 'avg': avg}
    fx_all[year]['current'] = new_current
    _save_fx_rates(fx_all)
    return jsonify({'ok': True, 'count': len(new_current)})


# ─── 결산연도 관리 ────────────────────────────────────────────────────────────

PERIOD_RE = re.compile(r'^(\d{4})-([1-4])Q$')


def _period_sort_key(p):
    """'2024-3Q' → (2024, 3). 형식 맞지 않으면 (0,0)."""
    m = PERIOD_RE.match(p or '')
    if not m:
        return (0, 0)
    return (int(m.group(1)), int(m.group(2)))


def _load_years():
    """years_config.json 로드. 없으면 현재연도 1분기로 생성.
    기존 순수 연도값('2024' 등)은 '2024-1Q'로 자동 마이그레이션.
    """
    import datetime
    default_period = f'{datetime.datetime.now().year}-1Q'

    if not YEARS_FILE.exists():
        data = {'years': [default_period], 'default': default_period, 'locked': []}
        with open(YEARS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return data

    with open(YEARS_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    data.setdefault('years', [])
    data.setdefault('locked', [])
    data.setdefault('fx_only_years', [])

    # 구 형식(순수 연도) → 분기 형식 마이그레이션
    migrated = False
    new_years = []
    for y in data['years']:
        if PERIOD_RE.match(y):
            new_years.append(y)
        elif re.fullmatch(r'\d{4}', str(y)):
            new_years.append(f'{y}-1Q')
            migrated = True
        else:
            new_years.append(y)
    data['years'] = sorted(set(new_years), key=_period_sort_key, reverse=True)

    if data.get('default') and re.fullmatch(r'\d{4}', str(data['default'])):
        data['default'] = f'{data["default"]}-1Q'
        migrated = True
    if not data.get('default'):
        data['default'] = data['years'][0] if data['years'] else None

    if migrated:
        with open(YEARS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print('[시작] 결산연도 데이터를 분기 형식으로 마이그레이션 완료')

    return data


def _save_years():
    with _years_filelock:
        _atomic_write_json(YEARS_FILE, YEARS_DATA)


YEARS_DATA = _load_years()


# ─── WCE 본사 입력값 (132행 이후 6개 테이블) 저장소 ───────────────────────────

WCE_FIRST_YEAR = '2025-4Q'  # 시스템 운영 시작 분기 — 이 분기는 기초금액 수기 입력


def _wce_key(year, company):
    return f'{year}::{company}'


def _prior_q4_period(year):
    """현재 year의 '전년도 4분기' 기간 문자열 반환. 첫해이거나 형식 오류면 None."""
    if year == WCE_FIRST_YEAR:
        return None
    m = PERIOD_RE.match(year or '')
    if not m:
        return None
    return f'{int(m.group(1)) - 1}-4Q'


def _prior_same_q_period(year):
    """현재 year의 '전년도 동기' 기간 문자열 반환. 예: '2026-1Q' → '2025-1Q'.
    회사별 전기비교 출력물(동기 매칭) 전용. WCE_FIRST_YEAR도 동일 규칙으로 산출.
    형식 오류면 None.
    """
    m = PERIOD_RE.match(year or '')
    if not m:
        return None
    return f"{int(m.group(1)) - 1}-{m.group(2)}Q"


def _get_same_q_fx_for_period(year):
    """회사별 전기비교 전용: 전년 동기에 등록된 'current' 환율을 반환.
    없으면 빈 dict. 비교용 보조분기(fx_only_years)에 입력된 환율을 가져오는 경로.
    """
    prior_same = _prior_same_q_period(year)
    if not prior_same:
        return {}
    return (_load_fx_rates().get(prior_same) or {}).get('current') or {}


def _find_uploaded_for(year, company):
    """주어진 (year, company)의 업로드 파일 entry 반환.
    동일 (year, company)에 여러 파일이 있으면 **가장 최근 업로드** 반환.
    없으면 None.
    """
    target = _norm_company_name(company)
    matches = [
        f for f in uploaded_files
        if f.get('year') == year
        and _norm_company_name(f.get('company', '')) == target
    ]
    if not matches:
        return None
    matches.sort(key=lambda f: f.get('uploaded_at') or '', reverse=True)
    return matches[0]


def _wce_local_full_for(year, company):
    """업로드 파일에서 wce_local_full(현지통화 raw 값) + currency + fx_avg 반환.
    반환: {'local': {tid: {code: {label: value}}}, 'currency': str, 'fx_avg': float, 'source_file': str|None}

    환율 우선순위:
      1) 중앙 환율 관리(fx_rates.json)에 설정된 당기 Avg Rate
      2) 패키지 파일에서 추출한 avg_rate
    """
    f = _find_uploaded_for(year, company)
    if not f:
        return {'local': {}, 'currency': None, 'fx_avg': None, 'source_file': None}
    ex = f.get('extracted') or {}
    currency = ex.get('currency')
    fx_avg = ex.get('fx_avg')

    # 중앙 관리 환율이 설정되어 있으면 우선 사용
    if currency and currency.upper() != 'KRW':
        central = _get_current_fx_for_period(year).get(currency.upper())
        if central and central.get('avg'):
            fx_avg = central['avg']

    return {
        'local': ex.get('wce_local_full') or {},
        'currency': currency,
        'fx_avg': fx_avg,
        'source_file': f.get('original_name'),
    }


def _lookup_local(local_data, schema_table_id, schema_code, schema_row_key):
    """schema 키로 로컬값 조회 (코드/라벨 별칭 자동 적용)."""
    tid = str(schema_table_id)
    local_code = wce_to_local_code(schema_code)
    local_label = wce_to_local_label(schema_row_key)
    return ((local_data.get(tid) or {}).get(local_code, {}) or {}).get(local_label, 0) or 0


def _wce_auto_re_cells_for(year, company):
    """자동 채움 셀의 KRW 환산값 반환.
    구조: {code: {row_key: krw_val}}.

    적용 대상:
      - 5번 이익잉여금: 당기순이익 / 보험수리적손익 / R/E조정
        · 행 합계 3500105(col 7) 우선, 없으면 3500104(col 6 Unappropriated) 폴백
      - 6번 비지배지분: 당기순이익 / 보험수리적손익
        · 단일 컬럼 FS32000000 (로컬 3600101)

    신규: wce_local_full × fx_avg 우선. 레거시: wce_local_re 폴백.
    """
    info = _wce_local_full_for(year, company)
    local = info['local']
    fx_avg = info['fx_avg'] or 0
    if local and fx_avg:
        out = {}
        # 5번 이익잉여금 — 행별 회계 귀속 코드:
        #   당기순이익 → 3500105 (Current Net Income, 자동 환산)
        # 자동 채움에서 제외 (수기 입력):
        #   보험수리적손익     → 3500104에 수기 입력
        #   지분법이익잉여금   → 3500104에 수기 입력 (구 'R/E조정')
        t5 = wce_get_table(5)
        T5_TARGET = {
            '당기순이익': '3500105',
        }
        for row_key, target_code in T5_TARGET.items():
            # 로컬값: 귀속 코드 컬럼 우선, 없으면 반대 컬럼 폴백
            # (col 7(3500105) 우선, 없으면 col 6(3500104))
            # abs() 금지 — 음수(순손실) 부호 반드시 유지
            if target_code == '3500104':
                total_v = (_lookup_local(local, 5, '3500104', row_key)
                           or _lookup_local(local, 5, '3500105', row_key))
            else:
                total_v = (_lookup_local(local, 5, '3500105', row_key)
                           or _lookup_local(local, 5, '3500104', row_key))
            if total_v:
                for col in t5['columns']:
                    code = col['code']
                    out.setdefault(code, {})[row_key] = (total_v * fx_avg) if code == target_code else 0.0

        # 회계상 위치하면 안 되는 셀: 강제 0 + readonly 마킹
        # (이전에 자동으로 채워졌다가 wce_overrides.json에 잘못 저장된 값 정리용)
        T5_FORCE_ZERO = [
            ('3500105', '보험수리적손익'),    # 3500104에만 (수기 입력)
            ('3500105', '지분법이익잉여금'),  # 3500104에만 (수기 입력)
        ]
        for (code, rk) in T5_FORCE_ZERO:
            out.setdefault(code, {})[rk] = 0.0

        # 6번 비지배지분 (Non-controlling Interest)
        t6 = wce_get_table(6)
        for row_key in ('당기순이익', '보험수리적손익'):
            v = _lookup_local(local, 6, 'FS32000000', row_key)
            if v:
                for col in t6['columns']:
                    code = col['code']
                    out.setdefault(code, {})[row_key] = v * fx_avg

        if out:
            return out
    # 레거시 폴백 (wce_local_re는 이미 KRW)
    f = _find_uploaded_for(year, company)
    re_data = ((f or {}).get('extracted') or {}).get('wce_local_re') or {}
    out = {}
    for row_key, code_dict in re_data.items():
        for code, val in (code_dict or {}).items():
            out.setdefault(code, {})[row_key] = val
    return out


def _apply_wce_to_aggregation(agg, year):
    """합산 결과의 자본 항목을 WCE 입력값으로 대체하고
    차대 불일치를 3400104(해외사업환산손익)로 자동 조정.

    적용 대상: WCE 데이터가 있는 회사만.
    미입력 회사는 기존 xlsm BS 값 그대로 유지.
    """
    wce_data = _load_wce()
    bs = agg['sheets'].get('BS', {})
    companies = agg.get('companies', [])

    # 수정된 코드 추적 (합계 재계산용)
    modified_codes = set()
    # 회사별 조정 내역 (응답에 포함)
    adj_log = []

    def _bs_val(code, company):
        return (bs.get(code, {}).get('by_company', {}).get(company, 0) or 0)

    def _set_bs_val(code, company, val):
        if code not in bs:
            bs[code] = {
                'kor': '', 'eng': '',
                'by_company': {c: 0 for c in companies},
                'compare_by_company': {c: 0 for c in companies},
                'total': 0, 'compare_total': 0,
            }
        bs[code].setdefault('by_company', {})[company] = val
        modified_codes.add(code)

    for company in companies:
        # _get_wce_for로 실시간 자동 채움(당기순이익·기초금액 합산 등) 적용된 데이터 사용
        # → 저장된 wce_overrides만 쓰면 자동값이 누락됨 (사용자가 페이지 열기만 하고 저장 안 한 경우)
        wce_full = _get_wce_for(year, company)
        wce_tables_data = wce_full['tables']

        # 의미있는 WCE 값이 있는지 검사 (저장값 + 자동값 모두 0이면 패키지 BS 그대로 사용)
        has_meaningful_data = any(
            v
            for tid_data in wce_tables_data.values()
            for code_data in tid_data.values()
            for v in code_data.values()
            if v
        )
        if not has_meaningful_data:
            continue

        # ① 각 테이블의 기말 잔액 계산 {code: value}
        wce_endings = {}
        for t in WCE_TABLES:
            tid = str(t['id'])
            endings = _compute_table_ending(t, wce_tables_data.get(tid) or {})
            for col in t['columns']:
                wce_endings[col['code']] = endings.get(col['code'], 0)

        # ② 차대 조정: Assets - Liabilities - WCE자본총계 → 3400104에 반영
        assets           = _bs_val('1000000', company)
        liab             = _bs_val('2000000', company)
        wce_equity_total = sum(wce_endings.values())
        adj              = assets - liab - wce_equity_total

        wce_endings['3400104'] = wce_endings.get('3400104', 0) + adj

        # ③ BS leaf 코드 업데이트
        # WCE 내부 코드(FS32000000=비지배지분)는 BS의 실제 코드(3600101)로 매핑
        WCE_TO_BS_CODE = {'FS32000000': '3600101'}
        for code, val in wce_endings.items():
            bs_code = WCE_TO_BS_CODE.get(code, code)
            _set_bs_val(bs_code, company, val)

        # ④ 자본 소계 코드 재계산 (3100000, 3200000 …)
        for group_code, children in WCE_EQUITY_GROUPS.items():
            group_val = sum(wce_endings.get(c, _bs_val(c, company)) for c in children)
            _set_bs_val(group_code, company, group_val)

        # ⑤ 자본총계(3000000) = 모든 leaf 합 (조정 후)
        equity_total_adj = sum(wce_endings.values())
        _set_bs_val('3000000', company, equity_total_adj)

        adj_log.append({
            'company': company,
            'assets': assets,
            'liabilities': liab,
            'wce_equity_before': wce_equity_total,
            'adjustment_3400104': adj,
            'equity_after': equity_total_adj,
            'balanced': abs(assets - liab - equity_total_adj) < 1,
        })

    # ⑥ 수정된 코드의 합계(total) 재계산
    for code in modified_codes:
        if code in bs:
            bs[code]['total'] = sum(v for v in bs[code].get('by_company', {}).values() if v)

    agg['sheets']['BS'] = bs
    agg['wce_adj_log'] = adj_log
    return agg


def _compute_wce_missing(local_data, fx_avg, current_tables):
    """현지값(×환율 추정 KRW) 대비 환산값(저장된 KRW) 미입력 셀 계산.
    - 로컬 != 0 인데 저장된 환산값 == 0 인 셀을 missing 으로 카운트
    - 자동 채워지는 셀(기초금액·자동 RE)은 제외
    반환: {'count': int, 'cells': [{'table_id', 'code', 'row_key', 'local'}]}
    """
    if not local_data or not fx_avg:
        return {'count': 0, 'cells': []}

    # 자동 셀 키 — 항상 자동 채워지므로 미입력 검출에서 제외
    # 5번: 당기순이익만 자동 (보험수리적손익·지분법이익잉여금은 수기 입력)
    AUTO_RE_ROWS_T5 = {'당기순이익'}
    AUTO_RE_ROWS_T6 = {'당기순이익', '보험수리적손익'}

    missing = []
    for t in WCE_TABLES:
        tid = str(t['id'])
        for col in t['columns']:
            code = col['code']
            for row in t['rows']:
                rk = row['key']
                # 자동 채움 행 제외
                if rk == '기초금액':
                    continue
                if tid == '5' and rk in AUTO_RE_ROWS_T5:
                    continue
                if tid == '6' and rk in AUTO_RE_ROWS_T6:
                    continue
                # 환산효과 같은 KRW-only 행은 로컬에 없을 수 있음 — 스킵
                if rk == '환산효과':
                    continue

                local_v = _lookup_local(local_data, t['id'], code, rk)
                if not local_v:
                    continue  # 로컬 값이 없으면 미입력 아님
                converted_v = (((current_tables.get(tid) or {}).get(code) or {}).get(rk, 0)) or 0
                if abs(converted_v) < 0.5:  # 사실상 0
                    missing.append({
                        'table_id': tid,
                        'code': code,
                        'row_key': rk,
                        'local': local_v,
                    })
    return {'count': len(missing), 'cells': missing}


def _compute_table_ending(table_def, table_data):
    """저장된 테이블 데이터로부터 코드별 기말금액(=기초+증감, 환산효과 제외) 계산.
    반환: {code: ending_value}
    """
    out = {}
    for col in table_def['columns']:
        code = col['code']
        ending = 0
        for r in table_def['rows']:
            if r['key'] == '환산효과':
                continue
            v = (table_data.get(code) or {}).get(r['key'], 0) or 0
            ending += float(v)
        out[code] = ending
    return out


def _load_wce():
    """wce_overrides.json 전체 로드. 없으면 빈 dict."""
    if not WCE_FILE.exists():
        return {}
    try:
        with _wce_filelock:
            with open(WCE_FILE, 'r', encoding='utf-8') as fp:
                return json.load(fp)
    except Exception as e:
        print(f'[경고] WCE 로드 실패: {e}')
        return {}


def _save_wce(data):
    with _wce_filelock:
        _atomic_write_json(WCE_FILE, data)


# ─── 환율 중앙 관리 ─────────────────────────────────────────────────────────

def _load_fx_rates():
    """fx_rates.json 전체 로드. 없으면 빈 dict."""
    if not FX_RATES_FILE.exists():
        return {}
    try:
        with _fx_rates_filelock:
            with open(FX_RATES_FILE, 'r', encoding='utf-8') as fp:
                return json.load(fp)
    except Exception as e:
        print(f'[경고] FX rates 로드 실패: {e}')
        return {}


def _save_fx_rates(data):
    with _fx_rates_filelock:
        _atomic_write_json(FX_RATES_FILE, data)


def _get_current_fx_for_period(year):
    """{currency: {spot, avg}} 반환 (당기 환율)."""
    return (_load_fx_rates().get(year) or {}).get('current') or {}


def _get_prior_fx_for_period(year):
    """{currency: {spot, avg}} 반환 (전기 환율).
    - WCE_FIRST_YEAR: fx_rates.json에 수동 저장된 'prior' 섹션 사용
    - 그 외: 전년도 4Q의 'current' 섹션을 자동으로 가져옴
    """
    fx = _load_fx_rates()
    if year == WCE_FIRST_YEAR:
        return (fx.get(year) or {}).get('prior') or {}
    prior_year = _prior_q4_period(year)
    if prior_year:
        return (fx.get(prior_year) or {}).get('current') or {}
    return {}


def _get_currencies_from_packages(year):
    """업로드된 패키지에서 해당 연도의 비-KRW 통화 목록 수집."""
    currencies = set()
    for f in uploaded_files:
        if f.get('year') != year:
            continue
        cur = (f.get('extracted') or {}).get('currency')
        if cur and cur.upper() != 'KRW':
            currencies.add(cur.upper())
    return sorted(currencies)


def _get_central_rates_for(year, currency):
    """특정 (연도, 통화)의 중앙 환율 dict 반환.

    반환: {'avg', 'spot_current', 'spot_prior', 'avg_prior'} 또는 None
    - 통화가 KRW면 None (환산 불필요, 1.0)
    - 중앙 환율에 아무것도 입력 안 됐으면 None (패키지 환율 사용)
    """
    if not currency or currency.upper() == 'KRW':
        return None
    cur_key = currency.upper()
    cur = _get_current_fx_for_period(year).get(cur_key) or {}
    pri = _get_prior_fx_for_period(year).get(cur_key) or {}
    rates = {
        'avg':          cur.get('avg'),
        'spot_current': cur.get('spot'),
        'spot_prior':   pri.get('spot'),
        'avg_prior':    pri.get('avg'),
    }
    if any(v is not None for v in rates.values()):
        return rates
    return None


# ─── 회사 마스터(업로드 대상 회사 + 회사별 통화) ──────────────────────────────
# 업로드 대상 회사 목록과 각 회사가 사용하는 통화를 한 곳에서 관리한다.
# 환율 값 자체는 기존 fx_rates.json(통화별·기간별)에 그대로 보관하고,
# 회사 → 통화 → 환율로 연결한다 (같은 통화 회사는 같은 환율을 공유).
COMPANY_MASTER_FILE = Path('company_master.json')
_company_master_filelock = threading.Lock()


def _seed_company_master():
    """company_master.json 최초 생성용 시드.
    1) 필수제출회사리스트.xlsx 의 회사명으로 목록 구성
    2) 회사별 통화는 업로드된 패키지(extracted.currency)에서 추정 (없으면 빈값)
    """
    cur_by_norm = {}
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or ''):
        co = (f.get('company') or '').strip()
        cur = ((f.get('extracted') or {}).get('currency') or '').strip().upper()
        if co and cur:
            cur_by_norm[_norm_company_name(co)] = cur     # 최신 업로드가 덮어씀

    names, seen = [], set()
    req_file = Path('필수제출회사리스트.xlsx')
    if req_file.exists():
        try:
            from openpyxl import load_workbook as _lw
            wb = _lw(str(req_file), data_only=True, read_only=True)
            for row in wb.active.iter_rows(values_only=True):
                co = str(row[0]).strip() if row and row[0] not in (None, '') else ''
                if not co or 'COMPANY' in co.upper():
                    continue
                key = _norm_company_name(co)
                if key and key not in seen:
                    seen.add(key)
                    names.append(co)
            wb.close()
        except Exception:
            pass

    companies = [
        {'name': n, 'currency': cur_by_norm.get(_norm_company_name(n), ''), 'active': True}
        for n in names
    ]
    return {'companies': companies}


def _save_company_master(companies):
    """회사 마스터 저장. companies: [{name, currency, active, since}]. 중복 회사명은 병합.
    since: 'YYYY-NQ' — 이 분기부터 제출대상(마감현황)에 포함. 비우면 항상 포함.
    """
    clean, seen = [], set()
    for c in (companies or []):
        if not isinstance(c, dict):
            continue
        name = str(c.get('name') or '').strip()
        if not name:
            continue
        key = _norm_company_name(name)
        if key in seen:
            continue
        seen.add(key)
        since = str(c.get('since') or '').strip().upper()
        if since and not PERIOD_RE.match(since):
            since = ''   # 형식(YYYY-NQ) 아니면 무시
        clean.append({
            'name': name,
            'currency': str(c.get('currency') or '').strip().upper(),
            'active': bool(c.get('active', True)),
            'since': since,
        })
    with _company_master_filelock:
        _atomic_write_json(COMPANY_MASTER_FILE, {'companies': clean})
    return clean


def _load_company_master():
    """회사 마스터 로드. 파일이 없으면 시드를 만들어 저장한 뒤 반환."""
    if not COMPANY_MASTER_FILE.exists():
        seeded = _seed_company_master()
        try:
            _save_company_master(seeded.get('companies') or [])
        except Exception:
            pass
        return seeded
    try:
        with _company_master_filelock:
            with open(COMPANY_MASTER_FILE, 'r', encoding='utf-8') as fp:
                data = json.load(fp) or {}
        if not isinstance(data, dict):
            data = {}
        data.setdefault('companies', [])
        return data
    except Exception as e:
        print(f'[경고] 회사 마스터 로드 실패: {e}')
        return {'companies': []}


def _company_required_names(active_only=True):
    """업로드 대상(필수제출) 회사명 목록 — 마스터 기준."""
    out = []
    for c in _load_company_master().get('companies') or []:
        if active_only and not c.get('active', True):
            continue
        n = (c.get('name') or '').strip()
        if n:
            out.append(n)
    return out


def _company_currency_map():
    """{norm_company_name: currency} — 마스터에 지정된 회사별 통화."""
    m = {}
    for c in _load_company_master().get('companies') or []:
        n = _norm_company_name(c.get('name') or '')
        cur = (c.get('currency') or '').strip().upper()
        if n and cur:
            m[n] = cur
    return m


def _get_wce_for(year, company):
    """특정 (연도, 회사)의 입력값 반환. 없으면 빈 구조.
    첫해(WCE_FIRST_YEAR)가 아니면 각 테이블의 '기초금액'을 전년 4분기 기말금액으로 덮어쓴다.
    또한 prior_meta(전년 Q4 입력 상태)도 함께 반환.
    """
    data = _load_wce()
    rec = data.get(_wce_key(year, company)) or {}
    tables = rec.get('tables') or wce_empty_overrides()

    is_first_year = (year == WCE_FIRST_YEAR)
    prior_year = _prior_q4_period(year)
    prior_meta = None

    # 신규 편입 회사 감지: 전년 분기에 패키지 파일 자체가 없으면 신규
    is_new_company = False
    if not is_first_year and prior_year:
        target_norm = _norm_company_name(company)
        prior_has_pkg = any(
            f.get('year') == prior_year
            and _norm_company_name(f.get('company', '')) == target_norm
            for f in uploaded_files
        )
        is_new_company = not prior_has_pkg

    # 신규 편입 회사는 prior 자동 채움 스킵 (전년 패키지 파일 없으니 prior 데이터가 신뢰 불가)
    if not is_first_year and prior_year and not is_new_company:
        prior_rec = data.get(_wce_key(prior_year, company))
        if prior_rec:
            prior_meta = {
                'year': prior_year,
                'updated_at': prior_rec.get('updated_at'),
                'updated_by': prior_rec.get('updated_by'),
                'has_data': True,
            }
            prior_tables = prior_rec.get('tables') or {}

            # 전년 시점의 자동 채움(당기순이익 등)을 prior_tables에 덮어씌움
            # → 저장값이 옛 환율 기반이더라도 항상 최신 환율 기준으로 기말 계산
            prior_auto = _wce_auto_re_cells_for(prior_year, company)
            if prior_auto:
                _code_to_tid = {col['code']: str(t['id']) for t in WCE_TABLES for col in t['columns']}
                for code, row_dict in prior_auto.items():
                    tid = _code_to_tid.get(code)
                    if not tid: continue
                    prior_tables.setdefault(tid, {})
                    prior_tables[tid].setdefault(code, {})
                    for row_key, val in row_dict.items():
                        prior_tables[tid][code][row_key] = val

            # 각 테이블의 기초금액을 전년 Q4 기말로 덮어쓰기
            # (단 자동값이 0이면 사용자 저장값 보존 — 빈 prior 데이터로 사용자 입력이 덮어써지는 것 방지)
            for t in WCE_TABLES:
                tid = str(t['id'])
                prior_table_data = prior_tables.get(tid) or {}
                endings = _compute_table_ending(t, prior_table_data)
                tables.setdefault(tid, {})
                for col in t['columns']:
                    code = col['code']
                    tables[tid].setdefault(code, {row['key']: 0 for row in t['rows']})
                    auto_beginning = endings.get(code, 0)
                    if auto_beginning:
                        tables[tid][code]['기초금액'] = auto_beginning
                    # else: 자동값 0이면 saved 또는 사용자 입력값 보존

            # 5번 이익잉여금 특별 처리:
            #   당기순이익(3500105)은 연도 마감 시 Unappropriated R/E(3500104)로 이체
            #   → 차기 3500104 기초 = 전년 3500104 기말 + 전년 3500105 기말
            #   → 차기 3500105 기초 = 0
            if '5' in tables:
                prior_t5 = prior_tables.get('5') or {}
                t5_endings = _compute_table_ending(wce_get_table(5), prior_t5)
                merged_104 = (t5_endings.get('3500104', 0) or 0) + (t5_endings.get('3500105', 0) or 0)
                t5_def = wce_get_table(5)
                tables['5'].setdefault('3500104', {row['key']: 0 for row in t5_def['rows']})
                tables['5'].setdefault('3500105', {row['key']: 0 for row in t5_def['rows']})
                if merged_104:
                    tables['5']['3500104']['기초금액'] = merged_104
                tables['5']['3500105']['기초금액'] = 0  # Current Net Income 기초는 항상 0
        else:
            prior_meta = {'year': prior_year, 'has_data': False}

    # 자동 채우기: 5번 이익잉여금(당기순이익/보험수리적손익/R/E조정) + 6번 비지배지분(당기순이익/보험수리적손익)
    auto_re = _wce_auto_re_cells_for(year, company)
    auto_cells = []  # [(table_id_str, code, row_key)] — 템플릿에서 readonly 표시용
    if auto_re:
        # 코드 → 테이블 ID 매핑 (스키마에서 동적으로 구성)
        code_to_tid = {}
        for t in WCE_TABLES:
            for col in t['columns']:
                code_to_tid[col['code']] = str(t['id'])

        # 자동 셀 덮어쓰기
        for code, row_dict in auto_re.items():
            tid = code_to_tid.get(code)
            if not tid:
                continue  # 알 수 없는 코드는 스킵
            t = wce_get_table(int(tid))
            tables.setdefault(tid, {})
            tables[tid].setdefault(code, {row['key']: 0 for row in t['rows']})
            for row_key, val in row_dict.items():
                tables[tid][code][row_key] = val
                auto_cells.append((tid, code, row_key))

    # 로컬(현지통화) 값 + 환율 정보
    local_info = _wce_local_full_for(year, company)
    auto_re_meta = {
        'has_data': bool(auto_re),
        'company': company,
        'year': year,
        'source_file': local_info.get('source_file') if auto_re else None,
    }

    # 셀별 로컬값 lookup 사전 (template에서 빠른 조회)
    local_lookup = {}
    if local_info['local']:
        for t in WCE_TABLES:
            tid = str(t['id'])
            for col in t['columns']:
                for row in t['rows']:
                    v = _lookup_local(local_info['local'], t['id'], col['code'], row['key'])
                    if v:
                        local_lookup[f'{tid}:{col["code"]}:{row["key"]}'] = v

    # 미입력 검출 (로컬 != 0, 환산값 == 0)
    missing = _compute_wce_missing(local_info['local'], local_info['fx_avg'], tables)

    return {
        'tables': tables,
        'summary': rec.get('summary') or {},
        'meta': ({'updated_at': rec.get('updated_at'),
                  'updated_by': rec.get('updated_by')} if rec else None),
        'is_first_year': is_first_year,
        'is_new_company': is_new_company,
        'prior_meta': prior_meta,
        'auto_cells': auto_cells,
        'auto_re_meta': auto_re_meta,
        'local_lookup': local_lookup,
        'currency': local_info.get('currency'),
        'fx_avg': local_info.get('fx_avg'),
        'missing': missing,
        'has_local': bool(local_info['local']),
    }


@app.route('/years', methods=['GET'])
@login_required
def list_years():
    return jsonify({
        'years': YEARS_DATA['years'],
        'default': YEARS_DATA.get('default'),
        'locked': YEARS_DATA.get('locked', []),
    })


def _is_locked(year):
    return year in (YEARS_DATA.get('locked') or [])


@app.route('/years/<year>/lock', methods=['POST'])
@require_permission('years.manage')
def lock_year(year):
    if year not in YEARS_DATA['years']:
        return jsonify({'error': '존재하지 않는 결산기간입니다.'}), 404
    locked = YEARS_DATA.setdefault('locked', [])
    if year not in locked:
        locked.append(year)
        _save_years()
    return jsonify({'ok': True, 'year': year, 'locked': True})


@app.route('/years/<year>/unlock', methods=['POST'])
@require_permission('years.manage')
def unlock_year(year):
    if year not in YEARS_DATA['years']:
        return jsonify({'error': '존재하지 않는 결산기간입니다.'}), 404
    locked = YEARS_DATA.setdefault('locked', [])
    if year in locked:
        locked.remove(year)
        _save_years()
    return jsonify({'ok': True, 'year': year, 'locked': False})


@app.route('/years', methods=['POST'])
@require_permission('years.manage')
def add_year():
    year_str = (request.form.get('year') or '').strip()
    quarter_str = (request.form.get('quarter') or '').strip()

    # 1) year+quarter 방식
    if year_str and quarter_str:
        if not re.fullmatch(r'\d{4}', year_str):
            return jsonify({'error': '연도는 4자리 숫자여야 합니다.'}), 400
        if quarter_str not in ('1', '2', '3', '4'):
            return jsonify({'error': '분기는 1~4 중 선택해야 합니다.'}), 400
        period = f'{year_str}-{quarter_str}Q'
    # 2) period 한 번에 전달하는 방식
    else:
        period = year_str
        if not PERIOD_RE.match(period):
            return jsonify({'error': '형식 오류: YYYY-NQ (예: 2024-1Q)'}), 400

    if period in YEARS_DATA['years']:
        return jsonify({'error': '이미 존재하는 결산기간입니다.'}), 400
    YEARS_DATA['years'].append(period)
    YEARS_DATA['years'].sort(key=_period_sort_key, reverse=True)
    if not YEARS_DATA.get('default'):
        YEARS_DATA['default'] = period
    _save_years()
    return jsonify({'ok': True, 'years': YEARS_DATA['years'], 'default': YEARS_DATA['default']})


@app.route('/years/<year>', methods=['DELETE'])
@require_permission('years.manage')
def delete_year(year):
    if year not in YEARS_DATA['years']:
        return jsonify({'error': '존재하지 않는 연도입니다.'}), 404
    if any(e.get('year') == year for e in uploaded_files):
        return jsonify({'error': '해당 연도에 업로드된 파일이 있어 삭제할 수 없습니다.'}), 400
    YEARS_DATA['years'].remove(year)
    if YEARS_DATA.get('default') == year:
        YEARS_DATA['default'] = YEARS_DATA['years'][0] if YEARS_DATA['years'] else None
    _save_years()
    return jsonify({'ok': True, 'years': YEARS_DATA['years'], 'default': YEARS_DATA.get('default')})


def _valid_year(y):
    return bool(y) and y in YEARS_DATA['years']


def _valid_fx_year(y):
    """환율 관리 전용 검증.
    결산기간(years) ∪ 비교용 보조분기(fx_only_years) 둘 다 허용.
    보조분기는 환율 입력만 가능하며 결산 흐름에는 노출되지 않는다.
    """
    if not y:
        return False
    return y in YEARS_DATA.get('years', []) or y in YEARS_DATA.get('fx_only_years', [])


def _norm_company_name(value):
    return re.sub(r'[\W_]+', '', str(value or '').casefold(), flags=re.UNICODE)


# ─── 상태 저장/복원 ──────────────────────────────────────────────────────────

def _atomic_write_json(path: Path, data, **dump_kw):
    """JSON을 임시 파일에 먼저 쓴 뒤 원자적으로 교체 (쓰기 도중 파일 손상 방지)."""
    tmp = path.with_suffix('.tmp')
    with open(tmp, 'w', encoding='utf-8') as fp:
        json.dump(data, fp, ensure_ascii=False, indent=2, default=str, **dump_kw)
    tmp.replace(path)  # os.replace() — 원자적 파일 교체


def _save_state_sync():
    """동기 저장 — 종료 직전 flush, 또는 백그라운드 쓰레드에서만 호출.

    - threading.Lock  : 동일 프로세스 내 동시 요청 직렬화
    - FileLock(.lock) : 복수 프로세스 간 배타 잠금 (timeout=10s)
    - 원자적 쓰기    : .tmp → rename 으로 부분 쓰기(JSON 손상) 원천 차단
    """
    with _state_threadlock:
        try:
            # uploaded_files는 mutable이므로 dump 도중 변경되지 않도록 얕은 복사
            snapshot = list(uploaded_files)
        except Exception as e:
            print(f'[경고] 상태 스냅샷 실패: {e}')
            return
    try:
        with _state_filelock:
            _atomic_write_json(STATE_FILE, snapshot)
    except Exception as e:
        print(f'[경고] 상태 저장 실패: {e}')


# 백그라운드 상태 저장 — 30MB JSON 쓰기(~2.5s)가 업로드 응답을 막지 않도록 비동기 처리.
# 다수의 dirty 신호는 한 번의 write로 합쳐진다 (coalescing).
_state_dirty = threading.Event()
_state_writer_started = False
_state_writer_lock = threading.Lock()


def _state_writer_loop():
    while True:
        _state_dirty.wait()
        # 짧은 coalesce 윈도우: 연속 업로드 요청들의 dirty 신호를 한 번에 처리
        time.sleep(0.4)
        _state_dirty.clear()
        _save_state_sync()


def _ensure_state_writer():
    global _state_writer_started
    if _state_writer_started:
        return
    with _state_writer_lock:
        if _state_writer_started:
            return
        t = threading.Thread(target=_state_writer_loop,
                             daemon=True, name='state-writer')
        t.start()
        _state_writer_started = True


def _save_state():
    """비차단 — dirty 신호만 세팅하고 즉시 반환. 백그라운드 쓰레드가 디스크에 flush."""
    _ensure_state_writer()
    _state_dirty.set()


# 인터프리터 종료 시 마지막 flush 보장
import atexit as _atexit
@_atexit.register
def _flush_state_on_exit():
    if _state_dirty.is_set():
        try:
            _save_state_sync()
        except Exception:
            pass


def _load_state():
    """디스크에서 업로드 목록을 복원. 실제 파일이 없어진 항목은 제거.
    'year' 필드가 없는 기존 항목은 기본 연도로 마이그레이션.
    """
    global uploaded_files
    if not STATE_FILE.exists():
        return
    try:
        with _state_filelock:
            with open(STATE_FILE, 'r', encoding='utf-8') as fp:
                data = json.load(fp)
        # 1차 필터: 디스크에 실제 파일이 존재하는 entry만
        on_disk = [e for e in data if Path(e.get('path', '')).exists()]
        dropped_missing_file = len(data) - len(on_disk)

        # 2차 필터: year 필드 검증 + 마이그레이션
        # — 등록된 결산기간(_valid_year)만 통과시킨다. 비어 있으면 default로 채우고,
        #   default 자체도 _valid_year 통과해야만 사용.
        # — 등록되지 않은 year(예: 옛 시기 default였다가 지금은 삭제된 분기, 오타 등)는
        #   격리해서 uploaded_files에 적재하지 않고 경고 로그만 남긴다.
        raw_default = YEARS_DATA.get('default') or ''
        default_year = raw_default if _valid_year(raw_default) else ''

        valid = []
        quarantined = []
        migrated = False
        for e in on_disk:
            yr = e.get('year') or ''
            if not yr:
                # year 비어 있으면 default로 보정 (단, default가 유효할 때만)
                if default_year:
                    e['year'] = default_year
                    migrated = True
                else:
                    quarantined.append(e)
                    continue
            elif not _valid_year(yr):
                # 등록되지 않은 year — 격리
                quarantined.append(e)
                continue

            # uploaded_at 보완
            if not e.get('uploaded_at'):
                try:
                    e['uploaded_at'] = datetime.fromtimestamp(
                        Path(e.get('path', '')).stat().st_mtime
                    ).strftime('%Y-%m-%d %H:%M:%S')
                except Exception:
                    e['uploaded_at'] = ''
                migrated = True

            valid.append(e)

        uploaded_files = valid
        print(f'[시작] 저장된 파일 {len(uploaded_files)}개 복원 완료')
        if dropped_missing_file:
            print(f'[시작] 디스크에 파일 없음으로 제외된 entry: {dropped_missing_file}개')
        if quarantined:
            print(f'[경고] 등록된 결산기간에 없는 year의 entry {len(quarantined)}개 격리됨 '
                  f'(uploaded_files에서 제외, 파일은 디스크에 그대로 유지):')
            for e in quarantined[:20]:
                print(f'  - year={e.get("year")!r} company={e.get("company")!r} '
                      f'file={e.get("original_name")!r} id={e.get("id")!r}')
            if len(quarantined) > 20:
                print(f'  ... (그 외 {len(quarantined)-20}건)')
        # 변경된 게 있으면 백그라운드 저장 트리거
        if dropped_missing_file or quarantined or migrated:
            _save_state()
    except Exception as e:
        print(f'[경고] 상태 복원 실패: {e}')
        uploaded_files = []


def allowed_file(filename):
    return Path(filename).suffix.lower() in ALLOWED_EXT


def safe_storage_name(original_name: str) -> str:
    """한글 등 유니코드 파일명도 안전하게 저장하기 위한 정리 (경로 분리자·제어문자 제거)."""
    name = Path(original_name).name
    name = re.sub(r'[\\/:*?"<>|\x00-\x1f]', '_', name)
    if len(name) > 120:
        stem, dot, ext = name.rpartition('.')
        name = (stem[:100] or 'file') + (dot + ext if dot else '')
    return name or 'file.xlsm'


# ─── 라우트 ──────────────────────────────────────────────────────────────────

def _sidebar_perms(uname):
    """사이드바·버튼 가시성 제어용 권한 플래그 묶음."""
    return {
        'users_manage':     _has_permission(uname, 'users.manage'),
        'years_manage':     _has_permission(uname, 'years.manage'),
        'wce_manage':       _has_permission(uname, 'wce.manage'),
        'fx_manage':        _has_permission(uname, 'fx.manage'),
        'consol_compute':   _has_permission(uname, 'consol.compute'),
        'consol_journal':   _has_permission(uname, 'consol.journal'),
        'files_upload':     _has_permission(uname, 'files.upload'),
        'files_delete':     _has_permission(uname, 'files.delete'),
        'files_reanalyze':  _has_permission(uname, 'files.reanalyze'),
        'aggregate_run':    _has_permission(uname, 'aggregate.run'),
        'package_verify':   _has_permission(uname, 'package.verify'),
        'note_aggregate':   _has_permission(uname, 'note.aggregate'),
        'cash_compute':     _has_permission(uname, 'cash.compute'),
        'cash_mapping':     _has_permission(uname, 'cash.mapping'),
        'distribute_run':   _has_permission(uname, 'distribute.run'),
        'distribute_admin': _has_permission(uname, 'distribute.admin'),
        'coa_audit':        _has_permission(uname, 'coa.audit'),
    }


def _closing_status_data(period):
    """마감현황 대시보드용 집계 — 순수 조회(읽기)만 수행, 상태를 바꾸지 않음.
    1) 환율 입력 여부  2) 패키지 제출/미제출  3) 연결조정분개 업로드 현황(연결그룹별)
    """
    # 2) 환율 입력 여부 — 해당 결산기간의 '당기' 환율이 하나라도 있으면 입력됨
    fx_all = _load_fx_rates()
    fx_entered = bool((fx_all.get(period) or {}).get('current'))

    # 3) 패키지 제출 현황 — 회사 마스터(active) 대비 업로드 여부
    #    '적용 시작 분기(since)'가 있으면 그 분기부터만 제출대상에 포함 (문자열 비교: YYYY-NQ 고정폭)
    master = _load_company_master().get('companies') or []
    active = []
    for c in master:
        if not c.get('active', True):
            continue
        since = (c.get('since') or '').strip()
        if since and period and period < since:
            continue   # 아직 적용 시작 전 → 제출대상 제외
        active.append(c)
    submitted_norms = {
        _norm_company_name(f.get('company'))
        for f in uploaded_files if f.get('year') == period
    }
    company_rows, missing = [], []
    for c in active:
        nm = (c.get('name') or '').strip()
        if not nm:
            continue
        ok = _norm_company_name(nm) in submitted_norms
        company_rows.append({'name': nm, 'submitted': ok})
        if not ok:
            missing.append(nm)
    company_rows.sort(key=lambda r: (r['submitted'], r['name']))   # 미제출 먼저
    submitted_count = sum(1 for r in company_rows if r['submitted'])

    # 4) 연결조정분개 업로드 현황 — 연결그룹별.
    #    업로드 원본파일(JOURNAL_DIR)은 파싱 후 남지 않을 수 있으므로,
    #    실제 분개 데이터(consol_journals.json)에 분개가 들어있는지로 판정한다.
    journal_rows = []
    for g in consol_list_groups():
        rec = consol_get_journal(g['id'], period) or {}
        has = bool(rec.get('adjustment_entries') or rec.get('intercompany_entries'))
        journal_rows.append({
            'name': g.get('name') or g.get('id'),
            'uploaded': has,
        })
    journal_rows.sort(key=lambda r: (r['uploaded'], r['name']))
    journal_done = sum(1 for r in journal_rows if r['uploaded'])

    return {
        'fx_entered':      fx_entered,
        'company_rows':    company_rows,
        'missing_companies': missing,
        'missing_count':   len(missing),
        'submitted_count': submitted_count,
        'total_companies': len(company_rows),
        'journal_rows':    journal_rows,
        'journal_done':    journal_done,
        'journal_total':   len(journal_rows),
    }


@app.route('/')
@login_required
def index():
    """메인페이지 — 마감현황 대시보드."""
    uname = session.get('username')
    period = YEARS_DATA.get('default') or (YEARS_DATA['years'][0] if YEARS_DATA.get('years') else None)
    data = _closing_status_data(period) if period else {
        'fx_entered': False, 'company_rows': [], 'missing_companies': [], 'missing_count': 0,
        'submitted_count': 0, 'total_companies': 0, 'journal_rows': [], 'journal_done': 0, 'journal_total': 0,
    }
    return render_template('closing_status.html',
                           username=uname,
                           is_admin=_is_admin(uname),
                           perms=_sidebar_perms(uname),
                           years=YEARS_DATA.get('years', []),
                           current_period=period,
                           locked_years=YEARS_DATA.get('locked', []),
                           **data)


@app.route('/closing-status/current-period', methods=['POST'])
@require_permission('years.manage')
def set_current_period():
    """현재 결산기간(기본 결산기간) 변경 — 관리자(years.manage) 전용."""
    body = request.get_json(silent=True) or {}
    period = (body.get('period') or request.form.get('period') or '').strip()
    if period not in YEARS_DATA.get('years', []):
        return jsonify({'error': '존재하지 않는 결산기간입니다.'}), 400
    YEARS_DATA['default'] = period
    _save_years()
    return jsonify({'ok': True, 'default': period})


@app.route('/package-upload')
@login_required
def package_upload_page():
    """패키지 업로드 — 기존 메인페이지(업로드/합산/파일목록) 내용."""
    uname = session.get('username')
    return render_template('index.html', files=uploaded_files,
                           username=uname,
                           is_admin=_is_admin(uname),
                           perms=_sidebar_perms(uname),
                           years=YEARS_DATA['years'],
                           default_year=YEARS_DATA.get('default'),
                           locked_years=YEARS_DATA.get('locked', []))


@app.route('/upload', methods=['POST'])
@login_required
@require_permission('files.upload')
def upload():
    if 'files' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400

    year = (request.args.get('year') or request.form.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산연도를 선택해주세요.'}), 400

    if _is_locked(year) and not _is_admin(session.get('username')):
        return jsonify({
            'error': f'{year} 결산기간은 마감되어 업로드할 수 없습니다. (관리자 문의)'
        }), 403

    results = []
    files = request.files.getlist('files')

    for f in files:
        save_path = None
        try:
            if not f.filename:
                continue
            if not allowed_file(f.filename):
                results.append({'name': f.filename, 'error': '지원하지 않는 파일 형식'})
                continue

            uid = str(uuid.uuid4())[:8]
            safe_name = safe_storage_name(f.filename)
            save_path = UPLOAD_DIR / f'{uid}_{safe_name}'

            f.save(str(save_path))

            # ─── 무결성 토큰 검증 (정식 배포 파일 여부) ───
            # 기본은 모두 차단. 관리자가 force_upload=true 로 명시적으로 켜야만 우회.
            is_admin_user = _is_admin(session.get('username'))
            force_upload = (str(request.form.get('force_upload') or
                                request.args.get('force_upload') or '').lower()
                            in ('1', 'true', 'yes', 'on'))
            allow_bypass = is_admin_user and force_upload
            bypass_warning = None
            tok = verify_dist_token(str(save_path))
            if not tok['ok']:
                msg_map = {
                    'token_missing':   '정식 배포 파일이 아닙니다. 사이드바 "배포용 생성"에서 받은 파일을 업로드해주세요.',
                    'sig_invalid':     '파일이 변조되었거나 다른 시스템에서 생성된 산출물입니다.',
                    'payload_corrupt': '토큰 데이터가 손상되었습니다. 새로 발급받아 주세요.',
                }
                err_msg = msg_map.get(tok.get('reason'), '토큰 검증 실패')
                if not allow_bypass:
                    save_path.unlink(missing_ok=True)
                    results.append({'name': f.filename, 'error': err_msg})
                    continue
                bypass_warning = f'토큰 검증 우회(관리자 강제): {err_msg}'
            else:
                # 토큰 유효 → 분기 일치 확인
                tok_year = (tok['payload'] or {}).get('year')
                if tok_year and tok_year != year:
                    err_msg = (f'이 파일은 {tok_year} 배포본입니다. {year}용 파일을 받아주세요. '
                               f'(옛 분기 파일 재활용 차단)')
                    if not allow_bypass:
                        save_path.unlink(missing_ok=True)
                        results.append({'name': f.filename, 'error': err_msg})
                        continue
                    bypass_warning = f'분기 불일치 우회(관리자 강제): {err_msg}'

            # 중앙 관리 환율이 입력돼 있으면 우선 적용 (없으면 패키지 환율)
            data = extract(
                str(save_path),
                central_rates_lookup=lambda cur: _get_central_rates_for(year, cur),
            )

            # ─── 토큰의 회사명과 추출된 회사명 일치 검증 ───
            if tok['ok'] and tok.get('payload'):
                tok_company = tok['payload'].get('company') or ''
                extracted_company = data.get('company') or ''
                if (tok_company and extracted_company
                        and _norm_company_name(tok_company) != _norm_company_name(extracted_company)):
                    err_msg = (f'이 파일은 "{tok_company}" 배포본입니다. '
                               f'"{extracted_company}" 으로 업로드할 수 없습니다.')
                    if not allow_bypass:
                        save_path.unlink(missing_ok=True)
                        results.append({'name': f.filename, 'error': err_msg})
                        continue
                    bypass_warning = f'회사 불일치 우회(관리자 강제): {err_msg}'

            # Index!C12 Error 개수 검증
            err_cnt = data.get('index_error_count')
            if err_cnt is not None and err_cnt != 0:
                save_path.unlink(missing_ok=True)
                results.append({
                    'name': f.filename,
                    'error': (f'Error {err_cnt}건 감지. 파일이 제출가능한 상태인지 Cover시트를 확인하세요 '
                              f'(Please check the Cover sheet to see if the file is ready for submission)')
                })
                continue

            # Cover!C9(year) / Cover!F9(quarter) 검증
            sel_m = re.match(r'^(\d{4})-([1-4])Q$', year)
            sel_year = sel_m.group(1) if sel_m else None
            sel_quarter = sel_m.group(2) if sel_m else None
            cy, cq = data.get('cover_year'), data.get('cover_quarter')

            if not cy or not cq:
                save_path.unlink(missing_ok=True)
                results.append({
                    'name': f.filename,
                    'error': f'Cover 시트에서 결산연도(C9)/분기(F9)를 읽을 수 없습니다. (year={cy}, quarter={cq})'
                })
                continue

            if cy != sel_year or cq != sel_quarter:
                save_path.unlink(missing_ok=True)
                results.append({
                    'name': f.filename,
                    'error': (f'선택한 결산기간({sel_year}년 {sel_quarter}분기)과 '
                              f'파일의 Cover 값({cy}년 {cq}분기)이 일치하지 않습니다.')
                })
                continue

            # 담당 회사 권한 확인
            if not _can_access_company(session.get('username'), data['company']):
                save_path.unlink(missing_ok=True)
                results.append({
                    'name': f.filename,
                    'error': f'담당 회사가 아닙니다: {data["company"]} (관리자에게 문의하세요)',
                })
                continue

            target_norm = _norm_company_name(data['company'])
            entry = {
                'id': uid,
                'original_name': f.filename,
                'path': str(save_path),
                'company': data['company'],
                'year': year,
                'uploaded_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'uploaded_by': session.get('username') or '',
                'extracted': data,
                'sheet_summary': {
                    s: len(rows) for s, rows in data['sheets'].items() if rows
                },
                'comment': '',
                'comment_updated_at': '',
                'comment_updated_by': '',
            }
            # 전역 uploaded_files 의 읽기-수정-쓰기는 threaded 서버에서 경쟁이 나므로
            # 스냅샷 저장과 동일한 _state_threadlock 으로 직렬화한다 (lost update 방지).
            replaced_old = []
            with _state_threadlock:
                # 동일 (회사, 연도) 기존 업로드가 있으면 자동 교체 (중복 누적 방지)
                # 교체 시 기존 코멘트는 새 entry로 인계 — 자회사가 수정본을 올려도 코멘트 보존.
                carry_comment = None  # 기존 코멘트 인계용
                new_uploads_list = []
                for old in uploaded_files:
                    if (old.get('year') == year
                            and _norm_company_name(old.get('company', '')) == target_norm):
                        # 코멘트가 있으면 인계 (가장 최근 매칭 1건만 사용)
                        if old.get('comment'):
                            carry_comment = {
                                'comment': old.get('comment') or '',
                                'comment_updated_at': old.get('comment_updated_at') or '',
                                'comment_updated_by': old.get('comment_updated_by') or '',
                            }
                        # 디스크 파일 정리
                        try:
                            Path(old.get('path', '')).unlink(missing_ok=True)
                        except Exception:
                            pass
                        replaced_old.append(old.get('original_name'))
                    else:
                        new_uploads_list.append(old)
                if carry_comment:
                    entry.update(carry_comment)
                uploaded_files[:] = new_uploads_list
                uploaded_files.append(entry)
            _save_state()

            # 부호·정합성 경고 — 차단은 아니고 사용자에게 알려서 확인하도록
            try:
                sign_warnings = validate_local_vs_value_signs(data)
            except Exception:
                sign_warnings = []

            results.append({
                'id': uid,
                'name': f.filename,
                'company': data['company'],
                'replaced': replaced_old,    # 교체된 구파일 이름 (UI에 표시 가능)
                'currency': data.get('currency'),
                'fx_rate': data.get('fx_rate'),
                'sheets': entry['sheet_summary'],
                # 토큰 검증 우회한 경우 사용자에게 경고 노출 (관리자만 발생)
                **({'warning': bypass_warning} if bypass_warning else {}),
                # 현지통화 vs 환산값 부호/정합성 경고 (있으면 모달에서 ⚠로 표시)
                **({'sign_warnings': sign_warnings} if sign_warnings else {}),
            })
        except Exception as e:
            err_text = f'{type(e).__name__}: {e}'
            print(f'[업로드 실패] {getattr(f, "filename", "?")}\n{traceback.format_exc()}',
                  file=sys.stderr, flush=True)
            if save_path is not None:
                try:
                    save_path.unlink(missing_ok=True)
                except Exception:
                    pass
            results.append({'name': getattr(f, 'filename', '알 수 없음'), 'error': err_text})

    return jsonify({'uploaded': results, 'total': len(uploaded_files)})


@app.route('/groups')
@login_required
def get_groups():
    """그룹정보.xlsx를 읽어 {그룹명: [회사명, ...]} 형태로 반환."""
    group_file = Path('그룹정보.xlsx')
    if not group_file.exists():
        return jsonify({'groups': {}})
    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(str(group_file), data_only=True, read_only=True)
        ws = wb.active
        groups = {}
        header_found = False
        for row in ws.iter_rows(values_only=True):
            # 헤더 행 탐색: A열에 'COMPANY' 포함된 행
            if not header_found:
                if row[0] and 'COMPANY' in str(row[0]).upper():
                    header_found = True
                continue
            company = str(row[0]).strip() if row[0] not in (None, '') else ''
            group   = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, '') else ''
            if company and group:
                groups.setdefault(group, []).append(company)
        wb.close()
        return jsonify({'groups': groups})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/submission-status')
@login_required
def submission_status():
    year = request.args.get('year')
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산기간을 선택해주세요.'}), 400

    # 업로드 대상 회사목록은 회사 마스터(company_master.json) 기준.
    # (최초 1회 필수제출회사리스트.xlsx 에서 자동 시드됨)
    required = _company_required_names(active_only=True)

    uploaded_by_company = {}
    for f in uploaded_files:
        if f.get('year') != year:
            continue
        key = _norm_company_name(f.get('company'))
        if not key:
            continue
        uploaded_by_company.setdefault(key, []).append({
            'company': f.get('company', ''),
            'name': f.get('original_name', ''),
            'uploaded_at': f.get('uploaded_at', ''),
        })

    submitted_items = []
    missing_items = []
    for company in required:
        uploads = uploaded_by_company.get(_norm_company_name(company), [])
        # 회사별 업로드 목록도 최신순으로 정렬
        uploads.sort(key=lambda u: u.get('uploaded_at') or '', reverse=True)
        entry = {
            'company': company,
            'submitted': bool(uploads),
            'uploads': uploads,
        }
        if uploads:
            submitted_items.append(entry)
        else:
            missing_items.append(entry)

    # 제출 회사: 가장 최근 업로드일시 기준 내림차순
    submitted_items.sort(
        key=lambda it: it['uploads'][0].get('uploaded_at') or '',
        reverse=True,
    )
    # 미제출 회사는 필수리스트 순서 유지 (이미 그 순서로 append됨)
    items = submitted_items + missing_items
    submitted_count = len(submitted_items)

    return jsonify({
        'year': year,
        'total': len(items),
        'submitted': submitted_count,
        'missing': len(items) - submitted_count,
        'items': items,
    })


@app.route('/files', methods=['GET'])
@login_required
def list_files():
    year = request.args.get('year')
    username = session.get('username')
    # 최신 업로드가 위로 오도록 uploaded_at 내림차순 정렬
    visible = [
        f for f in uploaded_files
        if (not year or f.get('year') == year)
        and _can_access_company(username, f.get('company', ''))
    ]
    visible.sort(key=lambda f: f.get('uploaded_at') or '', reverse=True)
    is_admin_user = _is_admin(username)
    return jsonify([
        {
            'id': f['id'],
            'name': f['original_name'],
            'company': f['company'],
            'year': f.get('year'),
            'currency': f['extracted'].get('currency'),
            'fx_rate': f['extracted'].get('fx_rate'),
            'uploaded_at': f.get('uploaded_at', ''),
            'uploaded_by': f.get('uploaded_by', ''),
            'sheets': f['sheet_summary'],
            'comment': f.get('comment', '') or '',
            'comment_updated_at': f.get('comment_updated_at', '') or '',
            'comment_updated_by': f.get('comment_updated_by', '') or '',
            # 클라이언트가 편집 버튼 노출 여부를 결정할 수 있도록 권한 플래그 동봉
            'can_edit_comment': (is_admin_user
                                 or (f.get('uploaded_by') == username and bool(username))),
        }
        for f in visible
    ])


@app.route('/files/reanalyze', methods=['POST'])
@login_required
@require_permission('files.reanalyze')
def reanalyze_all():
    """전체 재분석 (하위호환용 유지)."""
    year = request.args.get('year')
    results = {'ok': [], 'failed': []}
    targets = [e for e in uploaded_files if not year or e.get('year') == year]
    for entry in targets:
        try:
            entry_year = entry.get('year')
            data = extract(
                entry['path'],
                central_rates_lookup=lambda cur, _y=entry_year: _get_central_rates_for(_y, cur),
            )
            entry['extracted'] = data
            entry['company'] = data['company']
            entry['sheet_summary'] = {
                s: len(rows) for s, rows in data['sheets'].items() if rows
            }
            results['ok'].append(entry['original_name'])
        except Exception as e:
            print(f'[재분석 실패] {entry["original_name"]}: {e}',
                  file=sys.stderr, flush=True)
            results['failed'].append({'name': entry['original_name'], 'error': str(e)})
    _save_state()
    return jsonify(results)


@app.route('/files/<uid:uid>/reanalyze', methods=['POST'])
@login_required
@require_permission('files.reanalyze')
def reanalyze_one(uid):
    """단일 파일 재분석 — 프론트엔드 진행 모달용."""
    entry = next((f for f in uploaded_files if f['id'] == uid), None)
    if not entry:
        return jsonify({'error': '파일을 찾을 수 없습니다.'}), 404
    try:
        entry_year = entry.get('year')
        data = extract(
            entry['path'],
            central_rates_lookup=lambda cur: _get_central_rates_for(entry_year, cur),
        )
        entry['extracted'] = data
        entry['company'] = data['company']
        entry['sheet_summary'] = {
            s: len(rows) for s, rows in data['sheets'].items() if rows
        }
        _save_state()
        # 부호·정합성 경고 (업로드와 동일)
        try:
            sign_warnings = validate_local_vs_value_signs(data)
        except Exception:
            sign_warnings = []
        return jsonify({
            'ok': True,
            'company': data['company'],
            'currency': data.get('currency'),
            'fx_rate': data.get('fx_rate'),
            'sheets': entry['sheet_summary'],
            **({'sign_warnings': sign_warnings} if sign_warnings else {}),
        })
    except Exception as e:
        print(f'[재분석 실패] {entry["original_name"]}: {e}',
              file=sys.stderr, flush=True)
        return jsonify({'error': f'{type(e).__name__}: {e}'}), 500


@app.route('/files', methods=['DELETE'])
@require_permission('files.delete')
def delete_all_files():
    global uploaded_files
    year = request.args.get('year')
    if not year:
        return jsonify({'error': '연도를 지정해주세요.'}), 400
    with _state_threadlock:
        to_delete = [f for f in uploaded_files if f.get('year') == year]
        for entry in to_delete:
            Path(entry['path']).unlink(missing_ok=True)
        uploaded_files = [f for f in uploaded_files if f.get('year') != year]
    _save_state()
    return jsonify({'ok': True, 'deleted': len(to_delete)})


@app.route('/files/<uid:uid>/download')
@login_required
def download_uploaded_file(uid):
    entry = next((f for f in uploaded_files if f['id'] == uid), None)
    if not entry:
        return '파일을 찾을 수 없습니다.', 404
    if not _can_access_company(session.get('username'), entry.get('company', '')):
        return '담당 회사가 아닙니다.', 403
    path = Path(entry['path'])
    if not path.exists():
        return '파일이 서버에 존재하지 않습니다.', 404
    return send_file(str(path.resolve()), as_attachment=True,
                     download_name=entry['original_name'])


@app.route('/files/download-zip')
@login_required
def download_uploaded_files_zip():
    """업로드된 패키지를 ZIP으로 일괄 다운로드.

    쿼리:
      · year=YYYY-NQ   (필수)
      · ids=<JSON 배열> 또는 콤마구분 문자열  (선택 — 없으면 해당 기간 전체)

    동작:
      · 담당 회사가 아닌 파일은 자동 스킵 (사용자 권한 범위 안에서만 묶음)
      · 파일이 디스크에 없으면 스킵
      · 같은 이름이 충돌하면 ' (2).xlsm' 식으로 자동 넘버링
    """
    import zipfile
    year = (request.args.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산연도를 선택해주세요.'}), 400

    # ids 파라미터 파싱 — JSON 배열 또는 콤마구분 모두 허용
    raw_ids = (request.args.get('ids') or '').strip()
    id_set = None
    if raw_ids:
        try:
            parsed = json.loads(raw_ids)
            if isinstance(parsed, list):
                id_set = {str(x) for x in parsed}
        except Exception:
            id_set = {x.strip() for x in raw_ids.split(',') if x.strip()}

    username = session.get('username')
    targets = []
    for f in uploaded_files:
        if f.get('year') != year:
            continue
        if id_set is not None and f.get('id') not in id_set:
            continue
        if not _can_access_company(username, f.get('company', '')):
            continue
        p = Path(f['path'])
        if not p.exists():
            continue
        targets.append((f, p))

    if not targets:
        return jsonify({'error': '다운로드 가능한 파일이 없습니다.'}), 404

    buf = io.BytesIO()
    used_names = {}
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for entry, path in targets:
            name = entry.get('original_name') or path.name
            if name in used_names:
                used_names[name] += 1
                stem = Path(name).stem
                suffix = Path(name).suffix
                arc = f"{stem} ({used_names[name]}){suffix}"
            else:
                used_names[name] = 1
                arc = name
            zf.write(str(path.resolve()), arc)

    buf.seek(0)
    safe_year = re.sub(r'[\\/:*?"<>|]', '_', year)
    zip_name = f'패키지_{safe_year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
    return send_file(
        buf, as_attachment=True,
        download_name=zip_name,
        mimetype='application/zip',
    )


@app.route('/files/<uid:uid>', methods=['DELETE'])
@require_permission('files.delete')
def delete_file(uid):
    global uploaded_files
    with _state_threadlock:
        entry = next((f for f in uploaded_files if f['id'] == uid), None)
        if not entry:
            return jsonify({'error': '파일을 찾을 수 없습니다.'}), 404
        Path(entry['path']).unlink(missing_ok=True)
        uploaded_files = [f for f in uploaded_files if f['id'] != uid]
    _save_state()
    return jsonify({'ok': True})


# ─── 업로드 파일 코멘트 (업로드 후 편집·조회) ─────────────────────────────────
# 자회사 담당자가 업로드 후 특이사항을 기록할 수 있도록.
# 권한: 업로드한 본인(uploaded_by) 또는 관리자만 수정 가능. 조회는 누구나.
# 동일 (회사, 연도) 재업로드 시 코멘트는 업로드 핸들러에서 자동 인계됨.

def _can_edit_comment(entry: dict, username: str) -> bool:
    if not entry or not username:
        return False
    if _is_admin(username):
        return True
    return (entry.get('uploaded_by') or '') == username


@app.route('/files/<uid:uid>/comment', methods=['PUT'])
@login_required
def update_file_comment(uid):
    entry = next((f for f in uploaded_files if f['id'] == uid), None)
    if not entry:
        return jsonify({'error': '파일을 찾을 수 없습니다.'}), 404
    username = session.get('username')
    if not _can_access_company(username, entry.get('company', '')):
        return jsonify({'error': '해당 회사 파일에 접근 권한이 없습니다.'}), 403
    if not _can_edit_comment(entry, username):
        return jsonify({'error': '본인이 업로드한 파일이거나 관리자만 코멘트를 수정할 수 있습니다.'}), 403

    body = request.get_json(silent=True) or {}
    new_comment = (body.get('comment') or '').strip()
    # 길이 상한 — 너무 긴 텍스트로 _state.json 부풀리는 것 방지
    if len(new_comment) > 4000:
        return jsonify({'error': '코멘트는 4,000자를 넘을 수 없습니다.'}), 400

    entry['comment'] = new_comment
    entry['comment_updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    entry['comment_updated_by'] = username or ''
    _save_state()
    return jsonify({
        'ok': True,
        'comment': entry['comment'],
        'comment_updated_at': entry['comment_updated_at'],
        'comment_updated_by': entry['comment_updated_by'],
    })


@app.route('/aggregate', methods=['POST'])
@login_required
@require_permission('aggregate.run')
def run_aggregate():
    year = (request.args.get('year') or request.form.get('year') or '').strip()
    if not _valid_year(year):
        return jsonify({'error': '유효한 결산연도를 선택해주세요.'}), 400

    # 선택된 파일 ID 목록 (없으면 전체)
    ids_json = (request.form.get('ids') or request.args.get('ids') or '').strip()
    selected_ids = None
    if ids_json:
        try:
            selected_ids = set(json.loads(ids_json))
        except Exception:
            return jsonify({'error': '파일 선택 정보가 올바르지 않습니다.'}), 400

    files_for_year = [f for f in uploaded_files if f.get('year') == year]
    if selected_ids is not None:
        files_for_year = [f for f in files_for_year if f['id'] in selected_ids]

    if len(files_for_year) < 1:
        return jsonify({'error': '합산할 파일이 없습니다. 파일을 1개 이상 선택해주세요.'}), 400

    # 통화(국가) 기준 정렬: KRW(국내) → 그 외 통화 알파벳순 → 회사명 가나다순
    def _country_sort_key(f):
        cur = ((f.get('extracted') or {}).get('currency') or 'ZZZ').upper()
        co = f.get('company') or ''
        krw_priority = 0 if cur == 'KRW' else 1
        return (krw_priority, cur, co)
    files_for_year.sort(key=_country_sort_key)

    try:
        extracted_list = [f['extracted'] for f in files_for_year]
        agg = aggregate(extracted_list)

        # WCE 자본 항목 대체 + 3400104 차대 조정
        agg = _apply_wce_to_aggregation(agg, year)

        out_name = f'통합합산결과_{year}_{len(files_for_year)}개사.xlsx'
        out_path = RESULTS_DIR / out_name
        write_excel(agg, extracted_list, str(out_path))

        # 회사별 전기 비교 (동기 환율 매칭) — 별도 산출물
        compare_out_name = f'회사별_전기비교_{year}_{len(files_for_year)}개사.xlsx'
        compare_out_path = RESULTS_DIR / compare_out_name
        same_q_rates = _get_same_q_fx_for_period(year)
        prior_same_q_label = _prior_same_q_period(year)
        try:
            build_company_compare(
                year=year,
                files_for_year=files_for_year,
                same_q_rates=same_q_rates,
                prior_period_label=prior_same_q_label,
                output_path=str(compare_out_path),
                aggregated_bs=agg['sheets'].get('BS'),   # WCE 적용된 BS by_company
            )
        except Exception as e:
            print(f'[경고] 회사별 전기비교 엑셀 생성 실패: {e}')
            compare_out_name = None

        companies = agg['companies']
        bs_data  = agg['sheets'].get('BS', {})
        pl_data  = agg['sheets'].get('PL_MF', {})
        cf2_conn = agg['sheets'].get('CF2_연결', {})
        cf3_conn = agg['sheets'].get('CF3_연결', {})

        # 헬퍼: 특정 코드의 회사별+합계+비교합계
        def _pick(sheet, code, fallback_kor='', fallback_eng=''):
            info = sheet.get(code, {})
            return {
                'code': code,
                'kor': info.get('kor', fallback_kor),
                'eng': info.get('eng', fallback_eng),
                'companies': info.get('by_company', {}),
                'total': info.get('total', 0) or 0,
                'compare_total': info.get('compare_total', 0) or 0,
            }

        # ─── BS 미리보기: 자산/부채/자본 총계 + 차입금총계 + 내부/외부차입금 + 부채비율 ───
        # 기본 총계 (윗줄)
        total_assets = _pick(bs_data, '1000000', '자산총계', 'Total Assets')
        total_liab   = _pick(bs_data, '2000000', '부채총계', 'Total Liabilities')
        total_eq     = _pick(bs_data, '3000000', '자본총계', 'Total Equity')
        for row in (total_assets, total_liab, total_eq):
            row['dash_row'] = 0  # 윗줄
        bs_preview = [total_assets, total_liab, total_eq]

        # 차입금총계
        DEBT_CODES = ['2100201', '2100202', '2100203', '2100204', '2100205',
                      '2100301', '2100391', '2200101', '2200201', '2200291']
        debt_by_co     = {c: 0.0 for c in companies}
        debt_cmp_by_co = {c: 0.0 for c in companies}
        debt_total     = 0.0
        debt_cmp_total = 0.0
        for dc in DEBT_CODES:
            info = bs_data.get(dc, {})
            for co, v in (info.get('by_company', {}) or {}).items():
                debt_by_co[co] = debt_by_co.get(co, 0) + (v or 0)
            for co, v in (info.get('compare_by_company', {}) or {}).items():
                debt_cmp_by_co[co] = debt_cmp_by_co.get(co, 0) + (v or 0)
            debt_total     += info.get('total', 0) or 0
            debt_cmp_total += info.get('compare_total', 0) or 0

        # 내부차입금
        #   total        = CF2_연결 + CF3_연결 의 '기말' 행 합산 (당기 기말잔액)
        #   compare_total = CF2_연결 + CF3_연결 의 '기초' 행 합산 (전기 기말 = 당기 기초잔액)
        # CF1/CF2/CF3 키 형식: '{code}::{label}'
        internal_by_co     = {c: 0.0 for c in companies}
        internal_cmp_by_co = {c: 0.0 for c in companies}
        internal_total     = 0.0
        internal_cmp_total = 0.0
        for conn_sheet in (cf2_conn, cf3_conn):
            for key, info in conn_sheet.items():
                label = key.split('::', 1)[1] if '::' in key else ''
                if '기말' in label:
                    for co, v in (info.get('by_company', {}) or {}).items():
                        internal_by_co[co] = internal_by_co.get(co, 0) + (v or 0)
                    internal_total += info.get('total', 0) or 0
                elif '기초' in label:
                    for co, v in (info.get('by_company', {}) or {}).items():
                        internal_cmp_by_co[co] = internal_cmp_by_co.get(co, 0) + (v or 0)
                    internal_cmp_total += info.get('total', 0) or 0

        # 내부차입금 캡 처리: CF3 기초금액이 PY 시트 BS값과 불일치하는 경우
        # (소스 파일 데이터 오류) 내부차입금이 차입금총계를 초과할 수 없도록 보정
        for c in companies:
            # 당기: 내부 ≤ 총차입금
            d_cur = debt_by_co.get(c, 0) or 0
            i_cur = internal_by_co.get(c, 0) or 0
            if i_cur > d_cur:
                internal_by_co[c] = d_cur
            # 전기: 내부 ≤ 총차입금
            d_cmp = debt_cmp_by_co.get(c, 0) or 0
            i_cmp = internal_cmp_by_co.get(c, 0) or 0
            if i_cmp > d_cmp:
                internal_cmp_by_co[c] = d_cmp

        # 캡 적용 후 합계 재계산
        internal_total     = sum(internal_by_co.values())
        internal_cmp_total = sum(internal_cmp_by_co.values())

        # 외부차입금 = 차입금총계 - 내부차입금  (당기/전기 각각 동일 기간끼리 차감)
        external_by_co     = {c: (debt_by_co.get(c, 0) or 0) - (internal_by_co.get(c, 0) or 0)
                               for c in companies}
        external_cmp_by_co = {c: (debt_cmp_by_co.get(c, 0) or 0) - (internal_cmp_by_co.get(c, 0) or 0)
                               for c in companies}
        external_total     = debt_total     - internal_total
        external_cmp_total = debt_cmp_total - internal_cmp_total

        # 차입금 그룹 3장 (가운뎃줄) — chart_group으로 합산 차트 패널에서 하나로 묶임
        bs_preview.append({
            'code': '차입금총계', 'kor': '차입금총계', 'eng': 'Total Borrowings',
            'companies': debt_by_co, 'total': debt_total, 'compare_total': debt_cmp_total,
            'derived': True, 'dash_row': 1, 'chart_group': 'debt_composition',
        })
        bs_preview.append({
            'code': '내부차입금', 'kor': '내부차입금', 'eng': 'Internal Borrowings',
            'companies': internal_by_co, 'total': internal_total, 'compare_total': internal_cmp_total,
            'compare_by_company': internal_cmp_by_co,
            'derived': True, 'dash_row': 1, 'chart_group': 'debt_composition',
        })
        bs_preview.append({
            'code': '외부차입금', 'kor': '외부차입금', 'eng': 'External Borrowings',
            'companies': external_by_co, 'total': external_total, 'compare_total': external_cmp_total,
            'derived': True, 'dash_row': 1, 'chart_group': 'debt_composition',
        })

        # 부채비율 = 부채 / 자본
        liab_total  = bs_preview[1]['total']
        eq_total    = bs_preview[2]['total']
        liab_cmp    = bs_preview[1]['compare_total']
        eq_cmp      = bs_preview[2]['compare_total']
        debt_ratio_by_co = {}
        for co in companies:
            liab_co = bs_data.get('2000000', {}).get('by_company', {}).get(co, 0) or 0
            eq_co   = bs_data.get('3000000', {}).get('by_company', {}).get(co, 0) or 0
            debt_ratio_by_co[co] = (liab_co / eq_co * 100) if eq_co else None
        bs_preview.append({
            'code': '부채비율', 'kor': '부채비율 (부채/자본)',
            'eng': 'Debt Ratio (%)',
            'companies': debt_ratio_by_co,
            'total':         (liab_total / eq_total * 100) if eq_total else None,
            'compare_total': (liab_cmp   / eq_cmp   * 100) if eq_cmp   else None,
            'is_ratio': True, 'derived': True, 'dash_row': 2,  # 맨 아랫줄
        })

        # ─── PL 미리보기 ───
        # (사용자 요청의 45000000은 4500000의 오타로 가정 — 두 키 모두 시도)
        PL_CODES = [
            ('4100000', '매출액',         'Net Sales'),
            ('4200000', '매출원가',       'Cost of Sales'),
            ('4700001', '매출총이익',     'Gross Profit'),
            ('4300000', '판매비와관리비', 'SG&A'),
            ('4700002', '영업이익',       'Operating Income'),
            ('4400000', '영업외수익',     'Non-Op Income'),
            ('4500000', '영업외비용',     'Non-Op Expense'),
            ('4800001', '법인세비용',     'Income Tax'),
            ('4700004', '당기순이익',     'Net Income'),
        ]
        pl_preview = []
        for code, kor, eng in PL_CODES:
            info = pl_data.get(code)
            # 대체 코드 탐색 (예: 4500000 없으면 45000000)
            if info is None and code == '4500000':
                info = pl_data.get('45000000')
            row = {
                'code': code, 'kor': kor, 'eng': eng,
                'companies': (info or {}).get('by_company', {}) or {c: 0 for c in companies},
                'total': (info or {}).get('total', 0) or 0,
                'compare_total': (info or {}).get('compare_total', 0) or 0,
            }
            pl_preview.append(row)

        # ─── PL 기본 집계 ───────────────────────────────────────────────────
        sales = pl_data.get('4100000', {}) or {}
        op    = pl_data.get('4700002', {}) or {}
        ni    = pl_data.get('4700004', {}) or {}

        sales_by = sales.get('by_company', {}) or {}
        op_by    = op.get('by_company', {}) or {}
        ni_by    = ni.get('by_company', {}) or {}

        s_tot, s_cmp = (sales.get('total') or 0), (sales.get('compare_total') or 0)
        o_tot, o_cmp = (op.get('total') or 0),    (op.get('compare_total') or 0)
        n_tot, n_cmp = (ni.get('total') or 0),    (ni.get('compare_total') or 0)

        def _ratio_by_co(num_by, den_by):
            out = {}
            for c in companies:
                n = num_by.get(c, 0) or 0
                d = den_by.get(c, 0) or 0
                out[c] = (n / d * 100) if d else None
            return out

        def _mk(code, kor, eng, by_co, total, cmp_total, **kw):
            return {'code': code, 'kor': kor, 'eng': eng,
                    'companies': by_co, 'total': total, 'compare_total': cmp_total, **kw}

        # ─── 신규 지표: CF 기반 ─────────────────────────────────────────────
        cf_data = agg['sheets'].get('CF', {})

        # 1. 영업활동현금흐름 (CF 시트 "Ⅰ. 영업활동으로 인한 현금흐름" 레이블)
        ocf_info = None
        for _key, _info in cf_data.items():
            if '영업활동' in _info.get('kor', '') and '현금흐름' in _info.get('kor', ''):
                ocf_info = _info; break
        ocf_total = (ocf_info.get('total', 0) or 0) if ocf_info else 0
        ocf_cmp   = (ocf_info.get('compare_total', 0) or 0) if ocf_info else 0
        ocf_by_co = dict(ocf_info.get('by_company', {}) or {}) if ocf_info else {c: 0 for c in companies}

        # 2. 감가상각 합계 (EBITDA 구성 — CF 시트 코드 기반)
        DEPR_CODES = ['4300301', '5300301', '4300302', '5300302', '4300303', '5300303']
        depr_by_co = {c: 0.0 for c in companies}
        depr_total = 0.0
        for dc in DEPR_CODES:
            _info = cf_data.get(dc, {})
            for co, v in (_info.get('by_company', {}) or {}).items():
                depr_by_co[co] = depr_by_co.get(co, 0) + abs(v or 0)
            depr_total += abs(_info.get('total', 0) or 0)

        # EBITDA = 영업이익 + 감가상각 합계
        ebitda_by_co  = {c: (op_by.get(c, 0) or 0) + depr_by_co.get(c, 0) for c in companies}
        ebitda_total  = o_tot + depr_total
        ebitda_cmp    = None  # CF 전기 비교값 미확보 → delta 숨김

        # 3. 현금 (BS 1110101) 및 순차입금
        cash_info    = bs_data.get('1110101', {})
        cash_by_co   = {c: (cash_info.get('by_company', {}) or {}).get(c, 0) or 0 for c in companies}
        cash_total   = cash_info.get('total', 0) or 0
        cash_cmp     = cash_info.get('compare_total', 0) or 0

        net_debt_by_co = {c: (debt_by_co.get(c, 0) or 0) - cash_by_co.get(c, 0) for c in companies}
        net_debt_total = debt_total  - cash_total
        net_debt_cmp   = debt_cmp_total - cash_cmp

        # 4. 이자비용 (4500201+4500202) 및 이자보상배율
        INT_CODES = ['4500201', '4500202']
        int_by_co = {c: 0.0 for c in companies}
        int_total = 0.0
        int_cmp   = 0.0
        for ic in INT_CODES:
            _info = pl_data.get(ic, {})
            for co, v in (_info.get('by_company', {}) or {}).items():
                int_by_co[co] = int_by_co.get(co, 0) + abs(v or 0)
            int_total += abs(_info.get('total', 0) or 0)
            int_cmp   += abs(_info.get('compare_total', 0) or 0)

        icr_by_co = {c: ((op_by.get(c, 0) or 0) / int_by_co[c])
                     if int_by_co.get(c) else None for c in companies}
        icr_total = (o_tot / int_total) if int_total else None
        icr_cmp   = (o_cmp / int_cmp)   if int_cmp   else None

        # ─── 대시보드 3섹션 구조 ────────────────────────────────────────────
        dashboard = {
            # 수익성
            'profitability': [
                _mk('4100000',  '매출액',    'Net Sales',
                    sales_by, s_tot, s_cmp, dash_row=0),
                _mk('4700002',  '영업이익',  'Operating Income',
                    op_by, o_tot, o_cmp, dash_row=0),
                _mk('EBITDA',   'EBITDA',    'EBITDA',
                    ebitda_by_co, ebitda_total, ebitda_cmp,
                    derived=True, dash_row=0),
                _mk('4700004',  '당기순이익', 'Net Income',
                    ni_by, n_tot, n_cmp, dash_row=0),
                _mk('영업이익률', '영업이익률', 'Operating Margin (%)',
                    _ratio_by_co(op_by, sales_by),
                    (o_tot / s_tot * 100) if s_tot else None,
                    (o_cmp / s_cmp * 100) if s_cmp else None,
                    is_ratio=True, derived=True, dash_row=1, chart_group='margin_combo'),
                _mk('당기순이익률', '당기순이익률', 'Net Profit Margin (%)',
                    _ratio_by_co(ni_by, sales_by),
                    (n_tot / s_tot * 100) if s_tot else None,
                    (n_cmp / s_cmp * 100) if s_cmp else None,
                    is_ratio=True, derived=True, dash_row=1, chart_group='margin_combo'),
            ],
            # 재무안전성
            'safety': [
                _mk('1000000', '자산총계', 'Total Assets',
                    {c: bs_data.get('1000000',{}).get('by_company',{}).get(c, 0) for c in companies},
                    bs_data.get('1000000',{}).get('total', 0) or 0,
                    bs_data.get('1000000',{}).get('compare_total', 0) or 0, dash_row=0),
                _mk('2000000', '부채총계', 'Total Liabilities',
                    {c: bs_data.get('2000000',{}).get('by_company',{}).get(c, 0) for c in companies},
                    bs_data.get('2000000',{}).get('total', 0) or 0,
                    bs_data.get('2000000',{}).get('compare_total', 0) or 0, dash_row=0),
                _mk('3000000', '자본총계', 'Total Equity',
                    {c: bs_data.get('3000000',{}).get('by_company',{}).get(c, 0) for c in companies},
                    bs_data.get('3000000',{}).get('total', 0) or 0,
                    bs_data.get('3000000',{}).get('compare_total', 0) or 0, dash_row=0),
                _mk('부채비율', '부채비율', 'Debt Ratio (%)',
                    debt_ratio_by_co,
                    (liab_total / eq_total * 100) if eq_total else None,
                    (liab_cmp   / eq_cmp   * 100) if eq_cmp   else None,
                    is_ratio=True, derived=True, dash_row=0),
                # 차입금 그룹 (chart_group으로 합산 패널 병합)
                _mk('차입금총계', '차입금총계', 'Total Borrowings',
                    debt_by_co, debt_total, debt_cmp_total,
                    derived=True, dash_row=1, chart_group='debt_composition'),
                _mk('내부차입금', '내부차입금', 'Internal Borrowings',
                    internal_by_co, internal_total, internal_cmp_total,
                    derived=True, dash_row=1, chart_group='debt_composition'),
                _mk('외부차입금', '외부차입금', 'External Borrowings',
                    external_by_co, external_total, external_cmp_total,
                    derived=True, dash_row=1, chart_group='debt_composition'),
                _mk('순차입금', '순차입금', 'Net Debt',
                    net_debt_by_co, net_debt_total, net_debt_cmp,
                    derived=True, dash_row=1),
            ],
            # 현금흐름
            'cashflow': [
                _mk('OCF', '영업활동현금흐름', 'Operating Cash Flow',
                    ocf_by_co, ocf_total, ocf_cmp, dash_row=0),
                _mk('이자보상배율', '이자보상배율', 'Interest Coverage (x)',
                    icr_by_co, icr_total, icr_cmp,
                    is_multiple=True, derived=True, dash_row=0),
            ],
        }

        return jsonify({
            'ok': True,
            'companies': companies,
            'file': out_name,
            'download_url': url_for('download_result', filename=out_name),
            'compare_file': compare_out_name,
            'compare_download_url': (
                url_for('download_result', filename=compare_out_name)
                if compare_out_name else None
            ),
            'compare_prior_label': prior_same_q_label,
            'bs_preview': bs_preview,
            'pl_preview': pl_preview,
            'dashboard': dashboard,
            'wce_adj': agg.get('wce_adj_log', []),
        })
    except Exception as e:
        return _json_error(e)


@app.route('/download/<filename>')
@login_required
def download_result(filename):
    # 경로 탈출 방지: filename은 단순 파일명만 허용 (Windows에서 %5C→\ 디코딩 등 차단)
    if '/' in filename or '\\' in filename or '..' in filename:
        return '잘못된 파일명입니다.', 400
    path = RESULTS_DIR / filename
    if not path.exists():
        return '파일을 찾을 수 없습니다.', 404
    return send_file(str(path.resolve()), as_attachment=True, download_name=filename)


def _build_main_dashboard_excel(dashboard: dict, companies: list, period: str) -> bytes:
    """경영 대시보드 payload(profitability/safety/cashflow) → 시각화 엑셀.
    연결대시보드 엑셀(_build_dashboard_excel)과 동일한 톤·스타일."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    EOK = 1e8
    NUM_EOK  = '#,##0"억";(#,##0)"억";"-"'
    PCT_LIT  = '#,##0.0"%";(#,##0.0)"%";"-"'    # 값이 이미 ×100 된 퍼센트 숫자
    MULT_LIT = '#,##0.00"배";(#,##0.00)"배";"-"'

    C_BRAND='1F3864'; C_BRAND_LIGHT='D9E1F2'; C_BG_CARD='F4F6FA'
    C_ACCENT_BLUE='4472C4'; C_ACCENT_GREEN='70AD47'; C_ACCENT_ORANGE='ED7D31'
    C_ACCENT_RED='C0504D'; C_TEXT_MUTED='6B7280'

    HDR_FILL = PatternFill('solid', start_color=C_BRAND)
    HDR_FONT = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=11)
    SUB_FILL = PatternFill('solid', start_color=C_BRAND_LIGHT)
    SUB_FONT = Font(bold=True, color=C_BRAND, name='맑은 고딕', size=11)
    DATA_FONT = Font(name='맑은 고딕', size=10)
    TITLE_FONT = Font(bold=True, color=C_BRAND, size=18, name='맑은 고딕')
    SUBTITLE_FONT = Font(color=C_TEXT_MUTED, size=10, italic=True, name='맑은 고딕')
    THIN = Side(border_style='thin', color='BFBFBF')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _style_header_row(ws, row, cols, start_col=1):
        for c in range(start_col, start_col + cols):
            cell = ws.cell(row, c)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER

    def _style_data_cell(cell, fmt=None, align=None):
        cell.font = DATA_FONT
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
        if align:
            cell.alignment = align

    def _put_title(ws, title, subtitle, span_cols):
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 18
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = Alignment(vertical='center')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span_cols)
        ws['A2'] = subtitle
        ws['A2'].font = SUBTITLE_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span_cols)

    def _kpi_card(ws, row, col, span, label, value, value_fmt, compare, kind, accent):
        """KPI 카드 3행 블록 (라벨 / 값 / 전년대비)."""
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        lc = ws.cell(row, col, label)
        lc.font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=10)
        lc.fill = PatternFill('solid', start_color=accent)
        lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        lc.border = BORDER
        for cc in range(col + 1, col + span):
            ws.cell(row, cc).fill = PatternFill('solid', start_color=accent)
            ws.cell(row, cc).border = BORDER

        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + span - 1)
        vc = ws.cell(row + 1, col, value)
        vc.font = Font(bold=True, color=C_BRAND, name='맑은 고딕', size=18)
        vc.fill = PatternFill('solid', start_color=C_BG_CARD)
        vc.alignment = Alignment(horizontal='right', vertical='center', indent=2)
        vc.number_format = value_fmt
        vc.border = BORDER
        for cc in range(col + 1, col + span):
            ws.cell(row + 1, cc).fill = PatternFill('solid', start_color=C_BG_CARD)
            ws.cell(row + 1, cc).border = BORDER

        ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + span - 1)
        if compare is None or compare == 0 or value is None:
            txt = ' '; color = C_TEXT_MUTED
        else:
            diff = value - compare
            color = C_ACCENT_GREEN if diff >= 0 else C_ACCENT_RED
            arrow = '▲' if diff >= 0 else '▼'
            if kind == 'amount':
                pct = (diff / compare) if compare else 0
                txt = f'전년 {compare:,.0f}억  {arrow} {abs(pct)*100:.1f}%'
            elif kind == 'ratio':
                txt = f'전년 {compare:,.1f}%  {arrow} {abs(diff):.1f}%p'
            else:  # multiple
                txt = f'전년 {compare:,.2f}배  {arrow} {abs(diff):.2f}'
        cc = ws.cell(row + 2, col, txt)
        cc.font = Font(color=color, name='맑은 고딕', size=9)
        cc.fill = PatternFill('solid', start_color=C_BG_CARD)
        cc.alignment = Alignment(horizontal='right', vertical='center', indent=2)
        cc.border = BORDER
        for cc2 in range(col + 1, col + span):
            ws.cell(row + 2, cc2).fill = PatternFill('solid', start_color=C_BG_CARD)
            ws.cell(row + 2, cc2).border = BORDER
        ws.row_dimensions[row].height = 22
        ws.row_dimensions[row + 1].height = 34
        ws.row_dimensions[row + 2].height = 18

    def _kind(r):
        if r.get('is_ratio'):    return 'ratio'
        if r.get('is_multiple'): return 'multiple'
        return 'amount'

    def _find(section_key, code):
        for r in (dashboard.get(section_key) or []):
            if r.get('code') == code:
                return r
        return None

    def _disp(r):
        """표시값(전체, 전년) — 단위 환산."""
        k = _kind(r)
        tot = r.get('total'); cmp_ = r.get('compare_total')
        if k == 'amount':
            return (None if tot is None else tot / EOK,
                    None if cmp_ is None else cmp_ / EOK)
        return (tot, cmp_)   # ratio/multiple 은 그대로(이미 % / 배 숫자)

    wb = Workbook()

    # ── 시트 1: 개요 (KPI 카드) ──────────────────────────────────
    ws = wb.active
    ws.title = '개요'
    _put_title(ws, f'경영 대시보드 — {period}',
               '단위: 억원  ·  합산 기준  ·  전년 동기 대비 변동 표시', span_cols=8)

    # (행, 시작열, label, section, code, kind, accent)
    CARD_GRID = [
        (4, 1, '매출액',         'profitability', '4100000', 'amount',   C_ACCENT_BLUE),
        (4, 3, '영업이익',       'profitability', '4700002', 'amount',   C_ACCENT_GREEN),
        (4, 5, '당기순이익',     'profitability', '4700004', 'amount',   C_ACCENT_GREEN),
        (4, 7, 'EBITDA',         'profitability', 'EBITDA',  'amount',   C_ACCENT_BLUE),
        (8, 1, '자산총계',       'safety',        '1000000', 'amount',   C_ACCENT_BLUE),
        (8, 3, '부채총계',       'safety',        '2000000', 'amount',   C_ACCENT_ORANGE),
        (8, 5, '자본총계',       'safety',        '3000000', 'amount',   C_ACCENT_GREEN),
        (8, 7, '차입금총계',     'safety',        '차입금총계', 'amount', C_ACCENT_RED),
        (12, 1, '영업이익률',    'profitability', '영업이익률', 'ratio',  C_ACCENT_GREEN),
        (12, 3, '부채비율',      'safety',        '부채비율',   'ratio',  C_ACCENT_ORANGE),
        (12, 5, '영업활동현금흐름', 'cashflow',   'OCF',       'amount',  C_ACCENT_BLUE),
        (12, 7, '이자보상배율',  'cashflow',      '이자보상배율', 'multiple', C_ACCENT_RED),
    ]
    fmt_for = {'amount': NUM_EOK, 'ratio': PCT_LIT, 'multiple': MULT_LIT}
    for row, col, lbl, sec, code, kind, ac in CARD_GRID:
        r = _find(sec, code)
        if not r:
            continue
        val, cmp_ = _disp(r)
        _kpi_card(ws, row, col, span=2, label=lbl, value=val,
                  value_fmt=fmt_for[kind], compare=cmp_, kind=kind, accent=ac)
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ── 시트 2~4: 섹션별 표 + 차트 ───────────────────────────────
    SECTIONS = [
        ('profitability', '수익성',     C_ACCENT_GREEN),
        ('safety',        '재무안전성', C_ACCENT_BLUE),
        ('cashflow',      '현금흐름',   C_ACCENT_ORANGE),
    ]
    for sec_key, sec_title, _color in SECTIONS:
        rows = dashboard.get(sec_key) or []
        ws = wb.create_sheet(sec_title)
        _put_title(ws, sec_title, '전체 / 전년 / 증감 (금액 단위: 억원)', span_cols=6)

        amounts = [r for r in rows if _kind(r) == 'amount']
        ratios  = [r for r in rows if _kind(r) == 'ratio']
        mults   = [r for r in rows if _kind(r) == 'multiple']

        def _emit_block(r0, title, items, unit_fmt, delta_label, chart, chart_fmt):
            ws.cell(r0, 1, title).font = SUB_FONT
            ws.cell(r0, 1).fill = SUB_FILL
            ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=4)
            hdr = ['항목', '전체', '전년', delta_label]
            for i, h in enumerate(hdr, 1):
                ws.cell(r0 + 1, i, h)
            _style_header_row(ws, r0 + 1, 4)
            ds = r0 + 2
            for i, r in enumerate(items):
                rr = ds + i
                tot, cmp_ = _disp(r)
                ws.cell(rr, 1, r.get('kor') or r.get('code') or '')
                ws.cell(rr, 2, tot)
                ws.cell(rr, 3, cmp_)
                ws.cell(rr, 4, (tot - cmp_) if (tot is not None and cmp_ is not None) else None)
                _style_data_cell(ws.cell(rr, 1))
                _style_data_cell(ws.cell(rr, 2), unit_fmt)
                _style_data_cell(ws.cell(rr, 3), unit_fmt)
                _style_data_cell(ws.cell(rr, 4), unit_fmt)
            de = ds + len(items) - 1
            if chart and len(items) >= 1:
                ch = BarChart()
                ch.type = 'col'; ch.style = 11; ch.title = title
                ch.y_axis.number_format = chart_fmt
                data = Reference(ws, min_col=2, max_col=3, min_row=r0 + 1, max_row=de)
                cats = Reference(ws, min_col=1, min_row=ds, max_row=de)
                ch.add_data(data, titles_from_data=True)
                ch.set_categories(cats)
                ch.dataLabels = DataLabelList(showVal=True, showCatName=False,
                                              showSerName=False, showPercent=False,
                                              showLegendKey=False)
                ch.dataLabels.numFmt = chart_fmt
                ch.height = 9.0; ch.width = 14.0
                ws.add_chart(ch, f'F{r0}')
            return de

        cur = 4
        if amounts:
            cur = _emit_block(cur, '금액 지표 (억원)', amounts, NUM_EOK,
                              '증감(억)', True, '#,##0"억"') + 4
        if ratios:
            cur = _emit_block(cur, '비율 지표 (%)', ratios, PCT_LIT,
                              '증감(%p)', True, '#,##0.0"%"') + 4
        if mults:
            cur = _emit_block(cur, '배수 지표 (배)', mults, MULT_LIT,
                              '증감', False, '0.00"배"') + 4
        for c, w in enumerate([22, 14, 14, 14, 4, 14], 1):
            ws.column_dimensions[get_column_letter(c)].width = w

    # ── 시트 5: 회사별 상세 (금액 지표, 억원) ────────────────────
    ws = wb.create_sheet('회사별 상세')
    _put_title(ws, '회사별 상세', '개별 회사 단위 — 주요 금액 지표 (단위: 억원)', span_cols=8)

    CO_METRICS = [
        ('profitability', '4100000'),
        ('profitability', '4700002'),
        ('profitability', '4700004'),
        ('safety',        '1000000'),
        ('safety',        '2000000'),
        ('safety',        '3000000'),
        ('safety',        '차입금총계'),
        ('cashflow',      'OCF'),
    ]
    metric_cols = []
    for sec, code in CO_METRICS:
        r = _find(sec, code)
        if r and (r.get('companies') or {}):
            metric_cols.append(r)

    headers = ['회사'] + [f'{(r.get("kor") or r.get("code"))}(억)' for r in metric_cols]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    _style_header_row(ws, 4, len(headers))

    rr = 5
    for co in companies:
        ws.cell(rr, 1, co)
        _style_data_cell(ws.cell(rr, 1))
        for j, r in enumerate(metric_cols, start=2):
            v = (r.get('companies') or {}).get(co, 0) or 0
            ws.cell(rr, j, v / EOK)
            _style_data_cell(ws.cell(rr, j), NUM_EOK)
        rr += 1
    # 합계 행 (각 지표 total)
    ws.cell(rr, 1, '합계')
    ws.cell(rr, 1).font = Font(bold=True, color=C_BRAND, name='맑은 고딕', size=10)
    ws.cell(rr, 1).fill = SUB_FILL
    ws.cell(rr, 1).border = BORDER
    for j, r in enumerate(metric_cols, start=2):
        tot = r.get('total')
        ws.cell(rr, j, (tot / EOK) if tot is not None else None)
        c = ws.cell(rr, j)
        c.font = Font(bold=True, name='맑은 고딕', size=10, color=C_BRAND)
        c.fill = SUB_FILL
        c.border = BORDER
        c.number_format = NUM_EOK
        c.alignment = Alignment(horizontal='right', vertical='center')

    ws.column_dimensions['A'].width = 26
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = 'B5'

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@app.route('/dashboard/excel', methods=['POST'])
@login_required
@require_permission('aggregate.run')
def dashboard_excel():
    """경영 대시보드(메인 합산 결과)를 시각화 엑셀로 다운로드.
    합산 응답의 dashboard JSON 을 그대로 전송받아 재계산 없이 엑셀화."""
    body = request.get_json(silent=True) or {}
    dashboard = body.get('dashboard') or {}
    companies = body.get('companies') or []
    period = (body.get('period') or YEARS_DATA.get('default') or '').strip()
    if not isinstance(dashboard, dict) or not any(
            dashboard.get(k) for k in ('profitability', 'safety', 'cashflow')):
        return jsonify({'error': '대시보드 데이터가 없습니다. 먼저 합산을 실행해주세요.'}), 400

    data = _build_main_dashboard_excel(dashboard, companies, period)
    safe_period = re.sub(r'[\\/:*?"<>|]', '_', period) or 'period'
    fname = f'경영대시보드_{safe_period}.xlsx'
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── 연결정산 (연결조정분개 입력 + 최종 산출) ───────────────────────────────

def _norm_co_local(s):
    import re as _re
    return _re.sub(r'[\W_]+', '', str(s or '').casefold(), flags=_re.UNICODE)


def _files_for_group(year, company_names):
    """그룹 회사 리스트에 매칭되는 업로드 파일들 반환 (회사명 정규화 비교)."""
    targets = {_norm_co_local(c) for c in company_names}
    return [f for f in uploaded_files
            if f.get('year') == year
            and _norm_co_local(f.get('company', '')) in targets]


def _all_leaf_companies(group_id, period, _seen=None):
    """그룹과 모든 sub 그룹의 effective companies 합집합 (중복 제거, 순서 유지)."""
    _seen = set(_seen or [])
    if group_id in _seen:
        return []
    _seen.add(group_id)
    g = consol_get_group(group_id)
    if not g:
        return []
    out = list(consol_effective_companies(g, period))
    for inc_id in (g.get('included_groups') or []):
        out.extend(_all_leaf_companies(inc_id, period, _seen))
    seen, result = set(), []
    for c in out:
        if c not in seen:
            seen.add(c); result.append(c)
    return result


def _compute_global_sae_fund_adj(group_id, period):
    """글로벌세아 그룹 전용 — 직접 회사 + 모든 sub 그룹 leaf 회사들의
    CF1_연결/CF2_연결/CF3_연결 시트 데이터를 통합 aggregate 한 후
    cf_engine.compute_fund_adjustments_global_sae로 자금조정 dict 계산.

    반환: {cf_code: amount, ...} (해당 그룹의 파일이 하나도 없으면 빈 dict 반환)
    """
    all_companies = _all_leaf_companies(group_id, period)
    all_files = _files_for_group(period, all_companies)
    extracted_list = [f.get('extracted') for f in all_files if f.get('extracted')]
    if not extracted_list:
        return {}
    full_agg = aggregate(extracted_list)
    return cf_fund_adj_global_sae(full_agg)


@app.route('/consolidation')
@login_required
@require_permission('consol.compute')
def consolidation_index():
    """연결정산 메인 페이지."""
    year = request.args.get('year') or YEARS_DATA.get('default')
    if not _valid_year(year):
        year = YEARS_DATA.get('default')
    return render_template('consolidation.html',
                           year=year,
                           years=YEARS_DATA['years'],
                           locked_years=YEARS_DATA.get('locked', []),
                           username=session.get('username'),
                           is_admin=_is_admin(session.get('username')))


@app.route('/consolidation/template')
@login_required
def consolidation_template():
    """연결 COA 템플릿 JSON."""
    tpl = consol_load_template()
    return jsonify(tpl)


@app.route('/consolidation/groups', methods=['GET'])
@login_required
def consolidation_groups_list():
    """모든 그룹 + 각 그룹별 해당 연도 업로드 회사 수.
    + 현재 사용자의 분개 관리 권한(can_manage_journal) 플래그.
    담당회사 제한이 있는 사용자에게는 본인 담당회사가 포함된 그룹만 노출.
    """
    year = request.args.get('year') or YEARS_DATA.get('default')
    groups = consol_list_groups()
    uname = session.get('username')
    # 조회·실행은 모든 그룹 허용 → 목록은 필터하지 않음.
    # 분개 업로드/삭제 가능 여부만 그룹별 플래그로 표시(담당 그룹 기준).
    for g in groups:
        g['can_manage_journal'] = _can_manage_group_journal(uname, g.get('id'))
    if year and _valid_year(year):
        for g in groups:
            # 해당 기간에 그룹에 속한 회사만 (company_periods로 since/until 적용)
            active = consol_effective_companies(g, year)
            files = _files_for_group(year, active)
            g['active_companies'] = active
            g['file_count'] = len(files)
            g['matched_companies'] = sorted({f.get('company', '') for f in files})
    return jsonify({'groups': groups})


@app.route('/consolidation/groups', methods=['POST'])
@require_permission('consol.groups')
def consolidation_groups_create():
    data = request.get_json(silent=True) or {}
    try:
        g = consol_upsert_group(name=data.get('name', ''),
                                companies=data.get('companies', []),
                                group_id=data.get('id'),
                                included_groups=data.get('included_groups', []))
        return jsonify({'ok': True, 'group': g})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/consolidation/groups/<group_id>', methods=['DELETE'])
@require_permission('consol.groups')
def consolidation_groups_delete(group_id):
    consol_delete_group(group_id)
    return jsonify({'ok': True})


@app.route('/consolidation/prior/<group_id>', methods=['GET'])
@login_required
@require_permission('consol.compute')
def consolidation_prior_get(group_id):
    """그룹별 전년 연결값 수기 입력 데이터 조회."""
    if not consol_get_group(group_id):
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '해당 그룹에 접근 권한이 없습니다.'}), 403
    rec = consol_get_prior(group_id)
    return jsonify({
        'ok': True,
        'group_id': group_id,
        'bs_by_year': rec.get('bs_by_year') or {},
        'pl_by_period': rec.get('pl_by_period') or {},
        'bs_fields': PRIOR_BS_FIELDS,
        'pl_fields': PRIOR_PL_FIELDS,
        'updated_at': rec.get('updated_at') or '',
        'updated_by': rec.get('updated_by') or '',
    })


@app.route('/consolidation/prior/<group_id>', methods=['POST'])
@login_required
@require_permission('prior.edit')
def consolidation_prior_save(group_id):
    """그룹별 전년 연결값 수기 입력 저장. body: {bs_by_year, pl_by_period}."""
    if not consol_get_group(group_id):
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '해당 그룹에 접근 권한이 없습니다.'}), 403
    body = request.get_json(silent=True) or {}

    def _clean_numbers(d, allowed_fields):
        """{key: {field: value}} 구조에서 허용된 필드만, 숫자로 변환. 빈/null은 None으로."""
        out = {}
        for k, sub in (d or {}).items():
            if not isinstance(sub, dict):
                continue
            row = {}
            for f in allowed_fields:
                v = sub.get(f)
                if v is None or v == '':
                    row[f] = None
                else:
                    try:
                        row[f] = float(str(v).replace(',', '').strip())
                    except (ValueError, TypeError):
                        row[f] = None
            out[k] = row
        return out

    bs_by_year   = _clean_numbers(body.get('bs_by_year'),   PRIOR_BS_FIELDS)
    pl_by_period = _clean_numbers(body.get('pl_by_period'), PRIOR_PL_FIELDS)

    rec = consol_set_prior(group_id, bs_by_year, pl_by_period, session.get('username') or '')
    return jsonify({
        'ok': True,
        'updated_at': rec.get('updated_at') or '',
        'updated_by': rec.get('updated_by') or '',
    })


@app.route('/consolidation/journal/<group_id>/<period>', methods=['GET'])
@login_required
@require_permission('consol.compute')
def consolidation_journal_get(group_id, period):
    """현재 등록된 분개 조회 (두 종류)."""
    if not _valid_year(period):
        return jsonify({'error': '유효하지 않은 결산기간'}), 400
    if not consol_get_group(group_id):
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    rec = consol_get_journal(group_id, period)
    return jsonify(rec or {
        'group_id': group_id, 'period': period,
        'adjustment_entries': [], 'intercompany_entries': [],
    })


@app.route('/consolidation/journal/<group_id>/<period>', methods=['DELETE'])
@login_required
@require_permission('consol.journal')
def consolidation_journal_delete(group_id, period):
    """?type=adjustment 또는 ?type=intercompany 로 특정 종류만 삭제. 미지정 시 전체."""
    if not _can_manage_group_journal(session.get('username'), group_id):
        return jsonify({'error': '담당 그룹이 아닙니다. 배정된 연결그룹의 분개만 삭제할 수 있습니다. '
                                 '(담당 그룹 지정은 관리자에게 문의하세요)'}), 403
    if _is_locked(period) and not _is_admin(session.get('username')):
        return jsonify({
            'error': f'{period} 결산기간은 마감되어 분개를 삭제할 수 없습니다. (관리자 문의)'
        }), 403
    jtype = (request.args.get('type') or '').strip().lower() or None
    if jtype not in (None, 'adjustment', 'intercompany'):
        return jsonify({'error': '잘못된 분개 종류'}), 400
    consol_delete_journal(group_id, period, journal_type=jtype)
    # 전체 삭제 시 원본 분개 파일도 함께 제거 (부분 삭제는 보관 유지 — 다른 시트 데이터가 그대로 남아있을 수 있음)
    if jtype is None:
        for old in JOURNAL_DIR.glob(f'{group_id}_{period}.*'):
            try:
                old.unlink()
            except OSError:
                pass
    return jsonify({'ok': True})


def _find_journal_file(group_id, period):
    """업로드된 분개 원본 파일 경로 반환 (확장자 무관). 없으면 None."""
    for p in JOURNAL_DIR.glob(f'{group_id}_{period}.*'):
        if p.is_file():
            return p
    return None


@app.route('/consolidation/journal/<group_id>/<period>/download')
@login_required
@require_permission('consol.compute')
def consolidation_journal_download(group_id, period):
    """업로드된 원본 분개 파일을 그대로 다운로드.

    원본 파일이 없으면 (예전 업로드 / 파일 유실) 저장된 entries로부터 재생성된 양식을 대신 반환.
    """
    if not _valid_year(period):
        return '유효하지 않은 결산기간', 400
    g = consol_get_group(group_id)
    if not g:
        return '존재하지 않는 그룹', 404

    path = _find_journal_file(group_id, period)
    if path is not None:
        ext = path.suffix.lower()
        fname = f'분개_{g["name"]}_{period}{ext}'
        return send_file(path, as_attachment=True, download_name=fname)

    # 폴백: 원본 파일이 없으면 entries로부터 새로 생성
    rec = consol_get_journal(group_id, period) or {}
    if not (rec.get('adjustment_entries') or rec.get('intercompany_entries')):
        return '업로드된 분개가 없습니다.', 404
    data = consol_make_journal_template(
        g['name'], period,
        adjustment_entries=rec.get('adjustment_entries') or [],
        intercompany_entries=rec.get('intercompany_entries') or [],
    )
    fname = f'분개_{g["name"]}_{period}.xlsx'
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/consolidation/journal/<group_id>/<period>/template')
@login_required
@require_permission('consol.compute')
def consolidation_journal_template(group_id, period):
    """분개 입력용 빈 엑셀 양식 (2개 시트: 연결조정, 내부거래) 다운로드.

    업로드된 분개를 받으려면 별도 라우트 `/download` 사용.
    """
    if not _valid_year(period):
        return '유효하지 않은 결산기간', 400
    g = consol_get_group(group_id)
    if not g:
        return '존재하지 않는 그룹', 404
    data = consol_make_journal_template(
        g['name'], period,
        adjustment_entries=[],
        intercompany_entries=[],
    )
    fname = f'분개양식_{g["name"]}_{period}.xlsx'
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/consolidation/journal/<group_id>/<period>/upload', methods=['POST'])
@login_required
@require_permission('consol.journal')
def consolidation_journal_upload(group_id, period):
    """분개 엑셀 업로드 — 한 파일에 두 시트(연결조정/내부거래)를 모두 인식.

    선택적 ?mode=replace (기본) / ?mode=partial (시트에 데이터 있는 것만 갱신).
    """
    if not _valid_year(period):
        return jsonify({'error': '유효하지 않은 결산기간'}), 400
    if not consol_get_group(group_id):
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_manage_group_journal(session.get('username'), group_id):
        return jsonify({'error': '담당 그룹이 아닙니다. 배정된 연결그룹의 분개만 업로드할 수 있습니다. '
                                 '(담당 그룹 지정은 관리자에게 문의하세요)'}), 403
    if _is_locked(period) and not _is_admin(session.get('username')):
        return jsonify({
            'error': f'{period} 결산기간은 마감되어 분개를 업로드할 수 없습니다. (관리자 문의)'
        }), 403

    f = request.files.get('file')
    if not f:
        return jsonify({'error': '파일이 첨부되지 않았습니다.'}), 400
    upload_ext = Path(f.filename).suffix.lower()
    if upload_ext not in ALLOWED_EXT:
        return jsonify({'error': '엑셀 파일(.xlsx/.xlsm/.xls)만 허용됩니다.'}), 400

    mode = (request.args.get('mode') or 'replace').lower()

    raw_bytes = f.read()                          # 원본 파일 보관용으로 재사용
    try:
        parsed = consol_parse_journal_excel(raw_bytes)
    except Exception as e:
        return jsonify({'error': f'분개 파싱 실패: {e}'}), 400

    adj = parsed.get('adjustment_entries') or []
    inter = parsed.get('intercompany_entries') or []

    # 금액은 있는데 코드가 비어 있는 셀 검증
    chk_adj_codes = consol_validate_codes_present(adj)
    chk_int_codes = consol_validate_codes_present(inter)
    if not chk_adj_codes['ok'] or not chk_int_codes['ok']:
        SIDE_KR = {'debit': '차변', 'credit': '대변'}
        def _fmt_missing(items, sheet_label):
            return [
                {
                    'sheet': it.get('sheet') or sheet_label,
                    'row': it.get('row'),
                    'no': it.get('no'),
                    'side': SIDE_KR.get(it.get('side'), it.get('side')),
                    'amount': it.get('amount'),
                }
                for it in items
            ]
        return jsonify({
            'error': '금액이 입력되어 있는데 코드가 비어 있는 셀이 있어 업로드가 거부되었습니다. '
                     '해당 셀에 계정코드를 입력한 후 다시 업로드하세요.',
            'missing_codes': {
                'adjustment': _fmt_missing(chk_adj_codes['missing'], '연결조정'),
                'intercompany': _fmt_missing(chk_int_codes['missing'], '내부거래'),
            },
        }), 400

    # 저장 직전에는 검증용 보조 필드(_row, _sheet) 제거
    def _strip_aux(items):
        for it in items:
            it.pop('_row', None)
            it.pop('_sheet', None)
    _strip_aux(adj)
    _strip_aux(inter)

    # 종류별 차/대 균형 검증
    bal_adj = consol_validate_balance(adj)
    bal_int = consol_validate_balance(inter)
    if not bal_adj['ok'] or not bal_int['ok']:
        return jsonify({
            'error': '차변과 대변의 합이 일치하지 않아 업로드가 거부되었습니다.',
            'adjustment': {
                'ok': bal_adj['ok'],
                'total_debit': bal_adj['total_debit'],
                'total_credit': bal_adj['total_credit'],
                'diff': bal_adj['diff'],
                'entry_count': bal_adj['entry_count'],
            },
            'intercompany': {
                'ok': bal_int['ok'],
                'total_debit': bal_int['total_debit'],
                'total_credit': bal_int['total_credit'],
                'diff': bal_int['diff'],
                'entry_count': bal_int['entry_count'],
            },
        }), 400

    # 저장: partial 모드면 데이터 있는 시트만 교체, replace는 둘 다 교체
    if mode == 'partial':
        if adj:
            consol_set_journal_partial(group_id, period, 'adjustment', adj, session.get('username') or '')
        if inter:
            consol_set_journal_partial(group_id, period, 'intercompany', inter, session.get('username') or '')
        rec = consol_get_journal(group_id, period) or {}
    else:
        rec = consol_set_journal(group_id, period, adj, inter, session.get('username') or '')

    # 원본 분개 파일 보관 — 동일 그룹/기간의 기존 파일은 확장자와 무관하게 모두 정리 후 새로 저장
    try:
        for old in JOURNAL_DIR.glob(f'{group_id}_{period}.*'):
            try:
                old.unlink()
            except OSError:
                pass
        save_path = JOURNAL_DIR / f'{group_id}_{period}{upload_ext}'
        save_path.write_bytes(raw_bytes)
    except OSError:
        # 파일 시스템 오류로 원본 보관 실패해도 분개 자체는 저장 완료된 상태 → 경고만 로깅
        pass

    return jsonify({
        'ok': True,
        'mode': mode,
        'adjustment_count': len(adj),
        'intercompany_count': len(inter),
        'updated_at': rec.get('updated_at', ''),
    })


def _make_cash_result(ctx, period):
    """주어진 ctx(=`_compute_group_internal` 반환값)로 그룹의 현금정산표를 계산.
    cash_worksheet_compute 의 cf_compute 호출부와 동일한 인자 구성.
    상위 그룹의 rollup 컬럼에 sub의 "연결 최종" CF를 공급하기 위해 사용.
    실패 시 예외를 그대로 전파 — 호출부에서 try/except 처리.
    """
    agg = ctx['agg']
    companies = list((agg or {}).get('companies') or [])
    adj_entries    = ctx.get('adjustment_entries') or []
    inter_entries  = ctx.get('intercompany_entries') or []
    bridge_entries = ctx.get('bridge_entries') or []
    effective_adj = adj_entries + bridge_entries
    group_id = (ctx.get('group') or {}).get('id')

    # 전기(전년 4Q) 내부거래 — 차분 적용
    prior_period = cf_prior_year_4q(period)
    prior_inter_entries = []
    if prior_period and group_id:
        prior_rec = consol_get_journal(group_id, prior_period) or {}
        prior_inter_entries = prior_rec.get('intercompany_entries') or []

    manuals   = _cf_get_manuals(group_id, period)   if group_id else {}
    roundings = _cf_get_roundings(group_id, period) if group_id else {}

    # 연결정산표 최종 NI — sub에도 동일 plug 보정 적용 (sub 자체 NI 일치)
    target_final_ni = None
    try:
        for row in (ctx.get('result') or {}).get('rows') or []:
            if str(row.get('code') or '') == '4700004':
                target_final_ni = float(row.get('final') or 0)
                break
    except Exception:
        target_final_ni = None

    return cf_compute(agg, effective_adj, inter_entries, companies,
                      manual_adjustments=manuals,
                      rounding_adjustments=roundings,
                      prior_inter_entries=prior_inter_entries,
                      target_final_ni=target_final_ni,
                      rollups=(ctx.get('rollups') or None))


def _compute_group_internal(group_id, period, depth=0, _seen=None):
    """그룹 1개를 연결실행. 포함 그룹이 있으면 재귀적으로 먼저 실행해서 rollup 컬럼으로 추가.

    반환: {'result', 'bridge_info', 'bridge_entries', 'entries', 'files',
           'rollups': [{'group_id','name','rows_by_code'}, ...] }
    """
    if depth > 10:
        raise RuntimeError('연결 그룹 중첩 깊이가 너무 큽니다 (>10). 순환 참조 가능성.')
    _seen = set(_seen or [])
    if group_id in _seen:
        raise RuntimeError(f'그룹 순환 참조 감지: {group_id}')
    _seen.add(group_id)

    g = consol_get_group(group_id)
    if not g:
        raise ValueError(f'존재하지 않는 그룹: {group_id}')

    # 1) 직접 회사 매칭 — 해당 기간 멤버였던 회사만 (company_periods의 since/until)
    files = _files_for_group(period, consol_effective_companies(g, period))

    # 2) 포함 그룹 재귀 실행 → rollup 컬럼 생성
    rollups = []
    for inc_id in (g.get('included_groups') or []):
        inc_g = consol_get_group(inc_id)
        if not inc_g:
            continue
        sub = _compute_group_internal(inc_id, period, depth+1, _seen)
        # detail 행의 최종(P)값만 추출 — 연결정산표 rollup용 (BS/PL 코드)
        rows_by_code = {}
        for row in sub['result']['rows']:
            if row.get('kind') == 'detail' and row.get('code') and row.get('final'):
                rows_by_code[str(row['code'])] = float(row['final'])
        # 현금정산표 rollup용 — sub 그룹의 "연결 현금정산표 최종값"을 추출.
        # 과거에는 sub['agg'] CF 시트의 단순 total(직접 회사 raw 합)을 썼는데,
        # 그러면 sub 그룹의 연결조정/내부거래/수기조정/단수조정/target_final_ni
        # 보정이 전부 빠진 단순 합산값이 올라와 상역 합계가 틀어진다.
        # 대신 sub의 cf_compute 결과(행별 final = 연결 최종)를 사용.
        # 2단계 중첩(글로벌세아 → 상역 → 태림)도 sub의 cf_compute가 이미
        # sub의 rollups를 by_company로 반영하므로 별도 누적 불필요.
        cf_by_code = {}
        try:
            sub_cash = _make_cash_result(sub, period)
        except Exception:
            sub_cash = None

        if sub_cash:
            # 일반 섹션 행 (cf_code → final)
            for sec in (sub_cash.get('sections') or []):
                for crow in (sec.get('rows') or []):
                    ccode = str(crow.get('cf_code') or '').strip()
                    if not ccode:
                        continue
                    cf_by_code[ccode] = float(crow.get('final') or 0)
            # 라벨 행(Ⅴ.환율변동 / Ⅵ.기초현금 / Ⅶ.기말현금)은 cf_sheet 키 형태(LBL::)로
            # 보존해야 상위 그룹에서 매칭된다. sub_cash의 fx_effect/cash_begin/cash_end
            # final을 우선 사용하고, 매칭 안 되면 sub_cf_sheet의 total fallback.
            # 단, '당기순이익' 라벨은 NI 처리에서 final로 채울 것이므로 이 루프에서는 skip.
            sub_cf_sheet = ((sub.get('agg') or {}).get('sheets') or {}).get('CF') or {}
            label_finals = {
                'fx':    float((sub_cash.get('fx_effect') or {}).get('final') or 0),
                'begin': float((sub_cash.get('cash_begin') or {}).get('final') or 0),
                'end':   float((sub_cash.get('cash_end') or {}).get('final') or 0),
            }
            for code, info in sub_cf_sheet.items():
                if not (code and str(code).startswith('LBL::')):
                    continue
                label_text = str(code)[5:]
                if '당기순이익' in label_text:
                    # NI 라벨은 아래 NI 처리에서 final(연결 최종)로 주입 — 여기서 raw로 덮으면 안 됨
                    continue
                if ('환율변동' in label_text) or ('Ⅴ' in label_text):
                    cf_by_code[str(code)] = label_finals['fx']
                elif ('기초의현금' in label_text) or ('기초현금' in label_text) or ('Ⅵ' in label_text):
                    cf_by_code[str(code)] = label_finals['begin']
                elif ('기말의현금' in label_text) or ('기말현금' in label_text) or ('Ⅶ' in label_text):
                    cf_by_code[str(code)] = label_finals['end']
                else:
                    tot = (info or {}).get('total')
                    if tot is None:
                        bc = (info or {}).get('by_company') or {}
                        tot = sum(float(v or 0) for v in bc.values()) if bc else 0
                    try:
                        cf_by_code[str(code)] = float(tot or 0)
                    except (TypeError, ValueError):
                        pass
            # 당기순이익(NI) 행 — sections / LBL 라벨 처리 뒤에 마지막으로 주입.
            # _net_income_v2가 매칭한 cf_code(4900001 / 4700004 / 3500105 / LBL::당기순이익)
            # 그대로 사용해야 상위 그룹의 _net_income_v2가 동일 코드로 픽업한다.
            # 위치 순서: 다른 처리들이 같은 키를 raw 값으로 덮어쓰지 못하도록 맨 마지막.
            ni = sub_cash.get('net_income') or {}
            ni_code = str(ni.get('cf_code') or '').strip()
            if ni_code:
                cf_by_code[ni_code] = float(ni.get('final') or 0)
        else:
            # cf_compute 실패 시 안전 fallback — 기존 raw total 로직
            sub_cf_sheet = ((sub.get('agg') or {}).get('sheets') or {}).get('CF') or {}
            for code, info in sub_cf_sheet.items():
                if not code:
                    continue
                tot = (info or {}).get('total')
                if tot is None:
                    bc = (info or {}).get('by_company') or {}
                    tot = sum(float(v or 0) for v in bc.values()) if bc else 0
                try:
                    cf_by_code[str(code)] = float(tot or 0)
                except (TypeError, ValueError):
                    pass
            for sub_r in (sub.get('rollups') or []):
                for c, v in (sub_r.get('cf_by_code') or {}).items():
                    cf_by_code[c] = cf_by_code.get(c, 0.0) + float(v or 0)

        rollup_name = f'{inc_g["name"]}(연결)'
        rollups.append({
            'group_id': inc_id,
            'name': rollup_name,
            'rows_by_code': rows_by_code,
            'cf_by_code': cf_by_code,
            'sub_result': sub,        # 디버깅/감사용
        })

    if not files and not rollups:
        raise ValueError(f'{period} 기간에 업로드된 그룹 회사 파일도, 포함 그룹의 결과도 없습니다.')

    # 3) 직접 회사 합산
    def _sort_key(f):
        cur = ((f.get('extracted') or {}).get('currency') or 'ZZZ').upper()
        return (0 if cur == 'KRW' else 1, cur, f.get('company') or '')
    files.sort(key=_sort_key)

    extracted_list = [f['extracted'] for f in files]
    if extracted_list:
        agg = aggregate(extracted_list)
        try:
            agg = _apply_wce_to_aggregation(agg, period)
        except Exception:
            pass
        companies = agg['companies']
    else:
        # 직접 회사가 없고 포함 그룹만 있는 경우 (드문 케이스)
        agg = {'companies': [], 'sheets': {'BS': {}, 'PL_MF': {}}}
        companies = []

    # 4) 분개 (연결조정 + 내부거래) + 자동 매듭
    rec = consol_get_journal(group_id, period) or {}
    adj_entries   = rec.get('adjustment_entries')   or []
    inter_entries = rec.get('intercompany_entries') or []

    # 매듭 분개는 두 종류 통합 기준으로 만들고, '연결조정' 묶음에 추가
    bridge_entries, bridge_info = consol_auto_bridge(adj_entries, inter_entries)
    effective_adj = adj_entries + bridge_entries

    # 5) 계산 — rollup이 있으면 컬럼 추가
    if rollups:
        result = consol_compute_with_rollup(agg, effective_adj, inter_entries, companies, rollups)
    else:
        result = consol_compute(agg, effective_adj, inter_entries, companies)

    return {
        'group': g,
        'result': result,
        'bridge_info': bridge_info,
        'bridge_entries': bridge_entries,
        'adjustment_entries': adj_entries,
        'intercompany_entries': inter_entries,
        'files': files,
        'rollups': rollups,
        'agg': agg,
    }


@app.route('/consolidation/dashboard')
@login_required
@require_permission('consol.compute')
def consolidation_dashboard_index():
    """그룹별 대시보드 메인. 그룹/기간 선택 후 차트·지표 표시."""
    year = request.args.get('year') or YEARS_DATA.get('default')
    if not _valid_year(year):
        year = YEARS_DATA.get('default')
    group_id = request.args.get('group') or ''
    return render_template('consolidation_dashboard.html',
                           year=year, years=YEARS_DATA['years'],
                           group_id=group_id,
                           username=session.get('username'),
                           is_admin=_is_admin(session.get('username')))


def _build_dashboard_payload(ctx, period):
    """단일 ctx로부터 대시보드 데이터 dict를 빌드. 포함 그룹(rollup) 재귀 렌더링에 사용됨."""
    g = ctx['group']
    result = ctx['result']
    rows_by_code = {str(r['code']): r for r in result['rows'] if r.get('code')}

    def _row(code):
        r = rows_by_code.get(code) or {}
        return {
            'code': code,
            'name': r.get('name', ''),
            'sum': r.get('sum', 0) or 0,
            'dr_adj': r.get('dr_adj', 0) or 0,
            'cr_adj': r.get('cr_adj', 0) or 0,
            'dr_int': r.get('dr_int', 0) or 0,
            'cr_int': r.get('cr_int', 0) or 0,
            'final': r.get('final', 0) or 0,
        }

    # 핵심 지표
    bs_assets   = _row('1000000')
    bs_liab     = _row('2000000')
    bs_equity   = _row('3000000')
    pl_sales    = _row('4100000')
    pl_cogs     = _row('4200000')
    pl_gross    = _row('4700001')
    pl_op       = _row('4700002')
    pl_pretax   = _row('4700003')
    pl_ni       = _row('4700004')
    pl_ni_owner = _row('4900001')
    pl_ni_nci   = _row('4900002')

    # ─── 차입금 (총차입금 / 외부차입금 / 내부차입금) ─────────────────
    # 메인페이지 합산 대시보드와 동일한 BS 차입금 코드들
    DEBT_CODES = ['2100201', '2100202', '2100203', '2100204', '2100205',
                  '2100301', '2100391', '2200101', '2200201', '2200291']

    # 총차입금: rows_by_code의 sum / final 합산
    debt_total_sum   = sum(_row(c)['sum']   for c in DEBT_CODES)
    debt_total_final = sum(_row(c)['final'] for c in DEBT_CODES)

    # 외부차입금: 직접 회사별로 (BS 차입금 - CF2/CF3 기말) 합산 + 포함 그룹은 이미 연결완료된 BS 차입금
    agg_for_debt = ctx.get('agg') or {}
    bs_for_debt  = (agg_for_debt.get('sheets') or {}).get('BS')        or {}
    cf2_for_debt = (agg_for_debt.get('sheets') or {}).get('CF2_연결') or {}
    cf3_for_debt = (agg_for_debt.get('sheets') or {}).get('CF3_연결') or {}
    direct_companies = agg_for_debt.get('companies') or []

    external_total = 0.0
    for co in direct_companies:
        co_debt = 0.0
        for dc in DEBT_CODES:
            co_debt += (bs_for_debt.get(dc, {}).get('by_company', {}) or {}).get(co, 0) or 0
        co_internal = 0.0
        for conn_sheet in (cf2_for_debt, cf3_for_debt):
            for key, info in conn_sheet.items():
                label = key.split('::', 1)[1] if '::' in key else ''
                if '기말' in label:
                    co_internal += (info.get('by_company', {}) or {}).get(co, 0) or 0
        # 캡 처리: 내부차입금이 총차입금을 초과할 수 없음 (소스 파일 오차 보정)
        if co_internal > co_debt:
            co_internal = co_debt
        external_total += (co_debt - co_internal)

    # 포함 그룹(rollup): 이미 연결완료된 BS 차입금은 전부 외부차입금으로 간주
    for r in (ctx.get('rollups') or []):
        rollup_name = r['name']
        for dc in DEBT_CODES:
            external_total += (rows_by_code.get(dc, {}).get('companies', {}) or {}).get(rollup_name, 0) or 0

    # 은행차입금은 연결조정으로 변동되지 않음 (외부 채권자 대상이므로) → sum == final
    external_sum   = external_total
    external_final = external_total

    bs_debt_total = {'code': '차입금총계', 'name': '차입금총계',
                     'sum': debt_total_sum, 'final': debt_total_final,
                     'dr_adj': 0, 'cr_adj': 0, 'dr_int': 0, 'cr_int': 0}
    bs_debt_bank = {'code': '은행차입금', 'name': '은행차입금',
                    'sum': external_sum, 'final': external_final,
                    'dr_adj': 0, 'cr_adj': 0, 'dr_int': 0, 'cr_int': 0}

    # 회사별 통화 매핑 (그룹화에 사용)
    company_currency = {}
    for fr in (ctx['files'] or []):
        co = (fr.get('company') or '').strip()
        cur = ((fr.get('extracted') or {}).get('currency') or '').upper()
        if co:
            company_currency[co] = cur or 'KRW'

    rollup_names = {r['name'] for r in (ctx['rollups'] or [])}
    group_name = g.get('name', '')

    # 회사별 자산총계 / 매출액 (rollup 컬럼 포함)
    def _by_company(code):
        r = rows_by_code.get(code) or {}
        by_co = r.get('companies') or {}
        return [{'company': c, 'value': by_co.get(c, 0) or 0,
                 'currency': company_currency.get(c, '' if c in rollup_names else 'KRW'),
                 'is_rollup': c in rollup_names}
                for c in result['companies']]

    # 각 rollup이 해외 전용 그룹인지 판정 (소속 회사 통화가 전부 KRW가 아닐 때 → 해외)
    def _is_overseas_rollup(rollup_ctx):
        sub_files = (rollup_ctx.get('sub_result') or {}).get('files') or []
        if not sub_files:
            return False
        krw = 0
        non_krw = 0
        for f in sub_files:
            cur = ((f.get('extracted') or {}).get('currency') or '').upper()
            if cur == 'KRW' or not cur:
                krw += 1
            else:
                non_krw += 1
        return krw == 0 and non_krw > 0

    overseas_rollup_names = set()
    for r in (ctx['rollups'] or []):
        if _is_overseas_rollup(r):
            overseas_rollup_names.add(r['name'])

    # ── 대시보드 가상 합산 그룹 ──
    # 회사명에 패턴이 포함되면 해당 가상 그룹 버킷으로 묶음 (대시보드 표시 전용)
    VIRTUAL_BUCKETS = [
        {
            'name': 'Tegra(합산)',
            'patterns': ['Tegra', 'Southern Apparel', 'Decotex'],   # case-insensitive contains
            'type': 'virtual',
        },
    ]

    def _match_virtual_bucket(company_name):
        if not company_name:
            return None
        co_lower = company_name.lower()
        for vb in VIRTUAL_BUCKETS:
            for p in vb['patterns']:
                if p.lower() in co_lower:
                    return vb
        return None

    def _bucketize(per_company):
        """버킷: 본사(개별) / 실제 포함그룹 / 가상 합산(Tegra) / 기타 국내법인 / 기타 해외법인."""
        main_val = 0
        main_label = None
        rollup_buckets = []
        virtual_acc = {}   # 가상 그룹명 → {'value','count'}
        domestic_val = 0
        domestic_count = 0
        overseas_val = 0
        overseas_count = 0

        # 1) 실제 rollup 처리
        for r in (ctx['rollups'] or []):
            v = next((x['value'] for x in per_company if x['company'] == r['name']), 0)
            if r['name'] in overseas_rollup_names:
                overseas_val += v
                overseas_count += len((r.get('sub_result') or {}).get('files') or [])
            else:
                rollup_buckets.append({'label': r['name'], 'value': v, 'type': 'rollup'})

        # 2) 직접 회사 처리
        for x in per_company:
            co = x['company']
            v = x['value']
            if co in rollup_names:
                continue
            # 본사 탐지
            # 1) 완전 일치 (회사명 == 그룹명)
            # 2) "(개별)" 떼면 그룹명과 같음
            # 3) "(개별)" 버전이면서, "(개별)" 떼고 비교했을 때 그룹명이 회사명의 prefix 또는 suffix
            #    예: 상역 그룹 + "세아상역(개별)" → "세아상역".endswith("상역") = True → 본사 ✓
            #    예: 태림 그룹 + "㈜전주원파워(개별)" → "전주원파워"는 "태림"으로 시작·끝 안 함 → 본사 X
            #    "(개별)" 없는 단순 부분일치는 본사로 인정하지 않음 (예: 태림 그룹의 "태림판지"는 본사 X)
            co_stripped = co.replace('(개별)', '').strip().lstrip('㈜').strip()
            is_main = bool(group_name) and (
                co == group_name
                or co_stripped == group_name
                or (
                    '(개별)' in co
                    and (co_stripped.endswith(group_name) or co_stripped.startswith(group_name))
                )
            )
            if is_main and main_label is None:
                main_label = co
                main_val = v
                continue
            # 가상 합산 버킷
            vb = _match_virtual_bucket(co)
            if vb is not None:
                slot = virtual_acc.setdefault(vb['name'], {'value': 0, 'count': 0})
                slot['value'] += v
                slot['count'] += 1
                continue
            # 통화별
            cur = company_currency.get(co, 'KRW')
            if cur and cur != 'KRW':
                overseas_val += v
                overseas_count += 1
            else:
                domestic_val += v
                domestic_count += 1

        out = []
        if main_label is not None:
            out.append({'label': main_label, 'value': main_val, 'type': 'main'})
        out.extend(rollup_buckets)
        for vname, slot in virtual_acc.items():
            out.append({'label': vname, 'value': slot['value'],
                        'type': 'virtual', 'count': slot['count']})
        if domestic_count > 0:
            out.append({'label': f'기타 국내법인 ({domestic_count}개사)',
                        'value': domestic_val, 'type': 'domestic', 'count': domestic_count})
        if overseas_count > 0:
            out.append({'label': '기타 해외법인',
                        'value': overseas_val, 'type': 'overseas', 'count': overseas_count})
        return out

    # 검증행
    check_row = next((r for r in result['rows'] if r.get('kind') == 'check'), None)
    check_val = (check_row or {}).get('final', 0)

    # 비율 계산
    def _safe_div(a, b):
        return (a / b) if b else 0

    metrics = {
        'debt_ratio_sum': _safe_div(bs_liab['sum'], bs_equity['sum']),
        'debt_ratio_final': _safe_div(bs_liab['final'], bs_equity['final']),
        'op_margin_sum': _safe_div(pl_op['sum'], pl_sales['sum']),
        'op_margin_final': _safe_div(pl_op['final'], pl_sales['final']),
        'net_margin_sum': _safe_div(pl_ni['sum'], pl_sales['sum']),
        'net_margin_final': _safe_div(pl_ni['final'], pl_sales['final']),
    }

    # ─── 전년 연결값 (수기 입력) 주입: 각 KPI에 compare 필드 추가 ─────────
    prior_rec = consol_get_prior(g.get('id') or '')
    py = prior_year_of(period)
    pp = prior_period_of(period)
    prior_bs = (prior_rec.get('bs_by_year') or {}).get(py or '', {}) or {}
    prior_pl = (prior_rec.get('pl_by_period') or {}).get(pp or '', {}) or {}

    def _set_compare(kpi_dict, label, src):
        v = src.get(label)
        kpi_dict['compare'] = v if v is not None else None

    _set_compare(bs_assets,     '자산총계', prior_bs)
    _set_compare(bs_liab,       '부채총계', prior_bs)
    _set_compare(bs_equity,     '자본총계', prior_bs)
    _set_compare(bs_debt_total, '차입금총계', prior_bs)
    _set_compare(bs_debt_bank,  '은행차입금', prior_bs)
    _set_compare(pl_sales, '매출액',     prior_pl)
    _set_compare(pl_op,    '영업이익',   prior_pl)
    _set_compare(pl_ni,    '당기순이익', prior_pl)
    # 표시되지 않거나 전년 입력 항목 없음 → compare = None
    for x in (pl_cogs, pl_gross, pl_pretax, pl_ni_owner, pl_ni_nci):
        x['compare'] = None

    # 전년 기반 비율 (입력값이 모두 있을 때만)
    prior_bs_liab   = prior_bs.get('부채총계')
    prior_bs_equity = prior_bs.get('자본총계')
    prior_pl_sales  = prior_pl.get('매출액')
    prior_pl_op     = prior_pl.get('영업이익')
    prior_pl_ni     = prior_pl.get('당기순이익')
    metrics['debt_ratio_compare'] = (_safe_div(prior_bs_liab, prior_bs_equity)
                                     if prior_bs_liab is not None and prior_bs_equity else None)
    metrics['op_margin_compare']  = (_safe_div(prior_pl_op, prior_pl_sales)
                                     if prior_pl_op is not None and prior_pl_sales else None)
    metrics['net_margin_compare'] = (_safe_div(prior_pl_ni, prior_pl_sales)
                                     if prior_pl_ni is not None and prior_pl_sales else None)

    # 주요 자산 구성 (유동자산, 비유동자산)
    asset_breakdown = [
        {'name': '유동자산', 'sum': _row('1100000')['sum'], 'final': _row('1100000')['final']},
        {'name': '비유동자산', 'sum': _row('1200000')['sum'], 'final': _row('1200000')['final']},
    ]
    liab_breakdown = [
        {'name': '유동부채', 'sum': _row('2100000')['sum'], 'final': _row('2100000')['final']},
        {'name': '비유동부채', 'sum': _row('2200000')['sum'], 'final': _row('2200000')['final']},
    ]
    equity_breakdown = [
        {'name': '자본금', 'sum': _row('3100000')['sum'], 'final': _row('3100000')['final']},
        {'name': '자본잉여금', 'sum': _row('3200000')['sum'], 'final': _row('3200000')['final']},
        {'name': '자본조정', 'sum': _row('3300000')['sum'], 'final': _row('3300000')['final']},
        {'name': '기타포괄손익누계', 'sum': _row('3400000')['sum'], 'final': _row('3400000')['final']},
        {'name': '이익잉여금', 'sum': _row('3500000')['sum'], 'final': _row('3500000')['final']},
        {'name': '비지배지분', 'sum': _row('3600101')['sum'], 'final': _row('3600101')['final']},
    ]

    return {
        'ok': True,
        'group': g,
        'period': period,
        'prior_year': py,
        'prior_period': pp,
        'companies': result['companies'],
        'rollups': [
            {'group_id': r['group_id'], 'name': r['name'],
             'group_name': r['sub_result']['group']['name']}
            for r in ctx['rollups']
        ],
        'matched_files_count': len(ctx['files']),
        'adjustment_count': len(ctx['adjustment_entries']),
        'intercompany_count': len(ctx['intercompany_entries']),
        'bridge': {
            'applied': ctx['bridge_info']['applied'],
            'amount': ctx['bridge_info']['bridge_amount'],
        },
        'check_imbalance': check_val,
        'kpi': {
            'bs_assets': bs_assets,
            'bs_liab':   bs_liab,
            'bs_equity': bs_equity,
            'bs_debt_total':    bs_debt_total,
            'bs_debt_bank':     bs_debt_bank,
            'pl_sales':  pl_sales,
            'pl_cogs':   pl_cogs,
            'pl_gross':  pl_gross,
            'pl_op':     pl_op,
            'pl_pretax': pl_pretax,
            'pl_ni':     pl_ni,
            'pl_ni_owner': pl_ni_owner,
            'pl_ni_nci':   pl_ni_nci,
        },
        'metrics': metrics,
        'breakdown': {
            'asset': asset_breakdown,
            'liab':  liab_breakdown,
            'equity': equity_breakdown,
        },
        'by_company': {
            'assets':  _by_company('1000000'),
            'sales':   _by_company('4100000'),
            'op':      _by_company('4700002'),
            'ni':      _by_company('4700004'),
        },
        'by_company_grouped': {
            'assets':  _bucketize(_by_company('1000000')),
            'sales':   _bucketize(_by_company('4100000')),
            'op':      _bucketize(_by_company('4700002')),
            'ni':      _bucketize(_by_company('4700004')),
        },
    }


@app.route('/consolidation/dashboard/<group_id>/data')
@login_required
@require_permission('consol.compute')
def consolidation_dashboard_data(group_id):
    """대시보드용 핵심 지표/차트 데이터. 포함 그룹(rollup)은 동일 구성의 sub_dashboards로 함께 반환."""
    period = (request.args.get('period') or YEARS_DATA.get('default') or '').strip()
    if not _valid_year(period):
        return jsonify({'error': '유효하지 않은 기간'}), 400
    g = consol_get_group(group_id)
    if not g:
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '해당 그룹에 접근 권한이 없습니다.'}), 403

    try:
        ctx = _compute_group_internal(group_id, period)
    except (ValueError, RuntimeError) as e:
        return jsonify({'error': str(e)}), 400

    payload = _build_dashboard_payload(ctx, period)

    # 포함 그룹(rollup) 각각의 페이로드를 빌드 → 그 그룹의 by_company_grouped 슬라이스로
    # 메인 도넛의 단일 rollup 슬라이스를 펼쳐 치환한다.
    sub_payload_by_label = {}    # 예: '상역(연결)' → 상역 그룹 자체의 payload
    for r in (ctx.get('rollups') or []):
        sub_ctx = r.get('sub_result')
        if not sub_ctx:
            continue
        try:
            sub_payload_by_label[r['name']] = _build_dashboard_payload(sub_ctx, period)
        except Exception:
            pass

    def _expand_rollups(buckets, metric_key):
        """rollup 슬라이스를 그 그룹의 by_company_grouped[metric_key] 슬라이스로 치환."""
        out = []
        for bk in buckets:
            if bk.get('type') == 'rollup' and bk.get('label') in sub_payload_by_label:
                sub_p = sub_payload_by_label[bk['label']]
                sub_buckets = (sub_p.get('by_company_grouped') or {}).get(metric_key) or []
                if sub_buckets:
                    out.extend(sub_buckets)
                    continue
            out.append(bk)
        return out

    grouped = payload.get('by_company_grouped') or {}
    payload['by_company_grouped'] = {
        'assets': _expand_rollups(grouped.get('assets') or [], 'assets'),
        'sales':  _expand_rollups(grouped.get('sales')  or [], 'sales'),
        'op':     _expand_rollups(grouped.get('op')     or [], 'op'),
        'ni':     _expand_rollups(grouped.get('ni')     or [], 'ni'),
    }
    return jsonify(payload)


# ─── 대시보드 엑셀 다운로드 ──────────────────────────────────────────────────
def _build_dashboard_excel(payload: dict) -> bytes:
    """대시보드 payload → 시각화된 엑셀 워크북 (KPI 카드 + 도넛/막대 차트)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import DoughnutChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    EOK = 1e8  # 억원 단위 변환 (대시보드 화면과 동일)
    NUM = '#,##0;(#,##0);"-"'              # 원 단위 (정밀)
    NUM_EOK = '#,##0"억";(#,##0)"억";"-"'   # 억원 표시
    PCT = '0.00%;(0.00%);"-"'

    # 색 팔레트 — 대시보드와 톤 맞춤
    C_BRAND = '1F3864'
    C_BRAND_LIGHT = 'D9E1F2'
    C_BG_CARD = 'F4F6FA'
    C_ACCENT_BLUE = '4472C4'
    C_ACCENT_GREEN = '70AD47'
    C_ACCENT_ORANGE = 'ED7D31'
    C_ACCENT_RED = 'C0504D'
    C_TEXT_MUTED = '6B7280'

    HDR_FILL = PatternFill('solid', start_color=C_BRAND)
    HDR_FONT = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=11)
    SUB_FILL = PatternFill('solid', start_color=C_BRAND_LIGHT)
    SUB_FONT = Font(bold=True, color=C_BRAND, name='맑은 고딕', size=11)
    DATA_FONT = Font(name='맑은 고딕', size=10)
    TITLE_FONT = Font(bold=True, color=C_BRAND, size=18, name='맑은 고딕')
    SUBTITLE_FONT = Font(color=C_TEXT_MUTED, size=10, italic=True, name='맑은 고딕')
    THIN = Side(border_style='thin', color='BFBFBF')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()

    def _style_header_row(ws, row, cols, start_col=1):
        for c in range(start_col, start_col + cols):
            cell = ws.cell(row, c)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER

    def _style_data_cell(cell, fmt=None, align=None):
        cell.font = DATA_FONT
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
        if align:
            cell.alignment = align

    def _put_title(ws, title, subtitle, span_cols):
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 18
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = Alignment(vertical='center')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span_cols)
        ws['A2'] = subtitle
        ws['A2'].font = SUBTITLE_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span_cols)

    def _kpi_card(ws, row, col, span, label, value_eok, compare_eok=None,
                  accent=C_ACCENT_BLUE):
        """KPI 카드 — 3행 × span 열 머지 블록.
          row+0  라벨   (배경 accent, 흰글씨, 좌측 정렬)
          row+1  값(억) (흰배경, 큰 글씨, 우측 정렬)
          row+2  전년대비 변화 (작은 글씨)
        """
        # 라벨 행
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        lc = ws.cell(row, col, label)
        lc.font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=10)
        lc.fill = PatternFill('solid', start_color=accent)
        lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        lc.border = BORDER
        for cc in range(col + 1, col + span):
            ws.cell(row, cc).fill = PatternFill('solid', start_color=accent)
            ws.cell(row, cc).border = BORDER

        # 값 행
        ws.merge_cells(start_row=row + 1, start_column=col,
                       end_row=row + 1, end_column=col + span - 1)
        vc = ws.cell(row + 1, col, value_eok)
        vc.font = Font(bold=True, color=C_BRAND, name='맑은 고딕', size=18)
        vc.fill = PatternFill('solid', start_color=C_BG_CARD)
        vc.alignment = Alignment(horizontal='right', vertical='center', indent=2)
        vc.number_format = NUM_EOK
        vc.border = BORDER
        for cc in range(col + 1, col + span):
            ws.cell(row + 1, cc).fill = PatternFill('solid', start_color=C_BG_CARD)
            ws.cell(row + 1, cc).border = BORDER

        # 변화 행
        ws.merge_cells(start_row=row + 2, start_column=col,
                       end_row=row + 2, end_column=col + span - 1)
        if compare_eok is None or compare_eok == 0:
            txt = ' '
            color = C_TEXT_MUTED
        else:
            diff = value_eok - compare_eok
            pct = (diff / compare_eok) if compare_eok else 0
            arrow = '▲' if diff >= 0 else '▼'
            color = C_ACCENT_GREEN if diff >= 0 else C_ACCENT_RED
            txt = f'전년 {compare_eok:,.0f}억  {arrow} {abs(pct)*100:.1f}%'
        cc = ws.cell(row + 2, col, txt)
        cc.font = Font(color=color, name='맑은 고딕', size=9)
        cc.fill = PatternFill('solid', start_color=C_BG_CARD)
        cc.alignment = Alignment(horizontal='right', vertical='center', indent=2)
        cc.border = BORDER
        for cc2 in range(col + 1, col + span):
            ws.cell(row + 2, cc2).fill = PatternFill('solid', start_color=C_BG_CARD)
            ws.cell(row + 2, cc2).border = BORDER

        ws.row_dimensions[row].height = 22
        ws.row_dimensions[row + 1].height = 34
        ws.row_dimensions[row + 2].height = 18

    g = payload.get('group') or {}
    period = payload.get('period') or ''
    kpi = payload.get('kpi') or {}

    def _eok(d, key='final'):
        if not d: return 0
        v = d.get(key, 0) or 0
        return v / EOK

    # ─────────────────────────────────────────────────────────────
    # 시트 1: 개요 — KPI 카드 그리드 (대시보드 메인 화면 미러)
    # ─────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = '개요'
    _put_title(ws,
        f'연결 대시보드 — {g.get("name", "")} / {period}',
        f'단위: 억원  ·  최종(연결조정+내부거래 반영)  ·  전년 대비 변동 표시',
        span_cols=9)

    # 4컬럼 × 카드 그리드 (각 카드 = 2컬럼 폭)
    # 1줄: 자산 / 부채 / 자본 / 차입금
    # 2줄: 매출 / 영업이익 / 순이익 / 은행차입금
    card_grid = [
        # (행, 시작컬럼, label, kpi_dict, accent)
        (4,  1, '자산총계',   kpi.get('bs_assets'),     C_ACCENT_BLUE),
        (4,  3, '부채총계',   kpi.get('bs_liab'),       C_ACCENT_ORANGE),
        (4,  5, '자본총계',   kpi.get('bs_equity'),     C_ACCENT_GREEN),
        (4,  7, '차입금총계', kpi.get('bs_debt_total'), C_ACCENT_RED),
        (8,  1, '매출액',     kpi.get('pl_sales'),      C_ACCENT_BLUE),
        (8,  3, '영업이익',   kpi.get('pl_op'),         C_ACCENT_GREEN),
        (8,  5, '당기순이익', kpi.get('pl_ni'),         C_ACCENT_GREEN),
        (8,  7, '은행차입금', kpi.get('bs_debt_bank'),  C_ACCENT_ORANGE),
    ]
    for row, col, lbl, d, ac in card_grid:
        comp = (d or {}).get('compare')
        comp_eok = (comp / EOK) if comp is not None else None
        _kpi_card(ws, row, col, span=2, label=lbl,
                  value_eok=_eok(d), compare_eok=comp_eok, accent=ac)

    # 카드 그리드 컬럼 폭
    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 13

    # 그룹 메타 정보 표
    r0 = 13
    ws.cell(r0, 1, '그룹 정보').font = SUB_FONT
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=4)
    ws.cell(r0, 1).fill = SUB_FILL
    r0 += 1
    meta_rows = [
        ('당기 기간',     period),
        ('비교 전년',     payload.get('prior_year') or '-'),
        ('매칭 파일 수', payload.get('matched_files_count', 0)),
        ('연결조정 분개', payload.get('adjustment_count', 0)),
        ('내부거래 분개', payload.get('intercompany_count', 0)),
        ('자동매듭 적용', 'YES' if (payload.get('bridge') or {}).get('applied') else 'NO'),
        ('자동매듭 금액', (payload.get('bridge') or {}).get('amount', 0) or 0),
        ('검증 차이 (final)', payload.get('check_imbalance', 0) or 0),
    ]
    for i, (k, v) in enumerate(meta_rows):
        ws.cell(r0 + i, 1, k)
        ws.cell(r0 + i, 1).font = Font(bold=True, name='맑은 고딕', size=10, color=C_BRAND)
        ws.cell(r0 + i, 1).fill = SUB_FILL
        ws.cell(r0 + i, 1).border = BORDER
        c2 = ws.cell(r0 + i, 2, v)
        _style_data_cell(c2,
            NUM if isinstance(v, (int, float)) and not isinstance(v, bool) else None)
        c2.alignment = Alignment(
            horizontal='right' if isinstance(v, (int, float)) and not isinstance(v, bool) else 'left',
            vertical='center')

    # ─────────────────────────────────────────────────────────────
    # 시트 2: 회사별 구성 — 4개 도넛 차트 (대시보드의 도넛과 동일 데이터)
    # ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet('회사별 구성')
    _put_title(ws, '회사별 구성', '대시보드 도넛 차트 — 버킷 기준(본사·포함그룹·해외 등)', span_cols=6)

    bcg = payload.get('by_company_grouped') or {}
    # 메트릭별로 버킷 라벨 union 정렬
    METRICS = [
        ('assets', '자산총계'),
        ('sales',  '매출액'),
        ('op',     '영업이익'),
        ('ni',     '당기순이익'),
    ]

    # 데이터 표 + 차트 — 메트릭 블록을 22행 간격으로 세로 배치 (차트 높이 9cm ≈ 18행 + 여유)
    chart_anchors = ['F4', 'F26', 'F48', 'F70']
    block_start_row = 4
    for idx, (key, label) in enumerate(METRICS):
        buckets = [b for b in (bcg.get(key) or []) if (b.get('value') or 0) != 0]
        if not buckets:
            continue

        r = block_start_row + idx * 22
        # 섹션 헤더
        ws.cell(r, 1, label).font = SUB_FONT
        ws.cell(r, 1).fill = SUB_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

        ws.cell(r + 1, 1, '버킷');  ws.cell(r + 1, 2, '구분');
        ws.cell(r + 1, 3, '값(원)'); ws.cell(r + 1, 4, '값(억)')
        _style_header_row(ws, r + 1, 4)

        data_start = r + 2
        for i, b in enumerate(buckets):
            rr = data_start + i
            ws.cell(rr, 1, b.get('label') or '')
            ws.cell(rr, 2, b.get('type') or '')
            v = b.get('value', 0) or 0
            ws.cell(rr, 3, v)
            ws.cell(rr, 4, v / EOK)
            for c in range(1, 5):
                _style_data_cell(ws.cell(rr, c), NUM if c == 3 else (NUM_EOK if c == 4 else None))
        data_end = data_start + len(buckets) - 1

        # 도넛 차트
        chart = DoughnutChart()
        chart.title = label
        chart.style = 26
        chart.holeSize = 55
        labels_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        # 음수가 섞이면 도넛에 부적합 → 절댓값 컬럼(E열) 추가해서 차트에만 사용
        if any((ws.cell(rr, 3).value or 0) < 0 for rr in range(data_start, data_end + 1)):
            ws.cell(r + 1, 5, '|값|').font = HDR_FONT
            ws.cell(r + 1, 5).fill = HDR_FILL
            ws.cell(r + 1, 5).alignment = Alignment(horizontal='center')
            for rr in range(data_start, data_end + 1):
                ws.cell(rr, 5, abs(ws.cell(rr, 3).value or 0))
                _style_data_cell(ws.cell(rr, 5), NUM)
            data_ref = Reference(ws, min_col=5, min_row=data_start, max_row=data_end)
        else:
            data_ref = Reference(ws, min_col=3, min_row=data_start, max_row=data_end)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(labels_ref)
        # 라벨 — 퍼센트만, 슬라이스 밖에 표시. 값/계열명/카테고리 모두 끔(범례에 카테고리가 이미 있음).
        chart.dataLabels = DataLabelList(
            showVal=False,
            showCatName=False,
            showSerName=False,
            showPercent=True,
            showLegendKey=False,
            showBubbleSize=False,
        )
        chart.height = 9.0
        chart.width = 14.0
        ws.add_chart(chart, chart_anchors[idx])

    for c, w in enumerate([22, 16, 18, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ─────────────────────────────────────────────────────────────
    # 시트 3: 자산·부채·자본 구성 — 막대 차트 3개
    # ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet('자산·부채·자본')
    _put_title(ws, '자산 · 부채 · 자본 구성', '합산 vs 최종 비교 (단위: 억원)', span_cols=6)

    bd = payload.get('breakdown') or {}
    bd_blocks = [
        ('asset',  '자산 구성',  C_ACCENT_BLUE),
        ('liab',   '부채 구성',  C_ACCENT_ORANGE),
        ('equity', '자본 구성',  C_ACCENT_GREEN),
    ]

    row_anchor = 4
    bar_anchors = ['F4', 'F26', 'F48']
    for idx, (key, label, color) in enumerate(bd_blocks):
        items = bd.get(key) or []
        if not items:
            continue

        r = row_anchor + idx * 22
        ws.cell(r, 1, label).font = SUB_FONT
        ws.cell(r, 1).fill = SUB_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

        ws.cell(r + 1, 1, '항목'); ws.cell(r + 1, 2, '합산(억)'); ws.cell(r + 1, 3, '최종(억)')
        _style_header_row(ws, r + 1, 3)

        data_start = r + 2
        for i, it in enumerate(items):
            rr = data_start + i
            ws.cell(rr, 1, it.get('name') or '')
            ws.cell(rr, 2, (it.get('sum')   or 0) / EOK)
            ws.cell(rr, 3, (it.get('final') or 0) / EOK)
            _style_data_cell(ws.cell(rr, 1))
            _style_data_cell(ws.cell(rr, 2), NUM_EOK)
            _style_data_cell(ws.cell(rr, 3), NUM_EOK)
        data_end = data_start + len(items) - 1

        chart = BarChart()
        chart.type = 'col'        # 세로 막대 (가로 막대는 카테고리 라벨이 회전돼 안 보임)
        chart.style = 11
        chart.title = label
        chart.y_axis.title = None
        chart.x_axis.title = None
        chart.y_axis.number_format = '#,##0"억"'
        data = Reference(ws, min_col=2, max_col=3, min_row=r + 1, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        # 막대 위에 값 표시 — 억 단위
        chart.dataLabels = DataLabelList(
            showVal=True,
            showCatName=False,
            showSerName=False,
            showPercent=False,
            showLegendKey=False,
        )
        chart.dataLabels.numFmt = '#,##0"억"'
        chart.height = 9.0
        chart.width = 14.0
        ws.add_chart(chart, bar_anchors[idx])

    for c, w in enumerate([22, 16, 16, 12, 12, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ─────────────────────────────────────────────────────────────
    # 시트 4: 재무비율 — 막대 차트
    # ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet('재무비율')
    _put_title(ws, '주요 재무비율', '합산 / 최종 / 전년 비교', span_cols=6)

    for i, h in enumerate(['지표', '합산 기준', '최종 기준', '전년'], 1):
        ws.cell(4, i, h)
    _style_header_row(ws, 4, 4)
    m = payload.get('metrics') or {}
    ratio_rows = [
        ('부채비율',   m.get('debt_ratio_sum'),  m.get('debt_ratio_final'),  m.get('debt_ratio_compare')),
        ('영업이익률', m.get('op_margin_sum'),   m.get('op_margin_final'),   m.get('op_margin_compare')),
        ('순이익률',   m.get('net_margin_sum'),  m.get('net_margin_final'),  m.get('net_margin_compare')),
    ]
    for i, (label, s, f, c) in enumerate(ratio_rows, start=5):
        ws.cell(i, 1, label)
        ws.cell(i, 2, s)
        ws.cell(i, 3, f)
        ws.cell(i, 4, c)
        for col in range(1, 5):
            _style_data_cell(ws.cell(i, col), PCT if col > 1 else None)
    for c, w in enumerate([18, 16, 16, 16, 4, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    chart = BarChart()
    chart.type = 'col'
    chart.style = 11
    chart.title = '재무비율'
    data = Reference(ws, min_col=2, max_col=4, min_row=4, max_row=7)
    cats = Reference(ws, min_col=1, min_row=5, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 14
    chart.y_axis.number_format = PCT
    chart.dataLabels = DataLabelList(
        showVal=True,
        showCatName=False,
        showSerName=False,
        showPercent=False,
        showLegendKey=False,
    )
    chart.dataLabels.numFmt = PCT
    ws.add_chart(chart, 'F4')

    # ─────────────────────────────────────────────────────────────
    # 시트 5: KPI 상세 — 분개 효과까지 전 항목 (테이블)
    # ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet('KPI 상세')
    _put_title(ws, 'KPI 상세 — 분개 효과 포함', '합산 → 연결조정/내부거래 차·대 → 최종 → 전년', span_cols=10)

    KPI_HEADERS = ['구분', '항목', '코드',
                   '합산', '연결조정(차)', '연결조정(대)',
                   '내부거래(차)', '내부거래(대)',
                   '최종', '전년']
    for i, h in enumerate(KPI_HEADERS, 1):
        ws.cell(4, i, h)
    _style_header_row(ws, 4, len(KPI_HEADERS))

    BS_ROWS = [('재무상태표', '자산총계', kpi.get('bs_assets')),
               ('재무상태표', '부채총계', kpi.get('bs_liab')),
               ('재무상태표', '자본총계', kpi.get('bs_equity')),
               ('재무상태표', '차입금총계', kpi.get('bs_debt_total')),
               ('재무상태표', '은행차입금', kpi.get('bs_debt_bank'))]
    PL_ROWS = [('손익계산서', '매출액',       kpi.get('pl_sales')),
               ('손익계산서', '매출원가',     kpi.get('pl_cogs')),
               ('손익계산서', '매출총이익',   kpi.get('pl_gross')),
               ('손익계산서', '영업이익',     kpi.get('pl_op')),
               ('손익계산서', '세전이익',     kpi.get('pl_pretax')),
               ('손익계산서', '당기순이익',   kpi.get('pl_ni')),
               ('손익계산서', '지배지분순이익', kpi.get('pl_ni_owner')),
               ('손익계산서', '비지배지분순이익', kpi.get('pl_ni_nci'))]

    r = 5
    for section, label, d in (BS_ROWS + PL_ROWS):
        d = d or {}
        ws.cell(r, 1, section)
        ws.cell(r, 2, label)
        ws.cell(r, 3, d.get('code', ''))
        ws.cell(r, 4, d.get('sum', 0) or 0)
        ws.cell(r, 5, d.get('dr_adj', 0) or 0)
        ws.cell(r, 6, d.get('cr_adj', 0) or 0)
        ws.cell(r, 7, d.get('dr_int', 0) or 0)
        ws.cell(r, 8, d.get('cr_int', 0) or 0)
        ws.cell(r, 9, d.get('final', 0) or 0)
        compare = d.get('compare')
        ws.cell(r, 10, compare if compare is not None else None)
        for c in range(1, len(KPI_HEADERS) + 1):
            _style_data_cell(ws.cell(r, c), NUM if c >= 4 else None)
        # 섹션 색상 (구분 컬럼)
        ws.cell(r, 1).fill = SUB_FILL
        ws.cell(r, 1).font = Font(bold=True, color=C_BRAND, name='맑은 고딕', size=10)
        r += 1

    widths = [12, 18, 12, 16, 14, 14, 14, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'D5'

    # ─────────────────────────────────────────────────────────────
    # 시트 6: 회사별 상세 (개별 회사 단위)
    # ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet('회사별 상세')
    _put_title(ws, '회사별 상세', '개별 회사 단위 — 자산총계 / 매출액 / 영업이익 / 당기순이익', span_cols=7)

    bc = payload.get('by_company') or {}
    co_order = []
    seen = set()
    for key in ('assets', 'sales', 'op', 'ni'):
        for x in bc.get(key) or []:
            co = x.get('company')
            if co and co not in seen:
                seen.add(co)
                co_order.append((co, x.get('currency') or '', bool(x.get('is_rollup'))))

    headers = ['회사', '통화', '구분', '자산총계(억)', '매출액(억)', '영업이익(억)', '당기순이익(억)']
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    _style_header_row(ws, 4, len(headers))

    def _val(metric, co):
        for x in bc.get(metric) or []:
            if x.get('company') == co:
                return x.get('value', 0) or 0
        return 0

    r = 5
    for co, cur, is_rollup in co_order:
        ws.cell(r, 1, co)
        ws.cell(r, 2, cur)
        ws.cell(r, 3, '포함그룹(연결)' if is_rollup else '개별회사')
        ws.cell(r, 4, _val('assets', co) / EOK)
        ws.cell(r, 5, _val('sales',  co) / EOK)
        ws.cell(r, 6, _val('op',     co) / EOK)
        ws.cell(r, 7, _val('ni',     co) / EOK)
        for c in range(1, len(headers) + 1):
            _style_data_cell(ws.cell(r, c), NUM_EOK if c >= 4 else None)
        if is_rollup:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = PatternFill('solid', start_color=C_BRAND_LIGHT)
        r += 1

    widths2 = [28, 8, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths2, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'D5'

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@app.route('/consolidation/dashboard/<group_id>/excel')
@login_required
@require_permission('consol.compute')
def consolidation_dashboard_excel(group_id):
    """대시보드 내용을 엑셀(.xlsx)로 다운로드."""
    period = (request.args.get('period') or YEARS_DATA.get('default') or '').strip()
    if not _valid_year(period):
        return '유효하지 않은 기간', 400
    g = consol_get_group(group_id)
    if not g:
        return '존재하지 않는 그룹', 404
    if not _can_access_group(session.get('username'), group_id):
        return '해당 그룹에 접근 권한이 없습니다.', 403

    try:
        ctx = _compute_group_internal(group_id, period)
    except (ValueError, RuntimeError) as e:
        return f'대시보드 계산 실패: {e}', 400

    payload = _build_dashboard_payload(ctx, period)

    # rollup 펼침 적용 — JSON 응답과 동일한 구성
    sub_payload_by_label = {}
    for rl in (ctx.get('rollups') or []):
        sub_ctx = rl.get('sub_result')
        if sub_ctx:
            try:
                sub_payload_by_label[rl['name']] = _build_dashboard_payload(sub_ctx, period)
            except Exception:
                pass

    def _expand(buckets, metric_key):
        out = []
        for bk in buckets:
            if bk.get('type') == 'rollup' and bk.get('label') in sub_payload_by_label:
                sub_p = sub_payload_by_label[bk['label']]
                sub_b = (sub_p.get('by_company_grouped') or {}).get(metric_key) or []
                if sub_b:
                    out.extend(sub_b); continue
            out.append(bk)
        return out

    grouped = payload.get('by_company_grouped') or {}
    payload['by_company_grouped'] = {
        'assets': _expand(grouped.get('assets') or [], 'assets'),
        'sales':  _expand(grouped.get('sales')  or [], 'sales'),
        'op':     _expand(grouped.get('op')     or [], 'op'),
        'ni':     _expand(grouped.get('ni')     or [], 'ni'),
    }

    data = _build_dashboard_excel(payload)
    safe_name = re.sub(r'[\\/:*?"<>|]', '_', g.get('name') or 'group')
    fname = f'연결대시보드_{safe_name}_{period}.xlsx'
    return send_file(io.BytesIO(data), as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/consolidation/compute/<group_id>/<period>', methods=['POST'])
@login_required
@require_permission('consol.compute')
def consolidation_compute(group_id, period):
    """합산 + 분개 → 최종 산출. 결과 JSON + 엑셀 다운로드 URL.
    포함 그룹(included_groups)이 있으면 그 그룹들을 먼저 실행해 rollup 컬럼으로 합산.
    """
    if not _valid_year(period):
        return jsonify({'error': '유효하지 않은 결산기간'}), 400
    g = consol_get_group(group_id)
    if not g:
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '해당 그룹에 접근 권한이 없습니다.'}), 403

    try:
        ctx = _compute_group_internal(group_id, period)
    except (ValueError, RuntimeError) as e:
        return jsonify({'error': str(e)}), 400

    result = ctx['result']
    bridge_info = ctx['bridge_info']
    bridge_entries = ctx['bridge_entries']
    adj_entries = ctx['adjustment_entries']
    inter_entries = ctx['intercompany_entries']
    files = ctx['files']
    rollups = ctx['rollups']
    companies = result['companies']

    # 전년 4Q 연결정산표 최종값을 같은 그룹으로 재계산해 각 row에 prior_final로 주입
    prior_period = _prior_q4_period(period)
    prior_final_by_code = {}
    if prior_period and _valid_year(prior_period):
        try:
            prior_ctx = _compute_group_internal(group_id, prior_period)
            for prow in prior_ctx['result']['rows']:
                pcode = str(prow.get('code') or '').strip()
                if pcode and prow.get('kind') in ('detail', 'subtotal', 'formula', 'check'):
                    prior_final_by_code[pcode] = prow.get('final')
        except Exception:
            prior_period = None
            prior_final_by_code = {}
    else:
        prior_period = None

    for row in result['rows']:
        rcode = str(row.get('code') or '').strip()
        row['prior_final'] = prior_final_by_code.get(rcode) if rcode else None

    result['prior_period'] = prior_period

    try:

        # 엑셀 저장
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', g['name'])
        out_name = f'연결정산_{safe_name}_{period}.xlsx'
        out_path = RESULTS_DIR / out_name
        consol_write_excel(result, g['name'], period, str(out_path))

        # JSON 응답 (UI 표시용)
        return jsonify({
            'ok': True,
            'group': g,
            'period': period,
            'prior_period': prior_period,
            'companies': companies,
            'rows': result['rows'],
            'adjustment_count': len(adj_entries),
            'intercompany_count': len(inter_entries),
            'bridge': {
                'applied': bridge_info['applied'],
                'amount': bridge_info['bridge_amount'],
                'bs_dr': bridge_info['bs_dr'],
                'bs_cr': bridge_info['bs_cr'],
                'pl_dr': bridge_info['pl_dr'],
                'pl_cr': bridge_info['pl_cr'],
                'entries': bridge_entries,
            },
            'rollups': [
                {'group_id': r['group_id'],
                 'name': r['name'],
                 'group_name': r['sub_result']['group']['name'],
                 'companies_count': len(r['sub_result']['files']),
                 'adjustment_count': len(r['sub_result']['adjustment_entries']),
                 'intercompany_count': len(r['sub_result']['intercompany_entries']),
                 'bridge_applied': r['sub_result']['bridge_info']['applied'],
                 'bridge_amount': r['sub_result']['bridge_info']['bridge_amount']}
                for r in rollups
            ],
            'file': out_name,
            'download_url': url_for('download_result', filename=out_name),
            'matched_files': [
                {'company': f.get('company'),
                 'currency': (f.get('extracted') or {}).get('currency'),
                 'file': f.get('original_name')}
                for f in files
            ],
        })
    except Exception as e:
        return _json_error(e)


# ─────────────────────────────────────────────────────────────────────────────
# 현금정산표 (연결정산표 패턴 미러링, CF 시트 기반)
# ─────────────────────────────────────────────────────────────────────────────
CASH_MANUAL_PATH = Path('cash_worksheet_manuals.json')
# 현금정산표 ↔ 연결정산표 비교 대상 계정 리스트
CF_COMPARE_XLSX_PATHS = [
    Path(r'C:/연결시스템(26.1분기)/현표비교계정.xlsx'),
    Path('현표비교계정.xlsx'),
]


def _inject_consol_compare(cf_result, ctx):
    """cf_result의 각 detail 행에 연결정산표 비교값/차이를 주입.

    - 현표비교계정.xlsx의 코드만 대상
    - row['consol_value']  : 연결정산표 detail row의 final
    - row['consol_diff']   : cf row final − consol_value
    - row['has_compare']   : True
    cf_result['compare_codes'] 에 비교 코드 dict 저장.
    """
    compare = _cf_load_compare_codes()
    if not compare:
        cf_result['compare_codes'] = {}
        return
    # 연결정산표 detail 행 조회 index
    consol_by_code = {}
    for row in (ctx.get('result') or {}).get('rows') or []:
        code = str(row.get('code') or '')
        if code and row.get('kind') == 'detail':
            consol_by_code[code] = float(row.get('final') or 0)

    for sec in cf_result.get('sections', []):
        for row in sec.get('rows', []):
            code = str(row.get('cf_code') or '')
            if code in compare:
                cv = consol_by_code.get(code)
                row['has_compare']  = True
                row['consol_value'] = cv if cv is not None else 0.0
                row['consol_diff']  = float(row.get('final') or 0) - (cv if cv is not None else 0.0)
                row['consol_available'] = cv is not None

    cf_result['compare_codes'] = compare


def _cf_load_compare_codes():
    """현표비교계정.xlsx에서 비교 대상 코드 → 계정명 dict 반환.
    파일이 없으면 빈 dict.
    """
    out = {}
    for p in CF_COMPARE_XLSX_PATHS:
        if not p.exists():
            continue
        try:
            from openpyxl import load_workbook as _lw
            wb = _lw(str(p), data_only=True, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if not row or row[0] in (None, ''):
                    continue
                code = str(row[0]).strip()
                # 헤더 행 스킵
                if not (code and (code[0].isdigit() or code.upper().startswith('CF'))):
                    continue
                name = (row[1] if len(row) > 1 else '') or ''
                out[code] = str(name).strip()
            wb.close()
            break  # 첫 번째 존재하는 파일만
        except Exception:
            continue
    return out


def _cf_load_manuals():
    if not CASH_MANUAL_PATH.exists():
        return {}
    try:
        with open(CASH_MANUAL_PATH, encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}


def _cf_save_manuals(data):
    with open(CASH_MANUAL_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _cf_get_manuals(group_id, period):
    """{cf_code: amount} dict 반환. 없으면 빈 dict."""
    key = f'{group_id}__{period}'
    return (_cf_load_manuals().get(key) or {}).get('manuals') or {}


def _cf_get_comments(group_id, period):
    """{cf_code: text} dict 반환. 없으면 빈 dict."""
    key = f'{group_id}__{period}'
    return (_cf_load_manuals().get(key) or {}).get('comments') or {}


def _cf_get_roundings(group_id, period):
    """{cf_code: amount} dict 반환 — Q컬럼 단수조정. 없으면 빈 dict."""
    key = f'{group_id}__{period}'
    return (_cf_load_manuals().get(key) or {}).get('roundings') or {}


def _cf_set_manuals(group_id, period, manuals, username,
                    comments=None, roundings=None):
    """manuals = {cf_code: number}, comments = {cf_code: text},
    roundings = {cf_code: number} 저장.
    comments/roundings=None이면 기존값 보존.
    """
    key = f'{group_id}__{period}'
    data = _cf_load_manuals()
    prev = data.get(key) or {}
    if comments is None:
        comments_out = prev.get('comments') or {}
    else:
        comments_out = {str(k): str(v).strip() for k, v in (comments or {}).items()
                        if str(v or '').strip()}
    if roundings is None:
        roundings_out = prev.get('roundings') or {}
    else:
        roundings_out = {str(k): float(v or 0) for k, v in (roundings or {}).items()
                         if float(v or 0) != 0}
    data[key] = {
        'group_id': group_id,
        'period': period,
        'manuals': {str(k): float(v or 0) for k, v in (manuals or {}).items()},
        'comments':  comments_out,
        'roundings': roundings_out,
        'updated_by': username,
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    _cf_save_manuals(data)
    return data[key]


@app.route('/cash-worksheet')
@login_required
def cash_worksheet_index():
    """현금정산표 메인 페이지."""
    year = request.args.get('year') or YEARS_DATA.get('default')
    if not _valid_year(year):
        year = YEARS_DATA.get('default')
    return render_template('cash_worksheet.html',
                           year=year,
                           years=YEARS_DATA['years'],
                           username=session.get('username'),
                           is_admin=_is_admin(session.get('username')))


@app.route('/cash-worksheet/compute/<group_id>/<period>', methods=['POST'])
@login_required
@require_permission('cash.compute')
def cash_worksheet_compute(group_id, period):
    """그룹별 현금정산표 계산. consolidation_compute()와 동일한 ctx 사용."""
    if not _valid_year(period):
        return jsonify({'error': '유효하지 않은 결산기간'}), 400
    g = consol_get_group(group_id)
    if not g:
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '해당 그룹에 접근 권한이 없습니다.'}), 403

    try:
        ctx = _compute_group_internal(group_id, period)
    except (ValueError, RuntimeError) as e:
        return jsonify({'error': str(e)}), 400

    agg = ctx['agg']
    # 현금정산표용 companies는 직접 회사만 (rollup은 cf_compute에서 별도 컬럼으로 추가).
    # ctx['result']['companies']는 이미 rollup이 합쳐진 확장 리스트라 사용 금지.
    companies = list(agg.get('companies') or [])
    adj_entries   = ctx['adjustment_entries']
    inter_entries = ctx['intercompany_entries']
    bridge_entries = ctx['bridge_entries']
    effective_adj = adj_entries + bridge_entries

    # 전기(전년 4Q) 내부거래 분개 — BS 잔액 차분 산출용
    prior_period = cf_prior_year_4q(period)
    prior_inter_entries = []
    if prior_period:
        prior_rec = consol_get_journal(group_id, prior_period) or {}
        prior_inter_entries = prior_rec.get('intercompany_entries') or []

    manuals  = _cf_get_manuals(group_id, period)
    comments = _cf_get_comments(group_id, period)
    roundings = _cf_get_roundings(group_id, period)

    # 연결정산표 최종 NI 추출 — 4700004 당기순이익 행의 final.
    # 4700004는 PL 합계 formula 행 (= 매출 - 매출원가 + ... - 법인세). detail이 아님.
    # NI 행을 plug 보정해서 현금정산표 최종 NI = 연결정산표 최종 NI 가 되도록 사용.
    target_final_ni = None
    try:
        for row in (ctx['result'] or {}).get('rows') or []:
            if str(row.get('code') or '') == '4700004':
                target_final_ni = float(row.get('final') or 0)
                break
    except Exception:
        target_final_ni = None

    # 글로벌세아 그룹 전용 — CF1/CF2/CF3 연결범위회사 대여금/차입금 자금조정
    # ctx['agg']는 직접 회사만 합산되어 sub group(상역 등) 회사가 빠지므로
    # 모든 leaf 회사를 다시 모아 별도 aggregate 후 계산.
    fund_adj = (_compute_global_sae_fund_adj(group_id, period)
                if g.get('name') == '글로벌세아' else None)

    try:
        cf_result = cf_compute(agg, effective_adj, inter_entries, companies,
                               manual_adjustments=manuals,
                               rounding_adjustments=roundings,
                               prior_inter_entries=prior_inter_entries,
                               target_final_ni=target_final_ni,
                               rollups=(ctx.get('rollups') or None),
                               fund_adjustments=fund_adj)
    except Exception as e:
        return _json_error(e)

    # 연결정산표 비교 컬럼 주입 (현표비교계정.xlsx의 코드만)
    _inject_consol_compare(cf_result, ctx)

    # 엑셀 저장
    try:
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', g['name'])
        out_name = f'현금정산_{safe_name}_{period}.xlsx'
        out_path = RESULTS_DIR / out_name
        cf_write_excel(cf_result, g['name'], period, str(out_path))
        download_url = url_for('download_result', filename=out_name)
    except Exception as e:
        out_name = None
        download_url = None
        cf_result['_excel_error'] = str(e)

    return jsonify({
        'ok': True,
        'group': g,
        'period': period,
        'prior_period': prior_period,
        'prior_intercompany_count': len(prior_inter_entries),
        'result': cf_result,
        'adjustment_count': len(adj_entries),
        'intercompany_count': len(inter_entries),
        'bridge_count': len(bridge_entries),
        'manuals': manuals,
        'comments': comments,
        'roundings': roundings,
        'target_final_ni': target_final_ni,
        'file': out_name,
        'download_url': download_url,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 시계열 분석 (연도별 핵심지표) — 연결그룹 "글로벌세아" 전용 기본 표시
# ─────────────────────────────────────────────────────────────────────────────

# 손익 항목 연환산 계수: 분기 손익은 연초누적(YTD)으로 저장되므로
# 연환산 = YTD ÷ 경과분기수 × 4  (1Q→×4, 2Q→×2, 3Q→×1.33, 4Q→×1)
# 시계열 분석 — 선택 가능한 시작 결산기간 하한 (2025년 4분기부터)
TS_START_MIN = '2025-4Q'


def _ts_period_to_idx(p):
    """'2025-4Q' → 절대 분기 인덱스(2025*4 + (4-1)). 형식 오류 시 None."""
    m = PERIOD_RE.match(p or '')
    if not m:
        return None
    y, q = int(m.group(1)), int(m.group(2))
    return y * 4 + (q - 1)


def _ts_idx_to_period(idx):
    """절대 분기 인덱스 → '2025-4Q'."""
    y, q0 = divmod(idx, 4)
    return f'{y}-{q0 + 1}Q'


def _ts_quarter_options():
    """선택 가능한 분기 목록(오름차순). 2025-4Q부터 최신 결산기간까지."""
    start = _ts_period_to_idx(TS_START_MIN)
    maxp = start
    for y in YEARS_DATA.get('years') or []:
        i = _ts_period_to_idx(y)
        if i is not None and i > maxp:
            maxp = i
    return [_ts_idx_to_period(i) for i in range(start, maxp + 1)]


def _ts_annualize_factor(period):
    try:
        q = int(str(period).split('-')[1].lower().replace('q', '').strip())
    except Exception:
        return 1.0
    return (4.0 / q) if q else 1.0


# 현금정산표(cf_compute) 결과에서 상각비 합계(연환산 전) 추출.
#   감가상각비(4300301/5300301) + 무형자산상각비(4300302)
#   + 사용권자산상각비(4300303/5300303)
# cash_worksheet_compute() 와 동일한 입력으로 계산.
_TS_AMORT_CODES = {'4300301', '5300301', '4300302', '4300303', '5300303'}

def _ts_amort_from_cf(ctx, group_id, period):
    try:
        g = consol_get_group(group_id) or {}
        agg = ctx['agg']
        companies = list(agg.get('companies') or [])
        effective_adj = ctx['adjustment_entries'] + ctx['bridge_entries']
        inter_entries = ctx['intercompany_entries']

        prior_period = cf_prior_year_4q(period)
        prior_inter_entries = []
        if prior_period:
            prior_rec = consol_get_journal(group_id, prior_period) or {}
            prior_inter_entries = prior_rec.get('intercompany_entries') or []

        manuals   = _cf_get_manuals(group_id, period)
        roundings = _cf_get_roundings(group_id, period)

        target_final_ni = None
        for row in (ctx['result'] or {}).get('rows') or []:
            if str(row.get('code') or '') == '4700004':
                target_final_ni = float(row.get('final') or 0)
                break

        fund_adj = (_compute_global_sae_fund_adj(group_id, period)
                    if g.get('name') == '글로벌세아' else None)

        cf_result = cf_compute(agg, effective_adj, inter_entries, companies,
                               manual_adjustments=manuals,
                               rounding_adjustments=roundings,
                               prior_inter_entries=prior_inter_entries,
                               target_final_ni=target_final_ni,
                               rollups=(ctx.get('rollups') or None),
                               fund_adjustments=fund_adj)
    except Exception:
        return 0.0

    total = 0.0
    for section in (cf_result.get('sections') or []):
        for row in (section.get('rows') or []):
            if str(row.get('cf_code') or '') in _TS_AMORT_CODES:
                total += float(row.get('final') or 0)
    return total


_TS_DEBT_CODES = ['2100201', '2100202', '2100203', '2100204', '2100205',
                  '2100301', '2100391', '2200101', '2200201', '2200291']

def _ts_compute_period(group_id, period):
    """해당 기간의 연결 raw 수치(연환산 전) 반환. 패키지 데이터가 없으면 None."""
    try:
        ctx = _compute_group_internal(group_id, period)
    except (ValueError, RuntimeError):
        return None
    if not (ctx.get('files') or []):
        return None

    rows = (ctx.get('result') or {}).get('rows') or []
    rbc = {str(r.get('code')): r for r in rows if r.get('code')}

    def fin(code):
        return float((rbc.get(code) or {}).get('final') or 0)

    return {
        'sales':      fin('4100000'),  # 매출액
        'op':         fin('4700002'),  # 영업이익
        'ni':         fin('4700004'),  # 당기순이익
        'interest':   fin('4500201'),  # 이자비용
        'cash':       fin('1110101'),  # 현금및현금성자산
        'assets':     fin('1000000'),  # 총자산
        'liab':       fin('2000000'),  # 총부채
        'equity':     fin('3000000'),  # 총자본
        'debt_total': sum(fin(c) for c in _TS_DEBT_CODES),  # 총차입금
        'amort':      _ts_amort_from_cf(ctx, group_id, period),  # 상각비(현금정산표)
    }


def _ts_build_column(period, raw, factor):
    m = PERIOD_RE.match(period or '')
    if m:
        year = int(m.group(1))
        q = int(m.group(2))
        label = f"{str(year)[2:]}년 {q}Q"
    else:
        year = 0
        label = period or ''
    if not raw:
        return {'year': year, 'label': label, 'period': period,
                'has_data': False, 'metrics': {}}

    f = factor
    sales    = raw['sales'] * f
    op       = raw['op'] * f
    amort    = raw['amort'] * f
    ni       = raw['ni'] * f
    interest = raw['interest'] * f
    ebitda   = op + amort

    debt   = raw['debt_total']           # BS — 연환산 안 함
    cash   = raw['cash']
    assets = raw['assets']
    liab   = raw['liab']
    equity = raw['equity']

    return {
        'year': year, 'label': label, 'period': period, 'has_data': True,
        'metrics': {
            'sales':       sales,
            'op':          op,
            'op_margin':   (op / sales) if sales else None,           # 영업이익률
            'amort':       amort,
            'ni':          ni,
            'ebitda':      ebitda,
            'debt_total':  debt,
            'cash':        cash,
            'net_debt':    debt - cash,
            'interest':    interest,
            'assets':      assets,
            'liab':        liab,
            'equity':      equity,
            'debt_ratio':  (liab / equity) if equity else None,        # 부채비율
            'icr':         (ebitda / abs(interest)) if interest else None,  # 이자보상비율
        },
    }


# 시계열 표 행 정의 — 화면(timeseries.html ROWS)과 동일 순서.
#   kind: 'amt'(억원), 'pct'(부채비율 %), 'x'(이자보상비율 배)
_TS_EXCEL_ROWS = [
    ('sales',      '매출액',          'amt'),
    ('op',         '영업이익',        'amt'),
    ('op_margin',  '영업이익률(%)',   'pct'),
    ('amort',      '상각비',          'amt'),
    ('ni',         '당기순이익',      'amt'),
    ('ebitda',     'EBITDA',          'amt'),
    ('debt_total', '총차입금',        'amt'),
    ('cash',       '현금성자산',      'amt'),
    ('net_debt',   'Net Debt',        'amt'),
    ('interest',   '이자비용',        'amt'),
    ('assets',     '총자산',          'amt'),
    ('liab',       '총부채',          'amt'),
    ('equity',     '총자본',          'amt'),
    ('debt_ratio', '부채비율(%)',     'pct'),
    ('icr',        '이자보상비율(배)', 'x'),
]


def _ts_resolve_range(start, end):
    """시작/종료 결산기간 문자열 → (s_idx, e_idx) 또는 (None, 오류메시지).

    - start 미지정 시 TS_START_MIN, end 미지정 시 start 로 보정.
    - start 가 2025-4Q 이전이면 2025-4Q 로 끌어올림.
    """
    min_idx = _ts_period_to_idx(TS_START_MIN)
    s_idx = _ts_period_to_idx(start) if start else min_idx
    e_idx = _ts_period_to_idx(end) if end else s_idx
    if s_idx is None or e_idx is None:
        return None, '기간 형식 오류'
    if s_idx < min_idx:
        s_idx = min_idx
    if e_idx < s_idx:
        return None, '종료 결산기간이 시작 결산기간보다 빠릅니다.'
    if e_idx - s_idx > 40:
        return None, '한 번에 조회 가능한 기간은 최대 40분기입니다.'
    return (s_idx, e_idx), None


def _ts_columns_for_range(group_id, s_idx, e_idx):
    """s_idx~e_idx 분기별 시계열 컬럼 리스트."""
    cols = []
    for idx in range(s_idx, e_idx + 1):
        col_period = _ts_idx_to_period(idx)
        raw = _ts_compute_period(group_id, col_period)
        cols.append(_ts_build_column(col_period, raw, _ts_annualize_factor(col_period)))
    return cols


@app.route('/timeseries')
@login_required
def timeseries_index():
    """시계열 분석 페이지. 기본 대상 연결그룹 = 글로벌세아."""
    target = None
    for grp in consol_list_groups():
        if (grp.get('name') or '').strip() == '글로벌세아':
            target = grp
            break
    options = _ts_quarter_options()  # 오름차순, 2025-4Q부터 최신까지
    default_start = TS_START_MIN
    default_end = options[-1] if options else TS_START_MIN
    return render_template('timeseries.html',
                           options=options,
                           default_start=default_start, default_end=default_end,
                           group_id=(target or {}).get('id', ''),
                           group_name=(target or {}).get('name', '글로벌세아'),
                           username=session.get('username'),
                           is_admin=_is_admin(session.get('username')))


@app.route('/timeseries/<group_id>/data')
@login_required
def timeseries_data(group_id):
    """시작~종료 결산기간 사이의 분기별 시계열 데이터.

    query: start=2025-4Q, end=2026-1Q (둘 다 매분기 단위)
    하위호환: period 단일값이 오면 start=end=period 로 처리.
    """
    start = (request.args.get('start') or '').strip()
    end   = (request.args.get('end') or '').strip()
    if not start and not end:
        period = (request.args.get('period') or '').strip()
        start = end = period

    rng, err = _ts_resolve_range(start, end)
    if err:
        return jsonify({'error': err}), 400
    s_idx, e_idx = rng

    g = consol_get_group(group_id)
    if not g:
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '해당 그룹에 접근 권한이 없습니다.'}), 403

    columns = _ts_columns_for_range(group_id, s_idx, e_idx)

    return jsonify({'ok': True, 'group': g,
                    'start': _ts_idx_to_period(s_idx), 'end': _ts_idx_to_period(e_idx),
                    'columns': columns})


@app.route('/timeseries/<group_id>/download')
@login_required
def timeseries_download(group_id):
    """시작~종료 결산기간 시계열 표를 엑셀(.xlsx)로 다운로드.

    query: start=2025-4Q, end=2026-1Q (화면과 동일한 분기 컬럼/행)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    start = (request.args.get('start') or '').strip()
    end   = (request.args.get('end') or '').strip()
    rng, err = _ts_resolve_range(start, end)
    if err:
        return jsonify({'error': err}), 400
    s_idx, e_idx = rng

    g = consol_get_group(group_id)
    if not g:
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '해당 그룹에 접근 권한이 없습니다.'}), 403

    columns = _ts_columns_for_range(group_id, s_idx, e_idx)

    wb = Workbook()
    ws = wb.active
    ws.title = '시계열분석'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F3864')
    label_fill  = PatternFill('solid', fgColor='EBF0FA')
    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right', vertical='center')
    thin = Side(style='thin', color='D0D7E2')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ncol = 1 + len(columns)

    # 1행: 제목, 2행: 단위
    ws.cell(1, 1).value = (f"{g['name']} 시계열 분석  "
                           f"({_ts_idx_to_period(s_idx)} ~ {_ts_idx_to_period(e_idx)})")
    ws.cell(1, 1).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.cell(2, 1).value = '단위: 억원 (비율 제외) · 분기 결산의 손익 항목은 연환산'
    ws.cell(2, 1).font = Font(size=9, italic=True, color='888888')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)

    # 4행: 헤더 (구분 + 분기별 라벨/기간)
    hr = 4
    hc = ws.cell(hr, 1)
    hc.value = '구분'
    hc.font = header_font; hc.fill = header_fill; hc.alignment = center; hc.border = border
    for j, col in enumerate(columns, start=2):
        sub = col['period'] if col.get('has_data') else f"{col['period']} (데이터 없음)"
        cell = ws.cell(hr, j)
        cell.value = f"{col['label']}\n{sub}"
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border

    # 데이터 행
    r = hr + 1
    for key, label, kind in _TS_EXCEL_ROWS:
        lc = ws.cell(r, 1)
        lc.value = label; lc.font = bold; lc.fill = label_fill; lc.border = border
        for j, col in enumerate(columns, start=2):
            cell = ws.cell(r, j)
            cell.border = border; cell.alignment = right
            v = (col.get('metrics') or {}).get(key) if col.get('has_data') else None
            if v is None:
                continue
            if kind == 'amt':
                cell.value = round(float(v) / 1e8, 1)
                cell.number_format = '#,##0.0'
            elif kind == 'pct':
                cell.value = round(float(v) * 100, 1)
                cell.number_format = '#,##0.0"%"'
            elif kind == 'x':
                cell.value = round(float(v), 2)
                cell.number_format = '#,##0.00"배"'
        r += 1

    ws.column_dimensions['A'].width = 16
    for j in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13
    ws.row_dimensions[hr].height = 32
    ws.freeze_panes = 'B5'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = re.sub(r'[\\/:*?"<>|]', '_', g['name'])
    fname = (f"시계열분석_{safe}_"
             f"{_ts_idx_to_period(s_idx)}_{_ts_idx_to_period(e_idx)}.xlsx")
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/cash-worksheet/download-excel/<group_id>/<period>', methods=['GET'])
@login_required
@require_permission('cash.compute')
def cash_worksheet_download_excel(group_id, period):
    """현금정산표 엑셀 다운로드 — hide_zero 옵션 적용해 즉석 생성.

    GET query: hide_zero=1/true → 값 모두 0인 detail 행 제외
    """
    if not _valid_year(period):
        return jsonify({'error': '유효하지 않은 결산기간'}), 400
    g = consol_get_group(group_id)
    if not g:
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '해당 그룹에 접근 권한이 없습니다.'}), 403

    hide_zero = (request.args.get('hide_zero') or '').lower() in ('1', 'true', 'yes', 'on')

    try:
        ctx = _compute_group_internal(group_id, period)
    except (ValueError, RuntimeError) as e:
        return jsonify({'error': str(e)}), 400

    agg = ctx['agg']
    # 현금정산표용 companies는 직접 회사만 (rollup은 cf_compute에서 별도 컬럼으로 추가).
    # ctx['result']['companies']는 이미 rollup이 합쳐진 확장 리스트라 사용 금지.
    companies = list(agg.get('companies') or [])
    adj_entries   = ctx['adjustment_entries']
    inter_entries = ctx['intercompany_entries']
    bridge_entries = ctx['bridge_entries']
    effective_adj = adj_entries + bridge_entries

    prior_period = cf_prior_year_4q(period)
    prior_inter_entries = []
    if prior_period:
        prior_rec = consol_get_journal(group_id, prior_period) or {}
        prior_inter_entries = prior_rec.get('intercompany_entries') or []

    manuals   = _cf_get_manuals(group_id, period)
    roundings = _cf_get_roundings(group_id, period)

    target_final_ni = None
    try:
        for row in (ctx['result'] or {}).get('rows') or []:
            if str(row.get('code') or '') == '4700004':
                target_final_ni = float(row.get('final') or 0)
                break
    except Exception:
        target_final_ni = None

    # 글로벌세아 그룹 전용 — CF1/CF2/CF3 연결범위회사 대여금/차입금 자금조정
    # ctx['agg']는 직접 회사만 합산되어 sub group(상역 등) 회사가 빠지므로
    # 모든 leaf 회사를 다시 모아 별도 aggregate 후 계산.
    fund_adj = (_compute_global_sae_fund_adj(group_id, period)
                if g.get('name') == '글로벌세아' else None)

    try:
        cf_result = cf_compute(agg, effective_adj, inter_entries, companies,
                               manual_adjustments=manuals,
                               rounding_adjustments=roundings,
                               prior_inter_entries=prior_inter_entries,
                               target_final_ni=target_final_ni,
                               rollups=(ctx.get('rollups') or None),
                               fund_adjustments=fund_adj)
    except Exception as e:
        return _json_error(e)

    # 연결정산표 비교 컬럼 주입
    _inject_consol_compare(cf_result, ctx)

    try:
        bio = cf_write_excel(cf_result, g['name'], period, None, hide_zero=hide_zero)
    except Exception as e:
        return _json_error(e)

    safe_name = re.sub(r'[\\/:*?"<>|]', '_', g['name'])
    suffix = '_compact' if hide_zero else ''
    out_name = f'현금정산_{safe_name}_{period}{suffix}.xlsx'
    return send_file(
        bio,
        as_attachment=True,
        download_name=out_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/cash-worksheet/manual/<group_id>/<period>', methods=['GET'])
@login_required
@require_permission('cash.compute')
def cash_worksheet_manual_get(group_id, period):
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '권한이 없습니다.'}), 403
    return jsonify({
        'manuals':   _cf_get_manuals(group_id, period),
        'comments':  _cf_get_comments(group_id, period),
        'roundings': _cf_get_roundings(group_id, period),
    })


def _consol_cash_final(group_id, period, code='1110101'):
    """연결정산표 detail 행(현금및현금등가물)의 final 값. 실패 시 None."""
    try:
        ctx = _compute_group_internal(group_id, period)
    except Exception:
        return None
    for row in (ctx.get('result') or {}).get('rows') or []:
        if str(row.get('code') or '') == code and row.get('kind') == 'detail':
            return float(row.get('final') or 0)
    return None


@app.route('/cash-worksheet/validate/<group_id>/<period>', methods=['GET'])
@login_required
@require_permission('cash.compute')
def cash_worksheet_validate(group_id, period):
    """현금정산표 3가지 검증:
      1) 현금증감(Ⅳ) + 환율변동효과(Ⅴ) + 연결범위변동 + 기초현금(Ⅵ) = 기말현금(Ⅶ)
      2) 기초현금(Ⅵ) = 전년 4Q 연결정산표 1110101 (현금및현금등가물)
      3) 기말현금(Ⅶ) = 당기   연결정산표 1110101 (현금및현금등가물)
    """
    if not _valid_year(period):
        return jsonify({'error': '유효하지 않은 결산기간'}), 400
    g = consol_get_group(group_id)
    if not g:
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '해당 그룹에 접근 권한이 없습니다.'}), 403

    # 현금정산표 재계산 (compute 엔드포인트와 동일 흐름)
    try:
        ctx = _compute_group_internal(group_id, period)
    except (ValueError, RuntimeError) as e:
        return jsonify({'error': str(e)}), 400

    agg = ctx['agg']
    # 현금정산표용 companies는 직접 회사만 (rollup은 cf_compute에서 별도 컬럼으로 추가).
    # ctx['result']['companies']는 이미 rollup이 합쳐진 확장 리스트라 사용 금지.
    companies = list(agg.get('companies') or [])
    adj_entries   = ctx['adjustment_entries']
    inter_entries = ctx['intercompany_entries']
    bridge_entries = ctx['bridge_entries']
    effective_adj = adj_entries + bridge_entries

    prior_period = cf_prior_year_4q(period)
    prior_inter_entries = []
    if prior_period:
        prior_rec = consol_get_journal(group_id, prior_period) or {}
        prior_inter_entries = prior_rec.get('intercompany_entries') or []

    manuals   = _cf_get_manuals(group_id, period)
    roundings = _cf_get_roundings(group_id, period)

    target_final_ni = None
    for row in (ctx['result'] or {}).get('rows') or []:
        if str(row.get('code') or '') == '4700004':
            target_final_ni = float(row.get('final') or 0)
            break

    fund_adj = (_compute_global_sae_fund_adj(group_id, period)
                if g.get('name') == '글로벌세아' else None)
    cf_result = cf_compute(agg, effective_adj, inter_entries, companies,
                           manual_adjustments=manuals,
                           rounding_adjustments=roundings,
                           prior_inter_entries=prior_inter_entries,
                           target_final_ni=target_final_ni,
                           rollups=(ctx.get('rollups') or None),
                           fund_adjustments=fund_adj)

    # 현금정산표 4행 값
    net   = float((cf_result.get('net_cash')     or {}).get('final', 0) or 0)
    fxv   = float((cf_result.get('fx_effect')    or {}).get('final', 0) or 0)
    scp   = float((cf_result.get('scope_change') or {}).get('final', 0) or 0)
    beg   = float((cf_result.get('cash_begin')   or {}).get('final', 0) or 0)
    end   = float((cf_result.get('cash_end')     or {}).get('final', 0) or 0)

    # 연결정산표 현금및현금등가물(1110101) — 당기는 ctx['result']에서, 전년 4Q는 별도 계산
    consol_end_cash = None
    for row in (ctx['result'] or {}).get('rows') or []:
        if str(row.get('code') or '') == '1110101' and row.get('kind') == 'detail':
            consol_end_cash = float(row.get('final') or 0)
            break
    consol_begin_cash = (_consol_cash_final(group_id, prior_period)
                         if prior_period else None)

    # 허용 오차 — KRW 반올림 1원
    TOL = 0.5

    check1_lhs = net + fxv + scp + beg
    check1_diff = round(check1_lhs - end)
    check1_ok   = abs(check1_lhs - end) <= TOL

    check2_diff = (round(beg - consol_begin_cash)
                   if consol_begin_cash is not None else None)
    check2_ok   = (consol_begin_cash is not None
                   and abs(beg - consol_begin_cash) <= TOL)

    check3_diff = (round(end - consol_end_cash)
                   if consol_end_cash is not None else None)
    check3_ok   = (consol_end_cash is not None
                   and abs(end - consol_end_cash) <= TOL)

    return jsonify({
        'period':       period,
        'prior_period': prior_period,
        'checks': [
            {
                'name': '현금증감 등식',
                'formula': 'Ⅳ.현금증감 + Ⅴ.환율변동 + 연결범위변동 + Ⅵ.기초현금 = Ⅶ.기말현금',
                'ok': check1_ok,
                'components': {
                    'net_cash':     net,
                    'fx_effect':    fxv,
                    'scope_change': scp,
                    'cash_begin':   beg,
                    'cash_end':     end,
                    'lhs_sum':      check1_lhs,
                },
                'diff': check1_diff,
            },
            {
                'name': '기초현금 = 전년 4Q 연결정산표 현금및현금등가물(1110101)',
                'ok':   check2_ok,
                'available': consol_begin_cash is not None,
                'cf_value':     beg,
                'consol_value': consol_begin_cash,
                'consol_period': prior_period or '',
                'diff': check2_diff,
                'note': '' if consol_begin_cash is not None else
                        ('전년 4Q 데이터 없음 — 검증 스킵' if prior_period else '전기 기간 산출 실패'),
            },
            {
                'name': '기말현금 = 당기 연결정산표 현금및현금등가물(1110101)',
                'ok':   check3_ok,
                'available': consol_end_cash is not None,
                'cf_value':     end,
                'consol_value': consol_end_cash,
                'consol_period': period,
                'diff': check3_diff,
                'note': '' if consol_end_cash is not None else '연결정산표에서 1110101 행을 찾지 못함',
            },
        ],
    })


@app.route('/cash-worksheet/manual/<group_id>/<period>', methods=['POST'])
@login_required
@require_permission('cash.compute')
def cash_worksheet_manual_set(group_id, period):
    # 수기입력(수기조정·단수조정·코멘트) 저장은 관리자만 허용
    if not _is_admin(session.get('username')):
        return jsonify({'error': '현금정산표 수기입력은 관리자만 가능합니다.'}), 403
    if not _valid_year(period):
        return jsonify({'error': '유효하지 않은 결산기간'}), 400
    if not consol_get_group(group_id):
        return jsonify({'error': '존재하지 않는 그룹'}), 404
    if not _can_access_group(session.get('username'), group_id):
        return jsonify({'error': '해당 그룹에 접근 권한이 없습니다.'}), 403
    data = request.get_json(silent=True) or {}
    manuals = data.get('manuals') or {}
    # comments/roundings 키가 명시적으로 들어왔을 때만 갱신 (없으면 기존 값 보존)
    comments  = data.get('comments')  if 'comments'  in data else None
    roundings = data.get('roundings') if 'roundings' in data else None
    rec = _cf_set_manuals(group_id, period, manuals,
                          session.get('username') or '?',
                          comments=comments, roundings=roundings)
    return jsonify({'ok': True, 'record': rec})


@app.route('/cash-worksheet/mapping')
@login_required
def cash_worksheet_mapping():
    """매핑 뷰어. ?format=json 이면 raw JSON, 아니면 HTML 테이블."""
    try:
        mapping = cf_load_mapping()
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    if request.args.get('format') == 'json':
        return jsonify(mapping)

    # v2: COA 행 단위 표 — cf_by_code 인덱스 전달
    cf_by_code = {cl['cf_code']: cl for cl in (mapping.get('cf_lines') or [])}
    total_rows = sum(len(s.get('rows') or []) for s in mapping.get('sections', []))
    return render_template('cash_worksheet_mapping.html',
                           mapping=mapping,
                           cf_by_code=cf_by_code,
                           total_rows=total_rows,
                           username=session.get('username'),
                           is_admin=_is_admin(session.get('username')))


@app.route('/cash-worksheet/mapping/save', methods=['POST'])
@login_required
@require_permission('cash.mapping')
def cash_worksheet_mapping_save():
    """COA 매핑 일괄 저장. body: {coa: [...]}.

    각 COA 객체에서 사용하는 필드: code, adj_cf_code, adj_sign,
                                     inter_cf_code, inter_sign.
    """
    data = request.get_json(silent=True) or {}
    coa = data.get('coa') or []
    if not isinstance(coa, list) or not coa:
        return jsonify({'error': 'coa 배열이 비어있습니다.'}), 400

    lock = FileLock('cf_mapping_v2_draft.json.lock', timeout=10)
    try:
        with lock:
            saved = cf_save_mapping_v2(coa,
                                       version_label=data.get('version_label'))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return _json_error(e)
    return jsonify({
        'ok': True,
        'version': saved.get('version'),
        'updated_at': saved.get('updated_at'),
        'coa_count': len(saved.get('coa') or []),
    })


# ─────────────────────────────────────────────────────────────────────────────
# 배포용 패키지 생성
# ─────────────────────────────────────────────────────────────────────────────
DISTRIBUTE_RESULTS_DIR = RESULTS_DIR / 'distribute'
DISTRIBUTE_RESULTS_DIR.mkdir(parents=True, exist_ok=True)

# ─── 배포용 배치 자동 정리 ───────────────────────────────────────────────────
# 생성된 배포 파일은 서버에 영구 보존되므로, 보관 기간이 지난 배치 폴더를 자동 삭제한다.
DISTRIBUTE_RETENTION_DAYS = 30           # 이 일수가 지난 배치는 자동 삭제
_DISTRIBUTE_CLEANUP_INTERVAL = 3600      # 정리 작업 최소 실행 간격(초) — 요청마다 돌지 않게 스로틀
_distribute_cleanup_last = 0.0
_distribute_cleanup_lock = threading.Lock()


def _parse_manifest_dt(s):
    """manifest의 'YYYY-MM-DD HH:MM:SS' 문자열 → datetime. 실패하면 None."""
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), '%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        return None


def _cleanup_distribute_batches(force=False):
    """보관 기간(DISTRIBUTE_RETENTION_DAYS)이 지난 배포용 배치 폴더를 삭제.

    - 기준 시각: manifest의 updated_at(없으면 created_at), 둘 다 없으면 폴더 mtime.
    - 요청마다 돌지 않도록 _DISTRIBUTE_CLEANUP_INTERVAL 간격으로 스로틀(force=True면 무시).
    - 다중 워커/동시 호출 안전: 삭제 실패는 조용히 무시(이미 지워졌으면 OK).
    반환: 삭제한 배치 수.
    """
    global _distribute_cleanup_last
    now_ts = time.time()
    if not force:
        if now_ts - _distribute_cleanup_last < _DISTRIBUTE_CLEANUP_INTERVAL:
            return 0
        with _distribute_cleanup_lock:
            if now_ts - _distribute_cleanup_last < _DISTRIBUTE_CLEANUP_INTERVAL:
                return 0
            _distribute_cleanup_last = now_ts
    else:
        _distribute_cleanup_last = now_ts

    if not DISTRIBUTE_RESULTS_DIR.exists():
        return 0

    import shutil
    max_age_sec = DISTRIBUTE_RETENTION_DAYS * 86400
    now_dt = datetime.now()
    removed = 0
    for d in DISTRIBUTE_RESULTS_DIR.iterdir():
        if not d.is_dir():
            continue
        ref_dt = None
        mani_path = d / '_manifest.json'
        if mani_path.exists():
            try:
                with open(mani_path, 'r', encoding='utf-8') as fp:
                    mani = json.load(fp) or {}
                ref_dt = (_parse_manifest_dt(mani.get('updated_at'))
                          or _parse_manifest_dt(mani.get('created_at')))
            except Exception:
                ref_dt = None
        if ref_dt is None:
            try:
                ref_dt = datetime.fromtimestamp(d.stat().st_mtime)
            except Exception:
                continue
        if (now_dt - ref_dt).total_seconds() > max_age_sec:
            try:
                shutil.rmtree(d, ignore_errors=True)
                removed += 1
            except Exception as e:
                print(f'[배포정리] 삭제 실패 {d.name}: {e}', file=sys.stderr, flush=True)
    if removed:
        print(f'[배포정리] {removed}개 배치 삭제 (>{DISTRIBUTE_RETENTION_DAYS}일 경과)',
              file=sys.stderr, flush=True)
    return removed


def _year4_list():
    """YEARS_DATA의 분기 표기('2026-1Q')에서 4자리 연도만 unique 추출."""
    out = []
    seen = set()
    for y in YEARS_DATA.get('years', []):
        m = re.match(r'^(\d{4})', str(y))
        if m and m.group(1) not in seen:
            seen.add(m.group(1))
            out.append(m.group(1))
    return sorted(out, reverse=True)


def _default_year4():
    d = str(YEARS_DATA.get('default') or '')
    m = re.match(r'^(\d{4})', d)
    return m.group(1) if m else (_year4_list()[0] if _year4_list() else '')


def _is_distribute_owner(username, company_name) -> bool:
    """배포용 패키지 권한 — 자회사는 본인이 직접 담당한 회사만 (연결그룹 동료는 제외).
    관리자/무제한 사용자는 항상 True.
    """
    assigned = _assigned_companies(username)
    if assigned is None:           # 관리자 또는 미지정(=무제한)
        return True
    target = _norm_co(company_name)
    return any(_norm_co(c) == target for c in assigned)


def _accessible_companies_for(username):
    """배포용 패키지 화면에 보일 회사 목록.
    관리자: uploaded_files + 회사 마스터(업로드 대상 회사목록) 전체.
    자회사: 본인이 직접 담당한 회사만 (연결그룹 동료 제외).
    """
    seen = {}
    for f in uploaded_files:
        co = (f.get('company') or '').strip()
        if co:
            seen[_norm_company_name(co)] = co
    # 회사 마스터의 활성 회사 포함
    for n in _company_required_names(active_only=True):
        seen.setdefault(_norm_company_name(n), n)
    all_list = sorted(seen.values())

    if _is_admin(username):
        return all_list
    return [c for c in all_list if _is_distribute_owner(username, c)]


@app.route('/distribute')
@login_required
@require_permission('distribute.run')
def distribute_page():
    """배포 다운로드 (유저 모드) — 파일 생성/다운로드만."""
    uname = session.get('username')
    return render_template(
        'admin_distribute.html',
        mode='user',
        years=_year4_list(),
        default_year=_default_year4(),
        username=uname,
        is_admin=_is_admin(uname),
        companies=_accessible_companies_for(uname),
        retention_days=DISTRIBUTE_RETENTION_DAYS,
    )


@app.route('/distribute/admin')
@login_required
@require_permission('distribute.admin')
def distribute_admin_page():
    """배포 관리 (관리자 모드) — 템플릿 등록 / 분기 비밀번호 / 배포 오픈·폐쇄."""
    uname = session.get('username')
    # 결산기간별 배포 오픈/폐쇄 현재 상태 (화면에서 한눈에 확인·전환)
    open_periods = []
    for p in YEARS_DATA.get('years', []):
        m = re.match(r'^(\d{4})-([1-4])Q$', p)
        if m:
            open_periods.append({'period': p,
                                 'open': dbuilder.is_distribute_open(m.group(1), m.group(2))})
    return render_template(
        'admin_distribute.html',
        mode='admin',
        years=_year4_list(),
        default_year=_default_year4(),
        username=uname,
        is_admin=_is_admin(uname),
        companies=_accessible_companies_for(uname),
        retention_days=DISTRIBUTE_RETENTION_DAYS,
        open_periods=open_periods,
    )


@app.route('/distribute/templates', methods=['GET'])
@login_required
def distribute_list_templates():
    """등록된 빈 템플릿 목록 (조회는 누구나)."""
    return jsonify({'templates': dbuilder.list_templates()})


@app.route('/distribute/quarter-passwords', methods=['GET'])
@admin_required
def distribute_list_quarter_passwords():
    """등록된 분기별 시트 보호 비밀번호 목록 (배포 관리 권한 필요)."""
    return jsonify({'passwords': dbuilder.load_quarter_passwords()})


@app.route('/distribute/quarter-passwords', methods=['POST'])
@admin_required
def distribute_set_quarter_password():
    """분기별 비밀번호 설정/삭제 (배포 관리 권한 필요).
    body: {year: 'YYYY', quarter: '1'~'4', password: str (빈 문자열이면 삭제)}
    """
    data = request.get_json(force=True, silent=True) or {}
    year = str(data.get('year') or '').strip()
    quarter = str(data.get('quarter') or '').strip()
    password = str(data.get('password') or '')
    if not re.match(r'^\d{4}$', year):
        return jsonify({'error': '유효한 연도(YYYY)가 필요합니다.'}), 400
    if quarter not in ('1', '2', '3', '4'):
        return jsonify({'error': '분기(1~4)를 선택하세요.'}), 400
    if password and len(password) < 4:
        return jsonify({'error': '비밀번호는 4자 이상이어야 합니다.'}), 400
    updated = dbuilder.set_quarter_password(year, quarter, password)
    return jsonify({'ok': True, 'passwords': updated})


@app.route('/distribute/quarter-password/check', methods=['GET'])
@login_required
def distribute_check_quarter_password():
    """특정 분기에 비밀번호가 등록되어 있는지 여부만 반환 (누구나 조회 가능, 값은 미반환)."""
    year = (request.args.get('year') or '').strip()
    quarter = (request.args.get('quarter') or '').strip()
    if not re.match(r'^\d{4}$', year) or quarter not in ('1', '2', '3', '4'):
        return jsonify({'error': 'year/quarter 파라미터 오류'}), 400
    pwd = dbuilder.get_quarter_password(year, quarter)
    return jsonify({'has_password': bool(pwd)})


@app.route('/distribute/open-status', methods=['GET'])
@login_required
def distribute_open_status():
    """해당 분기 배포 오픈 여부 조회 (로그인 사용자 누구나). 값만 반환."""
    year = (request.args.get('year') or '').strip()
    quarter = (request.args.get('quarter') or '').strip()
    if not re.match(r'^\d{4}$', year) or quarter not in ('1', '2', '3', '4'):
        return jsonify({'error': 'year/quarter 파라미터 오류'}), 400
    return jsonify({'open': dbuilder.is_distribute_open(year, quarter)})


@app.route('/distribute/open', methods=['POST'])
@admin_required
def distribute_set_open():
    """분기별 배포 오픈/폐쇄 (배포 관리 권한 필요).
    body: {year: 'YYYY', quarter: '1'~'4', open: bool}
    """
    data = request.get_json(force=True, silent=True) or {}
    year = str(data.get('year') or '').strip()
    quarter = str(data.get('quarter') or '').strip()
    open_flag = bool(data.get('open'))
    if not re.match(r'^\d{4}$', year):
        return jsonify({'error': '유효한 연도(YYYY)가 필요합니다.'}), 400
    if quarter not in ('1', '2', '3', '4'):
        return jsonify({'error': '분기(1~4)를 선택하세요.'}), 400
    dbuilder.set_distribute_open(year, quarter, open_flag)
    return jsonify({'ok': True, 'open': open_flag})


def _strip_sheet_protection_bytes(data: bytes) -> bytes:
    """xlsx/xlsm(zip)의 각 워크시트 XML에서 <sheetProtection.../>(및 workbookProtection)를
    제거해 반환. openpyxl 재저장을 거치지 않으므로 서식·수식·매크로가 그대로 보존된다.
    """
    import io as _io, zipfile as _zip
    src = _io.BytesIO(data)
    out = _io.BytesIO()
    with _zip.ZipFile(src, 'r') as zin, _zip.ZipFile(out, 'w', _zip.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            content = zin.read(item.filename)
            nm = item.filename
            if nm.startswith('xl/worksheets/') and nm.endswith('.xml'):
                txt = content.decode('utf-8', 'ignore')
                txt = re.sub(r'<sheetProtection\b[^>]*?/>', '', txt)
                content = txt.encode('utf-8')
            elif nm == 'xl/workbook.xml':
                txt = content.decode('utf-8', 'ignore')
                txt = re.sub(r'<workbookProtection\b[^>]*?/>', '', txt)
                content = txt.encode('utf-8')
            zout.writestr(item, content)
    return out.getvalue()


@app.route('/distribute/unprotect', methods=['POST'])
@require_permission('distribute.admin')
def distribute_unprotect():
    """업로드한 배포용 파일들의 시트 보호를 일괄 해제해 ZIP으로 반환 (관리자 전용).
    form: files (여러 개, .xlsm/.xlsx). 파일 열기암호가 걸린 파일은 해제 불가(개별 스킵).
    """
    import io as _io, zipfile as _zip
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': '파일이 없습니다.'}), 400

    results = []
    used_names = {}
    out_zip = _io.BytesIO()
    with _zip.ZipFile(out_zip, 'w', _zip.ZIP_DEFLATED) as zf:
        for f in files:
            name = (f.filename or 'file.xlsm')
            if not name.lower().endswith(('.xlsm', '.xlsx')):
                results.append(f'[스킵] {name} — xlsx/xlsm 아님')
                continue
            try:
                unlocked = _strip_sheet_protection_bytes(f.read())
                # 파일명 중복 방지
                arc = name
                if arc in used_names:
                    used_names[arc] += 1
                    stem, dot, ext = arc.rpartition('.')
                    arc = f'{stem}({used_names[name]}){dot}{ext}' if dot else f'{arc}({used_names[name]})'
                else:
                    used_names[arc] = 0
                zf.writestr(arc, unlocked)
                results.append(f'[완료] {name}')
            except _zip.BadZipFile:
                results.append(f'[실패] {name} — 파일 열기암호가 걸렸거나 유효한 엑셀이 아님')
            except Exception as e:
                results.append(f'[실패] {name} — {type(e).__name__}: {e}')
        # 처리 결과 요약 텍스트 동봉
        zf.writestr('_보호해제_결과.txt', '\n'.join(results))

    if not any(r.startswith('[완료]') for r in results):
        return jsonify({'error': '보호 해제된 파일이 없습니다.', 'results': results}), 400

    out_zip.seek(0)
    return send_file(out_zip, as_attachment=True,
                     download_name='배포파일_보호해제.zip',
                     mimetype='application/zip')


@app.route('/distribute/template', methods=['POST'])
@admin_required
def distribute_upload_template():
    """빈 패키지 템플릿 업로드 — 배포 관리 권한 필요. form: year, file"""
    year = (request.form.get('year') or '').strip()
    if not re.match(r'^\d{4}$', year):
        return jsonify({'error': '유효한 연도(YYYY)를 선택하세요.'}), 400
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith('.xlsm'):
        return jsonify({'error': '.xlsm 파일만 등록 가능합니다.'}), 400
    try:
        out = dbuilder.save_template(f, year)
        return jsonify({'ok': True, 'path': str(out)})
    except Exception as e:
        return _json_error(e)


@app.route('/distribute/generate', methods=['POST'])
@login_required
@require_permission('distribute.run')
def distribute_generate():
    """
    body(JSON): {
      year: '2026', quarter: '1',
      companies: ['글로벌세아', ...],
    }
    시트 보호 암호: 관리자가 사전에 분기별로 등록한 비밀번호를 자동 사용.
                  등록 안 되어 있으면 400 에러.
    권한:
      · 관리자: 모든 회사 가능
      · 자회사: 본인 담당 회사 또는 같은 연결그룹의 동료 회사만
    반환: { batch_id, results: [...] }  — results에는 file_password 노출 안 함
    """
    _cleanup_distribute_batches()        # 보관기간 지난 배치 정리 (스로틀됨)
    uname = session.get('username')
    is_admin = _is_admin(uname)
    data = request.get_json(force=True, silent=True) or {}
    year = str(data.get('year') or '').strip()
    quarter = str(data.get('quarter') or '').strip()
    companies = data.get('companies') or []

    if not re.match(r'^\d{4}$', year):
        return jsonify({'error': '유효한 연도가 필요합니다.'}), 400
    if quarter not in ('1', '2', '3', '4'):
        return jsonify({'error': '분기(1~4)를 선택하세요.'}), 400
    if not isinstance(companies, list) or not companies:
        return jsonify({'error': '회사를 1개 이상 선택하세요.'}), 400

    # 배포 오픈 여부 — 폐쇄 상태면 자회사(비관리자)는 생성 불가 (환율 확정 전 방지)
    if not is_admin and not dbuilder.is_distribute_open(year, quarter):
        return jsonify({
            'error': f'{year}-{quarter}Q 배포가 아직 열리지 않았습니다. '
                     f'환율 확정 후 관리자가 배포를 열면 파일을 생성할 수 있습니다.'
        }), 403

    # 분기별 비밀번호 조회 — 미등록이면 생성 불가
    quarter_pwd = dbuilder.get_quarter_password(year, quarter)
    if not quarter_pwd:
        return jsonify({
            'error': f'{year}-{quarter}Q 시트 보호 비밀번호가 등록되지 않았습니다. '
                     f'관리자에게 분기 비밀번호 등록을 요청하세요.'
        }), 400

    # 권한 체크 — 자회사 사용자는 본인이 직접 담당한 회사만 (연결그룹 동료 제외)
    if not is_admin:
        denied = [c for c in companies if not _is_distribute_owner(uname, c)]
        if denied:
            return jsonify({
                'error': f'담당하지 않는 회사가 포함되어 있습니다: {", ".join(denied[:5])}'
            }), 403

    tpl = dbuilder.get_template_path(year)
    if not tpl:
        return jsonify({'error': f'{year}년 템플릿이 등록되지 않았습니다. 관리자에게 문의하세요.'}), 400

    # 배치 폴더 (소유자명 포함하여 격리)
    # 클라이언트가 batch_id를 보내면 재사용(같은 배치에 누적), 없으면 새로 생성
    safe_owner = re.sub(r'[^A-Za-z0-9_-]', '_', uname or 'anon')
    batch_id = (data.get('batch_id') or '').strip()
    if batch_id:
        # 보안: 외부에서 임의 경로 못 쓰도록 형식 검증
        if not re.match(r'^[0-9A-Za-z_\-]+$', batch_id):
            return jsonify({'error': '잘못된 batch_id'}), 400
    else:
        batch_id = f"{year}-{quarter}Q_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_owner}"
    batch_dir = DISTRIBUTE_RESULTS_DIR / batch_id
    batch_dir.mkdir(parents=True, exist_ok=True)

    # 기존 배치 사용 시 소유자 일치 확인 (관리자 제외)
    mani_path = batch_dir / '_manifest.json'
    if mani_path.exists() and not _is_admin(uname):
        try:
            with open(mani_path, 'r', encoding='utf-8') as fp:
                _prev = json.load(fp)
            if _prev.get('created_by') and _prev.get('created_by') != uname:
                return jsonify({'error': '해당 배치에 추가할 권한이 없습니다.'}), 403
        except Exception:
            pass

    target_year = int(year)
    target_quarter = int(quarter)

    results = []
    for co in companies:
        co = (co or '').strip()
        if not co:
            continue
        try:
            r = dbuilder.build_distribution_package(
                template_path=tpl,
                output_dir=batch_dir,
                company=co,
                target_year=target_year,
                target_quarter=target_quarter,
                uploaded_files=uploaded_files,
                file_password=None,                       # 파일 열기 암호 없음
                sheet_protect_password=quarter_pwd,        # 분기별 공통 비밀번호로 시트 보호
            )
        except Exception as e:
            r = {
                'ok': False, 'company': co, 'error': str(e),
                'output_path': '', 'file_password': '',
                'source_bs': '', 'source_pl': '', 'pl_scale': 1.0,
                'wrote_bs': 0, 'wrote_pl': 0,
                'missing_bs': [], 'missing_pl': [],
            }
        results.append(r)

    # 매니페스트 누적 (기존 + 신규) — file_password는 저장하지 않음 (분기 비번 사용)
    prev_mani = {}
    if mani_path.exists():
        try:
            with open(mani_path, 'r', encoding='utf-8') as fp:
                prev_mani = json.load(fp) or {}
        except Exception:
            prev_mani = {}
    prev_results = prev_mani.get('results') or []
    # 새 결과에서 file_password 제거 (manifest에도 분기 비번을 따로 저장하지 않음)
    clean_results = []
    for r in results:
        rr = dict(r)
        rr.pop('file_password', None)
        clean_results.append(rr)
    all_results = prev_results + clean_results
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    manifest = {
        'batch_id': batch_id,
        'year': year, 'quarter': quarter,
        'created_at': prev_mani.get('created_at') or now_str,
        'updated_at': now_str,
        'created_by': session.get('username'),
        'pwd_mode': 'quarter',
        'results': all_results,
    }
    try:
        with open(mani_path, 'w', encoding='utf-8') as fp:
            json.dump(manifest, fp, ensure_ascii=False, indent=2)
    except Exception:
        pass

    # CSV — 분기 비밀번호 방식이므로 비밀번호 컬럼 제거
    try:
        csv_path = batch_dir / '_생성목록.csv'
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as fp:
            import csv as _csv
            w = _csv.writer(fp)
            w.writerow(['회사명', '파일명', 'BS 출처', 'PL 출처', 'PL 배수',
                        'BS 채움', 'PL 채움', '미매칭 BS 코드 수', '미매칭 PL 코드 수', '오류'])
            for r in all_results:
                if not r.get('ok'):
                    w.writerow([r.get('company'), '', '', '', '', 0, 0, 0, 0, r.get('error') or ''])
                else:
                    w.writerow([
                        r['company'],
                        Path(r['output_path']).name,
                        r['source_bs'], r['source_pl'],
                        f"{r['pl_scale']:.4f}",
                        r['wrote_bs'], r['wrote_pl'],
                        len(r['missing_bs']), len(r['missing_pl']),
                        '',
                    ])
    except Exception:
        pass

    # 응답: file_password 제거 (모든 사용자) — 분기 비번은 관리자만 별도 페이지에서 조회
    response_results = []
    for r in results:
        rr = dict(r)
        rr.pop('file_password', None)
        response_results.append(rr)
    return jsonify({'ok': True, 'batch_id': batch_id, 'results': response_results})


@app.route('/distribute/download', methods=['GET'])
@login_required
@require_permission('distribute.run')
def distribute_download():
    """
    개별 파일 / CSV / ZIP 다운로드.
    Query: batch=<batch_id>, [file=<filename>] 또는 [kind=csv|zip]

    권한:
      · 관리자: 모든 배치·모든 파일 접근 가능
      · 자회사: 어떤 배치든 본인 담당 회사의 .xlsm만 가능 (CSV/ZIP/manifest 차단)
    """
    uname = session.get('username')
    is_admin = _is_admin(uname)

    batch = (request.args.get('batch') or '').strip()
    if not batch or '..' in batch or '/' in batch or '\\' in batch:
        return jsonify({'error': '잘못된 batch_id'}), 400
    batch_dir = DISTRIBUTE_RESULTS_DIR / batch
    if not batch_dir.exists() or not batch_dir.is_dir():
        return jsonify({'error': '배치를 찾을 수 없음'}), 404

    # 배포 폐쇄 분기의 파일은 자회사(비관리자)에게 다운로드 차단
    # (batch_id 접두사 'YYYY-NQ_' 로 분기 판별)
    if not is_admin:
        _pm = re.match(r'^(\d{4})-([1-4])Q', batch)
        if _pm and not dbuilder.is_distribute_open(_pm.group(1), _pm.group(2)):
            return jsonify({'error': '해당 분기 배포가 닫혀 있어 다운로드할 수 없습니다. 관리자에게 문의하세요.'}), 403

    kind = (request.args.get('kind') or '').strip().lower()
    file_name = (request.args.get('file') or '').strip()

    # CSV(비밀번호목록)는 항상 관리자만
    if kind == 'csv' and not is_admin:
        return jsonify({'error': '비밀번호가 포함된 파일은 관리자만 다운로드할 수 있습니다.'}), 403

    if kind == 'csv':
        p = batch_dir / '_비밀번호목록.csv'
        if not p.exists():
            return jsonify({'error': 'CSV 파일이 없습니다'}), 404
        return send_file(str(p), as_attachment=True,
                         download_name=f'{batch}_passwords.csv',
                         mimetype='text/csv')

    if kind == 'zip':
        # 일괄(ZIP) 다운로드.
        #   · 관리자 : 모든 .xlsm + _비밀번호목록.csv
        #   · 자회사 : 본인 담당 회사의 .xlsm 만 (비밀번호 CSV 제외)
        import zipfile, tempfile
        include_csv = False
        members = []   # (원본경로, zip 내부 이름)
        if is_admin:
            members = [(p, p.name) for p in sorted(batch_dir.glob('*.xlsm'))]
            include_csv = True
        else:
            try:
                with open(batch_dir / '_manifest.json', 'r', encoding='utf-8') as fp:
                    mani = json.load(fp)
            except Exception:
                return jsonify({'error': '배치 메타데이터를 확인할 수 없습니다.'}), 403
            for r in (mani.get('results') or []):
                if not r.get('ok'):
                    continue
                outp = r.get('output_path') or ''
                co = r.get('company')
                if not outp:
                    continue
                fpath = batch_dir / Path(outp).name
                if (fpath.exists() and fpath.suffix.lower() == '.xlsm'
                        and _is_distribute_owner(uname, co)):
                    members.append((fpath, fpath.name))
            if not members:
                return jsonify({'error': '다운로드할 수 있는 본인 담당 파일이 없습니다.'}), 404

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        tmp.close()
        try:
            with zipfile.ZipFile(tmp.name, 'w', zipfile.ZIP_DEFLATED) as zf:
                for src, arc in members:
                    zf.write(str(src), arc)
                if include_csv:
                    csv_p = batch_dir / '_비밀번호목록.csv'
                    if csv_p.exists():
                        zf.write(str(csv_p), '_비밀번호목록.csv')
            return send_file(tmp.name, as_attachment=True,
                             download_name=f'{batch}.zip',
                             mimetype='application/zip')
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    if file_name:
        # 경로 traversal 방지
        if '..' in file_name or '/' in file_name or '\\' in file_name:
            return jsonify({'error': '잘못된 파일명'}), 400

        # 자회사 사용자는 .xlsm 패키지 파일만 + 본인 담당 회사 파일만 허용
        if not is_admin:
            if not file_name.lower().endswith('.xlsm'):
                return jsonify({'error': '해당 파일은 관리자만 다운로드할 수 있습니다.'}), 403
            # manifest에서 파일명 → 회사명 매칭 후 권한 검사
            try:
                with open(batch_dir / '_manifest.json', 'r', encoding='utf-8') as fp:
                    mani = json.load(fp)
            except Exception:
                return jsonify({'error': '배치 메타데이터를 확인할 수 없습니다.'}), 403
            matched_co = None
            for r in (mani.get('results') or []):
                if not r.get('ok'):
                    continue
                outp = r.get('output_path') or ''
                if outp and Path(outp).name == file_name:
                    matched_co = r.get('company')
                    break
            if not matched_co:
                return jsonify({'error': '회사 정보를 확인할 수 없는 파일입니다.'}), 403
            if not _is_distribute_owner(uname, matched_co):
                return jsonify({'error': f'담당하지 않는 회사의 파일입니다: {matched_co}'}), 403

        p = batch_dir / file_name
        if not p.exists() or not p.is_file():
            return jsonify({'error': '파일을 찾을 수 없음'}), 404
        return send_file(str(p), as_attachment=True, download_name=p.name)

    return jsonify({'error': 'kind 또는 file 파라미터가 필요합니다'}), 400


@app.route('/distribute/recent', methods=['GET'])
@login_required
@require_permission('distribute.run')
def distribute_recent():
    """최근 생성한 배포용 배치 목록 — 페이지 재진입 시 결과표 복원용.

    생성된 파일과 _manifest.json 은 서버에 그대로 보존되므로, 다운로드 전에
    다른 페이지로 이동했다 돌아와도 이 API로 직전 결과를 다시 불러올 수 있다.

    권한:
      · 관리자: 모든 배치
      · 자회사: 본인이 생성한(created_by) 배치만
    각 배치의 manifest.results 를 그대로 포함 (file_password 는 애초에 미저장).
    """
    _cleanup_distribute_batches()        # 보관기간 지난 배치 정리 (스로틀됨)
    uname = session.get('username')
    is_admin = _is_admin(uname)
    try:
        limit = int(request.args.get('limit') or 20)
    except (TypeError, ValueError):
        limit = 20
    limit = max(1, min(limit, 100))

    batches = []
    if DISTRIBUTE_RESULTS_DIR.exists():
        for d in DISTRIBUTE_RESULTS_DIR.iterdir():
            if not d.is_dir():
                continue
            mani_path = d / '_manifest.json'
            if not mani_path.exists():
                continue
            try:
                with open(mani_path, 'r', encoding='utf-8') as fp:
                    mani = json.load(fp) or {}
            except Exception:
                continue
            if not is_admin and mani.get('created_by') != uname:
                continue
            results = mani.get('results') or []
            ok_n = sum(1 for r in results if r.get('ok'))
            batches.append({
                'batch_id':   mani.get('batch_id') or d.name,
                'year':       mani.get('year'),
                'quarter':    mani.get('quarter'),
                'created_at': mani.get('created_at'),
                'updated_at': mani.get('updated_at') or mani.get('created_at'),
                'created_by': mani.get('created_by'),
                'is_mine':    mani.get('created_by') == uname,
                'ok_count':   ok_n,
                'fail_count': len(results) - ok_n,
                'results':    results,
            })
    # 최신순 ('YYYY-MM-DD HH:MM:SS' 문자열은 사전식 정렬이 곧 시간순)
    batches.sort(key=lambda b: (b.get('updated_at') or ''), reverse=True)
    return jsonify({'batches': batches[:limit]})


# ─────────────────────────────────────────────────────────────────────────────
# COA Audit — 회사 패키지에서 발견되는 코드 중 등록 안 된 것 검출 (관리자 전용)
# ─────────────────────────────────────────────────────────────────────────────
_COA_AUDIT_PREFIX_RULES = [
    # (prefix tuple, section, section_full, section_sign, type)
    (('42', '43', '45', '48', '52', '53'), 'Ⅰ-2',
     'Ⅰ.  영업활동으로 인한 현금흐름 > 2. 현금의 유출이 없는 비용등의 가산', '+', ''),
    (('41', '44', '46'),                   'Ⅰ-3',
     'Ⅰ.  영업활동으로 인한 현금흐름 > 3. 현금의 유입이 없는 수익등의 차감', '-', ''),
]


def _coa_audit_section_for(code: str):
    """code의 prefix로 (section, section_full, section_sign, type) 추천. 매칭 없으면 None."""
    p2 = code[:2]
    for prefixes, sec, sec_full, sign, ty in _COA_AUDIT_PREFIX_RULES:
        if p2 in prefixes:
            return sec, sec_full, sign, ty
    return None


def _is_header_or_subtotal_code(code: str) -> bool:
    """헤더/subtotal 추정: 7자리 자연 코드 중 끝 5자리 이상이 모두 0.
    예: 1100000(유동자산), 4100000(매출액), 5300000(제조경비) → True
        5300103(개별계정) → False
    """
    if not code or not code[0].isdigit():
        return False
    return len(code) >= 6 and code[-5:] == '00000'


def _cf_lines_eligible_natural(code: str) -> bool:
    """자연 코드가 cf_lines 후보가 될 수 있는지.
    - PL prefix 41/42/43/44/45/46/48/52/53 (사용자 도메인 규칙: 비용/수익/MF)
    - 헤더/subtotal 제외
    BS prefix(1/2/3)는 cf_lines에 들어가지 않으므로 False.
    """
    if not code or not code[0].isdigit():
        return False
    if _is_header_or_subtotal_code(code):
        return False
    p2 = code[:2]
    return p2 in ('41', '42', '43', '44', '45', '46', '48', '52', '53')


def _coa_audit_recommend(code: str) -> dict:
    """코드에 대한 cf_lines 등록 권장도.
    {'level': 'line'|'review'|'ni', 'reason': '...'}

    line  : 자체 CF 라인으로 등록 권장 (영업외수익/비용, 중단사업)
    review: 사례별 판단 필요 (MF 제조경비, CF*)
    ni    : NI에 흡수되는 게 일반적 → cf_lines에 별도 행 불필요 (매출/원가/판관비/법인세)
    """
    if code.upper().startswith('CF'):
        return {'level': 'review', 'reason': 'WC 라인 — 영업자산부채 변동'}
    p2 = code[:2]
    if p2 in ('44', '45', '46'):
        return {'level': 'line',
                'reason': '영업외수익/비용/중단사업 — 자체 CF 라인이 일반적'}
    if p2 in ('52', '53'):
        return {'level': 'review',
                'reason': 'MF 제조경비 — 비현금 비용(상각비 등)만 자체 라인, 나머지는 NI 흡수'}
    if p2 in ('41', '42', '43', '48'):
        return {'level': 'ni',
                'reason': '매출/원가/판관비/법인세 — NI 흡수 (자체 라인 불필요)'}
    return {'level': 'review', 'reason': ''}


def _coa_audit_scan(year_filter=None):
    """업로드된 패키지를 스캔해 BS/PL_MF/CF 시트의 ref_code 수집.
    consol_template + cf_mapping_v2.coa + cf_mapping_v2.cf_lines 와 비교해 누락 보고.

    cf_lines 누락 후보 — 다음 중 하나:
      (a) CF* 접두 코드 (CF1/CF2/...)
      (b) PL/MF 자연 코드 (prefix 41~46/48/52/53), 단 헤더/subtotal 제외

    coa 누락 후보 — 자연 코드 중 cf_lines 후보 조건과 동일 (consol_template detail 미등록).

    반환:
      {
        'period_summary': {year: {scanned_files: N, companies: [...]}},
        'missing_in_cf_lines': [{code, label, prefix_section, ...}],
        'missing_in_coa':      [{code, label, prefix_section, ...}],
        'registered_summary':  {...},
      }

    year_filter : None이면 전체. 특정 year (예: '2026-1Q')이면 그 분기만.
    """
    # 1) 등록된 코드 집합
    try:
        with open('consol_template.json', encoding='utf-8') as f:
            tpl = json.load(f)
        tpl_detail = {str(r['code']) for r in tpl.get('rows', [])
                      if r.get('kind') == 'detail' and r.get('code')}
    except Exception:
        tpl_detail = set()

    try:
        cf_mapping = cf_load_mapping()
    except Exception:
        cf_mapping = {'coa': [], 'cf_lines': []}
    coa_codes = {c['code'] for c in (cf_mapping.get('coa') or [])}
    cf_line_codes = {cl['cf_code'] for cl in (cf_mapping.get('cf_lines') or [])
                     if cl.get('cf_code') not in ('_NI_', '_NONE_')}

    # 2) 업로드 파일 스캔
    seen = {}            # code -> {'label': str, 'seen_in': [(co, year, sheet)], 'any_nonzero': bool}
    period_summary = {}  # year -> {scanned_files, companies set}

    for f in uploaded_files:
        year = f.get('year') or ''
        if year_filter and year != year_filter:
            continue
        period_summary.setdefault(year, {'scanned_files': 0, 'companies': set()})
        period_summary[year]['scanned_files'] += 1
        co = f.get('company') or ''
        if co:
            period_summary[year]['companies'].add(co)

        sheets = ((f.get('extracted') or {}).get('sheets') or {})
        for sheet_name in ('BS', 'PL_MF', 'CF'):
            data = sheets.get(sheet_name) or {}
            for key, v in data.items():
                if not (isinstance(key, str) and key and not key.startswith('LBL::')):
                    continue
                # CF sheet의 'CF1xxx::라벨' 형식 같은 변형 제거 — '::' 앞이 코드
                code = key.split('::', 1)[0] if '::' in key else key
                # 영문/숫자/CF접두만 유효 코드로 인정
                if not (code and (code[0].isdigit() or code.upper().startswith('CF'))):
                    continue
                meta = seen.setdefault(code, {
                    'label': '', 'seen_in': [], 'any_nonzero': False
                })
                if not meta['label']:
                    meta['label'] = (v.get('kor') or '').strip()
                val = v.get('value') or 0
                if val:
                    meta['any_nonzero'] = True
                meta['seen_in'].append((co, year, sheet_name))

    # 3) 누락 분류 — 자연 코드(숫자 시작)와 CF* 분리
    missing_in_cf_lines = []
    missing_in_coa = []
    for code, meta in seen.items():
        is_cf_star = code.upper().startswith('CF')
        eligible_natural = (not is_cf_star) and _cf_lines_eligible_natural(code)
        # cf_lines 후보가 아니면 (BS 자연 코드 / 헤더 / 미해당 prefix) 스킵
        if not (is_cf_star or eligible_natural):
            continue

        in_coa = code in coa_codes
        in_cf_lines = code in cf_line_codes

        # cf_lines 누락 검출 — 합산 시 자동 조회되려면 등록 필요
        if not in_cf_lines:
            rec = _coa_audit_section_for(code) if not is_cf_star else None
            recommend = _coa_audit_recommend(code)
            missing_in_cf_lines.append({
                'code': code,
                'label': meta['label'],
                'prefix_section': rec[0] if rec else (
                    'Ⅰ-4' if is_cf_star else '미지정'
                ),
                'section_full': rec[1] if rec else (
                    'Ⅰ.  영업활동으로 인한 현금흐름 > 4. 영업활동으로 인한 자산부채의 변동'
                    if is_cf_star else ''
                ),
                'section_sign': rec[2] if rec else ('WC' if is_cf_star else ''),
                'type': rec[3] if rec else (
                    'asset' if code.startswith('CF1') else
                    'liability' if code.startswith('CF2') else ''
                ),
                'recommend':       recommend['level'],
                'recommend_reason': recommend['reason'],
                'seen_in':   [{'co': c, 'year': y, 'sheet': s} for c, y, s in meta['seen_in'][:20]],
                'seen_count': len(meta['seen_in']),
                'any_nonzero': meta['any_nonzero'],
                'is_cf_star':  is_cf_star,
            })

        # coa 누락 검출 — 매핑 편집을 위해서는 coa에도 필요 (PL/MF 자연 코드 한정)
        if eligible_natural and (not in_coa):
            missing_in_coa.append({
                'code': code,
                'label': meta['label'],
                'prefix_section': _coa_audit_section_for(code)[0]
                    if _coa_audit_section_for(code) else '미지정',
                'seen_count': len(meta['seen_in']),
                'any_nonzero': meta['any_nonzero'],
            })

    # period_summary의 set → list 직렬화
    out_period = {y: {'scanned_files': v['scanned_files'],
                      'companies': sorted(v['companies'])}
                  for y, v in period_summary.items()}

    return {
        'period_summary': out_period,
        'missing_in_cf_lines': sorted(missing_in_cf_lines,
                                      key=lambda x: (not x['any_nonzero'], x['code'])),
        'missing_in_coa':      sorted(missing_in_coa,
                                      key=lambda x: (not x['any_nonzero'], x['code'])),
        'registered_summary': {
            'consol_template_detail': len(tpl_detail),
            'cf_mapping_coa':         len(coa_codes),
            'cf_mapping_cf_lines':    len(cf_line_codes),
        },
        'year_filter': year_filter or '전체',
    }


@app.route('/admin/coa-audit')
@require_permission('coa.audit')
def admin_coa_audit_page():
    """COA Audit 페이지 (관리자 전용)."""
    return render_template(
        'admin_coa_audit.html',
        years=YEARS_DATA['years'],
        default_year=YEARS_DATA.get('default') or '',
        username=session.get('username'),
        is_admin=True,
    )


@app.route('/admin/coa-audit/scan', methods=['GET'])
@require_permission('coa.audit')
def admin_coa_audit_scan_api():
    year = request.args.get('year') or ''
    if year and not _valid_year(year):
        return jsonify({'error': '유효하지 않은 결산기간'}), 400
    report = _coa_audit_scan(year_filter=year or None)
    return jsonify(report)


@app.route('/admin/coa-audit/add-cf-lines', methods=['POST'])
@require_permission('coa.audit')
def admin_coa_audit_add_cf_lines():
    """선택한 누락 코드들을 cf_mapping_v2_draft.json의 cf_lines에 일괄 추가.

    body: {
      items: [
        {cf_code, name, section, section_full, section_sign, type}, ...
      ]
    }
    """
    data = request.get_json(silent=True) or {}
    items = data.get('items') or []
    if not isinstance(items, list) or not items:
        return jsonify({'error': 'items 배열이 비어있습니다.'}), 400

    lock = FileLock('cf_mapping_v2_draft.json.lock', timeout=10)
    try:
        with lock:
            mapping = cf_load_mapping()
            existing = {cl['cf_code'] for cl in (mapping.get('cf_lines') or [])}
            added = []
            skipped = []
            for it in items:
                code = str(it.get('cf_code') or '').strip()
                if not code:
                    skipped.append({'cf_code': code, 'reason': '코드 누락'})
                    continue
                if code in existing:
                    skipped.append({'cf_code': code, 'reason': '이미 존재'})
                    continue
                section = str(it.get('section') or '').strip()
                # prefix-section 일관성 검증 (자연 코드 한정)
                if code and code[0].isdigit():
                    p2 = code[:2]
                    if p2 in ('42','43','45','48','52','53') and section != 'Ⅰ-2':
                        skipped.append({'cf_code': code,
                                        'reason': f'비용계정 prefix {p2}xxxxx은 Ⅰ-2만 허용 (요청 {section})'})
                        continue
                    if p2 in ('41','44','46') and section != 'Ⅰ-3':
                        skipped.append({'cf_code': code,
                                        'reason': f'수익계정 prefix {p2}xxxxx은 Ⅰ-3만 허용 (요청 {section})'})
                        continue
                new_row = {
                    'cf_code': code,
                    'name':         str(it.get('name') or '').strip(),
                    'section':      section,
                    'section_full': str(it.get('section_full') or '').strip(),
                    'section_sign': str(it.get('section_sign') or '+').strip(),
                    'type':         str(it.get('type') or '').strip(),
                }
                mapping.setdefault('cf_lines', []).append(new_row)
                existing.add(code)
                added.append(code)

            # 백업 후 저장
            from pathlib import Path as _P
            mp = _P('cf_mapping_v2_draft.json')
            if mp.exists():
                import shutil
                shutil.copyfile(str(mp), str(mp.with_suffix('.json.bak')))
            with open('cf_mapping_v2_draft.json', 'w', encoding='utf-8') as fp:
                json.dump(mapping, fp, ensure_ascii=False, indent=2)
            mapping['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    except Exception as e:
        return _json_error(e)

    return jsonify({'ok': True, 'added': added, 'skipped': skipped,
                    'cf_lines_total': len(mapping.get('cf_lines') or [])})


_load_state()

if __name__ == '__main__':
    print("=" * 50)
    print("연결 재무보고 통합 시스템")
    print("브라우저에서 http://localhost:5000 접속")
    print("종료: 이 창에서 Ctrl+C")
    print("=" * 50)
    app.run(debug=False, host='0.0.0.0', port=5000, use_reloader=False, threaded=True,
            extra_files=None)
    # ↑ 템플릿 자동 반영 필요 시: debug=True, use_reloader=True 로 변경 (개발용)
