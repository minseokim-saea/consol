"""
배포용 패키지 생성 모듈.

흐름:
  1) 관리자가 빈 패키지 템플릿(.xlsm)을 업로드 → distribute_template/{year}.xlsm 저장
  2) 회사 선택 → 각 회사별로:
     · 전년 동분기 패키지에서 BS(KRW) · PL(local) 추출
     · 전년 동분기가 없으면 전년 Q4 사용 + PL × (n/4) 비례
     · 템플릿 복사 → Cover/PY 시트 채움 → PY 시트 잠금 → 파일 열기 암호
     · results/distribute/{target_year_q}/{회사명}.xlsm 저장
  3) 결과 CSV([회사명, 파일명, 비밀번호]) 함께 다운로드

PY 시트 입력 사양 (extractor.py:_extract_py_compare 참고):
  · D5↓ 코드 / E5↓ KRW 금액      (BS)
  · G5↓ 코드 / H5↓ 로컬통화 금액  (PL — 추출 시 avg_prior 곱해 KRW 환산됨)
"""
from __future__ import annotations
import json
import re
import csv
import io
import secrets
import shutil
import string
import tempfile
import zipfile
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

import msoffcrypto

# WCE 자본 입력 스키마 — app.py의 _apply_wce_to_aggregation와 동일 로직으로 활용
from wce_schema import WCE_TABLES, WCE_EQUITY_GROUPS


# ─────────────────────────────────────────────────────────────
# 경로
# ─────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "distribute_template"
RESULTS_DIR = BASE_DIR / "results" / "distribute"
UPLOADS_DIR = BASE_DIR / "uploads"
WCE_FILE = BASE_DIR / "wce_overrides.json"
QUARTER_PWD_FILE = BASE_DIR / "distribute_passwords.json"   # 관리자가 등록한 분기별 시트 보호 암호
FX_RATES_FILE = BASE_DIR / "fx_rates.json"

# Master 시트 환율 영역 (extractor.py / 템플릿 구조와 동일)
#   J(10)=Nation, K(11)=Currency, L(12)=Spot, M(13)=Avg, N(14)=Cross
#   당기  : 헤더 J4, 컬럼라벨 5행, 데이터 6~21
#   전년Q4: 헤더 J27, 컬럼라벨 28행, 데이터 29~45
MASTER_FX_CURRENT_ROWS = range(6, 22)
MASTER_FX_PRIOR_ROWS   = range(29, 46)
MASTER_FX_CUR_COL  = 11   # K
MASTER_FX_SPOT_COL = 12   # L
MASTER_FX_AVG_COL  = 13   # M

# WCE 내부 코드 → BS 시트 코드 매핑
WCE_TO_BS_CODE = {'FS32000000': '3600101'}

TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
RESULTS_DIR.mkdir(parents=True, exist_ok=True)


# ─────────────────────────────────────────────────────────────
# 비밀번호 생성
# ─────────────────────────────────────────────────────────────
_PWD_ALPHABET = string.ascii_letters + string.digits


def generate_password(length: int = 12) -> str:
    """영문 대소문자 + 숫자 length자리 무작위 비밀번호."""
    return ''.join(secrets.choice(_PWD_ALPHABET) for _ in range(length))


# ─────────────────────────────────────────────────────────────
# 템플릿 관리
# ─────────────────────────────────────────────────────────────
def save_template(uploaded_file_stream, year: str) -> Path:
    """관리자가 업로드한 빈 패키지 템플릿을 distribute_template/{year}.xlsm 으로 저장.
    year 예: '2026'
    """
    if not re.match(r'^\d{4}$', year):
        raise ValueError(f'유효하지 않은 연도: {year}')
    out = TEMPLATE_DIR / f"{year}.xlsm"
    if hasattr(uploaded_file_stream, 'save'):
        uploaded_file_stream.save(str(out))
    else:
        with open(out, 'wb') as f:
            shutil.copyfileobj(uploaded_file_stream, f)
    return out


def get_template_path(year: str) -> Path | None:
    """등록된 템플릿 경로. 없으면 None."""
    p = TEMPLATE_DIR / f"{year}.xlsm"
    return p if p.exists() else None


def list_templates() -> list[dict]:
    """등록된 템플릿 목록."""
    out = []
    for p in sorted(TEMPLATE_DIR.glob("*.xlsm")):
        out.append({
            'year': p.stem,
            'path': str(p),
            'size': p.stat().st_size,
            'mtime': p.stat().st_mtime,
        })
    return out


# ─────────────────────────────────────────────────────────────
# 전년 패키지 데이터 추출
# ─────────────────────────────────────────────────────────────
def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _code_str(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s or None


# 단순 셀 참조 수식 파서:
#   =Sheet!Col Row,  ='Sheet Name'!$Col$Row,  =SheetName!ColRow
_REF_RE = re.compile(
    r"^=\s*(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!\$?([A-Z]+)\$?(\d+)\s*$"
)


def parse_ref_formula(formula) -> tuple[str, str, int] | None:
    """수식 문자열에서 (sheet_name, col_letter, row_number)를 추출.
    단순 단일 셀 참조만 해석 (=Sheet!A1 형식). SUM/IF 등 복합 수식은 None.
    """
    if not isinstance(formula, str):
        return None
    m = _REF_RE.match(formula)
    if not m:
        return None
    sheet = m.group(1) or m.group(2)
    return (sheet, m.group(3), int(m.group(4)))


def extract_cells_from_package(package_path: Path,
                                refs: list[tuple[str, str, int]]) -> dict:
    """패키지 파일에서 지정한 셀들의 값(계산된 값)을 일괄 추출.
    refs: [(sheet, col_letter, row_number), ...]
    반환: {(sheet, col, row): value}
    """
    out = {}
    if not refs:
        return out
    wb = load_workbook(str(package_path), data_only=True, read_only=False)
    try:
        # 같은 시트는 묶어서 읽기 (사소한 최적화)
        for sn, col, row in refs:
            if sn not in wb.sheetnames:
                out[(sn, col, row)] = None
                continue
            try:
                v = wb[sn][f"{col}{row}"].value
            except Exception:
                v = None
            out[(sn, col, row)] = v
    finally:
        wb.close()
    return out


def extract_pl_mf_secondary(package_path: Path) -> 'OrderedDict[str, float]':
    """
    PL_MF 시트의 '제조원가/공사원가 명세서' 영역 (145행 부근 ~ 끝) 추출.
    구조 (149행 헤더): B=Code, C/D=Account, E=Local, F=Acct Policy Adj, G=Adjusted K-GAAP, N=KRW.
    PY!K열은 로컬통화 입력이므로 G열(Adjusted K-GAAP local)을 사용.

    반환: OrderedDict[code -> local_value]
    """
    from collections import OrderedDict
    out = OrderedDict()
    wb = load_workbook(str(package_path), data_only=True, read_only=True)
    try:
        if 'PL_MF' not in wb.sheetnames:
            return out
        ws = wb['PL_MF']
        # 145행~ 영역에서 첫 데이터 행 찾기: B열에 5xxxxxx 코드가 있는 행
        start_row = None
        for r in range(144, min(ws.max_row + 1, 180)):
            b = ws.cell(r, 2).value
            if _is_num(b) and int(b) >= 5100000:
                start_row = r
                break
        if start_row is None:
            return out
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if len(row) < 7:
                continue
            code = _code_str(row[1])
            if not (code and code[0].isdigit()):
                continue
            if code in out:
                continue
            if _is_num(row[6]):   # G열 (Adjusted K-GAAP local)
                out[code] = float(row[6])
    finally:
        wb.close()
    return out


def extract_py_sheet(package_path: Path) -> dict:
    """
    이미 사용된(또는 채워진) 패키지 파일의 PY 시트에서 모든 영역을 추출.
    신규 회사의 2~4Q 처리 시: 같은 연도 1Q 패키지의 PY를 재사용하기 위해 사용.

    반환:
      {
        'bs':       OrderedDict[code -> krw_value]    (D/E)
        'bs_local': OrderedDict[code -> local_value]  (A/B)
        'pl':       OrderedDict[code -> local_value]  (G/H, 누적 손익)
        'pl_mf_secondary': OrderedDict[code -> local_value]  (J/K, 누적 손익)
        'n_by_row': dict[row -> value]   (N열, 잔액)
        's_by_row': dict[row -> value]   (S열, 잔액)
      }
    """
    from collections import OrderedDict
    out = {
        'bs': OrderedDict(), 'bs_local': OrderedDict(),
        'pl': OrderedDict(), 'pl_mf_secondary': OrderedDict(),
        'n_by_row': {}, 's_by_row': {},
    }
    wb = load_workbook(str(package_path), data_only=True, read_only=True)
    try:
        if 'PY' not in wb.sheetnames:
            return out
        py = wb['PY']
        for r in range(5, py.max_row + 1):
            # A/B — BS Local
            a_code = _code_str(py.cell(r, 1).value)
            b_val = py.cell(r, 2).value
            if a_code and a_code[0].isdigit() and _is_num(b_val) and a_code not in out['bs_local']:
                out['bs_local'][a_code] = float(b_val)
            # D/E — BS KRW
            d_code = _code_str(py.cell(r, 4).value)
            e_val = py.cell(r, 5).value
            if d_code and d_code[0].isdigit() and _is_num(e_val) and d_code not in out['bs']:
                out['bs'][d_code] = float(e_val)
            # G/H — PL
            g_code = _code_str(py.cell(r, 7).value)
            h_val = py.cell(r, 8).value
            if g_code and g_code[0].isdigit() and _is_num(h_val) and g_code not in out['pl']:
                out['pl'][g_code] = float(h_val)
            # J/K — MF (제조원가)
            j_code = _code_str(py.cell(r, 10).value)
            k_val = py.cell(r, 11).value
            if j_code and j_code[0].isdigit() and _is_num(k_val) and j_code not in out['pl_mf_secondary']:
                out['pl_mf_secondary'][j_code] = float(k_val)
            # N — GAAP Diff PY (행 기준)
            n_val = py.cell(r, 14).value
            if _is_num(n_val):
                out['n_by_row'][r] = float(n_val)
            # S — CF PY (행 기준)
            s_val = py.cell(r, 19).value
            if _is_num(s_val):
                out['s_by_row'][r] = float(s_val)
    finally:
        wb.close()
    return out


def list_template_py_refs(template_path: Path) -> dict:
    """
    빈 템플릿의 PY 시트 O열(GAAP Diff CY)·T열(CF CY) 수식을 파싱하여
    "어느 셀에서 가져와야 하는지" 매핑을 반환.

    반환: {
      'n_refs': [(py_row, (sheet, col, row)), ...],  # PY!N{py_row} ← prior(sheet,col,row)
      's_refs': [(py_row, (sheet, col, row)), ...],  # PY!S{py_row} ← prior(sheet,col,row)
    }
    """
    out = {'n_refs': [], 's_refs': []}
    wb = load_workbook(str(template_path), keep_vba=True, data_only=False)
    try:
        if 'PY' not in wb.sheetnames:
            return out
        py = wb['PY']
        for r in range(4, py.max_row + 1):
            # O열 (GAAP Diff CY) → N열에 prior 값 기입
            ref_o = parse_ref_formula(py.cell(r, 15).value)
            if ref_o:
                out['n_refs'].append((r, ref_o))
            # T열 (CF CY) → S열에 prior 값 기입
            ref_t = parse_ref_formula(py.cell(r, 20).value)
            if ref_t:
                out['s_refs'].append((r, ref_t))
    finally:
        wb.close()
    return out


def extract_bs_pl_from_package(package_path: Path) -> dict:
    """
    완성된 패키지 .xlsm에서 BS · PL 당기값을 (local, KRW) 모두 추출.

    참고: extractor.py SHEET_CONFIG
      BS:    code_col=B(2), local_col=F(6), value_col=G(7) KRW, data_start_row=207
      PL_MF: code_col=B(2), local_col=G(7), value_col=N(14) KRW, data_start_row=10

    반환:
      {
        'company': str, 'currency': str, 'year': '2025', 'quarter': '4',
        'bs':       OrderedDict[code -> krw],     # PY  D/E 영역에 채울 값
        'bs_local': OrderedDict[code -> local],   # PY  A/B 영역(선택)에 채울 값
        'pl':       OrderedDict[code -> local],   # PY  G/H 영역에 채울 값 (로컬통화)
      }
    """
    wb = load_workbook(str(package_path), data_only=True, read_only=True)
    try:
        # Cover
        company, currency, year, quarter = None, None, None, None
        if 'Cover' in wb.sheetnames:
            cw = wb['Cover']
            v = cw['D11'].value
            if v is not None:
                company = ' '.join(str(v).split())
            v = cw['B28'].value
            if v is not None:
                currency = str(v).strip().upper()
            v = cw['C9'].value
            if v is not None:
                m = re.search(r'(\d{4})', str(v))
                if m:
                    year = m.group(1)
            v = cw['F9'].value
            if v is not None:
                m = re.search(r'[1-4]', str(v))
                if m:
                    quarter = m.group(0)

        # BS — F열 local, G열 KRW (순서 보존을 위해 OrderedDict)
        from collections import OrderedDict
        bs = OrderedDict()
        bs_local = OrderedDict()
        if 'BS' in wb.sheetnames:
            ws = wb['BS']
            for row in ws.iter_rows(min_row=207, values_only=True):
                # row index: B(2)=row[1], F(6)=row[5], G(7)=row[6]
                if len(row) < 7:
                    continue
                code = _code_str(row[1])
                if not (code and code[0].isdigit()):
                    continue
                if code not in bs and _is_num(row[6]):
                    bs[code] = float(row[6])
                if code not in bs_local and _is_num(row[5]):
                    bs_local[code] = float(row[5])

        # PL_MF — G열 local (PY H열은 로컬통화 입력)
        pl = OrderedDict()
        if 'PL_MF' in wb.sheetnames:
            ws = wb['PL_MF']
            for row in ws.iter_rows(min_row=10, values_only=True):
                if len(row) < 7:
                    continue
                code = _code_str(row[1])
                if not (code and code[0].isdigit()):
                    continue
                if code not in pl and _is_num(row[6]):
                    pl[code] = float(row[6])

        return {
            'company': company,
            'currency': currency,
            'year': year,
            'quarter': quarter,
            'bs': bs,
            'bs_local': bs_local,
            'pl': pl,
        }
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────
# 분기별 시트 보호 비밀번호 (관리자가 사전 등록)
# ─────────────────────────────────────────────────────────────
def _quarter_key(year: int | str, quarter: int | str) -> str:
    """예: (2026, 1) → '2026-1Q'."""
    return f"{int(year)}-{int(quarter)}Q"


def load_quarter_passwords() -> dict:
    """관리자가 등록한 분기별 비밀번호 dict 반환. 키: 'YYYY-NQ'."""
    if not QUARTER_PWD_FILE.exists():
        return {}
    try:
        with open(QUARTER_PWD_FILE, 'r', encoding='utf-8') as fp:
            data = json.load(fp) or {}
        return {k: str(v) for k, v in data.items() if v}
    except Exception:
        return {}


def save_quarter_passwords(data: dict) -> None:
    """분기별 비밀번호 dict를 디스크에 저장 (atomic)."""
    tmp = QUARTER_PWD_FILE.with_suffix('.tmp')
    with open(tmp, 'w', encoding='utf-8') as fp:
        json.dump(data, fp, ensure_ascii=False, indent=2)
    tmp.replace(QUARTER_PWD_FILE)


def get_quarter_password(year: int | str, quarter: int | str) -> str | None:
    """해당 분기에 등록된 시트 보호 비밀번호 반환. 미등록이면 None."""
    return load_quarter_passwords().get(_quarter_key(year, quarter))


def set_quarter_password(year: int | str, quarter: int | str, password: str) -> dict:
    """분기별 비밀번호 설정. 빈 문자열을 보내면 해당 분기 키를 삭제."""
    data = load_quarter_passwords()
    key = _quarter_key(year, quarter)
    if password:
        data[key] = str(password)
    else:
        data.pop(key, None)
    save_quarter_passwords(data)
    return data


# ─────────────────────────────────────────────────────────────
# WCE 자본 입력값 (전년 4Q) — BS 자본항목 대체용
# ─────────────────────────────────────────────────────────────
def _load_wce_overrides() -> dict:
    """wce_overrides.json 전체를 dict로 반환. 없으면 빈 dict."""
    if not WCE_FILE.exists():
        return {}
    try:
        with open(WCE_FILE, 'r', encoding='utf-8') as fp:
            return json.load(fp) or {}
    except Exception:
        return {}


def _compute_table_ending(table_def, table_data):
    """저장된 테이블 데이터로부터 코드별 기말잔액(=기초+증감, 환산효과 제외) 계산.
    app.py의 _compute_table_ending와 동일 로직.
    반환: {code: ending_value}
    """
    out = {}
    for col in table_def['columns']:
        code = col['code']
        ending = 0.0
        for r in table_def['rows']:
            if r['key'] == '환산효과':
                continue
            v = ((table_data.get(code) or {}).get(r['key'], 0) or 0)
            try:
                ending += float(v)
            except (TypeError, ValueError):
                pass
        out[code] = ending
    return out


def compute_wce_equity_bs(wce_record: dict | None) -> 'OrderedDict[str, float]':
    """
    한 회사의 WCE 입력 record(= overrides[year::company])로부터
    BS 자본 leaf 코드 + 자본 소계 + 자본총계를 모두 계산.

    반환: OrderedDict[code -> krw_value]   (없으면 빈 dict)
      · leaf 자본 코드 (3100101, 3100102, 3200101, ..., 3500105)
      · 비지배지분 3600101 (FS32000000 매핑)
      · 자본 소계 3100000, 3200000, 3300000, 3400000, 3500000
      · 자본총계 3000000 (모든 leaf 합)
    """
    from collections import OrderedDict
    out = OrderedDict()
    if not wce_record or not isinstance(wce_record, dict):
        return out
    tables = (wce_record.get('tables') or {})

    # 1) leaf 자본 코드의 기말잔액
    wce_endings = {}
    for t in WCE_TABLES:
        tid = str(t['id'])
        endings = _compute_table_ending(t, tables.get(tid) or {})
        for col in t['columns']:
            wce_endings[col['code']] = endings.get(col['code'], 0.0)

    # 2) WCE 내부 코드 → BS 코드 매핑 (FS32000000 → 3600101)
    for wce_code, val in wce_endings.items():
        bs_code = WCE_TO_BS_CODE.get(wce_code, wce_code)
        out[bs_code] = float(val)

    # 3) 자본 소계 코드 (3100000, 3200000, ...)
    for group_code, children in WCE_EQUITY_GROUPS.items():
        group_val = sum(out.get(WCE_TO_BS_CODE.get(c, c), 0.0) for c in children)
        out[group_code] = float(group_val)

    # 4) 자본총계 3000000 — 모든 WCE leaf 합 (소계 중복 제외)
    leaf_total = sum(wce_endings.values())
    out['3000000'] = float(leaf_total)

    return out


# ─────────────────────────────────────────────────────────────
# 전년 패키지 찾기
# ─────────────────────────────────────────────────────────────
def _norm_company(name: str) -> str:
    if not name:
        return ''
    return re.sub(r'\s+', '', str(name)).lower()


def find_prior_packages(uploaded_files: Iterable[dict], company: str,
                        target_year: int, target_quarter: int) -> dict:
    """
    전년 데이터 후보 패키지 경로를 반환.
      반환: {
        'same_q':       Path | None,   # 전년 동분기 패키지
        'q4':           Path | None,   # 전년 Q4 패키지
        'same_year_q1': Path | None,   # 같은 연도 1Q (신규 회사 2~4Q용)
      }
    """
    norm = _norm_company(company)
    same_q_key = f"{target_year - 1}-{target_quarter}Q"
    q4_key = f"{target_year - 1}-4Q"
    same_y_q1_key = f"{target_year}-1Q"   # 신규 회사 폴백용 (target_quarter > 1)

    same_q = None
    q4 = None
    same_year_q1 = None

    # 같은 회사 + 연도분기 매칭, 최신(uploaded_at) 우선
    for f in sorted(uploaded_files, key=lambda x: x.get('uploaded_at') or '', reverse=True):
        f_co = _norm_company(f.get('company') or '')
        if f_co != norm:
            continue
        f_year = f.get('year') or ''
        p = Path(f.get('path') or '')
        if not p.exists():
            continue
        if f_year == same_q_key and same_q is None:
            same_q = p
        if f_year == q4_key and q4 is None:
            q4 = p
        if target_quarter > 1 and f_year == same_y_q1_key and same_year_q1 is None:
            same_year_q1 = p
        if same_q and q4 and (same_year_q1 or target_quarter == 1):
            break

    return {'same_q': same_q, 'q4': q4, 'same_year_q1': same_year_q1}


def resolve_py_data(uploaded_files: Iterable[dict], company: str,
                    target_year: int, target_quarter: int) -> dict:
    """
    회사별 PY 시트에 채울 데이터를 결정.

    규칙:
      · BS:  항상 전년 Q4 패키지의 BS 사용 (잔액)
      · PL:  전년 동분기 패키지가 있으면 그대로
              없으면 전년 Q4 PL × (target_quarter / 4)  [누적 손익 기준 비례]

    반환:
      {
        'ok': bool,
        'reason': str,                   # 실패 사유 (ok=False일 때)
        'source_bs': str,                # 'YYYY-4Q' 표시
        'source_pl': str,                # 'YYYY-NQ' 표시
        'pl_scale': float,               # PL에 곱할 배수
        'bs': {code: value_krw},
        'pl': {code: value_local},       # 로컬통화 (PY H열용)
        'currency': str,
      }
    """
    found = find_prior_packages(uploaded_files, company, target_year, target_quarter)
    out = {
        'ok': False, 'reason': '',
        'source_bs': '', 'source_pl': '',
        'pl_scale': 1.0,
        'bs': {}, 'bs_local': {}, 'pl': {},
        'pl_mf_secondary': {},   # PY J/K (제조원가 영역)
        'prior_q4_path': None,   # PY N/S 채울 때 참조용 (정상 경로)
        'py_n_by_row': None,     # 신규 회사 경로용 — 행기준 dict
        'py_s_by_row': None,     # 신규 회사 경로용 — 행기준 dict
        'is_new_company': False,
        'wce_overrides_applied': 0,
        'currency': None,
    }

    # ── 신규 회사 경로 — 전년 Q4 없음 ────────────────────────────────
    if not found['q4']:
        # 케이스 A: 1Q 자체 — 전년 자료 없음 → 사용자가 PY 시트를 수기로 입력
        if target_quarter == 1:
            out['ok'] = True
            out['is_new_company'] = True
            out['source_bs'] = '신규 회사 — PY 시트 수기 입력 필요'
            out['source_pl'] = '신규 회사 — PY 시트 수기 입력 필요'
            out['pl_scale'] = 0.0
            return out
        # 케이스 B: 2~4Q — 같은 연도 1Q 패키지에서 PY 시트 재사용
        if found['same_year_q1']:
            py1 = extract_py_sheet(found['same_year_q1'])
            scale = float(target_quarter)   # 1Q PL × n (B 옵션)
            out['bs']              = dict(py1['bs'])
            out['bs_local']        = dict(py1['bs_local'])
            out['pl']              = {k: v * scale for k, v in py1['pl'].items()}
            out['pl_mf_secondary'] = {k: v * scale for k, v in py1['pl_mf_secondary'].items()}
            out['py_n_by_row']     = dict(py1['n_by_row'])    # 잔액 그대로
            out['py_s_by_row']     = dict(py1['s_by_row'])    # 잔액 그대로
            out['source_bs']       = f"신규 회사 — {target_year}-1Q 패키지 PY 그대로"
            out['source_pl']       = f"신규 회사 — {target_year}-1Q 패키지 PY × {target_quarter}"
            out['pl_scale']        = scale
            out['is_new_company']  = True
            out['ok']              = True
            return out
        # 케이스 C: 자료 전무
        out['reason'] = (f'전년({target_year - 1}) Q4 패키지가 없습니다. '
                         f'신규 회사라면 {target_year}-1Q 패키지를 먼저 업로드하세요.')
        return out
    bs_pkg = extract_bs_pl_from_package(found['q4'])
    out['bs'] = dict(bs_pkg['bs'])
    out['bs_local'] = dict(bs_pkg.get('bs_local') or {})
    out['source_bs'] = f"{target_year - 1}-4Q"
    out['currency'] = bs_pkg.get('currency')
    out['prior_q4_path'] = found['q4']

    # ── BS 자본 항목(3xxxxxx) 덮어쓰기: 전년 4Q WCE 본사 입력값으로 대체 ──
    # 패키지의 자본부 대신 WCE에서 산출된 기말잔액 사용 (D/E열 = BS(KRW)에 적용)
    out['wce_overrides_applied'] = 0
    prior_period = f"{target_year - 1}-4Q"
    wce_all = _load_wce_overrides()
    wce_key = f"{prior_period}::{company}"
    wce_rec = wce_all.get(wce_key)
    if wce_rec:
        wce_bs = compute_wce_equity_bs(wce_rec)
        # 3xxxxxx 자본 코드들을 WCE 값으로 덮어쓰기
        for code, val in wce_bs.items():
            if code.startswith('3'):
                out['bs'][code] = val
                out['wce_overrides_applied'] += 1
        out['source_bs'] += f" (자본부 WCE 적용: {out['wce_overrides_applied']}개)"

    # PL & PL_MF_secondary(제조원가): 동분기 우선
    if found['same_q']:
        pl_pkg = extract_bs_pl_from_package(found['same_q'])
        out['pl'] = pl_pkg['pl']
        out['pl_mf_secondary'] = extract_pl_mf_secondary(found['same_q'])
        out['source_pl'] = f"{target_year - 1}-{target_quarter}Q"
        out['pl_scale'] = 1.0
    else:
        # 전년 Q4 × n/4 fallback (PL/제조원가 모두 누적 손익이므로 분기 비례)
        scale = target_quarter / 4.0
        out['pl'] = {k: v * scale for k, v in bs_pkg['pl'].items()}
        sec_q4 = extract_pl_mf_secondary(found['q4'])
        out['pl_mf_secondary'] = {k: v * scale for k, v in sec_q4.items()}
        out['source_pl'] = f"{target_year - 1}-4Q × {target_quarter}/4"
        out['pl_scale'] = scale

    out['ok'] = True
    return out


# ─────────────────────────────────────────────────────────────
# 중앙 환율 로드 (배포용 Master 시트 자동 입력)
# ─────────────────────────────────────────────────────────────
_PERIOD_RE = re.compile(r'^(\d{4})-([1-4])Q$')


def _prior_q4_of(period: str) -> str | None:
    """'2026-1Q' → '2025-4Q'. 형식 오류면 None."""
    m = _PERIOD_RE.match(period or '')
    if not m:
        return None
    return f"{int(m.group(1)) - 1}-4Q"


def load_central_fx(period: str) -> dict:
    """fx_rates.json 에서 (period 당기, 전년4Q 또는 수동 prior) 환율 로드.

    반환: {'current': {CUR: {'spot','avg'}, ...}, 'prior': {...}}
    """
    out = {'current': {}, 'prior': {}}
    if not FX_RATES_FILE.exists():
        return out
    try:
        with open(FX_RATES_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception:
        return out

    out['current'] = (data.get(period) or {}).get('current') or {}

    # 1순위: 같은 period에 명시 저장된 prior (최초연도 케이스)
    explicit_prior = (data.get(period) or {}).get('prior')
    if explicit_prior:
        out['prior'] = explicit_prior
    else:
        # 2순위: 전년 4Q의 current
        pq4 = _prior_q4_of(period)
        if pq4:
            out['prior'] = (data.get(pq4) or {}).get('current') or {}
    return out


def _apply_fx_to_master(ws, fx_current: dict, fx_prior: dict) -> dict:
    """Master 시트의 환율 영역에 중앙 환율을 채워 넣는다.

    채움 규칙: K열 통화코드를 매칭 키로 사용, L=Spot, M=Avg 셀에 값 기록.
      · KRW 행은 건드리지 않음
      · 중앙 환율에 해당 통화가 없으면 해당 행 스킵 (기존 템플릿 값 유지)
      · spot / avg 중 None 인 항목은 해당 셀만 스킵
    """
    def _fill_block(rows, table):
        wrote = 0
        for r in rows:
            cur = ws.cell(r, MASTER_FX_CUR_COL).value
            if cur is None:
                continue
            key = str(cur).strip().upper()
            if not key or key == 'KRW':
                continue
            rec = table.get(key)
            if not rec:
                continue
            if rec.get('spot') is not None:
                ws.cell(r, MASTER_FX_SPOT_COL).value = float(rec['spot'])
                wrote += 1
            if rec.get('avg') is not None:
                ws.cell(r, MASTER_FX_AVG_COL).value = float(rec['avg'])
                wrote += 1
        return wrote

    return {
        'current': _fill_block(MASTER_FX_CURRENT_ROWS, fx_current or {}),
        'prior':   _fill_block(MASTER_FX_PRIOR_ROWS,   fx_prior or {}),
    }


# ─────────────────────────────────────────────────────────────
# 템플릿에 데이터 채우기
# ─────────────────────────────────────────────────────────────
UNPROTECTED_SHEETS = {'TB(m)'}   # 시트 보호에서 제외할 시트명 (자회사 입력 시트)


def fill_template(template_path: Path, output_path: Path,
                  company: str, target_year: str, target_quarter: str,
                  bs_data: dict, pl_data: dict,
                  bs_local_data: dict | None = None,
                  pl_mf_secondary_data: dict | None = None,
                  prior_q4_path: Path | None = None,
                  py_n_by_row: dict | None = None,
                  py_s_by_row: dict | None = None,
                  sheet_protect_password: str | None = None,
                  period: str | None = None) -> dict:
    """
    빈 템플릿을 복사하여 Cover · PY 시트를 채운 후 저장 (.xlsm).

    PY 시트 레이아웃 (4행=Code/Amount 헤더, 5행부터 데이터):
      A(1)/B(2):   BS(Local)  코드 / 금액
      D(4)/E(5):   BS(KRW)    코드 / 금액   ← extractor가 읽는 영역
      G(7)/H(8):   PL         코드 / 금액 (로컬통화)  ← extractor가 읽는 영역
      J(10)/K(11): MF         코드 / 금액 (PL_MF 제조원가, 로컬통화)
      M(13):       GAAP Diff. 코드 (템플릿에 이미 적혀 있음)
      N(14):       GAAP Diff. PY 금액 (BS!F열 GAAP 차이) — 템플릿 O열 수식이 알려줌
      Q(17)/R(18): CF 항목명/세부 (템플릿에 이미 적혀 있음)
      S(19):       CF PY 기말잔액 — 템플릿 T열 수식이 알려줌

    bs_data:              {code: krw_value}    → D5↓ 코드, E5↓ 금액
    bs_local_data:        {code: local_value}  → A5↓ 코드, B5↓ 금액
    pl_data:              {code: local_value}  → G5↓ 코드, H5↓ 금액
    pl_mf_secondary_data: {code: local_value}  → J5↓ 코드, K5↓ 금액
    prior_q4_path:        전년 4Q 패키지 경로  → PY!N{r}, S{r}을 채우는 데 사용

    시트 보호:
      sheet_protect_password 지정 시 TB(m) 시트를 제외한 모든 시트에 보호 설정.

    반환: {
      'wrote_bs', 'wrote_bs_local', 'wrote_pl', 'wrote_pl_mf_secondary',
      'wrote_n_refs', 'wrote_s_refs',
      'protected_sheets': [...],
      'missing_bs': [], 'missing_pl': []
    }
    """
    # 템플릿 복사
    shutil.copy2(str(template_path), str(output_path))

    wb = load_workbook(str(output_path), keep_vba=True)  # .xlsm 매크로 유지
    try:
        # Cover
        if 'Cover' in wb.sheetnames:
            cv = wb['Cover']
            cv['D11'] = company
            if target_year:
                cv['C9'] = int(target_year) if str(target_year).isdigit() else target_year
            if target_quarter:
                cv['F9'] = int(target_quarter) if str(target_quarter).isdigit() else target_quarter

        # Master 시트 환율 자동 입력 (중앙 환율 관리에서 가져옴)
        wrote_fx = {'current': 0, 'prior': 0}
        if 'Master' in wb.sheetnames:
            # period 우선, 없으면 target_year/target_quarter로 조립
            p = period
            if not p and target_year and target_quarter:
                p = f"{target_year}-{target_quarter}Q"
            if p:
                fx = load_central_fx(p)
                wrote_fx = _apply_fx_to_master(wb['Master'], fx['current'], fx['prior'])

        wrote_bs = 0
        wrote_pl = 0
        wrote_bs_local = 0
        wrote_pl_mf_sec = 0
        wrote_n_refs = 0
        wrote_s_refs = 0

        if 'PY' in wb.sheetnames:
            py = wb['PY']
            START = 5   # 데이터 시작 행

            # BS(KRW)  →  D(4) / E(5)
            r = START
            for code, val in (bs_data or {}).items():
                py.cell(r, 4).value = code
                py.cell(r, 5).value = val
                r += 1
                wrote_bs += 1

            # BS(Local)  →  A(1) / B(2)
            r = START
            for code, val in (bs_local_data or {}).items():
                py.cell(r, 1).value = code
                py.cell(r, 2).value = val
                r += 1
                wrote_bs_local += 1

            # PL  →  G(7) / H(8)
            r = START
            for code, val in (pl_data or {}).items():
                py.cell(r, 7).value = code
                py.cell(r, 8).value = val
                r += 1
                wrote_pl += 1

            # MF (PL_MF 제조원가)  →  J(10) / K(11)
            r = START
            for code, val in (pl_mf_secondary_data or {}).items():
                py.cell(r, 10).value = code
                py.cell(r, 11).value = val
                r += 1
                wrote_pl_mf_sec += 1

            # GAAP Diff. PY (N열) + CF PY (S열) 채우기
            # 1순위: py_n_by_row / py_s_by_row dict — 신규 회사 경로에서 1Q PY를 그대로 가져온 경우
            # 2순위: prior_q4_path — 정상 경로에서 빈 템플릿의 O/T 수식이 가리키는 셀을 전년 4Q에서 읽음
            if py_n_by_row or py_s_by_row:
                for r, v in (py_n_by_row or {}).items():
                    if _is_num(v):
                        py.cell(r, 14).value = float(v)
                        wrote_n_refs += 1
                for r, v in (py_s_by_row or {}).items():
                    if _is_num(v):
                        py.cell(r, 19).value = float(v)
                        wrote_s_refs += 1
            elif prior_q4_path and Path(prior_q4_path).exists():
                refs = {'n_refs': [], 's_refs': []}
                for r in range(4, py.max_row + 1):
                    ro = parse_ref_formula(py.cell(r, 15).value)   # O열
                    if ro:
                        refs['n_refs'].append((r, ro))
                    rt = parse_ref_formula(py.cell(r, 20).value)   # T열
                    if rt:
                        refs['s_refs'].append((r, rt))

                all_coords = list({c for _, c in refs['n_refs']}) + \
                             list({c for _, c in refs['s_refs']})
                values = extract_cells_from_package(Path(prior_q4_path), all_coords) \
                         if all_coords else {}

                for r, coord in refs['n_refs']:
                    v = values.get(coord)
                    if _is_num(v):
                        py.cell(r, 14).value = float(v)
                        wrote_n_refs += 1
                for r, coord in refs['s_refs']:
                    v = values.get(coord)
                    if _is_num(v):
                        py.cell(r, 19).value = float(v)
                        wrote_s_refs += 1

        # ── 시트 보호: TB(m) 시트를 제외한 모든 시트에 보호 설정 ──
        protected_sheets = []
        if sheet_protect_password:
            for sn in wb.sheetnames:
                if sn in UNPROTECTED_SHEETS:
                    continue
                ws = wb[sn]
                ws.protection.sheet = True
                ws.protection.password = sheet_protect_password
                ws.protection.enable()
                protected_sheets.append(sn)

        wb.save(str(output_path))

        return {
            'wrote_bs': wrote_bs,
            'wrote_pl': wrote_pl,
            'wrote_bs_local': wrote_bs_local,
            'wrote_pl_mf_secondary': wrote_pl_mf_sec,
            'wrote_n_refs': wrote_n_refs,
            'wrote_s_refs': wrote_s_refs,
            'wrote_fx_current': wrote_fx['current'],
            'wrote_fx_prior':   wrote_fx['prior'],
            'protected_sheets': protected_sheets,
            'missing_bs': [],
            'missing_pl': [],
        }
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────
# 파일 열기 암호 (msoffcrypto-tool)
# ─────────────────────────────────────────────────────────────
def encrypt_file(plain_path: Path, password: str, output_path: Path | None = None):
    """
    .xlsx / .xlsm 파일에 열기 암호 설정.
    output_path 없으면 plain_path를 in-place 갱신.
    """
    if output_path is None:
        output_path = plain_path
    # 임시 파일에 저장 후 교체
    tmp = Path(tempfile.mktemp(suffix=plain_path.suffix))
    try:
        with open(plain_path, 'rb') as f_in, open(tmp, 'wb') as f_out:
            office = msoffcrypto.OfficeFile(f_in)
            office.encrypt(password, f_out)
        shutil.move(str(tmp), str(output_path))
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except Exception:
                pass


# ─────────────────────────────────────────────────────────────
# 통합 함수: 한 회사 처리
# ─────────────────────────────────────────────────────────────
def build_distribution_package(
    template_path: Path,
    output_dir: Path,
    company: str,
    target_year: int,
    target_quarter: int,
    uploaded_files: Iterable[dict],
    file_password: str | None = None,
    sheet_protect_password: str | None = None,
) -> dict:
    """
    한 회사의 배포용 패키지 1개를 생성.

    반환: {
      'ok': bool,
      'company': str,
      'output_path': str,        # 생성된 파일 경로
      'file_password': str,      # 파일 열기 암호
      'source_bs': str,
      'source_pl': str,
      'pl_scale': float,
      'wrote_bs': int, 'wrote_pl': int,
      'missing_bs': [...], 'missing_pl': [...],
      'error': str | None,
    }
    """
    res = {
        'ok': False, 'company': company, 'error': None,
        'output_path': '', 'file_password': '',
        'source_bs': '', 'source_pl': '', 'pl_scale': 1.0,
        'wrote_bs': 0, 'wrote_pl': 0, 'missing_bs': [], 'missing_pl': [],
    }

    # 전년 데이터 결정
    py = resolve_py_data(uploaded_files, company, target_year, target_quarter)
    if not py['ok']:
        res['error'] = py['reason']
        return res

    res['source_bs'] = py['source_bs']
    res['source_pl'] = py['source_pl']
    res['pl_scale'] = py['pl_scale']

    # 출력 경로 — 파일명: {회사명}_{YY}Q{N}.xlsm  (예: 글로벌세아_26Q1.xlsm)
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_co = re.sub(r'[\\/:*?"<>|]', '_', company)
    yy = f"{int(target_year) % 100:02d}"
    output_path = output_dir / f"{safe_co}_{yy}Q{target_quarter}.xlsm"

    # 비밀번호 결정 — PY 시트 보호용
    sheet_pwd = sheet_protect_password or file_password or generate_password()

    # 채우기 (TB(m) 제외 전체 시트 보호 포함)
    try:
        fill_res = fill_template(
            template_path, output_path,
            company=company,
            target_year=str(target_year),
            target_quarter=str(target_quarter),
            bs_data=py['bs'],
            bs_local_data=py.get('bs_local'),
            pl_data=py['pl'],
            pl_mf_secondary_data=py.get('pl_mf_secondary'),
            prior_q4_path=py.get('prior_q4_path'),
            py_n_by_row=py.get('py_n_by_row'),
            py_s_by_row=py.get('py_s_by_row'),
            sheet_protect_password=sheet_pwd,
            period=f"{target_year}-{target_quarter}Q",
        )
        res.update({
            'wrote_bs': fill_res['wrote_bs'],
            'wrote_pl': fill_res['wrote_pl'],
            'wrote_pl_mf_secondary': fill_res.get('wrote_pl_mf_secondary', 0),
            'wrote_n_refs': fill_res.get('wrote_n_refs', 0),
            'wrote_s_refs': fill_res.get('wrote_s_refs', 0),
            'missing_bs': fill_res['missing_bs'],
            'missing_pl': fill_res['missing_pl'],
        })
    except Exception as e:
        res['error'] = f'템플릿 채우기 실패: {e}'
        return res

    res['ok'] = True
    res['output_path'] = str(output_path)
    res['file_password'] = sheet_pwd   # 시트 보호 비밀번호 (호출자가 안내용으로 사용)

    # 무결성 토큰 주입 — 업로드 시 옛 분기/타사 파일 재활용 차단용
    try:
        from distribute_token import generate_token, embed_token
        payload_str, sig = generate_token(
            company=company,
            year=f'{target_year}-{target_quarter}Q',
            template_version=str(template_path.stat().st_mtime_ns),
        )
        embed_token(output_path, payload_str, sig)
        res['token_embedded'] = True
    except Exception as e:
        print(f'[배포 토큰 주입 실패] {company}: {e}', flush=True)
        res['token_embedded'] = False

    return res


# ─────────────────────────────────────────────────────────────
# 결과 ZIP 묶기
# ─────────────────────────────────────────────────────────────
def make_results_zip(results: list[dict], zip_path: Path) -> Path:
    """
    생성된 배포용 파일들 + 비밀번호 CSV를 1개 ZIP으로 묶음.
    """
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    # CSV 만들기
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['회사명', '파일명', '파일 열기 암호', 'BS 출처', 'PL 출처',
                     'PL 배수', 'BS 채움', 'PL 채움', '미매칭 BS 코드 수',
                     '미매칭 PL 코드 수', '오류'])
    for r in results:
        if not r.get('ok'):
            writer.writerow([r.get('company'), '', '', '', '', '', 0, 0, 0, 0, r.get('error') or ''])
        else:
            writer.writerow([
                r['company'],
                Path(r['output_path']).name,
                r['file_password'],
                r['source_bs'], r['source_pl'],
                f"{r['pl_scale']:.4f}",
                r['wrote_bs'], r['wrote_pl'],
                len(r['missing_bs']), len(r['missing_pl']),
                '',
            ])
    csv_bytes = ('﻿' + buf.getvalue()).encode('utf-8')  # BOM (엑셀 한글 호환)

    with zipfile.ZipFile(str(zip_path), 'w', zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            if r.get('ok') and Path(r['output_path']).exists():
                zf.write(r['output_path'], Path(r['output_path']).name)
        zf.writestr('_비밀번호목록.csv', csv_bytes)
    return zip_path
