"""
현금정산표 엔진
- 합산 CF (환산 KRW) + 연결조정/내부거래 분개 + 수기조정 → 최종 CF
- 매핑(v2): COA 기반 — 각 BS/PL 코드를 연결조정용/내부거래용 CF 라인에 각각 매핑
  - cf_lines : 매핑 가능한 CF 행 목록 (가산/차감/WC/투자/재무 + 특수 _NI_, _NONE_)
  - coa      : [{code, name, section, sign,
                 adj_cf_code,   adj_sign,
                 inter_cf_code, inter_sign}]
"""
from __future__ import annotations
import io
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


MAPPING_PATH = Path(__file__).parent / 'cf_mapping_v2_draft.json'
MAPPING_PATH_LEGACY = Path(__file__).parent / 'cf_mapping_draft.json'

# 연결범위변동 행은 cf_code가 없는 plug 행이므로 manuals/roundings dict에서
# 매칭하기 위한 특수 키. UI도 이 키로 입력값을 전송.
SCOPE_CHANGE_KEY = '__SCOPE_CHANGE__'


def _to_num(v):
    try:
        if v is None or v == '':
            return 0.0
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def load_mapping():
    """매핑 JSON 로드 — v2 우선, 없으면 v1(legacy)."""
    path = MAPPING_PATH if MAPPING_PATH.exists() else MAPPING_PATH_LEGACY
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def is_v2_mapping(mapping):
    return isinstance(mapping, dict) and 'coa' in mapping and 'cf_lines' in mapping


def prior_year_4q(period: str) -> str | None:
    """'2026-1Q' / '2026-3Q' → '2025-4Q'. 항상 전년도 4분기.

    내부거래 BS 잔액 elim 차분 산출용 — 결산연도 기말 대비 변동.
    파싱 실패시 None.
    """
    if not period or '-' not in period:
        return None
    try:
        year = int(period.split('-', 1)[0])
        return f'{year - 1}-4Q'
    except ValueError:
        return None


def _cf_line_from_coa_prefix(coa_row):
    """COA 행의 PL prefix로 CF 라인 후보 메타 생성.
    Ⅰ-2 (가산): 42/43/45/48/52/53
    Ⅰ-3 (차감): 41/44/46
    그 외: None (CF 라인 후보 아님)
    """
    code = str(coa_row.get('code') or '')
    p2 = code[:2]
    if p2 in ('42', '43', '45', '48', '52', '53'):
        return {
            'cf_code': code,
            'name': coa_row.get('name', ''),
            'section': 'Ⅰ-2',
            'section_full': 'Ⅰ.  영업활동으로 인한 현금흐름 > 2. 현금의 유출이 없는 비용등의 가산',
            'section_sign': '+',
            'type': '',
        }
    if p2 in ('41', '44', '46'):
        return {
            'cf_code': code,
            'name': coa_row.get('name', ''),
            'section': 'Ⅰ-3',
            'section_full': 'Ⅰ.  영업활동으로 인한 현금흐름 > 3. 현금의 유입이 없는 수익등의 차감',
            'section_sign': '-',
            'type': '',
        }
    return None


def save_mapping_v2(coa_list, version_label=None):
    """v2 매핑의 coa 부분만 갱신해 저장. cf_lines는 그대로 유지.

    coa_list : [{code, name, section, sign, adj_cf_code, adj_sign,
                 inter_cf_code, inter_sign}]
    저장 직전 원본을 .bak로 복사 (있을 때).

    CF 코드 검증: 기존 cf_lines + 특수(_NI_/_NONE_) + COA의 PL prefix 규칙
    (42/43/45/48/52/53→Ⅰ-2, 41/44/46→Ⅰ-3). prefix로만 통과한 신규 코드는
    cf_lines에 자동 추가됨.

    반환: 저장된 dict.
    """
    cur = load_mapping()
    if not is_v2_mapping(cur):
        raise ValueError('현재 매핑이 v2 형식이 아닙니다. _build_cf_mapping_v2.py 로 먼저 생성하세요.')

    valid_cf_codes = {cl['cf_code'] for cl in cur.get('cf_lines') or []}
    valid_cf_codes.update({'_NI_', '_NONE_'})
    valid_signs = {'+', '-'}
    bs_adj_locked = bool((cur.get('policy') or {}).get('bs_adj_locked'))
    pl_inter_locked = bool((cur.get('policy') or {}).get('pl_inter_locked'))

    # 검증 + 정규화
    cur_by_code = {c['code']: c for c in cur['coa']}
    new_cf_lines = []          # prefix로 자동 생성될 cf_line들
    new_cf_codes_seen = set()  # 중복 추가 방지
    normalized = []
    errors = []

    def _accept_cf(cf_code, side_label, row_idx, coa_code):
        """cf_code가 valid인지 검사. cf_lines에 없지만 prefix 규칙을 만족하면
        자동 생성해 valid 풀에 추가. valid면 True, 아니면 errors에 기록 후 False.
        """
        if cf_code in valid_cf_codes:
            return True
        coa_ref = cur_by_code.get(cf_code)
        if coa_ref is None:
            errors.append(f'행 {row_idx+1} ({coa_code}): {side_label} CF "{cf_code}" — COA에 없는 코드')
            return False
        cf_meta = _cf_line_from_coa_prefix(coa_ref)
        if cf_meta is None:
            errors.append(
                f'행 {row_idx+1} ({coa_code}): {side_label} CF "{cf_code}" — '
                f'PL prefix 규칙(42/43/45/48/52/53→Ⅰ-2, 41/44/46→Ⅰ-3) 불일치')
            return False
        if cf_code not in new_cf_codes_seen:
            new_cf_lines.append(cf_meta)
            new_cf_codes_seen.add(cf_code)
        valid_cf_codes.add(cf_code)
        return True

    for i, c in enumerate(coa_list):
        code = str(c.get('code') or '').strip()
        if not code or code not in cur_by_code:
            errors.append(f'행 {i+1}: 알 수 없는 COA 코드 "{code}"')
            continue
        base = cur_by_code[code]
        adj_cf = str(c.get('adj_cf_code') or '_NONE_').strip()
        int_cf = str(c.get('inter_cf_code') or '_NONE_').strip()
        if not _accept_cf(adj_cf, '연결조정', i, code):
            continue
        if not _accept_cf(int_cf, '내부거래', i, code):
            continue
        adj_sign = str(c.get('adj_sign') or '+').strip()
        int_sign = str(c.get('inter_sign') or '+').strip()
        if adj_sign not in valid_signs or int_sign not in valid_signs:
            errors.append(f'행 {i+1} ({code}): 부호는 "+" 또는 "-"만 허용')
            continue

        # 정책: BS 연결조정 잠금 — _NONE_ 외 값 거부
        if bs_adj_locked and base.get('section') == 'BS' and adj_cf != '_NONE_':
            errors.append(
                f'행 {i+1} ({code} {base.get("name","")}): '
                f'BS 항목의 연결조정 매핑은 _NONE_ 으로 고정 (정책 bs_adj_locked). '
                f'요청값 "{adj_cf}" 거부.')
            continue
        # 정책: PL(IS) 내부거래 잠금 — _NI_ 외 값 거부
        if pl_inter_locked and base.get('section') == 'IS' and int_cf != '_NI_':
            errors.append(
                f'행 {i+1} ({code} {base.get("name","")}): '
                f'IS 항목의 내부거래 매핑은 _NI_ 으로 고정 (정책 pl_inter_locked). '
                f'요청값 "{int_cf}" 거부.')
            continue
        normalized.append({
            'code': code,
            'name': base.get('name', ''),
            'section': base.get('section', ''),
            'sign': base.get('sign', 'D'),
            'adj_cf_code':   adj_cf,
            'adj_sign':      adj_sign,
            'inter_cf_code': int_cf,
            'inter_sign':    int_sign,
        })

    if errors:
        raise ValueError('; '.join(errors[:5]) + (f' ... (외 {len(errors)-5}건)' if len(errors) > 5 else ''))

    # 누락 COA는 기존값 보존
    saved_codes = {c['code'] for c in normalized}
    for orig in cur['coa']:
        if orig['code'] not in saved_codes:
            normalized.append(orig)

    # 원래 순서대로 정렬 (cur 순서 기준)
    order = {c['code']: i for i, c in enumerate(cur['coa'])}
    normalized.sort(key=lambda c: order.get(c['code'], 10**9))

    cur['coa'] = normalized
    if new_cf_lines:
        cur.setdefault('cf_lines', []).extend(new_cf_lines)
    if version_label:
        cur['version'] = version_label
    cur['updated_at'] = __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 백업
    path = MAPPING_PATH if MAPPING_PATH.exists() else MAPPING_PATH_LEGACY
    if path.exists():
        bak = path.with_suffix(path.suffix + '.bak')
        import shutil
        shutil.copyfile(path, bak)

    # 저장
    with open(MAPPING_PATH, 'w', encoding='utf-8') as f:
        json.dump(cur, f, ensure_ascii=False, indent=2)
    return cur


def _index_journal(entries):
    """분개 리스트 → {code: {'dr': sum, 'cr': sum}}."""
    out = {}
    for e in entries or []:
        dc = str(e.get('debit_code') or '').strip()
        cc = str(e.get('credit_code') or '').strip()
        da = _to_num(e.get('debit_amt'))
        ca = _to_num(e.get('credit_amt'))
        if dc:
            out.setdefault(dc, {'dr': 0.0, 'cr': 0.0})['dr'] += da
        if cc:
            out.setdefault(cc, {'dr': 0.0, 'cr': 0.0})['cr'] += ca
    return out


def _sign_multiplier(sign):
    return -1.0 if sign == '-' else 1.0


def _apply_rollups_to_cf_agg(agg, companies, rollups):
    """포함 그룹(rollup)을 cf_sheet 의 별도 회사 컬럼으로 합쳐 넣는다.

    consol_engine.compute_with_rollup 과 동일한 패턴:
      · companies 리스트 끝에 rollup 그룹명들을 append
      · cf_sheet[code].by_company 에 {rollup_name: total} 추가

    rollups: [{'name': str, 'cf_by_code': {cf_code: total}}, ...]
    반환: (new_agg, extended_companies)
    """
    if not rollups:
        return agg, list(companies)

    extended_companies = list(companies) + [r['name'] for r in rollups]

    sheets = (agg.get('sheets') or {})
    cf_sheet = sheets.get('CF') or {}

    new_cf = {}
    for code, info in cf_sheet.items():
        new_info = dict(info or {})
        new_info['by_company'] = dict((info or {}).get('by_company') or {})
        new_cf[code] = new_info

    for r in rollups:
        name = r['name']
        for code, val in (r.get('cf_by_code') or {}).items():
            code_s = str(code)
            if code_s not in new_cf:
                # rollup 에만 존재하는 CF 코드도 행으로 살려두기
                new_cf[code_s] = {'kor': '', 'eng': '', 'by_company': {}, 'total': 0}
            v = float(val or 0)
            new_cf[code_s]['by_company'][name] = v
            # info['total']은 aggregator가 직접 회사 합으로 세팅한 값.
            # rollup 컬럼을 by_company에 추가했으므로 total에도 누적해야
            # _agg_cf_lookup의 row_sum이 태림/GIT 같은 rollup을 포함하게 된다.
            new_cf[code_s]['total'] = float(new_cf[code_s].get('total') or 0) + v

    new_sheets = dict(sheets)
    new_sheets['CF'] = new_cf
    new_agg = dict(agg)
    new_agg['sheets'] = new_sheets
    return new_agg, extended_companies


# 글로벌세아 자금조정 매핑 — CF1/CF2/CF3 시트 (연결범위회사) 기반
# (계정 라벨 매칭 키워드, row 라벨 매칭 키워드, CF 코드, 부호)
FUND_ADJ_GLOBAL_SAE_RULES = [
    # 단기대여금 (CF1)
    ('CF1_연결', ['단기대여금'],     ['단기'], ['장기', '유동성'], ['회수'],     'CF3111007', +1),
    ('CF1_연결', ['단기대여금'],     ['단기'], ['장기', '유동성'], ['신규대여'], 'CF4111007', -1),
    # 장기대여금 (CF1) — 유동성장기대여금 제외
    ('CF1_연결', ['장기대여금'],     ['장기'], ['유동성'],         ['회수'],     'CF3121005', +1),
    ('CF1_연결', ['장기대여금'],     ['장기'], ['유동성'],         ['신규대여'], 'CF4121004', -1),
    # 일반대차입금 (CF2)
    ('CF2_연결', ['일반대차입금'],   [],       [],                 ['신규차입'], 'CF5210004', -1),
    ('CF2_연결', ['일반대차입금'],   [],       [],                 ['상환'],     'CF6210004', +1),
    # 장기차입금 (CF3) — 유동성장기차입금 제외
    ('CF3_연결', ['장기차입금'],     ['장기'], ['유동성'],         ['신규차입'], 'CF5220002', -1),
    # 유동성장기차입금 (CF3)
    ('CF3_연결', ['유동성장기차입금'], ['유동성'], [],               ['상환'],     'CF6210001', +1),
]


def compute_fund_adjustments_global_sae(agg):
    """글로벌세아 그룹 전용 — agg.sheets 의 CF1_연결/CF2_연결/CF3_연결 시트에서
    연결범위회사 대상 대여금/차입금 변동을 합산하여 8개 CF 코드별 금액 dict 반환.

    aggregator 결과의 시트 구조:
      sheets[<CF1_연결|CF2_연결|CF3_연결>] = {
        '{code}::{label}' : {
          'kor': '단기대여금 / 신규대여',  # 블록 계정명 + ' / ' + 행 라벨
          'by_company': {co: 환산KRW}, 'total': 합계, ...
        }, ...
      }

    매칭 규칙 (계정 키워드 / 제외 키워드 / 행 라벨):
      - 단기대여금 / 회수      → CF3111007 (단기대여금의 감소)  +
      - 단기대여금 / 신규대여  → CF4111007 (단기대여금의 증가)  −
      - 장기대여금 / 회수      → CF3121005 (장기대여금의 감소)  +
      - 장기대여금 / 신규대여  → CF4121004 (장기대여금의 증가)  −
      - 일반대차입금 / 신규차입 → CF5210004 (일반대차입금의 차입) −
      - 일반대차입금 / 상환     → CF6210004 (일반대차입금의 상환) +
      - 장기차입금 / 신규차입   → CF5220002 (장기차입금의 차입)   −
      - 유동성장기차입금 / 상환 → CF6210001 (유동성장기차입금의 상환) +

    반환: {cf_code: amount, ...}  (값이 0인 항목도 포함되어 8개 키 유지)
    """
    sheets = (agg or {}).get('sheets') or {}
    out = {}

    for (sheet_name, include_kws, must_kws, exclude_kws,
         row_label_kws, cf_code, sign) in FUND_ADJ_GLOBAL_SAE_RULES:
        sheet = sheets.get(sheet_name) or {}
        total = 0.0
        for key, info in sheet.items():
            kor = ((info or {}).get('kor') or '').strip()
            if not kor:
                continue
            # '{블록계정명} / {행라벨}' 형식 — '/' 기준 분리
            if ' / ' in kor:
                acct_name, row_label = kor.split(' / ', 1)
            else:
                acct_name = kor
                row_label = ''
            acct_name = acct_name.strip()
            row_label = row_label.strip()
            # 계정명 매칭: include 키워드 모두 포함 + must 모두 포함 + exclude 하나도 없음
            if not all(k in acct_name for k in include_kws):
                continue
            if must_kws and not all(k in acct_name for k in must_kws):
                continue
            if any(k in acct_name for k in exclude_kws):
                continue
            # 행 라벨 매칭 (회수/신규대여/신규차입/상환 부분 일치)
            if not any(k in row_label for k in row_label_kws):
                continue
            total += float((info or {}).get('total', 0) or 0)
        out[cf_code] = sign * total

    # 매칭이 전혀 없는 코드도 키 유지 — UI/엑셀에서 행 자체는 표시되어야 함
    for (_s, _i, _m, _e, _r, cf_code, _sign) in FUND_ADJ_GLOBAL_SAE_RULES:
        out.setdefault(cf_code, 0.0)
    return out


def compute_cf(agg, adj_entries, inter_entries, companies, manual_adjustments=None,
               mapping=None, prior_inter_entries=None, target_final_ni=None,
               rounding_adjustments=None, rollups=None,
               fund_adjustments=None):
    """현금정산표 계산 진입점. v2 매핑이면 _compute_v2, 아니면 _compute_v1.

    agg                  : aggregator.aggregate() 결과 (CF 시트 포함)
    adj_entries          : 연결조정 분개 (당기)
    inter_entries        : 내부거래 분개 (당기)
    companies            : 회사 순서
    manual_adjustments   : {cf_code: amount} — L컬럼 (수기조정)
    rounding_adjustments : {cf_code: amount} — Q컬럼 (단수조정)
    mapping              : cf_mapping JSON (생략 시 자동 로드)
    prior_inter_entries  : 전기(전년 4Q) 내부거래 분개. None이면 0 처리.
                           내부거래 net = -(당기 dr-cr) + (전기 dr-cr) = (전기 - 당기)
                           ※ 부호 반전 — 자산↑(차변) 시 OCF↑ 가 되도록.
                           연결조정은 차분 적용하지 않음 (당기값 그대로).
    target_final_ni      : 연결정산표 최종 NI. 주어지면 NI 행의 adj 컬럼이
                           plug 값으로 자동 보정되어 최종 NI가 연결정산표와 일치.
                           v2 매핑에서만 적용. None이면 분개 기반 그대로.
    rollups              : 포함 그룹(included_groups) 합산 컬럼.
                           [{'name': '태림(연결)', 'cf_by_code': {cf_code: total}}, ...]
                           consol_engine.compute_with_rollup 과 동일한 처리 패턴.
    fund_adjustments     : {cf_code: amount} — O컬럼 (자금조정). dict 주어지면
                           각 행/소계에 fund_adj 필드 추가 + final에 가산되며
                           result['has_fund_adj']=True. None이면 컬럼 미사용.
                           글로벌세아 전용 — compute_fund_adjustments_global_sea() 참조.
    """
    if mapping is None:
        mapping = load_mapping()

    # rollup 이 있으면 cf_sheet/companies 확장 후 동일하게 진행
    if rollups:
        agg, companies = _apply_rollups_to_cf_agg(agg, companies, rollups)

    if is_v2_mapping(mapping):
        return _compute_v2(agg, adj_entries, inter_entries, companies,
                           manual_adjustments or {}, mapping,
                           prior_inter_entries=prior_inter_entries,
                           target_final_ni=target_final_ni,
                           rounding_adjustments=rounding_adjustments or {},
                           fund_adjustments=fund_adjustments)
    return _compute_v1(agg, adj_entries, inter_entries, companies,
                       manual_adjustments or {}, mapping)


def _compute_v1(agg, adj_entries, inter_entries, companies, manual, mapping):
    """구버전 (CF라인 → lookup_codes 역방향) 매핑."""
    cf_sheet = (agg.get('sheets') or {}).get('CF') or {}

    adj_idx   = _index_journal(adj_entries)
    inter_idx = _index_journal(inter_entries)

    out_sections = []

    # 영업/투자/재무 섹션 소계 누적기
    section_totals_by_kind = {
        'op_addback': {'sum': 0.0, 'manual': 0.0, 'adj': 0.0, 'inter': 0.0, 'final': 0.0},
        'op_subtract': {'sum': 0.0, 'manual': 0.0, 'adj': 0.0, 'inter': 0.0, 'final': 0.0},
        'op_wc': {'sum': 0.0, 'manual': 0.0, 'adj': 0.0, 'inter': 0.0, 'final': 0.0},
        'inv_in': {'sum': 0.0, 'manual': 0.0, 'adj': 0.0, 'inter': 0.0, 'final': 0.0},
        'inv_out': {'sum': 0.0, 'manual': 0.0, 'adj': 0.0, 'inter': 0.0, 'final': 0.0},
        'fin_in': {'sum': 0.0, 'manual': 0.0, 'adj': 0.0, 'inter': 0.0, 'final': 0.0},
        'fin_out': {'sum': 0.0, 'manual': 0.0, 'adj': 0.0, 'inter': 0.0, 'final': 0.0},
    }

    def _kind_of(roman, sub, sign_rule):
        if roman.startswith('Ⅰ') and sub and '가산' in sub:
            return 'op_addback'
        if roman.startswith('Ⅰ') and sub and '차감' in sub:
            return 'op_subtract'
        if roman.startswith('Ⅰ') and sub and '자산부채' in sub:
            return 'op_wc'
        if roman.startswith('Ⅱ') and sub and '유입' in sub:
            return 'inv_in'
        if roman.startswith('Ⅱ') and sub and '유출' in sub:
            return 'inv_out'
        if roman.startswith('Ⅲ') and sub and '유입' in sub:
            return 'fin_in'
        if roman.startswith('Ⅲ') and sub and '유출' in sub:
            return 'fin_out'
        return None

    for sec in mapping.get('sections', []):
        roman = sec.get('roman', '')
        sub = sec.get('sub', '')
        sign_rule = sec.get('sign_rule', '+')
        kind = _kind_of(roman, sub, sign_rule)

        # 차감/유출 섹션은 OCF/ICF 합계 시 행 final을 음수로 반영 (합산시트 -SUM 미러)
        # 가산/유입/WC는 그대로 합산
        subtotal_sign = -1.0 if sign_rule == '-' else 1.0

        out_rows = []
        sec_tot = {'sum': 0.0, 'manual': 0.0, 'adj': 0.0, 'inter': 0.0, 'final': 0.0}

        for row in sec.get('rows', []):
            cf_code = str(row.get('cf_code') or '').strip()
            name = row.get('name') or ''
            sign = row.get('sign') or '+'
            sign_mul = _sign_multiplier(sign)
            lookup_codes = [str(c) for c in (row.get('lookup_codes') or []) if c]

            # K: 합산 (환산 KRW) — agg CF에서 cf_code로 직접 조회
            cf_info = cf_sheet.get(cf_code) or {}
            by_co_raw = cf_info.get('by_company', {}) or {}
            row_companies = {c: float(by_co_raw.get(c, 0) or 0) for c in companies}
            row_sum = float(cf_info.get('total', 0) or 0)
            if row_sum == 0 and row_companies:
                row_sum = sum(row_companies.values())

            # L: 수기조정 — UI에서 cf_code 키로 입력
            row_manual = float(manual.get(cf_code, 0) or 0)

            # M, N: 분개 (lookup_codes로 분개 인덱스 조회 후 sign 반영)
            dr_adj = cr_adj = dr_int = cr_int = 0.0
            for code in lookup_codes:
                a = adj_idx.get(code, {})
                i = inter_idx.get(code, {})
                dr_adj += a.get('dr', 0.0)
                cr_adj += a.get('cr', 0.0)
                dr_int += i.get('dr', 0.0)
                cr_int += i.get('cr', 0.0)

            row_adj = sign_mul * (dr_adj - cr_adj)
            row_inter = sign_mul * (dr_int - cr_int)

            row_final = row_sum + row_manual + row_adj + row_inter

            out_rows.append({
                'cf_code': cf_code,
                'name': name,
                'sign': sign,
                'source': row.get('source'),
                'type': row.get('type'),
                'lookup_codes': lookup_codes,
                'companies': row_companies,
                'sum': row_sum,
                'manual': row_manual,
                'dr_adj': dr_adj, 'cr_adj': cr_adj,
                'dr_int': dr_int, 'cr_int': cr_int,
                'adj': row_adj,
                'inter': row_inter,
                'final': row_final,
            })

            sec_tot['sum']    += row_sum
            sec_tot['manual'] += row_manual
            sec_tot['adj']    += row_adj
            sec_tot['inter']  += row_inter
            sec_tot['final']  += row_final

        # 섹션 소계에 sign_rule 부호 적용 (합산 시트의 차감/유출 섹션 -SUM 효과)
        sec_tot_signed = {k: v * subtotal_sign for k, v in sec_tot.items()}

        out_sections.append({
            'roman': roman,
            'sub': sub,
            'sign_rule': sign_rule,
            'subtotal_sign': subtotal_sign,
            'rows': out_rows,
            'totals': sec_tot,           # 원시 합 (부호 미적용)
            'totals_signed': sec_tot_signed,  # 부호 적용 후
        })

        if kind:
            tot = section_totals_by_kind[kind]
            for k in tot:
                tot[k] += sec_tot_signed[k]

    # ── 1. 당기순이익 (코드 4900001 or 4700004) ────────────────────────────
    # agg.CF에서 "당기순이익" 라벨 또는 코드로 조회
    ni_row = _net_income_from_agg(cf_sheet, companies, adj_idx, inter_idx)

    # 영업/투자/재무 CF 합계
    op_cf = {
        'sum':    ni_row['sum']    + section_totals_by_kind['op_addback']['sum']
                                    + section_totals_by_kind['op_subtract']['sum']
                                    + section_totals_by_kind['op_wc']['sum'],
        'manual': ni_row['manual'] + section_totals_by_kind['op_addback']['manual']
                                    + section_totals_by_kind['op_subtract']['manual']
                                    + section_totals_by_kind['op_wc']['manual'],
        'adj':    ni_row['adj']    + section_totals_by_kind['op_addback']['adj']
                                    + section_totals_by_kind['op_subtract']['adj']
                                    + section_totals_by_kind['op_wc']['adj'],
        'inter':  ni_row['inter']  + section_totals_by_kind['op_addback']['inter']
                                    + section_totals_by_kind['op_subtract']['inter']
                                    + section_totals_by_kind['op_wc']['inter'],
        'final':  ni_row['final']  + section_totals_by_kind['op_addback']['final']
                                    + section_totals_by_kind['op_subtract']['final']
                                    + section_totals_by_kind['op_wc']['final'],
    }
    inv_cf = {k: (section_totals_by_kind['inv_in'][k] + section_totals_by_kind['inv_out'][k])
              for k in ('sum', 'manual', 'adj', 'inter', 'final')}
    fin_cf = {k: (section_totals_by_kind['fin_in'][k] + section_totals_by_kind['fin_out'][k])
              for k in ('sum', 'manual', 'adj', 'inter', 'final')}
    net_cash = {k: op_cf[k] + inv_cf[k] + fin_cf[k]
                for k in ('sum', 'manual', 'adj', 'inter', 'final')}

    return {
        'companies': list(companies),
        'sections': out_sections,
        'net_income': ni_row,
        'op_cf': op_cf,
        'inv_cf': inv_cf,
        'fin_cf': fin_cf,
        'net_cash': net_cash,
        'section_totals_by_kind': section_totals_by_kind,
    }


# ───────────────────────────────────────────────────────────────────────────
# v2 엔진 — COA 기반 라우팅
# ───────────────────────────────────────────────────────────────────────────

def _build_cf_line_index(cf_lines):
    """cf_code → cf_line meta 인덱스. 섹션 그룹핑도 함께 제공."""
    by_code = {l['cf_code']: l for l in cf_lines}
    return by_code


def _build_coa_index(coa_rows):
    """COA code → mapping spec 인덱스."""
    return {c['code']: c for c in coa_rows}


def _agg_cf_lookup(cf_sheet, cf_code, companies):
    """agg.CF 에서 cf_code 행의 by_company / total 추출."""
    info = cf_sheet.get(cf_code) or {}
    by_co = info.get('by_company', {}) or {}
    row_companies = {c: float(by_co.get(c, 0) or 0) for c in companies}
    row_sum = float(info.get('total', 0) or 0)
    if row_sum == 0 and row_companies:
        row_sum = sum(row_companies.values())
    return row_companies, row_sum


def _route_journal_to_cf(entries, coa_idx, mapping_key, sign_key):
    """분개 리스트 → {cf_code: {'dr': sum, 'cr': sum, 'signed': net}}.

    mapping_key : 'adj_cf_code' 또는 'inter_cf_code'
    sign_key    : 'adj_sign' 또는 'inter_sign'

    sign='+' : signed = +(dr - cr)
    sign='-' : signed = -(dr - cr)
    """
    out = {}
    unmatched = []
    for e in entries or []:
        for code_field, amt_field, side in (('debit_code', 'debit_amt', 'dr'),
                                            ('credit_code', 'credit_amt', 'cr')):
            code = str(e.get(code_field) or '').strip()
            amt = _to_num(e.get(amt_field))
            if not code or amt == 0:
                continue
            coa = coa_idx.get(code)
            if not coa:
                unmatched.append({'code': code, 'side': side, 'amt': amt,
                                  'memo': e.get('memo', '')})
                continue
            cf_code = coa.get(mapping_key) or '_NONE_'
            sign = coa.get(sign_key) or '+'
            slot = out.setdefault(cf_code, {'dr': 0.0, 'cr': 0.0,
                                            'signed': 0.0, 'sign': sign})
            slot[side] += amt
            sign_mul = _sign_multiplier(sign)
            slot['signed'] += sign_mul * (amt if side == 'dr' else -amt)
    return out, unmatched


def _compute_v2(agg, adj_entries, inter_entries, companies, manual, mapping,
                prior_inter_entries=None, target_final_ni=None,
                rounding_adjustments=None, fund_adjustments=None):
    """COA 기반 v2 매핑. cf_lines 순서대로 행을 출력, 각 행에 라우팅된 분개 합을 보임.

    prior_inter_entries 가 있으면 내부거래는 (전기-당기) 차분으로 계산.
    target_final_ni 가 주어지면 NI 행의 adj 컬럼이 plug 보정되어 최종 NI 일치.
    rounding_adjustments : {cf_code: amount} — Q컬럼(단수조정). final에 가산.
    fund_adjustments     : {cf_code: amount} — O컬럼(자금조정, 글로벌세아). final에 가산.
    """
    rounding = rounding_adjustments or {}
    fund_adj = fund_adjustments or {}
    has_fund_adj = fund_adjustments is not None
    cf_sheet = (agg.get('sheets') or {}).get('CF') or {}
    pl_sheet = (agg.get('sheets') or {}).get('PL_MF') or {}
    bs_sheet = (agg.get('sheets') or {}).get('BS') or {}
    cf_lines = mapping['cf_lines']
    coa_idx = _build_coa_index(mapping['coa'])
    # COA 이름 인덱스 — fallback용
    coa_name_by_code = {c['code']: (c.get('name') or '').strip()
                        for c in (mapping.get('coa') or [])}

    # 1) 분개 → CF 라우팅
    adj_by_cf,   adj_unmatched   = _route_journal_to_cf(adj_entries,   coa_idx,
                                                       'adj_cf_code',   'adj_sign')
    cur_inter_by_cf, inter_unmatched = _route_journal_to_cf(inter_entries, coa_idx,
                                                            'inter_cf_code', 'inter_sign')
    prior_inter_by_cf, prior_unmatched = _route_journal_to_cf(prior_inter_entries or [], coa_idx,
                                                              'inter_cf_code', 'inter_sign')

    # 내부거래 = 전기 - 당기 (cf_code 키 합집합으로 차분, 부호 반전)
    # 사용자 정의: -(당기 dr-cr) + (전기 dr-cr) = (전기 - 당기)
    inter_by_cf = {}
    all_keys = set(cur_inter_by_cf) | set(prior_inter_by_cf)
    for k in all_keys:
        cur   = cur_inter_by_cf.get(k,   {'dr': 0.0, 'cr': 0.0, 'signed': 0.0})
        prior = prior_inter_by_cf.get(k, {'dr': 0.0, 'cr': 0.0, 'signed': 0.0})
        inter_by_cf[k] = {
            'dr': prior['dr'] - cur['dr'],
            'cr': prior['cr'] - cur['cr'],
            'signed': prior['signed'] - cur['signed'],
            # 디버깅용 추가 메타
            'dr_cur':   cur['dr'],   'cr_cur':   cur['cr'],
            'dr_prior': prior['dr'], 'cr_prior': prior['cr'],
        }

    # 2) cf_lines 순회하며 행 구성 (특수 _NI_ / _NONE_ 제외)
    out_sections = []        # roman → sub → rows 구조와 동일하게 v1처럼 그룹핑
    section_buckets = {}     # (roman, sub) → {meta, rows}
    section_order = []       # 첫 등장 순서 보존

    for cl in cf_lines:
        cf_code = cl['cf_code']
        if cf_code in ('_NI_', '_NONE_'):
            continue

        # 섹션 키 (roman + sub)
        sec_full = cl.get('section_full', '')
        roman = sec_full.split(' > ')[0] if ' > ' in sec_full else sec_full
        sub   = sec_full.split(' > ')[1] if ' > ' in sec_full else ''
        sign_rule = cl.get('section_sign', '+')
        if sign_rule == 'WC':
            section_sign = 'WC'
            subtotal_sign = 1.0
        else:
            section_sign = sign_rule
            subtotal_sign = -1.0 if sign_rule == '-' else 1.0

        sec_key = (roman, sub)
        if sec_key not in section_buckets:
            section_buckets[sec_key] = {
                'roman': roman,
                'sub': sub,
                'sign_rule': sign_rule,
                'subtotal_sign': subtotal_sign,
                'rows': [],
                'totals': {'sum': 0.0, 'manual': 0.0, 'adj': 0.0,
                           'inter': 0.0, 'fund_adj': 0.0,
                           'rounding': 0.0, 'final': 0.0},
                'companies_total': {c: 0.0 for c in companies},
            }
            section_order.append(sec_key)

        # K: 합산
        row_companies, row_sum = _agg_cf_lookup(cf_sheet, cf_code, companies)
        # L: 수기조정
        row_manual = float(manual.get(cf_code, 0) or 0)
        # Q: 단수조정
        row_rounding = float(rounding.get(cf_code, 0) or 0)
        # O: 자금조정 (글로벌세아)
        row_fund_adj = float(fund_adj.get(cf_code, 0) or 0)
        # M: 연결조정 / N: 내부거래 — 라우팅된 net 값
        adj_slot   = adj_by_cf.get(cf_code,   {'dr': 0, 'cr': 0, 'signed': 0})
        inter_slot = inter_by_cf.get(cf_code, {'dr': 0, 'cr': 0, 'signed': 0})

        row_adj   = adj_slot.get('signed', 0.0)
        row_inter = inter_slot.get('signed', 0.0)
        row_final = row_sum + row_manual + row_adj + row_inter + row_fund_adj + row_rounding

        # 계정명 — 4단 fallback:
        #   ① cf_lines.name (명시적 지정)
        #   ② cf_sheet[code].kor (회사 CF 시트 라벨)
        #   ③ cf_mapping.coa[code].name (consol_template detail 명)
        #   ④ PL_MF/BS 시트의 kor (PL/MF 계정명)
        #   ⑤ cf_code 자체 (최후 fallback)
        row_name = (cl.get('name') or '').strip()
        if not row_name:
            row_name = ((cf_sheet.get(cf_code, {}) or {}).get('kor') or '').strip()
        if not row_name:
            row_name = coa_name_by_code.get(cf_code, '').strip()
        if not row_name:
            row_name = ((pl_sheet.get(cf_code, {}) or {}).get('kor') or '').strip()
        if not row_name:
            row_name = ((bs_sheet.get(cf_code, {}) or {}).get('kor') or '').strip()
        if not row_name:
            row_name = cf_code

        bucket = section_buckets[sec_key]
        bucket['rows'].append({
            'cf_code': cf_code,
            'name': row_name,
            'type': cl.get('type') or '',
            'companies': row_companies,
            'sum': row_sum,
            'manual': row_manual,
            'dr_adj': adj_slot.get('dr', 0),   'cr_adj': adj_slot.get('cr', 0),
            'dr_int': inter_slot.get('dr', 0), 'cr_int': inter_slot.get('cr', 0),
            'adj': row_adj,
            'inter': row_inter,
            'fund_adj': row_fund_adj,
            'rounding': row_rounding,
            'final': row_final,
        })
        t = bucket['totals']
        t['sum']      += row_sum
        t['manual']   += row_manual
        t['adj']      += row_adj
        t['inter']    += row_inter
        t['fund_adj'] += row_fund_adj
        t['rounding'] += row_rounding
        t['final']    += row_final
        # 회사별 누적 (합산 K 기준)
        for c in companies:
            bucket['companies_total'][c] += float(row_companies.get(c, 0) or 0)

    # 부호 적용 totals_signed 추가 + 누적기
    _zero_t = lambda: {'sum': 0.0, 'manual': 0.0, 'adj': 0.0,
                       'inter': 0.0, 'fund_adj': 0.0,
                       'rounding': 0.0, 'final': 0.0}
    _zero_co = lambda: {c: 0.0 for c in companies}
    section_totals_by_kind = {
        'op_addback':  _zero_t(),
        'op_subtract': _zero_t(),
        'op_wc':       _zero_t(),
        'inv_in':      _zero_t(),
        'inv_out':     _zero_t(),
        'fin_in':      _zero_t(),
        'fin_out':     _zero_t(),
    }
    section_companies_by_kind = {k: _zero_co() for k in section_totals_by_kind}

    def _kind_of(roman, sub):
        if roman.startswith('Ⅰ') and '가산' in (sub or ''):     return 'op_addback'
        if roman.startswith('Ⅰ') and '차감' in (sub or ''):     return 'op_subtract'
        if roman.startswith('Ⅰ') and '자산부채' in (sub or ''): return 'op_wc'
        if roman.startswith('Ⅱ') and '유입' in (sub or ''):     return 'inv_in'
        if roman.startswith('Ⅱ') and '유출' in (sub or ''):     return 'inv_out'
        if roman.startswith('Ⅲ') and '유입' in (sub or ''):     return 'fin_in'
        if roman.startswith('Ⅲ') and '유출' in (sub or ''):     return 'fin_out'
        return None

    for sec_key in section_order:
        bucket = section_buckets[sec_key]
        subt = bucket['subtotal_sign']
        bucket['totals_signed'] = {k: v * subt for k, v in bucket['totals'].items()}
        bucket['companies_signed'] = {c: v * subt
                                      for c, v in bucket['companies_total'].items()}
        out_sections.append(bucket)
        kind = _kind_of(bucket['roman'], bucket['sub'])
        if kind:
            for k in section_totals_by_kind[kind]:
                section_totals_by_kind[kind][k] += bucket['totals_signed'][k]
            for c in companies:
                section_companies_by_kind[kind][c] += bucket['companies_signed'][c]

    # 3) 당기순이익 — _NI_ 흡수 + agg.CF의 4900001 등
    ni_row = _net_income_v2(cf_sheet, companies, adj_by_cf, inter_by_cf, manual, rounding)
    ni_row.setdefault('fund_adj', 0.0)

    # 3-1) target_final_ni가 주어지면 NI 행의 adj를 plug 보정해 최종 NI 일치
    if target_final_ni is not None:
        target = float(target_final_ni)
        # final = sum + manual + adj + inter + fund_adj + rounding 이 target과 같아지도록 adj 보정
        plug = target - (ni_row['sum'] + ni_row['manual'] + ni_row['inter']
                         + ni_row.get('fund_adj', 0) + ni_row['rounding'])
        ni_row['adj_journal'] = ni_row['adj']         # 원본 분개 기반 adj 보존
        ni_row['adj_plug']    = plug - ni_row['adj']  # 추가된 plug 분
        ni_row['adj']   = plug
        ni_row['final'] = target
        ni_row['ni_target_applied'] = True
        ni_row['ni_target_value']   = target

    # 4) 영업/투자/재무 CF 합계
    _COL_KEYS = ('sum', 'manual', 'adj', 'inter', 'fund_adj', 'rounding', 'final')
    op_cf = {
        k: ni_row[k] + section_totals_by_kind['op_addback'][k]
                     + section_totals_by_kind['op_subtract'][k]
                     + section_totals_by_kind['op_wc'][k]
        for k in _COL_KEYS
    }
    inv_cf = {k: (section_totals_by_kind['inv_in'][k] + section_totals_by_kind['inv_out'][k])
              for k in _COL_KEYS}
    fin_cf = {k: (section_totals_by_kind['fin_in'][k] + section_totals_by_kind['fin_out'][k])
              for k in _COL_KEYS}
    net_cash = {k: op_cf[k] + inv_cf[k] + fin_cf[k]
                for k in _COL_KEYS}

    # 4-b) 회사별 소계 (합산 K 컬럼 기준)
    op_cf['companies'] = {
        c: (ni_row['companies'].get(c, 0)
            + section_companies_by_kind['op_addback'][c]
            + section_companies_by_kind['op_subtract'][c]
            + section_companies_by_kind['op_wc'][c])
        for c in companies
    }
    inv_cf['companies'] = {
        c: (section_companies_by_kind['inv_in'][c] + section_companies_by_kind['inv_out'][c])
        for c in companies
    }
    fin_cf['companies'] = {
        c: (section_companies_by_kind['fin_in'][c] + section_companies_by_kind['fin_out'][c])
        for c in companies
    }
    net_cash['companies'] = {
        c: op_cf['companies'][c] + inv_cf['companies'][c] + fin_cf['companies'][c]
        for c in companies
    }

    # 5) 현금증감 하단 — Ⅴ.환율변동효과 / 연결범위변동 / Ⅵ.기초현금 / Ⅶ.기말현금
    #    라벨 행은 L(수기)/Q(단수) 컬럼 사용자 편집 가능. (Ⅴ, scope: L+Q. 기초/기말: Q.)
    #    cf_engine은 manuals/roundings dict에 매칭되는 키가 있으면 final에 가산.
    fx_effect_row = _label_row_from_cf(
        cf_sheet, companies, ['환율변동', 'Ⅴ'],
        label_hint='Ⅴ. 환율변동효과',
        manuals=manual, roundings=rounding,
    )
    cash_begin_row = _label_row_from_cf(
        cf_sheet, companies, ['기초의현금', '기초현금', 'Ⅵ'],
        label_hint='Ⅵ. 기초의현금',
        manuals=None,  # 기초/기말은 수기조정(L) 비대상 — 단수(Q)만 가능
        roundings=rounding,
    )
    cash_end_row = _label_row_from_cf(
        cf_sheet, companies, ['기말의현금', '기말현금', 'Ⅶ'],
        label_hint='Ⅶ. 기말의현금',
        manuals=None,
        roundings=rounding,
    )
    # 연결범위변동 = 기말 − (기초 + 현금증감 + 환율변동) — plug (K 합산 기반)
    #   plug는 K 컬럼만 계산. 사용자가 L/Q 컬럼에 직접 편집해 P 컬럼 등식 잔차를 0으로 맞추는 구조.
    #   특수 키: SCOPE_CHANGE_KEY 로 manuals/roundings dict에서 추출.
    scope_key = SCOPE_CHANGE_KEY
    scope_manual   = float(manual.get(scope_key, 0)   or 0)
    scope_rounding = float(rounding.get(scope_key, 0) or 0)
    scope_change_sum = (
        cash_end_row['sum']
        - (cash_begin_row['sum'] + net_cash['sum'] + fx_effect_row['sum'])
    )
    scope_change_row = {
        'cf_code': scope_key,  # 매뉴얼 매칭용 특수 키
        'name': '연결범위변동',
        'companies': {c: 0.0 for c in companies},
        'sum': scope_change_sum,
        'manual': scope_manual, 'adj': 0.0, 'inter': 0.0, 'fund_adj': 0.0,
        'rounding': scope_rounding,
        'dr_adj': 0, 'cr_adj': 0, 'dr_int': 0, 'cr_int': 0,
        'final': scope_change_sum + scope_manual + scope_rounding,
        'is_plug': True,
        'plug_note': '기말 − (기초 + 현금증감 + 환율변동) 차액 (K 합산 기반 plug). L/Q 편집 가능.',
    }

    # Ⅴ/Ⅵ/Ⅶ 라벨 행에도 fund_adj=0 부여 (UI/엑셀 컬럼 정렬용)
    for _r in (fx_effect_row, cash_begin_row, cash_end_row):
        if _r is not None:
            _r.setdefault('fund_adj', 0.0)

    return {
        'companies': list(companies),
        'sections': out_sections,
        'net_income': ni_row,
        'op_cf': op_cf,
        'inv_cf': inv_cf,
        'fin_cf': fin_cf,
        'net_cash': net_cash,
        'fx_effect':    fx_effect_row,
        'scope_change': scope_change_row,
        'cash_begin':   cash_begin_row,
        'cash_end':     cash_end_row,
        'section_totals_by_kind': section_totals_by_kind,
        'adj_unmatched_codes':         adj_unmatched,
        'inter_unmatched_codes':       inter_unmatched,
        'prior_inter_unmatched_codes': prior_unmatched,
        'inter_delta_applied':         bool(prior_inter_entries),
        'has_fund_adj':                has_fund_adj,
        'mapping_version': mapping.get('version', '?'),
    }


def _label_row_from_cf(cf_sheet, companies, needles, label_hint='',
                       manuals=None, roundings=None):
    """cf_sheet의 LBL:: 라벨 중 needles 키워드 중 하나라도 포함된 첫 행 추출.
    반환: {cf_code, name, companies, sum, manual, adj, inter, rounding, final, dr_*, cr_*}

    cf_sheet 값 구조: {key: {'kor', 'by_company': {co: val}, 'total': val, ...}}
    L(manual)/Q(rounding) 컬럼은 manuals/roundings dict에 매칭되는 LBL 키가 있으면 적용.
    M(adj)/N(inter)는 분개 라우팅 대상이 아니므로 0.
    final = sum + manual + rounding.
    """
    manuals = manuals or {}
    roundings = roundings or {}
    info = None
    matched_key = ''
    matched_label = label_hint
    for key, v in (cf_sheet or {}).items():
        if not (isinstance(key, str) and key.startswith('LBL::')):
            continue
        label_text = key[5:]
        if any(n in label_text for n in needles):
            info = v
            matched_key = key
            matched_label = (v.get('kor') or '').strip() or label_text
            break

    by_co = (info or {}).get('by_company', {}) or {}
    row_companies = {c: float(by_co.get(c, 0) or 0) for c in companies}
    row_sum = float((info or {}).get('total', 0) or 0)
    if row_sum == 0 and row_companies:
        row_sum = sum(row_companies.values())

    row_manual   = float(manuals.get(matched_key, 0) or 0)   if matched_key else 0.0
    row_rounding = float(roundings.get(matched_key, 0) or 0) if matched_key else 0.0

    return {
        'cf_code': matched_key,  # LBL 키 그대로 노출 — manuals/roundings 매칭용
        'name': matched_label or label_hint,
        'companies': row_companies,
        'sum': row_sum,
        'manual': row_manual, 'adj': 0.0, 'inter': 0.0, 'rounding': row_rounding,
        'dr_adj': 0, 'cr_adj': 0, 'dr_int': 0, 'cr_int': 0,
        'final': row_sum + row_manual + row_rounding,
        'matched': info is not None,
    }


def _net_income_v2(cf_sheet, companies, adj_by_cf, inter_by_cf, manual, rounding=None):
    """v2 NI 행 — _NI_ 코드로 라우팅된 모든 분개 흡수."""
    rounding = rounding or {}
    NI_CODES = ['4900001', '4700004', '3500105']
    info = None
    matched_code = None
    for code in NI_CODES:
        if code in cf_sheet:
            info = cf_sheet[code]; matched_code = code; break
    if info is None:
        for key, v in cf_sheet.items():
            if isinstance(key, str) and key.startswith('LBL::') and '당기순이익' in key:
                info = v; matched_code = key; break

    by_co = (info or {}).get('by_company', {}) or {}
    row_companies = {c: float(by_co.get(c, 0) or 0) for c in companies}
    row_sum = float((info or {}).get('total', 0) or 0)
    if row_sum == 0 and row_companies:
        row_sum = sum(row_companies.values())

    adj_slot   = adj_by_cf.get('_NI_',   {'dr': 0, 'cr': 0, 'signed': 0})
    inter_slot = inter_by_cf.get('_NI_', {'dr': 0, 'cr': 0, 'signed': 0})
    row_manual   = float(manual.get('_NI_', 0) or 0)
    row_rounding = float(rounding.get('_NI_', 0) or 0)
    row_adj      = adj_slot.get('signed', 0.0)
    row_inter    = inter_slot.get('signed', 0.0)

    return {
        'cf_code': matched_code or '_NI_',
        'name': '당기순이익',
        'companies': row_companies,
        'sum': row_sum,
        'manual': row_manual,
        'dr_adj': adj_slot.get('dr', 0),   'cr_adj': adj_slot.get('cr', 0),
        'dr_int': inter_slot.get('dr', 0), 'cr_int': inter_slot.get('cr', 0),
        'adj': row_adj,
        'inter': row_inter,
        'rounding': row_rounding,
        'final': row_sum + row_manual + row_adj + row_inter + row_rounding,
    }


def _net_income_from_agg(cf_sheet, companies, adj_idx, inter_idx):
    """당기순이익 행을 agg.CF에서 추출. 코드 4900001 우선, 없으면 라벨 매칭."""
    NI_CODES = ['4900001', '4700004', '3500105']
    info = None
    matched_code = None
    for code in NI_CODES:
        if code in cf_sheet:
            info = cf_sheet[code]
            matched_code = code
            break
    if info is None:
        for key, v in cf_sheet.items():
            if isinstance(key, str) and key.startswith('LBL::'):
                if '당기순이익' in key:
                    info = v
                    matched_code = key
                    break

    by_co = (info or {}).get('by_company', {}) or {}
    row_companies = {c: float(by_co.get(c, 0) or 0) for c in companies}
    row_sum = float((info or {}).get('total', 0) or 0)
    if row_sum == 0 and row_companies:
        row_sum = sum(row_companies.values())

    # 분개 — NI 코드를 lookup
    dr_adj = cr_adj = dr_int = cr_int = 0.0
    for code in NI_CODES:
        a = adj_idx.get(code, {})
        i = inter_idx.get(code, {})
        dr_adj += a.get('dr', 0.0)
        cr_adj += a.get('cr', 0.0)
        dr_int += i.get('dr', 0.0)
        cr_int += i.get('cr', 0.0)

    # NI는 Cr-naturals 이지만 CF상 추가는 Dr-증가가 NI 증가가 되도록 sign='-' 적용
    # (당기순이익 차변 = 손실 증가 = NI 감소 → CF상도 감소)
    sign_mul = -1.0
    row_adj = sign_mul * (dr_adj - cr_adj)
    row_inter = sign_mul * (dr_int - cr_int)
    return {
        'cf_code': matched_code or '',
        'name': '당기순이익',
        'companies': row_companies,
        'sum': row_sum,
        'manual': 0.0,
        'dr_adj': dr_adj, 'cr_adj': cr_adj,
        'dr_int': dr_int, 'cr_int': cr_int,
        'adj': row_adj,
        'inter': row_inter,
        'final': row_sum + row_adj + row_inter,
    }


# ───────────────────────────────────────────────────────────────────────────
# 엑셀 출력
# ───────────────────────────────────────────────────────────────────────────

HDR_FILL = PatternFill('solid', start_color='1F3864')
HDR_FONT = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=11)
SEC_FILL = PatternFill('solid', start_color='EBF0FA')
SEC_FONT = Font(bold=True, color='1F3864', name='맑은 고딕', size=10)
SUM_FILL = PatternFill('solid', start_color='FFE699')
SUM_FONT = Font(bold=True, color='9C5700', name='맑은 고딕', size=10)
DATA_FONT = Font(name='맑은 고딕', size=10)
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0;(#,##0);"-"'


def _row_all_zero(row) -> bool:
    """detail 행이 K/L/M/N/O/Q/P 모두 0인지. 회사별 값도 모두 0이어야 True."""
    keys = ('sum', 'manual', 'adj', 'inter', 'fund_adj', 'rounding', 'final')
    if any(abs(float(row.get(k) or 0)) >= 0.5 for k in keys):
        return False
    for v in (row.get('companies') or {}).values():
        if abs(float(v or 0)) >= 0.5:
            return False
    return True


def write_cash_worksheet_excel(result, group_name, period, out_path, hide_zero=False):
    """현금정산표 엑셀 생성.
    hide_zero=True : K/L/M/N/(O)/Q/P + 회사별 값이 모두 0인 detail 행은 출력 생략.
                     섹션 소계 / NI / Ⅰ~Ⅳ total / Ⅴ~Ⅶ epilogue는 항상 표시.
    out_path=None  : openpyxl BytesIO 결과 반환 (디스크 저장 없음).
    result['has_fund_adj']=True : N(내부거래) 우측에 O. 자금조정 컬럼 추가 (글로벌세아 전용).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '현금정산표'

    companies = result['companies']
    has_fund_adj = bool(result.get('has_fund_adj'))

    # 헤더 — A=CF코드 / B=계정명 / C..= 회사별 / +합산 K / L 수기조정 / M 연결조정 / N 내부거래
    #        / [O 자금조정] / Q 단수조정 / P 최종 + [빈 spacer] + 연결정산표 + 차이
    n_co = len(companies)
    adj_headers = ['K. 합산', 'L. 수기조정', 'M. 연결조정', 'N. 내부거래']
    if has_fund_adj:
        adj_headers.append('O. 자금조정')
    adj_headers += ['Q. 단수조정', 'P. 최종']
    headers_top = (['CF코드', '계정명'] + companies +
                   adj_headers + ['', '연결정산표', '차이'])

    for i, h in enumerate(headers_top, 1):
        c = ws.cell(1, i, h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

    n_adj = len(adj_headers)   # 6 또는 7

    ws.row_dimensions[1].height = 32
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 32
    for ci in range(n_co):
        ws.column_dimensions[get_column_letter(3 + ci)].width = 14
    # 합계/조정 컬럼 + spacer 1 + 비교 2
    for ci in range(n_adj):
        ws.column_dimensions[get_column_letter(3 + n_co + ci)].width = 15
    ws.column_dimensions[get_column_letter(3 + n_co + n_adj)].width = 3   # spacer
    ws.column_dimensions[get_column_letter(3 + n_co + n_adj + 1)].width = 15  # 연결정산표
    ws.column_dimensions[get_column_letter(3 + n_co + n_adj + 2)].width = 15  # 차이

    row_idx = 2

    def _write_row(values, fill=None, font=None):
        nonlocal row_idx
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row_idx, ci, v if (v not in (None, 0, 0.0) or ci <= 2) else None)
            cell.border = BORDER
            cell.font = font or DATA_FONT
            if fill:
                cell.fill = fill
            if ci > 2:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left' if ci == 2 else 'center')
        row_idx += 1

    def _adj_cells(t):
        """K/L/M/N/[O]/Q/P 셀 값 리스트. has_fund_adj 면 O 자금조정 포함."""
        cells = [t.get('sum', 0), t.get('manual', 0), t.get('adj', 0), t.get('inter', 0)]
        if has_fund_adj:
            cells.append(t.get('fund_adj', 0))
        cells += [t.get('rounding', 0), t.get('final', 0)]
        return cells

    def _totals_row(label, t, fill, font, co_values=None):
        co_values = co_values or {}
        _write_row(
            ['', label] +
            [co_values.get(co, 0) for co in companies] +
            _adj_cells(t or {}) +
            [None, None, None],  # spacer / 연결정산표 / 차이
            fill=fill, font=font,
        )

    def _detail_row(row):
        if hide_zero and _row_all_zero(row):
            return
        compare_tail = [None, None, None]
        if row.get('has_compare'):
            cv = row.get('consol_value')
            cd = row.get('consol_diff')
            compare_tail = [None,
                            cv if cv is not None else 0,
                            cd if cd is not None else 0]
        _write_row(
            [row['cf_code'], '  ' + row['name']] +
            [row['companies'].get(co, 0) for co in companies] +
            _adj_cells(row) +
            compare_tail
        )

    # ── 섹션을 로마숫자별로 그룹핑 ───────────────────────────────────
    # mapping의 sections 순서는 보장됨 (Ⅰ→Ⅱ→Ⅲ)
    from collections import OrderedDict
    roman_groups = OrderedDict()
    for sec in result['sections']:
        roman_groups.setdefault(sec['roman'], []).append(sec)

    roman_totals = {
        'Ⅰ': result['op_cf'],
        'Ⅱ': result['inv_cf'],
        'Ⅲ': result['fin_cf'],
    }

    ni = result['net_income']

    for roman, sections in roman_groups.items():
        roman_key = roman[:1]
        totals = roman_totals.get(roman_key, {})

        # 각 로마 합계 헤더는 섹션 맨 위.
        #   Ⅰ : Ⅰ. 영업활동... → 1. 당기순이익 → 가산/차감/WC
        #   Ⅱ : Ⅱ. 투자활동... → 1. 유입 → 2. 유출
        #   Ⅲ : Ⅲ. 재무활동... → 1. 유입 → 2. 유출
        _totals_row(roman, totals, fill=SUM_FILL, font=SUM_FONT,
                    co_values=(totals or {}).get('companies'))

        # 영업CF 구간: 합계 헤더 뒤에 1. 당기순이익
        if roman_key == 'Ⅰ':
            _write_row(
                [ni.get('cf_code'), '  1. 당기순이익'] +
                [ni['companies'].get(co, 0) for co in companies] +
                _adj_cells(ni) +
                [None, None, None],  # spacer / 연결정산표 / 차이
                fill=SEC_FILL, font=SEC_FONT,
            )

        for sec in sections:
            ts = sec['totals_signed']
            sub_label = '  ' + (sec.get('sub') or '')
            _totals_row(sub_label, ts, fill=SEC_FILL, font=SEC_FONT,
                        co_values=sec.get('companies_signed'))
            for row in sec['rows']:
                _detail_row(row)

    # Ⅳ. 현금의 증감
    _totals_row('Ⅳ. 현금의 증감', result['net_cash'], fill=SUM_FILL, font=SUM_FONT,
                co_values=(result['net_cash'] or {}).get('companies'))

    # 현금증감 하단 4행 — Ⅴ.환율변동효과, 연결범위변동, Ⅵ.기초현금, Ⅶ.기말현금
    def _epilogue_row(label, row, fill=None, font=None):
        if not row:
            return
        _write_row(
            [row.get('cf_code') or '', label] +
            [row.get('companies', {}).get(co, 0) for co in companies] +
            _adj_cells(row) +
            [None, None, None],  # spacer / 연결정산표 / 차이
            fill=fill, font=font,
        )

    _epilogue_row('Ⅴ. 환율변동효과',  result.get('fx_effect'))
    _epilogue_row('연결범위변동',     result.get('scope_change'))
    _epilogue_row('Ⅵ. 기초의현금',    result.get('cash_begin'))
    _epilogue_row('Ⅶ. 기말의현금',    result.get('cash_end'),
                  fill=SUM_FILL, font=SUM_FONT)

    ws.freeze_panes = 'C2'

    # 제목 시트
    info = wb.create_sheet('정보', 0)
    info['A1'] = f'현금정산표 — {group_name} / {period}'
    info['A1'].font = Font(bold=True, size=16, name='맑은 고딕', color='1F3864')
    info['A3'] = '컬럼 설명'
    info['A3'].font = Font(bold=True, name='맑은 고딕')
    descs = [
        ('K. 합산',     'aggregator.CF (각사 환산 KRW 합산)'),
        ('L. 수기조정', '사용자가 직접 입력하는 보정 (UI 입력)'),
        ('M. 연결조정', '연결조정 분개 (투자자본 + 미실현 통합)'),
        ('N. 내부거래', '내부거래 분개'),
    ]
    if has_fund_adj:
        descs.append(('O. 자금조정',
                      '글로벌세아 그룹 전용 — CF1/CF2/CF3 시트의 연결범위회사 '
                      '대여금/차입금 변동 합산'))
    descs += [
        ('Q. 단수조정', '단수절사/반올림 보정 (UI 입력)'),
        ('P. 최종',     ('K + L + M + N + ' + ('O + ' if has_fund_adj else '') +
                        'Q — 행 단위 / 섹션 소계는 부호 적용 후')),
        ('연결정산표',  '현표비교계정.xlsx의 코드만 — 연결정산표 detail 행의 final 값'),
        ('차이',        'P. 최종 − 연결정산표 (현금정산표 vs 연결정산표 검증용)'),
        ('Ⅴ. 환율변동효과', '비-KRW 회사 환산 차액 (extractor가 Ⅶ−Ⅵ−Ⅳ로 자동 보정)'),
        ('연결범위변동', '기말 − (기초 + 현금증감 + 환율변동) plug. 신규 편입/제외 영향'),
        ('Ⅵ. 기초의현금', 'CF 시트 라벨 매칭 (각사 전기 spot rate 환산 합산)'),
        ('Ⅶ. 기말의현금', 'CF 시트 라벨 매칭 (각사 당기 spot rate 환산 합산)'),
    ]
    for i, (k, v) in enumerate(descs, 4):
        info[f'A{i}'] = k
        info[f'B{i}'] = v
    info.column_dimensions['A'].width = 14
    info.column_dimensions['B'].width = 80

    if out_path is None:
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio
    wb.save(out_path)
    return out_path
